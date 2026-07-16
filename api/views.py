from django.shortcuts import render,  redirect, get_object_or_404
from django.http import FileResponse, JsonResponse, HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.db.models import Sum, Q, F, Count, Value, CharField, DecimalField, DateField, ExpressionWrapper
from django.db.models.functions import ExtractYear, ExtractMonth, Concat, Coalesce, TruncYear, TruncMonth, Cast
from django.contrib.auth.models import User
from django.conf import settings

from rest_framework.response import Response
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework import generics
from rest_framework.authentication import TokenAuthentication, SessionAuthentication
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.pagination import PageNumberPagination

from dashboard.models import Inventario, Order, Product, ArticulosOrdenados, ArticulosparaSurtir
from compras.models import Compra, Proveedor_direcciones, Moneda, Proveedor, ArticuloComprado
from solicitudes.models import Proyecto, Subproyecto
from requisiciones.models import Requis, ArticulosRequisitados, Salidas
from tesoreria.models import Saldo_Cuenta, Pago, Cuenta
from tesoreria.filters import Matriz_Pago_Filter

from user.models import Profile, Distrito
from .serializers import  CompraTablaLiteSerializer, ProveedorDireccionesSerializer, ProyectoSerializer, SubProyectoSerializer, MonedaSerializer
from .serializers import ProfileSerializer, DistritoSerializer, RequisicionSerializer, ProveedorSerializer, OrdenSerializer, Compra_tabla_Serializer
from .serializers import InventarioSerializer, ProductSerializer, Articulos_Ordenados_Serializer,Articulos_para_Surtir_Serializer, Articulos_Requisitados_Serializer, Articulo_Comprado_Serializer
from .serializers import PagoControlBancosSerializer, ReporteSolicitudesSerializer

import requests

from user.models import CustomUser, Empresa

from rest_framework.views import APIView
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


from compras.views import generar_pdf_nueva
from rest_framework import status
from user.decorators import perfil_seleccionado_required
from api.models import TablaFestivos
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
import logging

logger = logging.getLogger("user.middleware")

#import openai

from openai import OpenAI
#import os

import mysql.connector
client = OpenAI(
    organization='org-9Legd0seRBYosepjlvTnzipq',
    project='proj_82cTfrUnAMXikaj5cdr1Dk5a',
    api_key = settings.OPENAI_API_KEY,
    )


# Create your views here.
@api_view(['GET'])
@authentication_classes([SessionAuthentication,TokenAuthentication])
@permission_classes([IsAuthenticated])
def monedas_api(request):
    monedas = Moneda.objects.all()
    #page = request.query_params.get('page', 1)
    #per_page = request.query_params.get('per_page', 20)
    #
    #ordering = request.query_params.get('ordering')

    #if ordering:
    #    monedas = monedas.order_by(ordering)
        
    
    serialized_monedas = MonedaSerializer(monedas, many=True)

    #paginator = Paginator(monedas, per_page=per_page)
    #try: 
    #    monedas = paginator.page(number=page)
    #except EmptyPage:
    #    monedas = []
        
    return Response(serialized_monedas.data)

# Create your views here.
@api_view(['GET'])
@authentication_classes([SessionAuthentication,TokenAuthentication])
@permission_classes([IsAuthenticated])
def profiles_api(request):
    profiles = Profile.objects.all()
        
    serialized_profiles = ProfileSerializer(profiles, many=True)
        
    return Response(serialized_profiles.data)

# Create your views here.
@api_view(['GET'])
@authentication_classes([SessionAuthentication,TokenAuthentication])
@permission_classes([IsAuthenticated])
def proyectos_api(request):
    proyectos = Proyecto.objects.filter(activo = True)
    #page = request.query_params.get('page', 1)
    #per_page = request.query_params.get('per_page', 20)
    #
    #ordering = request.query_params.get('ordering')

    #if ordering:
    #    proyectos = Proyecto.order_by(ordering)
        
    
    serialized_proyectos = ProyectoSerializer(proyectos, many=True)

    #paginator = Paginator(proyectos, per_page=per_page)
    #try: 
    #    proyectos = paginator.page(number=page)
    #except EmptyPage:
    #    proyectos = []
        
    return Response(serialized_proyectos.data)

@api_view(['GET'])
@authentication_classes([SessionAuthentication,TokenAuthentication])
@permission_classes([IsAuthenticated])
def subproyectos_api(request):
    subproyectos = Subproyecto.objects.filter(activo = True)
    #page = request.query_params.get('page', 1)
    #per_page = request.query_params.get('per_page', 20)
    #
    #ordering = request.query_params.get('ordering')

    #if ordering:
    #    subproyectos = Subproyecto.order_by(ordering)
        
    
    serialized_subproyectos = SubProyectoSerializer(subproyectos, many=True)

    #paginator = Paginator(subproyectos, per_page=per_page)
    #try: 
    #    subproyectos = paginator.page(number=page)
    #except EmptyPage:
    #    subproyectos = []
        
    return Response(serialized_subproyectos.data)


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def productos_api(request):
    productos = Product.objects.all()
    serializer = ProductSerializer(productos, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def inventario_api(request):
    inventario = Inventario.objects.all()
    serializer = InventarioSerializer(inventario, many=True)
    return Response(serializer.data)

@api_view(["GET"])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def solicitudes_api(request):
    solicitudes = Order.objects.filter(complete=True).order_by("id")
    serializer = OrdenSerializer(solicitudes, many=True)
    return Response(serializer.data)

@api_view(["GET"])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def productos_solicitados_api(request):
    articulos = ArticulosOrdenados.objects.all().order_by("id")
    serializer = Articulos_Ordenados_Serializer(articulos, many=True)
    return Response(serializer.data)

@api_view(["GET"])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def productos_surtir_api(request):
    productos_surtir = ArticulosparaSurtir.objects.all().order_by("id")
    serializer = Articulos_para_Surtir_Serializer(productos_surtir, many=True)
    return Response(serializer.data)


@api_view(["GET"])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def requisiciones_api(request):
    requisiciones = Requis.objects.filter(complete=True).order_by("id")
    serializer = RequisicionSerializer(requisiciones, many=True)
    return Response(serializer.data)

@api_view(["GET"])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def productos_requisitados_api(request):
    productos_requisitados = ArticulosRequisitados.objects.all().order_by("id")
    serializer = Articulos_Requisitados_Serializer(productos_requisitados, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def Compra_tabla_api(request):

      # Trae TODO pero lo paginamos
    compras_qs = Compra.objects.filter(complete=True).order_by('id')

    page = int(request.query_params.get('page', 1))
    per_page = int(request.query_params.get('per_page', 1000))  # ajustable

    paginator = Paginator(compras_qs, per_page=per_page)

    try:
        compras_page = paginator.page(number=page)
    except EmptyPage:
        # Sin más datos -> lista vacía
        return Response([])

    serialized_compras = Compra_tabla_Serializer(compras_page, many=True)
    #serialized_compras = Compra_tabla_Serializer(compras_qs, many=True)
        
    return Response(serialized_compras.data)



@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def CompraAPI(request):
    compras_qs = (
        Compra.objects
        .filter(complete=True)
        .select_related(
            'req',
            'req__orden',
            'req__orden__distrito',
            'req__orden__proyecto',
            'req__orden__subproyecto',
            'proveedor',
            'proveedor__nombre',
            'moneda',
        )
        .order_by('-folio')
    )

    # filtros básicos opcionales
    folio = request.query_params.get('folio')
    proveedor = request.query_params.get('proveedor')
    distrito = request.query_params.get('distrito')
    proyecto = request.query_params.get('proyecto')
    fecha_desde = request.query_params.get('fecha_desde')
    fecha_hasta = request.query_params.get('fecha_hasta')

    if folio:
        compras_qs = compras_qs.filter(folio__icontains=folio)

    if proveedor:
        compras_qs = compras_qs.filter(proveedor__nombre__razon_social__icontains=proveedor)

    if distrito:
        compras_qs = compras_qs.filter(req__orden__distrito__nombre__icontains=distrito)

    if proyecto:
        compras_qs = compras_qs.filter(req__orden__proyecto__nombre__icontains=proyecto)

    if fecha_desde:
        compras_qs = compras_qs.filter(created_at__date__gte=fecha_desde)

    if fecha_hasta:
        compras_qs = compras_qs.filter(created_at__date__lte=fecha_hasta)

    page = int(request.query_params.get('page', 1))
    per_page = int(request.query_params.get('per_page', 50))

    paginator = Paginator(compras_qs, per_page=per_page)

    try:
        compras_page = paginator.page(number=page)
    except EmptyPage:
        return Response({
            "results": [],
            "page": page,
            "per_page": per_page,
            "total": paginator.count,
            "total_pages": paginator.num_pages,
            "has_next": False,
            "has_previous": False,
        })

    serializer = CompraTablaLiteSerializer(compras_page.object_list, many=True)

    return Response({
        "results": serializer.data,
        "page": page,
        "per_page": per_page,
        "total": paginator.count,
        "total_pages": paginator.num_pages,
        "has_next": compras_page.has_next(),
        "has_previous": compras_page.has_previous(),
    })


@api_view(['GET'])
@authentication_classes([SessionAuthentication])
@permission_classes([IsAuthenticated])
def compras_resumen_api(request):
    qs = (
        Compra.objects
        .filter(
            complete=True,
            solo_servicios=True,
        )
        .exclude(req__orden__distrito__nombre__in=[
            'BRASIL',
            'ALTAMIRA ALTERNATIVO',
            'VH SECTOR 6',
        ])
        .exclude(autorizado1=False)
        .exclude(autorizado2=False)
    )

    distrito = request.query_params.get('distrito')
    proveedor = request.query_params.get('proveedor')
    proyecto = request.query_params.get('proyecto')
    subproyecto = request.query_params.get('subproyecto')
    anio = request.query_params.get('anio')
    mes = request.query_params.get('mes')
    pagada = request.query_params.get('pagada')
    fecha_desde = request.query_params.get('fecha_desde')
    fecha_hasta = request.query_params.get('fecha_hasta')

    if distrito:
        qs = qs.filter(req__orden__distrito__nombre=distrito)

    if proveedor:
        qs = qs.filter(proveedor__nombre__razon_social__icontains=proveedor)

    if proyecto:
        qs = qs.filter(req__orden__proyecto__nombre__icontains=proyecto)

    if subproyecto:
        qs = qs.filter(req__orden__subproyecto__nombre__icontains=subproyecto)

    if anio:
        qs = qs.filter(created_at__year=anio)

    if mes:
        qs = qs.filter(created_at__month=mes)

    if pagada in ['true', 'false']:
        qs = qs.filter(pagada=(pagada == 'true'))

    if fecha_desde:
        qs = qs.filter(created_at__date__gte=fecha_desde)

    if fecha_hasta:
        qs = qs.filter(created_at__date__lte=fecha_hasta)

    proveedores = list(
    qs.order_by()
    .values(
        proveedor_master_id=F('proveedor__nombre_id'),
        proveedor_nombre=F('proveedor__nombre__razon_social'),
    )
    .annotate(
        total_compras=Count('id'),
        monto_total=Coalesce(
            Sum('costo_oc'),
            Value(Decimal('0.00')),
            output_field=DecimalField(max_digits=14, decimal_places=2)
        ),
        monto_pagado_total=Coalesce(
            Sum('monto_pagado'),
            Value(Decimal('0.00')),
            output_field=DecimalField(max_digits=14, decimal_places=2)
        ),
        compras_pagadas=Count('id', filter=Q(pagada=True)),
        compras_no_pagadas=Count('id', filter=Q(pagada=False)),
    )
    .order_by('-monto_total')
)

    productos_por_proveedor = list(
        ArticuloComprado.objects
        .filter(oc__in=qs)
        .order_by()
        .values(
            proveedor_fk=F('oc__proveedor__nombre_id'),
            producto_nombre=Coalesce(
                F('producto__producto__articulos__producto__producto__nombre'),
                Value('SIN DESCRIPCIÓN'),
                output_field=CharField()
            ),
        )
        .annotate(
            total_ocs=Count('oc', distinct=True),
            cantidad_total=Coalesce(
                Sum('cantidad'),
                Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=14, decimal_places=2)
            ),
            monto_total=Coalesce(
                Sum(
                    ExpressionWrapper(
                        F('cantidad') * F('precio_unitario'),
                        output_field=DecimalField(max_digits=20, decimal_places=2)
                    )
                ),
                Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=20, decimal_places=2)
            )
        )
        .order_by('proveedor_fk', '-monto_total')
    )

    mapa_productos = {}
    for item in productos_por_proveedor:
        proveedor_id = item["proveedor_fk"]
        mapa_productos.setdefault(proveedor_id, []).append({
            "producto": item["producto_nombre"],
            "total_ocs": item["total_ocs"],
            "cantidad_total": item["cantidad_total"],
            "monto_total": item["monto_total"],
        })

    resultado = []
    for prov in proveedores:
        resultado.append({
            "proveedor_id": prov["proveedor_master_id"],
            "proveedor": prov["proveedor_nombre"],
            "total_compras": prov["total_compras"],
            "monto_total": prov["monto_total"],
            "monto_pagado_total": prov["monto_pagado_total"],
            "compras_pagadas": prov["compras_pagadas"],
            "compras_no_pagadas": prov["compras_no_pagadas"],
            "productos": mapa_productos.get(prov["proveedor_master_id"], []),
        })
    
    return Response(resultado)

from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook

@api_view(['GET'])
@authentication_classes([SessionAuthentication])
@permission_classes([IsAuthenticated])
def compras_resumen_excel(request):
    print("entra excel")

    data = construir_compras_resumen(request)
    print("data construida:", len(data))

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Compras"

    # 🎨 ESTILOS
    title_fill = PatternFill(fill_type="solid", fgColor="163A5F")
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    header_font = Font(bold=True)
    body_font = Font(size=11)

    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # 🧾 TÍTULO
    ws.merge_cells("A1:E1")
    ws["A1"] = "Resumen de Compras"
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # 📌 HEADERS
    headers = [
        "Proveedor",
        "Servicio",
        "Total OCs",
        "Cantidad Total",
        "Monto Total",
    ]

    header_row = 3

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # 📊 DATOS
    current_row = header_row + 1

    for prov in data:
        proveedor = prov["proveedor"]

        for prod in prov["productos"]:
            ws.cell(row=current_row, column=1, value=proveedor)
            ws.cell(row=current_row, column=2, value=prod["producto"])
            ws.cell(row=current_row, column=3, value=prod["total_ocs"])
            ws.cell(row=current_row, column=4, value=float(prod["cantidad_total"]))
            ws.cell(row=current_row, column=5, value=float(prod["monto_total"]))

            for col in range(1, 6):
                cell = ws.cell(row=current_row, column=col)
                cell.font = body_font
                cell.border = border

                if col in [3, 4, 5]:
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")

            current_row += 1

    # 💰 FORMATO NUMÉRICO
    for row in range(header_row + 1, current_row):
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        ws.cell(row=row, column=5).number_format = '$#,##0.00'

    # 🔍 AUTOFILTRO
    ws.auto_filter.ref = f"A{header_row}:E{current_row - 1}"

    # ❄️ CONGELAR ENCABEZADO
    ws.freeze_panes = "A4"

    # 📏 ANCHOS DE COLUMNA
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18

    # ALTURA
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[3].height = 22

    # 💾 GUARDAR
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    contenido = output.getvalue()
    print("bytes excel:", len(contenido))

    response = HttpResponse(
        contenido,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="resumen_compras.xlsx"'
    response["Content-Length"] = str(len(contenido))

    print("response listo")
    return response


#@api_view(['GET'])
#@permission_classes([IsAuthenticated])
def construir_compras_resumen(request):
    print("1. entra helper")

    qs = (
        Compra.objects
        .filter(
            complete=True,
            solo_servicios=True,
        )
        .exclude(req__orden__distrito__nombre__in=[
            'BRASIL',
            'ALTAMIRA ALTERNATIVO',
            'VH SECTOR 6',
        ])
        .exclude(autorizado1=False)
        .exclude(autorizado2=False)
    )

    print("2. queryset base armado")

    distrito = request.query_params.get('distrito')
    proveedor = request.query_params.get('proveedor')
    proyecto = request.query_params.get('proyecto')
    subproyecto = request.query_params.get('subproyecto')
    anio = request.query_params.get('anio')
    mes = request.query_params.get('mes')
    pagada = request.query_params.get('pagada')
    fecha_desde = request.query_params.get('fecha_desde')
    fecha_hasta = request.query_params.get('fecha_hasta')

    print("3. params leídos")

    if distrito:
        qs = qs.filter(req__orden__distrito__nombre=distrito)
    if proveedor:
        qs = qs.filter(proveedor__nombre__razon_social__icontains=proveedor)
    if proyecto:
        qs = qs.filter(req__orden__proyecto__nombre__icontains=proyecto)
    if subproyecto:
        qs = qs.filter(req__orden__subproyecto__nombre__icontains=subproyecto)
    if anio:
        qs = qs.filter(created_at__year=anio)
    if mes:
        qs = qs.filter(created_at__month=mes)
    if pagada in ['true', 'false']:
        qs = qs.filter(pagada=(pagada == 'true'))
    if fecha_desde:
        qs = qs.filter(created_at__date__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(created_at__date__lte=fecha_hasta)

    print("4. filtros aplicados")

    print("5. count qs:", qs.count())

    proveedores = list(
        qs.order_by()
        .values(
            proveedor_master_id=F('proveedor__nombre_id'),
            proveedor_nombre=F('proveedor__nombre__razon_social'),
        )
        .annotate(
            total_compras=Count('id'),
            monto_total=Coalesce(
                Sum('costo_oc'),
                Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=14, decimal_places=2)
            ),
            monto_pagado_total=Coalesce(
                Sum('monto_pagado'),
                Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=14, decimal_places=2)
            ),
            compras_pagadas=Count('id', filter=Q(pagada=True)),
            compras_no_pagadas=Count('id', filter=Q(pagada=False)),
        )
        .order_by('-monto_total')
    )

    print("6. proveedores listos:", len(proveedores))

    productos_por_proveedor = list(
        ArticuloComprado.objects
        .filter(oc__in=qs)
        .order_by()
        .values(
            proveedor_fk=F('oc__proveedor__nombre_id'),
            producto_nombre=Coalesce(
                F('producto__producto__articulos__producto__producto__nombre'),
                Value('SIN DESCRIPCIÓN'),
                output_field=CharField()
            ),
        )
        .annotate(
            total_ocs=Count('oc', distinct=True),
            cantidad_total=Coalesce(
                Sum('cantidad'),
                Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=14, decimal_places=2)
            ),
            monto_total=Coalesce(
                Sum(
                    ExpressionWrapper(
                        F('cantidad') * F('precio_unitario'),
                        output_field=DecimalField(max_digits=20, decimal_places=2)
                    )
                ),
                Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=20, decimal_places=2)
            )
        )
        .order_by('proveedor_fk', '-monto_total')
    )

    print("7. productos listos:", len(productos_por_proveedor))

    mapa_productos = {}
    for item in productos_por_proveedor:
        proveedor_id = item["proveedor_fk"]
        mapa_productos.setdefault(proveedor_id, []).append({
            "producto": item["producto_nombre"],
            "total_ocs": item["total_ocs"],
            "cantidad_total": item["cantidad_total"],
            "monto_total": item["monto_total"],
        })

    print("8. mapa listo")

    resultado = []
    for prov in proveedores:
        resultado.append({
            "proveedor_id": prov["proveedor_master_id"],
            "proveedor": prov["proveedor_nombre"],
            "total_compras": prov["total_compras"],
            "monto_total": prov["monto_total"],
            "monto_pagado_total": prov["monto_pagado_total"],
            "compras_pagadas": prov["compras_pagadas"],
            "compras_no_pagadas": prov["compras_no_pagadas"],
            "productos": mapa_productos.get(prov["proveedor_master_id"], []),
        })

    print("9. resultado listo:", len(resultado))
    return resultado


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def compras_resumen_chart(request):
    qs = (
        Compra.objects
        .filter(
            complete=True,
            solo_servicios=True,
        )
        .exclude(req__orden__distrito__nombre__in=[
            'BRASIL',
            'ALTAMIRA ALTERNATIVO',
            'VH SECTOR 6',
        ])
        .exclude(autorizado1=False)
        .exclude(autorizado2=False)
    )

    distrito = request.query_params.get('distrito')
    proveedor = request.query_params.get('proveedor')
    proyecto = request.query_params.get('proyecto')
    subproyecto = request.query_params.get('subproyecto')
    anio = request.query_params.get('anio')
    mes = request.query_params.get('mes')
    pagada = request.query_params.get('pagada')
    fecha_desde = request.query_params.get('fecha_desde')
    fecha_hasta = request.query_params.get('fecha_hasta')
    top = int(request.query_params.get('top', 10))

    if distrito:
        qs = qs.filter(req__orden__distrito__nombre=distrito)

    if proveedor:
        qs = qs.filter(proveedor__nombre__razon_social__icontains=proveedor)

    if proyecto:
        qs = qs.filter(req__orden__proyecto__nombre__icontains=proyecto)

    if subproyecto:
        qs = qs.filter(req__orden__subproyecto__nombre__icontains=subproyecto)

    if anio:
        qs = qs.filter(created_at__year=anio)

    if mes:
        qs = qs.filter(created_at__month=mes)

    if pagada in ['true', 'false']:
        qs = qs.filter(pagada=(pagada == 'true'))

    if fecha_desde:
        qs = qs.filter(created_at__date__gte=fecha_desde)

    if fecha_hasta:
        qs = qs.filter(created_at__date__lte=fecha_hasta)

    data = list(
        qs.values(
            label=F('proveedor__nombre__razon_social')
        )
        .annotate(
            monto_total=Coalesce(
                Sum('costo_oc'),
                Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=14, decimal_places=2)
            )
        )
        .order_by('-monto_total')[:top]
    )

    return Response(data)

@api_view(['GET'])
@authentication_classes([SessionAuthentication])
@permission_classes([IsAuthenticated])
def compras_resumen_chart_proveedores(request):
    qs = (
        Compra.objects
        .filter(
            complete=True,
            solo_servicios=True,
        )
        .exclude(req__orden__distrito__nombre__in=[
            'BRASIL',
            'ALTAMIRA ALTERNATIVO',
            'VH SECTOR 6',
        ])
        .exclude(autorizado1=False)
        .exclude(autorizado2=False)
    )

    distrito = request.query_params.get('distrito')
    proveedor = request.query_params.get('proveedor')
    proyecto = request.query_params.get('proyecto')
    subproyecto = request.query_params.get('subproyecto')
    anio = request.query_params.get('anio')
    mes = request.query_params.get('mes')
    pagada = request.query_params.get('pagada')
    fecha_desde = request.query_params.get('fecha_desde')
    fecha_hasta = request.query_params.get('fecha_hasta')
    top = int(request.query_params.get('top', 10))

    if distrito:
        qs = qs.filter(req__orden__distrito__nombre=distrito)

    if proveedor:
        qs = qs.filter(proveedor__nombre__razon_social__icontains=proveedor)

    if proyecto:
        qs = qs.filter(req__orden__proyecto__nombre__icontains=proyecto)

    if subproyecto:
        qs = qs.filter(req__orden__subproyecto__nombre__icontains=subproyecto)

    if anio:
        qs = qs.filter(created_at__year=anio)

    if mes:
        qs = qs.filter(created_at__month=mes)

    if pagada in ['true', 'false']:
        qs = qs.filter(pagada=(pagada == 'true'))

    if fecha_desde:
        qs = qs.filter(created_at__date__gte=fecha_desde)

    if fecha_hasta:
        qs = qs.filter(created_at__date__lte=fecha_hasta)

    data = list(
        qs.values(
            label=F('proveedor__nombre__razon_social')
        )
        .annotate(
            monto_total=Coalesce(
                Sum('costo_oc'),
                Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=14, decimal_places=2)
            )
        )
        .order_by('-monto_total')[:top]
    )

    return Response(data)

@api_view(["GET"])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def productos_comprados_api(request):
    productos_comprados = ArticuloComprado.objects.all().order_by("id")
    serializer = Articulo_Comprado_Serializer(productos_comprados, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@authentication_classes([SessionAuthentication,TokenAuthentication])
@permission_classes([IsAuthenticated])
def proveedor_direccion_api(request):
    #registra el acceso a la api
    #print(f"Usuario autenticado: {request.user}")
    #user = request.user
    #ip_address = request.META.get('REMOTE_ADDR')
    #logger.info(f"GET {request.path} by {user.first_name} {user.last_name} from {ip_address}")
    
    proveedores = Proveedor_direcciones.objects.filter(completo = True)
    #page = request.query_params.get('page', 1)
    #per_page = request.query_params.get('per_page', 20)
    #
    #ordering = request.query_params.get('ordering')

    #if ordering:
    #    proveedores = Proveedor_direcciones.order_by(ordering)
        
    #paginator = Paginator(proveedores, per_page=per_page)
    #try: 
    #    proveedores = paginator.page(number=page)
    #except EmptyPage:
    #    proveedores = []
    serialized_proveedores = ProveedorDireccionesSerializer(proveedores, many=True)
        
    return Response(serialized_proveedores.data)

@api_view(['GET'])
@authentication_classes([SessionAuthentication,TokenAuthentication])
@permission_classes([IsAuthenticated])
def proveedor_api(request):
    #registra el acceso a la api
    #print(f"Usuario autenticado: {request.user}")
    #user = request.user
    #ip_address = request.META.get('REMOTE_ADDR')
    #logger.info(f"GET {request.path} by {user.first_name} {user.last_name} from {ip_address}")
    
    proveedores = Proveedor.objects.filter(completo = True)
    #page = request.query_params.get('page', 1)
    #per_page = request.query_params.get('per_page', 20)
    #
    #ordering = request.query_params.get('ordering')

    #if ordering:
    #    proveedores = Proveedor.order_by(ordering)
        
    #paginator = Paginator(proveedores, per_page=per_page)
    #try: 
    #    proveedores = paginator.page(number=page)
    #except EmptyPage:
    #    proveedores = []
    serialized_proveedores = ProveedorSerializer(proveedores, many=True)
        
    return Response(serialized_proveedores.data)

@api_view(['GET'])
@authentication_classes([SessionAuthentication])
@permission_classes([IsAuthenticated])
def distritos_api(request):
    #registra el acceso a la api
    #print(f"Usuario autenticado: {request.user}")
    #user = request.user
    #ip_address = request.META.get('REMOTE_ADDR')
    #logger.info(f"GET {request.path} by {user.first_name} {user.last_name} from {ip_address}")
    
    distritos = Distrito.objects.filter(status = True).exclude(nombre__in=['BRASIL','ALTAMIRA ALTERNATIVO','MATRIZ ALTERNATIVO','VH SECTOR 6'])
    #page = request.query_params.get('page', 1)
    #per_page = request.query_params.get('per_page', 20)
    #
    #ordering = request.query_params.get('ordering')

    #if ordering:
    #    distritos = Distrito.order_by(ordering)
        
    #paginator = Paginator(distritos, per_page=per_page)
    #try: 
    #    distritos = paginator.page(number=page)
    #except EmptyPage:
    #    distritos = []
    serialized_distritos = DistritoSerializer(distritos, many=True)
        
    return Response(serialized_distritos.data)



#url = 'https://vordcab.cloud/apiapp/perfiles/'
#token = 'defa1b040b2e8acf4d9ab20127e87d820eb913b9'
@perfil_seleccionado_required
def obtener_perfiles(request):
    actualizado = False
    empleados_actualizados = []  # Lista para almacenar los usuarios actualizados

    if request.method == 'POST':
        actualizado = True  # Actualizar mensaje en template
        #url = 'http://127.0.0.1:9000/apiapp/perfiles/'
        #token = 'f36cf2df116c3aeab68b9ee948331f382f5edcc0'
        url = 'https://vordcab.cloud/apiapp/perfiles/'
        token = 'defa1b040b2e8acf4d9ab20127e87d820eb913b9'
        headers = {
            'Authorization': f'Token {token}'
        }

        # Hacer la solicitud a la API con los encabezados
        response = requests.get(url, headers=headers)
        
        if response.status_code == 200:
            # Procesar el JSON de la respuesta
            data = response.json()

            # Iterar sobre cada perfil recibido en el JSON
            for perfil in data:
                correo_vordcab = perfil.get('correo_vordcab')
                baja = perfil.get('baja', False)  # Por defecto, es False si no existe
                nivel_data = perfil.get('nivel')  # Obtener datos del nivel
                empresa = perfil.get('empresa', {}).get('empresa')

                if empresa == 'VORDCAB':
                    empresa = 'Grupo Vordcab S.A. de C.V.'

                # Extraer el nivel del JSON, si existe
                nivel = nivel_data['nivel']['nivel'] if nivel_data and nivel_data.get('nivel') else None

                try:
                    # Buscar al usuario por su correo electrónico
                    usuario = User.objects.get(email=correo_vordcab)
                    custom_user = CustomUser.objects.get(staff=usuario)  # Obtener el CustomUser relacionado

                    # Determinar si se necesita una actualización
                    estado_anterior = usuario.is_active
                    nivel_anterior = custom_user.nivel
                    empresa_anterior = custom_user.empresa.nombre if custom_user.empresa else None

                    cambio = False  # variable que indica si hubo un cambio

                    # Actualizar el estado si ha cambiado
                    if baja:
                        if usuario.is_active:
                            usuario.is_active = False
                            cambio = True
                    else:
                        if not usuario.is_active:
                            usuario.is_active = True
                            cambio = True

                    # Convertir nivel a float si no es None y ha cambiado
                    if nivel is not None and custom_user.nivel != float(nivel):
                        custom_user.nivel = float(nivel)
                        cambio = True

                    # Compara la empresa actual con la empresa del JSON
                    if empresa and (empresa_anterior is None or empresa_anterior.strip().lower() != empresa.strip().lower()):
                        try:
                            nueva_empresa = Empresa.objects.get(nombre=empresa)
                            custom_user.empresa = nueva_empresa
                            cambio = True
                        except Empresa.DoesNotExist:
                            print(f"Empresa con nombre {empresa} no encontrada.")

                    # Si hubo un cambio, guardar el usuario y agregarlo a la lista de actualizados
                    if cambio:
                        usuario.save()
                        custom_user.save()
                        empleados_actualizados.append({
                            'nombre': usuario.get_full_name(),
                            'correo': usuario.email,
                            'activo': usuario.is_active,
                            'activo_anterior': estado_anterior,
                            'nivel': custom_user.nivel,
                            'nivel_anterior': nivel_anterior,
                            'empresa': custom_user.empresa.nombre if custom_user.empresa else None,
                            'empresa_anterior': empresa_anterior,
                        })
                
                except User.DoesNotExist:
                    # Si no existe el usuario con ese correo
                    print(f"Usuario con correo {correo_vordcab} no encontrado.")

    return render(request, 'api/perfiles_lista.html', {'empleados_actualizados': empleados_actualizados, 'actualizado': actualizado})

@api_view(['GET'])
@authentication_classes([TokenAuthentication]) #Si quieres decargar el pdf desde el nodo desabilita este decorador de token
@permission_classes([IsAuthenticated])
def descargar_pdf_oc(request, pk):
    user = request.user
    ip_address = request.META.get('REMOTE_ADDR')
    logger.info(f"GET {request.path} by {user.first_name} {user.last_name} from {ip_address}")


    try:
        # Intentar obtener la orden de compra por su id
        compra = Compra.objects.get(id=pk)
    except Compra.DoesNotExist:
        # Si no existe la OC, devolver un mensaje de éxito con estado 200 ya que si uso 404 me manda al middleware
        return Response(
            {"detail": "La OC que intenta traer no existe, pero la solicitud fue procesada correctamente."},
            status=status.HTTP_200_OK
        )

    # Generar el PDF si la OC existe
    buf = generar_pdf_nueva(compra)

    # Devolver el PDF como respuesta
    return FileResponse(buf, as_attachment=True, filename='oc_' + str(compra.folio) + '.pdf')

@perfil_seleccionado_required
def tabla_festivos(request):
    datos = TablaFestivos.objects.all()
    
    if request.method == 'POST':
        url = 'https://vordcab.cloud/apiapp/festivos_actual/'
        api_key = 'defa1b040b2e8acf4d9ab20127e87d820eb913b9'
        #url = 'http://127.0.0.1:9000/apiapp/festivos_actual/'
        #api_key = 'f36cf2df116c3aeab68b9ee948331f382f5edcc0'
        # Hacer la solicitud a tu API de festivos
        headers = {
            'Authorization': f'Token {api_key}',
        }
        response = requests.get(url, headers=headers)
        
        if response.status_code == 200:
            # Procesar el JSON de la respuesta
            data = response.json()
            
            # Recorrer la lista de días festivos recibidos
            for holiday in data:
                holiday_date_str = holiday.get('dia_festivo')  # Obtener la fecha en formato YYYY-MM-DD
                if holiday_date_str:
                    # Crear o actualizar el objeto en la base de datos directamente con el formato YYYY-MM-DD
                    TablaFestivos.objects.update_or_create(dia_festivo=holiday_date_str)
            
            messages.success(request, 'Has actualizado correctamente los días festivos')
            return redirect('tabla_festivos')
        else:
            messages.error(request, 'Ocurrió un error al recibir la respuesta de la API')
    
    context = {
        'datos': datos,
    }

    return render(request, 'api/tabla_festivos.html', context)

def chatbot_view2(request):
    #print('chatbot')
    if request.method == "POST":
        #print('chatbot_post')
        user_message = request.POST.get("message", "").lower()
      
        print(user_message)

        #procesos = {
        #    "solicitud": {
        #        "pasos": (
        #           "Para hacer una solicitud en 'SAVIA 2.0', sigue estos pasos:\n"
        #            "1. Accede al módulo de **Solicitudes** desde el menú principal.\n"
        #            "2. Haz clic en el botón **Nueva Solicitud**.\n"
        #            "3. Llena el formulario con los datos requeridos.\n"
        #            "4. Adjunta los documentos necesarios.\n"
        #            "5. Haz clic en **Enviar** para guardar y enviar la solicitud."
        #        ),
        #        "video": "https://www.ejemplo.com/tutorial-solicitudes"
        #    },
        #    "reporte": {
        #        "pasos": (
        #            "Para generar un reporte:\n"
        #            "1. Ve al módulo de **Reportes**.\n"
        #            "2. Selecciona el tipo de reporte.\n"
        #            "3. Define el rango de fechas.\n"
        #            "4. Haz clic en **Generar** para descargar el reporte."
        #        ),
        #        "video": "https://www.ejemplo.com/tutorial-reportes"
        #    }
        #}

        #if "solicitud" in user_message:
        #    bot_reply = f"{procesos['solicitud']['pasos']}\n\nVideo: {procesos['solicitud']['video']}"
        #elif "reporte" in user_message:
        #    bot_reply = f"{procesos['reporte']['pasos']}\n\nVideo: {procesos['reporte']['video']}"
        #else:
        sql_generation_response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role":"system",
                    "content":(
                        "Eres un experto en MySQL y conoces la estructura de la base de las tablas de SAVIA2 donde la tabla order(solicitud) está relacionada"
                        "utiliza la información dada para construir consultas SQL basada en preguntas en lenguaje natural"
                        "La relación entre tablas es dashboard_product(id) << dashboard_inventario(producto_id) << dashboard_articulosordenados(producto_id) << dashboard_order(alias: solicitud)"
                        "<< requisiciones_requis << compras_compra << entradas_entrada -- Todas los campos distrito están ligados a una tabla llamada user_distrito"
                        "En la tabla user_distrito la columna nombre es la variable del nombre del distrito.Distrito está ligado a las tablas dashboard_inventario y dashboard_order"
                        "En la tabla dashboard_inventario el importe_producto = (cantidad + cantidad_apartada) * price"
                        "En la tabla dashboard_inventario el valor_inventario =  SUM(importe_producto)"
                        "Solo genera la consulta SQL, no incluyas texto adicional. Solo SQL sintaxis por favor"
                    )
                },
                {
                    "role": "user", 
                    "content": f"Genera una consulta SQL para: {user_message}"
                }
            ],
            temperature=0,
            max_tokens=200
        )
        print(sql_generation_response)
        sql_query = sql_generation_response.choices[0].message.content
        sql_query = sql_query.replace("sql","").replace("```sql", "").replace("```", "").strip()
        print(f"Consulta SQL generada: {sql_query}")
        conn_savia2 = mysql.connector.connect(
            host='localhost', 
            user='root', 
            password='peruzzi25', 
            database='savia2'
        )
        db_cursor = conn_savia2.cursor()
            
        db_cursor.execute(sql_query)
        result = db_cursor.fetchall()
        conn_savia2.close()
        # Convertir el resultado en un formato más legible
        result_text = f"El resultado de tu consulta es: {result}"
        #bot_reply = response.choices[0].message.content
        #print(bot_reply)
                #except Exception as e:
                #    bot_reply = "I'm sorry, there was an error processing your request."

        natural_language_response =client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "system", "content": (
                    "Eres un asistente que convierte resultados de consultas SQL en respuestas naturales para los usuarios."
                )
                },
                {"role": "user", "content": f"El resultado del query es: {result}"}
                ],
                temperature=0.3,
                max_tokens=180
            )
        bot_reply = natural_language_response.choices[0].message.content
        print(bot_reply)
        return JsonResponse({"response": bot_reply})
    

def chatbot_view(request):
    #print('chatbot')
    if request.method == "POST":
        #print('chatbot_post')
        user_message = request.POST.get("message", "").lower()
      
        print(user_message)
        if 'status' and 'solicitud' and 'folio' and 'distrito' in user_message:
            #folio_number = 
            
            status_solicitud()
        else:
            orm_generation_response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {
                        "role":"system",
                        "content":(
                            "Eres un experto en Django ORM. Conoces la estructura de los modelos en un proyecto Django."
                            "Tu tarea es generar consultas utilizando el ORM de Django"
                            "Si algún modelo o relación es necesario para construir la consulta, genera las clases de modelos correspondientes."
                            "La relación entre modelos es: "
                            "dashboard.models Product(id) -> dashboard.models Inventario(FK producto_id, fields: cantidad_apartada, cantidad, price, distrito) -> dashboard.models ArticuloOrdenado(FK producto_id) -> dashboard.Order(alias: Solicitud, FK distrito_id, fields folio) -> requisiciones.models Requis(id, FK orden_id) -> compras.models Compra(id, FK req_id)"
                            "el modelo user.Distrito está relacionado con dashboard.models.Inventario y dashboard.models.Order a través de una FK llamada distrito_id."
                            "En Inventario, el valor del producto se calcula como: importe_producto = (cantidad + cantidad_apartada) * price."
                            "El valor total del inventario por distrito se calcula como: SUM(importe_producto)."
                            "Cuando se proporciona el folio de una solicitud el estatus se refiere a interpretar lo siguiente"
                            "### Reglas para Interpretar el Estado de las Solicitudes:"
                            "1. Si `order.autorizar` es `None`, describe: 'La solicitud con folio {order.folio} no ha sido autorizada aún.'"
                            "2. Si `order.autorizar` es `False`, describe: 'La solicitud con folio {order.folio} está cancelada.'"
                            "3. Si `order.autorizar` es `True`, describe: 'La solicitud con folio {order.folio} ha sido autorizada.'"
                            "\n\n"
                            "1. Si `order.requis.exists()` es `True`, recorre todas las requisiciones con un ciclo `for`."
                            "2. Recuerda que el modelo requi proviene de requisiciones.models Para cada requisición (`requi`) en `order.requis.all()`:"
                            "    a. Si `requi.autorizar` es `None`, describe: 'La requisición con folio {requi.folio} no ha sido autorizada aún.'"
                            "    b. Si `requi.autorizar` es `False`, describe: 'La requisición con folio {requi.folio} está cancelada.'"
                            "    c. Si `requi.autorizar` es `True`, describe: 'La requisición con folio {requi.folio} ha sido autorizada.'"
                            "\n\n"
                            "### Respuesta Esperada:"
                            "Siempre responde en formato de código Python utilizando el ORM de Django, los modelos ya existen solo hay que importarlos para hacer la consulta"
                            "Tu respuesta debe incluir todas las importaciones necesarias para que el código funcione correctamente. No incluyas comentarios explicativos solo código funcional"
                            "Por ejemplo, asegúrate de importar desde 'django.db.models' funciones como 'F', 'Sum', 'Count', 'Value', 'Case', etc."
                            "Siempre asigna el resultado a una variable llamada 'resultado'."
                            #"Puedes intentar desarrollar el código de un gráfico sí el usuario así lo solicita el usuario utilizando plotly y es posible"
                        )
                    },
                    {
                        "role": "user", 
                        "content": f"Genera una consulta SQL para: {user_message}"
                    }
                ],
                temperature=0,
                max_tokens=100
            )
            print(orm_generation_response)
            orm_query = orm_generation_response.choices[0].message.content
            print(orm_query)
            orm_query = orm_query.replace("```python","").replace("```","").strip()
            print(f"Consulta ORM generada: {orm_query}")
            # Paso 2: Ejecutar el código dinámicamente
            local_variables = {}
            exec(orm_query, globals(), local_variables)
            print(local_variables)
            result = local_variables.get('resultado', None)

        natural_language_response =client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "system", "content": (
                        "Eres un asistente que convierte resultados de consultas del ORM de Django en respuestas naturales para los usuarios"
                        "En caso de ser necesario en la respuesta se prefiere el uso de nombre que de ID's"
                        "Reproducir los status de manera completa de acuerdo al resultado"
                        "Dar formato de tabla atractivo en la medida que aplique en html"
                        #"si el campo se refiere a cantidad no es dinero por lo tanto no se formatea, si la respuesta incluye un monto o valor que se infiera que es dinero incluir formato de pesos"
                    )
                    },
                    {
                        "role": "user", "content": f"El resultado del query es: {result}"
                    }
            ],
                    temperature=0.3,
                    max_tokens=200
                )
        bot_reply = natural_language_response.choices[0].message.content
        print(bot_reply)
        return JsonResponse({"response": bot_reply})
    
def status_solicitud(folio, distrito):
    order = Order.objects.filter(folio=folio, distrito__nombre=distrito).first()

    if order:
        if order.autorizar is None:
            resultado = f'La solicitud con folio {order.folio} no ha sido autorizada aún.'
        elif order.autorizar is False:
            resultado = f'La solicitud con folio {order.folio} está cancelada.'
        elif order.autorizar is True:
            resultado = f'La solicitud con folio {order.folio} ha sido autorizada.'

            requisiciones = order.requis.all()
            for requi in requisiciones:
                if requi.autorizar is None:
                    resultado += f'\nLa requisición con folio {requi.folio} no ha sido autorizada aún.'
                elif requi.autorizar is False:
                    resultado += f'\nLa requisición con folio {requi.folio} está cancelada.'
                elif requi.autorizar is True:
                    resultado += f'\nLa requisición con folio {requi.folio} ha sido autorizada.'
    else:
        resultado = f'No se encontró la solicitud con folio {folio} en el distrito {distrito}.'


def obtener_pagos_control_bancos(cuenta):
    cuenta_saldos = Saldo_Cuenta.objects.filter(cuenta=cuenta).order_by('-fecha_inicial')
    ultimo_saldo = cuenta_saldos.filter(hecho=True).first() if cuenta_saldos.exists() else None

    if ultimo_saldo is not None:
        fecha_saldo = ultimo_saldo.fecha_inicial
        pagos = Pago.objects.filter(
            cuenta=cuenta,
            eliminado=False,
            hecho=True,
            pagado_real__gte=fecha_saldo
        ).order_by('pagado_real', 'pagado_hora', '-tipo__id')
    else:
        pagos = Pago.objects.filter(
            cuenta=cuenta,
            hecho=True,
            eliminado=False
        ).order_by('pagado_real', 'pagado_hora', '-tipo__id')

    return pagos, ultimo_saldo

def calcular_saldos_control_bancos(cuenta, pagos, start_date=None, end_date=None):
    saldo_obj = (
        Saldo_Cuenta.objects
        .filter(cuenta=cuenta, hecho=True)
        .order_by('-fecha_inicial')
        .first()
    )

    saldo_base = (saldo_obj.monto_inicial if saldo_obj else Decimal('0.00')) or Decimal('0.00')
    fecha_base = saldo_obj.fecha_inicial if saldo_obj else None

    if isinstance(start_date, str) and start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()

    if isinstance(end_date, str) and end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    if fecha_base and start_date and start_date > fecha_base:
        intermedios = pagos.filter(
            pagado_real__gte=fecha_base,
            pagado_real__lt=start_date
        ).aggregate(
            cargos=Coalesce(
                Sum('monto', filter=Q(tipo__isnull=True) | Q(tipo__nombre='CARGO')),
                Decimal('0.00')
            ),
            abonos=Coalesce(
                Sum('monto', filter=Q(tipo__isnull=False) & ~Q(tipo__nombre='CARGO')),
                Decimal('0.00')
            ),
        )
        saldo_trasladado = saldo_base - intermedios['cargos'] + intermedios['abonos']
        inicio_periodo = start_date
    else:
        saldo_trasladado = saldo_base
        inicio_periodo = fecha_base if fecha_base else None

    fin_periodo = end_date or date.today()

    movimientos_cargos = Decimal('0.00')
    movimientos_abonos = Decimal('0.00')

    if inicio_periodo:
        periodo = pagos.filter(
            pagado_real__gte=inicio_periodo,
            pagado_real__lte=fin_periodo
        ).aggregate(
            cargos=Coalesce(
                Sum('monto', filter=Q(tipo__isnull=True) | ~Q(tipo__nombre='ABONO')),
                Decimal('0.00')
            ),
            abonos=Coalesce(
                Sum('monto', filter=Q(tipo__isnull=False) & Q(tipo__nombre='ABONO')),
                Decimal('0.00')
            ),
        )
        movimientos_cargos = periodo['cargos']
        movimientos_abonos = periodo['abonos']

    saldo_final = saldo_trasladado - movimientos_cargos + movimientos_abonos

    return {
        'saldo_base': saldo_base,
        'fecha_base': fecha_base,
        'saldo_trasladado': saldo_trasladado,
        'saldo_final': saldo_final,
        'start_date': start_date,
        'end_date': end_date,
        'movimientos_cargos': movimientos_cargos,
        'movimientos_abonos': movimientos_abonos,
    }

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def control_bancos_api(request, pk):
    cuenta = get_object_or_404(Cuenta, id=pk)

    pagos, ultimo_saldo = obtener_pagos_control_bancos(cuenta)

    myfilter = Matriz_Pago_Filter(request.GET, queryset=pagos)
    pagos_filtrados = myfilter.qs

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    calculos = calcular_saldos_control_bancos(
        cuenta=cuenta,
        pagos=pagos_filtrados,
        start_date=start_date,
        end_date=end_date
    )

    paginator = PageNumberPagination()
    paginator.page_size = 25
    page = paginator.paginate_queryset(pagos_filtrados, request)

    serializer = PagoControlBancosSerializer(page, many=True, context={'request': request})

    return paginator.get_paginated_response({
        'cuenta': {
            'id': cuenta.id,
            'nombre': str(cuenta),
        },
        'ultimo_saldo': {
            'fecha_inicial': ultimo_saldo.fecha_inicial if ultimo_saldo else None,
            'monto_inicial': ultimo_saldo.monto_inicial if ultimo_saldo else None,
        } if ultimo_saldo else None,
        'resumen': {
            'saldo_base': calculos['saldo_base'],
            'fecha_base': calculos['fecha_base'],
            'saldo_trasladado': calculos['saldo_trasladado'],
            'movimientos_cargos': calculos['movimientos_cargos'],
            'movimientos_abonos': calculos['movimientos_abonos'],
            'saldo_final': calculos['saldo_final'],
            'start_date': calculos['start_date'],
            'end_date': calculos['end_date'],
        },
        'pagos': serializer.data,
    })

@api_view(["GET"])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def reporte_solicitudes_api(request):
    last_id = int(request.query_params.get("last_id", 0))
    limit = int(request.query_params.get("limit", 2000))
    fecha_inicio = request.query_params.get("fecha_inicio")
    print(fecha_inicio)
    salidas_qs = (
        Salidas.objects
        .select_related(
            "vale_salida",
            "vale_salida__solicitud",
            "vale_salida__solicitud__distrito",
            "vale_salida__solicitud__staff__staff__staff",
            "vale_salida__solicitud__activo",
            "producto__articulos__producto",
               # Relaciones de la entrada y la OC
            "entrada",
            "entrada__oc",
            "entrada__oc__moneda",
            
        )
        .prefetch_related(
            "producto__articulosrequisitados_set__req",
            "producto__articulosrequisitados_set__articulocomprado_set",
            "producto__articulosrequisitados_set__articulocomprado_set__entradaarticulo_set__entrada",
        )
        .exclude(vale_salida__solicitud__distrito__nombre__in=[
            'BRASIL',
            'ALTAMIRA ALTERNATIVO',
            'VH SECTOR 6',
        ])
        .exclude(vale_salida__folio__isnull=True)

    )
    #print(salidas_qs)
    

    if fecha_inicio:
        print('estoy entrando a filtrar por fecha')
        salidas_qs = salidas_qs.filter(
            vale_salida__created_at__gte=fecha_inicio
        )

        #print(salidas_qs)

    salidas_qs = salidas_qs.filter(
        id__gt=last_id
    ).order_by("id")

    #print(salidas_qs)
    salidas = salidas_qs[:limit]
    #print(salidas)
    data = []

    def decimal_seguro(valor, default="0"):
        try:
            if valor in (None, ""):
                return Decimal(default)

            return Decimal(str(valor))
        except (InvalidOperation, TypeError, ValueError):
            return Decimal(default)

    for salida in salidas:
        vale = salida.vale_salida
        order = vale.solicitud if vale else None
        material_recibido_por = ""

        if vale and vale.material_recibido_por:
            perfil_recibe = vale.material_recibido_por

        try:
            usuario_recibe = perfil_recibe.staff.staff
            material_recibido_por = (
                f"{usuario_recibe.first_name} "
                f"{usuario_recibe.last_name}"
            ).strip()
        except AttributeError:
            material_recibido_por = str(perfil_recibe)

        precio_salida = decimal_seguro(salida.precio)

        precio_producto = Decimal("0")
        precio_catalogo = Decimal("0")

        if salida.producto:
            precio_producto = decimal_seguro(salida.producto.precio)

            try:
                precio_catalogo = decimal_seguro(
                    salida.producto.articulos.producto.price
                )
            except AttributeError:
                precio_catalogo = Decimal("0")


        if precio_salida > 0:
            precio_original = precio_salida
        elif precio_producto > 0:
            precio_original = precio_producto
        elif precio_catalogo > 0:
            precio_original = precio_catalogo
        else:
            precio_original = Decimal("0")


        moneda = "PESOS"
        tipo_cambio = Decimal("1")

        if salida.entrada_id:
            try:
                entrada_salida = salida.entrada
                oc = entrada_salida.oc

                if oc:
                    if oc.moneda and oc.moneda.nombre:
                        moneda = str(oc.moneda.nombre).strip().upper()

                    tipo_cambio = decimal_seguro(
                        oc.tipo_de_cambio,
                        default="1"
                    )

                    if tipo_cambio <= 0:
                        tipo_cambio = Decimal("1")

            except AttributeError:
                moneda = "PESOS"
                tipo_cambio = Decimal("1")

        if moneda in ("DOLARES", "DÓLARES", "USD"):
            precio_unitario = precio_original * tipo_cambio
        else:
            precio_unitario = precio_original
        
        articulo_para_surtir = salida.producto
        articulo_ordenado = articulo_para_surtir.articulos if articulo_para_surtir else None

        req = None
        entrada = None

        if articulo_para_surtir:
            for art_req in articulo_para_surtir.articulosrequisitados_set.all():
                req = art_req.req

                art_comp = art_req.articulocomprado_set.first()
                if art_comp:
                    entrada_articulo = art_comp.entradaarticulo_set.first()
                    if entrada_articulo:
                        entrada = entrada_articulo.entrada

                if req or entrada:
                    break

        solicitante = ""
        try:
            user = order.staff.staff.staff
            solicitante = f"{user.first_name} {user.last_name}".strip()
        except Exception:
            pass

        material = ""
        try:
            material = articulo_ordenado.producto.producto.nombre
        except Exception:
            material = ""

        item = {
            "salida_id": salida.id,
            "distrito": order.distrito.nombre if order and order.distrito else "",
            "material_recibido_por": material_recibido_por,

            "quien_solicita": solicitante,
            "economico": order.activo.eco_unidad if order and order.activo and order.activo.eco_unidad else "NA",
            "folio_solicitud": str(order.folio) if order and order.folio else "",
            "fecha_solicitud": order.created_at.strftime("%d/%m/%Y") if order and order.created_at else "",

            "numero_requisicion": str(req.folio) if req and req.folio else "",
            "fecha_requisicion": req.created_at.strftime("%d/%m/%Y") if req and req.created_at else "",
            "fecha_autorizacion_requisicion": req.approved_at.strftime("%d/%m/%Y") if req and req.approved_at else "",
            "status_de_autorizacion": "Autorizada" if req and req.autorizar else "Pendiente",

            "fecha_llegada_almacen": entrada.entrada_date.strftime("%d/%m/%Y") if entrada and entrada.entrada_date else "",
            "fecha_entrega_de_almacen": salida.vale_salida.created_at.strftime("%d/%m/%Y") if salida.vale_salida and salida.vale_salida.created_at else "",

            "material_o_servicio_solicitado": material,
            "cantidad_de_material": salida.cantidad or 0,
            "precio_unitario": precio_unitario,
        }

        serializer = ReporteSolicitudesSerializer(data=item)
        serializer.is_valid(raise_exception=True)

        data.append(serializer.data)

    return Response({
        "results": data,
        "last_id": data[-1]["salida_id"] if data else None,
        "has_more": len(data) == limit,
    })