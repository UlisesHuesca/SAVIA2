from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, FileResponse, Http404, JsonResponse
from django.core.mail import EmailMessage, BadHeaderError
import socket
from smtplib import SMTPException
from django.core.paginator import Paginator
from django.core.files.base import ContentFile
from django.db.models import Count, Q, Case, When, Value, CharField, Sum, DecimalField, F
from django.db.models.functions import Concat, Coalesce
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.conf import settings
from django.utils.timezone import localtime
from django.utils.dateparse import parse_date
from django.urls import reverse, NoReverseMatch
from user.models import Distrito, Empresa
from compras.models import ArticuloComprado, Compra, TipoPrioridad
from compras.forms import CompraForm
from compras.filters import CompraFilter
from compras.views import dof, attach_oc_pdf, attach_antisoborno_pdf, attach_codigo_etica_pdf, attach_aviso_privacidad_pdf, attach_politica_proveedor, generar_pdf #convert_excel_matriz_compras
from dashboard.models import Subproyecto, Producto_Calidad
from .models import Pago, Cuenta, Facturas, Comprobante_saldo_favor, Saldo_Cuenta, Tipo_Pago, Complemento_Pago
from gastos.models import Solicitud_Gasto, Articulo_Gasto, Factura
from gastos.views import render_pdf_gasto, crear_pdf_cfdi_gasto
from viaticos.views import generar_pdf_viatico
from viaticos.models import Solicitud_Viatico, Viaticos_Factura, Concepto_Viatico
from finanzas.models import Exhibit, Linea_Exhibit
from requisiciones.views import get_image_base64
from .forms import PagoForm, Facturas_Form, Facturas_Completas_Form, Saldo_Form, ComprobanteForm, TxtForm, CompraSaldo_Form, Cargo_Abono_Form, Cargo_Abono_Tipo_Form, Saldo_Inicial_Form, Transferencia_Form, UploadFileForm, UploadComplementoForm, Cargo_Abono_No_Documento_Form
from .filters import PagoFilter, Matriz_Pago_Filter
from viaticos.filters import Solicitud_Viatico_Filter
from gastos.filters import Solicitud_Gasto_Filter
from user.models import Profile
from .utils import extraer_texto_de_pdf, encontrar_variables, extraer_texto_pdf_prop
import pytz  # Si est√°s utilizando pytz para manejar zonas horarias
from io import BytesIO
from num2words import num2words
import json
import qrcode
import tempfile
from PIL import Image
from django.utils import timezone
from django.urls import reverse
import re
from openpyxl.styles import numbers
from openpyxl.chart import BarChart, Reference
from .tasks import validar_lote_facturas
from urllib.parse import urlencode


from datetime import date, datetime, timedelta
from decimal import Decimal
import decimal
import os
import io
import zipfile
import xml.etree.ElementTree as ET




#Excel stuff
import xlsxwriter
from xlsxwriter.utility import xl_col_to_name

from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill
from openpyxl.utils import get_column_letter
import datetime as dt

#PDF generator
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.colors import Color, black, blue, red, white
from reportlab.lib.units import cm
from reportlab.lib.pagesizes import letter
from reportlab.rl_config import defaultPageSize
from compras.tasks import convert_excel_matriz_compras_task
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_RIGHT
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Frame, PageBreak
from bs4 import BeautifulSoup
from user.decorators import perfil_seleccionado_required
import subprocess
import paramiko
import logging
from PyPDF2 import PdfMerger

#Para conectar con API del SAT
from zeep import Client
import time
# Configurar logger
LOG_PATH = '/home/savia/logs/pagos_sftp.log'
os.makedirs(os.path.dirname(LOG_PATH), exist_ok=True)
logging.basicConfig(
    filename=LOG_PATH,
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)


@perfil_seleccionado_required
def compras_por_pagar(request):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    almacenes_distritos = set(usuario.almacen.values_list('distrito__id', flat=True))
    tipos_prioridad = TipoPrioridad.objects.all()
    if usuario.tipo.cuentas_por_pagar:
        compras = Compra.objects.filter(autorizado2=True, para_pago = False, pagada=False, regresar_oc = False, cerrar_sin_pago_completo = False, req__orden__distrito__in = almacenes_distritos).order_by('-folio')
        
    
    
    #compras = Compra.objects.filter(autorizado2=True, pagada=False).order_by('-folio')
    myfilter = CompraFilter(request.GET, queryset=compras)
    compras = myfilter.qs
    
    p = Paginator(compras, 50)
    page = request.GET.get('page')
    compras_list = p.get_page(page)
    
    if request.method == 'POST' and 'btnReporte' in request.POST:
        #if usuario.tipo.tesoreria or usuario.tipo.finanzas:
        return convert_excel_matriz_compras_tesoreria(compras)
        #else:
        #    return convert_excel_matriz_compras_autorizadas(compras)
       
    
    if request.method == 'POST':
        compra_ids = request.POST.getlist('compra_ids')
        print(compra_ids)
        if compra_ids:
            for compra_id in compra_ids:
                parcial = request.POST.get(f'parcial_{compra_id}')
                print(parcial)
                  # Asegurarte de que monto no sea None y que sea un n√∫mero v√°lido
                if parcial:
                    try:
                        parcial = float(parcial)
                    except ValueError:
                        parcial = 0  # O alg√∫n valor por defecto en caso de error
                Compra.objects.filter(id=compra_id).update(para_pago=True, parcial = parcial, tesorero =usuario)
            # Despu√©s de la actualizaci√≥n, redirige para restablecer el conteo y sumatoria
            return redirect('compras-por-pagar')

    context= {
        'usuario':usuario,
        'compras':compras,
        'myfilter':myfilter,
        'compras_list':compras_list,
        'tipos_prioridad': tipos_prioridad,
        }

    return render(request, 'tesoreria/compras_por_pagar.html',context)

def actualizar_prioridad(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        compra_id = data.get('compra_id')
        tipo_prioridad_id = data.get('tipo_prioridad_id')
        compra = Compra.objects.get(pk=compra_id)
        tipo = TipoPrioridad.objects.get(pk=tipo_prioridad_id)
        compra.tipo_prioridad = tipo
        compra.save()
        return JsonResponse({'success': True, 'tipo': tipo.nombre})
    return JsonResponse({'success': False}, status=400)

# Create your views here.
@perfil_seleccionado_required
def compras_autorizadas(request):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    almacenes_distritos = set(usuario.almacen.values_list('distrito__id', flat=True))
    
    if usuario.tipo.tesoreria == True:
        if usuario.tipo.rh:
            compras = Compra.objects.none()
        elif usuario.tipo.tesoreria: 
            compras = Compra.objects.filter(
                Q(tesorero__isnull=True) | Q(tesorero__tipo__tesoreria=True), # üëà Filtra que quien envi√≥ a pago sea Tesorer√≠a
                para_pago=True,
                pagada=False,
                cerrar_sin_pago_completo = False,
                autorizado2=True, 
                req__orden__distrito__in = almacenes_distritos
            ).annotate(
                total_facturas=Count('facturas', filter=Q(facturas__oc__isnull=False)),
                autorizadas=Count(Case(When(Q(facturas__autorizada=True, facturas__oc__isnull=False), then=Value(1))))
            ).order_by('-folio')
    elif usuario.tipo.finanzas:
        compras = Compra.objects.filter(
            para_pago=True,
            pagada=False,
            cerrar_sin_pago_completo = False,
            autorizado2=True, 
            tesorero__tipo__finanzas=True,  # üëà Filtra que quien envi√≥ a pago sea Finanzas
            req__orden__distrito = usuario.distritos
        ).annotate(
            total_facturas=Count('facturas', filter=Q(facturas__oc__isnull=False)),
            autorizadas=Count(Case(When(Q(facturas__autorizada=True, facturas__oc__isnull=False), then=Value(1))))
        ).order_by('-folio')
        
    
   
    
    #compras = Compra.objects.filter(autorizado2=True, pagada=False).order_by('-folio')
    myfilter = CompraFilter(request.GET, queryset=compras)
    compras = myfilter.qs
    
    p = Paginator(compras, 50)
    page = request.GET.get('page')
    compras_list = p.get_page(page)

    for compra in compras_list:
        if compra.total_facturas == 0:
            compra.estado_facturas = 'sin_facturas'
        elif compra.autorizadas == compra.total_facturas:
            compra.estado_facturas = 'todas_autorizadas'
        else:
            compra.estado_facturas = 'pendientes'

    if request.method == 'POST' and 'btnReporte' in request.POST:
        if usuario.tipo.tesoreria:
            return convert_excel_matriz_compras_tesoreria(compras)
        else:
            return convert_excel_matriz_compras_autorizadas(compras)
    

    context= {
        'compras':compras,
        'myfilter':myfilter,
        'compras_list':compras_list,
        }

    return render(request, 'tesoreria/compras_autorizadas.html',context)


# Create your views here.
@perfil_seleccionado_required
def tiempo_proceso_autorizacion(request):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    almacenes_distritos = set(usuario.almacen.values_list('distrito__id', flat=True))
    num_almacenes = usuario.almacen.count()
    if usuario.tipo.nombre == "Admin":  
        compras = Compra.objects.filter(
            autorizado2 = True, 
            req__orden__distrito__in = almacenes_distritos
        ).order_by('-folio')#.annotate(
   
    else: 
        compras = Compra.objects.filter(
            autorizado2 = True, 
            req__orden__distrito__in = almacenes_distritos
        ).order_by('-folio')#.annotate(
     
    myfilter = CompraFilter(request.GET, queryset=compras)
    compras = myfilter.qs
    
    p = Paginator(compras, 50)
    page = request.GET.get('page')
    compras_list = p.get_page(page)

    if request.method == 'POST' and 'btnReporte' in request.POST:
        return convert_excel_matriz_tiempo_proceso(compras)
    

    context= {
        'num_almacenes': num_almacenes,
        'compras':compras,
        'myfilter':myfilter,
        'compras_list':compras_list,
        }

    return render(request, 'tesoreria/tiempos_proceso.html',context)

@perfil_seleccionado_required
def transferencia_cuentas(request, pk):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    tipos_pago = Tipo_Pago.objects.all()
    transferencia = tipos_pago.get(id = 3)
    abono = tipos_pago.get(id = 2)
    transaccion, created = Pago.objects.get_or_create(tesorero = usuario, hecho=False, tipo = transferencia, cuenta=None)
    transaccion2, created = Pago.objects.get_or_create(tesorero = usuario, hecho=False, tipo = abono)
    form = Cargo_Abono_Form(instance=transaccion)
    form_transferencia = Transferencia_Form(prefix='abono')
    #pk_cuenta = request.GET.get('cuenta')
    cuenta = get_object_or_404(Cuenta, id=pk)
    #cuenta = Cuenta.objects.get(id=pk_cuenta)
    error_messages = []

    #form.fields['tipo_pago'].queryset = Tipo_Pago.objects.get(id = 3)
    print(Tipo_Pago.objects.filter(id=3))
    cuentas = Cuenta.objects.filter(moneda__nombre = 'PESOS')
      
    cuentas_para_select2 = [
        {'id': cuenta.id,
         'text': str(cuenta.cuenta) +' '+ str(cuenta.moneda), 
         'moneda': str(cuenta.moneda),
        } for cuenta in cuentas]

    if request.method == 'POST':
        if "envio" in request.POST:
            form = Cargo_Abono_Form(request.POST, request.FILES, instance = transaccion)
            form_transferencia = Transferencia_Form(request.POST, instance = transaccion2, prefix='abono')
            
            if form.is_valid() and form_transferencia.is_valid():
                cargo = form.save(commit=False)
                cargo.pagado_date = date.today()
                cargo.tipo = Tipo_Pago.objects.get(id = 3)
                cargo.pagado_hora = datetime.now().time() 
                cargo.hecho = True
                
                CUENTA_VOS = "0167229585"
                abono = form_transferencia.save(commit=False)
                abono.monto = cargo.monto
                abono.tipo = Tipo_Pago.objects.get(id = 2)
                abono.comentario = f"{cargo.comentario} (Relacionado con cuenta {cargo.cuenta})"
                abono.pagado_real = cargo.pagado_real
                abono.pagado_date = date.today()
                abono.pagado_hora = datetime.now().time()
                abono.hecho = True
                abono.distrito = cargo.cuenta.encargado.distritos
                abono.save()

                cargo.distrito = abono.cuenta.encargado.distritos
                if str(abono.cuenta.cuenta) == CUENTA_VOS:
                    cargo.comentario = f"VOS - {cargo.comentario}"
                else:
                    cargo.comentario = f"{cargo.comentario} (Relacionado con cuenta {abono.cuenta})"

                cargo.save()
                messages.success(request,f'{usuario.staff.staff.first_name}, Has agregado correctamente el traspaso')
                return redirect('control-bancos', pk = cuenta.id)
            else:
                for field, errors in form.errors.items():
                    error_messages.append(f"{field}: {errors.as_text()}")
                for field, errors in form_transferencia.errors.items():
                    error_messages.append(f"{field}: {errors.as_text()}")

    context= {
        'form':form,
        'form_transferencia': form_transferencia,
        'cuenta': cuenta,
        'cuentas_para_select2': cuentas_para_select2,
        'error_messages': error_messages,
    }

    return render(request, 'tesoreria/transferencia_cuentas.html',context)

@perfil_seleccionado_required
def cargo_abono(request, pk):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    enproceso = Tipo_Pago.objects.get(id = 3)
    cuenta = get_object_or_404(Cuenta, id=pk)
    transaccion, created = Pago.objects.get_or_create(tesorero = usuario, hecho=False, cuenta = cuenta, tipo = enproceso)
    form = Cargo_Abono_No_Documento_Form(instance=transaccion)
    cargo_abono_solo = True
    cuentas = Cuenta.objects.filter(moneda__nombre = 'PESOS')
    distritos = Distrito.objects.exclude(id__in = [7,8,16]).exclude(status=False) #7 MATRIZ ALTERNATIVO, 8 ALTAMIRA ALTERNATIVO,16 BRASIL
    form.fields['tipo'].queryset = Tipo_Pago.objects.filter(id__in=[1, 2])
    form.fields['distrito'].queryset = distritos
   

    if request.method == 'POST':
        if "envio" in request.POST:
            form = Cargo_Abono_No_Documento_Form(request.POST, instance = transaccion)
            if form.is_valid():
                pago = form.save(commit = False)
                pago.pagado_date = date.today()
                pago.pagado_hora = datetime.now().time()
                pago.hecho = True
                pago.save()   
                return redirect('control-bancos', pk = cuenta.id)
            else:
                error_str = form.errors.as_text()
                messages.error(request, f'{usuario.staff.staff.first_name}, el formulario tiene errores: {error_str}')

    context= {
        'form':form,
        'cuenta': cuenta,
        'cargo_abono_solo': cargo_abono_solo,
    }

    return render(request, 'tesoreria/cargo_abono.html',context)

@perfil_seleccionado_required
def cargo_abono_documento(request, pk):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    enproceso = Tipo_Pago.objects.get(id = 3)
    cuenta = get_object_or_404(Cuenta, id=pk)
    transaccion, created = Pago.objects.get_or_create(tesorero = usuario, hecho=False, cuenta = cuenta, tipo = enproceso)
    form = Cargo_Abono_Tipo_Form(instance=transaccion)
    distritos = Distrito.objects.exclude(id__in = [7,8,16]).exclude(status=False) #7 MATRIZ ALTERNATIVO, 8 ALTAMIRA ALTERNATIVO,16 BRASIL
    
    cuentas = Cuenta.objects.filter(moneda__nombre = 'PESOS')
    form.fields['tipo'].queryset = Tipo_Pago.objects.filter(id__in=[1, 2])
    form.fields['distrito'].queryset = distritos

    if request.method == 'POST':
        if "envio" in request.POST:
            form = Cargo_Abono_Tipo_Form(request.POST, request.FILES, instance = transaccion)
            if form.is_valid():
                pago = form.save(commit = False)
                pago.pagado_date = date.today()
                pago.pagado_hora = datetime.now().time()
                pago.hecho = True
                #Se elimina el concepto del movimiento directo a la cuenta, todos son movimientos separados que suman y restan cuando deba sacarse el c√°lculo
                #cuenta = Cuenta.objects.get(cuenta = pago.cuenta.cuenta, moneda = pago.cuenta.moneda)               
                pago.save()   
                return redirect('control-bancos', pk = cuenta.id)
            else:
                error_str = form.errors.as_text()
                messages.error(request, f'{usuario.staff.staff.first_name}, el formulario tiene errores: {error_str}')

    context= {
        'form':form,
        #'form_transferencia': form_transferencia,
        'cuenta': cuenta,
    }

    return render(request, 'tesoreria/cargo_abono.html',context)


@perfil_seleccionado_required
def saldo_inicial(request, pk):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    cuenta = Cuenta.objects.get(id = pk)
    saldo, created = Saldo_Cuenta.objects.get_or_create(hecho=False, cuenta = cuenta)
    form = Saldo_Inicial_Form(instance = saldo)

   
      
    #cuentas_para_select2 = [
    #    {'id': cuenta.id,
    #     'text': str(cuenta.cuenta) +' '+ str(cuenta.moneda), 
    #     'moneda': str(cuenta.moneda),
    #    } for cuenta in cuentas]
    
    if request.method == 'POST' and "envio" in request.POST:
        form = Saldo_Inicial_Form(request.POST, instance = saldo)
        if form.is_valid():
            saldo = form.save(commit = False)
            saldo.updated = date.today()
            #saldo.pagado_hora = datetime.now().time()
            saldo.hecho = True
            saldo.updated_by = usuario
            #Se elimina el concepto del movimiento directo a la cuenta, todos son movimientos separados que suman y restan cuando deba sacarse el c√°lculo
            #cuenta = Cuenta.objects.get(cuenta = pago.cuenta.cuenta, moneda = pago.cuenta.moneda)               
            saldo.save()   
            messages.success(request,f'{usuario.staff.staff.first_name}, Has agregado correctamente el saldo inicial de la cuenta')
            return redirect('control-bancos', pk = cuenta.id)
        else:
            messages.error(request,f'{usuario.staff.staff.first_name}, No est√° validando')

    context = {
        'cuenta':cuenta,
        'form':form,
    }

    return render(request, 'tesoreria/saldo_inicial.html',context)



def prellenar_formulario(request):
    print("M√©todo:", request.method)
    #print("X-Requested-With:", request.headers.get('X-Requested-With'))

    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        pdf_content = request.FILES.get('comprobante_pago')
        
        if not pdf_content:
            return JsonResponse({'error': 'No file uploaded'}, status=400)
        
        pdf_content = pdf_content.read()
        texto_extraido = extraer_texto_de_pdf(pdf_content)
        print("Texto extra√≠do:", texto_extraido)
        datos_extraidos = encontrar_variables(texto_extraido)
        print("Datos extra√≠dos:", datos_extraidos)
        
        fecha_str = datos_extraidos.get('fecha', '').strip()

        fecha_formato_correcto = None  # Valor por defecto en caso de que no se pueda procesar la fecha
        
        if fecha_str:
            try:
                fecha_obj = datetime.strptime(fecha_str, '%d/%m/%Y')
                fecha_formato_correcto = fecha_obj.strftime('%Y-%m-%d')
            except ValueError:
                # Opcional: Agregar alguna forma de logging o notificaci√≥n de que la fecha no es v√°lida
                print('Se lo llev√≥ madres')
                pass
        
        numero_cuenta_extraido = datos_extraidos.get('cuenta_retiro', '').strip().lstrip('0')
        cuenta_objeto = None
        
        if numero_cuenta_extraido:
            try:
                cuenta_objeto = Cuenta.objects.get(cuenta__contains=numero_cuenta_extraido)
            except Cuenta.DoesNotExist:
                # Manejar el caso donde la cuenta no existe
                return JsonResponse({'error': 'Account not found'}, status=404)
        
        divisa_cuenta_extraida = datos_extraidos.get('divisa_cuenta', '').strip()
        
        datos_para_formulario = {
            'monto': datos_extraidos.get('importe_operacion', '').replace('MXP', '').replace(',', '').strip() or None,
            'pagado_real': fecha_formato_correcto,  # Valor procesado o None
            'cuenta': cuenta_objeto.id if cuenta_objeto else None,
            'divisa_cuenta': divisa_cuenta_extraida or None,
            'hora_operacion': datos_extraidos.get('hora_operacion', '') or None,
        }
        
        return JsonResponse(datos_para_formulario)
    
    return JsonResponse({'error': 'Invalid request'}, status=400)




@perfil_seleccionado_required
def compras_pagos(request, pk):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    compra = Compra.objects.get(id=pk)
    productos = ArticuloComprado.objects.filter(oc=pk)
    productos_criticos = productos.filter(producto__producto__articulos__producto__producto__critico=True)
    pagos = Pago.objects.filter(oc=compra.id, hecho=True) #.aggregate(Sum('monto'))
    sub = Subproyecto.objects.get(id=compra.req.orden.subproyecto.id)
    pagos_alt = Pago.objects.filter(oc=compra.id, hecho=True)
    suma_pago = 0
    suma_pago_usd = 0
    
    for pago in pagos:
        
        if pago.oc.moneda.nombre == "DOLARES":
            if pago.cuenta.moneda.nombre == "PESOS":
                suma_pago = suma_pago + pago.monto
                monto_pago_usd = pago.monto/pago.tipo_de_cambio
                suma_pago_usd = suma_pago_usd + monto_pago_usd
            else:
                suma_pago = suma_pago + pago.monto * (pago.tipo_de_cambio or compra.tipo_de_cambio)
                suma_pago_usd = suma_pago_usd + pago.monto
        else:
            suma_pago = suma_pago + pago.monto

    if compra.moneda.nombre == 'PESOS':
        cuentas = Cuenta.objects.filter(moneda__nombre = 'PESOS')
        remanente = compra.costo_plus_adicionales - suma_pago
    if compra.moneda.nombre == 'DOLARES':
        cuentas = Cuenta.objects.all()
        remanente = compra.costo_plus_adicionales - suma_pago_usd


    cuentas_para_select2 = [
        {'id': cuenta.id,
         'text': str(cuenta.cuenta) +' '+ str(cuenta.moneda), 
         'moneda': str(cuenta.moneda),
        } for cuenta in cuentas]


    pago, created = Pago.objects.get_or_create(tesorero = usuario, oc__req__orden__distrito = usuario.distritos, oc=compra, hecho=False)
    form = PagoForm(instance=pago)

    base_url = reverse('compras-autorizadas')
    filtros = {
        'proveedor': request.GET.get('proveedor', ''),
        'distrito': request.GET.get('distrito', ''),
        'start_date': request.GET.get('start_date', ''),
        'end_date': request.GET.get('end_date', ''),
    }
    # Codificar los par√°metros
    query_string = urlencode(filtros)
    #print('query_string:',filtros)
    
   
   
    
    redirect_url = f"{base_url}?{query_string}" if query_string else base_url
    print('redirect_url',redirect_url)
    
    if request.method == 'POST':
        if "envio" in request.POST:
            form = PagoForm(request.POST, request.FILES or None, instance = pago)
            if form.is_valid():
                pago = form.save(commit = False)
                pago.pagado_date = date.today()
                #pago.pagado_hora = datetime.now().time()
                pago.hecho = True
                #Traigo la cuenta que se capturo en el form
                cuenta = Cuenta.objects.get(cuenta = pago.cuenta.cuenta, moneda = pago.cuenta.moneda)
                #La utilizo para sacar la informaci√≥n de todos los pagos relacionados con esa cuenta y sumarlos

                # Actualizo el saldo de la cuenta, no es necesario actualizar el saldo de la cuenta
                monto_actual = pago.monto
                if compra.moneda.nombre == "PESOS":
                    sub.gastado = sub.gastado + monto_actual
                
                if compra.moneda.nombre == "DOLARES":
                    if pago.cuenta.moneda.nombre == "PESOS": #Si la cuenta es en pesos
                        #Estoy aca
                        sub.gastado = sub.gastado + monto_actual * pago.tipo_de_cambio
                        monto_actual = monto_actual/pago.tipo_de_cambio #Lo convierto a dolares
                        suma_pago = suma_pago_usd
                    
                    if pago.cuenta.moneda.nombre == "DOLARES":
                        tipo_de_cambio = decimal.Decimal(dof())
                        sub.gastado = sub.gastado + monto_actual * tipo_de_cambio
                    #actualizar la cuenta de la que se paga
                
                print('monto_actual:',monto_actual)
                monto_total_pagado= monto_actual + suma_pago
                print('monto_total_pagado:',monto_total_pagado)
                
                compra.monto_pagado = monto_total_pagado
                costo_oc = compra.costo_plus_adicionales 
                monto_parcial = compra.parcial
                
                print('costo_oc:',costo_oc)
                print('monto_total_pagado',monto_total_pagado,)
                print('monto_parcial:', monto_parcial)
                print(monto_total_pagado - monto_parcial)
                TOLERANCIA = Decimal(0.2)
                if monto_actual <= 0:
                    messages.error(request,f'El pago {monto_actual} debe ser mayor a 0')
                #monto_total_pagado
                elif (monto_total_pagado <= abs(costo_oc + TOLERANCIA)) or (monto_total_pagado <= abs(monto_parcial + TOLERANCIA)):
                    print('dentro',monto_total_pagado - monto_parcial)
                    if abs(monto_actual - monto_parcial) <= TOLERANCIA:
                        compra.para_pago = False
                    if abs(monto_total_pagado - costo_oc) <= TOLERANCIA:
                        compra.pagada = True
                    archivo_oc = attach_oc_pdf(request, compra.id)
                    pdf_antisoborno = attach_antisoborno_pdf(request)
                    pdf_privacidad = attach_aviso_privacidad_pdf(request)
                    pdf_etica = attach_codigo_etica_pdf(request)
                    pdf_politica_proveedor = attach_politica_proveedor(request)
                    static_path = settings.STATIC_ROOT
                    img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
                    img_path2 = os.path.join(static_path,'images','logo_vordcab.jpg')
                    image_base64 = get_image_base64(img_path)
                    logo_v_base64 = get_image_base64(img_path2)
                    articulos_html = """
                    <table border="1" style="border-collapse: collapse; width: 100%;">
                        <thead>
                            <tr>
                                <th>C√≥digo</th>
                                <th>Producto Nombre</th>
                                <th>Requerimiento</th>
                            </tr>
                        </thead>        
                        <tbody>
                    """
                    productos_criticos = productos_criticos
                    for articulo in productos_criticos:
                        producto = articulo.producto.producto.articulos.producto.producto
                        pc = getattr(producto, 'producto_calidad', None)
                        if not pc:
                            # no hay ficha de calidad para este producto
                            continue

                        reqs = pc.requerimientos_calidad.select_related('requerimiento').all()

                        if reqs.exists():
                            for requerimiento in reqs:
                                articulos_html += f"""
                                    <tr>
                                        <td>{producto.codigo}</td>
                                        <td>{producto.nombre}
                                        <td>{requerimiento.requerimiento}</td>
                                    </tr>
                                """
                        else:
                            articulos_html += f"""
                                <tr>
                                    <td>{producto.codigo}</td>
                                    <td>{producto.nombre}</td>
                                    <td>Sin requerimiento</td>
                                </tr>
                            """
                    articulos_html += """
                        </tbody>
                    </table>
                    """
                    #if compra.cond_de_pago.nombre == "CONTADO":
                    pagos = Pago.objects.filter(oc=compra, hecho=True)
                    html_message = f"""
                    <html>
                        <head>
                            <meta charset="UTF-8">
                        </head>
                        <body style="font-family: Arial, sans-serif; color: #333; background-color: #f4f4f4; margin: 0; padding: 0;">
                            <table width="100%" cellspacing="0" cellpadding="0" style="background-color: #f4f4f4; padding: 20px;">
                                <tr>
                                    <td align="center">
                                        <table width="600px" cellspacing="0" cellpadding="0" style="background-color: #ffffff; padding: 20px; border-radius: 10px;">
                                            <tr>
                                                <td align="center">
                                                    <img src="data:image/jpeg;base64,{logo_v_base64}" alt="Logo" style="width: 100px; height: auto;" />
                                                </td>
                                            </tr>
                                            <tr>
                                                <td style="padding: 20px;">
                                                    <p style="font-size: 18px; text-align: justify;">
                                                        <p>Estimado {compra.req.orden.staff.staff.staff.first_name} {compra.req.orden.staff.staff.staff.last_name},</p>
                                                    </p>
                                                    <p style="font-size: 16px; text-align: justify;">
                                                        Est√°s recibiendo este correo porque tu OC {compra.folio} | RQ: {compra.req.folio} |Sol: {compra.req.orden.folio} ha sido pagada por {pago.tesorero.staff.staff.first_name} {pago.tesorero.staff.staff.last_name},</p>
                                                    <p>El siguiente paso del sistema: Recepci√≥n por parte de Almac√©n</p>
                                                    </p>
                                                    <p style="text-align: center; margin: 20px 0;">
                                                        <img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width: 50px; height: auto; border-radius: 50%;" />
                                                    </p>
                                                    <p style="font-size: 14px; color: #999; text-align: justify;">
                                                        Este mensaje ha sido autom√°ticamente generado por SAVIA 2.0
                                                    </p>
                                                </td>
                                            </tr>
                                        </table>
                                    </td>
                                </tr>
                            </table>
                        </body>
                    </html>
                    """
                    try:
                        email = EmailMessage(
                        f'OC Pagada {compra.folio}|RQ: {compra.req.folio} |Sol: {compra.req.orden.folio}',
                        body=html_message,
                        from_email = settings.DEFAULT_FROM_EMAIL,
                        to= [compra.req.orden.staff.staff.staff.email,],
                        headers={'Content-Type': 'text/html'}
                        )
                        email.content_subtype = "html " # Importante para que se interprete como HTML
                        email.send()
                    except (BadHeaderError, SMTPException, socket.gaierror) as e:
                        error_message = f'Correo de notificaci√≥n 1: No enviado'
                    html_message2 = f"""
                    <html>
                        <head>
                            <meta charset="UTF-8">
                        </head>
                        <body style="font-family: Arial, sans-serif; color: #333; background-color: #f4f4f4; margin: 0; padding: 0;">
                            <table width="100%" cellspacing="0" cellpadding="0" style="background-color: #f4f4f4; padding: 20px;">
                                <tr>
                                    <td align="center">
                                        <table width="600px" cellspacing="0" cellpadding="0" style="background-color: #ffffff; padding: 20px; border-radius: 10px;">
                                            <tr>
                                                <td align="center">
                                                    <img src="data:image/jpeg;base64,{logo_v_base64}" alt="Logo" style="width: 100px; height: auto;" />
                                                </td>
                                            </tr>
                                            <tr>
                                                <td style="padding: 20px;">
                                                    <p style="font-size: 18px; text-align: justify;">
                                                        <p>Estimado(a) {compra.proveedor.contacto}| Proveedor {compra.proveedor.nombre}:,</p>
                                                    </p>
                                                    <p style="font-size: 16px; text-align: justify;">
                                                        Est√°s recibiendo este correo porque has sido seleccionado para surtirnos la OC adjunta con folio: {compra.folio}.</p>
                                                        <p>&nbsp;</p>
                                                        <p> Atte. {compra.creada_por.staff.staff.first_name} {compra.creada_por.staff.staff.last_name}</p> 
                                                        <p>GRUPO VORDCAB S.A. de C.V.</p>
                                                        {f"{articulos_html}" if productos_criticos.exists() else ""}
                                                    </p>
                                                    <p style="text-align: center; margin: 20px 0;">
                                                        <img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width: 50px; height: auto; border-radius: 50%;" />
                                                    </p>
                                                    <p style="font-size: 14px; color: #999; text-align: justify;">
                                                        Este mensaje ha sido autom√°ticamente generado por SAVIA 2.0
                                                    </p>
                                                </td>
                                            </tr>
                                        </table>
                                    </td>
                                </tr>
                            </table>
                        </body>
                    </html>
                    """
                    try:
                        email = EmailMessage(
                        f'Compra Autorizada {compra.folio}|SAVIA',
                        body=html_message2,
                        from_email =settings.DEFAULT_FROM_EMAIL,
                        to= ['ulises_huesc@hotmail.com', compra.creada_por.staff.staff.email, compra.proveedor.email,],
                        headers={'Content-Type': 'text/html'}
                        )
                        #print('Correo enviadoooooooooooo')
                        #print(compra.creada_por.staff.staff.email)
                        email.content_subtype = "html " # Importante para que se interprete como HTML
                        if compra.entrada_completa == False:
                            email.attach(f'OC_folio_{compra.folio}.pdf',archivo_oc,'application/pdf')
                        email.attach(f'Politica_antisoborno.pdf', pdf_antisoborno, 'application/pdf')
                        email.attach(f'Aviso_de_privacidad.pdf', pdf_privacidad, 'application/pdf')
                        email.attach(f'Codigo_de_etica.pdf', pdf_etica, 'application/pdf')
                        email.attach(f'Politica_de_proveedor.pdf', pdf_politica_proveedor, 'application/pdf')
                        #email.attach('Pago.pdf',request.FILES['comprobante_pago'].read(),'application/pdf')
                        archivo_comprobante = request.FILES.get('comprobante_pago')
                        if archivo_comprobante:  # Verifica que el archivo exista en el request
                            archivo_contenido = archivo_comprobante.read()
                            email.attach('Pago.pdf', archivo_contenido, 'application/pdf')
                        for archivo in pagos:
                            if archivo.comprobante_pago:  # Verificar que el archivo exista
                                    with open(archivo.comprobante_pago.path, 'rb') as file:  # Abrir el archivo
                                        archivo_contenido = file.read()  # Leer el contenido
                                        nombre_archivo = f'Pago_{archivo.id}.pdf'  
                                        email.attach(nombre_archivo, archivo_contenido, 'application/pdf') 
                        # Adjuntar los archivos con nombres personalizados
                        #for articulo in productos:
                            #producto = articulo.producto.producto.articulos.producto.producto
                           
                            #if not getattr(producto, "critico", False):
                            #    continue  # solo interesa si es cr√≠tico

                            # Intenta obtener la ficha de calidad sin que truene
                            #pc = getattr(producto, "producto_calidad", None)
                            #if not pc:
                                # Producto cr√≠tico sin ficha de calidad ‚Üí lo saltamos
                            #   continue
                            #requerimientos = producto.producto_calidad.requerimientos_calidad.all()
                            #contador = 1  # Contador para evitar nombres duplicados
                                
                            #if not requerimientos.exists():
                            #    continue  # Si no hay requerimientos, saltar al siguiente producto

                            #for requerimiento in requerimientos:
                                #archivo_path = requerimiento.url.path
                                #nombre_archivo = f"{producto.codigo}_requerimiento_{contador}{os.path.splitext(archivo_path)[1]}"
                                    
                                # Abrir el archivo en modo binario y adjuntarlo directamente
                                #with open(archivo_path, 'rb') as archivo:
                                #    email.attach(nombre_archivo, archivo.read())

                                #contador += 1  # Incrementar el contador para el siguiente archivo
                        email.send()
                        messages.success(request,f'Gracias por registrar tu pago, {usuario.staff.staff.first_name}')
                    except (BadHeaderError, SMTPException, socket.gaierror) as e:
                        error_message = f'Gracias por registrar tu pago, {usuario.staff.staff.first_name} Atencion: el correo de notificaci√≥n no ha sido enviado debido a un error: {e}'
                        messages.warning(request, error_message)
                else:
                    messages.error(request,f'El monto total pagado es mayor que el costo de la compra o que el monto parcial {monto_actual} - {suma_pago} - {monto_total_pagado} - {monto_parcial} - {costo_oc}')
                    return redirect(redirect_url)
                pago.save()
                compra.save()
                form.save()
                sub.save()
                cuenta.save()
                messages.success(request,f'Gracias por registrar tu pago, {usuario.staff.staff.first_name}')
                return redirect(redirect_url)#No content to render nothing and send a "signal" to javascript in order to close window
                #elif monto_pagado > compra.costo_oc:
                #    messages.error(request,f'El monto total pagado es mayor que el costo de la compra {monto_pagado} > {compra.costo_oc}')
            else:
                form = PagoForm()
                messages.error(request,f'{usuario.staff.staff.first_name}, No se pudo subir tu documento')
        if "cerrar_sin_pago" in request.POST:
            compra.comentario_cierre = request.POST.get('comentario_cierre')
            compra.cerrar_sin_pago_completo = True
            compra.fecha_cierre = date.today()
            compra.persona_cierre = usuario  # Aseg√∫rate de tener esta variable ya disponible en tu vista
            compra.para_pago = False  # (si aplica para cerrar la compra)
            compra.save()
            messages.success(request, f'Compra {compra.folio} cerrada sin pago completo.')
            return redirect('compras-autorizadas')  # o donde desees redirigir
        #    messages.error(request,f'{usuario.staff.staff.first_name}, No est√° validando')

    context= {
        'compra':compra,
        'pago':pago,
        'form':form,
        'monto':suma_pago,
        'suma_pago_usd': suma_pago_usd,
        'pagos_alt':pagos_alt,
        'cuentas_para_select2': cuentas_para_select2,
        'remanente':remanente,
    }

    return render(request, 'tesoreria/compras_pagos.html',context)

@perfil_seleccionado_required
def edit_pago(request, pk):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    pago = Pago.objects.get(id=pk)
    print('pago_id',pk)
     
    compra = Compra.objects.get(id = pago.oc.id)
    productos = ArticuloComprado.objects.filter(oc=compra.id)
    pagos = Pago.objects.filter(oc=compra.id, hecho=True) #.aggregate(Sum('monto'))
    sub = Subproyecto.objects.get(id=compra.req.orden.subproyecto.id)
    pagos_alt = Pago.objects.filter(oc=compra.id, hecho=True)
    suma_pago = 0

    for item in pagos:
        if item.oc.moneda.nombre == "DOLARES":
            if item.cuenta.moneda.nombre == "PESOS":
                monto_pago = item.monto/item.tipo_de_cambio
                suma_pago = suma_pago + monto_pago
            else:
                suma_pago = suma_pago + item.monto
        else:
            suma_pago = suma_pago + item.monto


    if compra.moneda.nombre == 'PESOS':
        cuentas = Cuenta.objects.filter(moneda__nombre = 'PESOS')
    if compra.moneda.nombre == 'DOLARES':
        cuentas = Cuenta.objects.all()

    remanente = compra.costo_oc - suma_pago
    # Verificar si es un POST para guardar los cambios
    print(pago)
    if request.method == "POST":
        form = PagoForm(request.POST, request.FILES or None, instance=pago)
        if "btn_actualizar" in request.POST:
            if form.is_valid():
                form.save()
                messages.success(request,f'Has actualizado el pago {pago.id} de manera satisfactoria')
                # Redirigir al usuario a donde quieras luego de guardar los cambios
                return redirect('compras-autorizadas')
        if "btn_eliminar" in request.POST:
            print('ya estoy aqu√≠')
            if compra.moneda.nombre == "PESOS":
                sub.gastado = sub.gastado - pago.monto
                compra.monto_pagado = compra.monto_pagado - pago.monto
            if compra.moneda.nombre == "DOLARES":
                if pago.cuenta.moneda.nombre == "PESOS": #Si la cuenta es en pesos
                    sub.gastado = sub.gastado - pago.monto * pago.tipo_de_cambio
                    compra.monto_pagado = compra.monto_pagado - pago.monto * pago.tipo_de_cambio
                if pago.cuenta.moneda.nombre == "DOLARES":
                    tipo_de_cambio = decimal.Decimal(dof())
                    sub.gastado = sub.gastado - pago.monto * tipo_de_cambio
                    compra.monto_pagado = compra.monto_pagado - pago.monto
                
                messages.success(request,f'Has eliminado el pago {pago.id} de manera satisfactoria')
            pago.delete()
            return redirect('compras-autorizadas')
    else:
        # Si no es un POST, simplemente carga el formulario con el objeto Pago
        form = PagoForm(instance=pago)


    context= {
        'compra':compra,
        'pago':pago,
        'form':form,
        'monto':suma_pago,
        'suma_pagos': suma_pago,
        'pagos_alt':pagos_alt,
        'cuentas':cuentas,
        'remanente':remanente,
    }
    return render(request, 'tesoreria/compras_pagos.html', context)


@perfil_seleccionado_required
def edit_comprobante_pago(request, pk):
    pago = Pago.objects.get(id = pk)
    #print(pago.id)
    form = ComprobanteForm(instance = pago)

    if request.method == 'POST':
        form = ComprobanteForm(request.POST, request.FILES, instance=pago)
        if form.is_valid():
            form.save()
            return HttpResponse(status=204) #No content to render nothing and send a "signal" to javascript in order to close window
    
    context = {
        'pago':pago,
        'form':form, 
    }
    
    return render(request, 'tesoreria/edit_comprobante_pago.html',context)


@perfil_seleccionado_required
def saldo_a_favor(request, pk):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    compra = Compra.objects.get(id=pk)
    pagos = Pago.objects.filter(oc=compra.id, hecho=True) #.aggregate(Sum('monto'))
    saldo, created = Comprobante_saldo_favor.objects.get_or_create(oc=compra, hecho=False)
    form2 = Saldo_Form(instance = saldo)
    form = CompraSaldo_Form(instance = compra)
    suma_pago = 0
    for item in pagos:
        if item.oc.moneda.nombre == "DOLARES":
            if item.cuenta.moneda.nombre == "PESOS":
                monto_pago = item.monto/item.tipo_de_cambio
                suma_pago = suma_pago + monto_pago
            else:
                suma_pago = suma_pago + item.monto
        else:
            suma_pago = suma_pago + item.monto

    remanente = compra.costo_oc - suma_pago


    if request.method == 'POST':
        form2 = Saldo_Form(request.POST, request.FILES, instance = saldo)
        form = CompraSaldo_Form( request.POST, instance = compra)
        if form.is_valid() and form2.is_valid():
            form.save()
            saldo = form2.save(commit=False)
            saldo.subido_por = usuario
            saldo.fecha_subido = date.today()
            saldo.hora_subido = datetime.now().time()
            saldo.hecho = True
            saldo.save()
            if remanente <= compra.saldo_a_favor:
                compra.pagada = True
            compra.save()
            messages.success(request,f'El saldo se ha registrado correctamente, {usuario.staff.staff.first_name}')
            return HttpResponse(status=204) 

    context= {
        'compra':compra,
        'monto':suma_pago,
        'remanente':remanente,
        'form':form,
        'form2':form2,
    }

    return render(request, 'tesoreria/saldo_a_favor.html',context)

# Create your views here.
@perfil_seleccionado_required
def matriz_pagos(request):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    almacenes_distritos = set(usuario.almacen.values_list('distrito__id', flat=True))
    if usuario.tipo.rh == True:
        pagos = Pago.objects.filter(
        gasto__distrito__in = almacenes_distritos, gasto__autorizar2 = True, gasto__tipo__tipo__in = ['APOYO DE MANTENIMIENTO', 'APOYO DE RENTA'] , 
        hecho=True
        ).annotate(
        # Detectar la relaci√≥n que tiene facturas
        total_facturas=Count(
            'oc__facturas', filter=Q(oc__facturas__hecho=True)
        ) + Count(
            'gasto__facturas__hecho', filter=Q(gasto__facturas__hecho=True)
        ) + Count(
            'viatico__facturas__hecho', filter=Q(viatico__facturas__hecho=True)
        ),
        autorizadas=Count(
            Case(
                When(Q(oc__facturas__autorizada=True, oc__facturas__hecho=True), then=Value(1))
            )
        ) + Count(
            Case(
                When(Q(gasto__facturas__autorizada=True, gasto__facturas__hecho=True), then=Value(1))
            )
        ) + Count(
            Case(
                When(Q(viatico__facturas__autorizada=True, viatico__facturas__hecho=True), then=Value(1))
            )
        ),
        ).order_by('-pagado_real')
    elif usuario.distritos.nombre == "MATRIZ":
        pagos = Pago.objects.filter(
        Q(oc__req__orden__distrito__in =almacenes_distritos) & Q(oc__autorizado2=True) | 
        Q(viatico__distrito__in = almacenes_distritos) & Q(viatico__autorizar2=True) |
        Q(gasto__distrito__in = almacenes_distritos) & Q(gasto__autorizar2 = True)|
        Q(gasto__tipo__tipo = 'NOMINA'), 
        hecho=True
        ).annotate(
        # Detectar la relaci√≥n que tiene facturas
        total_facturas=Count(
            'oc__facturas', filter=Q(oc__facturas__hecho=True)
        ) + Count(
            'gasto__facturas__hecho', filter=Q(gasto__facturas__hecho=True)
        ) + Count(
            'viatico__facturas__hecho', filter=Q(viatico__facturas__hecho=True)
        ),
        autorizadas=Count(
            Case(
                When(Q(oc__facturas__autorizada=True, oc__facturas__hecho=True), then=Value(1))
            )
        ) + Count(
            Case(
                When(Q(gasto__facturas__autorizada=True, gasto__facturas__hecho=True), then=Value(1))
            )
        ) + Count(
            Case(
                When(Q(viatico__facturas__autorizada=True, viatico__facturas__hecho=True), then=Value(1))
            )
        ),
        ).order_by('-pagado_real')
    else:    
        pagos = Pago.objects.filter(
        Q(oc__req__orden__distrito__in =almacenes_distritos) & Q(oc__autorizado2=True) | 
        Q(viatico__distrito__in = almacenes_distritos) & Q(viatico__autorizar2=True) |
        Q(gasto__distrito__in = almacenes_distritos) & Q(gasto__autorizar2 = True), 
        hecho=True
        ).exclude(
            Q(gasto__tipo__tipo = "NOMINA")
        ).annotate(
        # Detectar la relaci√≥n que tiene facturas
        total_facturas=Count(
            'oc__facturas', filter=Q(oc__facturas__hecho=True)
        ) + Count(
            'gasto__facturas__hecho', filter=Q(gasto__facturas__hecho=True)
        ) + Count(
            'viatico__facturas__hecho', filter=Q(viatico__facturas__hecho=True)
        ),
        autorizadas=Count(
            Case(
                When(Q(oc__facturas__autorizada=True, oc__facturas__hecho=True), then=Value(1))
            )
        ) + Count(
            Case(
                When(Q(gasto__facturas__autorizada=True, gasto__facturas__hecho=True), then=Value(1))
            )
        ) + Count(
            Case(
                When(Q(viatico__facturas__autorizada=True, viatico__facturas__hecho=True), then=Value(1))
            )
        ),
        ).order_by('-pagado_real')
    myfilter = Matriz_Pago_Filter(request.GET, queryset=pagos)
    pagos = myfilter.qs
    

    
    #Los distritos se definen de forma "din√°mica" de acuerdo a los almacenes que tiene el usuario en el perfil
    distritos = Distrito.objects.filter(id__in=almacenes_distritos)
    tesoreros = Profile.objects.filter(tipo__nombre__in = ["Tesoreria","Tesoreria_Documentos" ], st_activo = True, distritos__in = almacenes_distritos)
    #Set up pagination
    p = Paginator(pagos, 50)
    page = request.GET.get('page')
    pagos_list = p.get_page(page)
    for pago in pagos_list:
        if pago.gasto:
            articulos_gasto = Articulo_Gasto.objects.filter(gasto=pago.gasto)

            proyectos = set()
            subproyectos = set()

            for articulo in articulos_gasto:
                if articulo.proyecto:
                    proyectos.add(str(articulo.proyecto.nombre))
                if articulo.subproyecto:
                    subproyectos.add(str(articulo.subproyecto.nombre))
            pago.proyectos = ', '.join(proyectos)
            pago.subproyectos = ', '.join(subproyectos)
        
            #print('pago.proyectos:', pago.proyectos)

    if request.method == 'POST': 
        if 'btnReporte' in request.POST:
            return convert_excel_matriz_pagos(pagos)
        elif 'btnReporteXML' in request.POST:
            fecha_inicio = parse_date(request.POST.get('fecha_inicio'))
            fecha_fin = parse_date(request.POST.get('fecha_fin'))
            distrito_id = request.POST.get('distrito')
            tesorero_id = request.POST.get('tesorero')
            folio = request.POST.get('folio')
            print('reporte_xml')
            if usuario.distritos.nombre == "MATRIZ" and usuario.tipo.documentos == False:
                pagos = Pago.objects.filter(hecho=True)
                if fecha_inicio and fecha_fin:
                    pagos = Pago.objects.filter(Q(pagado_real__range=[fecha_inicio, fecha_fin])|Q(pagado_date__range=[fecha_inicio, fecha_fin]), hecho = True)
              
                    if distrito_id:
                        pagos = pagos.filter(
                            Q(gasto__distrito_id=distrito_id) |
                            Q(oc__req__orden__distrito_id=distrito_id) |
                            Q(viatico__distrito_id=distrito_id)
                        )

                    if tesorero_id:
                        pagos = pagos.filter(tesorero_id=tesorero_id)

                if folio:
                    pagos = Pago.objects.filter(hecho = True)
                    pagos = pagos.filter(
                        Q(gasto__folio=folio) |
                        Q(oc__folio=folio) |
                        Q(viatico__folio=folio)
                    )
                    if distrito_id:
                        pagos = pagos.filter(
                            Q(gasto__distrito_id=distrito_id) |
                            Q(oc__req__orden__distrito_id=distrito_id) |
                            Q(viatico__distrito_id=distrito_id)
                        )

                    if tesorero_id:
                        pagos = pagos.filter(tesorero_id=tesorero_id)
            else:
                pagos = pagos.filter(
                    Q(pagado_real__range=[fecha_inicio, fecha_fin])|Q(pagado_date__range=[fecha_inicio, fecha_fin]),
                    Q(gasto__distrito__nombre=usuario.distritos.nombre) |
                    Q(oc__req__orden__distrito__nombre=usuario.distritos.nombre) |
                    Q(viatico__distrito__nombre=usuario.distritos.nombre),
                    hecho = True
                )
             
            datos_xml_lista = []
            for pago in pagos:
                if pago.gasto:
                    gasto = pago.gasto
                    for factura in gasto.facturas.all():
                        beneficiario = factura.solicitud_gasto.colaborador.staff.staff.first_name + ' ' + factura.solicitud_gasto.colaborador.staff.staff.last_name  if factura.solicitud_gasto.colaborador else factura.solicitud_gasto.staff.staff.staff.first_name + ' ' + factura.solicitud_gasto.staff.staff.staff.last_name
                        fecha_subida = factura.fecha_subida.astimezone(tz=None).replace(tzinfo=None) if factura.fecha_subida else 'No disponible'      
                        if factura.archivo_xml:     
                            datos_xml_lista.append(extraer_datos_xml_carpetas(factura.archivo_xml.path, f"G{gasto.folio}", fecha_subida, gasto.distrito.nombre, beneficiario, "NA", factura))    
                elif pago.oc:
                    oc = pago.oc
                    for factura in oc.facturas.all():
                        if factura.factura_xml:
                            datos_xml_lista.append(extraer_datos_xml_carpetas(factura.factura_xml.path, f"OC{oc.folio}", factura.fecha_subido, oc.req.orden.distrito.nombre, "NA", "NA", factura))
                elif pago.viatico:
                    viatico = pago.viatico
                    for factura in viatico.facturas.all():
                        beneficiario = factura.solicitud_viatico.colaborador.staff.staff.first_name + ' ' + factura.solicitud_viatico.colaborador.staff.staff.last_name  if factura.solicitud_viatico.colaborador else factura.solicitud_gasto.staff.staff.staff.first_name + ' ' + factura.solicitud_gasto.staff.staff.staff.last_name
                        fecha_subida = factura.fecha_subido.astimezone(tz=None).replace(tzinfo=None) if factura.fecha_subido else 'No disponible' # Formato YYYY-MM-DD
                        if factura.factura_xml:
                            datos_xml_lista.append(extraer_datos_xml_carpetas(factura.factura_xml.path, f"V{viatico.folio}", fecha_subida, viatico.distrito.nombre, beneficiario, "NA", factura))
            output = generar_excel_xmls(datos_xml_lista)
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=reporte_facturas.xlsx'
            response.set_cookie('descarga_iniciada', 'true', max_age=20)
            return response
        elif 'btnDescargarFacturas' in request.POST:
            fecha_inicio = parse_date(request.POST.get('fecha_inicio'))
            fecha_fin = parse_date(request.POST.get('fecha_fin'))
            distrito_id = request.POST.get('distrito')
            tesorero_id = request.POST.get('tesorero')
            folio = request.POST.get('folio')
            validar_sat = request.POST.get('validacion') == 'on'
            tipo_documento = request.POST.get('tipo_documento') 
            #print('tipo_documento:', tipo_documento)
            
            print(usuario.tipo.documentos)
            if usuario.distritos.nombre == "MATRIZ" and usuario.tipo.documentos == False:
                print('Como entr√≥?')
                pagos = Pago.objects.filter(hecho=True)
                if fecha_inicio and fecha_fin:
                    #pagos = Pago.objects.filter(Q(pagado_real__range=[fecha_inicio, fecha_fin])|Q(pagado_date__range=[fecha_inicio, fecha_fin]), hecho = True)
                    pagos = Pago.objects.filter(pagado_real__range=[fecha_inicio, fecha_fin], hecho = True) #Se modifica para que ya solo descargue con respecto a la --fecha real--
                    
                    if tipo_documento == "gastos":
                        pagos = pagos.filter(gasto__isnull=False)
                    elif tipo_documento == "compras":
                        pagos = pagos.filter(oc__isnull=False)
                    elif tipo_documento == "viaticos":
                        pagos = pagos.filter(viatico__isnull=False)

                    if distrito_id:
                        pagos = pagos.filter(
                            Q(gasto__distrito_id=distrito_id) |
                            Q(oc__req__orden__distrito_id=distrito_id) |
                            Q(viatico__distrito_id=distrito_id)
                        )

                    if tesorero_id:
                        pagos = pagos.filter(tesorero_id=tesorero_id)

                if folio:
                    pagos = Pago.objects.filter(hecho = True)
                    pagos = pagos.filter(
                        Q(gasto__folio=folio) |
                        Q(oc__folio=folio) |
                        Q(viatico__folio=folio)
                    )
                    if distrito_id:
                        pagos = pagos.filter(
                            Q(gasto__distrito_id=distrito_id) |
                            Q(oc__req__orden__distrito_id=distrito_id) |
                            Q(viatico__distrito_id=distrito_id)
                        )

                    if tesorero_id:
                        pagos = pagos.filter(tesorero_id=tesorero_id)
            else:
                #print('Aqu√≠ tambi√©n entr√≥')
                pagos = pagos.filter(
                    Q(pagado_real__range=[fecha_inicio, fecha_fin])|Q(pagado_date__range=[fecha_inicio, fecha_fin]),
                    Q(gasto__distrito=usuario.distritos) |
                    Q(oc__req__orden__distrito=usuario.distritos) |
                    Q(viatico__distrito=usuario.distritos)
                )
                if tipo_documento == "gastos":
                    pagos = pagos.filter(gasto__isnull=False)
                elif tipo_documento == "compras":
                    pagos = pagos.filter(oc__isnull=False)
                elif tipo_documento == "viaticos":
                    pagos = pagos.filter(viatico__isnull=False)


            if validar_sat:
                ids_gastos = set()
                ids_compras = set()
                ids_viaticos = set()
                #print(f"Pagos: {pagos.count()}")
                for pago in pagos:
                    if pago.gasto:
                        ids_gastos.update(pago.gasto.facturas.values_list('id', flat=True))
                    elif pago.oc:
                        ids_compras.update(pago.oc.facturas.values_list('id', flat=True))
                    elif pago.viatico:
                        ids_viaticos.update(pago.viatico.facturas.values_list('id', flat=True))
                print(ids_gastos, ids_compras, ids_viaticos)
                validar_lote_facturas.delay(list(ids_gastos), list(ids_compras), list(ids_viaticos))
            else:
                zip_buffer = BytesIO()
                datos_xml_lista = []
                processed_docs = set()

                with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
                    #zip_file.mkdir("GENERAL_PDFs")
                    #zip_file.mkdir("GENERAL_XMLs")

                    for pago in pagos:
                        carpeta = None
                        if pago.gasto:
                            gasto = pago.gasto
                            carpeta = f'{pago.pagado_real}_GASTO_{gasto.folio}_{gasto.distrito.nombre}'
                            #zip_file.mkdir(carpeta)
                            for factura in gasto.facturas.all():
                                beneficiario = factura.solicitud_gasto.colaborador.staff.staff.first_name + ' ' + factura.solicitud_gasto.colaborador.staff.staff.last_name  if factura.solicitud_gasto.colaborador else factura.solicitud_gasto.staff.staff.staff.first_name + ' ' + factura.solicitud_gasto.staff.staff.staff.last_name
                                fecha_subida = factura.fecha_subida.astimezone(tz=None).replace(tzinfo=None) if factura.fecha_subida else 'No disponible'
                                if factura.archivo_pdf:
                                    zip_file.write(factura.archivo_pdf.path, os.path.join(carpeta, os.path.basename(factura.archivo_pdf.path)))
                                    uuid = factura.uuid if factura.uuid else 'SIN_UUID'
                                    zip_file.write(factura.archivo_pdf.path, f"GENERAL_PDFs/{factura.id}_{uuid}.pdf")
                                if factura.archivo_xml:
                                    zip_file.write(factura.archivo_xml.path, os.path.join(carpeta, os.path.basename(factura.archivo_xml.path)))
                                    uuid = factura.uuid if factura.uuid else 'SIN_UUID'
                                    gen_path = f"GENERAL_XMLs/{factura.id}_{uuid}.xml"
                                    zip_file.write(factura.archivo_xml.path, gen_path)
                                    datos_xml_lista.append(extraer_datos_xml_carpetas(factura.archivo_xml.path, f"G{gasto.folio}", fecha_subida, gasto.distrito.nombre, beneficiario, gen_path, factura))
                            if gasto.id not in processed_docs:
                                pdf_buf = render_pdf_gasto(gasto.id)
                                zip_file.writestr(os.path.join(carpeta, f'GASTO_{gasto.folio}.pdf'), pdf_buf.getvalue())
                                processed_docs.add(gasto.id)
                        elif pago.oc:
                            oc = pago.oc
                            carpeta = f'{pago.pagado_real}_COMPRA_{oc.folio}_{oc.req.orden.distrito.nombre}'
                            #zip_file.mkdir(carpeta)
                            for factura in oc.facturas.all():
                                if factura.factura_pdf:
                                    zip_file.write(factura.factura_pdf.path, os.path.join(carpeta, os.path.basename(factura.factura_pdf.path)))
                                    uuid = factura.uuid if factura.uuid else 'SIN_UUID'
                                    zip_file.write(factura.factura_pdf.path, f"GENERAL_PDFs/{factura.id}_{uuid}.pdf")
                                if factura.factura_xml:
                                    zip_file.write(factura.factura_xml.path, os.path.join(carpeta, os.path.basename(factura.factura_xml.path)))
                                    uuid = factura.uuid if factura.uuid else 'SIN_UUID'
                                    gen_path = f"GENERAL_XMLs/{factura.id}_{uuid}.xml"
                                    zip_file.write(factura.factura_xml.path, gen_path)
                                    datos_xml_lista.append(extraer_datos_xml_carpetas(factura.factura_xml.path, f"OC{oc.folio}", factura.fecha_subido, oc.req.orden.distrito.nombre, "NA", gen_path, factura))
                                for complemento in factura.complementos.all():
                                    if complemento.complemento_pdf:     #Encarpeta el complemento_pdf
                                        complemento_file_name = os.path.basename(complemento.complemento_pdf.path)
                                        zip_file.write(complemento.complemento_pdf.path, os.path.join(carpeta, complemento_file_name))
                                    if complemento.complemento_xml:     #Encarpeta el complemento_xml
                                        complemento_file_name = os.path.basename(complemento.complemento_xml.path)
                                        zip_file.write(complemento.complemento_xml.path, os.path.join(carpeta, complemento_file_name))
                            if oc.id not in processed_docs:
                                pdf_buf = generar_pdf(oc)
                                zip_file.writestr(os.path.join(carpeta, f'OC_{oc.folio}.pdf'), pdf_buf.getvalue())
                                processed_docs.add(oc.id)
                                     # üöÄ Incluir los complementos de pago relacionados en la misma carpeta
                              
                        elif pago.viatico:
                            viatico = pago.viatico
                            carpeta = f'{pago.pagado_real}_VIATICO_{viatico.folio}_{viatico.distrito.nombre}'
                            #zip_file.mkdir(carpeta)
                            for factura in viatico.facturas.all():
                                beneficiario = factura.solicitud_viatico.colaborador.staff.staff.first_name + ' ' + factura.solicitud_viatico.colaborador.staff.staff.last_name  if factura.solicitud_viatico.colaborador else factura.solicitud_gasto.staff.staff.staff.first_name + ' ' + factura.solicitud_gasto.staff.staff.staff.last_name
                                fecha_subida = factura.fecha_subido.astimezone(tz=None).replace(tzinfo=None) if factura.fecha_subido else 'No disponible' # Formato YYYY-MM-DD
                                if factura.factura_pdf:
                                    zip_file.write(factura.factura_pdf.path, os.path.join(carpeta, os.path.basename(factura.factura_pdf.path)))
                                    uuid = factura.uuid if factura.uuid else 'SIN_UUID'
                                    zip_file.write(factura.factura_pdf.path, f"GENERAL_PDFs/{factura.id}_{uuid}.pdf")
                                if factura.factura_xml:
                                    zip_file.write(factura.factura_xml.path, os.path.join(carpeta, os.path.basename(factura.factura_xml.path)))
                                    uuid = factura.uuid if factura.uuid else 'SIN_UUID'
                                    gen_path = f"GENERAL_XMLs/{factura.id}_{uuid}.xml"
                                    zip_file.write(factura.factura_xml.path, gen_path)
                                    datos_xml_lista.append(extraer_datos_xml_carpetas(factura.factura_xml.path, f"V{viatico.folio}", fecha_subida, viatico.distrito.nombre, beneficiario, gen_path, factura))
                            if viatico.id not in processed_docs:
                                pdf_buf = generar_pdf_viatico(viatico.id)
                                zip_file.writestr(os.path.join(carpeta, f'VIATICO_{viatico.folio}.pdf'), pdf_buf.getvalue())
                                processed_docs.add(viatico.id)

                        if pago.comprobante_pago and carpeta is not None:
                            fecha_pago = pago.pagado_real.strftime('%Y-%m-%d') if pago.pagado_real else 'SIN_FECHA'
                            if pago.gasto:
                                folio = f'G{pago.gasto.folio}'
                                if pago.gasto.colaborador:
                                    pago_nombre = f'{pago.gasto.colaborador.staff.staff.first_name}_{pago.gasto.colaborador.staff.staff.last_name}'
                                else:
                                    pago_nombre = f'{pago.gasto.staff.staff.staff.first_name}_{pago.gasto.staff.staff.staff.last_name}'
                            elif pago.oc:
                                folio = f'OC{oc.folio}'
                                pago_nombre = f'{pago.oc.proveedor.nombre.razon_social}'
                            elif pago.viatico:
                                folio = f'V{viatico.folio}'
                                if pago.viatico.colaborador:
                                    pago_nombre = f'{pago.viatico.colaborador.staff.staff.first_name}_{pago.viatico.colaborador.staff.staff.last_name}'
                                else:
                                    pago_nombre = f'{pago.viatico.staff.staff.staff.first_name}_{pago.viatico.staff.staff.staff.last_name}'
                            
                            monto = f"{pago.monto:.2f}".replace('.', '_')
                            nuevo_nombre = f'{fecha_pago}_{folio}_{pago_nombre}_{monto}'
                           
                            zip_file.write(pago.comprobante_pago.path, os.path.join(carpeta, f'{nuevo_nombre}.pdf'))

                        # Excel de resumen
                    output = BytesIO()
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Resumen XML"
                    columnas = ['Distrito','Folio','Fecha subida','Fecha factura', 'Raz√≥n Social', 'Folio Fiscal (UUID)', 
                                'Monto Total Factura', 'Tipo de Moneda', 'Forma de pago','M√©todo de Pago',
                                'Receptor (Empresa) Nombre', 'Beneficiario', 'Archivo', 'Tipo de Documento','Fecha Validaci√≥n SAT', 'EstadoSAT']
                    ws.append(columnas)
                    for dato in datos_xml_lista:
                        ws.append([dato.get(col, '') for col in columnas])
                    for col in ['G']:
                        for row in range(2, ws.max_row + 1):
                            ws[f"{col}{row}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                    wb.save(output)
                    zip_file.writestr("GENERAL_XMLs/reporte_facturas.xlsx", output.getvalue())

                zip_buffer.seek(0)
                response = HttpResponse(zip_buffer, content_type='application/zip')
                response.set_cookie('descarga_iniciada', 'true', max_age=20)
                response['Content-Disposition'] = 'attachment; filename=pagos.zip'
                return response
        
        elif 'btnDescargar' in request.POST:
            validar_sat = request.POST.get('validacion') == 'on'
            fecha_inicio = parse_date(request.POST.get('fecha_inicio'))
            fecha_fin = parse_date(request.POST.get('fecha_fin'))
            distrito_id = request.POST.get('distrito')
            tesorero_id = request.POST.get('tesorero')
            folio = request.POST.get('folio')
            print(folio)
            tipo_documento = request.POST.get('tipo_documento')

            facturas_gastos = Factura.objects.none()
            facturas_compras = Facturas.objects.none()
            facturas_viaticos = Viaticos_Factura.objects.none()
            
            if usuario.distritos.nombre == "MATRIZ":
                if tipo_documento in ["", "gastos"]:
                    facturas_gastos = Factura.objects.filter(Q(solicitud_gasto__pagosg__pagado_real__range=[fecha_inicio, fecha_fin])|Q(solicitud_gasto__pagosg__pagado_date__range=[fecha_inicio, fecha_fin]))
                if tipo_documento in ["", "compras"]:    
                    facturas_compras = Facturas.objects.filter(Q(oc__pagos__pagado_real__range=[fecha_inicio, fecha_fin])|Q(oc__pagos__pagado_date__range=[fecha_inicio, fecha_fin]), hecho = True)
                if tipo_documento in ["", "viaticos"]:
                    facturas_viaticos = Viaticos_Factura.objects.filter(Q(solicitud_viatico__pagosv__pagado_real__range=[fecha_inicio, fecha_fin])|Q(solicitud_viatico__pagosv__pagado_date__range=[fecha_inicio, fecha_fin]))

                if distrito_id:
                    facturas_gastos = facturas_gastos.filter(solicitud_gasto__distrito_id=distrito_id)
                    facturas_compras = facturas_compras.filter(oc__req__orden__distrito_id=distrito_id)
                    facturas_viaticos = facturas_viaticos.filter(solicitud_viatico__distrito_id=distrito_id)

                if tesorero_id:
                    facturas_gastos = facturas_gastos.filter(solicitud_gasto__pagosg__tesorero__id=tesorero_id)
                    facturas_compras = facturas_compras.filter(oc__pagos__tesorero__id=tesorero_id)
                    facturas_viaticos = facturas_viaticos.filter(solicitud_viatico__pagosv__tesorero__id=tesorero_id)
                
                if folio:
                    #print('folio',int(folio))
                    #viatico = Solicitud_Viatico.objects.get(folio = folio)
                    facturas_gastos = Factura.objects.filter(solicitud_gasto__folio = folio)
                    facturas_compras = Facturas.objects.filter(oc__folio= folio)
                    facturas_viaticos = Viaticos_Factura.objects.filter(solicitud_viatico__folio = folio)
                    #print('viatico',viatico)
                    #print('facturas_gastos',facturas_gastos)
                    #print('facturas_compras',facturas_compras)
                    #print('facturas_viaticos',facturas_viaticos)

            else:
                facturas_gastos = Factura.objects.filter(solicitud_gasto__approbado_fecha2__range=[fecha_inicio, fecha_fin], solicitud_gasto__distrito = usuario.distritos)
                facturas_compras = Facturas.objects.filter(oc__autorizado_at_2__range=[fecha_inicio, fecha_fin], oc__req__orden__distrito = usuario.distritos)
                facturas_viaticos = Viaticos_Factura.objects.filter(solicitud_viatico__approved_at2__range=[fecha_inicio, fecha_fin], solicitud_viatico__distrito = usuario.distritos)

            if validar_sat:
                ids_gastos = list(facturas_gastos.values_list('id', flat=True))
                ids_compras = list(facturas_compras.values_list('id', flat=True))
                ids_viaticos = list(facturas_viaticos.values_list('id', flat=True))
                #print(ids_viaticos)

                validar_lote_facturas.delay(ids_gastos, ids_compras, ids_viaticos)
                
                
                #for factura in facturas_gastos:
                #    if factura.archivo_xml:
                #        extraer_datos_validacion(factura.archivo_xml.path, factura)
                #for factura in facturas_compras:
                #    if factura.factura_xml:
                #        extraer_datos_validacion(factura.factura_xml.path, factura)
                
                #for factura in facturas_viaticos:
                #    if factura.factura_xml:
                #        extraer_datos_validacion(factura.factura_xml.path, factura)  

            else:
                zip_buffer = BytesIO()
                processed_ocs = set()  # Mant√©n un conjunto de OCs procesadas
                processed_gastos = set()  # Mant√©n un conjunto de gastos procesados
                processed_viaticos = set()  # Mant√©n un conjunto de vi√°ticos procesados
                processed_pagos = set()  # Mant√©n un conjunto de pagos procesados
                datos_xml_lista = []  # Lista para el resumen en Excel

                with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
                    #Se agrgean carpetas generales
                    general_pdfs_folder = "GENERAL_PDFs"
                    general_xmls_folder = "GENERAL_XMLs"
                    
                    #Se procesan facturas de gastos
                    for factura in facturas_gastos:
                        folder_name = f'GASTO_{factura.solicitud_gasto.folio}_{factura.solicitud_gasto.distrito.nombre}'
                        if factura.archivo_pdf:   
                            file_name = os.path.basename(factura.archivo_pdf.path)
                            zip_file.write(factura.archivo_pdf.path, os.path.join(folder_name, file_name))
                            if factura.archivo_xml:
                                # Guardar en la carpeta GENERAL_PDFs con nombre id_uuid.pdf
                                uuid_str = factura.uuid if factura.uuid else 'SIN_UUID'
                                general_file_name = f'{factura.id}_{uuid_str}.pdf'
                            else:
                                general_file_name = file_name
                            zip_file.write(factura.archivo_pdf.path, os.path.join(general_pdfs_folder, general_file_name)) #Est√° l√≠nea guarda en el zip general de pdf
                        
                        beneficiario = factura.solicitud_gasto.colaborador.staff.staff.first_name + ' ' + factura.solicitud_gasto.colaborador.staff.staff.last_name  if factura.solicitud_gasto.colaborador else factura.solicitud_gasto.staff.staff.staff.first_name + ' ' + factura.solicitud_gasto.staff.staff.staff.last_name
                        distrito = factura.solicitud_gasto.distrito.nombre  # Obtener distrito de la factura
                        folio = 'G' + str(factura.solicitud_gasto.folio)
                        fecha_subida = factura.fecha_subida.astimezone(tz=None).replace(tzinfo=None) if factura.fecha_subida else 'No disponible'
                        if factura.archivo_xml:
                            file_name = os.path.basename(factura.archivo_xml.path)
                            zip_file.write(factura.archivo_xml.path, os.path.join(folder_name, file_name))
                            uuid_str = factura.uuid if factura.uuid else 'SIN_UUID'
                            general_file_name = f'{factura.id}_{uuid_str}.xml'

                            zip_file.write(factura.archivo_xml.path, os.path.join(general_xmls_folder, general_file_name)) #Est√° l√≠nea guarda en el zip general de xml's
                            datos_xml_lista.append(extraer_datos_xml_carpetas(factura.archivo_xml.path, folio, fecha_subida, distrito, beneficiario, general_file_name, factura))

                       

                        if factura.solicitud_gasto.id not in processed_gastos:
                            buf = render_pdf_gasto(factura.solicitud_gasto.id)
                            gasto_file_name = f'GASTO_{factura.solicitud_gasto.folio}.pdf'
                            zip_file.writestr(os.path.join(folder_name, gasto_file_name), buf.getvalue())
                            processed_gastos.add(factura.solicitud_gasto.id)

                        pagos = Pago.objects.filter(gasto=factura.solicitud_gasto)
                        for pago in pagos:
                            if pago.comprobante_pago and pago.id not in processed_pagos:
                                texto_pago = extraer_texto_pdf_prop(pago.comprobante_pago)
                                variables_pago = encontrar_variables(texto_pago)

                                fecha_pago = variables_pago.get('fecha', '').replace('/', '-')
                                if not pago.pagado_real:
                                    if fecha_pago:
                                        pago.pagado_real = fecha_pago
                                        pago.save()
                                titular_cuenta_2 = variables_pago.get('titular_cuenta_2', '')
                                importe_operacion = variables_pago.get('importe_operacion', '').split('.')[0].replace(',', '')

                                # Validamos si todas las variables son v√°lidas:
                                if fecha_pago and fecha_pago != 'No disponible' and titular_cuenta_2 and titular_cuenta_2 != 'No disponible' and importe_operacion and importe_operacion != 'No disponible':
                                    pago_file_name = f'{fecha_pago} {titular_cuenta_2} ${importe_operacion}.pdf'
                                else:
                                    # Si no, conservamos el nombre original
                                    pago_file_name = os.path.basename(pago.comprobante_pago.path)
                                
                                #pago_file_name = os.path.basename(pago.comprobante_pago.path)
                                zip_file.write(pago.comprobante_pago.path, os.path.join(folder_name, f'{pago_file_name}'))
                                processed_pagos.add(pago.id)
                    
                    #Se procesan facturas de compras
                    for factura in facturas_compras:
                        folder_name = f'COMPRA_{factura.oc.folio}_{factura.oc.req.orden.distrito.nombre}'
                        if factura.factura_pdf:
                            #folder_name = f'COMPRA_{factura.oc.folio}_{factura.oc.req.orden.distrito.nombre}'
                            file_name = os.path.basename(factura.factura_pdf.path)
                            zip_file.write(factura.factura_pdf.path, os.path.join(folder_name, file_name))
                            if factura.factura_xml:
                                # Guardar en la carpeta GENERAL_PDFs con nombre id_uuid.pdf
                                uuid_str = factura.uuid if factura.uuid else 'SIN_UUID'
                                general_file_name = f'{factura.id}_{uuid_str}.pdf'
                            else:
                                general_file_name = file_name

                            zip_file.write(factura.factura_pdf.path, os.path.join(general_pdfs_folder, file_name))

                        beneficiario = "NA"
                        distrito = factura.oc.req.orden.distrito.nombre  # Obtener distrito de la factura
                        folio = factura.oc.folio
                        fecha_subida = factura.fecha_subido if factura.fecha_subido else 'No disponible'
                        if factura.factura_xml:
                            file_name = os.path.basename(factura.factura_xml.path)
                            zip_file.write(factura.factura_xml.path, os.path.join(folder_name, file_name))
                            uuid_str = factura.uuid if factura.uuid else 'SIN_UUID'
                            general_file_name = f'{factura.id}_{uuid_str}.xml'
                            
                            zip_file.write(factura.factura_xml.path, os.path.join(general_xmls_folder, general_file_name))
                            datos_xml_lista.append(extraer_datos_xml_carpetas(factura.factura_xml.path, folio, fecha_subida, distrito, beneficiario, general_file_name, factura))
                        
                        # Incluir la ficha de pago
                        pagos = Pago.objects.filter(oc=factura.oc)
                        for pago in pagos:
                            if pago.comprobante_pago and pago.id not in processed_pagos:
                                texto_pago = extraer_texto_pdf_prop(pago.comprobante_pago)
                                variables_pago = encontrar_variables(texto_pago)
                                
                                fecha_obj = variables_pago.get('fecha')
                                fecha_pago = ''
                                if fecha_obj:
                                # Si es objeto datetime.date, lo convertimos a string con formato
                                    if isinstance(fecha_obj, datetime.date):
                                        fecha_pago = fecha_obj.strftime('%d-%m-%Y')
                                    else:
                                        # Si ya es cadena (por ejemplo, por error de extracci√≥n)
                                        fecha_pago = str(fecha_obj).replace('/', '-')
                                titular_cuenta_2 = variables_pago.get('titular_cuenta_2', '')
                                importe_operacion = variables_pago.get('importe_operacion', '').split('.')[0].replace(',', '')

                                if not pago.pagado_real:
                                    if fecha_pago:
                                        pago.pagado_real = fecha_pago
                                        pago.save()


                                if fecha_pago and fecha_pago != 'No disponible' and titular_cuenta_2 and titular_cuenta_2 != 'No disponible' and importe_operacion and importe_operacion != 'No disponible':
                                    pago_file_name = f'{fecha_pago} {titular_cuenta_2} ${importe_operacion}.pdf'
                                else:
                                    pago_file_name = os.path.basename(pago.comprobante_pago.path)
                                #pago_file_name = os.path.basename(pago.comprobante_pago.path)

                                zip_file.write(pago.comprobante_pago.path, os.path.join(folder_name, f'{pago_file_name}'))
                                processed_pagos.add(pago.id) 
                        
                        # Generar e incluir la OC en el ZIP solo si no ha sido procesada
                        if factura.oc.id not in processed_ocs:
                            buf = generar_pdf(factura.oc)
                            oc_file_name = f'OC_{factura.oc.folio}.pdf'
                            zip_file.writestr(os.path.join(folder_name, oc_file_name), buf.getvalue())
                            processed_ocs.add(factura.oc.id)
                        
                      
                    

                    for factura in facturas_viaticos:
                        folder_name = f'VIATICO_{factura.solicitud_viatico.folio}_{factura.solicitud_viatico.distrito.nombre}'
                        if factura.factura_pdf:

                            file_name = os.path.basename(factura.factura_pdf.path)
                            zip_file.write(factura.factura_pdf.path, os.path.join(folder_name, file_name))
                            if factura.factura_xml:
                                # Guardar en la carpeta GENERAL_PDFs con nombre id_uuid.pdf
                                uuid_str = factura.uuid if factura.uuid else 'SIN_UUID'
                                general_file_name = f'{factura.id}_{uuid_str}.pdf'
                            else:
                                general_file_name = file_name
                            zip_file.write(factura.factura_pdf.path, os.path.join(general_pdfs_folder, general_file_name))

                        beneficiario = factura.solicitud_viatico.colaborador.staff.staff.first_name + ' ' + factura.solicitud_viatico.colaborador.staff.staff.last_name  if factura.solicitud_viatico.colaborador else factura.solicitud_gasto.staff.staff.staff.first_name + ' ' + factura.solicitud_gasto.staff.staff.staff.last_name
                        distrito = factura.solicitud_viatico.distrito.nombre  # Obtener distrito de la factura
                        folio = 'V' + str(factura.solicitud_viatico.folio)
                        fecha_subida = factura.fecha_subido.astimezone(tz=None).replace(tzinfo=None) if factura.fecha_subido else 'No disponible' # Formato YYYY-MM-DD
                        if factura.factura_xml:
                            file_name = os.path.basename(factura.factura_xml.path)
                            zip_file.write(factura.factura_xml.path, os.path.join(folder_name, file_name))
                            uuid_str = factura.uuid if factura.uuid else 'SIN_UUID'
                            general_file_name = f'{factura.id}_{uuid_str}.xml'

                            zip_file.write(factura.factura_xml.path, os.path.join(general_xmls_folder, general_file_name))
                            datos_xml_lista.append(extraer_datos_xml_carpetas(factura.factura_xml.path, folio, fecha_subida, distrito, beneficiario, general_file_name, factura))

                        if factura.solicitud_viatico.id not in processed_viaticos:
                            buf = generar_pdf_viatico(factura.solicitud_viatico.id)
                            viatico_file_name = f'VIATICO_{factura.solicitud_viatico.folio}.pdf'
                            zip_file.writestr(os.path.join(folder_name, viatico_file_name), buf.getvalue())
                            processed_viaticos.add(factura.solicitud_viatico.id)

                    
                        
                        pagos = Pago.objects.filter(viatico=factura.solicitud_viatico)
                        for pago in pagos:
                            if pago.comprobante_pago and pago.id not in processed_pagos:
                                texto_pago = extraer_texto_pdf_prop(pago.comprobante_pago)
                                variables_pago = encontrar_variables(texto_pago)

                                fecha_obj = variables_pago.get('fecha')
                                fecha_pago = ''

                                if fecha_obj:
                                    # Si es objeto datetime.date, lo convertimos a string con formato
                                    if isinstance(fecha_obj, datetime.date):
                                        fecha_pago = fecha_obj.strftime('%d-%m-%Y')
                                    else:
                                        # Si ya es cadena (por ejemplo, por error de extracci√≥n)
                                        fecha_pago = str(fecha_obj).replace('/', '-')
                                titular_cuenta_2 = variables_pago.get('titular_cuenta_2', '')
                                importe_operacion = variables_pago.get('importe_operacion', '').split('.')[0].replace(',', '')

                                if fecha_pago and fecha_pago != 'No disponible' and titular_cuenta_2 and titular_cuenta_2 != 'No disponible' and importe_operacion and importe_operacion != 'No disponible':
                                    pago_file_name = f'{fecha_pago} {titular_cuenta_2} ${importe_operacion}.pdf'
                                else:
                                    pago_file_name = os.path.basename(pago.comprobante_pago.path)

                                if not pago.pagado_real:
                                    if fecha_pago:
                                        pago.pagado_real = fecha_pago
                                        pago.save()
                                
                                #pago_file_name = os.path.basename(pago.comprobante_pago.path)
                                zip_file.write(pago.comprobante_pago.path, os.path.join(folder_name, f'{pago_file_name}'))
                                processed_pagos.add(pago.id)

                    # Crear archivo Excel con los datos extra√≠dos
                    output = BytesIO()
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Resumen XML"

                    columnas = ['Distrito','Folio','Fecha subida','Fecha factura', 'Raz√≥n Social', 'Folio Fiscal (UUID)', 
                                'Monto Total Factura', 'Tipo de Moneda', 'Forma de pago','M√©todo de Pago',
                                'Receptor (Empresa) Nombre', 'Beneficiario', 'Archivo', 'Tipo de Documento','Fecha Validaci√≥n SAT', 'EstadoSAT'
                                ]
                    ws.append(columnas)

                    for dato in datos_xml_lista:
                        ws.append([dato.get(col, '') for col in columnas])

                    # Formatos de Excel
                    for col in ['G']:  # Monto Total Factura
                        for row in range(2, ws.max_row + 1):
                            ws[f"{col}{row}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

                    wb.save(output)
                    zip_file.writestr("GENERAL_XMLs/reporte_facturas.xlsx", output.getvalue())

                zip_buffer.seek(0)
                response = HttpResponse(zip_buffer, content_type='application/zip')
                response.set_cookie('descarga_iniciada', 'true', max_age=20)
                response['Content-Disposition'] = 'attachment; filename=facturas.zip'
                return response
        elif 'validacion' in request.POST:
            pago_ids = request.POST.getlist('pago_ids')
            pagos = Pago.objects.filter(id__in=pago_ids)

            for pago in pagos:
                facturas = []
                if pago.gasto:
                    facturas = pago.gasto.facturas.filter(hecho=True)
                elif pago.oc:
                    facturas = pago.oc.facturas.filter(hecho=True)
                elif pago.viatico:
                    facturas = pago.viatico.facturas.filter(hecho=True)

                for factura in facturas:
                    if not factura.autorizada:
                        factura.autorizada = True
                        factura.autorizada_el = timezone.now()  # Opcional
                        factura.save()

            messages.success(request, f'{len(pago_ids)} pagos validados correctamente.')
            return redirect('matriz-pagos') 
        elif 'btnImprimir' in request.POST:
            pago_ids = request.POST.getlist('pago_ids')
            pagos = Pago.objects.filter(id__in=pago_ids)

            if not pagos.exists():
                return HttpResponse("No se seleccionaron pagos v√°lidos.", content_type="text/plain")

            merger = PdfMerger()

            for pago in pagos:
                # 1. Comprobante de pago
                if pago.comprobante_pago and os.path.exists(pago.comprobante_pago.path):
                    merger.append(pago.comprobante_pago.path, import_outline=False)

                # 2. Car√°tula + facturas
                if pago.gasto:
                    buffer = render_pdf_gasto(pago.gasto.id)
                    facturas = pago.gasto.facturas.filter(hecho=True)
                elif pago.oc:
                    buffer = generar_pdf(pago.oc)
                    facturas = pago.oc.facturas.filter(hecho=True)
                elif pago.viatico:
                    buffer = generar_pdf_viatico(pago.viatico.id)
                    facturas = pago.viatico.facturas.filter(hecho=True)
                else:
                    buffer = None
                    facturas = []

                # Car√°tula (guardar buffer en archivo temporal)
                if buffer:
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_caratula:
                        temp_caratula.write(buffer.read())
                        temp_caratula.flush()
                        caratula_path = temp_caratula.name
                    merger.append(caratula_path, import_outline=False)

                # Facturas
                for factura in facturas:
                    if pago.gasto and factura.archivo_pdf and os.path.exists(factura.archivo_pdf.path):
                        merger.append(factura.archivo_pdf.path, import_outline=False)
                    elif (pago.oc or pago.viatico) and factura.factura_pdf and os.path.exists(factura.factura_pdf.path):
                        merger.append(factura.factura_pdf.path, import_outline=False)

            # Guardar PDF combinado final en archivo temporal
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_final:
                merger.write(temp_final.name)
                temp_file_path = temp_final.name

            merger.close()

            # Guardar ruta del PDF final en la sesi√≥n
            request.session['temp_pdf_path'] = temp_file_path
            return redirect('mostrar-pdf')
        elif 'enviar_a_control' in request.POST:
            ids = request.POST.getlist('compra_ids')
            if ids:
                pagos = Pago.objects.filter(id__in=ids)
                for pago in pagos:
                    pago.control_documentos = True
                    pago.fecha_control_documentos = datetime.today()
                    pago.save()

            return redirect('matriz-pagos')  
            
           
           
    for pago in pagos_list:
        if pago.total_facturas == 0:
            pago.estado_facturas = 'sin_facturas'
        elif pago.autorizadas == pago.total_facturas:
            pago.estado_facturas = 'todas_autorizadas'
        else:
            pago.estado_facturas = 'pendientes'
        
       
        
    context= {
        'pagos_list':pagos_list,
        'pagos':pagos,
        'myfilter':myfilter,
        'tesoreros':tesoreros,
        'distritos':distritos,
        'usuario':usuario,
        }

    return render(request, 'tesoreria/matriz_pagos.html',context)


@perfil_seleccionado_required
def control_documentos(request):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    almacenes_distritos = set(usuario.almacen.values_list('distrito__id', flat=True))
    exhibits = Exhibit.objects.all().order_by('-created_at')  # o alg√∫n filtro √∫til

    pagos = Pago.objects.filter(
        Q(oc__req__orden__distrito__in =almacenes_distritos) & Q(oc__autorizado2=True) | 
        Q(viatico__distrito__in = almacenes_distritos) & Q(viatico__autorizar2=True) |
        Q(gasto__distrito__in = almacenes_distritos) & Q(gasto__autorizar2 = True), 
        control_documentos = True,
        hecho=True
        ).annotate(
        # Detectar la relaci√≥n que tiene facturas
        total_facturas=Count(
            'oc__facturas', filter=Q(oc__facturas__hecho=True)
        ) + Count(
            'gasto__facturas__hecho', filter=Q(gasto__facturas__hecho=True)
        ) + Count(
            'viatico__facturas__hecho', filter=Q(viatico__facturas__hecho=True)
        ),
        autorizadas=Count(
            Case(
                When(Q(oc__facturas__autorizada=True, oc__facturas__hecho=True), then=Value(1))
            )
        ) + Count(
            Case(
                When(Q(gasto__facturas__autorizada=True, gasto__facturas__hecho=True), then=Value(1))
            )
        ) + Count(
            Case(
                When(Q(viatico__facturas__autorizada=True, viatico__facturas__hecho=True), then=Value(1))
            )
        ),
        ).order_by('-pagado_real')
    myfilter = Matriz_Pago_Filter(request.GET, queryset=pagos)
    pagos = myfilter.qs
    #Los distritos se definen de forma "din√°mica" de acuerdo a los almacenes que tiene el usuario en el perfil
    distritos = Distrito.objects.filter(id__in=almacenes_distritos)
    tesoreros = Profile.objects.filter(tipo__nombre__in = ["Tesoreria","Tesoreria_Documentos","Admin" ], st_activo = True, distritos__in = almacenes_distritos)
    cuentas = Cuenta.objects.filter(distrito__in=almacenes_distritos)
    empresas = Empresa.objects.all()
    #Set up pagination
    p = Paginator(pagos, 50)
    page = request.GET.get('page')
    pagos_list = p.get_page(page)
    

    if request.method == 'POST': 
        if 'btnReporte' in request.POST:
            return convert_excel_matriz_pagos(pagos)
        elif 'btnReporteXML' in request.POST:
            fecha_inicio = parse_date(request.POST.get('fecha_inicio'))
            fecha_fin = parse_date(request.POST.get('fecha_fin'))
            distrito_id = request.POST.get('distrito')
            tesorero_id = request.POST.get('tesorero')
            folio = request.POST.get('folio')

            if usuario.distritos.nombre == "MATRIZ":
                pagos = Pago.objects.filter(hecho=True)
                if fecha_inicio and fecha_fin:
                    pagos = Pago.objects.filter(Q(pagado_real__range=[fecha_inicio, fecha_fin])|Q(pagado_date__range=[fecha_inicio, fecha_fin]), hecho = True)
              
                    if distrito_id:
                        pagos = pagos.filter(
                            Q(gasto__distrito_id=distrito_id) |
                            Q(oc__req__orden__distrito_id=distrito_id) |
                            Q(viatico__distrito_id=distrito_id)
                        )

                    if tesorero_id:
                        pagos = pagos.filter(tesorero_id=tesorero_id)

                if folio:
                    pagos = Pago.objects.filter(hecho = True)
                    pagos = pagos.filter(
                        Q(gasto__folio=folio) |
                        Q(oc__folio=folio) |
                        Q(viatico__folio=folio)
                    )
                    if distrito_id:
                        pagos = pagos.filter(
                            Q(gasto__distrito_id=distrito_id) |
                            Q(oc__req__orden__distrito_id=distrito_id) |
                            Q(viatico__distrito_id=distrito_id)
                        )

                    if tesorero_id:
                        pagos = pagos.filter(tesorero_id=tesorero_id)
            else:
                pagos = pagos.filter(
                    Q(pagado_real__range=[fecha_inicio, fecha_fin])|Q(pagado_date__range=[fecha_inicio, fecha_fin]),
                    Q(gasto__distrito=usuario.distritos) |
                    Q(oc__req__orden__distrito=usuario.distritos) |
                    Q(viatico__distrito=usuario.distritos)
                )
             
            datos_xml_lista = []
            for pago in pagos:
                if pago.gasto:
                    gasto = pago.gasto
                    for factura in gasto.facturas.all():
                        beneficiario = factura.solicitud_gasto.colaborador.staff.staff.first_name + ' ' + factura.solicitud_gasto.colaborador.staff.staff.last_name  if factura.solicitud_gasto.colaborador else factura.solicitud_gasto.staff.staff.staff.first_name + ' ' + factura.solicitud_gasto.staff.staff.staff.last_name
                        fecha_subida = factura.fecha_subida.astimezone(tz=None).replace(tzinfo=None) if factura.fecha_subida else 'No disponible'      
                        if factura.archivo_xml:     
                            datos_xml_lista.append(extraer_datos_xml_carpetas(factura.archivo_xml.path, f"G{gasto.folio}", fecha_subida, gasto.distrito.nombre, beneficiario, "NA", factura))    
                elif pago.oc:
                    oc = pago.oc
                    for factura in oc.facturas.all():
                        if factura.factura_xml:
                            datos_xml_lista.append(extraer_datos_xml_carpetas(factura.factura_xml.path, f"OC{oc.folio}", factura.fecha_subido, oc.req.orden.distrito.nombre, "NA", "NA", factura))
                elif pago.viatico:
                    viatico = pago.viatico
                    for factura in viatico.facturas.all():
                        beneficiario = factura.solicitud_viatico.colaborador.staff.staff.first_name + ' ' + factura.solicitud_viatico.colaborador.staff.staff.last_name  if factura.solicitud_viatico.colaborador else factura.solicitud_gasto.staff.staff.staff.first_name + ' ' + factura.solicitud_gasto.staff.staff.staff.last_name
                        fecha_subida = factura.fecha_subido.astimezone(tz=None).replace(tzinfo=None) if factura.fecha_subido else 'No disponible' # Formato YYYY-MM-DD
                        if factura.factura_xml:
                            datos_xml_lista.append(extraer_datos_xml_carpetas(factura.factura_xml.path, f"V{viatico.folio}", fecha_subida, viatico.distrito.nombre, beneficiario, "NA", factura))
            output = generar_excel_xmls(datos_xml_lista)
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=reporte_facturas.xlsx'
            response.set_cookie('descarga_iniciada', 'true', max_age=20)
            return response
                       
        elif 'btnDescargarFacturas' in request.POST:
            fecha_inicio = parse_date(request.POST.get('fecha_inicio'))
            fecha_fin = parse_date(request.POST.get('fecha_fin'))
            distrito_id = request.POST.get('distrito')
            tesorero_id = request.POST.get('tesorero')
            empresa_id = request.POST.get('empresa')
            tipo_documento_id = request.POST.get('tipo_documento')
            cuenta_bancaria_id = request.POST.get('cuenta_bancaria')
            folio = request.POST.get('folio')
            validar_sat = request.POST.get('validacion') == 'on'
            

            if usuario.distritos.nombre == "MATRIZ":
                pagos = Pago.objects.filter(hecho=True)
                if fecha_inicio and fecha_fin:
                    pagos = Pago.objects.filter(Q(pagado_real__range=[fecha_inicio, fecha_fin])|Q(pagado_date__range=[fecha_inicio, fecha_fin]), hecho = True, control_documentos = True)

                    if cuenta_bancaria_id:
                        pagos = pagos.filter(cuenta__id = cuenta_bancaria_id)
                    
                    if empresa_id:
                        pagos = pagos.filter(cuenta__empresa__id = empresa_id)

                    if distrito_id:
                        pagos = pagos.filter(
                            Q(gasto__distrito_id=distrito_id) |
                            Q(oc__req__orden__distrito_id=distrito_id) |
                            Q(viatico__distrito_id=distrito_id)
                        )

                    if tesorero_id:
                        pagos = pagos.filter(tesorero_id=tesorero_id)

                if folio:
                    pagos = Pago.objects.filter(hecho = True)
                    pagos = pagos.filter(
                        Q(gasto__folio=folio) |
                        Q(oc__folio=folio) |
                        Q(viatico__folio=folio)
                    )
                    if distrito_id:
                        pagos = pagos.filter(
                            Q(gasto__distrito_id=distrito_id) |
                            Q(oc__req__orden__distrito_id=distrito_id) |
                            Q(viatico__distrito_id=distrito_id)
                        )

                    if tesorero_id:
                        pagos = pagos.filter(tesorero_id=tesorero_id)
            else:
                pagos = pagos.filter(
                    Q(pagado_real__range=[fecha_inicio, fecha_fin])|Q(pagado_date__range=[fecha_inicio, fecha_fin]),
                    Q(gasto__distrito=usuario.distritos) |
                    Q(oc__req__orden__distrito=usuario.distritos) |
                    Q(viatico__distrito=usuario.distritos),
                    control_documentos = True
                )
            


            if validar_sat:
                ids_gastos = set()
                ids_compras = set()
                ids_viaticos = set()
                #print(f"Pagos: {pagos.count()}")
                for pago in pagos:
                    if pago.gasto:
                        ids_gastos.update(pago.gasto.facturas.values_list('id', flat=True))
                    elif pago.oc:
                        ids_compras.update(pago.oc.facturas.values_list('id', flat=True))
                    elif pago.viatico:
                        ids_viaticos.update(pago.viatico.facturas.values_list('id', flat=True))
                print(ids_gastos, ids_compras, ids_viaticos)
                validar_lote_facturas.delay(list(ids_gastos), list(ids_compras), list(ids_viaticos))
            else:
                zip_buffer = BytesIO()
                datos_xml_lista = []
                processed_docs = set()

                with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
                    #zip_file.mkdir("GENERAL_PDFs")
                    #zip_file.mkdir("GENERAL_XMLs")

                    for pago in pagos:
                        if pago.gasto:
                            gasto = pago.gasto
                            texto_pago = extraer_texto_pdf_prop(pago.comprobante_pago)
                            variables_pago = encontrar_variables(texto_pago)
                            fecha_str = variables_pago.get('fecha')
                            
                            fecha_pago = ''

                            if not pago.pagado_real:
                                fecha_obj = datetime.strptime(fecha_str, '%d/%m/%Y').date()
                                if fecha_obj:
                                    if isinstance(fecha_obj, date):
                                        fecha_pago = fecha_obj.strftime('%d-%m-%Y')  # Para usar en nombre de archivo, etc.
                                    else:
                                        fecha_pago = str(fecha_obj).replace('/', '-')
                                    pago.pagado_real = fecha_obj
                                    pago.save()
                            carpeta = f'{pago.pagado_real}_GASTO_{gasto.folio}_{gasto.distrito.nombre}'
                            #zip_file.mkdir(carpeta)
                            for factura in gasto.facturas.all():
                                beneficiario = factura.solicitud_gasto.colaborador.staff.staff.first_name + ' ' + factura.solicitud_gasto.colaborador.staff.staff.last_name  if factura.solicitud_gasto.colaborador else factura.solicitud_gasto.staff.staff.staff.first_name + ' ' + factura.solicitud_gasto.staff.staff.staff.last_name
                                fecha_subida = factura.fecha_subida.astimezone(tz=None).replace(tzinfo=None) if factura.fecha_subida else 'No disponible'
                                if factura.archivo_pdf:
                                    zip_file.write(factura.archivo_pdf.path, os.path.join(carpeta, os.path.basename(factura.archivo_pdf.path)))
                                    uuid = factura.uuid if factura.uuid else 'SIN_UUID'
                                    zip_file.write(factura.archivo_pdf.path, f"GENERAL_PDFs/{factura.id}_{uuid}.pdf")
                                if factura.archivo_xml:
                                    zip_file.write(factura.archivo_xml.path, os.path.join(carpeta, os.path.basename(factura.archivo_xml.path)))
                                    uuid = factura.uuid if factura.uuid else 'SIN_UUID'
                                    gen_path = f"GENERAL_XMLs/{factura.id}_{uuid}.xml"
                                    zip_file.write(factura.archivo_xml.path, gen_path)
                                    datos_xml_lista.append(extraer_datos_xml_carpetas(factura.archivo_xml.path, f"G{gasto.folio}", fecha_subida, gasto.distrito.nombre, beneficiario, gen_path, factura))
                                    if not factura.archivo_pdf or not os.path.exists(factura.archivo_pdf.path):
                                        # Si no hay PDF, generamos uno
                                        ruta_pdf = crear_pdf_cfdi_gasto(factura)
                                        zip_file.write(ruta_pdf, os.path.join(carpeta, os.path.basename(ruta_pdf)))
                                        uuid = factura.uuid if factura.uuid else 'SIN_UUID'
                                        zip_file.write(ruta_pdf, f"GENERAL_PDFs/{factura.id}_{uuid}.pdf")

                            if gasto.id not in processed_docs:
                                pdf_buf = render_pdf_gasto(gasto.id)
                                zip_file.writestr(os.path.join(carpeta, f'GASTO_{gasto.folio}.pdf'), pdf_buf.getvalue())
                                processed_docs.add(gasto.id)
                        elif pago.oc:
                            oc = pago.oc
                            if not pago.pagado_real:
                                fecha_obj = datetime.strptime(fecha_str, '%d/%m/%Y').date()
                                if fecha_obj:
                                    if isinstance(fecha_obj, date):
                                        fecha_pago = fecha_obj.strftime('%d-%m-%Y')  # Para usar en nombre de archivo, etc.
                                    else:
                                        fecha_pago = str(fecha_obj).replace('/', '-')
                                    pago.pagado_real = fecha_obj
                                    pago.save()
                            carpeta = f'{pago.pagado_real}_COMPRA_{oc.folio}_{oc.req.orden.distrito.nombre}'
                            #zip_file.mkdir(carpeta)
                            for factura in oc.facturas.all():
                                if factura.factura_pdf:
                                    zip_file.write(factura.factura_pdf.path, os.path.join(carpeta, os.path.basename(factura.factura_pdf.path)))
                                    uuid = factura.uuid if factura.uuid else 'SIN_UUID'
                                    zip_file.write(factura.factura_pdf.path, f"GENERAL_PDFs/{factura.id}_{uuid}.pdf")
                                if factura.factura_xml:
                                    zip_file.write(factura.factura_xml.path, os.path.join(carpeta, os.path.basename(factura.factura_xml.path)))
                                    uuid = factura.uuid if factura.uuid else 'SIN_UUID'
                                    gen_path = f"GENERAL_XMLs/{factura.id}_{uuid}.xml"
                                    zip_file.write(factura.factura_xml.path, gen_path)
                                    datos_xml_lista.append(extraer_datos_xml_carpetas(factura.factura_xml.path, f"OC{oc.folio}", factura.fecha_subido, oc.req.orden.distrito.nombre, "NA", gen_path, factura))
                                    if not factura.factura_pdf or not os.path.exists(factura.factura_pdf.path):
                                        # Si no hay PDF, generamos uno
                                        ruta_pdf = crear_pdf_cfdi_buffer(factura)
                                        zip_file.write(ruta_pdf, os.path.join(carpeta, os.path.basename(ruta_pdf)))
                                        uuid = factura.uuid if factura.uuid else 'SIN_UUID'
                                        zip_file.write(ruta_pdf, f"GENERAL_PDFs/{factura.id}_{uuid}.pdf")
                                for complemento in factura.complementos.all():
                                    if complemento.complemento_pdf:     #Encarpeta el complemento_pdf
                                        complemento_file_name = os.path.basename(complemento.complemento_pdf.path)
                                        zip_file.write(complemento.complemento_pdf.path, os.path.join(carpeta, complemento_file_name))
                                    if complemento.complemento_xml:     #Encarpeta el complemento_xml
                                        complemento_file_name = os.path.basename(complemento.complemento_xml.path)
                                        zip_file.write(complemento.complemento_xml.path, os.path.join(carpeta, complemento_file_name))
                            if oc.id not in processed_docs:
                                pdf_buf = generar_pdf(oc)
                                zip_file.writestr(os.path.join(carpeta, f'OC_{oc.folio}.pdf'), pdf_buf.getvalue())
                                processed_docs.add(oc.id)
                                     # üöÄ Incluir los complementos de pago relacionados en la misma carpeta
                              
                        elif pago.viatico:
                            viatico = pago.viatico
                            if not pago.pagado_real:
                                fecha_obj = datetime.strptime(fecha_str, '%d/%m/%Y').date()
                                if fecha_obj:
                                    if isinstance(fecha_obj, date):
                                        fecha_pago = fecha_obj.strftime('%d-%m-%Y')  # Para usar en nombre de archivo, etc.
                                    else:
                                        fecha_pago = str(fecha_obj).replace('/', '-')
                                    pago.pagado_real = fecha_obj
                                    pago.save()
                            carpeta = f'{pago.pagado_real}_VIATICO_{viatico.folio}_{viatico.distrito.nombre}'
                            #zip_file.mkdir(carpeta)
                            for factura in viatico.facturas.all():
                                beneficiario = factura.solicitud_viatico.colaborador.staff.staff.first_name + ' ' + factura.solicitud_viatico.colaborador.staff.staff.last_name  if factura.solicitud_viatico.colaborador else factura.solicitud_gasto.staff.staff.staff.first_name + ' ' + factura.solicitud_gasto.staff.staff.staff.last_name
                                fecha_subida = factura.fecha_subido.astimezone(tz=None).replace(tzinfo=None) if factura.fecha_subido else 'No disponible' # Formato YYYY-MM-DD
                                if factura.factura_pdf:
                                    zip_file.write(factura.factura_pdf.path, os.path.join(carpeta, os.path.basename(factura.factura_pdf.path)))
                                    uuid = factura.uuid if factura.uuid else 'SIN_UUID'
                                    zip_file.write(factura.factura_pdf.path, f"GENERAL_PDFs/{factura.id}_{uuid}.pdf")
                                if factura.factura_xml:
                                    zip_file.write(factura.factura_xml.path, os.path.join(carpeta, os.path.basename(factura.factura_xml.path)))
                                    uuid = factura.uuid if factura.uuid else 'SIN_UUID'
                                    gen_path = f"GENERAL_XMLs/{factura.id}_{uuid}.xml"
                                    zip_file.write(factura.factura_xml.path, gen_path)
                                    datos_xml_lista.append(extraer_datos_xml_carpetas(factura.factura_xml.path, f"V{viatico.folio}", fecha_subida, viatico.distrito.nombre, beneficiario, gen_path, factura))
                                    if not factura.factura_pdf or not os.path.exists(factura.factura_pdf.path):
                                        # Si no hay PDF, generamos uno
                                        ruta_pdf = crear_pdf_cfdi_gasto(factura)
                                        zip_file.write(ruta_pdf, os.path.join(carpeta, os.path.basename(ruta_pdf)))
                                        uuid = factura.uuid if factura.uuid else 'SIN_UUID'
                                        zip_file.write(ruta_pdf, f"GENERAL_PDFs/{factura.id}_{uuid}.pdf")
                            if viatico.id not in processed_docs:
                                pdf_buf = generar_pdf_viatico(viatico.id)
                                zip_file.writestr(os.path.join(carpeta, f'VIATICO_{viatico.folio}.pdf'), pdf_buf.getvalue())
                                processed_docs.add(viatico.id)

                        if pago.comprobante_pago:
                            fecha_pago = pago.pagado_real.strftime('%Y-%m-%d') if pago.pagado_real else 'SIN_FECHA'
                            if pago.gasto:
                                folio = f'G{pago.gasto.folio}'
                                if pago.gasto.colaborador:
                                    pago_nombre = f'{pago.gasto.colaborador.staff.staff.first_name}_{pago.gasto.colaborador.staff.staff.last_name}'
                                else:
                                    pago_nombre = f'{pago.gasto.staff.staff.staff.first_name}_{pago.gasto.staff.staff.staff.last_name}'
                            elif pago.oc:
                                folio = f'OC{oc.folio}'
                                pago_nombre = f'{pago.oc.proveedor.nombre.razon_social}'
                            elif pago.viatico:
                                folio = f'V{viatico.folio}'
                                if pago.viatico.colaborador:
                                    pago_nombre = f'{pago.viatico.colaborador.staff.staff.first_name}_{pago.viatico.colaborador.staff.staff.last_name}'
                                else:
                                    pago_nombre = f'{pago.viatico.staff.staff.staff.first_name}_{pago.viatico.staff.staff.staff.last_name}'
                            
                            monto = f"{pago.monto:.2f}".replace('.', '_')
                            nuevo_nombre = f'{fecha_pago}_{folio}_{pago_nombre}_{monto}'
                           
                            zip_file.write(pago.comprobante_pago.path, os.path.join(carpeta, f'{nuevo_nombre}.pdf'))
                    output = generar_excel_xmls(datos_xml_lista)
                    zip_file.writestr("GENERAL_XMLs/reporte_facturas.xlsx", output.getvalue())

                zip_buffer.seek(0)
                response = HttpResponse(zip_buffer, content_type='application/zip')
                response.set_cookie('descarga_iniciada', 'true', max_age=20)
                response['Content-Disposition'] = 'attachment; filename=pagos.zip'
                return response
        
        elif 'btnDescargar' in request.POST:
            validar_sat = request.POST.get('validacion') == 'on'
            fecha_inicio = parse_date(request.POST.get('fecha_inicio'))
            fecha_fin = parse_date(request.POST.get('fecha_fin'))
            distrito_id = request.POST.get('distrito')
            tesorero_id = request.POST.get('tesorero')
            folio = request.POST.get('folio')
            print(folio)
            tipo_documento = request.POST.get('tipo_documento')

            facturas_gastos = Factura.objects.none()
            facturas_compras = Facturas.objects.none()
            facturas_viaticos = Viaticos_Factura.objects.none()
            
            if usuario.distritos.nombre == "MATRIZ":
                if tipo_documento in ["", "gastos"]:
                    facturas_gastos = Factura.objects.filter(Q(solicitud_gasto__pagosg__pagado_real__range=[fecha_inicio, fecha_fin])|Q(solicitud_gasto__pagosg__pagado_date__range=[fecha_inicio, fecha_fin]))
                if tipo_documento in ["", "compras"]:    
                    facturas_compras = Facturas.objects.filter(Q(oc__pagos__pagado_real__range=[fecha_inicio, fecha_fin])|Q(oc__pagos__pagado_date__range=[fecha_inicio, fecha_fin]), hecho = True)
                if tipo_documento in ["", "viaticos"]:
                    facturas_viaticos = Viaticos_Factura.objects.filter(Q(solicitud_viatico__pagosv__pagado_real__range=[fecha_inicio, fecha_fin])|Q(solicitud_viatico__pagosv__pagado_date__range=[fecha_inicio, fecha_fin]))

                if distrito_id:
                    facturas_gastos = facturas_gastos.filter(solicitud_gasto__distrito_id=distrito_id)
                    facturas_compras = facturas_compras.filter(oc__req__orden__distrito_id=distrito_id)
                    facturas_viaticos = facturas_viaticos.filter(solicitud_viatico__distrito_id=distrito_id)

                if tesorero_id:
                    facturas_gastos = facturas_gastos.filter(solicitud_gasto__pagosg__tesorero__id=tesorero_id)
                    facturas_compras = facturas_compras.filter(oc__pagos__tesorero__id=tesorero_id)
                    facturas_viaticos = facturas_viaticos.filter(solicitud_viatico__pagosv__tesorero__id=tesorero_id)
                
                if folio:
                    #print('folio',int(folio))
                    #viatico = Solicitud_Viatico.objects.get(folio = folio)
                    facturas_gastos = Factura.objects.filter(solicitud_gasto__folio = folio)
                    facturas_compras = Facturas.objects.filter(oc__folio= folio)
                    facturas_viaticos = Viaticos_Factura.objects.filter(solicitud_viatico__folio = folio)
                 

            else:
                facturas_gastos = Factura.objects.filter(solicitud_gasto__approbado_fecha2__range=[fecha_inicio, fecha_fin], solicitud_gasto__distrito = usuario.distritos)
                facturas_compras = Facturas.objects.filter(oc__autorizado_at_2__range=[fecha_inicio, fecha_fin], oc__req__orden__distrito = usuario.distritos)
                facturas_viaticos = Viaticos_Factura.objects.filter(solicitud_viatico__approved_at2__range=[fecha_inicio, fecha_fin], solicitud_viatico__distrito = usuario.distritos)

            if validar_sat:
                ids_gastos = list(facturas_gastos.values_list('id', flat=True))
                ids_compras = list(facturas_compras.values_list('id', flat=True))
                ids_viaticos = list(facturas_viaticos.values_list('id', flat=True))
                #print(ids_viaticos)

                validar_lote_facturas.delay(ids_gastos, ids_compras, ids_viaticos)
                
            else:
                zip_buffer = BytesIO()
                processed_ocs = set()  # Mant√©n un conjunto de OCs procesadas
                processed_gastos = set()  # Mant√©n un conjunto de gastos procesados
                processed_viaticos = set()  # Mant√©n un conjunto de vi√°ticos procesados
                processed_pagos = set()  # Mant√©n un conjunto de pagos procesados
                datos_xml_lista = []  # Lista para el resumen en Excel

                with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
                    #Se agrgean carpetas generales
                    general_pdfs_folder = "GENERAL_PDFs"
                    general_xmls_folder = "GENERAL_XMLs"
                    
                    #Se procesan facturas de gastos
                    for factura in facturas_gastos:
                        folder_name = f'GASTO_{factura.solicitud_gasto.folio}_{factura.solicitud_gasto.distrito.nombre}'
                        if factura.archivo_pdf:   
                            file_name = os.path.basename(factura.archivo_pdf.path)
                            zip_file.write(factura.archivo_pdf.path, os.path.join(folder_name, file_name))
                            if factura.archivo_xml:
                                # Guardar en la carpeta GENERAL_PDFs con nombre id_uuid.pdf
                                uuid_str = factura.uuid if factura.uuid else 'SIN_UUID'
                                general_file_name = f'{factura.id}_{uuid_str}.pdf'
                            else:
                                general_file_name = file_name
                            zip_file.write(factura.archivo_pdf.path, os.path.join(general_pdfs_folder, general_file_name)) #Est√° l√≠nea guarda en el zip general de pdf
                        
                        beneficiario = factura.solicitud_gasto.colaborador.staff.staff.first_name + ' ' + factura.solicitud_gasto.colaborador.staff.staff.last_name  if factura.solicitud_gasto.colaborador else factura.solicitud_gasto.staff.staff.staff.first_name + ' ' + factura.solicitud_gasto.staff.staff.staff.last_name
                        distrito = factura.solicitud_gasto.distrito.nombre  # Obtener distrito de la factura
                        folio = 'G' + str(factura.solicitud_gasto.folio)
                        fecha_subida = factura.fecha_subida.astimezone(tz=None).replace(tzinfo=None) if factura.fecha_subida else 'No disponible'
                        if factura.archivo_xml:
                            file_name = os.path.basename(factura.archivo_xml.path)
                            zip_file.write(factura.archivo_xml.path, os.path.join(folder_name, file_name))
                            uuid_str = factura.uuid if factura.uuid else 'SIN_UUID'
                            general_file_name = f'{factura.id}_{uuid_str}.xml'

                            zip_file.write(factura.archivo_xml.path, os.path.join(general_xmls_folder, general_file_name)) #Est√° l√≠nea guarda en el zip general de xml's
                            datos_xml_lista.append(extraer_datos_xml_carpetas(factura.archivo_xml.path, folio, fecha_subida, distrito, beneficiario, general_file_name, factura))

                       

                        if factura.solicitud_gasto.id not in processed_gastos:
                            buf = render_pdf_gasto(factura.solicitud_gasto.id)
                            gasto_file_name = f'GASTO_{factura.solicitud_gasto.folio}.pdf'
                            zip_file.writestr(os.path.join(folder_name, gasto_file_name), buf.getvalue())
                            processed_gastos.add(factura.solicitud_gasto.id)

                        pagos = Pago.objects.filter(gasto=factura.solicitud_gasto)
                        for pago in pagos:
                            if pago.comprobante_pago and pago.id not in processed_pagos:
                                texto_pago = extraer_texto_pdf_prop(pago.comprobante_pago)
                                variables_pago = encontrar_variables(texto_pago)

                                fecha_obj = variables_pago.get('fecha')
                                fecha_pago = ''

                                if fecha_obj:
                                    # Si es objeto datetime.date, lo convertimos a string con formato
                                    if isinstance(fecha_obj, datetime.date):
                                        fecha_pago = fecha_obj.strftime('%d-%m-%Y')
                                    else:
                                        # Si ya es cadena (por ejemplo, por error de extracci√≥n)
                                        fecha_pago = str(fecha_obj).replace('/', '-')
                                    pago.pagado_real = fecha_pago
                                    pago.save()
                                titular_cuenta_2 = variables_pago.get('titular_cuenta_2', '')
                                importe_operacion = variables_pago.get('importe_operacion', '').split('.')[0].replace(',', '')

                                # Validamos si todas las variables son v√°lidas:
                                if fecha_pago and fecha_pago != 'No disponible' and titular_cuenta_2 and titular_cuenta_2 != 'No disponible' and importe_operacion and importe_operacion != 'No disponible':
                                    pago_file_name = f'{fecha_pago} {titular_cuenta_2} ${importe_operacion}.pdf'
                                else:
                                    # Si no, conservamos el nombre original
                                    pago_file_name = os.path.basename(pago.comprobante_pago.path)
                                
                                #pago_file_name = os.path.basename(pago.comprobante_pago.path)
                                zip_file.write(pago.comprobante_pago.path, os.path.join(folder_name, f'{pago_file_name}'))
                                processed_pagos.add(pago.id)
                    
                    #Se procesan facturas de compras
                    for factura in facturas_compras:
                        folder_name = f'COMPRA_{factura.oc.folio}_{factura.oc.req.orden.distrito.nombre}'
                        if factura.factura_pdf:
                            #folder_name = f'COMPRA_{factura.oc.folio}_{factura.oc.req.orden.distrito.nombre}'
                            file_name = os.path.basename(factura.factura_pdf.path)
                            zip_file.write(factura.factura_pdf.path, os.path.join(folder_name, file_name))
                            if factura.factura_xml:
                                # Guardar en la carpeta GENERAL_PDFs con nombre id_uuid.pdf
                                uuid_str = factura.uuid if factura.uuid else 'SIN_UUID'
                                general_file_name = f'{factura.id}_{uuid_str}.pdf'
                            else:
                                general_file_name = file_name

                            zip_file.write(factura.factura_pdf.path, os.path.join(general_pdfs_folder, file_name))

                        beneficiario = "NA"
                        distrito = factura.oc.req.orden.distrito.nombre  # Obtener distrito de la factura
                        folio = factura.oc.folio
                        fecha_subida = factura.fecha_subido if factura.fecha_subido else 'No disponible'
                        if factura.factura_xml:
                            file_name = os.path.basename(factura.factura_xml.path)
                            zip_file.write(factura.factura_xml.path, os.path.join(folder_name, file_name))
                            uuid_str = factura.uuid if factura.uuid else 'SIN_UUID'
                            general_file_name = f'{factura.id}_{uuid_str}.xml'
                            
                            zip_file.write(factura.factura_xml.path, os.path.join(general_xmls_folder, general_file_name))
                            datos_xml_lista.append(extraer_datos_xml_carpetas(factura.factura_xml.path, folio, fecha_subida, distrito, beneficiario, general_file_name, factura))
                        
                        # Incluir la ficha de pago
                        pagos = Pago.objects.filter(oc=factura.oc)
                        for pago in pagos:
                            if pago.comprobante_pago and pago.id not in processed_pagos:
                                texto_pago = extraer_texto_pdf_prop(pago.comprobante_pago)
                                variables_pago = encontrar_variables(texto_pago)
                                fecha_obj = variables_pago.get('fecha')
                                fecha_pago = ''

                                if fecha_obj:
                                    # Si es objeto datetime.date, lo convertimos a string con formato
                                    if isinstance(fecha_obj, datetime.date):
                                        fecha_pago = fecha_obj.strftime('%d-%m-%Y')
                                    else:
                                        # Si ya es cadena (por ejemplo, por error de extracci√≥n)
                                        fecha_pago = str(fecha_obj).replace('/', '-')
                                titular_cuenta_2 = variables_pago.get('titular_cuenta_2', '')
                                importe_operacion = variables_pago.get('importe_operacion', '').split('.')[0].replace(',', '')


                                if fecha_pago and fecha_pago != 'No disponible' and titular_cuenta_2 and titular_cuenta_2 != 'No disponible' and importe_operacion and importe_operacion != 'No disponible':
                                    pago_file_name = f'{fecha_pago} {titular_cuenta_2} ${importe_operacion}.pdf'
                                else:
                                    pago_file_name = os.path.basename(pago.comprobante_pago.path)
                                #pago_file_name = os.path.basename(pago.comprobante_pago.path)

                                zip_file.write(pago.comprobante_pago.path, os.path.join(folder_name, f'{pago_file_name}'))
                                processed_pagos.add(pago.id) 
                        
                        # Generar e incluir la OC en el ZIP solo si no ha sido procesada
                        if factura.oc.id not in processed_ocs:
                            buf = generar_pdf(factura.oc)
                            oc_file_name = f'OC_{factura.oc.folio}.pdf'
                            zip_file.writestr(os.path.join(folder_name, oc_file_name), buf.getvalue())
                            processed_ocs.add(factura.oc.id)
                        
                      
                    

                    for factura in facturas_viaticos:
                        folder_name = f'VIATICO_{factura.solicitud_viatico.folio}_{factura.solicitud_viatico.distrito.nombre}'
                        if factura.factura_pdf:

                            file_name = os.path.basename(factura.factura_pdf.path)
                            zip_file.write(factura.factura_pdf.path, os.path.join(folder_name, file_name))
                            if factura.factura_xml:
                                # Guardar en la carpeta GENERAL_PDFs con nombre id_uuid.pdf
                                uuid_str = factura.uuid if factura.uuid else 'SIN_UUID'
                                general_file_name = f'{factura.id}_{uuid_str}.pdf'
                            else:
                                general_file_name = file_name
                            zip_file.write(factura.factura_pdf.path, os.path.join(general_pdfs_folder, general_file_name))

                        beneficiario = factura.solicitud_viatico.colaborador.staff.staff.first_name + ' ' + factura.solicitud_viatico.colaborador.staff.staff.last_name  if factura.solicitud_viatico.colaborador else factura.solicitud_gasto.staff.staff.staff.first_name + ' ' + factura.solicitud_gasto.staff.staff.staff.last_name
                        distrito = factura.solicitud_viatico.distrito.nombre  # Obtener distrito de la factura
                        folio = 'V' + str(factura.solicitud_viatico.folio)
                        fecha_subida = factura.fecha_subido.astimezone(tz=None).replace(tzinfo=None) if factura.fecha_subido else 'No disponible' # Formato YYYY-MM-DD
                        if factura.factura_xml:
                            file_name = os.path.basename(factura.factura_xml.path)
                            zip_file.write(factura.factura_xml.path, os.path.join(folder_name, file_name))
                            uuid_str = factura.uuid if factura.uuid else 'SIN_UUID'
                            general_file_name = f'{factura.id}_{uuid_str}.xml'

                            zip_file.write(factura.factura_xml.path, os.path.join(general_xmls_folder, general_file_name))
                            datos_xml_lista.append(extraer_datos_xml_carpetas(factura.factura_xml.path, folio, fecha_subida, distrito, beneficiario, general_file_name, factura))

                        if factura.solicitud_viatico.id not in processed_viaticos:
                            buf = generar_pdf_viatico(factura.solicitud_viatico.id)
                            viatico_file_name = f'VIATICO_{factura.solicitud_viatico.folio}.pdf'
                            zip_file.writestr(os.path.join(folder_name, viatico_file_name), buf.getvalue())
                            processed_viaticos.add(factura.solicitud_viatico.id)

                    
                        
                        pagos = Pago.objects.filter(viatico=factura.solicitud_viatico)
                        for pago in pagos:
                            if pago.comprobante_pago and pago.id not in processed_pagos:
                                texto_pago = extraer_texto_pdf_prop(pago.comprobante_pago)
                                variables_pago = encontrar_variables(texto_pago)

                                fecha_pago = variables_pago.get('fecha', '').replace('/', '-')
                                titular_cuenta_2 = variables_pago.get('titular_cuenta_2', '')
                                importe_operacion = variables_pago.get('importe_operacion', '').split('.')[0].replace(',', '')

                                if fecha_pago and fecha_pago != 'No disponible' and titular_cuenta_2 and titular_cuenta_2 != 'No disponible' and importe_operacion and importe_operacion != 'No disponible':
                                    pago_file_name = f'{fecha_pago} {titular_cuenta_2} ${importe_operacion}.pdf'
                                else:
                                    pago_file_name = os.path.basename(pago.comprobante_pago.path)
                                
                                #pago_file_name = os.path.basename(pago.comprobante_pago.path)
                                zip_file.write(pago.comprobante_pago.path, os.path.join(folder_name, f'{pago_file_name}'))
                                processed_pagos.add(pago.id)

                    output = generar_excel_xmls(datos_xml_lista)
                    zip_file.writestr("GENERAL_XMLs/reporte_facturas.xlsx", output.getvalue())

                zip_buffer.seek(0)
                response = HttpResponse(zip_buffer, content_type='application/zip')
                response.set_cookie('descarga_iniciada', 'true', max_age=20)
                response['Content-Disposition'] = 'attachment; filename=facturas.zip'
                return response
        elif 'asignar_exhibit' in request.POST:
            print('entro')
            exhibit_id = request.POST.get('exhibit_id')
            exhibit = Exhibit.objects.get(id=exhibit_id)
            pagos = Pago.objects.filter(id__in=request.POST.getlist('compra_ids'))
            for pago in pagos:
                pago.exhibit = exhibit
                pago.save()
            return redirect('control-documentos')
    
    #for pago in pagos_list:
    #    if pago.total_facturas == 0:
    #        pago.estado_facturas = 'sin_facturas'
    #    elif pago.autorizadas == pago.total_facturas:
    #        pago.estado_facturas = 'todas_autorizadas'
    #    else:
    #        pago.estado_facturas = 'pendientes'
    #    if 'enviar_a_control' in request.POST:
    #        ids = request.POST.getlist('compra_ids')
    #        if ids:
    #            pagos = Pago.objects.filter(id__in=ids)
    #            for pago in pagos:
    #                pago.control_documentos = True
    #                pago.fecha_control_documentos = datetime.today()
    #                pago.save()

    #       return redirect('control-documentos')  # Ajusta a donde quieres redirigir
    context= {
        'pagos_list':pagos_list,
        'pagos':pagos,
        'myfilter':myfilter,
        'tesoreros':tesoreros,
        'distritos':distritos,
        'cuentas':cuentas,
        'usuario':usuario,
        'empresas':empresas,
        'exhibits': exhibits,
        }

    return render(request, 'tesoreria/control_documentos.html',context)

@perfil_seleccionado_required
def control_cuentas(request):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    
    if usuario.distritos.nombre == "MATRIZ" and usuario.tipo.supervisor:
        if usuario.tipo.tesoreria:
            cuentas = Cuenta.objects.filter(encargado__tipo__tesoreria = True)
        elif usuario.tipo.finanzas:
            cuentas = Cuenta.objects.filter(encargado__tipo__finanzas = True)
    else:
        cuentas = Cuenta.objects.filter(encargado = usuario)
    
    context= {
        'cuentas': cuentas,
        }

    return render(request, 'tesoreria/control_cuentas.html',context)


def sum_firmada(qs):
    return qs.aggregate(total=Coalesce(Sum(Case(
        When(tipo__isnull=True, then=-F('monto')),
        When(tipo__nombre='CARGO', then=-F('monto')),
        # expl√≠cito por si usas ABONO
        When(tipo__nombre='ABONO', then=F('monto')),
        default=F('monto'),
        output_field=DecimalField(max_digits=18, decimal_places=2),
    )), Value(Decimal('0.00'))))['total'] or Decimal('0.00')


@perfil_seleccionado_required
def control_bancos(request, pk):
    pk_profile = request.session.get('selected_profile_id')
    #usuario = Profile.objects.get(id = pk_profile)
    # Obtener la cuenta seleccionada en el filtro
    
    cuenta = Cuenta.objects.get(id=pk)
    cuenta_saldos = Saldo_Cuenta.objects.filter(cuenta=cuenta).order_by('-fecha_inicial')
    ultimo_saldo = cuenta_saldos.filter(hecho =True).first() if cuenta_saldos.exists() else None

    if ultimo_saldo is not None:
        fecha_saldo = ultimo_saldo.fecha_inicial
        pagos = Pago.objects.filter(
            cuenta = cuenta,
            hecho=True,
            pagado_real__gte= fecha_saldo # Filtrar pagos hechos despu√©s del √∫ltimo saldo
        ).order_by('pagado_real',  'pagado_hora','-tipo__id')  
    else:
        pagos = Pago.objects.filter(cuenta = cuenta, hecho= True).order_by('pagado_real', 'pagado_hora','-tipo__id')  
    
    myfilter = Matriz_Pago_Filter(request.GET, queryset=pagos)
    pagos = myfilter.qs

      # Valores para mostrar solo cuando se calcule
    saldo_final = None
    saldo_trasladado = None
    start_date = None
    end_date = None

    p = Paginator(pagos, 25)
    page = request.GET.get('page')
    pagos_list = p.get_page(page)

    if request.method == 'POST':
        start_date = request.GET.get('start_date')
        if 'btnReporte' in request.POST:
            
            #pagos = pagos.order_by('pagado_real')
            return convert_excel_control_bancos(cuenta.id, pagos, ultimo_saldo, start_date)
        elif 'btnRecalcular' in request.POST:
            start_date_raw = request.POST.get('start_date') or None
            end_date_raw = request.POST.get('end_date') or None

            # Si tu campo pagado_real es DateField, esto es suficiente:
            start_date = start_date_raw or None
            end_date = end_date_raw or None

            # √öltimo saldo (base)
            saldo_obj = (Saldo_Cuenta.objects
                        .filter(cuenta=cuenta, hecho=True)
                        .order_by('-fecha_inicial')
                        .first())

            saldo_base = (saldo_obj.monto_inicial if saldo_obj else Decimal('0.00')) or Decimal('0.00')
            fecha_base = saldo_obj.fecha_inicial if saldo_obj else None

            # 1) TRASLADO: si hay start_date > fecha_base, ajusta saldo base con pagos intermedios
            if fecha_base and start_date and start_date > str(fecha_base):
                intermedios = pagos.filter(
                    pagado_real__gte=fecha_base,
                    pagado_real__lt=start_date
                ).aggregate(
                    cargos=Coalesce(Sum('monto', filter=Q(tipo__isnull=True) | Q(tipo__nombre='CARGO')), Decimal('0.00')),
                    abonos=Coalesce(Sum('monto', filter=Q(tipo__isnull=False) & ~Q(tipo__nombre='CARGO')), Decimal('0.00')),
                )
                saldo_trasladado = saldo_base - intermedios['cargos'] + intermedios['abonos']
                inicio_periodo = start_date
            else:
                # sin start_date o start_date <= saldo base => no hay traslado
                saldo_trasladado = saldo_base
                inicio_periodo = str(fecha_base) if fecha_base else None

            # 2) MOVIMIENTOS DEL PER√çODO
            # - Si no hay end_date, usamos hoy
            fin_periodo = end_date or date.today()

            movimientos_cargos = Decimal('0.00')
            movimientos_abonos = Decimal('0.00')
            print(inicio_periodo, fin_periodo)
            if inicio_periodo:  # s√≥lo si tenemos un inicio
                periodo = pagos.filter(
                    pagado_real__gte=inicio_periodo,
                    pagado_real__lte=fin_periodo
                ).aggregate(
                    cargos=Coalesce(Sum('monto', filter=Q(tipo__isnull=True) | ~Q(tipo__nombre='ABONO')), Decimal('0.00')),
                    abonos=Coalesce(Sum('monto', filter=Q(tipo__isnull=False) & Q(tipo__nombre='ABONO')), Decimal('0.00')),
                )
                movimientos_cargos = periodo['cargos']
                movimientos_abonos = periodo['abonos']

            # 3) SALDO FINAL
            saldo_final = saldo_trasladado - movimientos_cargos + movimientos_abonos

    context= {
        'pagos_list':pagos_list,
        'cuenta': cuenta,
        'pagos':pagos,
        'myfilter':myfilter,
        # Solo aparecen si se calcul√≥
        'saldo_final': saldo_final,
        'saldo_trasladado': saldo_trasladado,
        'start_date': start_date,
        'end_date': end_date,
        }

    return render(request, 'tesoreria/control_bancos.html',context)


def generar_excel_xmls(datos_xml_lista):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen XML"

    columnas = [
        'Distrito', 'Folio', 'Fecha subida', 'Fecha factura', 'Raz√≥n Social', 'Folio Fiscal (UUID)',
        'Monto Total Factura', 'Tipo de Moneda', 'Forma de pago', 'M√©todo de Pago',
        'Receptor (Empresa) Nombre', 'Beneficiario', 'Archivo', 'Tipo de Documento',
        'Fecha Validaci√≥n SAT', 'EstadoSAT'
    ]
    ws.append(columnas)

    for dato in datos_xml_lista:
        ws.append([dato.get(col, '') for col in columnas])

    # Aplicar formato monetario a la columna G (Monto Total Factura)
    for row in range(2, ws.max_row + 1):
        ws[f"G{row}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    wb.save(output)
    output.seek(0)
    return output


def eliminar_caracteres_invalidos(archivo_xml):
    # Definir la expresi√≥n regular para encontrar caracteres inv√°lidos
    regex = re.compile(r'[^\x09\x0A\x0D\x20-\uD7FF\uE000-\uFFFD\u10000-\u10FFFF]')

    # Leer el contenido del archivo XML
    xml_content = archivo_xml.read().decode('utf-8')

    if xml_content.startswith("o;?"):
        print('Detectado "o;?" en el inicio del XML')
        xml_content = xml_content[3:]

    # Eliminar caracteres inv√°lidos seg√∫n la expresi√≥n regular
    xml_content = regex.sub('', xml_content)

    # Volver a posicionar el puntero del archivo al principio
    archivo_xml.seek(0)

    # Guardar el contenido modificado en el archivo original
    archivo_xml.write(xml_content.encode('utf-8'))
    archivo_xml.truncate()  # Asegurarse de que no quede contenido sobrante

    print('Contenido corregido guardado exitosamente.')

    # Retornar el archivo con el contenido modificado
    return archivo_xml

def extraer_datos_del_xml(archivo_xml):
    try:
        # Parsear el archivo XML
        archivo_xml.seek(0)
        tree = ET.parse(archivo_xml)
        root = tree.getroot()
    except (ET.ParseError, FileNotFoundError) as e:
        print(f"Error al parsear el archivo XML: {e}")
        return None, None  # Si ocurre un error, devuelve None
    
    # Identificar la versi√≥n del XML y el espacio de nombres
    version = root.tag
    ns = {}
    if 'http://www.sat.gob.mx/cfd/3' in version:
        ns = {
            'cfdi': 'http://www.sat.gob.mx/cfd/3',
            'tfd': 'http://www.sat.gob.mx/TimbreFiscalDigital',
            'if': 'https://www.interfactura.com/Schemas/Documentos',
        }
    elif 'http://www.sat.gob.mx/cfd/4' in version:
        ns = {
            'cfdi': 'http://www.sat.gob.mx/cfd/4',
            'tfd': 'http://www.sat.gob.mx/TimbreFiscalDigital',
            'if': 'https://www.interfactura.com/Schemas/Documentos',
        }
    else:
        print(f"Versi√≥n del documento XML no reconocida")
        return None, None
    

    rfc_receptor = None
    receptor = root.find('cfdi:Receptor', ns)
    if receptor is not None:
        rfc_receptor = receptor.get('Rfc')
    else:
        print("Receptor no encontrado")
    # Buscar el complemento donde se encuentra el UUID y la fecha de timbrado
    complemento = root.find('cfdi:Complemento', ns)
    if complemento is not None:
        timbre_fiscal = complemento.find('tfd:TimbreFiscalDigital', ns)
        if timbre_fiscal is not None:
            uuid = timbre_fiscal.get('UUID')
            fecha_timbrado = timbre_fiscal.get('FechaTimbrado') or root.get('Fecha')
        else:
            print("Timbre Fiscal Digital no encontrado")
            return None, None, None
    else:
        print("Complemento no encontrado")
        return None, None, None
    
    return uuid, fecha_timbrado, rfc_receptor
    
import xml.etree.ElementTree as ET

def extraer_datos_del_complemento(ruta_xml):
    try:
        # Parsear el archivo XML
        tree = ET.parse(ruta_xml)
        root = tree.getroot()
    except (ET.ParseError, FileNotFoundError) as e:
        print(f"Error al parsear el archivo XML: {e}")
        return None, None  # Si ocurre un error, devolver None
    
    # Definir espacios de nombres
    ns = {
        'cfdi': 'http://www.sat.gob.mx/cfd/4',
        'tfd': 'http://www.sat.gob.mx/TimbreFiscalDigital',
        'pago20': 'http://www.sat.gob.mx/Pagos20'
    }
    
    # Buscar el UUID en TimbreFiscalDigital
    uuid = None
    complemento = root.find('cfdi:Complemento', ns)
    if complemento is not None:
        timbre_fiscal = complemento.find('tfd:TimbreFiscalDigital', ns)
        if timbre_fiscal is not None:
            uuid = timbre_fiscal.get('UUID', '')

    # Buscar el IdDocumento dentro de DoctoRelacionado
    # Obtener todos los IdDocumento
    ids_documentos = []
    pagos = complemento.find('pago20:Pagos', ns) if complemento is not None else None
    if pagos is not None:
        for pago in pagos.findall('pago20:Pago', ns):
            doctos = pago.findall('pago20:DoctoRelacionado', ns)
            for docto in doctos:
                id_doc = docto.get('IdDocumento')
                if id_doc:
                    ids_documentos.append(id_doc)

    return uuid, ids_documentos


def extraer_datos_xml_carpetas(xml_file, folio, fecha_subida, distrito, beneficiario, nombre_general, factura):
    """Extrae los datos clave de un archivo XML CFDI, compatible con diferentes versiones, incluyendo complementos de pago."""
    try:
        tree = ET.parse(xml_file)
        root = tree.getroot()
    except ET.ParseError as e:
        print(f"Error al parsear {xml_file}: {e}")
        return {
            'Folio': folio,
            'Archivo': nombre_general,
            'Error': f"Archivo XML inv√°lido: {e}"
        }

    # Detectar la versi√≥n del CFDI
    version = root.get("Version", "3.3")

    # Definir los espacios de nombres seg√∫n la versi√≥n
    ns = {
        'cfdi': f'http://www.sat.gob.mx/cfd/{version[0]}',
        'tfd': 'http://www.sat.gob.mx/TimbreFiscalDigital',
        'pago20': 'http://www.sat.gob.mx/Pagos20'
    }

    emisor = root.find("cfdi:Emisor", ns)
    receptor = root.find("cfdi:Receptor", ns)
    complemento = root.find("cfdi:Complemento/tfd:TimbreFiscalDigital", ns)
    pagos = root.find("cfdi:Complemento/pago20:Pagos", ns)

    # Verificar si es un complemento de pago
    es_complemento_pago = pagos is not None

    # Obtener datos generales
    fecha_emision = root.get('Fecha', '')
    fecha_emision_excel = datetime.strptime(fecha_emision, "%Y-%m-%dT%H:%M:%S") if fecha_emision else None

    # Datos espec√≠ficos para complemento de pago
    if es_complemento_pago:
        pago = pagos.find("pago20:Pago", ns)
        if pago is not None:
            moneda = pago.get("MonedaP", "")
            monto_total = float(pago.get("Monto", "0"))
            forma_pago = pago.get("FormaDePagoP", "")
            metodo_pago = pago.get("MetodoPago", "N/A")  
        else:
            moneda = ""
            monto_total = 0
            forma_pago = ""
            metodo_pago = ""

        tipo_documento = "Complemento de Pago"
    else:
        moneda = root.get('Moneda', '')
        monto_total = float(root.get('Total', '0'))
        forma_pago = root.get('FormaPago', '')
        metodo_pago = root.get('MetodoPago', '')
        tipo_documento = "Factura"

    datos = {
        'Fecha subida': fecha_subida,
        'Beneficiario': beneficiario,
        'Folio': folio,
        'Distrito': distrito,  # Se agrega el distrito
        'Tipo de Documento': tipo_documento,
        'Fecha factura': fecha_emision_excel,
        'Raz√≥n Social': emisor.get('Nombre') if emisor is not None else '',
        'Folio Fiscal (UUID)': complemento.get('UUID') if complemento is not None else '',
        'Monto Total Factura': monto_total,
        'Tipo de Moneda': moneda,
        'M√©todo de Pago': metodo_pago,
        'Forma de pago': forma_pago,
        'Receptor (Empresa) Nombre': receptor.get('Nombre') if receptor is not None else '',
        'Archivo': nombre_general,
        'EstadoSAT': factura.estado_sat or '',
        'Fecha Validaci√≥n SAT': timezone.localtime(factura.fecha_validacion_sat).strftime("%Y-%m-%d %H:%M:%S") if factura.fecha_validacion_sat else '',
    }
    return datos



def generar_archivo_zip(facturas, compra):
    nombre = compra.folio if compra.folio else ''
    zip_filename = f'facturas_compragasto-{nombre}.zip'
    
    # Crear un archivo zip en memoria
    in_memory_zip = io.BytesIO()

    with zipfile.ZipFile(in_memory_zip, 'w') as zip_file:
        for factura in facturas:
            if factura.factura_pdf:
                pdf_path = factura.factura_pdf.path
                zip_file.write(pdf_path, os.path.basename(pdf_path))
            if factura.factura_xml:
                # Generar el PDFreader
                response = generar_cfdi(None, factura.id)
                pdf_filename = f"{factura.id}.pdf" if factura.id else f"factura_{factura.id}.pdf"
                # A√±adir el contenido del PDF al ZIP
                zip_file.writestr(pdf_filename, response.content)
                #A√±adir el xml
                xml_path = factura.factura_xml.path
                zip_file.write(xml_path, os.path.basename(xml_path))

    # Resetear el puntero del archivo en memoria
    in_memory_zip.seek(0)

    return in_memory_zip, zip_filename

@perfil_seleccionado_required
def matriz_facturas_nomodal(request, pk):
    #print('estoy en matriz_facturas_nomodal')
    pk_perfil = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk_perfil)

    try:
        if perfil.tipo.nombre == "PROVEEDOR_EXTERNO":
            base_url = reverse('matriz-oc-proveedores')
            compra = get_object_or_404(Compra, id=pk, proveedor__nombre = perfil.proveedor)
        else:
            compra = get_object_or_404(Compra, id=pk)
            next_url = request.GET.get('next', 'matriz-compras')
            try:
                base_url = reverse(next_url)
            except NoReverseMatch:
                base_url = next_url

    except Http404:
        messages.error(request, "No tienes acceso a esta orden de compra.")
        #return redirect(next_url)
    facturas = Facturas.objects.filter(oc = compra, hecho=True)
    pagos = Pago.objects.filter(oc = compra)
    form = Facturas_Completas_Form(instance=compra)
    # Construir los par√°metros de filtro
    print('base_url',base_url)
    filtros = {
        'proveedor': request.GET.get('proveedor', ''),
        'distrito': request.GET.get('distrito', ''),
        'start_date': request.GET.get('start_date', ''),
        'end_date': request.GET.get('end_date', ''),
    }
    # Codificar los par√°metros
    query_string = urlencode(filtros)
    #print('query_string:',filtros)
    
    for pago in pagos:
        fecha_pdf = None
        if pago.comprobante_pago:
            try:
                texto = extraer_texto_pdf_prop(pago.comprobante_pago)
                variables = encontrar_variables(texto)
                fecha_pdf = variables.get('fecha', 'No disponible')
            except Exception as e:
                fecha_pdf = f"Error: {str(e)}"

        pago.fecha_pdf = fecha_pdf  # Asignar el valor a la variable de instancia
        #print(pago.fecha_pdf)  # Imprimir el valor de fecha_pdf
   
    if request.method == 'POST':
        #connector = '&' if '?' in base_url else '?'
        #print(query_string)
        redirect_url = f"{base_url}?{query_string}" if query_string else base_url
        #print('imprimiendo',redirect_url)
        form = Facturas_Completas_Form(request.POST, instance=compra)
        if "btn_factura_completa" in request.POST:
            fecha_hora = datetime.today()
            for factura in facturas:
                checkbox_name = f'autorizar_factura_{factura.id}'
                #print("Nombre del checkbox esperado:", checkbox_name)  # Imprimir el nombre esperado
                if checkbox_name in request.POST:
                    factura.autorizada = True
                    factura.autorizada_por = perfil
                    factura.autorizada_el = fecha_hora
                else:
                    factura.autorizada = False
                factura.save()
            if form.is_valid():
                form.save()
                messages.success(request,'Has cambiado el status de facturas completas')
                return redirect(redirect_url)
            else:
                messages.error(request,'No est√° validando')
        elif "btn_descargar_todo" in request.POST:
            in_memory_zip, zip_filename = generar_archivo_zip(facturas, compra)
            response = HttpResponse(in_memory_zip, content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
            return response
        elif 'salir' in request.POST:
            return redirect(redirect_url)

    context={
        'pagos':pagos,
        'form':form,
        'facturas':facturas,
        'compra':compra,
        }

    return render(request, 'tesoreria/matriz_factura_no_modal.html', context)

@perfil_seleccionado_required
def matriz_complementos(request, pk):
    print('estoy en matriz_complementos')
    pk_perfil = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk_perfil)

    try:
        if perfil.tipo.nombre == "PROVEEDOR_EXTERNO":
            next_url = 'matriz-oc-proveedores'
            print(next_url)
            factura = get_object_or_404(Facturas, id=pk, oc__proveedor__nombre = perfil.proveedor)
        else:
            factura = get_object_or_404(Facturas, id=pk)
            next_url = request.GET.get('next','matriz-compras')
    except Http404:
        messages.error(request, "No tienes acceso a esta orden de compra.")
        return redirect(next_url)
    complementos = Complemento_Pago.objects.filter(facturas=factura, hecho=True)
    #pagos = Factura.objects.filter(oc = compra)
    #form = Facturas_Completas_Form(instance=compra)
    

    if request.method == 'POST':
        #form = Facturas_Completas_Form(request.POST, instance=compra)
        #if "btn_factura_completa" in request.POST:
        #    fecha_hora = datetime.today()
        #    for factura in facturas:
        #        checkbox_name = f'autorizar_factura_{factura.id}'
                #print("Nombre del checkbox esperado:", checkbox_name)  # Imprimir el nombre esperado
        #        if checkbox_name in request.POST:
        #            factura.autorizada = True
        #            factura.autorizada_por = perfil
        #            factura.autorizada_el = fecha_hora
        #        else:
        #            factura.autorizada = False
        #        factura.save()
        #    if form.is_valid():
        #        form.save()
        #        messages.success(request,'Haz cambiado el status de facturas completas')
        #        return redirect(next_url)
        #    else:
        #        messages.error(request,'No est√° validando')
        #elif "btn_descargar_todo" in request.POST:
        #    in_memory_zip, zip_filename = generar_archivo_zip(facturas, compra)
        #    response = HttpResponse(in_memory_zip, content_type='application/zip')
        #    response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
        #    return response
        print(request.POST)
        if 'salir' in request.POST:
            return redirect(next_url)

    context={
        #'pagos':pagos,
        #'form':form,
        'factura':factura,
        'complementos':complementos,
        }

    return render(request, 'tesoreria/matriz_complementos.html', context)



def guardar_factura(factura, archivo_procesado, nombre_archivo,  uuid_extraido, fecha_timbrado_extraida, usuario, comentario):
    factura.factura_xml.save(nombre_archivo, archivo_procesado)
    factura.uuid = uuid_extraido
    factura.fecha_timbrado = fecha_timbrado_extraida
    factura.hecho = True
    factura.fecha_subido = date.today()
    factura.hora_subido = datetime.now().time()
    factura.subido_por = usuario
    factura.comentario = comentario
    factura.save()

@perfil_seleccionado_required
def factura_nueva(request, pk):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    compra = Compra.objects.get(id = pk)
    #facturas = Facturas.objects.filter(pago = pago, hecho=True)

    form = UploadFileForm()

    if request.method == 'POST':
        if 'btn_registrar' in request.POST:
            form = UploadFileForm(request.POST, request.FILES or None)
            if form.is_valid():
                archivos_pdf = request.FILES.getlist('archivo_pdf')
                archivos_xml = request.FILES.getlist('archivo_xml')
                if not archivos_pdf and not archivos_xml:
                    messages.error(request, 'Debes subir al menos un archivo PDF o XML.')
                    return HttpResponse(status=204)
                
                # Iterar sobre el n√∫mero m√°ximo de archivos en cualquiera de las listas
                max_len = max(len(archivos_pdf), len(archivos_xml))
                facturas_registradas = []
                facturas_duplicadas = []
                facturas_mes_invalido = []  # Lista para facturas fuera del mes
                comentario = request.POST.get('comentario', '')  # Extraer el comentario
                
                fecha_actual = datetime.today()
                mes_actual = fecha_actual.month
                a√±o_actual = fecha_actual.year
                print(fecha_actual)

                for i in range(max_len):
                    archivo_pdf = archivos_pdf[i] if i < len(archivos_pdf) else None
                    archivo_xml = archivos_xml[i] if i < len(archivos_xml) else None
                    factura, created = Facturas.objects.get_or_create(oc=compra, hecho=False)
                    if archivo_xml:
                        archivo_procesado = eliminar_caracteres_invalidos(archivo_xml)

                        # Guardar temporalmente para extraer datos
                        #factura_temp = Factura(archivo_xml=archivo_xml)
                        #factura_temp.archivo_xml.save(archivo_xml.name, archivo_procesado, save=False)
                        RFC_RECEPTOR_ESPERADO = "GVO020226811"
                        uuid_extraido, fecha_timbrado_extraida, rfc_receptor = extraer_datos_del_xml(archivo_procesado)
                        if rfc_receptor and rfc_receptor != RFC_RECEPTOR_ESPERADO:
                            messages.error(request, f"RFC receptor inv√°lido ({rfc_receptor}). Se esperaba {RFC_RECEPTOR_ESPERADO}.")
                            break # Saltar al siguiente archivo si el RFC no coincide
                        if fecha_timbrado_extraida:
                            try:
                                # Si la fecha incluye la hora, parsearla correctamente
                                fecha_timbrado_dt = datetime.strptime(fecha_timbrado_extraida, "%Y-%m-%dT%H:%M:%S")
                            except ValueError:
                                print(f"‚ö†Ô∏è Error: Formato de fecha desconocido -> {fecha_timbrado_extraida}")
                                fecha_timbrado_dt = None
                        else:
                            print("‚ö†Ô∏è Advertencia: No se extrajo ninguna fecha de timbrado")
                            fecha_timbrado_dt = None
                        mes_factura = fecha_timbrado_dt.month  # Obtener el mes de la factura
                        a√±o_factura = fecha_timbrado_dt.year  # Obtener el a√±o de la factura

                        # Validar que el mes y el a√±o de la factura sea el mismo que el actual
                        # Se quita por el momento la restricci√≥n del mes
                        #if mes_factura != mes_actual or a√±o_factura != a√±o_actual:
                        #    facturas_mes_invalido.append(uuid_extraido)
                        #    continue  # Saltar la factura si no cumple la condici√≥n
                        # Verificar si ya existe una factura con el mismo UUID y fecha de timbrado en cualquiera de las tablas
                        factura_existente = Factura.objects.filter(uuid=uuid_extraido, fecha_timbrado=fecha_timbrado_extraida).first()
                        facturas_existentes = Facturas.objects.filter(uuid=uuid_extraido, fecha_timbrado=fecha_timbrado_extraida).first()
                        viaticos_factura_existente = Viaticos_Factura.objects.filter(uuid=uuid_extraido, fecha_timbrado=fecha_timbrado_extraida).first()

                        if factura_existente or facturas_existentes or viaticos_factura_existente:
                            # Si una factura existente se encuentra, verificamos si su solicitud no est√° aprobada
                            if factura_existente and (factura_existente.solicitud_gasto.autorizar is False or factura_existente.solicitud_gasto.autorizar2 is False):
                                factura_existente.delete()
                                guardar_factura(factura, archivo_procesado, archivo_xml.name, uuid_extraido, fecha_timbrado_extraida, usuario, comentario)

                            elif facturas_existentes and (facturas_existentes.oc.autorizado1 is False or facturas_existentes.oc.autorizado2 is False):
                                facturas_existentes.delete()
                                guardar_factura(factura, archivo_procesado, archivo_xml.name, uuid_extraido, fecha_timbrado_extraida, usuario, comentario)

                            elif viaticos_factura_existente and (viaticos_factura_existente.solicitud_viatico.autorizar is False or viaticos_factura_existente.solicitud_viatico.autorizar2 is False):
                                viaticos_factura_existente.delete()
                                guardar_factura(factura, archivo_procesado, archivo_xml.name, uuid_extraido, fecha_timbrado_extraida, usuario, comentario)

                            else:
                                # Si no cumple las condiciones de eliminaci√≥n, consideramos la factura duplicada
                                facturas_duplicadas.append(uuid_extraido)
                                continue  # Saltar al siguiente archivo si se encuentra duplicado
                        else:
                            # Si no existe ninguna factura, guardar la nueva
                            guardar_factura(factura, archivo_procesado, archivo_xml.name, uuid_extraido, fecha_timbrado_extraida, usuario, comentario)
                            #messages.success(request, 'Las facturas se registraron de manera exitosa')
                    if archivo_pdf:
                        factura.factura_pdf = archivo_pdf
                        factura.hecho = True
                        factura.fecha_subido = date.today()
                        factura.hora_subido = datetime.now().time()
                        factura.subido_por = usuario
                        factura.comentario = comentario
                        factura.save()
                      
                        facturas_registradas.append(uuid_extraido if archivo_xml else f"Factura PDF {archivo_pdf.name}")
                    #messages.success(request, 'Los facturas se registraron de manera exitosa')
                     # Mensajes de √©xito o duplicados
                #return HttpResponse(status=204)
                # Mensajes de √©xito o advertencias
                if facturas_duplicadas:
                    messages.warning(request, f'Las siguientes no se pudieron subir porque ya estaban registradas: {", ".join(facturas_duplicadas)}')
                if facturas_mes_invalido:
                    messages.error(request, f'Las siguientes facturas no se pudieron registrar porque no corresponden al mes y a√±o actual: {", ".join(facturas_mes_invalido)}')
                #return HttpResponse(status=204)

            else:
                messages.error(request,'No se pudo subir tu documento')

    context={
        'form':form,
        'compra':compra,
        }

    return render(request, 'tesoreria/registrar_nueva_factura.html', context)


@perfil_seleccionado_required
def complemento_nuevo(request, pk):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    #factura = Facturas.objects.get(id = pk)
    #facturas = Facturas.objects.filter(pago = pago, hecho=True)

    form = UploadComplementoForm()

    if request.method == 'POST' and 'btn_registrar' in request.POST:
        form = UploadComplementoForm(request.POST, request.FILES or None)
        if form.is_valid():
            archivos_pdf = request.FILES.getlist('complemento_pdf')
            archivos_xml = request.FILES.getlist('complemento_xml')
           
            if not archivos_pdf and not archivos_xml:
                messages.error(request, 'Debes subir al menos un archivo PDF o XML.')
                return HttpResponse(status=204)

            # Listas de seguimiento
            complementos_invalidos = []
            complementos_duplicados = []
            complementos_registrados = []
            pdf_sin_complemento = []
            comentario = request.POST.get('comentario', '')
            #print(comentario)
            
            # Determinar el n√∫mero m√°ximo de archivos a procesar
            max_len = max(len(archivos_pdf), len(archivos_xml))

            for i in range(max_len):
                archivo_pdf = archivos_pdf[i] if i < len(archivos_pdf) else None
                archivo_xml = archivos_xml[i] if i < len(archivos_xml) else None
                complemento_final = None  # Variable para almacenar el complemento en el que se trabajar√°

                # Procesar XML si est√° presente
                if archivo_xml:
                    try:
                        archivo_procesado = eliminar_caracteres_invalidos(archivo_xml)

                        # Guardar el archivo XML en un archivo temporal para procesarlo
                        with tempfile.NamedTemporaryFile(delete=False, suffix='.xml') as tmp:
                            for chunk in archivo_procesado.chunks():
                                tmp.write(chunk)
                            tmp_path = tmp.name

                        try:
                            # Extraer datos desde el archivo temporal
                            uuid_complemento, uuids_facturas = extraer_datos_del_complemento(tmp_path)
                        finally:
                            # Asegurar que el archivo temporal se borre aunque falle
                            os.remove(tmp_path)


                        # Validaciones de UUID y relaci√≥n con factura
                        if not uuid_complemento:
                            complementos_invalidos.append(archivo_xml.name)
                            continue

                        complemento_existente = Complemento_Pago.objects.filter(uuid=uuid_complemento).first()
                        print(complemento_existente)
                        if complemento_existente:
                            complementos_duplicados.append(uuid_complemento)
                            complemento_final = complemento_existente  # Reusar complemento existente
                        else:
                            # Crear nuevo complemento sin facturas a√∫n
                            complemento_final = Complemento_Pago.objects.create(
                                complemento_xml=archivo_xml,
                                uuid=uuid_complemento,
                                subido_por=usuario,
                                fecha_subido=date.today(),
                                hora_subido=datetime.now().time(),
                                comentario=comentario,
                                hecho=True
                            )
                            # Llamar la property que extrae los UUIDs de facturas
                            #info_xml = complemento_final.emisor
                            #if info_xml and 'doctos_relacionados_uuids' in info_xml:
                            #    uuids_facturas = info_xml['doctos_relacionados_uuids']
                            facturas_relacionadas = Facturas.objects.filter(uuid__in=uuids_facturas)

                            if facturas_relacionadas.exists():
                                complemento_final.facturas.set(facturas_relacionadas)
                                complementos_registrados.append(uuid_complemento)
                            else:
                                complemento_final.delete()  # limpia si no hay facturas v√°lidas
                                complementos_invalidos.append(f"No se encontraron facturas relacionadas con UUIDs: {', '.join(uuids_facturas)}")
                                continue
                            #else:
                            #    complemento_final.delete()
                             #   complementos_invalidos.append(f"{archivo_xml.name} no contiene facturas relacionadas.")
                             #   continue


                          

                    except Exception as e:
                        messages.error(request, f"Error al procesar {archivo_xml.name}: {e}")
                        continue

                # Procesar PDF y asociarlo con el mismo complemento
                if archivo_pdf:
                    if complemento_final:
                        complemento_final.complemento_pdf = archivo_pdf  # ‚úÖ ADHERIR PDF AL COMPLEMENTO EXISTENTE
                        complemento_final.save()
                        complementos_registrados.append(f"PDF: {archivo_pdf.name}")
                    else:
                        pdf_sin_complemento.append(archivo_pdf.name)  # üìå Registrar PDFs sin complemento

            # Generar mensajes para el usuario
            if complementos_registrados:
                messages.success(request, f'Se han registrado los siguientes complementos: {", ".join(complementos_registrados)}')
            if complementos_duplicados:
                messages.warning(request, f'Los siguientes complementos ya estaban registrados y no se duplicaron: {", ".join(complementos_duplicados)}')
            if complementos_invalidos:
                messages.error(request, f'Los siguientes archivos no tienen factura relacionada o est√°n mal estructurados: {", ".join(complementos_invalidos)}')
            if pdf_sin_complemento:
                messages.error(request, f'Los siguientes archivos PDF no tienen un complemento de pago asociado y no se guardaron: {", ".join(pdf_sin_complemento)}')

        else:
            messages.error(request, 'No se pudo subir tu documento.')

    context={
        'form':form,
        }

    return render(request, 'tesoreria/registrar_nuevo_complemento.html', context)

@perfil_seleccionado_required
def factura_compra_edicion(request, pk):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    factura = Facturas.objects.get(id = pk)
    #facturas = Facturas.objects.filter(pago = pago, hecho=True)
    #factura, created = Facturas.objects.get_or_create(pago=pago, hecho=False)
    form = Facturas_Form(instance= factura)

    if request.method == 'POST':
        if 'btn_edicion' in request.POST:
            form = Facturas_Form(request.POST or None, request.FILES or None, instance = factura)
            if form.is_valid():
                factura = form.save(commit = False)
                factura.subido_por = usuario
                factura.save()
                form.save()
                messages.success(request,'Las facturas se subieron de manera exitosa')
            else:
                messages.error(request,'No se pudo subir tu documento')


    context={
        'factura':factura,
        'form':form,
        }

    return render(request, 'tesoreria/factura_compra_edicion.html', context)

@perfil_seleccionado_required
def factura_eliminar(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk_perfil)
    factura = Facturas.objects.get(id = pk)
    compra = factura.oc
    comentario = request.POST.get('comentario')
    # Obtener el par√°metro `next` de la URL
    next_url = request.GET.get('next', None)

    # Construir la URL de la matriz de facturas de vi√°ticos
    matriz_url = reverse('matriz-facturas-nomodal', args=[compra.id])

    static_path = settings.STATIC_ROOT
    img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
    img_path2 = os.path.join(static_path,'images','logo_vordcab.jpg')
    image_base64 = get_image_base64(img_path)
    logo_v_base64 = get_image_base64(img_path2)
    # Crear el mensaje HTML
    html_message = f"""
    <html>
        <head>
            <meta charset="UTF-8">
        </head>
        <body style="font-family: Arial, sans-serif; color: #333; background-color: #f4f4f4; margin: 0; padding: 0;">
            <table width="100%" cellspacing="0" cellpadding="0" style="background-color: #f4f4f4; padding: 20px;">
                <tr>
                    <td align="center">
                        <table width="600px" cellspacing="0" cellpadding="0" style="background-color: #ffffff; padding: 20px; border-radius: 10px;">
                            <tr>
                                <td align="center">
                                    <img src="data:image/jpeg;base64,{logo_v_base64}" alt="Logo" style="width: 100px; height: auto;" />
                                </td>
                            </tr>
                            <tr>
                                <td style="padding: 20px;">
                                    <p style="font-size: 18px; text-align: justify;">
                                        <p>Estimado {factura.subido_por.staff.staff.first_name} {factura.subido_por.staff.staff.last_name},</p>
                                    </p>
                                    <p style="font-size: 16px; text-align: justify;">
                                        Est√°s recibiendo este correo porque tu factura subida el: <strong>{factura.fecha_subido}</strong> en la compra <strong>{factura.oc.folio}</strong> ha sido eliminada.</p>
                                    <p>Comentario:</p>
                                    {comentario}
                                    </p>
                                <p style="font-size: 16px; text-align: justify;">
                                    Att: {perfil.staff.staff.first_name} {perfil.staff.staff.last_name}
                                </p>
                                    <p style="text-align: center; margin: 20px 0;">
                                        <img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width: 50px; height: auto; border-radius: 50%;" />
                                    </p>
                                    <p style="font-size: 14px; color: #999; text-align: justify;">
                                        Este mensaje ha sido autom√°ticamente generado por SAVIA 2.0
                                    </p>
                                </td>
                            </tr>
                        </table>
                    </td>
                </tr>
            </table>
        </body>
    </html>
    """
    try:
        email = EmailMessage(
            f'Factura eliminada',
            body=html_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[compra.creada_por.staff.staff.email],
            headers={'Content-Type': 'text/html'}
            )
        email.content_subtype = "html " # Importante para que se interprete como HTML
        if factura.factura_pdf:
            pdf_path = factura.factura_pdf.path
            if os.path.exists(pdf_path):  # Verificar si el archivo realmente existe
                with open(pdf_path, 'rb') as pdf_file:
                    email.attach(factura.factura_pdf.name, pdf_file.read(), 'application/pdf')
            else:
                print(f"El archivo PDF no se encuentra en la ruta: {pdf_path}")

        if factura.factura_xml:
            xml_path = factura.factura_xml.path
            if os.path.exists(xml_path):  # Verificar si el archivo realmente existe
                with open(xml_path, 'rb') as xml_file:
                    email.attach(factura.factura_xml.name, xml_file.read(), 'application/xml')
            else:
                print(f"El archivo XML no se encuentra en la ruta: {xml_path}")

        email.send()
        messages.success(request, f'La factura {factura.id} ha sido eliminada exitosamente')
    except (BadHeaderError, SMTPException, socket.gaierror) as e:
        error_message = f'La factura {factura.id} ha sido eliminada, pero el correo no ha sido enviado debido a un error: {e}'
        messages.success(request, error_message)
    factura.delete()

    # Redirigir a 'matriz-facturas-viaticos' con el par√°metro `next` si existe
    if next_url:
        return redirect(f'{matriz_url}?next={next_url}')
    else:
        return redirect(matriz_url)
    

@perfil_seleccionado_required
def complemento_eliminar(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk_perfil)
    complemento = Complemento_Pago.objects.get(id = pk)
    factura = complemento.facturas.first()
    comentario = request.POST.get('comentario')
    # Obtener el par√°metro `next` de la URL
    next_url = request.GET.get('next', None)

    # Construir la URL de la matriz de facturas de vi√°ticos
    matriz_url = reverse('matriz-complementos', args=[factura.id])

    static_path = settings.STATIC_ROOT
    img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
    img_path2 = os.path.join(static_path,'images','logo_vordcab.jpg')
    image_base64 = get_image_base64(img_path)
    logo_v_base64 = get_image_base64(img_path2)
    # Crear el mensaje HTML
    html_message = f"""
    <html>
        <head>
            <meta charset="UTF-8">
        </head>
        <body style="font-family: Arial, sans-serif; color: #333; background-color: #f4f4f4; margin: 0; padding: 0;">
            <table width="100%" cellspacing="0" cellpadding="0" style="background-color: #f4f4f4; padding: 20px;">
                <tr>
                    <td align="center">
                        <table width="600px" cellspacing="0" cellpadding="0" style="background-color: #ffffff; padding: 20px; border-radius: 10px;">
                            <tr>
                                <td align="center">
                                    <img src="data:image/jpeg;base64,{logo_v_base64}" alt="Logo" style="width: 100px; height: auto;" />
                                </td>
                            </tr>
                            <tr>
                                <td style="padding: 20px;">
                                    <p style="font-size: 18px; text-align: justify;">
                                        <p>Estimado {complemento.subido_por.staff.staff.first_name} {complemento.subido_por.staff.staff.last_name},</p>
                                    </p>
                                    <p style="font-size: 16px; text-align: justify;">
                                        Est√°s recibiendo este correo porque tu complemento subido el: <strong>{complemento.fecha_subido}</strong> en la compra <strong>{factura.oc.folio}</strong> ha sido eliminado.</p>
                                    <p>Comentario:</p>
                                    {comentario}
                                    </p>
                                <p style="font-size: 16px; text-align: justify;">
                                    Att: {perfil.staff.staff.first_name} {perfil.staff.staff.last_name}
                                </p>
                                    <p style="text-align: center; margin: 20px 0;">
                                        <img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width: 50px; height: auto; border-radius: 50%;" />
                                    </p>
                                    <p style="font-size: 14px; color: #999; text-align: justify;">
                                        Este mensaje ha sido autom√°ticamente generado por SAVIA 2.0
                                    </p>
                                </td>
                            </tr>
                        </table>
                    </td>
                </tr>
            </table>
        </body>
    </html>
    """
    try:
        email = EmailMessage(
            f'Complemento eliminado',
            body=html_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[factura.oc.creada_por.staff.staff.email],
            headers={'Content-Type': 'text/html'}
            )
        email.content_subtype = "html " # Importante para que se interprete como HTML
        if complemento.complemento_pdf:
            pdf_path = complemento.complemento_pdf.path
            if os.path.exists(pdf_path):  # Verificar si el archivo realmente existe
                with open(pdf_path, 'rb') as pdf_file:
                    email.attach(complemento.complemento_pdf.name, pdf_file.read(), 'application/pdf')
            else:
                print(f"El archivo PDF no se encuentra en la ruta: {pdf_path}")

        if complemento.complemento_xml:
            xml_path = complemento.complemento_xml.path
            if os.path.exists(xml_path):  # Verificar si el archivo realmente existe
                with open(xml_path, 'rb') as xml_file:
                    email.attach(complemento.complemento_xml.name, xml_file.read(), 'application/xml')
            else:
                print(f"El archivo XML no se encuentra en la ruta: {xml_path}")

        email.send()
        messages.success(request, f'El complemento de pago {complemento.id} ha sido eliminado exitosamente')
    except (BadHeaderError, SMTPException, socket.gaierror) as e:
        error_message = f'El complemento {complemento.id} ha sido eliminado, pero el correo no ha sido enviado debido a un error: {e}'
        messages.success(request, error_message)
    complemento.delete()

    # Redirigir a 'matriz-facturas-viaticos' con el par√°metro `next` si existe
    if next_url:
        return redirect(f'{matriz_url}?next={next_url}')
    else:
        return redirect(matriz_url)

@perfil_seleccionado_required
def mis_gastos(request):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    
    if usuario.sustituto:
        usuario = Profile.objects.filter(staff = usuario.staff, tipo__subdirector = True, st_activo = True).first()
    gastos = Solicitud_Gasto.objects.filter(
        Q(staff = usuario) |Q(colaborador = usuario), 
        complete=True
        ).order_by('-folio')
    myfilter = Solicitud_Gasto_Filter(request.GET, queryset=gastos)
    gastos = myfilter.qs



    for gasto in gastos:
        articulos_gasto = Articulo_Gasto.objects.filter(gasto=gasto)

        proyectos = set()
        subproyectos = set()
        gasto.creado_reciente = (gasto.approbado_fecha2 >= timezone.now() - timedelta(days=30)) if gasto.approbado_fecha2 else True
        for articulo in articulos_gasto:
            if articulo.proyecto:
                proyectos.add(str(articulo.proyecto.nombre))
            if articulo.subproyecto:
                subproyectos.add(str(articulo.subproyecto.nombre))

        gasto.proyectos = ', '.join(proyectos)
        gasto.subproyectos = ', '.join(subproyectos)

    p = Paginator(gastos, 20)
    page = request.GET.get('page')
    gastos_list = p.get_page(page)

    if request.method =='POST' and 'btnExcel' in request.POST:
        return convert_excel_gasto(gastos)
    
    usuario_view = True
    print('usuario_view:',usuario_view)
    context= {
        'gastos':gastos,
        'myfilter':myfilter,
        'gastos_list': gastos_list,
        'usuario_view': usuario_view,
        }

    return render(request, 'tesoreria/mis_gastos.html',context)

@perfil_seleccionado_required
def mis_viaticos(request):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    if usuario.sustituto:
        usuario = Profile.objects.filter(staff = usuario.staff, tipo__subdirector = True, st_activo = True).first()
    
    viaticos = Solicitud_Viatico.objects.filter(Q(staff = usuario) |Q(colaborador = usuario), complete=True).order_by('-folio')
    myfilter = Solicitud_Viatico_Filter(request.GET, queryset=viaticos)
    viaticos = myfilter.qs



    for viatico in viaticos:
        viatico.creado_reciente = (viatico.approved_at2 >= timezone.now() - timedelta(days=30)) if viatico.approved_at2 else True

    p = Paginator(viaticos, 20)
    page = request.GET.get('page')
    viaticos_list = p.get_page(page)
        
    context= {
        'viaticos':viaticos,
        'myfilter':myfilter,
        'viaticos_list': viaticos_list,
        }

    return render(request, 'tesoreria/mis_viaticos.html',context)

@perfil_seleccionado_required
def mis_comprobaciones_gasto(request):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    a√±o_actual = datetime.now().year
    a√±o_anterior = a√±o_actual - 1
    inicio = datetime(a√±o_anterior, 1, 1)
    fin = datetime(a√±o_actual, 12, 31, 23, 59, 59)

    print(a√±o_actual)
    print(a√±o_anterior)
    # QuerySet base (por usuario, fechas, autorizaciones, etc.)
    gastos = Solicitud_Gasto.objects.filter(
        Q(staff=usuario) | Q(colaborador=usuario),
        autorizar2=True,
        created_at__range=(inicio, fin),
        complete=True
    ).annotate(
        facturas_hechas=Count('facturas', filter=Q(facturas__hecho=True))
    ).order_by('-folio')
    
    #).annotate(total_facturas=Count('facturas', filter=Q(facturas__hecho=True)),autorizadas=Count(Case(When(Q(facturas__hecho=True), then=Value(1))))
    #            ).order_by('-folio')
    print('gastos:',gastos)
    
    myfilter = Solicitud_Gasto_Filter(request.GET, queryset=gastos)
    gastos = myfilter.qs

    
   
    suma = decimal.Decimal(0)
    total_monto_gastos = decimal.Decimal(0)
    total_todas_facturas = decimal.Decimal(0)
    for gasto in gastos:
        
        suma = decimal.Decimal('0')
       
        total_monto_gastos += gasto.get_total_solicitud
        for factura in gasto.facturas.all():
            if factura.archivo_xml and factura.hecho: 
                try:
                    if factura.emisor is not None:
                        suma += decimal.Decimal(factura.emisor['total'])
                except FileNotFoundError:
                    # Ignorar o registrar si el archivo XML no existe
                    pass
        gasto.suma_total_facturas = suma
        total_todas_facturas += gasto.suma_total_facturas

    p = Paginator(gastos, 25)
    page = request.GET.get('page')
    gastos_list = p.get_page(page)

    if request.method =='POST':
        if 'btnExcel' in request.POST:
            return convert_comprobacion_gastos_to_xls2(gastos, a√±o_actual,total_todas_facturas,total_monto_gastos)
            
        if 'btnImprimir' in request.POST or 'btnCorreo' in request.POST:
            # Obtener los IDs de los gastos seleccionados
            ids = request.POST.getlist('gastos')
            print(ids)
            tabla_gastos = """
            <table border="1" cellpadding="6" cellspacing="0" style="border-collapse: collapse; font-size: 14px; width: 100%;">
                <thead style="background-color: #eaeaea;">
                    <tr>
                        <th>Folio</th>
                        <th>Monto Solicitado</th>
                        <th>Monto Comprobado</th>
                        <th>Diferencia</th>
                    </tr>
                </thead>
                <tbody>
            """
            gastos_enviar = Solicitud_Gasto.objects.filter(id__in=ids).prefetch_related('facturas', 'articulos')
            merger = PdfMerger()
            folios_gastos_enviados = []
           
            for gasto in gastos_enviar:
                # 1. Generar la car√°tula en memoria
                gasto.suma_total_facturas = decimal.Decimal(0)
                buffer = render_pdf_gasto(gasto.id)
                with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_caratula:
                    temp_caratula.write(buffer.read())
                    caratula_path = temp_caratula.name
                merger.append(caratula_path)
                folios_gastos_enviados.append(str(gasto.folio))
                # 2. Comprobantes de pago (justo despu√©s de la car√°tula)
                for pago in gasto.pagosg.filter(hecho=True):
                    #print('esta entrando pago',pago)
                    if pago.comprobante_pago and os.path.exists(pago.comprobante_pago.path):
                        try:
                            merger.append(pago.comprobante_pago.path, import_outline=False)
                        except Exception as e:
                            print(f"Error al agregar comprobante de pago para gasto {gasto.folio}: {e}")

                suma = 0
                for factura in gasto.facturas.all():
                    
                    if factura.archivo_pdf and factura.hecho: 
                        path = factura.archivo_pdf.path
                        if os.path.exists(path):  # ‚úÖ Validaci√≥n que te falt√≥
                            merger.append(path)
                        else:
                            print(f"Archivo no encontrado: {path}")
                    elif factura.archivo_xml and factura.hecho:
                        try:
                            buffer = crear_pdf_cfdi_buffer(factura)  # <-- aqu√≠ llamas a tu funci√≥n que genera el PDF desde XML
                            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_pdf:
                                temp_pdf.write(buffer.read())
                                temp_pdf.flush()
                                merger.append(temp_pdf.name, import_outline=False)
                        except Exception as e:
                            print(f"Error al generar PDF desde XML para factura {factura.id}: {str(e)}") 

                    if factura.archivo_xml and factura.hecho: 
                        try:
                            if factura.emisor is not None:
                                suma += decimal.Decimal(factura.emisor['total'])
                        except FileNotFoundError:
                            # Ignorar o registrar si el archivo XML no existe
                            pass
                        gasto.suma_total_facturas = suma    
                    
                    
                
                comprobado = gasto.suma_total_facturas
                monto = gasto.get_total_solicitud  # o gasto.monto si lo prefieres
                diferencia = comprobado - monto
                tabla_gastos += f"""
                    <tr>
                        <td>{gasto.folio}</td>
                        <td>${monto:,.2f}</td>
                        <td>${comprobado:,.2f}</td>
                        <td style="color: {'green' if diferencia > 0 else 'red' if diferencia < 0 else 'black'};">
                            ${diferencia:,.2f}
                        </td>
                    </tr>
                """


                 # ‚ö†Ô∏è Salimos del bucle antes de escribir o cerrar
            if not merger.pages:
                return HttpResponse("No hay facturas v√°lidas para imprimir.", content_type="text/plain")

           
           
            if 'btnImprimir' in request.POST:
                tipo = 'gasto'
                acuse_buffer = generar_acuse_recibo(usuario, gastos_enviar, tipo)
                with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_acuse:
                    temp_acuse.write(acuse_buffer.read())
                    acuse_path = temp_acuse.name
                merger.append(acuse_path)
            
            
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_file:
                merger.write(temp_file.name)
                temp_file_path = temp_file.name

           
            merger.close()
            print('temp_file_path:', temp_file_path)
            if 'btnImprimir' in request.POST:
                #return FileResponse(open(temp_file_path, 'rb'), content_type='application/pdf')
                # Guarda la ruta del archivo temporal en la sesi√≥n
                request.session['temp_pdf_path'] = temp_file_path
                return redirect('mostrar-pdf')
            if 'btnCorreo' in request.POST:
                # 2. Leer el contenido del archivo temporal
                with open(temp_file_path, 'rb') as f:
                    contenido_pdf = f.read()
                nombre_archivo = f"Comprobacion_gastos_{'_'.join(folios_gastos_enviados)}.pdf"
                static_path = settings.STATIC_ROOT
                img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
                img_path2 = os.path.join(static_path,'images','logo_vordcab.jpg')
                image_base64 = get_image_base64(img_path)
                logo_v_base64 = get_image_base64(img_path2)
                html_message = f"""
                <html>
                    <head>
                        <meta charset="UTF-8">
                    </head>
                    <body style="font-family: Arial, sans-serif; color: #333; background-color: #f4f4f4; margin: 0; padding: 0;">
                        <table width="100%" cellspacing="0" cellpadding="0" style="background-color: #f4f4f4; padding: 20px;">
                            <tr>
                                <td align="center">
                                    <table width="600px" cellspacing="0" cellpadding="0" style="background-color: #ffffff; padding: 20px; border-radius: 10px;">
                                        <tr>
                                            <td align="center">
                                                <img src="data:image/jpeg;base64,{logo_v_base64}" alt="Logo" style="width: 100px; height: auto;" />
                                            </td>
                                        </tr>
                                        <tr>
                                            <td style="padding: 20px;">
                                                <p style="font-size: 18px; text-align: justify;">
                                                    A quien corresponda,
                                                </p>
                                                <p style="font-size: 18px; text-align: justify;">
                                                    Adjunto la comprobaci√≥n de gastos de los siguientes folios: {', '.join(folios_gastos_enviados)},
                                                </p>
                                            </td>
                                        </tr>
                                        <tr>
                                            <td style="padding: 20px;">
                                            {tabla_gastos}
                                            </td>
                                        </tr>
                                    </table>
                                    <table width="600px" cellspacing="0" cellpadding="0" style="background-color: #ffffff; padding: 20px; border-radius: 10px;">
                                        <tr>
                                            <td>
                                                <p style="font-size: 18px; text-align: justify;">
                                                Atte.
                                                </p>
                                                <p style="font-size: 18px; text-align: justify;">
                                                {request.user.get_full_name()}
                                                </p>
                                                <p style="text-align: center; margin: 20px 0;">
                                                    <img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width: 50px; height: auto; border-radius: 50%;" />
                                                </p>
                                                <p style="font-size: 14px; color: #999; text-align: justify;">
                                                    Este mensaje ha sido generado por SAVIA 2.0
                                                </p>
                                            </td>
                                        </tr>
                                    </table>
                                </td>
                            </tr>
                        </table>
                    </body>
                </html>
                """
                # Lista base con el remitente y alguien m√°s si quieres
                correos = [request.user.email]
                #otros_correos = []
                # Si se marc√≥ el checkbox de RH
                if request.POST.get('enviarRH'):
                    personal_rh = Profile.objects.filter(tipo__rh=True, distritos=usuario.distritos, st_activo = True, tipo__documentos = True)
                    for persona in personal_rh:
                        if persona.staff.staff.email:
                            correos.append(persona.staff.staff.email)

                # Si se marc√≥ el checkbox de contabilidad y tesorer√≠a
                if request.POST.get('enviarContabilidad'):
                    personal_ct = Profile.objects.filter(tipo__tesoreria=True, tipo__rh = False, distritos=usuario.distritos, st_activo = True, tipo__documentos = True)
                    for persona in personal_ct:
                        if persona.staff.staff.email:
                            correos.append(persona.staff.staff.email)

                # Elimina duplicados
                correos = list(set(correos))
                print('correos:',correos)
                email = EmailMessage(
                subject=f"Comprobaci√≥n de Gastos - {request.user.get_full_name()} - G{', '.join(folios_gastos_enviados)}",
                body=html_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to = correos,   
                headers={'Content-Type': 'text/html'}
            )
            email.attach(nombre_archivo, contenido_pdf, 'application/pdf')
            # Adjuntar los XMLs de las facturas
            for gasto in gastos_enviar:
                for factura in gasto.facturas.all():
                    if factura.archivo_xml and factura.hecho:
                        try:
                            path = factura.archivo_xml.path
                            if os.path.exists(path):
                                nombre_xml = f"G{gasto.folio}_F{factura.uuid}.xml"
                                with open(path, 'rb') as f:
                                    email.attach(nombre_xml, f.read(), 'application/xml')
                            else:
                                print(f"XML no encontrado: {path}")
                        except Exception as e:
                            print(f"Error al adjuntar XML: {e}")
            email.content_subtype = "html"
            email.send()

            messages.success(request, "El correo fue enviado exitosamente.")
            return redirect('mis-comprobaciones-gasto')
        
    
    context= {
        'gastos':gastos,
        'total_todas_facturas':total_todas_facturas,
        'total_monto_gastos':total_monto_gastos,
        'a√±o_actual':str(a√±o_actual),
        'a√±o_anterior':str(a√±o_anterior),
        'myfilter':myfilter,
        'gastos_list': gastos_list,
        }

    return render(request, 'tesoreria/mis_comprobaciones_gasto.html',context)

@perfil_seleccionado_required
def mis_comprobaciones_viaticos(request):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    a√±o_actual = datetime.now().year
    a√±o_anterior = a√±o_actual - 1
    inicio = datetime(a√±o_anterior, 1, 1)
    fin = datetime(a√±o_actual, 12, 31, 23, 59, 59)



    viaticos = Solicitud_Viatico.objects.filter(
        Q(staff=usuario) | Q(colaborador=usuario),
        autorizar2=True,
        created_at__range=(inicio, fin),
        complete=True
    ).annotate(
        total_facturas=Count('facturas', filter=Q(facturas__hecho=True))                    
    ).order_by('-folio')
    print(viaticos)
    suma = decimal.Decimal(0)
    total_monto_viaticos = decimal.Decimal(0)
    total_todas_facturas = decimal.Decimal(0)
    for viatico in viaticos:
        suma = decimal.Decimal('0')
        total_monto_viaticos += viatico.get_total
        for factura in viatico.facturas.all():
            if factura.factura_xml and factura.hecho: 
                try:
                    if factura.emisor is not None:
                        suma += decimal.Decimal(factura.emisor['total'])
                except FileNotFoundError:
                    # Ignorar o registrar si el archivo XML no existe
                    pass
        viatico.suma_total_facturas = suma
        total_todas_facturas += viatico.suma_total_facturas

    
    if request.method =='POST':
        if 'btnExcel' in request.POST:
            return convert_comprobacion_viaticos_to_xls2(viaticos, a√±o_actual,total_todas_facturas,total_monto_viaticos)
        if 'btnImprimir' in request.POST or 'btnCorreo' in request.POST:
            # Obtener los IDs de los gastos seleccionados
            ids = request.POST.getlist('gastos')
            print(ids)
            tabla_viaticos = """
            <table border="1" cellpadding="6" cellspacing="0" style="border-collapse: collapse; font-size: 14px; width: 100%;">
                <thead style="background-color: #eaeaea;">
                    <tr>
                        <th>Folio</th>
                        <th>Monto Solicitado</th>
                        <th>Monto Comprobado</th>
                        <th>Diferencia</th>
                    </tr>
                </thead>
                <tbody>
            """
            viaticos_enviar = Solicitud_Viatico.objects.filter(id__in=ids).prefetch_related('facturas', 'conceptos')
            merger = PdfMerger()
            folios_viaticos_enviados = []
           
            for viatico in viaticos_enviar:
                # 1. Generar la car√°tula en memoria
                buffer =generar_pdf_viatico(viatico.id) # >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>><
                with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_caratula:
                    temp_caratula.write(buffer.read())
                    caratula_path = temp_caratula.name
                merger.append(caratula_path)
                folios_viaticos_enviados.append(str(viatico.folio))
                # 2. Comprobantes de pago (justo despu√©s de la car√°tula)
                for pago in viatico.pagosv.filter(hecho=True):
                    #print('esta entrando pago',pago)
                    if pago.comprobante_pago and os.path.exists(pago.comprobante_pago.path):
                        try:
                            merger.append(pago.comprobante_pago.path, import_outline=False)
                        except Exception as e:
                            print(f"Error al agregar comprobante de pago para gasto {viatico.folio}: {e}")

                suma = 0
                for factura in viatico.facturas.all():
                    
                    if factura.factura_pdf and factura.hecho: 
                        path = factura.factura_pdf.path
                        if os.path.exists(path):  # ‚úÖ Validaci√≥n que te falt√≥
                            merger.append(path)
                        else:
                            print(f"Archivo no encontrado: {path}")
                    elif factura.factura_xml and factura.hecho:
                        try:
                            buffer = crear_pdf_cfdi_buffer(factura)  # <-- aqu√≠ llamas a tu funci√≥n que genera el PDF desde XML
                            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_pdf:
                                temp_pdf.write(buffer.read())
                                temp_pdf.flush()
                                merger.append(temp_pdf.name, import_outline=False)
                        except Exception as e:
                            print(f"Error al generar PDF desde XML para factura {factura.id}: {str(e)}") 

                    if factura.factura_xml and factura.hecho: 
                        try:
                            if factura.emisor is not None:
                                suma += decimal.Decimal(factura.emisor['total'])
                        except FileNotFoundError:
                            # Ignorar o registrar si el archivo XML no existe
                            pass
                        viatico.suma_total_facturas = suma    
                    
                    
                
                comprobado = viatico.suma_total_facturas
                monto = viatico.get_total  # o gasto.monto si lo prefieres
                diferencia = comprobado - monto
                tabla_viaticos += f"""
                    <tr>
                        <td>{viatico.folio}</td>
                        <td>${monto:,.2f}</td>
                        <td>${comprobado:,.2f}</td>
                        <td style="color: {'green' if diferencia > 0 else 'red' if diferencia < 0 else 'black'};">
                            ${diferencia:,.2f}
                        </td>
                    </tr>
                """


                 # ‚ö†Ô∏è Salimos del bucle antes de escribir o cerrar
            if not merger.pages:
                return HttpResponse("No hay facturas v√°lidas para imprimir.", content_type="text/plain")

           
           
            if 'btnImprimir' in request.POST:
                tipo = 'viatico'
                acuse_buffer = generar_acuse_recibo(usuario, viaticos_enviar, tipo)
                with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_acuse:
                    temp_acuse.write(acuse_buffer.read())
                    acuse_path = temp_acuse.name
                merger.append(acuse_path)
            
            
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_file:
                merger.write(temp_file.name)
                temp_file_path = temp_file.name

           
            merger.close()
            print('temp_file_path:', temp_file_path)
            if 'btnImprimir' in request.POST:
                #return FileResponse(open(temp_file_path, 'rb'), content_type='application/pdf')
                # Guarda la ruta del archivo temporal en la sesi√≥n
                request.session['temp_pdf_path'] = temp_file_path
                return redirect('mostrar-pdf')
            if 'btnCorreo' in request.POST:
                # 2. Leer el contenido del archivo temporal
                with open(temp_file_path, 'rb') as f:
                    contenido_pdf = f.read()
                nombre_archivo = f"Comprobacion_viaticos_{'_'.join(folios_viaticos_enviados)}.pdf"
                static_path = settings.STATIC_ROOT
                img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
                img_path2 = os.path.join(static_path,'images','logo_vordcab.jpg')
                image_base64 = get_image_base64(img_path)
                logo_v_base64 = get_image_base64(img_path2)
                html_message = f"""
                <html>
                    <head>
                        <meta charset="UTF-8">
                    </head>
                    <body style="font-family: Arial, sans-serif; color: #333; background-color: #f4f4f4; margin: 0; padding: 0;">
                        <table width="100%" cellspacing="0" cellpadding="0" style="background-color: #f4f4f4; padding: 20px;">
                            <tr>
                                <td align="center">
                                    <table width="600px" cellspacing="0" cellpadding="0" style="background-color: #ffffff; padding: 20px; border-radius: 10px;">
                                        <tr>
                                            <td align="center">
                                                <img src="data:image/jpeg;base64,{logo_v_base64}" alt="Logo" style="width: 100px; height: auto;" />
                                            </td>
                                        </tr>
                                        <tr>
                                            <td style="padding: 20px;">
                                                <p style="font-size: 18px; text-align: justify;">
                                                    A quien corresponda,
                                                </p>
                                                <p style="font-size: 18px; text-align: justify;">
                                                    Adjunto la comprobaci√≥n de vi√°ticos de los siguientes folios: {', '.join(folios_viaticos_enviados)},
                                                </p>
                                            </td>
                                        </tr>
                                        <tr>
                                            <td style="padding: 20px;">
                                            {tabla_viaticos}
                                            </td>
                                        </tr>
                                    </table>
                                    <table width="600px" cellspacing="0" cellpadding="0" style="background-color: #ffffff; padding: 20px; border-radius: 10px;">
                                        <tr>
                                            <td>
                                                <p style="font-size: 18px; text-align: justify;">
                                                Atte.
                                                </p>
                                                <p style="font-size: 18px; text-align: justify;">
                                                {request.user.get_full_name()}
                                                </p>
                                                <p style="text-align: center; margin: 20px 0;">
                                                    <img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width: 50px; height: auto; border-radius: 50%;" />
                                                </p>
                                                <p style="font-size: 14px; color: #999; text-align: justify;">
                                                    Este mensaje ha sido generado por SAVIA 2.0
                                                </p>
                                            </td>
                                        </tr>
                                    </table>
                                </td>
                            </tr>
                        </table>
                    </body>
                </html>
                """
                # Lista base con el remitente y alguien m√°s si quieres
                correos = [request.user.email]
                #otros_correos = []
                # Si se marc√≥ el checkbox de RH
                if request.POST.get('enviarRH'):
                    personal_rh = Profile.objects.filter(tipo__rh=True, distritos=usuario.distritos, st_activo = True, tipo__documentos = True)
                    for persona in personal_rh:
                        if persona.staff.staff.email:
                            correos.append(persona.staff.staff.email)

                # Si se marc√≥ el checkbox de contabilidad y tesorer√≠a
                if request.POST.get('enviarContabilidad'):
                    personal_ct = Profile.objects.filter(tipo__tesoreria=True, tipo__rh = False, distritos=usuario.distritos, st_activo = True, tipo__documentos = True)
                    for persona in personal_ct:
                        if persona.staff.staff.email:
                            correos.append(persona.staff.staff.email)

                # Elimina duplicados
                correos = list(set(correos))
                print('correos:',correos)
                email = EmailMessage(
                subject=f"Comprobaci√≥n de Vi√°ticos - {request.user.get_full_name()} - G{', '.join(folios_viaticos_enviados)}",
                body=html_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to = correos,   
                headers={'Content-Type': 'text/html'}
            )
            email.attach(nombre_archivo, contenido_pdf, 'application/pdf')
            # Adjuntar los XMLs de las facturas
            for viatico in viaticos_enviar:
                for factura in viatico.facturas.all():
                    if factura.factura_xml and factura.hecho:
                        try:
                            path = factura.factura_xml.path
                            if os.path.exists(path):
                                nombre_xml = f"V{viatico.folio}_F{factura.uuid}.xml"
                                with open(path, 'rb') as f:
                                    email.attach(nombre_xml, f.read(), 'application/xml')
                            else:
                                print(f"XML no encontrado: {path}")
                        except Exception as e:
                            print(f"Error al adjuntar XML: {e}")
            email.content_subtype = "html"
            email.send()

            messages.success(request, "El correo fue enviado exitosamente.")
            return redirect('mis-comprobaciones-gasto')
    

    context= {
        'viaticos':viaticos,
        'total_todas_facturas':total_todas_facturas,
        'total_monto_viaticos':total_monto_viaticos,
        'a√±o_actual':str(a√±o_actual),
        'a√±o_anterior':str(a√±o_anterior),
        #'myfilter':myfilter,
        }

    return render(request, 'tesoreria/mis_comprobaciones_viaticos.html',context)

#@perfil_seleccionado_required
def mostrar_pdf(request):
    pdf_path = request.session.get('temp_pdf_path')

    if not pdf_path or not os.path.exists(pdf_path):
        return HttpResponse("Archivo PDF no encontrado.", status=404)

    return FileResponse(open(pdf_path, 'rb'), content_type='application/pdf')


def convert_comprobacion_gastos_to_xls2(entradas, a√±o_actual, total_todas_facturas, total_monto_gastos):
    # Crea un objeto BytesIO para guardar el archivo Excel
    output = BytesIO()

    # Crea un libro de trabajo y a√±ade una hoja
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet("Mis_gastos_" + str(a√±o_actual))

     
    date_format = workbook.add_format({'num_format': 'dd/mm/yyyy'})
    # Define los estilos
    head_style = workbook.add_format({'bold': True, 'font_color': 'FFFFFF', 'bg_color': '333366', 'font_name': 'Arial', 'font_size': 11})
    body_style = workbook.add_format({'font_name': 'Calibri', 'font_size': 10})
    money_style = workbook.add_format({'num_format': '$ #,##0.00', 'font_name': 'Calibri', 'font_size': 10})
    date_style = workbook.add_format({'num_format': 'dd/mm/yyyy', 'font_name': 'Calibri', 'font_size': 10})
    percent_style = workbook.add_format({'num_format': '0.00%', 'font_name': 'Calibri', 'font_size': 10})
    messages_style = workbook.add_format({'font_name':'Arial Narrow', 'font_size':11})

    #columns = ['Folio Solicitud', 'Solicitante', 'Almacenista','Proyecto', 'Subproyecto', 'Fecha creaci√≥n','Productos','Tipo','Autorizada','Fecha autorizaci√≥n','Comentario']
    columns = ['Folio Gasto','Comentario', 'Solicitante', 'Importe','Monto XML',]

    columna_max = len(columns)+2

    worksheet.write(0, columna_max - 1, 'Reporte Creado Autom√°ticamente por SAVIA 2.0 Vordcab. UH', messages_style)
    worksheet.write(1, columna_max - 1, 'Software desarrollado por Grupo Vordcab S.A. de C.V.', messages_style)
    worksheet.write(2, columna_max - 1, 'Monto total de facturas:', messages_style)
    worksheet.write(2, columna_max, total_todas_facturas, messages_style)
    worksheet.write(3, columna_max - 1, 'Monto total de importe: ', messages_style)
    worksheet.write(3, columna_max, total_monto_gastos, messages_style)
    worksheet.set_column(columna_max - 1, columna_max, 30)  # Ajusta el ancho de las columnas nuevas

    for i, column in enumerate(columns):
        worksheet.write(0, i, column, head_style)
        worksheet.set_column(i, i, 15)  # Ajusta el ancho de las columnas

    row_num = 0
    for gasto in entradas:
        row_num += 1
        # Crear la lista de productos con nombre y cantidad
        #productos_lista = [
        #    f"{producto['producto__producto__nombre']} (Cantidad: {producto['cantidad']})"
        #    for producto in dev.solicitud.productos.values('producto__producto__nombre', 'cantidad')
        #]
        # Unir la lista en una cadena
        #productos_str = ", ".join(productos_lista)

        row = [
            gasto.folio,
            gasto.comentario,
            f"{gasto.staff.staff.staff.first_name} {gasto.staff.staff.staff.last_name}",
            gasto.get_total_solicitud,
            gasto.suma_total_facturas,
        ]
        
        for col_num, cell_value in enumerate(row):
        # Define el formato por defecto
            cell_format = body_style

            # Finalmente, escribe la celda con el valor y el formato correspondiente
            worksheet.write(row_num, col_num, cell_value, cell_format)

      
        #worksheet.write_formula(row_num, 19, f'=IF(ISBLANK(R{row_num+1}), L{row_num+1}, L{row_num+1}*R{row_num+1})', money_style)
    
   
    workbook.close()

    # Construye la respuesta
    output.seek(0)

    response = HttpResponse(
        output.read(), 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    response['Content-Disposition'] = f'attachment; filename=Mis_gastos_{a√±o_actual} {dt.date.today()}.xlsx'

      # Establecer una cookie para indicar que la descarga ha iniciado
    response.set_cookie('descarga_iniciada', 'true', max_age=20)  # La cookie expira en 20 segundos
    output.close()
    return response

def convert_comprobacion_viaticos_to_xls2(entradas, a√±o_actual, total_todas_facturas, total_monto_viaticos):
    # Crea un objeto BytesIO para guardar el archivo Excel
    output = BytesIO()

    # Crea un libro de trabajo y a√±ade una hoja
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet("Mis_viaticos_" + str(a√±o_actual))

     
    date_format = workbook.add_format({'num_format': 'dd/mm/yyyy'})
    # Define los estilos
    head_style = workbook.add_format({'bold': True, 'font_color': 'FFFFFF', 'bg_color': '333366', 'font_name': 'Arial', 'font_size': 11})
    body_style = workbook.add_format({'font_name': 'Calibri', 'font_size': 10})
    money_style = workbook.add_format({'num_format': '$ #,##0.00', 'font_name': 'Calibri', 'font_size': 10})
    date_style = workbook.add_format({'num_format': 'dd/mm/yyyy', 'font_name': 'Calibri', 'font_size': 10})
    percent_style = workbook.add_format({'num_format': '0.00%', 'font_name': 'Calibri', 'font_size': 10})
    messages_style = workbook.add_format({'font_name':'Arial Narrow', 'font_size':11})

    #columns = ['Folio Solicitud', 'Solicitante', 'Almacenista','Proyecto', 'Subproyecto', 'Fecha creaci√≥n','Productos','Tipo','Autorizada','Fecha autorizaci√≥n','Comentario']
    columns = ['Folio Viatico','Motivo', 'Solicitante', 'Importe','Monto XML',]

    columna_max = len(columns)+2

    worksheet.write(0, columna_max - 1, 'Reporte Creado Autom√°ticamente por SAVIA 2.0 Vordcab. UH', messages_style)
    worksheet.write(1, columna_max - 1, 'Software desarrollado por Grupo Vordcab S.A. de C.V.', messages_style)
    worksheet.write(2, columna_max - 1, 'Monto total de facturas:', messages_style)
    worksheet.write(2, columna_max, total_todas_facturas, messages_style)
    worksheet.write(3, columna_max - 1, 'Monto total de importe: ', messages_style)
    worksheet.write(3, columna_max, total_monto_viaticos, messages_style)
    worksheet.set_column(columna_max - 1, columna_max, 30)  # Ajusta el ancho de las columnas nuevas

    for i, column in enumerate(columns):
        worksheet.write(0, i, column, head_style)
        worksheet.set_column(i, i, 15)  # Ajusta el ancho de las columnas

    row_num = 0
    for viatico in entradas:
        row_num += 1
        # Crear la lista de productos con nombre y cantidad
        #productos_lista = [
        #    f"{producto['producto__producto__nombre']} (Cantidad: {producto['cantidad']})"
        #    for producto in dev.solicitud.productos.values('producto__producto__nombre', 'cantidad')
        #]
        # Unir la lista en una cadena
        #productos_str = ", ".join(productos_lista)

        row = [
            viatico.folio,
            viatico.motivo,
            f"{viatico.staff.staff.staff.first_name} {viatico.staff.staff.staff.last_name}",
            viatico.get_total,
            viatico.suma_total_facturas,
        ]
        
        for col_num, cell_value in enumerate(row):
        # Define el formato por defecto
            cell_format = body_style

            # Finalmente, escribe la celda con el valor y el formato correspondiente
            worksheet.write(row_num, col_num, cell_value, cell_format)

      
        #worksheet.write_formula(row_num, 19, f'=IF(ISBLANK(R{row_num+1}), L{row_num+1}, L{row_num+1}*R{row_num+1})', money_style)
    
   
    workbook.close()

    # Construye la respuesta
    output.seek(0)

    response = HttpResponse(
        output.read(), 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    response['Content-Disposition'] = f'attachment; filename=Mis_viaticos_{a√±o_actual} {dt.date.today()}.xlsx'
      # Establecer una cookie para indicar que la descarga ha iniciado
    response.set_cookie('descarga_iniciada', 'true', max_age=20)  # La cookie expira en 20 segundos
    output.close()
    return response

def convert_excel_matriz_compras_autorizadas(compras):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Pendientes_de_pago_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Compras Autorizadas')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')
    percent_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(percent_style)

    columns = ['Folio','Fecha Autorizaci√≥n','Proyecto','Subproyecto','Distrito','Proveedor','C. Pago',
               'Importe', 'Moneda','Tipo de cambio','Total en Pesos','Fecha Creaci√≥n','Recibida']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16
        if col_num == 5: #Columna del proveedor
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        if col_num == 2:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 20

    columna_max = len(columns)+2

    # Agregar los mensajes
    ws.cell(column = columna_max, row = 1, value='{Reporte Creado Autom√°ticamente por SAVIA 2.0. UH}').style = messages_style
    ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}').style = messages_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 30

    # Agregar los encabezados de las nuevas columnas debajo de los mensajes
    ws.cell(row=3, column = columna_max, value="Total de OC's").style = head_style
    ws.cell(row=4, column = columna_max, value="Sumatoria de Pagos Pendientes").style = head_style
   

    # Asumiendo que las filas de datos comienzan en la fila 2 y terminan en row_num
    ws.cell(row=3, column=columna_max + 1, value=f"=COUNTA(A:A)-1").style = body_style
    ws.cell(row=4, column=columna_max + 1, value=f"=SUM(K:K)").style = money_resumen_style
  
    
   
    
    for compra in compras:
        row_num = row_num + 1    
        #productos = ArticuloComprado.objects.filter(oc = compra)

        # Unir los nombres de los productos en una sola cadena separada por comas
        #productos_texto = ', '.join([producto.nombre for producto in productos])
        # Manejar autorizado_at_2
        if compra.autorizado_at_2 and isinstance(compra.autorizado_at_2, datetime):
        # Si autorizado_at_2 es timezone-aware, convi√©rtelo a timezone-naive
            autorizado_at_2_naive = compra.autorizado_at_2.astimezone(pytz.utc).replace(tzinfo=None)
        else:
            autorizado_at_2_naive = ''
        
        # Manejar created_at
        if compra.created_at and isinstance(compra.created_at, datetime):
        # Si created_at es timezone-aware, convi√©rtelo a timezone-naive
            created_at_naive = compra.created_at.astimezone(pytz.utc).replace(tzinfo=None)
        else:
            created_at_naive = ''

        recibida = "Recibida" if compra.entrada_completa else "No Recibida"

        row = [
            compra.folio,
            autorizado_at_2_naive,
            compra.req.orden.proyecto.nombre,
            compra.req.orden.subproyecto.nombre,
            compra.req.orden.distrito.nombre,
            compra.proveedor.nombre.razon_social,
            compra.cond_de_pago.nombre,
            compra.costo_plus_adicionales,
            compra.moneda.nombre,
            compra.tipo_de_cambio if compra.tipo_de_cambio else '',
            f'=IF(J{row_num}="",H{row_num},H{row_num}*J{row_num})',  # Calcula total en pesos usando la f√≥rmula de Excel
            created_at_naive,
            recibida,
        ]

    
        for col_num in range(len(row)):
            (ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style
            if col_num == 1 or col_num == 11:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style
            if col_num == 7 or col_num == 9 or col_num == 10:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = money_style
       
    
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)
    response.set_cookie('descarga_iniciada', 'true', max_age=20)  # La cookie expira en 20 segundos
    return(response)

def convert_excel_matriz_tiempo_proceso(compras):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Tiempos_proceso_sol-oc' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Compras Autorizadas')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')
    percent_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(percent_style)

    columns = ['Folio OC','Folio Req','Folio Sol','Distrito','Fecha Creaci√≥n Sol','Fecha Autorizaci√≥n Sol', 'Fecha Creaci√≥n Req',
               'Fecha Autorizaci√≥n Req','Fecha Creaci√≥n OC','Fecha Autorizaci√≥n OC','Fecha Autorizaci√≥n OC 2','Tiempo Autorizaci√≥n Sol (horas)',
               'Tiempo Proceso Req (horas)', 'Tiempo Autorizaci√≥n Req (horas)', 'Tiempo Proceso OC (horas)', 'Tiempo Proceso OC Autorizaci√≥n 1 (horas)',
               'Tiempo OC Autorizaci√≥n 2 (horas)', 'Total',
               ]

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16
        if col_num == 5: #Columna del proveedor
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        if col_num == 2:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 20

    columna_max = len(columns)+2

    # Agregar los mensajes
    ws.cell(column = columna_max, row = 1, value='{Reporte Creado Autom√°ticamente por SAVIA 2.0. UH}').style = messages_style
    ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}').style = messages_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 30

    # Agregar los encabezados de las nuevas columnas debajo de los mensajes
    ws.cell(row=3, column = columna_max, value="Total de OC's").style = head_style
    #ws.cell(row=4, column = columna_max, value="Sumatoria de Pagos Pendientes").style = head_style
   

    # Asumiendo que las filas de datos comienzan en la fila 2 y terminan en row_num
    ws.cell(row=3, column=columna_max + 1, value=f"=COUNTA(A:A)-1").style = body_style
    #ws.cell(row=4, column=columna_max + 1, value=f"=SUM(R:R)").style = money_resumen_style
    
    for compra in compras:
        row_num = row_num + 1    
       
        if compra.autorizado_at_2 and isinstance(compra.autorizado_at_2, datetime):
        # Si autorizado_at_2 es timezone-aware, convi√©rtelo a timezone-naive
            autorizado_at_2_naive = compra.autorizado_at_2.astimezone(pytz.utc).replace(tzinfo=None)
        else:
            autorizado_at_2_naive = ''
        
        # Manejar created_at
        if compra.created_at and isinstance(compra.created_at, datetime):
        # Si created_at es timezone-aware, convi√©rtelo a timezone-naive
            created_at_naive = compra.created_at.astimezone(pytz.utc).replace(tzinfo=None)
        else:
            created_at_naive = ''

        fecha_creacion_orden = (
            localtime(compra.req.orden.created_at).date()
            if compra.req.orden and compra.req.orden.created_at
            else ''
)
        row = [
            compra.folio,
            compra.req.folio,
            compra.req.orden.folio,
            compra.req.orden.distrito.nombre,
            fecha_creacion_orden,
            compra.req.orden.approved_at.date() if compra.req.orden.approved_at  else '',
            compra.req.created_at.date() if compra.req.created_at else '',
            compra.req.approved_at,
            created_at_naive.date(),
            compra.autorizado_at.date() if compra.autorizado_at else '',
            autorizado_at_2_naive.date(),
        ]

    
        for col_num in range(len(row)):
            (ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style
            if col_num in [4, 5, 6, 7, 8, 9, 10,11]:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style
        
        tiempo_solicitud = ws.cell(row=row_num, column=12, value=f"=F{row_num}-E{row_num}")
        tiempo_solicitud.style = body_style  # o usa money_style si corresponde
        tiempo_req_sol = ws.cell(row=row_num, column=13, value=f"=G{row_num}-F{row_num}")
        tiempo_req_sol.style = body_style  # o usa money_style si corresponde
        tiempo_autorizacion_req = ws.cell(row=row_num, column=14, value=f"=H{row_num}-G{row_num}")
        tiempo_autorizacion_req.style = body_style  # o usa money_style si corresponde
        tiempo_oc_req = ws.cell(row=row_num, column=15, value=f"=I{row_num}-H{row_num}")
        tiempo_oc_req.style = body_style  # o usa money_style si corresponde
        tiempo_oc_autorizacion_1 = ws.cell(row=row_num, column=16, value=f"=J{row_num}-I{row_num}")
        tiempo_oc_autorizacion_1.style = body_style  # o usa money_style si corresponde
        tiempo_oc_autorizacion_2 = ws.cell(row=row_num, column=17,  value=f"=K{row_num}-J{row_num}")    
        tiempo_total = ws.cell(row=row_num, column=18, value=f"=SUM(L{row_num}:Q{row_num})")
        tiempo_total.style = body_style  # o money_style si aplica

    
    
    sheet = wb['Sheet']
    wb.remove(sheet)
    ws_resumen = wb.create_sheet(title="Resumen")
    ws_resumen.column_dimensions['A'].width = 30  # o el n√∫mero que necesites
    # Encabezados
    headers = ["Tiempo Sol - Autorizaci√≥n", "Tiempo Almac√©n (Req.)", "Tiempo Autorizaci√≥n Req",
            "Tiempo Compras (OC)", "OC Autorizaci√≥n Suppte", "OC Autorizaci√≥n Gerente", "Total"]
    ws_resumen.append(["Etapa", "Promedio (d√≠as)"])  # encabezado de la tabla
    for i, titulo in enumerate(headers, start=1):
        col_letra = chr(76 + i - 1)  # L=76 en ASCII
        col_index = 12 + i - 1       # de la columna L (12) a la R (18)
        formula = f"=AVERAGE('{ws.title}'!{col_letra}2:{col_letra}{row_num})"
        ws_resumen.append([titulo, formula])

    # Crear el gr√°fico de barras
    chart = BarChart()
    chart.title = "Promedio de tiempos por etapa"
    chart.y_axis.title = "D√≠as"
    chart.x_axis.title = "Etapas"
    chart.width = 20
    chart.height = 15

    data = Reference(ws_resumen, min_col=2, min_row=2, max_row=8)  # valores
    cats = Reference(ws_resumen, min_col=1, min_row=2, max_row=8)  # categor√≠as
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)

    # Insertar el gr√°fico en la hoja
    ws_resumen.add_chart(chart, "D2")

    wb.save(response)
    response.set_cookie('descarga_iniciada', 'true', max_age=20)  # La cookie expira en 20 segundos
    return(response)


def convert_excel_matriz_compras_tesoreria(compras):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Pendientes_de_pago_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Compras Autorizadas')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')
    percent_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(percent_style)
    number_style = NamedStyle(name='number_style',number_format ='##0')
    number_style.font = Font(name ='Calibri', size = 10)

    columns = ['A√±o','Prioridad','Folio OC','Fecha Creaci√≥n','Fecha Autorizaci√≥n OC','Proyecto','Subproyecto','Distrito',
               'Proveedor','Producto','Banco', 'Cuenta Bancaria','Clabe','Moneda','Tipo de cambio','Importe','Total en Pesos','Importe Pagado',
               'Importe Restante','C. Pago', 'D√≠as de Cr√©dito','Recibida','Fecha Entrada','Factura','Folio UUID', 'Fecha Timbrado']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16
        if col_num == 5: #Columna del proveedor
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        if col_num == 2:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 20

    columna_max = len(columns)+2

    # Agregar los mensajes
    ws.cell(column = columna_max, row = 1, value='{Reporte Creado Autom√°ticamente por SAVIA 2.0. UH}').style = messages_style
    ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}').style = messages_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 30

    # Agregar los encabezados de las nuevas columnas debajo de los mensajes
    ws.cell(row=3, column = columna_max, value="Total de OC's").style = head_style
    ws.cell(row=4, column = columna_max, value="Sumatoria de Pagos Pendientes").style = head_style
   

    # Asumiendo que las filas de datos comienzan en la fila 2 y terminan en row_num
    ws.cell(row=3, column=columna_max + 1, value=f"=COUNTA(A:A)-1").style = body_style
    ws.cell(row=4, column=columna_max + 1, value=f"=SUM(S:S)").style = money_resumen_style
  
    
   
    
    for compra in compras:
        row_num = row_num + 1    
        productos = ArticuloComprado.objects.filter(oc = compra)
        ultima_fecha_entrada = compra.vale_entrada.filter(entrada_date__isnull=False).order_by('-entrada_date').values_list('entrada_date', flat=True).first()
        if ultima_fecha_entrada:
            ultima_fecha_entrada_naive = ultima_fecha_entrada.replace(tzinfo=None)
        else:
            ultima_fecha_entrada_naive = None

        # Unir los nombres de los productos en una sola cadena separada por comas
        productos_texto = ', '.join([producto.producto.producto.articulos.producto.producto.nombre for producto in productos])
        # Manejar autorizado_at_2
        if compra.autorizado_at_2 and isinstance(compra.autorizado_at_2, datetime):
        # Si autorizado_at_2 es timezone-aware, convi√©rtelo a timezone-naive
            autorizado_at_2_naive = compra.autorizado_at_2.astimezone(pytz.utc).replace(tzinfo=None)
            a√±o = autorizado_at_2_naive.year
        else:
            autorizado_at_2_naive = ''
            a√±o = ''

        if compra.tipo_prioridad:
            prioridad = compra.tipo_prioridad
        else:
            prioridad = '' 
        # Manejar created_at
        if compra.created_at and isinstance(compra.created_at, datetime):
        # Si created_at es timezone-aware, convi√©rtelo a timezone-naive
            created_at_naive = compra.created_at.astimezone(pytz.utc).replace(tzinfo=None)
        else:
            created_at_naive = ''

        if compra.facturas.filter(factura_xml__isnull=False).exists():
            tiene_facturas = 'S√≠'
            uuids = compra.facturas.filter(factura_xml__isnull=False, uuid__isnull=False).values_list('uuid', flat=True)
            uuid_string = "///".join(uuids)
            fechas_timbrado = compra.facturas.filter(factura_xml__isnull=False, fecha_timbrado__isnull=False).values_list('fecha_timbrado', flat=True)
            fecha_timbrado_string = ">>>".join([fecha.strftime("%d/%m/%Y") for fecha in fechas_timbrado])
        else:
            tiene_facturas = 'No'
            uuid_string = ''  # O None, seg√∫n lo que necesites
            fecha_timbrado_string = ''

        recibida = "Recibida" if compra.entrada_completa else "No Recibida"

        row = [
            a√±o,
            prioridad,
            compra.folio,
            created_at_naive,
            autorizado_at_2_naive,
            compra.req.orden.proyecto.nombre,
            compra.req.orden.subproyecto.nombre,
            compra.req.orden.distrito.nombre,
            compra.proveedor.nombre.razon_social,
            productos_texto,
            compra.proveedor.banco.nombre,
            compra.proveedor.cuenta,
            compra.proveedor.clabe,
            compra.moneda.nombre,
            compra.tipo_de_cambio if compra.tipo_de_cambio else '',
            compra.costo_plus_adicionales,
            # Calcula total en pesos usando la f√≥rmula de Excel
            f'=IF(O{row_num}="",P{row_num},O{row_num}*P{row_num})', 
            compra.monto_pagado,
            f'=Q{row_num} - R{row_num}',
            compra.cond_de_pago.nombre,
            compra.dias_de_credito if compra.dias_de_credito else '',
            recibida,
            ultima_fecha_entrada_naive or '',
            tiene_facturas,
            uuid_string,
            fecha_timbrado_string,
            
        ]

    
        for col_num in range(len(row)):
            (ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style
            if col_num in [0,2]:
                (ws.cell(row= row_num, column = col_num+1, value=row[col_num])).style = number_style
            if col_num in [3, 4, 28]:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style
            if col_num in [14, 15, 16, 17, 18]:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = money_style
       
    
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)
    response.set_cookie('descarga_iniciada', 'true', max_age=20)  # La cookie expira en 20 segundos
    return(response)


def convert_excel_matriz_pagos(pagos):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Matriz_pagos_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Pagos')
    #Comenzar en la fila 1
    row_num = 1

    # Funci√≥n para manejar los IDs de las compras, gastos o vi√°ticos
    def get_transaction_id(pago):
        if pago.oc:
            return 'OC'+str(pago.oc.folio)
        elif pago.gasto:
            return 'G'+str(pago.gasto.folio)
        elif pago.viatico:
            return 'V'+str(pago.viatico.folio)
        else:
            return None

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')
    percent_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(percent_style)

    columns = ['Distrito','Compra/Gasto','Solicitado','Autorizado','Fecha Creaci√≥n','Fecha Autorizaci√≥n','Proyecto','Subproyecto','Proveedor/Colaborador',
               'Producto/Concepto','Importe', 'Moneda','Tipo de cambio', 'Total en Pesos','Fecha de pago', 'Tiene Facturas',]

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16
        if col_num == 5:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 25

    columna_max = len(columns)+2

    # Agregar los mensajes
    ws.cell(column = columna_max, row = 1, value='{Reporte Creado Autom√°ticamente por SAVIA 2.0. UH}').style = messages_style
    ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}').style = messages_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 30

    # Agregar los encabezados de las nuevas columnas debajo de los mensajes
    ws.cell(row=3, column = columna_max, value="Total de Pagos").style = head_style
    ws.cell(row=4, column = columna_max, value="Sumatoria de Pagos").style = head_style
   

    # Asumiendo que las filas de datos comienzan en la fila 2 y terminan en row_num
    ws.cell(row=3, column=columna_max + 1, value=f"=COUNTA(A:A)-1").style = body_style
    ws.cell(row=4, column=columna_max + 1, value=f"=SUM(N:N)").style = money_resumen_style
  

   # Aqu√≠ debes extraer el conjunto completo de pagos en lugar de solo ciertos valores
    
    for pago in pagos:
        row_num = row_num + 1
        # Define los valores de las columnas bas√°ndote en el tipo de pago
        if pago.oc:
            articulos_compra = ArticuloComprado.objects.filter(oc=pago.oc)
            solicitado = pago.oc.req.orden.staff.staff.staff.first_name + ' ' + pago.oc.req.orden.staff.staff.staff.last_name if pago.oc else ''
            distrito = pago.oc.req.orden.distrito.nombre
            proveedor = pago.oc.proveedor
            #facturas_completas = pago.oc.facturas_completas
            cuenta_moneda = pago.cuenta.moneda.nombre if pago.cuenta else None
            proyectos = pago.oc.req.orden.proyecto.nombre if pago.oc else ''
            subproyectos = pago.oc.req.orden.subproyecto.nombre if pago.oc else ''
            autorizado = pago.oc.oc_autorizada_por2.staff.staff.first_name + ' ' + pago.oc.oc_autorizada_por2.staff.staff.last_name if pago.oc.oc_autorizada_por2 else ''
            fecha_creacion = pago.oc.created_at.replace(tzinfo=None)
            fecha_autorizacion = (
                pago.oc.autorizado_at_2.replace(tzinfo=None).strftime("%Y-%m-%d")
                if pago.oc.autorizado_at_2 else ''
            )
            productos = set()
            for articulo in articulos_compra:
                if articulo.producto:
                    productos.add(str(articulo.producto.producto.articulos.producto.producto.nombre))
            productos = ', '.join(productos)


            if pago.oc.facturas.filter(factura_xml__isnull=False).exists():
                tiene_facturas = 'S√≠'
            else:
                tiene_facturas = 'No'

            if cuenta_moneda == 'PESOS':
                tipo_de_cambio = ''
            elif cuenta_moneda == 'DOLARES':
                 tipo_de_cambio = pago.tipo_de_cambio or pago.oc.tipo_de_cambio or 17
            else:
                tipo_de_cambio = ''  # default si no se cumplen las condiciones anteriores
        elif pago.gasto:
            distrito = pago.gasto.distrito.nombre
            solicitado = pago.gasto.staff.staff.staff.first_name + ' ' + pago.gasto.staff.staff.staff.last_name
            fecha_creacion = pago.gasto.created_at.replace(tzinfo=None)

            fecha_autorizacion = (
                pago.gasto.approbado_fecha2.replace(tzinfo=None)
                if pago.gasto.approbado_fecha2 else ''
            )
            if pago.gasto.colaborador:
                proveedor = pago.gasto.colaborador.staff.staff.first_name + ' ' + pago.gasto.colaborador.staff.staff.last_name
            else:
                proveedor = pago.gasto.staff.staff.staff.first_name + ' ' + pago.gasto.staff.staff.staff.last_name
            
            articulos_gasto = Articulo_Gasto.objects.filter(gasto=pago.gasto)
            if pago.gasto.distrito.nombre == 'MATRIZ':
                autorizado = pago.gasto.superintendente.staff.staff.first_name + ' ' + pago.gasto.superintendente.staff.staff.last_name if pago.gasto.superintendente else ''
            else:
                autorizado = pago.gasto.autorizado_por2.staff.staff.first_name + ' ' + pago.gasto.autorizado_por2.staff.staff.last_name if pago.gasto.autorizado_por2 else ''
            proyectos = set()
            subproyectos = set()
            productos = set()

            for articulo in articulos_gasto:
                if articulo.proyecto:
                    proyectos.add(str(articulo.proyecto.nombre))
                if articulo.subproyecto:
                    subproyectos.add(str(articulo.subproyecto.nombre))
                if articulo.producto:
                    productos.add(str(articulo.producto.nombre))
            proyectos = ', '.join(proyectos)
            subproyectos = ', '.join(subproyectos)
            productos = ', '.join(productos)
            #facturas_completas = pago.gasto.facturas_completas
            tipo_de_cambio = '' # Asume que no se requiere tipo de cambio para gastos
            if pago.gasto.facturas.exists():
                
                tiene_facturas = 'S√≠'
            else:
                tiene_facturas = 'No'

           
            
        elif pago.viatico:
            articulos_viatico = Concepto_Viatico.objects.filter(viatico=pago.viatico)
            proyectos = pago.viatico.proyecto.nombre if pago.viatico else ''
            subproyectos = pago.viatico.subproyecto.nombre if pago.viatico else ''
            distrito = pago.viatico.distrito.nombre
            fecha_creacion = pago.viatico.created_at.replace(tzinfo=None)
            fecha_autorizacion = (
                pago.viatico.approved_at.replace(tzinfo=None).strftime("%Y-%m-%d")
                if pago.viatico and pago.viatico.approved_at
                else ''
            )
            autorizado = pago.viatico.gerente.staff.staff.first_name + ' ' + pago.viatico.gerente.staff.staff.last_name if pago.viatico.gerente else ''
            if pago.viatico.colaborador:
                proveedor = pago.viatico.colaborador.staff.staff.first_name + ' ' + pago.viatico.colaborador.staff.staff.last_name
            else:
                proveedor = pago.viatico.staff.staff.staff.first_name + ' ' + pago.viatico.staff.staff.staff.last_name
            solicitado = pago.viatico.staff.staff.staff.first_name + ' ' + pago.viatico.staff.staff.staff.last_name
            #facturas_completas = pago.viatico.facturas_completas
            tipo_de_cambio = '' # Asume que no se requiere tipo de cambio para vi√°ticos
            
            if pago.viatico.facturas.exists():
                tiene_facturas = 'S√≠'
            else:
                tiene_facturas = 'No'
            productos = set()
            for articulo in articulos_viatico:
                if articulo.producto:
                    productos.add(str(articulo.producto.nombre))
            productos = ', '.join(productos)
        else:
            proveedor = None
            #facturas_completas = None
            tipo_de_cambio = ''


       

        row = [
            distrito,
            get_transaction_id(pago),
            solicitado,
            autorizado,
            fecha_creacion,
            fecha_autorizacion,
            proyectos,
            subproyectos,
            proveedor,
            productos,
            pago.monto,
            pago.oc.moneda.nombre if pago.oc else 'PESOS',  # Modificaci√≥n aqu√≠
            tipo_de_cambio,
            f'=IF(M{row_num}="",K{row_num},K{row_num}*M{row_num})',  # Calcula total en pesos usando la f√≥rmula de Excel
            pago.pagado_date.replace(tzinfo=None),
            tiene_facturas,
        ]

    
        for col_num in range(len(row)):
            (ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style
            if col_num in (4, 5, 14):
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style
            if col_num in (10, 12, 13):
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = money_style
       
    
    sheet = wb['Sheet']
    wb.remove(sheet)
    response.set_cookie('descarga_iniciada', 'true', max_age=20)  # La cookie expira en 20 segundos
    wb.save(response)

    return(response)

def mass_payment_view(request):
    if request.method == 'POST':
        request.session['compras_ids'] = request.POST.getlist('compra_id')
        return redirect('layout_pagos')  # No pasamos 'ids' porque usaremos la sesi√≥n

# Si necesitas pasar las IDs como parte del contexto a un nuevo template puedes hacerlo as√≠:
def layout_pagos(request):
    compras_ids = request.session.get('compras_ids', [])
    compras_ids = [int(id) for id in compras_ids if str(id).isdigit()]
    compras = Compra.objects.filter(id__in=compras_ids)
    cuentas_disponibles = Cuenta.objects.all()

    if request.method == 'POST':
        try:
            root = ET.Element('Document', {
                'xmlns:xsi': 'http://www.w3.org/2001/XMLSchema-instance',
                'xmlns': 'urn:iso:std:iso:20022:tech:xsd:pain.001.001.03'
            })

            cstmr_cdt_trf_initn = ET.SubElement(root, 'CstmrCdtTrfInitn')
            grp_hdr = ET.SubElement(cstmr_cdt_trf_initn, 'GrpHdr')
            now_local = datetime.now()
            ET.SubElement(grp_hdr, 'MsgId').text = 'AUTO' +  now_local.strftime('%Y%m%d%H%M%S')
            ET.SubElement(grp_hdr, 'CreDtTm').text = now_local.isoformat()
            ET.SubElement(grp_hdr, 'NbOfTxs').text = str(len(compras))

            initg_pty = ET.SubElement(grp_hdr, 'InitgPty')
            id_ = ET.SubElement(initg_pty, 'Id')
            org_id = ET.SubElement(id_, 'OrgId')
            othr = ET.SubElement(org_id, 'Othr')
            ET.SubElement(othr, 'Id').text = 'VORDCA00H2H'
            
            for count, compra in enumerate(compras, start=1):
                cuenta_pago_id = request.POST.get(f'cuenta_{count}')
                cuenta_pago = cuentas_disponibles.get(id=cuenta_pago_id)

                
            
                monto = float(request.POST.get(f'monto_{count}', '0'))
                

                pmt_inf = ET.SubElement(cstmr_cdt_trf_initn, 'PmtInf')
                ET.SubElement(pmt_inf, 'PmtInfId').text = f'Pmt-{compra.id}'
                ET.SubElement(pmt_inf, 'PmtMtd').text = 'TRF'

                pmt_tp_inf = ET.SubElement(pmt_inf, 'PmtTpInf')
                svc_lvl = ET.SubElement(pmt_tp_inf, 'SvcLvl')
                ET.SubElement(svc_lvl, 'Cd').text = 'URGP'

                ET.SubElement(pmt_inf, 'ReqdExctnDt').text = now_local.strftime('%Y-%m-%d')

                dbtr = ET.SubElement(pmt_inf, 'Dbtr')
                dbtr_id = ET.SubElement(dbtr, 'Id')
                dbtr_org_id = ET.SubElement(dbtr_id, 'OrgId')
                dbtr_othr = ET.SubElement(dbtr_org_id, 'Othr')
                ET.SubElement(dbtr_othr, 'Id').text = '123456789'

                dbtr_acct = ET.SubElement(pmt_inf, 'DbtrAcct')
                dbtr_acct_id = ET.SubElement(dbtr_acct, 'Id')
                dbtr_acct_othr = ET.SubElement(dbtr_acct_id, 'Othr')
                ET.SubElement(dbtr_acct_othr, 'Id').text = str(cuenta_pago.cuenta)
                
                #ET.SubElement(dbtr_acct, 'Ccy').text = compra.moneda.nombre

                dbtr_agt = ET.SubElement(pmt_inf, 'DbtrAgt')
                fin_instn_id = ET.SubElement(dbtr_agt, 'FinInstnId')
                ET.SubElement(fin_instn_id, 'BIC').text = 'BCMRMXMM'

                cdt_trf_tx_inf = ET.SubElement(pmt_inf, 'CdtTrfTxInf')

                pmt_id = ET.SubElement(cdt_trf_tx_inf, 'PmtId')
                instr_id = f'INST-{compra.id}'
                ET.SubElement(pmt_id, 'InstrId').text = instr_id
                ET.SubElement(pmt_id, 'EndToEndId').text = instr_id

                amt = ET.SubElement(cdt_trf_tx_inf, 'Amt')
                if compra.moneda.nombre == "PESOS":
                    moneda = "MXN"
                ET.SubElement(amt, 'InstdAmt', Ccy=moneda).text = f"{monto:.2f}"

                ET.SubElement(cdt_trf_tx_inf, 'ChrgBr').text = 'DEBT'

                cdtr_agt = ET.SubElement(cdt_trf_tx_inf, 'CdtrAgt')
                fin_instn_id_cdtr = ET.SubElement(cdtr_agt, 'FinInstnId')
                bic_banco_receptor = compra.proveedor.banco.bic if compra.proveedor.banco.bic else 'BICDESCONOCIDO'
                ET.SubElement(fin_instn_id_cdtr, 'BIC').text = bic_banco_receptor

                cdtr = ET.SubElement(cdt_trf_tx_inf, 'Cdtr')
                ET.SubElement(cdtr, 'Nm').text = compra.proveedor.nombre.razon_social

                PstlAdr = ET.SubElement(cdtr,'PstlAdr')
                ET.SubElement(PstlAdr, 'StrtNm').text = compra.proveedor.domicilio
                if compra.proveedor.estado:
                    ET.SubElement(PstlAdr, 'TwnNm').text = compra.proveedor.estado.nombre
                if compra.proveedor.nombre.extranjero == False:
                    ET.SubElement(PstlAdr, 'Ctry').text = 'MX'
                else:
                    ET.SubElement(PstlAdr, 'Ctry').text = 'EX'

                cdtr_id = ET.SubElement(cdtr, 'Id')
                cdtr_org_id = ET.SubElement(cdtr_id, 'OrgId')
                cdtr_othr = ET.SubElement(cdtr_org_id, 'Othr')
                ET.SubElement(cdtr_othr, 'Id').text = compra.proveedor.nombre.rfc


                cdtr_acct = ET.SubElement(cdt_trf_tx_inf, 'CdtrAcct')
                cdtr_acct_id = ET.SubElement(cdtr_acct, 'Id')
                cdtr_acct_othr = ET.SubElement(cdtr_acct_id, 'Othr')
                ET.SubElement(cdtr_acct_othr, 'Id').text = str(compra.proveedor.cuenta)

                rmt_inf = ET.SubElement(cdt_trf_tx_inf, 'RmtInf')
                ET.SubElement(rmt_inf, 'Ustrd').text = f"F-{compra.folio}"

            xml_bytes = ET.tostring(root, encoding='utf-8', method='xml')
            #logging.info("XML generado (primeros 500 caracteres):")
            #logging.info(xml_bytes.decode()[:500])
            # Generar secuencial √∫nico (persistente en un archivo)
            secuencial_file = '/home/savia/pagos_xml/secuencial.txt'

            # Leer el √∫ltimo secuencial
            if not os.path.exists(secuencial_file):
                ultimo_secuencial = 0
            else:
                with open(secuencial_file, 'r') as f:
                    ultimo_secuencial = int(f.read().strip())

            # Incrementar
            nuevo_secuencial = (ultimo_secuencial + 1) % 1000  # M√°ximo 3 d√≠gitos: 000-999

            # Guardar el nuevo secuencial para la pr√≥xima vez
            with open(secuencial_file, 'w') as f:
                f.write(str(nuevo_secuencial))

            # Formatear a 3 d√≠gitos
            secuencia = '{:03d}'.format(nuevo_secuencial)
            bei = 'VORDCA00H2H'
            country = 'MX'
            #print(country)
            fecha_actual = datetime.now().strftime('%Y%m%d')
            #print(fecha_actual)
            extension = 'CAN'
            nombre_base = f'{bei}_{country}_{fecha_actual}{secuencia}'
            nombre_final = f'{nombre_base}.{extension}' 
            #print(nombre_final)
            # Guardar XML en disco
            xml_path = '/home/savia/pagos_xml/temporal.xml'
            with open(xml_path, 'wb') as f:
                f.write(xml_bytes)
            logging.info(f'Archivo XML generado: {xml_path}')

            # Encriptar el archivo XML con GPG
            encrypted_path = f'/home/savia/pagos_encrypted/{nombre_final}'
            #print(encrypted_path)
            
            subprocess.run([
                '/usr/bin/gpg', '--yes', '--batch', '--trust-model', 'always',
                '--output', encrypted_path,
                '--encrypt', '--sign', 
                '--recipient', 'gruvor1i', 
                '--local-user', 'A5B3FE060FE7283919E6B10732C9AA4231DB66B8',  # clave privada GPG
                xml_path
            ], check=True)
            logging.info(f'Archivo encriptado: {encrypted_path}')
            # 5. Enviar por SFTP

            host = os.getenv("BBVA_SFTP_HOST")
            port = int(os.getenv("BBVA_PORT"))
            username = os.getenv("BBVA_UP")
            password = os.getenv("BBVA_PP")
            remote_path = '/'

            transport = paramiko.Transport((host, port))
            transport.connect(username=username, password=password)
            sftp = paramiko.SFTPClient.from_transport(transport)
            sftp.chdir(remote_path)
            sftp.put(encrypted_path, 'pagos.xml.gpg')
            sftp.close()
            transport.close()
            logging.info(f'Archivo enviado a BBVA SFTP ({host}:{port}{remote_path})')

            messages.success(request, 'Archivo encriptado y enviado por SFTP a BBVA correctamente.')
            return redirect('compras-autorizadas')  # Cambiar por el nombre real de tu vista
        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()

            # Mostrar en pantalla
            messages.error(request, f"‚ùå Ocurri√≥ un error al ejecutar GPG.")
            messages.error(request, f"üîß STDOUT: {e.stdout}")
            messages.error(request, f"üí• STDERR: {e.stderr}")
            logger = logging.getLogger('pagos_sftp')
            logger.error("Error al firmar y cifrar el XML con GPG", exc_info=True)
            return redirect('compras-autorizadas')


    context = {
        'compras': compras,
        'cuentas_disponibles': cuentas_disponibles,
    }

    return render(request, 'tesoreria/layout_pagos.html', context)


def descargar_respuestas_bbva():
    host = os.getenv("BBVA_SFTP_HOST")
    port = int(os.getenv("BBVA_PORT"))
    username = os.getenv("BBVA_UG")   
    password = os.getenv("BBVA_PG")     
    remote_path = '/'                 
    local_path = '/home/savia/pagos_respuestas/'

    os.makedirs(local_path, exist_ok=True)

    try:
        transport = paramiko.Transport((host, port))
        transport.connect(username=username, password=password)
        sftp = paramiko.SFTPClient.from_transport(transport)
        sftp.chdir(remote_path)

        archivos = sftp.listdir()
        logging.info(f'Se encontraron {len(archivos)} archivo(s) en {remote_path}.')

        logging.info("üìÇ Listado completo del directorio remoto:")
        for entry in sftp.listdir_attr(remote_path):
            tipo = 'üìÅ Carpeta' if str(entry.longname).startswith('d') else 'üìÑ Archivo'
            logging.info(f"{tipo}: {entry.filename}")

        for archivo in archivos:
            remote_file = f"{remote_path}{archivo}" if not remote_path.endswith('/') else f"{remote_path}{archivo}"
            local_file = os.path.join(local_path, archivo)

            sftp.get(remote_file, local_file)
            logging.info(f'Archivo encontrado en SFTP: {archivo}')
            logging.info(f'Archivo descargado: {archivo} ‚Üí {local_file}')

            # Si es archivo .pgp, desencriptarlo
            if archivo.endswith('.pgp') or archivo.endswith('.gpg'):
                desencriptar_pgp(local_file)

        sftp.close()
        transport.close()
        logging.info('Conexi√≥n cerrada correctamente despu√©s de descargar archivos.')

    except Exception as e:
        logging.error(f'Error al descargar archivos desde BBVA: {str(e)}')

def es_respuesta_bbva(path_xml):
    try:
        tree = ET.parse(path_xml)
        root = tree.getroot()
        claves = ['respuesta', 'ack', 'estatus', 'resultado', 'codigo']
        return any(clave in root.tag.lower() for clave in claves)
    except Exception as e:
        logging.warning(f"No se pudo analizar {path_xml} como XML: {str(e)}")
        return False

def desencriptar_pgp(archivo_pgp):
    global desencriptados, errores
    archivo_xml = archivo_pgp.replace('.pgp', '.xml').replace('.gpg', '.xml')
    if os.path.exists(archivo_xml):
        logging.info(f"Archivo ya desencriptado: {archivo_xml}")
        return
    try:
        logging.info(f"Desencriptando {archivo_pgp}...")
        result = subprocess.run(
            ['gpg', '--batch', '--yes', '--output', archivo_xml, '--decrypt', archivo_pgp],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            check=True
        )
        logging.info(f"Desencriptado exitosamente: {archivo_xml}")
        desencriptados += 1
        if es_respuesta_bbva(archivo_xml):
            logging.info(f"‚úÖ Archivo identificado como respuesta BBVA: {archivo_xml}")
        else:
            logging.info(f"‚ÑπÔ∏è Archivo desencriptado pero no parece respuesta BBVA: {archivo_xml}")
    except subprocess.CalledProcessError as e:
        logging.error(f"Error desencriptando {archivo_pgp}: {e.stderr.decode()}")
        errores += 1
    
def update_comentario_control(request):
    data= json.loads(request.body)
    pk = data["pago_id"]
    dato = data["dato"]
    tipo = data["tipo"]
    pago = Pago.objects.get(id=pk)
    if tipo == "comentario": 
        pago.comentario = dato
    if tipo == "hora":
        pago.pagado_hora = dato
    pago.save()
    # Construye un objeto de respuesta que incluya el dato y el tipo.
    response_data = {
        'dato': dato,
        'tipo': tipo
    }

    return JsonResponse(response_data, safe=False)




def escanear_sftp_recursivo(sftp, remote_path, local_path):
    global descargados, errores
    os.makedirs(local_path, exist_ok=True)
    logging.info(f"üìÅ Visitando carpeta: {remote_path}")

    try:
        for item in sftp.listdir_attr(remote_path):
            remote_item_path = os.path.join(remote_path, item.filename)
            local_item_path = os.path.join(local_path, item.filename)

            if str(item.longname).startswith('d'):
                escanear_sftp_recursivo(sftp, remote_item_path, local_item_path)
            else:
                logging.info(f"üìÑ Archivo encontrado: {remote_item_path}")
                try:
                    sftp.get(remote_item_path, local_item_path)
                    logging.info(f"Archivo descargado: {remote_item_path} ‚Üí {local_item_path}")
                    descargados += 1
                    if item.filename.endswith(('.pgp', '.gpg')):
                        desencriptar_pgp(local_item_path)
                except Exception as e:
                    logging.error(f"Error descargando {remote_item_path}: {str(e)}")
                    errores += 1
    except IOError as e:
        logging.warning(f"No se pudo acceder a {remote_path}: {str(e)}")
        errores += 1


def escanear_todo_bbva():
    global descargados, desencriptados, errores
    descargados = 0
    desencriptados = 0
    errores = 0

    host = os.getenv("BBVA_SFTP_HOST")
    port = int(os.getenv("BBVA_PORT"))
    username = os.getenv("BBVA_UG")   
    password = os.getenv("BBVA_PG")     
    remote_root = '/'                 
    local_root = '/home/savia/pagos_respuestas/'

    try:
        transport = paramiko.Transport((host, port))
        transport.connect(username=username, password=password)
        sftp = paramiko.SFTPClient.from_transport(transport)
        logging.info(f"üîç Iniciando escaneo completo en SFTP desde: {remote_root}")

        escanear_sftp_recursivo(sftp, remote_root, local_root)

        sftp.close()
        transport.close()
        logging.info("‚úÖ Escaneo completo finalizado correctamente.")
        logging.info(f"üì¶ Archivos descargados: {descargados}")
        logging.info(f"üîê Archivos desencriptados: {desencriptados}")
        logging.info(f"‚ùå Errores durante el proceso: {errores}")

    except Exception as e:
        logging.error(f"‚ùå Error en escaneo de SFTP: {str(e)}")
        
def convert_excel_control_bancos(cuenta_id, pagos, saldo_inicial_objeto,  start_date_str=None):
    # Paso 1: determinar la fecha de inicio real
    cuenta =  Cuenta.objects.get(id = cuenta_id)
    start_date = None

    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass  # fecha inv√°lida o malformada

    if not start_date and saldo_inicial_objeto:
        start_date = saldo_inicial_objeto.fecha_inicial

    if not start_date and pagos.exists():
        start_date = pagos.order_by('pagado_real').first().pagado_real

    # Paso 2: calcular el saldo inicial ajustado
    if saldo_inicial_objeto is None:
        saldo_inicial = 0
        fecha_saldo_inicial = "No definido"
    else:
        saldo_inicial = saldo_inicial_objeto.monto_inicial
        fecha_saldo_inicial = saldo_inicial_objeto.fecha_inicial

        if start_date and start_date > saldo_inicial_objeto.fecha_inicial:
            pagos_intermedios = Pago.objects.filter(
                cuenta=cuenta,
                hecho=True,
                pagado_real__gte=saldo_inicial_objeto.fecha_inicial,
                pagado_real__lt=start_date
            )

            total_intermedios = sum(
                p.monto if p.tipo is not None and p.tipo.nombre == "ABONO"
                else -p.monto
                for p in pagos_intermedios
            )
            print(total_intermedios)
            saldo_inicial += total_intermedios
            fecha_saldo_inicial = start_date
    print(start_date)
    #print(saldo_inicial_objeto.fecha_inicial)
    pagos = pagos.order_by('pagado_real', 'pagado_hora')
    static_path = settings.STATIC_ROOT
    img_path2 = os.path.join(static_path, 'images', 'logo_vordcab.jpg')
    # Crea un objeto BytesIO para guardar el archivo Excel
    output = BytesIO()

    # Crea un libro de trabajo y a√±ade una hoja
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet("Matriz_Compras")
     # Ajustar la altura de las filas 1 y 2
   
     
    date_format = workbook.add_format({'num_format': 'dd/mm/yyyy'})
    # Define los estilos
    head_style = workbook.add_format({'bold': True, 'font_color': 'FFFFFF', 'bg_color': '#16324F', 'font_name': 'Arial', 'font_size': 11})
    body_style = workbook.add_format({'font_name': 'Calibri', 'font_size': 10})
    money_style = workbook.add_format({'num_format': '$ #,##0.00', 'font_name': 'Calibri', 'font_size': 10})
    date_style = workbook.add_format({'num_format': 'dd/mm/yyyy', 'font_name': 'Calibri', 'font_size': 10})
    percent_style = workbook.add_format({'num_format': '0.00%', 'font_name': 'Calibri', 'font_size': 10})
    messages_style = workbook.add_format({'font_name':'Arial Narrow', 'font_size':11})
    header_format = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
    cell_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1})
    d_cell_format = workbook.add_format({'num_format': 'dd/mm/yyyy','align': 'center', 'valign': 'vcenter', 'border': 1})
    title_format = workbook.add_format({'font_name': 'Calibri', 'font_size': 18, 'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
    vordcab_format = workbook.add_format({'font_color': 'FFFFFF', 'bg_color': '#16324F','font_name': 'Calibri', 'font_size': 18, 'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
    h_money_style = workbook.add_format({'num_format': '$ #,##0.00', 'font_name': 'Calibri', 'font_size': 10, 'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
    # Ajustar el ancho de las columnas
    worksheet.set_column('A:A', 20)  # Fecha
    #worksheet.set_column('B:B', 20)  # Empresa
    worksheet.set_column('B:B', 35)  # Empresa/Proveedor
    worksheet.set_column('C:C', 25)  # Cuenta
    worksheet.set_column('D:D', 30)  # Concepto/Servicio
    worksheet.set_column('E:E', 25)  # Contrato
    worksheet.set_column('F:F', 25)  # Sector
    worksheet.set_column('G:G', 20)  # Distrito
    worksheet.set_column('H:H', 20)  # Monto
    worksheet.set_column('I:I', 20)  # Saldo
    worksheet.set_column('J:J', 20)  # Saldo
    # worksheet.set_column('K:K', 20)  # Saldo

    worksheet.set_row(0, 40)  # Fila 1 (√≠ndice 0) con altura 40
    worksheet.set_row(1, 30)  # Fila 2 (√≠ndice 1) con altura 30

    # Insertar el logo en la hoja de trabajo
    worksheet.insert_image('A1', img_path2, {'x_scale': 1, 'y_scale': 1})

    # Agregar y fusionar celdas para el encabezado
    worksheet.write('I1', 'Preparado Por:', header_format)
    worksheet.write('I2', 'SUBD FIN', cell_format)
    worksheet.write('J1', 'Aprobaci√≥n', header_format)
    worksheet.write('J2', 'DG', cell_format)

    worksheet.merge_range('C1:H2', 'CONTROL DE BANCOS', title_format)
    worksheet.merge_range('A3:B3', 'N√∫mero de documento', header_format)
    worksheet.merge_range('A4:B4', 'SEOV-TES-N4-01.03', cell_format)
    
    worksheet.merge_range('C3:D3', 'Clasificaci√≥n del documento', header_format)
    worksheet.merge_range('C4:D4', 'Controlado', cell_format)
    worksheet.merge_range('E3:F3', 'Nivel del documento', header_format)
    worksheet.merge_range('E4:F4', 'N5', cell_format)
    
    worksheet.write('G3', 'Revisi√≥n No.', header_format)
    worksheet.write('G4', '000', cell_format)
    worksheet.write('H3', 'Fecha de emisi√≥n', header_format)
    worksheet.write('H4', '12/09/2022', d_cell_format)
    worksheet.merge_range('I3:J3', 'Fecha Revisi√≥n', header_format)
    worksheet.merge_range('I4:J4', '', cell_format)
    #worksheet.write('I3', 'Fecha de emisi√≥n', header_format)
    
    worksheet.merge_range('A5:J8', 'GRUPO VORDCAB, S.A. DE C.V.', vordcab_format)
  
    if not pagos.exists():
        worksheet.merge_range('A9:B9', 'INSTITUCI√ìN BANCARIA: NO HAY PAGOS DISPONIBLES', header_format)
    else:
        worksheet.merge_range('A9:B9', 'INSTITUCI√ìN BANCARIA: '+ str(cuenta.banco.nombre), header_format)
   
    
    worksheet.merge_range('A10:B10', 'CUENTA BANCARIA: '+ str(cuenta.cuenta), header_format)
    worksheet.merge_range('A11:B11', 'DISTRITO: ' + str(cuenta.encargado.distritos), header_format)
    worksheet.merge_range('A12:B12', 'RESPONSABLE DE CUENTA: ' + str(cuenta.encargado.staff.staff.first_name)+ ' '+ str(cuenta.encargado.staff.staff.last_name), header_format)

    worksheet.write('H9', 'PERIODO:', header_format)
    #worksheet.write('I9', 'MES', cell_format)
    #worksheet.write('J9', 'A√ëO', cell_format)
   
    
    worksheet.write('I10', 'SALDO INICIAL' , header_format)
    worksheet.write('G10', 'Fecha Saldo Inicial')
    worksheet.write('H10', fecha_saldo_inicial, date_style)
    worksheet.write('J10', saldo_inicial, h_money_style)
    worksheet.write('I11', 'SALDO FINAL', header_format)

    columns = ['Fecha','Empresa/Colaborador','Folio','Concepto/Servicio','Proyecto','Subproyecto','Distrito','Cargo','Abono','Saldo']

    columna_max = len(columns)+2

    #worksheet.write(0, columna_max - 1, 'Reporte Creado Autom√°ticamente por SAVIA Vordcab. UH', messages_style)
    #worksheet.write(1, columna_max - 1, 'Software desarrollado por Grupo Vordcab S.A. de C.V.', messages_style)
    worksheet.set_column(columna_max - 1, columna_max, 30)  # Ajusta el ancho de las columnas nuevas
    
    for col_num, header in enumerate(columns):
        worksheet.write(12, col_num, header, head_style)

    folios_unicos = set()  # Mantener un conjunto de folios √∫nicos

    row_num = 13
    for pago in pagos:
        proveedor = pago.detalles_comprobante.get('titular_cuenta_2','No disponible') 
        print('proveedor:',proveedor)
        fecha = pago.pagado_real
        empresa = pago.cuenta.empresa.nombre
        if proveedor == "No disponible":
            if hasattr(pago, 'oc') and pago.oc :
                proveedor = pago.oc.proveedor.nombre.razon_social
            elif hasattr(pago, 'gasto') and pago.gasto:
                if pago.gasto.colaborador:
                    proveedor = f"{pago.gasto.colaborador.staff.staff.first_name} {pago.gasto.colaborador.staff.staff.last_name}"
                else:
                    proveedor = f"{pago.gasto.staff.staff.staff.first_name} {pago.gasto.staff.staff.staff.last_name}"
            elif hasattr(pago, 'viatico') and pago.viatico:
                if pago.viatico.colaborador:
                    proveedor = f"{pago.viatico.colaborador.staff.staff.first_name} {pago.viatico.colaborador.staff.staff.last_name}"
                else:
                    proveedor = f"{pago.viatico.staff.staff.staff.first_name} {pago.viatico.staff.staff.staff.last_name}"
            elif pago.empresa_beneficiario:
                proveedor = pago.empresa_beneficiario
        #else:
        #    if pago.tesorero:
        #        proveedor = pago.detalles_comprobante.get('titular_cuenta_2', 'No disponible') 
        #    else:
        #        proveedor = "No disponible"

        cuenta = pago.cuenta.cuenta
        concepto_servicio = pago.detalles_comprobante.get('motivo_pago')
        #if motivo and motivo != 'No disponible':
            # usar motivo
        #    print(motivo)
        #else:
        #    print("No hay motivo de pago disponible")
      
        
        if hasattr(pago, 'oc') and pago.oc:
            folio = f"OC{pago.oc.folio}"
            contrato = pago.oc.req.orden.proyecto.nombre
            sector = pago.oc.req.orden.subproyecto.nombre
        elif hasattr(pago, 'gasto') and pago.gasto:
            folio = f"G{pago.gasto.folio}"
            articulos_gasto = Articulo_Gasto.objects.filter(gasto=pago.gasto)
            proyectos = set()
            subproyectos = set()
            for articulo in articulos_gasto:
                if articulo.proyecto:
                    proyectos.add(str(articulo.proyecto.nombre))
                if articulo.subproyecto:
                    subproyectos.add(str(articulo.subproyecto.nombre))
            contrato = ', '.join(proyectos)
            sector = ', '.join(subproyectos)
       
            
        elif hasattr(pago, 'viatico') and pago.viatico:
            folio = f"V{pago.viatico.folio}"
            contrato = pago.viatico.proyecto.nombre
            sector = pago.viatico.subproyecto.nombre
        else:
            concepto_servicio = str(pago.tipo)
            folio = f'NA - {pago.tipo.nombre}'
            contrato = ''
            sector = ''
       

        # Determinar contrato y sector
        if pago.comentario:
            comentarios = pago.comentario
        elif hasattr(pago, 'oc') and pago.oc:
            comentarios = (pago.oc.req.orden.comentario or '').upper()
        elif hasattr(pago, 'viatico') and pago.viatico:            
            comentarios = (pago.viatico.comentario_general or '').upper()
        elif hasattr(pago, 'gasto') and pago.gasto:
            
            if pago.gasto.comentario:
                comentarios = pago.gasto.comentario.strip().upper()
            else:
                comentarios_articulos = [
                    (a.comentario or '').strip().upper()
                    for a in articulos_gasto if a.comentario
                ]
                comentarios = ', '.join(comentarios_articulos) if comentarios_articulos else 'NO HAY COMENTARIOS DISPONIBLES'
        else:
            comentarios = 'NO HAY COMENTARIOS DISPONIBLES'
            
                    
        distrito = pago.oc.req.orden.distrito.nombre if hasattr(pago, 'oc') and pago.oc else (pago.gasto.distrito.nombre if hasattr(pago, 'gasto') and pago.gasto else (pago.viatico.subproyecto.nombre if hasattr(pago, 'viatico') and pago.viatico else (pago.distrito.nombre if pago.distrito else '')))
        cargo = ''
        abono = ''
        if pago.tipo == None or pago.tipo.nombre == "CARGO" or pago.tipo.nombre == "TRANSFERENCIA":
            cargo = pago.monto
        elif pago.tipo.nombre == "ABONO":
            abono = pago.monto
        #saldo = pago.saldo
        
   

        # Escribir los datos en el archivo Excel
        worksheet.write(row_num, 0, fecha.strftime('%d/%m/%Y') if fecha else '', date_style)
        worksheet.write(row_num, 1, empresa)
        worksheet.write(row_num, 1, proveedor)
        worksheet.write(row_num, 2, folio)
        worksheet.write(row_num, 3, comentarios)

        worksheet.write(row_num, 4, contrato)
        worksheet.write(row_num, 5, sector)
        worksheet.write(row_num, 6, distrito)
        worksheet.write(row_num, 7, cargo, money_style)
        worksheet.write(row_num, 8, abono, money_style)
        #worksheet.write(row_num, 9, comentarios)
        # Saldo en la columna 9 (√≠ndice 9 = columna J)
        if row_num <= 12:
            # Primera fila de saldo, usa saldo inicial
            # Ya est√° escrito en J13
            pass
        elif row_num == 13:
            # Fila 14: saldo inicial en J10 - cargo actual + abono actual
            fila_actual_excel = row_num + 1  # Excel indexa desde 1
            celda_saldo_inicial = 'J10'
            celda_cargo_actual = f'H{fila_actual_excel}'
            celda_abono_actual = f'I{fila_actual_excel}'

            formula_saldo = f'={celda_saldo_inicial} - {celda_cargo_actual} + {celda_abono_actual}'
            worksheet.write_formula(row_num, 9, formula_saldo, money_style)

        else:
            # Desde la fila 14 en adelante, calcula el saldo din√°mico
            fila_actual_excel = row_num + 1  # Excel indexa desde 1
            fila_anterior_excel = fila_actual_excel - 1

            # Celdas relevantes
            celda_saldo_anterior = f'J{fila_anterior_excel}'
            celda_cargo_actual = f'H{fila_actual_excel}'
            celda_abono_actual = f'I{fila_actual_excel}'

            # F√≥rmula: saldo anterior - cargo + abono
            formula_saldo = f'={celda_saldo_anterior} - {celda_cargo_actual} + {celda_abono_actual}'
            worksheet.write_formula(row_num, 9, formula_saldo, money_style)

        
        row_num += 1

    last_filled_row = row_num
    worksheet.write_formula('J11', f'J{last_filled_row}', h_money_style)
    #worksheet.write_formula('I9', f'={last_filled_cell}', h_money_style)
    worksheet.write_formula('I9', 'A14', date_style)
    worksheet.write_formula('J9', f'A{last_filled_row}', date_style)
     # Agregar el marco general desde A1 hasta J12
    border_format = workbook.add_format({
        'top': 1,
        'bottom': 1,
        'left': 1,
        'right': 1
    })
    

    # Aplicar el borde derecho
    #for row in range(12):
        #worksheet.write(row, 9, '', border_format)
   
    workbook.close()

    # Construye la respuesta
    output.seek(0)

    response = HttpResponse(
        output.read(), 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    response['Content-Disposition'] = f'attachment; filename=Control_Bancos_{dt.date.today()}.xlsx'
      # Establecer una cookie para indicar que la descarga ha iniciado
    response.set_cookie('descarga_iniciada', 'true', max_age=20)  # La cookie expira en 20 segundos
    output.close()
    return response


def cfdi_compras(request, pk):
    factura = Facturas.objects.get(id=pk)
    data = factura.emisor
    # Verificar y asignar un valor predeterminado para impuestos si es None
    if data['impuestos'] is None:
        data['impuestos'] = 0.0

    # Verificar y asignar un valor predeterminado para total si es None
    if data['total'] is None:
        data['total'] = 0.0

    # Verificar y asignar un valor predeterminado para subtotal si es None
    if data['subtotal'] is None:
        data['subtotal'] = 0.0
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    if not data:
        return HttpResponse("Error al parsear el archivo XML", status=400)

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    
    # Generar c√≥digo QR
    qr_data = f"https://verificacfdi.facturaelectronica.sat.gob.mx/default.aspx?id={data['uuid']}&re={data['rfc_emisor']}&rr={data['rfc_receptor']}&tt={data['total']}&fe={data['sello_cfd'][-8:]}"
    qr_img = qrcode.make(qr_data)
    
    with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as temp_file:
        qr_img.save(temp_file)
        temp_file.seek(0)
        qr_x = 500
        qr_y = height - 700
        qr_size = 2.75 * cm
        c.drawImage(temp_file.name, qr_x, qr_y, qr_size, qr_size)

    # T√≠tulo
    c.setFillColor(prussian_blue)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(30, height - 40, "FACTURA GENERADA POR SAVIA 2.0")

    # Datos del Emisor
    c.setFillColor(black)
    c.setFont("Helvetica-Bold", 12)
    c.drawString(30, height - 80, "Datos del Emisor:")
    
    c.setFont("Helvetica", 8)
    alineado_x = 30
    alineado_y = height - 100
    alineado_y2 = alineado_y
    line_height = 12

    c.drawString(alineado_x, alineado_y, f"RFC: {data['rfc_emisor']}")
    alineado_y -= line_height
    c.drawString(alineado_x, alineado_y, f"Nombre: {data['nombre_emisor']}")
    alineado_y -= line_height
    c.drawString(alineado_x, alineado_y, f"R√©gimen Fiscal: {data['regimen_fiscal_emisor']}")
    alineado_y -= line_height
    c.drawString(alineado_x, alineado_y, f"Lugar de Expedici√≥n: {data['lugar_expedicion']}")
    alineado_y -= line_height
    c.drawString(alineado_x, alineado_y, f"Fecha y hora de expedici√≥n: {data['fecha']}")
    alineado_y -= line_height
    c.drawString(alineado_x, alineado_y, f"Moneda: {data['moneda']}")
    alineado_y -= line_height
    c.drawString(alineado_x, alineado_y, f"Forma de Pago: {data['forma_pago']}")

    # Datos del Receptor
    alineado_y -= 2 * line_height
    c.setFont("Helvetica-Bold", 12)
    c.drawString(alineado_x + 350, height - 80, "Datos del Receptor:")
    
    c.setFont("Helvetica", 8)
    alineado_y -= line_height
    c.drawString(alineado_x + 350, alineado_y2, f"RFC: {data['rfc_receptor']}")
    alineado_y2 -= line_height
    c.drawString(alineado_x + 350, alineado_y2, f"Nombre: {data['nombre_receptor']}")
    alineado_y2 -= line_height
    c.drawString(alineado_x + 350, alineado_y2, f"R√©gimen Fiscal: {data['regimen_fiscal_receptor']}")
    alineado_y2 -= line_height
    c.drawString(alineado_x + 350, alineado_y2, f"R√©gimen Fiscal: {data['codigo_postal']}")
    alineado_y2 -= line_height
    c.drawString(alineado_x + 350, alineado_y2, f"Uso del CFDI: {data['uso_cfdi']}")

    # Conceptos (Tabla)
    alineado_y -= line_height
    # Configuraci√≥n del estilo para los p√°rrafos
    styles = getSampleStyleSheet()
    styleN = styles['Normal']
    styleN.wordWrap = 'CJK'  # Ajusta autom√°ticamente el texto
    # Crear un estilo personalizado
    custom_style = ParagraphStyle(
        'CustomStyle',
        parent=styleN,
        fontSize=6,  # Ajusta el tama√±o del texto aqu√≠
        leading=7,   # Ajusta el interlineado aqu√≠ si es necesario
    )

    # Preparamos los datos de la tabla
    table_data = [["CANT", "CLAVE", "CONCEPTO", "U DE M", "P.U.", "IMPORTE", "IMPUESTO", "TIPO TASA"]]
    for item in data['resultados']:
        descripcion = item['descripcion']
        cantidad = float(item['cantidad'])
        unidad = item['unidad']
        valor_unitario = float(item['precio'])
        importe = float(item['importe'])
        # Verificar y convertir solo si el valor no es 'N/A'
         # Inicializar las variables impuesto y tasa
        impuesto = item['impuesto']
        tasa = item['tasa_cuota']
        if impuesto != 'N/A':
            impuesto = float(impuesto)
        else:
            impuesto = 0.0  # o cualquier valor predeterminado que consideres adecuado
        
        if tasa != 'N/A':
            tasa = float(tasa)
        else:
            tasa = 0.0  # o cualquier valor predeterminado que consideres adecuado
        clave = item['clave']
         # Crear un p√°rrafo para la descripci√≥n
        descripcion_paragraph = Paragraph(descripcion, custom_style)
        unidad_paragraph = Paragraph(unidad, custom_style)
        table_data.append([
            f"{cantidad:.2f}",
            clave,
            descripcion_paragraph,
            unidad_paragraph,
            f"{valor_unitario:,.2f}",
            f"{importe:,.2f}",
            f"{impuesto:,.2f}",
            f"{tasa:.2f}",
        ])

    # Crear la tabla
    table = Table(table_data, colWidths=[1.0 * cm, 1.5 * cm, 8.5 * cm, 1.5 * cm, 2 * cm, 2 * cm, 1.5 * cm, 1.5 * cm, 1.5 * cm])
    table.setStyle(TableStyle([
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 0), (-1, 0), prussian_blue),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
    ]))

    # Guardar la tabla en el PDF
    table.wrapOn(c, width, height)
    table.drawOn(c, alineado_x, alineado_y - len(table_data) * line_height)

    # Ajustar el alineado_y para seguir escribiendo debajo de la tabla
    alineado_y -= len(table_data) * line_height + 2 * line_height

    # Totales
    c.setFont("Helvetica-Bold", 12)
   
    c.setFont("Helvetica", 10)
    alineado_y -= line_height

     # Importe con letra
    alineado_y -= 2 * line_height
    c.drawString(alineado_x, alineado_y, "Importe con Letra:")
    total_letras = num2words(float(data['total']), lang='es', to='currency', currency='MXN')
    c.drawString(alineado_x, alineado_y - 10, total_letras)
    #c.drawRightString(alineado_x, alineado_y , f"{data['importe_con_letra']}")
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.setFillColor(prussian_blue)
    c.rect(alineado_x + 390 ,alineado_y - 50,110,62, fill=True, stroke=False) #Barra azul superior | Subtotal
    c.setFillColor(white)
    c.drawRightString(alineado_x + 500, alineado_y , f"Subtotal:")
    c.setFillColor(black)
    c.drawRightString(alineado_x + 555, alineado_y, f"{float(data['subtotal']):,.2f}")
    alineado_y -= line_height
    c.setFillColor(white)
    c.drawRightString(alineado_x + 500, alineado_y, f"Impuestos trasladados:")
    c.setFillColor(black)
    c.drawRightString(alineado_x + 555, alineado_y, f"{float(data['impuestos']):,.2f}")
    alineado_y -= line_height
    if data['iva_retenido'] > 0:
        c.setFillColor(white)
        c.drawRightString(alineado_x + 500, alineado_y, f"Impuestos retenidos:")
        c.setFillColor(black)
        c.drawRightString(alineado_x + 555, alineado_y, f"{float(data['iva_retenido']):,.2f}")
        alineado_y -= line_height
    if data['isr_retenido'] > 0:
        c.setFillColor(white)
        c.drawRightString(alineado_x + 500, alineado_y, f"ISR:")
        c.setFillColor(black)
        c.drawRightString(alineado_x + 555, alineado_y, f"{float(data['isr_retenido']):,.2f}")
        alineado_y -= line_height
    c.setFillColor(white)
    c.drawRightString(alineado_x + 500, alineado_y, f"Total:")
    c.setFillColor(black)
    c.drawRightString(alineado_x + 555, alineado_y, f"{float(data['total']):,.2f}")
    # Otros detalles
    

    otros_detalles = [
        ["Folio Fiscal", "Fecha y Hora de Certificaci√≥n", "No. Certificado Digital", "M√©todo de Pago"],
        [data['uuid'], data['fecha_timbrado'], data['no_certificado'], data['metodo_pago']]
    ]
    detalles_table = Table(otros_detalles, colWidths=[5 * cm, 5 * cm, 4.5 * cm, 4.5 * cm])
    detalles_table.setStyle(TableStyle([
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, 0), (-1, 0), prussian_blue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
    ]))

    # Guardar la tabla de detalles en el PDF
    detalles_table.wrapOn(c, width, height)
    detalles_table.drawOn(c, alineado_x, 180)
    alineado_y -= 4 * line_height
     # Utilizar Paragraph para las l√≠neas largas
    styles = getSampleStyleSheet()
    styleN = styles["BodyText"]
    styleN.fontSize = 6
    c.setFont("Helvetica", 6)
    c.line(30,177,580,177)
    c.drawString(alineado_x, 170, f"ESTE DOCUMENTO ES UNA REPRESENTACI√ìN IMPRESA DE UN CFDI v4.0")
    
    # Reducir el ancho de los p√°rrafos
    reduced_width = width * 0.7  # Ajusta este valor seg√∫n sea necesario

    sello_cfd_paragraph = Paragraph(f"Sello Digital del CFDI: {data['sello_cfd']}", styleN)
    sello_cfd_paragraph.wrapOn(c,  reduced_width, line_height * 4)
    sello_cfd_paragraph.drawOn(c, alineado_x, 130)
    alineado_y -= line_height * 5
    
    sello_sat_paragraph = Paragraph(f"Sello del SAT: {data['sello_sat']}", styleN)
    sello_sat_paragraph.wrapOn(c,  reduced_width, line_height * 4)
    sello_sat_paragraph.drawOn(c, alineado_x, 90)
    alineado_y -= line_height * 3
    c.drawString(alineado_x, 40, f"No. serie CSD SAT {data['no_certificadoSAT']}")

    sello_cfd_paragraph = Paragraph(f"Cadena Original del complemento de certificaci√≥n digital del SAT: {data['cadena_original']}", styleN)
    sello_cfd_paragraph.wrapOn(c,  reduced_width, line_height * 4)
    sello_cfd_paragraph.drawOn(c, alineado_x, 50)
    alineado_y -= line_height * 5
    
   

    c.showPage()
    c.save()

    buffer.seek(0)
    return buffer

def generar_cfdi(request, pk):
    #
    buffer = cfdi_compras(None, pk)
    # Crear la respuesta HTTP con el PDF
    factura = Facturas.objects.get(id=pk)
    folio_fiscal = factura.emisor.get('uuid', f'factura_{factura.id}')
    return HttpResponse(buffer, content_type='application/pdf', headers={
        'Content-Disposition': f'attachment; filename="{folio_fiscal}.pdf"'
    })
    # Crear la respuesta HTTP con el PDF
    #folio_fiscal = data['uuid']
def crear_pdf_cfdi_buffer(factura):
    buffer = cfdi_compras(None, factura.id)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_file:
        tmp_file.write(buffer.read())
        return tmp_file.name

def generar_qr(data):
    # URL del acceso al servicio
    url = "https://verificacfdi.facturaelectronica.sat.gob.mx/default.aspx"
    
    # Construcci√≥n de la cadena de datos para el QR
    qr_data = f"{url}?id={data['uuid']}&re={data['rfc_emisor']}&rr={data['rfc_receptor']}&tt={float(data['total']):.6f}&fe={data['sello_cfd'][-8:]}"
    
    # Generar el c√≥digo QR
    qr = qrcode.make(qr_data)
    
    # Guardar el QR como imagen temporal
    qr_img = io.BytesIO()
    qr.save(qr_img, format='PNG')
    qr_img.seek(0)
    
    return qr_img



def convert_excel_gasto(gastos):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Gastos_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Gastos')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')
    percent_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(percent_style)

    columns = ['Folio','Fecha Autorizaci√≥n','Distrito','Proyectos','Subproyectos','Colaborador','Solicitado para',
               'Importe','Fecha Creaci√≥n','Status','Autorizado por','Facturas','Status Pago']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16
        if col_num == 5: #Columna del proveedor
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        if col_num == 2:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 20

    columna_max = len(columns)+2

    # Agregar los mensajes
    ws.cell(column = columna_max, row = 1, value='{Reporte Creado Autom√°ticamente por SAVIA 2.0. UH}').style = messages_style
    ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}').style = messages_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 30

    # Agregar los encabezados de las nuevas columnas debajo de los mensajes
    ws.cell(row=3, column = columna_max, value="Total de Gastos").style = head_style
    ws.cell(row=4, column = columna_max, value="Sumatoria de Pagos Pendientes").style = head_style
   

    # Asumiendo que las filas de datos comienzan en la fila 2 y terminan en row_num
    ws.cell(row=3, column=columna_max + 1, value=f"=COUNTA(A:A)-1").style = body_style
    ws.cell(row=4, column=columna_max + 1, value=f"=SUM(F:F)").style = money_resumen_style
  

   
    
    for gasto in gastos:
        row_num = row_num + 1    
        
        # Manejar autorizado_at_2
        if gasto.approbado_fecha2 and isinstance(gasto.approbado_fecha2, datetime):
        # Si autorizado_at_2 es timezone-aware, convi√©rtelo a timezone-naive
            autorizado_at_2_naive = gasto.approbado_fecha2.astimezone(pytz.utc).replace(tzinfo=None)
        else:
            autorizado_at_2_naive = ''
        
        # Manejar created_at
        if gasto.created_at and isinstance(gasto.created_at, datetime):
        # Si created_at es timezone-aware, convi√©rtelo a timezone-naive
           created_at_naive = gasto.created_at.astimezone(pytz.utc).replace(tzinfo=None)
        else:
            created_at_naive = ''

        if gasto.pagada:
            pagada = "Con Pago"
        else:
            pagada = "Sin Pago"

        if gasto.facturas.exists():
            facturas = "Con Facturas"
        else:
            facturas = "Sin Facturas"
        
        if gasto.autorizar2:
            status = "Autorizado"
            
            if gasto.distrito.nombre == "MATRIZ":
                autorizado_por = str(gasto.superintendente.staff.staff.first_name) + ' ' + str(gasto.superintendente.staff.staff.last_name)
            elif gasto.autorizado_por2:
                autorizado_por = str(gasto.autorizado_por2.staff.staff.first_name) + ' ' + str(gasto.autorizado_por2.staff.staff.last_name)
            else:
                autorizado_por = "NR"
        elif gasto.autorizar2 == False:
            status = "Cancelado"
            if gasto.distrito.nombre == "MATRIZ":
                autorizado_por = str(gasto.superintendente.staff.staff.first_name) + ' ' + str(gasto.superintendente.staff.staff.last_name)
            elif gasto.autorizado_por2:
                autorizado_por = str(gasto.autorizado_por2.staff.staff.first_name) + ' ' + str(gasto.autorizado_por2.staff.staff.last_name)
            else:
                autorizado_por = "NR"
        elif gasto.autorizar:
            autorizado_por =str(gasto.superintendente.staff.staff.first_name) + ' ' + str(gasto.superintendente.staff.staff.last_name)
            status = "Autorizado | Falta una autorizaci√≥n"
            if gasto.facturas:
                facturas = gasto.facturas.exists()
            else:
                facturas = False
        elif gasto.autorizar == False:
            status = "Cancelado"
            autorizado_por = str(gasto.superintendente.staff.staff.last_name)
        else:
            autorizado_por = "Faltan autorizaciones"
            status = "Faltan autorizaciones"

        proyectos = set()
        subproyectos = set()
        articulos_gasto = Articulo_Gasto.objects.filter(gasto=gasto)
        for articulo in articulos_gasto:
            if articulo.proyecto:
                proyectos.add(str(articulo.proyecto.nombre))
            if articulo.subproyecto:
                subproyectos.add(str(articulo.subproyecto.nombre))

        proyectos_str = ', '.join(proyectos)
        subproyectos_str = ', '.join(subproyectos)

        row = [
            gasto.folio,
            autorizado_at_2_naive,
            gasto.distrito.nombre,
            proyectos_str,
            subproyectos_str,
            gasto.staff.staff.staff.first_name + ' ' + gasto.staff.staff.staff.last_name,
            gasto.colaborador.staff.staff.first_name + ' '  + gasto.colaborador.staff.staff.last_name if gasto.colaborador else '',
            gasto.get_total_solicitud,
            created_at_naive,
            status,
            autorizado_por,
            facturas,
            pagada
            #f'=IF(I{row_num}="",G{row_num},I{row_num}*G{row_num})',  # Calcula total en pesos usando la f√≥rmula de Excel
            #created_at_naive,
        ]

    
        for col_num in range(len(row)):
            (ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style
            if col_num ==1 or col_num == 6:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style
            if col_num == 5:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = money_style
       
    
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)


def generar_acuse_recibo(usuario, gastos_enviar, tipo):
      #Configuration of the PDF object
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)

    styles = getSampleStyleSheet()
    normal_style = styles['Normal']
    custom_style = ParagraphStyle(
        name='CustomStyle',
        parent=normal_style,
        fontSize=10,
        alignment=1,  # Alineaci√≥n centrada (0 = izquierda, 1 = centro, 2 = derecha)
    )
    #Azul Vordcab
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)
    caja_iso = 760
    #Elaborar caja
    #c.line(caja_iso,500,caja_iso,720)



    #Encabezado
    c.drawString(420,caja_iso,'Preparado por:')
    c.drawString(420,caja_iso-10,'S2')
    c.drawString(520,caja_iso,'Aprobaci√≥n')
    c.drawString(520,caja_iso-10,'---')
    c.drawString(150,caja_iso-20,'N√∫mero de documento')
    c.drawString(160,caja_iso-30,'--------')
    c.drawString(245,caja_iso-20,'Clasificaci√≥n del documento')
    c.drawString(275,caja_iso-30,'--------')
    c.drawString(355,caja_iso-20,'Nivel del documento')
    c.drawString(380,caja_iso-30, '--------')
    c.drawString(440,caja_iso-20,'Revisi√≥n No.')
    c.drawString(452,caja_iso-30,'000')
    c.drawString(510,caja_iso-20,'Fecha de Emisi√≥n')
    c.drawString(525,caja_iso-30,'04/2024')

    caja_proveedor = caja_iso - 65
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(150,750,250,20, fill=True, stroke=False) #Barra azul superior Solicitud
    c.rect(20,caja_proveedor - 8,565,20, fill=True, stroke=False) #Barra azul superior Proveedor | Detalle
    c.rect(20,615,565,2, fill=True, stroke=False) #Linea posterior horizontal
    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    if tipo == 'gasto':
        c.drawCentredString(280,755,'Acuse de Recibo Gastos')
    if tipo == 'viatico':
        c.drawCentredString(280,755,'Acuse de Recibo Vi√°ticos')
    c.setLineWidth(.3) #Grosor
    c.line(20,caja_proveedor-8,20,615) #Eje Y donde empieza, Eje X donde empieza, donde termina eje y,donde termina eje x (LINEA 1 contorno)
    c.line(584,caja_proveedor-8,584,615) #Linea 2 contorno
    c.drawInlineImage('static/images/logo_vordcab.jpg',45,730, 3 * cm, 1.5 * cm) #Imagen vortec

    c.setFillColor(white)
    c.setFont('Helvetica-Bold',11)
    #c.drawString(120,caja_proveedor,'Infor')
    c.drawString(300,caja_proveedor, 'Detalles')
    inicio_central = 300
    #c.line(inicio_central,caja_proveedor-25,inicio_central,520) #Linea Central de caja Proveedor | Detalle
    c.setFillColor(black)
    c.setFont('Helvetica',9)
    c.drawString(30,caja_proveedor-20,'Present√≥:')
    c.drawString(30,caja_proveedor-40,'Distrito:')
    c.drawString(30,caja_proveedor-60,'Fecha:')
    # Segunda columna del encabezado
    #c.drawString(30,caja_proveedor-80,'Empresa')


    folios_str = ', '.join(str(g.folio) for g in gastos_enviar)
    fecha_hoy = datetime.today().strftime("%d/%m/%Y")
    c.setFont('Helvetica-Bold',12)
    c.drawString(300,caja_proveedor-20,'FOLIOS:')
    c.setFillColor(rojo)
    c.setFont('Helvetica-Bold',12)
    c.drawString(350,caja_proveedor-20, folios_str)

    c.setFillColor(black)
    c.setFont('Helvetica',9)
    c.drawString(100,caja_proveedor-20, usuario.staff.staff.first_name+' '+ usuario.staff.staff.last_name)
    c.drawString(100,caja_proveedor-40, usuario.distritos.nombre)
    c.drawString(100,caja_proveedor-60, fecha_hoy)

    # --- Tabla de resumen de gastos ---
    tabla_gastos_data = [['Folio', 'Monto Solicitado', 'Monto Comprobado', 'Diferencia']]
    for gasto in gastos_enviar:
        if tipo == 'gasto':
            monto = gasto.get_total_solicitud
            comprobado = gasto.suma_total_facturas if hasattr(gasto, 'suma_total_facturas') else Decimal('0.00')
        if tipo == 'viatico':
            monto = gasto.get_total
            comprobado = gasto.suma_total_facturas if hasattr(gasto, 'suma_total_facturas') else Decimal('0.00')
        diferencia = comprobado - monto
        color = colors.black
        if diferencia > 0:
            color = colors.green
        elif diferencia < 0:
            color = colors.red

        tabla_gastos_data.append([
            str(gasto.folio),
            f"${monto:,.2f}",
            f"${comprobado:,.2f}",
            Paragraph(f"${diferencia:,.2f}", ParagraphStyle('diff', textColor=color, alignment=TA_RIGHT))
        ])

    # Crear la tabla
    tabla = Table(tabla_gastos_data, colWidths=[3*cm, 4*cm, 4*cm, 5*cm])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (1,1), (-1,-1), 'RIGHT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))

    # Posici√≥n de la tabla
    tabla.wrapOn(c, 500, 200)
    tabla.drawOn(c, 80, 500)  # Ajusta esta posici√≥n si necesitas moverla m√°s arriba/abajo


    c.setFont("Helvetica", 10)
    c.drawString(100, 280, "Nombre y firma de quien recibe:")
    c.line(100, 245, 500, 245)

    c.drawString(100, 230, "Fecha de recepci√≥n:")
    c.line(100, 205, 300, 205)

    c.drawString(100, 190, "Comentarios u observaciones:")
    c.rect(100, 100, 400, 80)  # cuadro para observaciones (de 300 a 380)

    # L√≠nea final del documento (opcional)
    c.setFont("Helvetica-Oblique", 8)
    c.drawCentredString(300, 80, "Este acuse de recibo corresponde a la comprobaci√≥n de gastos generada por SAVIA 2.0.")
    c.showPage()
    c.save()

    buffer.seek(0)
    return buffer