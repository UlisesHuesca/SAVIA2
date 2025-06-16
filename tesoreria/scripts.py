from compras.models import Compra
from xml.etree.ElementTree import ParseError
from tesoreria.models import Pago, Facturas
from gastos.models import Solicitud_Gasto, Factura
from viaticos.models import Viaticos_Factura 
from django.db.models import F, Sum, Q, Count
from django.db.models.functions import ExtractYear
import mysql.connector
from collections import defaultdict
import os
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
import fitz  # PyMuPDF
import re


def update_pagado_real():
    pagos = Pago.objects.all()

    for pago in pagos:
        if pago.comprobante_pago and not pago.pagado_real:
            detalles = pago.detalles_comprobante
            fecha = detalles.get('fecha', None)
            if fecha:
                try:
                    # Convierte la fecha al formato de Django
                    pagado_real_date = datetime.strptime(fecha, '%d/%m/%Y').date()
                    pago.pagado_real = pagado_real_date
                    pago.save()
                    print(f"Actualizado pago {pago.id} con fecha {pagado_real_date}")
                except ValueError as e:
                    print(f"Error al convertir la fecha para el pago {pago.id}: {e}")
            else:
                print(f"No se encontró fecha en el comprobante de pago {pago.id}")
        else:
            if not pago.comprobante_pago:
                print(f"Pago {pago.id} no tiene comprobante de pago")
            else:
                print(f"Pago {pago.id} ya tiene una fecha en pagado_real")


def verificar_compras_colocadas():
    print('Procesando...')
    # Obtener todas las compras que cumplen con las condiciones
    compras = Compra.objects.filter(autorizado2=True, pagada=False, req__orden__distrito__id = 2)
    print(f"Encontradas {compras.count()} compras para verificar.")
    
    count = 0
    for compra in compras:
        # Obtener la suma de los montos de los pagos relacionados con la compra
        total = Pago.objects.filter(oc=compra).aggregate(cantidad_sumada=Sum('monto'))
        
        monto_total = total['cantidad_sumada'] if total['cantidad_sumada'] is not None else 0
        
        if monto_total >= compra.costo_oc:  # Asumiendo que "monto" es el campo de la Compra con el monto total
            count += 1
            print(f"Compra con ID: {compra.id} ha sido marcada como pagada. {count}")
            compra.pagada = True
            compra.save()

#verificar_compras_colocadas()
def verificar_gastos_pagados():
    print('Procesando...')
    # Obtener todas las compras que cumplen con las condiciones
    gastos = Solicitud_Gasto.objects.filter(autorizar2=True, pagada=False)
    print(f"Encontradas {gastos.count()} gastos para verificar.")
    
    count = 0
    for gasto in gastos:
        # Obtener la suma de los montos de los pagos relacionados con la compra
        total = Pago.objects.filter(gasto=gasto).aggregate(cantidad_sumada=Sum('monto'))
        monto_total = total['cantidad_sumada'] if total['cantidad_sumada'] is not None else 0
        
        if monto_total >= gasto.get_total_solicitud:  # Asumiendo que "monto" es el campo de la Compra con el monto total
            count += 1
            print(f"Gasto con ID: {gasto.id} ha sido marcada como pagada. {count}")
            gasto.pagada = True
            gasto.save()


def actualizar_gastos_factura_con_pdf():
    conn_savia1 = mysql.connector.connect(
        host='localhost', 
        user='root', 
        password='*$HbAq*/4528*', 
        database='SAVIA1'
    )
    cursor_savia1 = conn_savia1.cursor()

    conn_savia2_default = mysql.connector.connect(
        host='localhost', 
        user='root', 
        password='*$HbAq*/4528*', 
        database='savia2_default'
    )
    cursor_savia2_default = conn_savia2_default.cursor()

    consulta_facturas = """
    SELECT f.IDGASTO, f.indice, f.ruta_factura
    FROM facturasgastostb f
    JOIN gastostb g ON g.IDGASTO = f.IDGASTO
    WHERE f.ruta_factura IS NOT NULL AND f.IDGASTO > 16349 AND g.IDALMACEN = 2
    ORDER BY f.IDGASTO, f.indice;
    """
    cursor_savia1.execute(consulta_facturas)
    facturas = cursor_savia1.fetchall()

    idgasto_actual = None

    for idgasto, indice, ruta_factura in facturas:
        if idgasto_actual != idgasto:
            idgasto_actual = idgasto
            # Obtener todos los registros de gastos_factura para el IDGASTO actual
            cursor_savia2_default.execute("""
                SELECT id FROM gastos_factura WHERE solicitud_gasto_id = %s ORDER BY id;
            """, (idgasto + 2000,))
            gastos_ids = cursor_savia2_default.fetchall()
            print(gastos_ids)

        # Asumiendo que el indice corresponde al orden de los registros obtenidos
        if indice - 1 < len(gastos_ids):
            gasto_id = gastos_ids[indice - 1][0]  # Obtiene el id basado en el índice
            cursor_savia2_default.execute("""
                UPDATE gastos_factura SET archivo_pdf = %s WHERE id = %s;
            """, (ruta_factura, gasto_id))
            print(ruta_factura, gasto_id)

    conn_savia2_default.commit()

    cursor_savia1.close()
    conn_savia1.close()
    cursor_savia2_default.close()
    conn_savia2_default.close()

    print("Las actualizaciones se han completado.")

def actualizar_gastos_factura_con_xml():
    conn_savia1 = mysql.connector.connect(
        host='localhost', 
        user='root', 
        password='peruzzi25',  # Cambia [TU_PASSWORD] por la contraseña real
        database='savia_activo_a'
    )
    cursor_savia1 = conn_savia1.cursor()

    conn_savia2 = mysql.connector.connect(
        host='localhost', 
        user='root', 
        password='peruzzi25',  # Cambia [TU_PASSWORD] por la contraseña real
        database='savia2'
    )
    cursor_savia2 = conn_savia2.cursor()

    consulta_facturas = """
    SELECT f.IDGASTO, f.indice, f.ruta_xml  # Asumiendo que existe ruta_xml en la tabla
    FROM facturasgastostb f
    JOIN gastostb g ON g.IDGASTO = f.IDGASTO
    WHERE f.ruta_xml IS NOT NULL AND f.IDGASTO > 16349 AND g.IDALMACEN = 1
    ORDER BY f.IDGASTO, f.indice;
    """
    cursor_savia1.execute(consulta_facturas)
    facturas = cursor_savia1.fetchall()

    idgasto_actual = None

    for idgasto, indice, ruta_xml in facturas:
        if idgasto_actual != idgasto:
            idgasto_actual = idgasto
            cursor_savia2.execute("""
                SELECT id FROM gastos_factura WHERE solicitud_gasto_id = %s ORDER BY id;
            """, (idgasto + 6000,))
            gastos_ids = cursor_savia2.fetchall()
            print(gastos_ids)

        if indice - 1 < len(gastos_ids):
            gasto_id = gastos_ids[indice - 1][0]  # Obtiene el id basado en el índice
            #if gasto_id not in (19095, 19096):  # Excluir los ids no deseados
            cursor_savia2.execute("""
            UPDATE gastos_factura SET archivo_xml = %s WHERE id = %s;
            """, (ruta_xml, gasto_id))
            print(ruta_xml, gasto_id)

    conn_savia2.commit()

    cursor_savia1.close()
    conn_savia1.close()
    cursor_savia2.close()
    conn_savia2.close()

    print("Las actualizaciones de XML se han completado.")

# No olvides cambiar '[TU_PASSWORD]' por tu contraseña real antes de ejecutar el script.


def actualizar_gastos_factura_con_xml_general():
    conn_savia1 = mysql.connector.connect(
        host='localhost', 
        user='root', 
        password='*$HbAq*/4528*',  # Asegúrate de usar la contraseña correcta
        database='SAVIA1'
    )
    cursor_savia1 = conn_savia1.cursor()

    conn_savia2_default = mysql.connector.connect(
        host='localhost', 
        user='root', 
        password='*$HbAq*/4528*',  # Asegúrate de usar la contraseña correcta
        database='savia2_default'
    )
    cursor_savia2_default = conn_savia2_default.cursor()

    consulta_facturas = """
    SELECT f.IDGASTO, f.indice, f.ruta_xml
    FROM facturasgastostb f
    JOIN gastostb g ON g.IDGASTO = f.IDGASTO
    WHERE f.ruta_xml IS NOT NULL AND f.IDGASTO <= 16349
    ORDER BY f.IDGASTO, f.indice;
    """
    cursor_savia1.execute(consulta_facturas)
    facturas = cursor_savia1.fetchall()

    for idgasto, indice, ruta_xml in facturas:
        # Obtén el gasto_id basado en el IDGASTO y el índice proporcionado
        cursor_savia2_default.execute("""
            SELECT id FROM gastos_factura WHERE solicitud_gasto_id = %s ORDER BY id LIMIT %s,1;
        """, (idgasto, indice - 1))  # Ajusta el OFFSET según el índice
        gasto_factura = cursor_savia2_default.fetchone()

        if gasto_factura:
            gasto_id = gasto_factura[0]
            cursor_savia2_default.execute("""
                UPDATE gastos_factura SET archivo_xml = %s WHERE id = %s;
            """, (ruta_xml, gasto_id))
            print(ruta_xml, gasto_id)

    conn_savia2_default.commit()

    cursor_savia1.close()
    conn_savia1.close()
    cursor_savia2_default.close()
    conn_savia2_default.close()

    print("Las actualizaciones de XML se han completado.")




def migrate_blob_to_files():
    # Establecer conexión con la base de datos
    conn = mysql.connector.connect(
    host='localhost',
    user='root',
    password='peruzzi25',
    database='savia_activo_a'
    )
    cursor = conn.cursor()
    #Se crea el directorio
    if not os.path.exists('media/comprobantes/'):
        os.makedirs('media/comprobantes/')

    # Consultar el campo BLOB
    query = "SELECT IDPAGO, FOLIO, COMPROBANTE FROM savia_activo_a.pagostb"
    cursor.execute(query)
    rows = cursor.fetchall()

    for row in rows:
        IDPAGO, FOLIO, COMPROBANTE = row
        

        # Verifica si comprobante_blob no es None
        if COMPROBANTE is not None:
        # Define el nombre del archivo y la ruta
            file_name = f"{IDPAGO}.pdf"  # Suponiendo que es un PDF, ajusta si es necesario
            path = f"media/comprobantes/{file_name}"
        
            #Guarda el BLOB como archivo físico
            with open(path, 'wb') as file:
                file.write(COMPROBANTE)

            # Actualiza la columna con la ruta del archivo en la base de datos
            cursor.execute("UPDATE savia_activo_a.pagostb SET ruta_comprobante=%s WHERE IDPAGO=%s", (path, IDPAGO))
        else:
            print(f"El registro con ID {IDPAGO} tiene un comprobante BLOB nulo o vacío.")   

    conn.commit()
    cursor.close()
    conn.close()

    print("Migración completada")

def migrate_compras_to_directory():
    conn = mysql.connector.connect(
        host='localhost',
        user='root',
        password='peruzzi25',
        database='savia_activo_a'
    )
    cursor = conn.cursor()

    # Consultar el campo BLOB
    query = "SELECT IDFACTURACOMPRA, FACTURA, XML, IDCOMPRA FROM savia_activo_a.facturascomprastb ORDER BY IDCOMPRA, IDFACTURACOMPRA"
    cursor.execute(query)
    rows = cursor.fetchall()

    counters = {}
    for row in rows:
        IDFACTURACOMPRA, FACTURA, XML, IDCOMPRA = row

        # Para cada IDCOMPRA, inicializar o actualizar el contador
        if IDCOMPRA not in counters:
            counters[IDCOMPRA] = {'pdf': 0, 'xml': 0}

        if FACTURA:
            counters[IDCOMPRA]['pdf'] += 1
            indice = counters[IDCOMPRA]['pdf']
            file_name = f"{IDCOMPRA}-PDF-{counters[IDCOMPRA]['pdf']}.pdf"
            path = f"facturas/{file_name}"
            with open(path, 'wb') as file:
                file.write(FACTURA)
            cursor.execute("UPDATE savia_activo_a.facturascomprastb SET ruta_factura=%s, indice=%s WHERE IDFACTURACOMPRA=%s", (path, indice, IDFACTURACOMPRA))

        if XML:
            counters[IDCOMPRA]['xml'] += 1
            indice = counters[IDCOMPRA]['xml']
            file_name = f"{IDCOMPRA}-XML-{counters[IDCOMPRA]['xml']}.xml"
            path = f"xml/{file_name}"
            with open(path, 'wb') as file:
                file.write(XML)
            cursor.execute("UPDATE savia_activo_a.facturascomprastb SET ruta_xml=%s, indice=%s WHERE IDFACTURACOMPRA=%s", (path, indice, IDFACTURACOMPRA))

    conn.commit()
    cursor.close()
    conn.close()

    print("Migración completada")

def migrate_gastosfacturas_to_files():
    # Establecer conexión con la base de datos
    conn = mysql.connector.connect(
        host='localhost',
        user='root',
        password='peruzzi25',
        database='savia_activo_a'
    )
    cursor = conn.cursor()
    print('Iniciando proceso')
    # Crear el directorio si no existe
    directory = 'media/'
    if not os.path.exists(directory):
        os.makedirs(directory)

    # Consultar los campos BLOB
    query = "SELECT IDGASTO, PDF, PDFCOMPROBACION, FACTURA, XML FROM savia_activo_a.gastostb"
    cursor.execute(query)
    rows = cursor.fetchall()

    for row in rows:
        IDGASTO, PDF, PDFCOMPROBACION, FACTURA, XML = row
        
        # Guardar PDF
        if PDF is not None:
            pdf_path = f"{directory}/gastos_pdf/PDF_ID{IDGASTO}.pdf"
            pdf_path_r = f"/gastos_pdf/PDF_ID{IDGASTO}.pdf"
            with open(pdf_path, 'wb') as file:
                file.write(PDF)
            cursor.execute("UPDATE savia_activo_a.gastostb SET ruta_pdf=%s WHERE IDGASTO=%s", (pdf_path_r, IDGASTO))

        #Guarda comprabación
        if PDFCOMPROBACION is not None:
            pdfc_path = f"{directory}/gastos_comprobacion/PDFC_ID{IDGASTO}.pdf"
            pdfc_path_r = f"/gastos_comprobacion/PDFC_ID{IDGASTO}.pdf"
            with open(pdfc_path, 'wb') as file:
                file.write(PDFCOMPROBACION)
            cursor.execute("UPDATE savia_activo_a.gastostb SET ruta_comprobacion = %s WHERE IDGASTO=%s",(pdfc_path_r, IDGASTO))

        # Guardar FACTURA
        if FACTURA is not None:
            factura_path = f"{directory}/gastos_factura/FACTURA_ID{IDGASTO}.pdf"
            factura_path_r = f"/gastos_factura/FACTURA_ID{IDGASTO}.pdf"
            with open(factura_path, 'wb') as file:
                file.write(FACTURA)
            cursor.execute("UPDATE savia_activo_a.gastostb SET ruta_factura=%s WHERE IDGASTO=%s", (factura_path_r,IDGASTO))
        
        # Guardar XML
        if XML is not None:
            xml_path = f"{directory}/gastos_xml/XML_ID{IDGASTO}.xml"
            xml_path_r = f"/gastos_xml/XML_ID{IDGASTO}.xml"
            with open(xml_path, 'wb') as file:
                file.write(XML)
            cursor.execute("UPDATE savia_activo_a.gastostb SET ruta_xml=%s WHERE IDGASTO=%s", (xml_path_r, IDGASTO))

    conn.commit()
    cursor.close()
    conn.close()

    print("Migración completada")


def get_file_type_from_signature(blob_data):
    if blob_data.startswith(b'%PDF-'):
        return 'pdf'
    elif blob_data.startswith(b'<?xml'):
        return 'xml'
    elif blob_data[:4] == b'\xFF\xD8\xFF\xE0':
        return 'jpg'
    # Agrega más condiciones aquí para otros tipos de archivo
    else:
        return 'unknown'

def process_db_records():
    # Conexión a la base de datos y obtención de los primeros 10 registros
    import mysql.connector

    conn = mysql.connector.connect(
        host='localhost',
        user='root',
        password='peruzzi25',
        database='saviauno'
    )
    cursor = conn.cursor()

    query = "SELECT FACTURA FROM saviados.gastostb LIMIT 100,200"
    cursor.execute(query)
    rows = cursor.fetchall()

    for row in rows:
        for data in row:
            if data:
                print(get_file_type_from_signature(data))

    cursor.close()
    conn.close()

def migrate_tablafacturasgastos_to_files():
    # Establecer conexión con la base de datos
    conn = mysql.connector.connect(
        host='localhost',
        user='root',
        password='peruzzi25',
        database='saviauno_pr'
    )
    cursor = conn.cursor()
    
    # Crear el directorio si no existe
    directory = 'media/'
    if not os.path.exists(directory):
        os.makedirs(directory)

    # Consultar los campos BLOB
    query = "SELECT IDFACTURAGASTO, FACTURA, XML, IDGASTO FROM saviauno_pr.facturasgastostb"
    cursor.execute(query)
    rows = cursor.fetchall()

    for row in rows:
        IDFACTURAGASTO, FACTURA, XML, IDGASTO = row
        
        #Guardar FACTURA
        if FACTURA is not None:
            factura_path = f"{directory}/gastos_facturatb/FACTURA_{IDFACTURAGASTO}_{IDGASTO}.pdf"
            factura_path_r = f"/gastos_facturatb/FACTURA_{IDFACTURAGASTO}_{IDGASTO}.pdf"
            with open(factura_path, 'wb') as file:
                file.write(FACTURA)
            cursor.execute("UPDATE saviauno_pr.facturasgastostb SET ruta_factura=%s WHERE IDFACTURAGASTO=%s ", (factura_path_r,IDFACTURAGASTO))
        
        # Guardar XML
        if XML is not None:
            xml_path = f"{directory}/gastos_xmltb/XML_{IDFACTURAGASTO}_{IDGASTO}.xml"
            xml_path_r =f"/gastos_xmltb/XML_{IDFACTURAGASTO}_{IDGASTO}.xml"
            with open(xml_path, 'wb') as file:
                file.write(XML)
            cursor.execute("UPDATE saviauno_pr.facturasgastostb SET ruta_xml=%s WHERE IDFACTURAGASTO=%s", (xml_path_r, IDFACTURAGASTO))

    conn.commit()
    cursor.close()
    conn.close()

    print("Migración completada")



def migrate_tablafacturasgastos_to_files_v2():
    # Establecer conexión con la base de datos
    conn = mysql.connector.connect(
        host='localhost',
        user='root',
        password='peruzzi25',
        database='savia_activo_a'
    )
    cursor = conn.cursor()
    
    # Crear el directorio si no existe
    directory = 'media/'
    if not os.path.exists(directory):
        os.makedirs(directory)

    # Consultar los campos BLOB
    query = "SELECT IDFACTURAGASTO, FACTURA, XML, IDGASTO FROM savia_activo_a.facturasgastostb ORDER BY IDGASTO, IDFACTURAGASTO"
    cursor.execute(query)
    rows = cursor.fetchall()

    # Diccionarios para manejar los índices de cada IDGASTO
    indice_factura_por_idgasto = {}
    indice_xml_por_idgasto = {}

    for row in rows:
        IDFACTURAGASTO, FACTURA, XML, IDGASTO = row
        
        # Incrementar o inicializar índice para FACTURA si existe
        if FACTURA is not None:
            if IDGASTO not in indice_factura_por_idgasto:
                indice_factura_por_idgasto[IDGASTO] = 1
            else:
                indice_factura_por_idgasto[IDGASTO] += 1

            factura_path = f"{directory}gastos_facturatb/FACTURA_{IDGASTO}_{indice_factura_por_idgasto[IDGASTO]}.pdf"
            factura_path_r = f"/gastos_facturatb/FACTURA_{IDGASTO}_{indice_factura_por_idgasto[IDGASTO]}.pdf"
            with open(factura_path, 'wb') as file:
                file.write(FACTURA)
            cursor.execute("UPDATE savia_activo_a.facturasgastostb SET ruta_factura=%s, indice=%s WHERE IDFACTURAGASTO=%s", (factura_path_r, indice_factura_por_idgasto[IDGASTO], IDFACTURAGASTO))
        
        # Incrementar o inicializar índice para XML si existe
        if XML is not None:
            if IDGASTO not in indice_xml_por_idgasto:
                indice_xml_por_idgasto[IDGASTO] = 1
            else:
                indice_xml_por_idgasto[IDGASTO] += 1

            xml_path = f"{directory}gastos_xmltb/XML_{IDGASTO}_{indice_xml_por_idgasto[IDGASTO]}.xml"
            xml_path_r = f"/gastos_xmltb/XML_{IDGASTO}_{indice_xml_por_idgasto[IDGASTO]}.xml"
            with open(xml_path, 'wb') as file:
                file.write(XML)
            cursor.execute("UPDATE savia_activo_a.facturasgastostb SET ruta_xml=%s, indice=%s WHERE IDFACTURAGASTO=%s", (xml_path_r, indice_xml_por_idgasto[IDGASTO], IDFACTURAGASTO))

    conn.commit()
    cursor.close()
    conn.close()

    print("Migración completada")

def actualizar_facturas():
    facturas = Facturas.objects.filter(uuid__isnull=True)  # Filtra facturas sin UUID guardado
    for factura in facturas:
        if factura.factura_xml and os.path.exists(factura.factura_xml.path):  # Verifica si el archivo XML existe
            try:
                data = factura.emisor  # Llama a la propiedad que ya tienes
                if data:  # Verifica si se pudo obtener el UUID y la fecha
                    uuid = data.get('uuid')
                    fecha_timbrado = data.get('fecha_timbrado')

                    if uuid and fecha_timbrado:
                        factura.uuid = uuid
                        factura.fecha_timbrado = fecha_timbrado
                        factura.save()
                        print(f'Actualizada Factura ID {factura.id}: UUID {uuid}')
                    else:
                        print(f'No se pudo obtener UUID y fecha para Factura ID {factura.id}')
                else:
                    print(f'El archivo XML de la factura ID {factura.id} no contiene la información esperada.')

            except (ParseError, FileNotFoundError) as e:
                print(f"Error al procesar el archivo XML para la factura ID {factura.id}: {e}")
                continue  # Salta al siguiente registro si ocurre un error

        else:
            print(f'El archivo XML no existe para la factura ID {factura.id}.')
            continue  # Salta al siguiente registro si el archivo XML no existe


def corregir_rutas_facturas():
    # Obtener todas las facturas
    facturas = Facturas.objects.all()
    
    # Definir los directorios no deseados
    rutas_invalidas = ['facturas/xml/', 'xml/']

    # Iterar sobre cada factura
    for factura in facturas:
        # Verificar el campo archivo_xml
        if factura.factura_xml and factura.factura_xml.name in rutas_invalidas:
            print(f"Corrigiendo ruta XML para la factura ID {factura.id}: {factura.factura_xml.name}")
            factura.factura_xml = None  # Simplemente asignar None sin intentar eliminar el archivo
            factura.save()

        # Verificar el campo archivo_pdf
        if factura.factura_pdf and factura.factura_pdf.name in rutas_invalidas:
            print(f"Corrigiendo ruta PDF para la factura ID {factura.id}: {factura.factura_pdf.name}")
            factura.factura_pdf = None  # Simplemente asignar None sin intentar eliminar el archivo
            factura.save()

    print("Corrección de rutas completada.")

def actualizar_facturas_gastos():
    facturas = Factura.objects.filter(uuid__isnull=True)  #Gastos Filtra facturas sin UUID guardado
    for factura in facturas:
        print(factura.id)
        if factura.archivo_xml and os.path.exists(factura.archivo_xml.path):  # Verifica si el archivo XML existe
            try:
                data = factura.emisor  # Llama a la propiedad que ya tienes
                if data:  # Verifica si se pudo obtener el UUID y la fecha
                    uuid = data.get('uuid')
                    fecha_timbrado = data.get('fecha_timbrado')

                    if uuid and fecha_timbrado:
                        factura.uuid = uuid
                        factura.fecha_timbrado = fecha_timbrado
                        factura.save()
                        print(f'Actualizada Factura ID {factura.id}: UUID {uuid}')
                    else:
                        print(f'No se pudo obtener UUID y fecha para Factura ID {factura.id}')
                else:
                    print(f'El archivo XML de la factura ID {factura.id} no contiene la información esperada.')

            except (ParseError, FileNotFoundError) as e:
                print(f"Error al procesar el archivo XML para la factura ID {factura.id}: {e}")
                continue  # Salta al siguiente registro si ocurre un error

        else:
            print(f'El archivo XML no existe para la factura ID {factura.id}.')
            continue  # Salta al siguiente registro si el archivo XML no existe

def actualizar_facturas_viaticos():
    facturas = Viaticos_Factura.objects.filter(uuid__isnull=True)  #Gastos Filtra facturas sin UUID guardado
    for factura in facturas:
        print(factura.id)
        if factura.factura_xml and os.path.exists(factura.factura_xml.path):  # Verifica si el archivo XML existe
            try:
                data = factura.emisor  # Llama a la propiedad que ya tienes
                if data:  # Verifica si se pudo obtener el UUID y la fecha
                    uuid = data.get('uuid')
                    fecha_timbrado = data.get('fecha_timbrado')

                    if uuid and fecha_timbrado:
                        factura.uuid = uuid
                        factura.fecha_timbrado = fecha_timbrado
                        factura.save()
                        print(f'Actualizada Factura ID {factura.id}: UUID {uuid}')
                    else:
                        print(f'No se pudo obtener UUID y fecha para Factura ID {factura.id}')
                else:
                    print(f'El archivo XML de la factura ID {factura.id} no contiene la información esperada.')

            except (ParseError, FileNotFoundError) as e:
                print(f"Error al procesar el archivo XML para la factura ID {factura.id}: {e}")
                continue  # Salta al siguiente registro si ocurre un error

        else:
            print(f'El archivo XML no existe para la factura ID {factura.id}.')
            continue  # Salta al siguiente registro si el archivo XML no existe

def generar_informe_duplicados_por_anio():
    # Obtener facturas duplicadas sin usar ExtractYear
    facturas_duplicadas = Viaticos_Factura.objects.values('uuid', 'fecha_timbrado').annotate(uuid_count=Count('uuid')).filter(uuid_count__gt=1)
    
    total_facturas_duplicadas = 0
    total_grupos_duplicados = facturas_duplicadas.count()
    
    # Diccionario para almacenar el conteo de facturas duplicadas por año
    conteo_por_anio = {}

    with open('informe_facturas_duplicadas_por_anio.txt', 'w') as file:
        file.write(f"Total de grupos con UUID duplicados: {total_grupos_duplicados}\n")
        
        # Agrupar por año manualmente y mostrar facturas duplicadas
        current_year = None
        for factura_grupo in facturas_duplicadas:
            fecha_timbrado = factura_grupo['fecha_timbrado']
            if fecha_timbrado:
                anio = fecha_timbrado.year  # Extraer el año directamente
            else:
                continue  # Saltar si no hay fecha de timbrado

            uuid_duplicado = factura_grupo['uuid']
            facturas_con_uuid = Viaticos_Factura.objects.filter(uuid=uuid_duplicado)

            # Imprimir el año cuando cambia y actualizar el conteo
            if anio != current_year:
                if current_year is not None:
                    file.write(f"\nNúmero de facturas repetidas en el año {current_year}: {conteo_por_anio[current_year]}\n")
                file.write(f"\nFacturas repetidas en el año {anio}:\n")
                current_year = anio
                conteo_por_anio[anio] = 0  # Inicializar el conteo para este año

            # Escribir UUID duplicado y detalles de cada factura
            file.write(f"UUID duplicado: {uuid_duplicado}\n")
            for factura in facturas_con_uuid:
                file.write(f"  Factura ID: {factura.id}, Fecha: {factura.fecha_timbrado}\n")
            
            # Contar las facturas duplicadas, excluyendo la primera
            conteo_por_anio[anio] += len(facturas_con_uuid) - 1
            total_facturas_duplicadas += len(facturas_con_uuid) - 1

        # Escribir el conteo del último año procesado
        if current_year is not None:
            file.write(f"\nNúmero de facturas repetidas en el año {current_year}: {conteo_por_anio[current_year]}\n")

        file.write(f"\nTotal de facturas duplicadas (excluyendo la primera en cada grupo): {total_facturas_duplicadas}\n")

def generar_reporte_excel():
    # Crear un nuevo libro de trabajo y una hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturas Duplicadas"

    # Definir los encabezados
    encabezados = ['Fecha Timbrado', 'Folio Solicitud', 'Colaborador (Nombre)', 'Colaborador (Apellido)', 'UUID']
    ws.append(encabezados)

    # Obtener facturas duplicadas
    facturas_duplicadas = Facturas.objects.values('uuid', 'fecha_timbrado', 'solicitud_gasto__folio','solicitud_gasto','solicitud_gasto__distrito__nombre').annotate(uuid_count=Count('uuid')).filter(uuid_count__gt=1)

    for factura_grupo in facturas_duplicadas:
        facturas_con_uuid = Facturas.objects.filter(uuid=factura_grupo['uuid'])
        
        for factura in facturas_con_uuid:
            # Extraer la información requerida
            fecha_timbrado = factura.fecha_timbrado
            if fecha_timbrado and fecha_timbrado.tzinfo:
                fecha_timbrado = fecha_timbrado.replace(tzinfo=None)  # Eliminar la información de la zona horaria
            solicitud_gasto = factura.solicitud_gasto.folio if factura.solicitud_gasto else "N/A"
            distrito = factura.solicitud_gasto.distrito.nombre
            colaborador_nombre = f"{factura.solicitud_gasto.staff.staff.staff.first_name} {factura.solicitud_gasto.staff.staff.staff.last_name}"
            uuid = factura.uuid

            # Agregar la información a una nueva fila en el Excel
            ws.append([fecha_timbrado, solicitud_gasto, colaborador_nombre, distrito, uuid])

    # Ajustar el ancho de las columnas
    for col in range(1, len(encabezados) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].auto_size = True

    # Guardar el archivo Excel
    wb.save("informe_facturas_duplicadas.xlsx")

def eliminar_facturas_duplicadas_compras():
    # Buscar facturas duplicadas basadas en el UUID
    facturas_duplicadas = Facturas.objects.values('uuid').annotate(uuid_count=Count('uuid')).filter(uuid_count__gt=1)

    # Contador de facturas eliminadas
    total_eliminadas = 0

    # Procesar cada grupo de facturas duplicadas
    for factura_grupo in facturas_duplicadas:
        # Buscar todas las facturas con el mismo UUID
        facturas_con_uuid = Facturas.objects.filter(uuid=factura_grupo['uuid']).order_by('fecha_timbrado')

        # Mantener solo la primera factura (la más antigua, por ejemplo)
        factura_a_conservar = facturas_con_uuid.first()

        # Eliminar las demás facturas
        facturas_a_eliminar = facturas_con_uuid.exclude(id=factura_a_conservar.id)
        total_eliminadas += facturas_a_eliminar.count()

        # Eliminar las facturas duplicadas
        facturas_a_eliminar.delete()

        print(f"Eliminadas {facturas_a_eliminar.count()} facturas duplicadas con UUID: {factura_grupo['uuid']}")

    print(f"Total de facturas eliminadas: {total_eliminadas}")


def eliminar_facturas_duplicadas():
    # Buscar facturas duplicadas basadas en el UUID
    facturas_duplicadas = Factura.objects.values('uuid').annotate(uuid_count=Count('uuid')).filter(uuid_count__gt=1)

    # Contador de facturas eliminadas
    total_eliminadas = 0

    # Procesar cada grupo de facturas duplicadas
    for factura_grupo in facturas_duplicadas:
        # Buscar todas las facturas con el mismo UUID
        facturas_con_uuid = Factura.objects.filter(uuid=factura_grupo['uuid']).order_by('fecha_timbrado')

        # Mantener solo la primera factura (la más antigua, por ejemplo)
        factura_a_conservar = facturas_con_uuid.first()

        # Eliminar las demás facturas
        facturas_a_eliminar = facturas_con_uuid.exclude(id=factura_a_conservar.id)
        total_eliminadas += facturas_a_eliminar.count()

        # Eliminar las facturas duplicadas
        facturas_a_eliminar.delete()

        print(f"Eliminadas {facturas_a_eliminar.count()} facturas duplicadas con UUID: {factura_grupo['uuid']}")

    print(f"Total de facturas eliminadas: {total_eliminadas}")

def eliminar_facturas_por_id(uuid_especifico):
    # Filtrar todas las facturas que tienen el UUID específico
    facturas_con_uuid = Factura.objects.filter(uuid=uuid_especifico).order_by('fecha_timbrado')

    if facturas_con_uuid.count() > 1:
        # Mantener solo la primera factura (la más antigua, por ejemplo)
        factura_a_conservar = facturas_con_uuid.first()

        # Eliminar las demás facturas
        facturas_a_eliminar = facturas_con_uuid.exclude(id=factura_a_conservar.id)
        total_eliminadas = facturas_a_eliminar.count()

        # Eliminar las facturas duplicadas
        facturas_a_eliminar.delete()

        print(f"Eliminadas {total_eliminadas} facturas duplicadas con UUID: {uuid_especifico}")
    else:
        print(f"No hay facturas duplicadas para el UUID: {uuid_especifico}")

def eliminar_facturas_duplicadas_viaticos():
    # Buscar facturas duplicadas basadas en el UUID
    facturas_duplicadas = Viaticos_Factura.objects.values('uuid').annotate(uuid_count=Count('uuid')).filter(uuid_count__gt=1)

    # Contador de facturas eliminadas
    total_eliminadas = 0

    # Procesar cada grupo de facturas duplicadas
    for factura_grupo in facturas_duplicadas:
        # Buscar todas las facturas con el mismo UUID
        facturas_con_uuid = Viaticos_Factura.objects.filter(uuid=factura_grupo['uuid']).order_by('fecha_timbrado')

        # Mantener solo la primera factura (la más antigua, por ejemplo)
        factura_a_conservar = facturas_con_uuid.first()

        # Eliminar las demás facturas
        facturas_a_eliminar = facturas_con_uuid.exclude(id=factura_a_conservar.id)
        total_eliminadas += facturas_a_eliminar.count()

        # Eliminar las facturas duplicadas
        facturas_a_eliminar.delete()

        print(f"Eliminadas {facturas_a_eliminar.count()} facturas duplicadas con UUID: {factura_grupo['uuid']}")

    print(f"Total de facturas eliminadas: {total_eliminadas}")

def extraer_texto_de_pdf(pdf_path):
    with fitz.open(pdf_path) as doc:
        texto = ""
        for pagina in doc:
            texto += pagina.get_text() + "\n"
    return texto

def extraer_hora_operacion(texto):
    # Busca una hora en formato HH:mm:ss
    match = re.search(r'(?:Hora de captura en el canal|Hora):\s*(\d{2}:\d{2}:\d{2})', texto)
    if match:
        try:
            return datetime.strptime(match.group(1), "%H:%M:%S").time()
        except ValueError:
            return None
    return None

def actualizar_horas_faltantes():
    pagos = Pago.objects.filter(comprobante_pago__isnull=False)

    for pago in pagos:
        print(f"Procesando pago ID {pago.id}")
        try:
            texto = extraer_texto_de_pdf(pago.comprobante_pago.path)
            hora = extraer_hora_operacion(texto)
            if hora:
                pago.pagado_hora = hora
                pago.save()
                print(f"✓ Hora actualizada: {hora}")
            else:
                print("⚠️ Hora no encontrada.")
        except Exception as e:
            print(f"❌ Error con pago ID {pago.id}: {e}")