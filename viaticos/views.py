from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Max, Q
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.mail import EmailMessage, BadHeaderError
from smtplib import SMTPException
import socket
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse, FileResponse
from django.conf import settings

from user.models import Profile
from solicitudes.models import Proyecto, Subproyecto, Operacion
from dashboard.models import Inventario, Product
from requisiciones.views import get_image_base64
from tesoreria.models import Cuenta, Pago, Facturas
from .models import Solicitud_Viatico, Concepto_Viatico, Viaticos_Factura, Puntos_Intermedios
from .forms import Solicitud_ViaticoForm, Concepto_ViaticoForm, Pago_Viatico_Form, Viaticos_Factura_Form, Puntos_Intermedios_Form, UploadFileForm, Cancelacion_viatico_Form
from tesoreria.forms import Facturas_Viaticos_Form
#from tesoreria.views import eliminar_caracteres_invalidos, extraer_datos_del_xml
from gastos.models import Factura, ValeRosa
#from tesoreria.views import generar_cfdi
from .filters import Solicitud_Viatico_Filter
from user.decorators import perfil_seleccionado_required, tipo_usuario_requerido
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from api.models import TablaFestivos
from django.db.models import Count, Q, Case, When, Value, CharField
from django.http import HttpResponseRedirect


import io
import json
import os
import datetime as dt
import pytz
import qrcode
import re
from num2words import num2words
import tempfile
from datetime import date, datetime, timedelta

#PDF generator
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.colors import Color, black, blue, red, white
from reportlab.lib.units import cm
from reportlab.lib.pagesizes import letter, portrait
from reportlab.rl_config import defaultPageSize 
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Frame
from bs4 import BeautifulSoup
import zipfile
from django.urls import reverse

#Excel stuff
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill
from openpyxl.utils import get_column_letter
import xlsxwriter
from io import BytesIO
import xml.etree.ElementTree as ET

#Se duplican funciones porque me arroja una referencia circular
#######################################################################
# Create your views here.
def eliminar_caracteres_invalidos(archivo_xml):
    # Definir la expresión regular para encontrar caracteres inválidos
    regex = re.compile(r'[^\x09\x0A\x0D\x20-\uD7FF\uE000-\uFFFD\u10000-\u10FFFF]')

    # Leer el contenido del archivo XML
    xml_content = archivo_xml.read().decode('utf-8')

    if xml_content.startswith("o;?"):
        print('Detectado "o;?" en el inicio del XML')
        xml_content = xml_content[3:]

    # Eliminar caracteres inválidos según la expresión regular
    xml_content = regex.sub('', xml_content)

    # Volver a posicionar el puntero del archivo al principio
    archivo_xml.seek(0)

    # Guardar el contenido modificado en el archivo original
    archivo_xml.write(xml_content.encode('utf-8'))
    archivo_xml.truncate()  # Asegurarse de que no quede contenido sobrante

    print('Contenido corregido guardado exitosamente.')

    # Retornar el archivo con el contenido modificado
    return archivo_xml

def extraer_datos_del_xml(archivo_xml):
    try:
        # Parsear el archivo XML
        archivo_xml.seek(0)
        tree = ET.parse(archivo_xml)
        root = tree.getroot()
    except (ET.ParseError, FileNotFoundError) as e:
        print(f"Error al parsear el archivo XML: {e}")
        return None, None  # Si ocurre un error, devuelve None
    
    # Identificar la versión del XML y el espacio de nombres
    version = root.tag
    ns = {}
    if 'http://www.sat.gob.mx/cfd/3' in version:
        ns = {
            'cfdi': 'http://www.sat.gob.mx/cfd/3',
            'tfd': 'http://www.sat.gob.mx/TimbreFiscalDigital',
            'if': 'https://www.interfactura.com/Schemas/Documentos',
        }
    elif 'http://www.sat.gob.mx/cfd/4' in version:
        ns = {
            'cfdi': 'http://www.sat.gob.mx/cfd/4',
            'tfd': 'http://www.sat.gob.mx/TimbreFiscalDigital',
            'if': 'https://www.interfactura.com/Schemas/Documentos',
        }
    else:
        print(f"Versión del documento XML no reconocida")
        return None, None
    
    # Buscar el complemento donde se encuentra el UUID y la fecha de timbrado
    complemento = root.find('cfdi:Complemento', ns)
    if complemento is not None:
        timbre_fiscal = complemento.find('tfd:TimbreFiscalDigital', ns)
        if timbre_fiscal is not None:
            uuid = timbre_fiscal.get('UUID')
            fecha_timbrado = timbre_fiscal.get('FechaTimbrado') or root.get('Fecha')
            return uuid, fecha_timbrado  # Devolver UUID y fecha de timbrado
        else:
            print("Timbre Fiscal Digital no encontrado")
            return None, None
    else:
        print("Complemento no encontrado")
        return None, None
################################################################################################################################


# Create your views here.
@perfil_seleccionado_required
def solicitud_viatico(request):
    colaborador = Profile.objects.all()
    pk_perfil = request.session.get('selected_profile_id')
    usuario = colaborador.get(id = pk_perfil)
    proyectos = Proyecto.objects.filter(~Q(status_de_entrega__status = "INACTIVO"), activo=True, distrito = usuario.distritos)
    #subproyectos = Subproyecto.objects.all()
    viatico, created = Solicitud_Viatico.objects.get_or_create(complete= False, staff = usuario)
    colaboradores = colaborador.filter(distritos = usuario.distritos)
    puntos = Puntos_Intermedios.objects.filter(solicitud = viatico)
    error_messages = {}
    

    if usuario.distritos.nombre == "MATRIZ":
        superintendentes = colaborador.filter(tipo__subdirector = True, distritos = usuario.distritos, st_activo =True, sustituto__isnull = True)
    elif usuario.tipo.superintendente and not usuario.tipo.nombre == "Admin":
        superintendentes = colaborador.filter(tipo__superintendente = True, distritos = usuario.distritos, st_activo = True).exclude(tipo__nombre="Admin")
    else:
        superintendentes = colaborador.filter(tipo__superintendente = True, distritos = usuario.distritos, st_activo = True).exclude(tipo__nombre="Admin")

    max_folio = Solicitud_Viatico.objects.all().aggregate(Max('folio'))['folio__max']
    folio_probable = max_folio + 1

    proyectos_para_select2 = [
        {
            'id': item.id, 
            'text': str(item.nombre)
        } for item in proyectos
    ]

    superintendentes_para_select2 = [
        {
            'id': super.id, 
            'text': str(super.staff.staff.first_name) + (' ') + str(super.staff.staff.last_name)
        } for super in superintendentes
    ]

    colaboradores_para_select2 = [
        {
            'id': item.id,
            'text': str(item.staff.staff.first_name) + (' ') + str(item.staff.staff.last_name)
        } for item in colaboradores
    ]

    form = Solicitud_ViaticoForm(instance = viatico)
    form2 = Puntos_Intermedios_Form()

    if request.method =='POST':
        if "btn_agregar" in request.POST:
            form = Solicitud_ViaticoForm(request.POST, instance=viatico)
            max_folio = Solicitud_Viatico.objects.all().aggregate(Max('folio'))['folio__max']
            nuevo_folio = (max_folio or 0) + 1
            #abrev= usuario.distrito.abreviado
            
            if form.is_valid():
                transporte_seleccionado = request.POST.get('tipo')
                cuenta_alterna = request.POST.get('toggleBankFields', 'false')  # 'true' si está marcado, 'false' si no
                numero_emergencia = request.POST.get('toggleEmergencyPhone', 'false')  # 'true' si está marcado, 'false' si no
                # Convertir a booleanos
                cuenta_alterna_marcada = (cuenta_alterna == 'true')  # True si marcado, False si no
                numero_emergencia_marcado = (numero_emergencia == 'true')  # True si marcado, False si no
                if cuenta_alterna_marcada == False:
                    viatico.banco = None
                    viatico.cuenta_bancaria = None 
                    viatico.clabe = None
                if numero_emergencia_marcado == False:
                    viatico.phone = None
                comentario_transporte = viatico.transporte
                viatico = form.save(commit=False)
                viatico.complete = True
                viatico.created_at =  datetime.now()
                #viatico.created_at_time = datetime.now().time()
                viatico.staff =  usuario
                viatico.distrito = usuario.distritos
                viatico.folio = nuevo_folio
                viatico.transporte = transporte_seleccionado + '-'  + comentario_transporte
                if usuario.distritos.nombre == "MATRIZ":
                    viatico.gerente = viatico.superintendente
                else:
                    viatico.gerente = colaborador.get(tipo__gerente = True, distritos = usuario.distritos, st_activo = True)
                if not viatico.colaborador:
                    viatico.colaborador = usuario
                viatico.save()
                messages.success(request, f'La solicitud {viatico.folio} ha sido creada')
                return redirect('solicitudes-viaticos')
            else:
                for field, errors in form.errors.items():
                    error_messages[field] = errors.as_text()
        if "btn_punto" in request.POST:
            form2 = Puntos_Intermedios_Form(request.POST)
            #abrev= usuario.distrito.abreviado
            if form2.is_valid():
                punto = form2.save(commit=False)
                punto.solicitud = viatico
                punto.save()
                messages.success(request, f'El punto intermedio se ha agregado correctamente')
                return redirect('solicitud-viatico')
            else:
                for field, errors in form2.errors.items():
                    error_messages[field] = errors.as_text()


    context= {
        'proyectos_para_select2':proyectos_para_select2,
        'error_messages': error_messages,
        'superintendentes_para_select2':superintendentes_para_select2,
        'colaboradores_para_select2':colaboradores_para_select2,
        'form':form,
        'form2':form2,
        'puntos':puntos,
        'viatico':viatico,
        'folio_probable': folio_probable,
        #'superintendentes':superintendentes,
        #'proyectos':proyectos,
        #'subproyectos':subproyectos,
    }
    return render(request, 'viaticos/crear_viaticos.html', context)

def eliminar_punto(request):
    data= json.loads(request.body)
    id = data["id"]
    punto = Puntos_Intermedios.objects.get(id=id)
    punto.delete()
    response_data = {
            'action': 'Item was removed',
        }


    return JsonResponse(response_data)

@perfil_seleccionado_required
def viaticos_pendientes_autorizar(request):
    #Autoriza
    #obtengo el id de usuario, lo paso como argumento a id de profiles para obtener el objeto profile que coindice con ese usuario_id
    pk_perfil = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk_perfil)

    if perfil.sustituto:
        perfil = Profile.objects.filter(staff=perfil.staff, tipo=perfil.tipo, distritos=perfil.distritos).first()


    if perfil.distritos.nombre == "MATRIZ":
        viaticos = Solicitud_Viatico.objects.filter(complete=True, autorizar = None, distrito = perfil.distritos, superintendente = perfil).order_by('-folio')
    else:
        viaticos = Solicitud_Viatico.objects.filter(complete=True, autorizar = None, distrito = perfil.distritos).order_by('-folio')

    myfilter=Solicitud_Viatico_Filter(request.GET, queryset=viaticos)
    viaticos = myfilter.qs

    #Set up pagination
    p = Paginator(viaticos, 10)
    page = request.GET.get('page')
    ordenes_list = p.get_page(page)

    #if request.method =='POST' and 'btnExcel' in request.POST:

        #return convert_excel_solicitud_matriz(solicitudes)

    context= {
        'ordenes_list':ordenes_list,
        'myfilter':myfilter,
        }

    return render(request, 'viaticos/pendientes_autorizar_viaticos.html', context)

@perfil_seleccionado_required
def viaticos_pendientes_autorizar2(request):
    #obtengo el id de usuario, lo paso como argumento a id de profiles para obtener el objeto profile que coindice con ese usuario_id
    colaborador = Profile.objects.all()
    pk_perfil = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk_perfil)

    
    if perfil.sustituto:
        perfil = Profile.objects.filter(staff=perfil.staff, tipo=perfil.tipo, distritos=perfil.distritos).first()

    if perfil.distritos.nombre == "MATRIZ":
        if perfil.tipo.subdirector and perfil.tipo.dg:
            viaticos = Solicitud_Viatico.objects.filter(
                Q(superintendente = perfil) | Q(colaborador__tipo__subdirector = True), 
                complete=True, autorizar = True, 
                montos_asignados=True, autorizar2 = None, 
                distrito = perfil.distritos
                ).order_by('-folio')
        elif perfil.tipo.subdirector:
            viaticos = Solicitud_Viatico.objects.filter(
                complete=True, 
                autorizar = True, 
                montos_asignados=True, 
                autorizar2 = None, 
                distrito = perfil.distritos, 
                superintendente = perfil
            ).exclude(
                Q(colaborador=perfil) | Q(staff=perfil)
            ).order_by('-folio')
        else:
            viaticos = Solicitud_Viatico.objects.filter(complete=True, autorizar = True, montos_asignados=True, autorizar2 = None, distrito = perfil.distritos, superintendente = perfil).order_by('-folio')
    else:
        viaticos = Solicitud_Viatico.objects.filter(complete=True, autorizar = True, montos_asignados=True, autorizar2 = None, distrito = perfil.distritos).order_by('-folio')

    myfilter=Solicitud_Viatico_Filter(request.GET, queryset=viaticos)
    viaticos = myfilter.qs

    #Set up pagination
    p = Paginator(viaticos, 10)
    page = request.GET.get('page')
    ordenes_list = p.get_page(page)

    #if request.method =='POST' and 'btnExcel' in request.POST:

        #return convert_excel_solicitud_matriz(solicitudes)

    context= {
        'ordenes_list':ordenes_list,
        'myfilter':myfilter,
        }

    return render(request, 'viaticos/pendientes_autorizar_viaticos2.html', context)


@perfil_seleccionado_required
def detalles_viaticos(request, pk):
    viatico = Solicitud_Viatico.objects.get(id=pk)

    context= {
        'viatico': viatico,
        }

    return render(request, 'viaticos/detalles_viaticos.html', context)

@perfil_seleccionado_required
def detalles_viaticos2(request, pk):
    viatico = Solicitud_Viatico.objects.get(id=pk)
    conceptos = Concepto_Viatico.objects.filter(viatico = viatico, completo = True)

    context= {
        'viatico': viatico,
        'conceptos':conceptos,
        }

    return render(request, 'viaticos/detalles_viaticos_montos.html', context)

@perfil_seleccionado_required
def detalles_viaticos3(request, pk):
    viatico = Solicitud_Viatico.objects.get(id=pk)
    conceptos = Concepto_Viatico.objects.filter(viatico = viatico, completo = True)

    context= {
        'viatico': viatico,
        'conceptos':conceptos,
        }

    return render(request, 'viaticos/detalles_viaticos_montos2.html', context)

@perfil_seleccionado_required
def autorizar_viaticos(request, pk):
    colaborador = Profile.objects.all()
    pk_perfil = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk_perfil)
    viatico = Solicitud_Viatico.objects.get(id = pk)

    if request.method =='POST' and 'btn_autorizar' in request.POST:
        viatico.autorizar = True
        viatico.approved_at = datetime.now()
        #viatico.approved_at_time = datetime.now().time()
        viatico.save()
        messages.success(request, f'{perfil.staff.staff.first_name} {perfil.staff.staff.last_name} has autorizado la solicitud {viatico.folio}')
        return redirect ('viaticos-pendientes-autorizar')


    context = {
        'viatico': viatico,
    }

    return render(request,'viaticos/autorizar_viaticos.html', context)

@perfil_seleccionado_required
def autorizar_viaticos2(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk_perfil)
    viatico = Solicitud_Viatico.objects.get(id = pk)
    conceptos = Concepto_Viatico.objects.filter(viatico = viatico, completo = True)

    if request.method =='POST' and 'btn_autorizar' in request.POST:
        viatico.autorizar2 = True
        viatico.approved_at2 = datetime.now()
        #viatico.approved_at_time2 = datetime.now().time()
        viatico.gerente = perfil
        viatico.save()
        messages.success(request, f'{perfil.staff.staff.first_name} {perfil.staff.staff.last_name} has autorizado la solicitud {viatico.id}')
        return redirect ('viaticos-pendientes-autorizar2')


    context = {
        'viatico': viatico,
        'conceptos': conceptos,
    }

    return render(request,'viaticos/autorizar_viaticos2.html', context)


@perfil_seleccionado_required
def cancelar_viaticos(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk_perfil)
    viatico = Solicitud_Viatico.objects.get(id = pk)
    form = Cancelacion_viatico_Form(instance= viatico)


    if request.method =='POST':
        form =  Cancelacion_viatico_Form(request.POST, instance = viatico)
        if form.is_valid():
            viatico = form.save(commit = False)
            viatico.autorizar = False
            viatico.approved_at = datetime.now()
            #viatico.approved_at_time = datetime.now().time()
            viatico.superintendente = perfil
            viatico.save()
            messages.info(request, f'{perfil.staff.staff.first_name} {perfil.staff.staff.last_name} has cancelado la solicitud {viatico.folio}')
            return HttpResponse(status=204)

    context = {
        'form':form,
        'viatico': viatico,
    }

    return render(request,'viaticos/cancelar_viaticos.html', context)

@perfil_seleccionado_required
def cancelar_viaticos2(request, pk):
    colaborador = Profile.objects.all()
    pk_perfil = request.session.get('selected_profile_id')
    perfil = colaborador.get(id = pk_perfil)
    viatico = Solicitud_Viatico.objects.get(id = pk)
    conceptos = Concepto_Viatico.objects.filter(viatico = viatico, completo = True)
    form = Cancelacion_viatico_Form(instance= viatico)

    if request.method =='POST':
        form =  Cancelacion_viatico_Form(request.POST, instance = viatico)
        if form.is_valid():
            viatico = form.save(commit = False)
            viatico.autorizar2 = False
            viatico.approbado_fecha2 = datetime.now()
            #viatico.approved_at_time2 = datetime.now().time()
            viatico.gerente = perfil
            viatico.save()
            messages.info(request, f'{perfil.staff.staff.first_name} {perfil.staff.staff.last_name} has cancelado la solicitud de viático {viatico.folio}')
            return HttpResponse(status=204)

    context = {
        'form': form,
        'viatico': viatico,
        'conceptos': conceptos,
    }


    return render(request,'viaticos/cancelar_viaticos2.html', context)

@perfil_seleccionado_required
def solicitudes_viaticos(request):
    #obtengo el id de usuario, lo paso como argumento a id de profiles para obtener el objeto profile que coindice con ese usuario_id
    colaborador = Profile.objects.all()
    pk_perfil = request.session.get('selected_profile_id')
    perfil = colaborador.get(id = pk_perfil)

    if perfil.tipo.nombre == "Admin" or perfil.tipo.nombre == "Control" or perfil.tipo.nombre == "Gerente" or perfil.tipo.superintendente == True:
        viaticos = Solicitud_Viatico.objects.filter(complete=True, distrito = perfil.distritos).order_by('-folio')
    else:
        viaticos = Solicitud_Viatico.objects.filter(complete=True, staff = perfil).order_by('-folio')

    myfilter=Solicitud_Viatico_Filter(request.GET, queryset=viaticos)
    viaticos = myfilter.qs

    #Set up pagination
    p = Paginator(viaticos, 10)
    page = request.GET.get('page')
    ordenes_list = p.get_page(page)

    if request.method =='POST' and 'btnExcel' in request.POST:

        return convert_excel_viatico(viaticos)

    context= {
        'ordenes_list':ordenes_list,
        'myfilter':myfilter,
        }

    return render(request, 'viaticos/solicitudes_viaticos.html', context)


@perfil_seleccionado_required
@tipo_usuario_requerido('tesoreria')
def viaticos_autorizados(request):

    #obtengo el id de usuario, lo paso como argumento a id de profiles para obtener el objeto profile que coindice con ese usuario_id
    colaborador = Profile.objects.all()
    pk_perfil = request.session.get('selected_profile_id')
    perfil = colaborador.get(id = pk_perfil)
    #Este es un filtro por perfil supervisor o superintendente, es decir puede ver todo lo del distrito
    #if perfil.tipo.superintendente == True:
    #    solicitudes = Solicitud_viatico.objects.filter(complete=True, staff__distrito=perfil.distrito).order_by('-folio')
    #elif perfil.tipo.supervisor == True:
    #    solicitudes = Solicitud_viatico.objects.filter(complete=True, staff__distrito=perfil.distrito, supervisor=perfil).order_by('-folio')
    #else:
    viaticos = Solicitud_Viatico.objects.filter(complete=True, distrito = perfil.distritos, autorizar = True, montos_asignados = False).order_by('-folio')

    myfilter=Solicitud_Viatico_Filter(request.GET, queryset=viaticos)
    viaticos = myfilter.qs

    #Set up pagination
    p = Paginator(viaticos, 10)
    page = request.GET.get('page')
    ordenes_list = p.get_page(page)

    #if request.method =='POST' and 'btnExcel' in request.POST:

        #return convert_excel_solicitud_matriz(solicitudes)

    context= {
        'ordenes_list':ordenes_list,
        'myfilter':myfilter,
        }

    return render(request, 'viaticos/viaticos_autorizados.html', context)

@perfil_seleccionado_required
@tipo_usuario_requerido('tesoreria')
def asignar_montos(request, pk):
    colaborador = Profile.objects.all()
    pk_perfil = request.session.get('selected_profile_id')
    usuario = colaborador.get(id = pk_perfil)
    viatico = Solicitud_Viatico.objects.get(id = pk)
    viatico_query= Solicitud_Viatico.objects.filter(id = pk)
    concepto, created = Concepto_Viatico.objects.get_or_create(completo = False, staff=usuario)

    conceptos = Concepto_Viatico.objects.filter(viatico = viatico, completo = True)
    error_messages = {}

    concepto_viatico = Product.objects.filter(viatico = True)

    form = Concepto_ViaticoForm()
    form.fields['producto'].queryset = concepto_viatico

    if request.method =="POST":
        if "btn_producto" in request.POST:
            form = Concepto_ViaticoForm(request.POST, instance=concepto)
            if form.is_valid():
                concepto = form.save(commit=False)
                concepto.viatico = viatico
                concepto.completo = True
                concepto.save()
                messages.success(request,'Se ha agregado un concepto de viático con éxito')
                return redirect('asignar-montos', pk=viatico.id)
            else:
                for field, errors in form.errors.items():
                    error_messages[field] = errors.as_text()
                form.fields['producto'].queryset = concepto_viatico
                form.fields['viatico'].queryset = viatico_query
        if "btn_asignar" in request.POST:
            conceptos = concepto_viatico.count()
            if conceptos > 0:
                viatico.montos_asignados = True
                viatico.save()
                messages.success(request,'Has agregado montos al viático con éxito')
                return redirect('viaticos_autorizados')
            else:
                messages.error(request,'No tienes conceptos agregados')



    context= {
        'error_messages':error_messages,
        'viatico':viatico,
        'conceptos':conceptos,
        'form':form,
    }

    return render(request, 'viaticos/asignar_montos.html', context)

@perfil_seleccionado_required
def delete_viatico(request, pk):
    concepto = Concepto_Viatico.objects.get(id=pk)
    messages.success(request,f'El articulo {concepto.producto} ha sido eliminado exitosamente')
    concepto.delete()

    return redirect('asignar-montos', pk=concepto.viatico.id)

@perfil_seleccionado_required
def viaticos_autorizados_pago(request):
    #obtengo el id de usuario, lo paso como argumento a id de profiles para obtener el objeto profile que coindice con ese usuario_id
    colaborador = Profile.objects.all()
    pk_perfil = request.session.get('selected_profile_id')
    perfil = colaborador.get(id = pk_perfil)

    #Este es un filtro por perfil supervisor o superintendente, es decir puede ver todo lo del distrito
    if perfil.tipo.tesoreria:
        if perfil.tipo.rh:
            viaticos = Solicitud_Viatico.objects.none()
        else:
            viaticos = Solicitud_Viatico.objects.filter(complete=True,distrito=perfil.distritos,autorizar=True,autorizar2=True,pagada=False, cerrar_sin_pago_completo = False).annotate(
                total_facturas=Count('facturas', filter=Q(facturas__hecho=True)),autorizadas=Count(Case(When(Q(facturas__autorizada=True, facturas__hecho=True), then=Value(1))))
                ).order_by('-folio')

    myfilter=Solicitud_Viatico_Filter(request.GET, queryset=viaticos)
    viaticos = myfilter.qs

    #Set up pagination
    p = Paginator(viaticos, 10)
    page = request.GET.get('page')
    viaticos_list = p.get_page(page)

    #if request.method =='POST' and 'btnExcel' in request.POST:

        #return convert_excel_solicitud_matriz(solicitudes)
    # Calcular el estado de las facturas basado en los conteos
    for viatico in viaticos_list:
        if viatico.total_facturas == 0:
            viatico.estado_facturas = 'sin_facturas'
        elif viatico.autorizadas == viatico.total_facturas:
            viatico.estado_facturas = 'todas_autorizadas'
        else:
            viatico.estado_facturas = 'pendientes'
    context= {
        'viaticos_list':viaticos_list,
        'myfilter':myfilter,
        }

    return render(request, 'viaticos/viaticos_autorizados_pago.html', context)

@perfil_seleccionado_required
def viaticos_pagos(request, pk):
    colaborador = Profile.objects.all()
    pk_perfil = request.session.get('selected_profile_id')
    usuario = colaborador.get(id = pk_perfil)
    viatico = Solicitud_Viatico.objects.get(id=pk)
    conceptos = Concepto_Viatico.objects.filter(viatico=viatico)
    pagos = Pago.objects.filter(viatico=viatico, hecho=True)
    cuentas = Cuenta.objects.filter(moneda__nombre = 'PESOS')
    pago, created = Pago.objects.get_or_create(tesorero = usuario, viatico__distrito = usuario.distritos, hecho=False, viatico=viatico)
    form = Pago_Viatico_Form()
    remanente = viatico.get_total - viatico.monto_pagado
    error_messages = []
#'text': str(super.staff.staff.first_name) + (' ') + str(super.staff.staff.last_name)
    cuentas_para_select2 = [
        {'id': cuenta.id,
         'text': str(cuenta.cuenta) + (' ') + str(cuenta.moneda), 
        } for cuenta in cuentas]

    if request.method == 'POST':
        if "myBtn" in request.POST:
            form = Pago_Viatico_Form(request.POST or None, request.FILES or None, instance = pago)

            if form.is_valid():
                print('Formulario válido')
                pago = form.save(commit = False)
                #pago.viatico = viatico
                pago.pagado_date = datetime.now()
                #pago.pagado_hora = datetime.now().time()
                pago.hecho = True
                total_pagado = round(viatico.monto_pagado  + pago.monto, 2)
                total_sol = round(viatico.get_total,2)
                if total_sol == total_pagado:
                    flag = True
                    print('flag')
                else:
                    flag = False
                    print('No flag')
                if total_pagado > viatico.get_total:
                    print('Entonces que')
                    messages.error(request,f'{usuario.staff.staff.first_name}, el monto introducido más los pagos anteriores superan el monto total del viático')
                else:
                    print('Donde escapa')
                    if flag:
                        viatico.pagada = True
                        viatico.save()
                    pago.save()
                    pagos = Pago.objects.filter(viatico=viatico, hecho=True)
                    static_path = settings.STATIC_ROOT
                    img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
                    img_path2 = os.path.join(static_path,'images','logo_vordcab.jpg')
            
                    image_base64 = get_image_base64(img_path)
                    logo_v_base64 = get_image_base64(img_path2)
                    # Crear el mensaje HTML
                    html_message = f"""
                    <html>
                        <head>
                            <meta charset="UTF-8">
                        </head>
                        <body>
                            <p><img src="data:image/jpeg;base64,{logo_v_base64}" alt="Imagen" style="width:100px;height:auto;"/></p>
                            <p>Estimado {viatico.staff.staff.staff.first_name} {viatico.staff.staff.staff.last_name},</p>
                            <p>Estás recibiendo este correo porque el viático solicitado: {viatico.folio} ha sido pagado,</p>
                            <p>por {pago.tesorero.staff.staff.first_name} {pago.tesorero.staff.staff.last_name}.</p>
                            <p>Buen viaje!</p>
                            <p><img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width:50px;height:auto;border-radius:50%"/></p>
                            <p>Este mensaje ha sido automáticamente generado por SAVIA 2.0</p>
                        </body>
                    </html>
                    """
                    archivo_viatico = attach_viatico_pdf(request, viatico.id)
                    try:
                        email = EmailMessage(
                            f'Viatico Autorizado {viatico.folio}',
                            body=html_message,
                            from_email = settings.DEFAULT_FROM_EMAIL,
                            to = ['ulises_huesc@hotmail.com',viatico.staff.staff.staff.email],
                            headers={'Content-Type': 'text/html'}
                            )
                        #if pagos.count() > 0:
                        #for pago in pagos:
                            #email.attach(f'Pago_folio_{pago.id}.pdf',pago.comprobante_pago.path,'application/pdf')
                        
                        email.content_subtype = "html " # Importante para que se interprete como HTML
                        email.attach(f'folio:{viatico.folio}.pdf',archivo_viatico,'application/pdf')
                        email.attach('Pago.pdf',pago.comprobante_pago.read(),'application/pdf')
                        email.send()
                        messages.success(request,f'Gracias por registrar tu pago, {usuario.staff.staff.first_name}')
                    except (BadHeaderError, SMTPException, socket.gaierror) as e:
                        error_message = f'{usuario.staff.staff.first_name}, Has generado el pago correctamente pero el correo de notificación no ha sido enviado debido a un error: {e}'
                        messages.success(request, error_message)
                    #Este código es para enviar correo informativo a cada uno de los RH's del distrito del usuario
                    personal_rh = colaborador.filter(distritos = viatico.staff.distritos, tipo__rh =True)
                    for persona in personal_rh:
                        html_message = f"""
                        <html>
                            <head>
                                <meta charset="UTF-8">
                            </head>
                            <body>
                                <p><img src="data:image/jpeg;base64,{logo_v_base64}" alt="Imagen" style="width:100px;height:auto;"/></p>
                                <p>Estimado {persona.staff.staff.first_name} {persona.staff.staff.last_name},</p>
                                <p>Para notificarte que el viático: {viatico.folio} ha sido pagado y se considere para los efectos y fines que para el departamento de RH sean aplicables</p>
                                <p>Tesorero que paga:{pago.tesorero.staff.staff.first_name} {pago.tesorero.staff.staff.last_name}.</p>
                                <p><img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width:50px;height:auto;border-radius:50%"/></p>
                                <p>Este mensaje ha sido automáticamente generado por SAVIA 2.0</p>
                            </body>
                        </html>
                        """
                        try:
                            email = EmailMessage(
                                f'Viatico Autorizado {viatico.folio} |Correo informativo para RH',
                                body=html_message,
                                from_email = settings.DEFAULT_FROM_EMAIL,
                                to = ['ulises_huesc@hotmail.com',persona.staff.staff.email],
                                headers={'Content-Type': 'text/html'}
                                )
                            #if pagos.count() > 0:
                            #for pago in pagos:
                                #email.attach(f'Pago_folio_{pago.id}.pdf',pago.comprobante_pago.path,'application/pdf')
                            email.content_subtype = "html " # Importante para que se interprete como HTML
                            email.attach(f'folio:{viatico.folio}.pdf',archivo_viatico,'application/pdf')
                            email.send()    
                        except (BadHeaderError, SMTPException, socket.gaierror) as e:
                            error_message = f'{usuario.staff.staff.first_name}, Has generado el pago correctamente pero el correo de notificación no ha sido enviado debido a un error: {e}'
                            messages.success(request, error_message)
                    return redirect('viaticos-autorizados-pago')
        if "cerrar_sin_pago" in request.POST:
            viatico.comentario_cierre = request.POST.get('comentario_cierre')
            viatico.cerrar_sin_pago_completo = True
            viatico.fecha_cierre = date.today()
            viatico.persona_cierre = usuario  # Asegúrate de tener esta variable ya disponible en tu vista
            viatico.save()
            messages.success(request, f'Viatico {viatico.folio} cerrada sin pago completo.')
            return redirect('pago-gastos-autorizados')
        else:
            print('No está entrando')
            for field, errors in form.errors.items():
                error_messages.append(f"{field}: {errors.as_text()}")
            
            print('Error messages:', error_messages)

    context= {
        'viatico':viatico,
        'pago':pago,
        'form':form,
        'conceptos': conceptos,
        'pagos':pagos,
        'cuentas_para_select2':cuentas_para_select2,
        'error_messages': error_messages,
        'remanente':remanente,
    }

    return render(request,'viaticos/viaticos_pagos.html',context)



@perfil_seleccionado_required
def facturas_viaticos(request, pk):
    colaborador = Profile.objects.all()
    pk_perfil = request.session.get('selected_profile_id')
    usuario = colaborador.get(id = pk_perfil)

    concepto = Concepto_Viatico.objects.get(id = pk)
    viatico = Solicitud_Viatico.objects.get(id = concepto.viatico.id)
    facturas = Viaticos_Factura.objects.filter(solicitud_viatico = viatico, hecho=True)
    factura, created = Viaticos_Factura.objects.get_or_create(solicitud_viatico = viatico, hecho=False)

    form = Viaticos_Factura_Form()

    if request.method == 'POST':
        if "btn_factura" in request.POST:
            form = Viaticos_Factura_Form(request.POST or None, request.FILES or None, instance = factura)
            if form.is_valid():
                factura = form.save(commit = False)
                factura.fecha_subido = date.today()
                #factura.hora_subido = datetime.now().time()
                factura.hecho = True
                factura.subido_por = usuario

                factura.save()
                messages.success(request,'Haz registrado tu factura')
                return redirect('facturas-viaticos', pk= concepto.id) #No content to render nothing and send a "signal" to javascript in order to close window
            else:
                messages.error(request,'No está validando')


    context={
        'concepto':concepto,
        'form':form,
        'facturas':facturas,
        'viatico':viatico,
        }

    return render(request, 'viaticos/matriz_facturas.html', context)

def guardar_factura(factura, archivo_procesado, nombre_archivo, uuid_extraido, fecha_timbrado_extraida, usuario):
    factura.factura_xml.save(nombre_archivo, archivo_procesado)
    factura.uuid = uuid_extraido
    factura.fecha_timbrado = fecha_timbrado_extraida
    factura.hecho = True
    factura.fecha_subido = datetime.now()
    factura.subido_por = usuario
    factura.save()

@perfil_seleccionado_required
def factura_nueva_viatico(request, pk):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    viatico = Solicitud_Viatico.objects.get(id = pk)
    
    #facturas = Facturas.objects.filter(pago = pago, hecho=True)
    #factura, created = Viaticos_Factura.objects.get_or_create(solicitud_viatico=viatico, hecho=False)
    

    #form = Viaticos_Factura_Form(instance=factura)
    form = UploadFileForm()

    if request.method == 'POST':
        if 'btn_registrar' in request.POST:
            form = UploadFileForm(request.POST, request.FILES or None)
            if form.is_valid():
                archivos_pdf = request.FILES.getlist('factura_pdf')
                archivos_xml = request.FILES.getlist('factura_xml')
                if not archivos_pdf and not archivos_xml:
                    messages.error(request, 'Debes subir al menos un archivo PDF o XML.')
                    return HttpResponse(status=204)
                
                # Iterar sobre el número máximo de archivos en cualquiera de las listas
                max_len = max(len(archivos_pdf), len(archivos_xml))
                facturas_registradas = []
                facturas_duplicadas = []

                for i in range(max_len):
                    archivo_pdf = archivos_pdf[i] if i < len(archivos_pdf) else None
                    archivo_xml = archivos_xml[i] if i < len(archivos_xml) else None
                    factura, created = Viaticos_Factura.objects.get_or_create(solicitud_viatico=viatico, hecho=False)
                    if archivo_xml:
                        archivo_procesado = eliminar_caracteres_invalidos(archivo_xml)

                        # Guardar temporalmente para extraer datos
                        #factura_temp = Viaticos_Factura(factura_xml=archivo_xml)
                        #factura_temp.factura_xml.save(archivo_xml.name, archivo_procesado, save=False)
                        #print(factura_temp)
                        uuid_extraido, fecha_timbrado_extraida = extraer_datos_del_xml(archivo_xml)

                        # Verificar si ya existe una factura con el mismo UUID y fecha de timbrado en cualquiera de las tablas
                        factura_existente = Factura.objects.filter(uuid=uuid_extraido, fecha_timbrado=fecha_timbrado_extraida).first()
                        facturas_existentes = Facturas.objects.filter(uuid=uuid_extraido, fecha_timbrado=fecha_timbrado_extraida).first()
                        viaticos_factura_existente = Viaticos_Factura.objects.filter(uuid=uuid_extraido, fecha_timbrado=fecha_timbrado_extraida).first()

                        if factura_existente or facturas_existentes or viaticos_factura_existente:
                            # Si una factura existente se encuentra, verificamos si su solicitud no está aprobada
                            if factura_existente and (factura_existente.solicitud_gasto.autorizar is False or factura_existente.solicitud_gasto.autorizar2 is False):
                                factura_existente.delete()
                                guardar_factura(factura, archivo_xml, uuid_extraido, fecha_timbrado_extraida, usuario)

                            elif facturas_existentes and (facturas_existentes.oc.autorizado1 is False or facturas_existentes.oc.autorizado2 is False):
                                facturas_existentes.delete()
                                guardar_factura(factura, archivo_xml, uuid_extraido, fecha_timbrado_extraida, usuario)

                            elif viaticos_factura_existente and (viaticos_factura_existente.solicitud_viatico.autorizar is False or viaticos_factura_existente.solicitud_viatico.autorizar2 is False):
                                viaticos_factura_existente.delete()
                                guardar_factura(factura, archivo_xml, uuid_extraido, fecha_timbrado_extraida, usuario)

                            else:
                                # Si no cumple las condiciones de eliminación, consideramos la factura duplicada
                                facturas_duplicadas.append(uuid_extraido)
                                continue  # Saltar al siguiente archivo si se encuentra duplicado
                        else:
                            # Si no existe ninguna factura, guardar la nueva
                            guardar_factura(factura, archivo_procesado, archivo_xml.name, uuid_extraido, fecha_timbrado_extraida, usuario)
                            #messages.success(request, 'Las facturas se registraron de manera exitosa')
                    if archivo_pdf:
                        factura.factura_pdf = archivo_pdf
                        factura.hecho = True
                        factura.fecha_subido = datetime.now()
                        factura.subido_por = usuario
                        factura.save()
                      
                        facturas_registradas.append(uuid_extraido if archivo_xml else f"Factura PDF {archivo_pdf.name}")
                    #messages.success(request, 'Los facturas se registraron de manera exitosa')
                     # Mensajes de éxito o duplicados
                #return HttpResponse(status=204)
                if facturas_registradas:
                    messages.success(request, f'Se han registrado las siguientes facturas: {", ".join(facturas_registradas)}')
                if facturas_duplicadas:
                    messages.warning(request, f'Las siguientes no se pudieron subir porque ya estaban registradas: {", ".join(facturas_duplicadas)}') 
            else:
                messages.error(request,'No se pudo subir tu documento')


    context={
        'form': form, 
        'viatico':viatico,
    }

    return render(request, 'viaticos/registrar_nueva_factura_viatico.html', context)


def generar_archivo_zip(facturas, viatico):
    nombre = viatico.folio if viatico.folio else ''
    zip_filename = f'facturas_compraviatico-{nombre}.zip'
    
    # Crear un archivo zip en memoria
    in_memory_zip = io.BytesIO()

    with zipfile.ZipFile(in_memory_zip, 'w') as zip_file:
        for factura in facturas:
            if factura.factura_pdf:
                pdf_path = factura.factura_pdf.path
                zip_file.write(pdf_path, os.path.basename(pdf_path))
            if factura.factura_xml:
                # Generar el PDFreader
                response = generar_cfdi_viaticos(None, factura.id)
                pdf_filename = f"{factura.id}.pdf" if factura.id else f"factura_{factura.id}.pdf"
                # Añadir el contenido del PDF al ZIP
                zip_file.writestr(pdf_filename, response.content)
                #Añadir el xml
                xml_path = factura.factura_xml.path
                zip_file.write(xml_path, os.path.basename(xml_path))

    # Resetear el puntero del archivo en memoria
    in_memory_zip.seek(0)

    return in_memory_zip, zip_filename

def calcular_fecha_limite(fecha_inicio, dias_habiles, festivos):
    fecha_actual = fecha_inicio
    dias_contados = 0

    while dias_contados < dias_habiles:
        fecha_actual += timedelta(days=1)

        # Verifica si es sábado o domingo
        if fecha_actual.weekday() in (5, 6):
            continue

        # Verifica si la fecha actual es un día festivo
        if fecha_actual in festivos:
            continue

        dias_contados += 1

    return fecha_actual

@perfil_seleccionado_required
def matriz_facturas_viaticos(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk_perfil)
    viatico = Solicitud_Viatico.objects.get(id = pk)
    concepto_viatico = Concepto_Viatico.objects.filter(viatico = viatico)
    #print(concepto_viatico)
    facturas = Viaticos_Factura.objects.filter(solicitud_viatico =viatico, hecho=True)
    form = Facturas_Viaticos_Form(instance=viatico)
    next_url = request.GET.get('next', 'mis-viaticos')
    # Obtener días festivos de la base de datos
    festivos = set(TablaFestivos.objects.values_list('dia_festivo', flat=True))

    # Calcular la fecha límite
    fecha_limite = calcular_fecha_limite(viatico.fecha_retorno, 10, festivos) #Se define en 10 díash habiles por intrucción del contador Heriberto 23/10/24

    #print('Fecha limite retorno:', viatico.fecha_retorno)
    print('Fecha limite para ingresar factura:', fecha_limite)
    #print('Fecha actual:', date.today())
    fuera_de_tiempo = date.today() > fecha_limite
    #print(fuera_de_tiempo)

    if request.method == 'POST':
        form = Facturas_Viaticos_Form(request.POST, instance=viatico)
        if "btn_factura_completa" in request.POST:
            # Procesar los checkboxes de las facturas
            #print("Contenido de request.POST:", request.POST)  # Para ver qué datos se están recibiendo
            fecha_hora = datetime.today()
            for factura in facturas:
                checkbox_name = f'autorizar_factura_{factura.id}'
                #print("Nombre del checkbox esperado:", checkbox_name)  # Imprimir el nombre esperado
                if checkbox_name in request.POST:
                    factura.autorizada = True
                    factura.autorizada_por = perfil
                    factura.autorizada_el = fecha_hora
                else:
                    factura.autorizada = False
                factura.save()
            if form.is_valid():
                form.save()
                messages.success(request,'Haz cambiado el status de facturas completas')
                return redirect(next_url)
            else:
                messages.error(request,'No está validando')
        elif "btn_descargar_todo" in request.POST:
            in_memory_zip, zip_filename = generar_archivo_zip(facturas, viatico)
            response = HttpResponse(in_memory_zip, content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
            return response
        elif 'salir' in request.POST:
            return redirect(next_url)

    context={
        'next_url':next_url,
        'facturas':facturas,
        'form':form,
        'conceptos_viatico': concepto_viatico,
        'viatico': viatico,
        'fuera_de_tiempo':fuera_de_tiempo,
        }

    return render(request, 'viaticos/matriz_facturas_viaticos.html', context)

@perfil_seleccionado_required
def eliminar_factura_viatico(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk_perfil)
    # Obtener la factura y el viático relacionado
    factura = Viaticos_Factura.objects.get(id=pk)
    viatico = factura.solicitud_viatico
    comentario = request.POST.get('comentario')

    # Obtener el parámetro `next` de la URL
    next_url = request.GET.get('next', None)

    # Construir la URL de la matriz de facturas de viáticos
    matriz_url = reverse('matriz-facturas-viaticos', args=[viatico.id])

    static_path = settings.STATIC_ROOT
    img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
    img_path2 = os.path.join(static_path,'images','logo_vordcab.jpg')
    image_base64 = get_image_base64(img_path)
    logo_v_base64 = get_image_base64(img_path2)
    # Crear el mensaje HTML
    html_message = f"""
    <html>
        <head>
            <meta charset="UTF-8">
        </head>
        <body style="font-family: Arial, sans-serif; color: #333; background-color: #f4f4f4; margin: 0; padding: 0;">
            <table width="100%" cellspacing="0" cellpadding="0" style="background-color: #f4f4f4; padding: 20px;">
                <tr>
                    <td align="center">
                        <table width="600px" cellspacing="0" cellpadding="0" style="background-color: #ffffff; padding: 20px; border-radius: 10px;">
                            <tr>
                                <td align="center">
                                    <img src="data:image/jpeg;base64,{logo_v_base64}" alt="Logo" style="width: 100px; height: auto;" />
                                </td>
                            </tr>
                            <tr>
                                <td style="padding: 20px;">
                                    <p style="font-size: 18px; text-align: justify;">
                                        <p>Estimado {factura.solicitud_viatico.staff.staff.staff.first_name} {factura.solicitud_viatico.staff.staff.staff.last_name},</p>
                                    </p>
                                    <p style="font-size: 16px; text-align: justify;">
                                        Estás recibiendo este correo porque tu factura subida el: <strong>{factura.fecha_subido.date()}</strong> en el viatico <strong>{viatico.folio}</strong> ha sido eliminada.</p>
                                    <p>Comentario:</p>
                                    {comentario}
                                    </p>
                                <p style="font-size: 16px; text-align: justify;">
                                    Att: {perfil.staff.staff.first_name} {perfil.staff.staff.last_name}
                                </p>
                                    <p style="text-align: center; margin: 20px 0;">
                                        <img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width: 50px; height: auto; border-radius: 50%;" />
                                    </p>
                                    <p style="font-size: 14px; color: #999; text-align: justify;">
                                        Este mensaje ha sido automáticamente generado por SAVIA 2.0
                                    </p>
                                </td>
                            </tr>
                        </table>
                    </td>
                </tr>
            </table>
        </body>
    </html>
    """
    try:
        email = EmailMessage(
            f'Factura eliminada',
            body=html_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[viatico.staff.staff.staff.email],
            headers={'Content-Type': 'text/html'}
            )
        email.content_subtype = "html " # Importante para que se interprete como HTML
        if factura.factura_pdf:
            pdf_path = factura.factura_pdf.path
            if os.path.exists(pdf_path):  # Verificar si el archivo realmente existe
                with open(pdf_path, 'rb') as pdf_file:
                    email.attach(factura.factura_pdf.name, pdf_file.read(), 'application/pdf')
            else:
                print(f"El archivo PDF no se encuentra en la ruta: {pdf_path}")

        if factura.factura_xml:
            xml_path = factura.factura_xml.path
            if os.path.exists(xml_path):  # Verificar si el archivo realmente existe
                with open(xml_path, 'rb') as xml_file:
                    email.attach(factura.factura_xml.name, xml_file.read(), 'application/xml')
            else:
                print(f"El archivo XML no se encuentra en la ruta: {xml_path}")

        email.send()
        messages.success(request, f'La factura {factura.id} ha sido eliminada exitosamente')
    except (BadHeaderError, SMTPException, socket.gaierror) as e:
        error_message = f'La factura {factura.id} ha sido eliminada, pero el correo no ha sido enviado debido a un error: {e}'
        messages.success(request, error_message)
    factura.delete()

    # Redirigir a 'matriz-facturas-viaticos' con el parámetro `next` si existe
    if next_url:
        return redirect(f'{matriz_url}?next={next_url}')
    else:
        return redirect(matriz_url)

@perfil_seleccionado_required
def factura_viatico_edicion(request, pk):
    usuario = Profile.objects.get(staff__id=request.user.id)
    factura = Viaticos_Factura.objects.get(id = pk)

    form = UploadFileForm()

    if request.method == 'POST':
        if 'btn_edicion' in request.POST:
            form = Viaticos_Factura_Form(request.POST or None, request.FILES or None, instance = factura)
            if form.is_valid():
                factura = form.save(commit = False)
                factura.subido_por = usuario
                factura.save()
                form.save()
                messages.success(request,'Las facturas se subieron de manera exitosa')
            else:
                messages.error(request,'No se pudo subir tu documento')


    context={
        'factura':factura,
        'form':form,
        }

    return render(request, 'viaticos/factura_viatico_edicion.html', context)



def render_pdf_viatico(request, pk):
    viatico = get_object_or_404(Solicitud_Viatico, id=pk)
    buf = generar_pdf_viatico(viatico.id)
    return FileResponse(buf, as_attachment=True, filename='V_' + str(viatico.folio) + '.pdf')

def attach_viatico_pdf(request, pk):
    viatico = get_object_or_404(Solicitud_Viatico, id=pk)
    buf = generar_pdf_viatico(viatico.id)

    return buf.getvalue()

def generar_pdf_viatico(pk):
    #Configuration of the PDF object
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    #Here ends conf.
    viatico = Solicitud_Viatico.objects.get(id=pk)
    conceptos = Concepto_Viatico.objects.filter(viatico = viatico)
    facturas = Viaticos_Factura.objects.filter(solicitud_viatico = viatico, hecho = True, autorizada=True)
    vales = ValeRosa.objects.filter(viatico = viatico)

    #Configuraciones por default 
    styles = getSampleStyleSheet()
    width, height = letter

   #Azul Vordcab
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)
    caja_iso = 760
    #Elaborar caja
    #c.line(caja_iso,500,caja_iso,720)



    #Encabezado
    c.drawString(420,caja_iso,'Preparado por:')
    c.drawString(420,caja_iso-10,'SUP. ADMON')
    c.drawString(520,caja_iso,'Aprobación')
    c.drawString(520,caja_iso-10,'SUB ADM')
    #c.drawString(150,caja_iso-20,'Número de documento')
    #c.drawString(160,caja_iso-30,'F-ADQ-N4-01.02')
    #c.drawString(245,caja_iso-20,'Clasificación del documento')
    #c.drawString(275,caja_iso-30,'Controlado')
    #c.drawString(355,caja_iso-20,'Nivel del documento')
    #c.drawString(380,caja_iso-30, 'N5')
    #c.drawString(440,caja_iso-20,'Revisión No.')
    #c.drawString(452,caja_iso-30,'000')
    #c.drawString(510,caja_iso-20,'Fecha de Emisión')
    #c.drawString(525,caja_iso-30,'01/2024')

    caja_proveedor = caja_iso - 50
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(150,750,250,20, fill=True, stroke=False) #Barra azul superior Solicitud
    c.rect(20,caja_proveedor - 8,565,20, fill=True, stroke=False) #Barra azul superior Proveedor | Detalle
    c.rect(20,460,565,2, fill=True, stroke=False) #Linea posterior horizontal 
    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(280,755,'Comprobación de Viáticos')
    c.setLineWidth(.3) #Grosor
    c.line(20,caja_proveedor-8,20,460) #Eje Y donde empieza, Eje X donde empieza, donde termina eje y,donde termina eje x (LINEA 1 contorno)
    c.line(585,caja_proveedor-8,585,460) #Linea 2 contorno
    c.drawInlineImage('static/images/logo_vordcab.jpg',45,730, 3 * cm, 1.5 * cm) #Imagen vortec

    c.setFillColor(white)
    c.setFont('Helvetica-Bold',11)
    #c.drawString(120,caja_proveedor,'Infor')
    c.drawString(300,caja_proveedor, 'Detalles')
    inicio_central = 300
    #c.line(inicio_central,caja_proveedor-25,inicio_central,520) #Linea Central de caja Proveedor | Detalle
    c.setFillColor(black)
    c.setFont('Helvetica',9)
    # Primera columna
    c.drawString(30,caja_proveedor-20,'Solicitó:')
    c.drawString(30,caja_proveedor-40,'Distrito:')
    c.drawString(30,caja_proveedor-60,'Proyecto:')
    c.drawString(30,caja_proveedor-80,'Subproyecto:')
    c.drawString(30,caja_proveedor-100,'Lugar de partida:')
    c.drawString(30,caja_proveedor-120,'Lugar de Comisión:')
    c.drawString(30,caja_proveedor-140,'Fecha de partida:')
    c.drawString(30,caja_proveedor-160,'Fecha de retorno:')
    c.drawString(30,caja_proveedor-180,'Comisión:')
    # Segunda columna del encabezado
    c.drawString(320,caja_proveedor-20,'Banco:')
    c.drawString(320,caja_proveedor-40,'Cuenta:')
    c.drawString(320,caja_proveedor-60,'Clabe:')
    c.drawString(320,caja_proveedor-80,'Colaborador:')
    c.drawString(320,caja_proveedor-100,'Nivel:')
    c.drawString(320, caja_proveedor-120,'Transporte:')
    c.drawString(320, caja_proveedor-140,'Hospedaje:')
    c.drawString(320, caja_proveedor-160,'Empresa:')

    c.drawString(320,caja_proveedor-200,'Fecha de Elaboración:')



    
    c.setFont('Helvetica-Bold',12)
    c.drawString(500,caja_proveedor-20,'FOLIO:')
    c.setFillColor(rojo)
    c.setFont('Helvetica-Bold',12)
    c.drawString(540,caja_proveedor-20, str(viatico.folio))

    c.setFillColor(black)
    c.setFont('Helvetica',9)
    c.drawString(120,caja_proveedor-20, viatico.staff.staff.staff.first_name+' '+ viatico.staff.staff.staff.last_name)
    c.drawString(120,caja_proveedor-40, viatico.staff.distritos.nombre)
    c.drawString(120,caja_proveedor-60, viatico.proyecto.nombre)
    c.drawString(120,caja_proveedor-80, viatico.subproyecto.nombre)
   
    c.drawString(120,caja_proveedor-100, viatico.lugar_partida)
    #c.drawString(120,caja_proveedor-120, viatico.lugar_comision)
    style = styles['Normal']
    y_inicial = caja_proveedor - 120
    paragraph = Paragraph(viatico.lugar_comision, style)
    paragraph.wrapOn(c, 200, height)
    longitud_umbral = 38
    num_lineas_esperadas = len(viatico.lugar_comision) / longitud_umbral
    if num_lineas_esperadas > 1:
        y_inicial -= 10
    paragraph.drawOn(c, 120, y_inicial)

    
    c.drawString(120,caja_proveedor-140, viatico.fecha_partida.strftime("%d/%m/%Y"))
    
    if viatico.fecha_retorno:
        c.drawString(120,caja_proveedor-160, viatico.fecha_retorno.strftime("%d/%m/%Y"))
    if viatico.motivo:
        #c.drawString(120,caja_proveedor-180, viatico.motivo)
        y_inicial = caja_proveedor - 180
        style = styles['Normal']
        paragraph = Paragraph(viatico.motivo, style)
        paragraph.wrapOn(c, 450, height)
        longitud_umbral = 75
        num_lineas_esperadas = len(viatico.motivo) / longitud_umbral
        if num_lineas_esperadas > 1:
            y_inicial -= 12
        paragraph.drawOn(c, 120, y_inicial)
    # Segunda Columna del encabezado
   
    if viatico.colaborador:
        c.drawString(380, caja_proveedor-80, viatico.colaborador.staff.staff.first_name+' '+viatico.colaborador.staff.staff.last_name)
        c.drawString(380, caja_proveedor-100, str(viatico.colaborador.staff.nivel))
        if viatico.colaborador.staff.banco:
            c.drawString(380,caja_proveedor-20, viatico.colaborador.staff.banco.nombre)
        else:
            c.drawString(380,caja_proveedor-20, "Sin registro")
        if viatico.colaborador.staff.cuenta_bancaria:
            c.drawString(380,caja_proveedor-40,viatico.colaborador.staff.cuenta_bancaria)
        else:
            c.drawString(380,caja_proveedor-40, "Sin registro")
        if viatico.colaborador.staff.clabe:
            c.drawString(380,caja_proveedor-60,viatico.colaborador.staff.clabe)
        else:
            c.drawString(380,caja_proveedor-60, "Sin registro")
        if viatico.colaborador.staff.empresa:
            c.drawString(380,caja_proveedor-160,viatico.colaborador.staff.empresa.nombre)
        else:
            c.drawString(380,caja_proveedor-160, "Sin registro")
    else:
        c.drawString(380, caja_proveedor-80,'Solicitante')
        c.drawString(380, caja_proveedor-100, viatico.staff.staff.staff.first_name+' '+viatico.staff.staff.staff.last_name)
        if viatico.staff.staff.banco:
            c.drawString(380,caja_proveedor-20, viatico.staff.staff.banco.nombre)
        else:
            c.drawString(380,caja_proveedor-20, "Sin registro")
        if viatico.staff.staff.cuenta_bancaria:
            c.drawString(380,caja_proveedor-40,viatico.staff.staff.cuenta_bancaria)
        else:
            c.drawString(380,caja_proveedor-40, "Sin registro")
        if viatico.staff.staff.clabe:
            c.drawString(380,caja_proveedor-60,viatico.staff.staff.clabe)
        else:
            c.drawString(380,caja_proveedor-60, "Sin registro")
        if viatico.staff.staff.empresa:
            c.drawString(380,caja_proveedor-160,viatico.staff.staff.empresa.nombre)
        else:
            c.drawString(380,caja_proveedor-160, "Sin registro")
   
    
    #c.drawString(380, caja_proveedor-120, str(viatico.transporte))
    style = styles['Normal']
    y_inicial = caja_proveedor - 120
    paragraph = Paragraph(viatico.transporte, style)
    paragraph.wrapOn(c, 200, height)
    longitud_umbral = 38
    num_lineas_esperadas = len(viatico.transporte) / longitud_umbral
    if num_lineas_esperadas > 1:
        y_inicial -= 10
    paragraph.drawOn(c, 380, y_inicial)
    


    if viatico.hospedaje:
        c.drawString(380, caja_proveedor-140, "Requerido")
    else:
        c.drawString(380, caja_proveedor-140, "No Requerido")
    if viatico.approved_at:
        c.drawString(430,caja_proveedor-200, viatico.approved_at.strftime("%d/%m/%Y"))
    


    #Create blank list
    data =[]

    data.append(['''Código''', '''Nombre''', '''Cantidad''','''Precio''', '''Subtotal''','''Comentario'''])


    high = 440
    total = 0

    for concepto in conceptos:
        # Convert to Decimal and round to two decimal places
        cantidad_redondeada = Decimal(concepto.cantidad).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        precio_unitario_redondeado = Decimal(concepto.precio).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        subtotal = Decimal(cantidad_redondeada * precio_unitario_redondeado).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        total += subtotal
        data.append([
            concepto.producto.codigo, 
            concepto.producto.nombre,
            cantidad_redondeada, 
            precio_unitario_redondeado,
            subtotal, 
            concepto.comentario,
            ])
        high = high - 18


    c.setFillColor(prussian_blue)
    c.rect(20,30,565,30, fill=True, stroke=False)
    c.setFillColor(white)
    #Primer renglón
    c.drawCentredString(70,48,'Clasificación:')
    c.drawCentredString(140,48,'Nivel:')
    c.drawCentredString(240,48,'Preparado por:')
    c.drawCentredString(350,48,'Aprobado:')
    c.drawCentredString(450,48,'Fecha emisión:')
    c.drawCentredString(550,48,'Rev:')
    #Segundo renglón
    c.drawCentredString(70,34,'Controlado')
    c.drawCentredString(140,34,'N5')
    c.drawCentredString(240,34,'SEOV-ALM-N4-01-01')
    c.drawCentredString(350,34,'SUB ADM')
    c.drawCentredString(450,34,'24/Oct/2018')
    c.drawCentredString(550,34,'001')

    c.setFillColor(black)
    
    
    styleN = styles["BodyText"]

    if viatico.comentario_general is not None:
        comentario = viatico.comentario_general
    else:
        comentario = "No hay comentarios"

    
   
    # Crear un marco (frame) en la posición específica
    frame = Frame(35, 0, width-40, high-65, id='normal')
    options_conditions_paragraph = Paragraph(comentario, styleN)
    # Agregar el párrafo al marco
    frame.addFromList([options_conditions_paragraph], c)
    c.setFillColor(prussian_blue)
    c.rect(20,30,500,30, fill=True, stroke=False)
    c.setFillColor(white)
    # Personalizar el estilo de los párrafos
    custom_style = ParagraphStyle(
    'CustomStyle',
        parent=styles['BodyText'],
        fontSize=6,  # Reducir el tamaño de la fuente a 6
        leading=8,   # Aumentar el espacio entre líneas para asegurar que el texto no se superponga
        alignment=TA_LEFT,  # Alineación del texto
        # Puedes añadir más ajustes si es necesario
    )
    for i, row in enumerate(data):
        for j, item in enumerate(row):
            if i!=0 and j == 6:
                data[i][j] = Paragraph(item, custom_style)

    table = Table(data, colWidths=[1.5 * cm, 6.5 * cm, 2 * cm, 2 * cm, 1.7 * cm, 6 * cm,])
    table_style = TableStyle([ #estilos de la tabla
        ('INNERGRID',(0,0),(-1,-1), 0.25, colors.white),
        ('BOX',(0,0),(-1,-1), 0.25, colors.black),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        #ENCABEZADO
        ('TEXTCOLOR',(0,0),(-1,0), white),
        ('FONTSIZE',(0,0),(-1,0), 10),
        ('BACKGROUND',(0,0),(-1,0), prussian_blue),
        #CUERPO
        ('TEXTCOLOR',(0,1),(-1,-1), colors.black),
        ('FONTSIZE',(0,1),(-1,-1), 6),
        ])
    table.setStyle(table_style)

    #pdf size
    table.wrapOn(c, width, height)
    table.drawOn(c, 20, high)
    # Crear una lista de datos para la tabla secundaria
    data_secundaria = []
    data_secundaria.append(['Proyecto', 'Subproyecto'])  # Encabezados de la tabla secundaria

    #c.setFillColor(prussian_blue)
    #c.drawString(290, high-20, 'Total:')
    #c.drawString(320, high-20, '$' + str(total))
    total_vales = 0
    for vale in vales:
        if vale.esta_aprobado:
            total_vales += float(vale.monto)

    total_facturas = 0
    suma_total = Decimal('0.00')
    data_facturas = [['Datos de XML', 'Nombre', 'Monto']]  # Encabezados de la tabla de facturas
    # Iterar sobre cada factura y sumar el total
    for factura in facturas:
        if factura.factura_xml:
            datos_emisor = factura.emisor  # Llamar a la propiedad 'emisor' (Esto devuelve el diccionario de la propiedad)
            # Acceder directamente a los datos de XML
            resultados = datos_emisor.get('resultados', [])
            nombre = datos_emisor.get('nombre', 'No disponible')
            total_xml_str = datos_emisor.get('total', '0.00')  # Obtener el total o usar '0.00' como predeterminado
            #Para el total de todas las factuas
            total = datos_emisor.get('total', 0.0)  # Obtener el total o usar 0.0 si no está disponible
            total_facturas += float(total)  # Sumar el total al total general
            try:
                total_factura = Decimal(total_xml_str)  # Convertir a Decimal
            except (InvalidOperation, ValueError):
                total_factura = Decimal('0.00')  # Si no es convertible, usar 0.00
            
            # Sumar al total acumulado
            suma_total += total_factura

            # Añadir los datos a la lista
            data_facturas.append([
                Paragraph(str(resultados), custom_style), 
                Paragraph(nombre, custom_style),
                Paragraph(f"${total_factura:,.2f}", custom_style)  # Formatear el total como una cadena de texto
            ])
    #Parrafó de totales
    data_totales = []
    total_comprobado = total_facturas + total_vales
    diferencia_totales = total_comprobado - float(viatico.get_total)
    if diferencia_totales > 0:
        color_diferencia = colors.green
    elif diferencia_totales < 0:
        color_diferencia = colors.red
    else:
        color_diferencia = colors.black 
    total_str = "${:,.2f}".format(total_facturas)  # Convierte Decimal a string y formatea
    # 4. Posición de la tabla de facturas en el PDF
    # Asumiendo que 'y_pos' es la posición Y después de dibujar la tabla secundaria y cualquier otro contenido
    
    data_totales = [
    ['Total solicitado', 'Total Facturado', 'Total No deducible', 'Total comprobado', 'Saldo A cargo/Favor en Pesos'],  # Encabezados
    ['$' + str(viatico.get_total), f"${total_facturas:,.2f}",f"{total_vales:,.2f}", f"{total_comprobado:,.2f}",Paragraph(f'${diferencia_totales:,.2f}', ParagraphStyle('CustomStyle', textColor=color_diferencia))]
    ]

    # Estilo para la tabla secundaria
    table_secundaria_style = TableStyle([
        ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
        ('BOX', (0,0), (-1,-1), 0.25, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BACKGROUND', (0,0), (1,0), colors.grey),  # Fondo gris para los encabezados
        ('TEXTCOLOR', (0,0), (1,0), colors.whitesmoke),  # Texto blanco para los encabezados
        ('TEXTCOLOR', (0,1), (-1,-1), colors.black),  # Texto negro para el cuerpo
        ('FONTSIZE', (0,0), (-1,-1), 8),  # Tamaño de fuente para toda la tabla
        # Añade aquí más estilos si lo necesitas
    ])

    table_totales = Table(data_totales, colWidths=[3 * cm, 3 * cm, 3* cm, 3*cm, 5 * cm])  # Ajusta las medidas según necesites
    table_totales.setStyle(table_secundaria_style)
    # Añadir filas de proyectos y subproyectos
   
    table_totales.wrapOn(c, width, height)
    y_totales_pos = high-10 - (len(data_totales) * 15 ) 
    table_totales.drawOn(c, 20, y_totales_pos)


    c.setFillColor(prussian_blue)
    c.rect(25,high-70,540,20, fill=True, stroke=False)
    c.setFillColor(white)
    c.drawCentredString(320,high-65,'Comentario General')
    c.setFillColor(black)
    c.drawCentredString(230,high-190, viatico.staff.staff.staff.first_name +' '+ viatico.staff.staff.staff.last_name)
    c.line(180,high-195,280,high-195)
    c.drawCentredString(230,high-205, 'Solicitado')
   
    c.setFillColor(black)
    c.drawCentredString(410,high-190, viatico.superintendente.staff.staff.first_name +' '+ viatico.superintendente.staff.staff.last_name)
    c.line(360,high-195,460,high-195)
    c.drawCentredString(410,high-205,'Aprobado por')
    
    

    c.showPage()
    c.save()
    buf.seek(0)

    return buf

def convert_excel_viatico(viaticos):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = viaticos_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='viaticos')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')
    percent_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(percent_style)

    columns = ['Folio','Fecha Autorización','Distrito','Colaborador','Solicitado para','Proyecto','Subproyecto',
               'Importe','Fecha Creación','Status','Autorizado por','Tiene_Facturas','Status de Pago']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16
        if col_num == 5: #Columna del proveedor
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        if col_num == 2:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 20

    columna_max = len(columns)+2

    # Agregar los mensajes
    ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por SAVIA 2.0. UH}').style = messages_style
    ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}').style = messages_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 30

    # Agregar los encabezados de las nuevas columnas debajo de los mensajes
    ws.cell(row=3, column = columna_max, value="Total de viaticos").style = head_style
    ws.cell(row=4, column = columna_max, value="Sumatoria de Viáticos").style = head_style
   

    # Asumiendo que las filas de datos comienzan en la fila 2 y terminan en row_num
    ws.cell(row=3, column=columna_max + 1, value=f"=COUNTA(A:A)-1").style = body_style
    ws.cell(row=4, column=columna_max + 1, value=f"=SUM(H:H)").style = money_resumen_style
  

   
    
    for viatico in viaticos:
        row_num = row_num + 1    
        
        # Manejar autorizado_at_2
        if viatico.approved_at2 and isinstance(viatico.approved_at2, datetime):
        # Si autorizado_at_2 es timezone-aware, conviértelo a timezone-naive
            autorizado_at_2_naive = viatico.approved_at2.astimezone(pytz.utc).replace(tzinfo=None)
        else:
            autorizado_at_2_naive = ''
        
        # Manejar created_at
        if viatico.created_at and isinstance(viatico.created_at, datetime):
        # Si created_at es timezone-aware, conviértelo a timezone-naive
           created_at_naive = viatico.created_at.astimezone(pytz.utc).replace(tzinfo=None)
        else:
            created_at_naive = ''

        
        if viatico.pagada:
            pagada = "Tiene Pago"
        else: 
            pagada ="No tiene pago"

        if viatico.facturas.filter(factura_xml__isnull=False).exists():
            tiene_facturas = 'Sí'
        else:
            tiene_facturas = 'No'
        
        if viatico.autorizar2:
            status = "Autorizado"
            
            if viatico.distrito.nombre == "MATRIZ":
                if viatico.superintendente:
                    autorizado_por = str(viatico.superintendente.staff.staff.first_name) + ' ' +str(viatico.superintendente.staff.staff.last_name)
                else:
                    autorizado_por = "NR"
            else:
                if viatico.gerente:
                    autorizado_por = str(viatico.gerente.staff.staff.first_name) + ' ' + str(viatico.gerente.staff.staff.last_name)
                else:
                    autorizado_por = "NR"
        elif viatico.autorizar2 == False:
            status = "Cancelado"
            if viatico.distrito.nombre == "MATRIZ":
                if viatico.superintendente:
                    autorizado_por = str(viatico.superintendente.staff.staff.first_name) + ' ' +str(viatico.superintendente.staff.staff.last_name)
                else:
                    autorizado_por = "NR"
            else:
                autorizado_por =   str(viatico.gerente.staff.staff.first_name) + ' ' + str(viatico.gerente.staff.staff.last_name)
        elif viatico.autorizar:
            autorizado_por =str(viatico.superintendente.staff.staff.first_name) + ' ' + str(viatico.superintendente.staff.staff.last_name)
            status = "Autorizado | Falta una autorización"
        elif viatico.autorizar == False:
            status = "Cancelado"
            if viatico.superintendente:
                autorizado_por = str(viatico.superintendente.staff.staff.first_name) + ' ' +str(viatico.superintendente.staff.staff.last_name)
            else:
                autorizado_por = "NR"
        else:
            autorizado_por = "Faltan autorizaciones"
            status = "Faltan autorizaciones"

        row = [
            viatico.folio,
            autorizado_at_2_naive,
            viatico.distrito.nombre,
            viatico.staff.staff.staff.first_name + ' ' + viatico.staff.staff.staff.last_name,
            viatico.colaborador.staff.staff.first_name + ' '  + viatico.colaborador.staff.staff.last_name if viatico.colaborador else '',
            viatico.proyecto.nombre,
            viatico.subproyecto.nombre,
            viatico.get_total,
            created_at_naive,
            status,
            autorizado_por,
            tiene_facturas,
            pagada,
            #f'=IF(I{row_num}="",G{row_num},I{row_num}*G{row_num})',  # Calcula total en pesos usando la fórmula de Excel
            #created_at_naive,
        ]

    
        for col_num in range(len(row)):
            (ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style
            if col_num ==1 or col_num == 8:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style
            if col_num == 7:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = money_style
       
    
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)

def generar_cfdi_buffer(request, pk):
    factura = Viaticos_Factura.objects.get(id=pk)
    data = factura.emisor
    # Verificar y asignar un valor predeterminado para impuestos si es None
    if data['impuestos'] is None:
        data['impuestos'] = 0.0

    # Verificar y asignar un valor predeterminado para total si es None
    if data['total'] is None:
        data['total'] = 0.0

    # Verificar y asignar un valor predeterminado para subtotal si es None
    if data['subtotal'] is None:
        data['subtotal'] = 0.0
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    if not data:
        return HttpResponse("Error al parsear el archivo XML", status=400)

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    
    # Generar código QR
    qr_data = f"https://verificacfdi.facturaelectronica.sat.gob.mx/default.aspx?id={data['uuid']}&re={data['rfc_emisor']}&rr={data['rfc_receptor']}&tt={data['total']}&fe={data['sello_cfd'][-8:]}"
    qr_img = qrcode.make(qr_data)
    
    with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as temp_file:
        qr_img.save(temp_file)
        temp_file.seek(0)
        qr_x = 500
        qr_y = height - 700
        qr_size = 2.75 * cm
        c.drawImage(temp_file.name, qr_x, qr_y, qr_size, qr_size)

    # Título
    c.setFillColor(prussian_blue)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(30, height - 40, "FACTURA GENERADA POR SAVIA 2.0")

    # Datos del Emisor
    c.setFillColor(black)
    c.setFont("Helvetica-Bold", 12)
    c.drawString(30, height - 80, "Datos del Emisor:")
    
    c.setFont("Helvetica", 8)
    alineado_x = 30
    alineado_y = height - 100
    alineado_y2 = alineado_y
    line_height = 12

    c.drawString(alineado_x, alineado_y, f"RFC: {data['rfc_emisor']}")
    alineado_y -= line_height
    c.drawString(alineado_x, alineado_y, f"Nombre: {data['nombre_emisor']}")
    alineado_y -= line_height
    c.drawString(alineado_x, alineado_y, f"Régimen Fiscal: {data['regimen_fiscal_emisor']}")
    alineado_y -= line_height
    c.drawString(alineado_x, alineado_y, f"Lugar de Expedición: {data['lugar_expedicion']}")
    alineado_y -= line_height
    c.drawString(alineado_x, alineado_y, f"Fecha y hora de expedición: {data['fecha']}")
    alineado_y -= line_height
    c.drawString(alineado_x, alineado_y, f"Moneda: {data['moneda']}")
    alineado_y -= line_height
    c.drawString(alineado_x, alineado_y, f"Forma de Pago: {data['forma_pago']}")

    # Datos del Receptor
    alineado_y -= 2 * line_height
    c.setFont("Helvetica-Bold", 12)
    c.drawString(alineado_x + 350, height - 80, "Datos del Receptor:")
    
    c.setFont("Helvetica", 8)
    alineado_y -= line_height
    c.drawString(alineado_x + 350, alineado_y2, f"RFC: {data['rfc_receptor']}")
    alineado_y2 -= line_height
    c.drawString(alineado_x + 350, alineado_y2, f"Nombre: {data['nombre_receptor']}")
    alineado_y2 -= line_height
    c.drawString(alineado_x + 350, alineado_y2, f"Régimen Fiscal: {data['regimen_fiscal_receptor']}")
    alineado_y2 -= line_height
    c.drawString(alineado_x + 350, alineado_y2, f"Régimen Fiscal: {data['codigo_postal']}")
    alineado_y2 -= line_height
    c.drawString(alineado_x + 350, alineado_y2, f"Uso del CFDI: {data['uso_cfdi']}")

    # Conceptos (Tabla)
    alineado_y -= line_height
    # Configuración del estilo para los párrafos
    styles = getSampleStyleSheet()
    styleN = styles['Normal']
    styleN.wordWrap = 'CJK'  # Ajusta automáticamente el texto
    # Crear un estilo personalizado
    custom_style = ParagraphStyle(
        'CustomStyle',
        parent=styleN,
        fontSize=6,  # Ajusta el tamaño del texto aquí
        leading=7,   # Ajusta el interlineado aquí si es necesario
    )

    # Preparamos los datos de la tabla
    table_data = [["CANT", "CLAVE", "CONCEPTO", "U DE M", "P.U.", "IMPORTE", "IMPUESTO", "TIPO TASA"]]
    for item in data['resultados']:
        descripcion = item['descripcion']
        cantidad = float(item['cantidad'])
        unidad = item['unidad']
        valor_unitario = float(item['precio'])
        importe = float(item['importe'])
        # Verificar y convertir solo si el valor no es 'N/A'
         # Inicializar las variables impuesto y tasa
        impuesto = item['impuesto']
        tasa = item['tasa_cuota']
        if impuesto != 'N/A':
            impuesto = float(impuesto)
        else:
            impuesto = 0.0  # o cualquier valor predeterminado que consideres adecuado
        
        if tasa != 'N/A':
            tasa = float(tasa)
        else:
            tasa = 0.0  # o cualquier valor predeterminado que consideres adecuado
        clave = item['clave_prod_serv']
         # Crear un párrafo para la descripción
        descripcion_paragraph = Paragraph(descripcion, custom_style)
        unidad_paragraph = Paragraph(unidad, custom_style)
        table_data.append([
            f"{cantidad:.2f}",
            clave,
            descripcion_paragraph,
            unidad_paragraph,
            f"{valor_unitario:,.2f}",
            f"{importe:,.2f}",
            f"{impuesto:,.2f}",
            f"{tasa:.2f}",
        ])

    # Crear la tabla
    table = Table(table_data, colWidths=[1.0 * cm, 1.5 * cm, 8.5 * cm, 1.5 * cm, 2 * cm, 2 * cm, 1.5 * cm, 1.5 * cm, 1.5 * cm])
    table.setStyle(TableStyle([
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 0), (-1, 0), prussian_blue),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
    ]))

    # Guardar la tabla en el PDF
    table.wrapOn(c, width, height)
    table.drawOn(c, alineado_x, alineado_y - len(table_data) * line_height)

    # Ajustar el alineado_y para seguir escribiendo debajo de la tabla
    alineado_y -= len(table_data) * line_height + 2 * line_height

    # Totales
    c.setFont("Helvetica-Bold", 12)
   
    c.setFont("Helvetica", 10)
    alineado_y -= line_height

     # Importe con letra
    alineado_y -= 2 * line_height
    c.drawString(alineado_x, alineado_y, "Importe con Letra:")
    total_letras = num2words(float(data['total']), lang='es', to='currency', currency='MXN')
    c.drawString(alineado_x, alineado_y - 10, total_letras)
    #c.drawRightString(alineado_x, alineado_y , f"{data['importe_con_letra']}")
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.setFillColor(prussian_blue)
    c.rect(alineado_x + 390 ,alineado_y - 50,110,62, fill=True, stroke=False) #Barra azul superior | Subtotal
    c.setFillColor(white)
    c.drawRightString(alineado_x + 500, alineado_y , f"Subtotal:")
    c.setFillColor(black)
    c.drawRightString(alineado_x + 555, alineado_y, f"{float(data['subtotal']):,.2f}")
    alineado_y -= line_height
    c.setFillColor(white)
    c.drawRightString(alineado_x + 500, alineado_y, f"Impuestos trasladados:")
    c.setFillColor(black)
    c.drawRightString(alineado_x + 555, alineado_y, f"{float(data['impuestos']):,.2f}")
    alineado_y -= line_height
    if data['iva_retenido'] > 0:
        c.setFillColor(white)
        c.drawRightString(alineado_x + 500, alineado_y, f"Impuestos retenidos:")
        c.setFillColor(black)
        c.drawRightString(alineado_x + 555, alineado_y, f"{float(data['iva_retenido']):,.2f}")
        alineado_y -= line_height
    if data['isr_retenido'] > 0:
        c.setFillColor(white)
        c.drawRightString(alineado_x + 500, alineado_y, f"ISR:")
        c.setFillColor(black)
        c.drawRightString(alineado_x + 555, alineado_y, f"{float(data['isr_retenido']):,.2f}")
        alineado_y -= line_height
    c.setFillColor(white)
    c.drawRightString(alineado_x + 500, alineado_y, f"Total:")
    c.setFillColor(black)
    c.drawRightString(alineado_x + 555, alineado_y, f"{float(data['total']):,.2f}")
    # Otros detalles
    

    otros_detalles = [
        ["Folio Fiscal", "Fecha y Hora de Certificación", "No. Certificado Digital", "Método de Pago"],
        [data['uuid'], data['fecha_timbrado'], data['no_certificado'], data['metodo_pago']]
    ]
    detalles_table = Table(otros_detalles, colWidths=[5 * cm, 5 * cm, 4.5 * cm, 4.5 * cm])
    detalles_table.setStyle(TableStyle([
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, 0), (-1, 0), prussian_blue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
    ]))

    # Guardar la tabla de detalles en el PDF
    detalles_table.wrapOn(c, width, height)
    detalles_table.drawOn(c, alineado_x, 180)
    alineado_y -= 4 * line_height
     # Utilizar Paragraph para las líneas largas
    styles = getSampleStyleSheet()
    styleN = styles["BodyText"]
    styleN.fontSize = 6
    c.setFont("Helvetica", 6)
    c.line(30,177,580,177)
    c.drawString(alineado_x, 170, f"ESTE DOCUMENTO ES UNA REPRESENTACIÓN IMPRESA DE UN CFDI v4.0")
    
    # Reducir el ancho de los párrafos
    reduced_width = width * 0.7  # Ajusta este valor según sea necesario

    sello_cfd_paragraph = Paragraph(f"Sello Digital del CFDI: {data['sello_cfd']}", styleN)
    sello_cfd_paragraph.wrapOn(c,  reduced_width, line_height * 4)
    sello_cfd_paragraph.drawOn(c, alineado_x, 130)
    alineado_y -= line_height * 5
    
    sello_sat_paragraph = Paragraph(f"Sello del SAT: {data['sello_sat']}", styleN)
    sello_sat_paragraph.wrapOn(c,  reduced_width, line_height * 4)
    sello_sat_paragraph.drawOn(c, alineado_x, 90)
    alineado_y -= line_height * 3
    c.drawString(alineado_x, 40, f"No. serie CSD SAT {data['no_certificadoSAT']}")

    sello_cfd_paragraph = Paragraph(f"Cadena Original del complemento de certificación digital del SAT: {data['cadena_original']}", styleN)
    sello_cfd_paragraph.wrapOn(c,  reduced_width, line_height * 4)
    sello_cfd_paragraph.drawOn(c, alineado_x, 50)
    alineado_y -= line_height * 5
    
   

    c.showPage()
    c.save()

    buffer.seek(0)
    # Crear la respuesta HTTP con el PDF
    folio_fiscal = data['uuid']
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{folio_fiscal}.pdf"'

    return response


def generar_cfdi_viaticos(request, pk):
    factura = Factura.objects.get(id=pk)
    buffer = generar_cfdi_buffer(factura)
    # Crear la respuesta HTTP con el PDF
    folio_fiscal = factura.emisor.get('uuid', f'factura_{factura.id}')
    return HttpResponse(buffer, content_type='application/pdf', headers={
        'Content-Disposition': f'attachment; filename="{folio_fiscal}.pdf"'
    })

def crear_pdf_cfdi_viaticos(factura):
    buffer = generar_cfdi_buffer(factura)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_file:
        tmp_file.write(buffer.read())
        return tmp_file.name