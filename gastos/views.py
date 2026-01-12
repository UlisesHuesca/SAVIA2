from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse, FileResponse
from django.core.paginator import Paginator
from django.db.models.functions import Concat
from django.db.models import Sum, Q, Prefetch, Max, Value,Count, When, Case,DecimalField

from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.mail import EmailMessage
from django.conf import settings
from django.urls import reverse
from django.core.mail import EmailMessage, BadHeaderError
import traceback
import calendar


import zipfile

from smtplib import SMTPException
import logging
import socket
from .models import Solicitud_Gasto, Articulo_Gasto, Entrada_Gasto_Ajuste, Conceptos_Entradas, Factura, Tipo_Gasto, ValeRosa, TipoArchivoSoporte, ArchivoSoporte
from .forms import Solicitud_GastoForm, Articulo_GastoForm, Articulo_Gasto_Edit_Form, Pago_Gasto_Form, Entrada_Gasto_AjusteForm, Conceptos_EntradasForm, UploadFileForm, FacturaForm, Autorizacion_Gasto_Form, Vale_Rosa_Form
from .filters import Solicitud_Gasto_Filter, Conceptos_EntradasFilter
from user.models import Profile, Distrito, Empresa 
from dashboard.models import Inventario, Order, ArticulosparaSurtir, ArticulosOrdenados, Tipo_Orden, Product
from solicitudes.models import Proyecto, Subproyecto, Operacion
from tesoreria.models import Pago, Cuenta, Facturas, Tipo_Pago
from compras.models import Proveedor_direcciones
from tesoreria.forms import Facturas_Gastos_Form, Transferencia_Form, Cargo_Abono_Form

from compras.views import attach_oc_pdf
from tesoreria.utils import extraer_texto_de_pdf, encontrar_variables, extraer_bloques_formato_2, detectar_formato_pdf, encontrar_variables_bloques
from requisiciones.views import get_image_base64
from viaticos.models import Viaticos_Factura, Solicitud_Viatico
import qrcode
from num2words import num2words
import tempfile
#PDF generator
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.colors import Color, black, blue, red, white
from reportlab.lib.units import cm
from reportlab.lib.pagesizes import letter, portrait
from reportlab.rl_config import defaultPageSize 
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Frame
from bs4 import BeautifulSoup
from django.utils import timezone
#Excel stuff
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill
from openpyxl.utils import get_column_letter
import xlsxwriter
from io import BytesIO


from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from datetime import date, datetime, timedelta
import json
import xml.etree.ElementTree as ET
import decimal
import os
import io
import datetime as dt
import pytz
import re
from user.decorators import perfil_seleccionado_required

logger = logging.getLogger(__name__)

#Se duplican funciones porque me arroja una referencia circular
#######################################################################
# Create your views here.
def eliminar_caracteres_invalidos(archivo_xml):
    # Definir la expresión regular para encontrar caracteres inválidos
    regex = re.compile(r'[^\x09\x0A\x0D\x20-\uD7FF\uE000-\uFFFD\u10000-\u10FFFF]')

    # Leer el contenido del archivo XML
    xml_content = archivo_xml.read().decode('utf-8')

    if xml_content.startswith("o;?"):
        print('Detectado "o;?" en el inicio del XML')
        xml_content = xml_content[3:]

    # Eliminar caracteres inválidos según la expresión regular
    xml_content = regex.sub('', xml_content)

    # Volver a posicionar el puntero del archivo al principio
    archivo_xml.seek(0)

    # Guardar el contenido modificado en el archivo original
    archivo_xml.write(xml_content.encode('utf-8'))
    archivo_xml.truncate()  # Asegurarse de que no quede contenido sobrante

    print('Contenido corregido guardado exitosamente.')

    # Retornar el archivo con el contenido modificado
    return archivo_xml

def extraer_datos_del_xml(archivo_xml):
    try:
        # Parsear el archivo XML
        archivo_xml.seek(0)
        tree = ET.parse(archivo_xml)
        root = tree.getroot()
    except (ET.ParseError, FileNotFoundError) as e:
        print(f"Error al parsear el archivo XML: {e}")
        return None, None  # Si ocurre un error, devuelve None
    
    # Identificar la versión del XML y el espacio de nombres
    version = root.tag
    ns = {}
    if 'http://www.sat.gob.mx/cfd/3' in version:
        ns = {
            'cfdi': 'http://www.sat.gob.mx/cfd/3',
            'tfd': 'http://www.sat.gob.mx/TimbreFiscalDigital',
            'if': 'https://www.interfactura.com/Schemas/Documentos',
        }
    elif 'http://www.sat.gob.mx/cfd/4' in version:
        ns = {
            'cfdi': 'http://www.sat.gob.mx/cfd/4',
            'tfd': 'http://www.sat.gob.mx/TimbreFiscalDigital',
            'if': 'https://www.interfactura.com/Schemas/Documentos',
        }
    else:
        print(f"Versión del documento XML no reconocida")
        return None, None
    

    rfc_receptor = None
    receptor = root.find('cfdi:Receptor', ns)
    if receptor is not None:
        rfc_receptor = receptor.get('Rfc')
    else:
        print("Receptor no encontrado")
    # Buscar el complemento donde se encuentra el UUID y la fecha de timbrado
    complemento = root.find('cfdi:Complemento', ns)
    if complemento is not None:
        timbre_fiscal = complemento.find('tfd:TimbreFiscalDigital', ns)
        if timbre_fiscal is not None:
            uuid = timbre_fiscal.get('UUID')
            fecha_timbrado = timbre_fiscal.get('FechaTimbrado') or root.get('Fecha')
        else:
            print("Timbre Fiscal Digital no encontrado")
            return None, None, None
    else:
        print("Complemento no encontrado")
        return None, None, None
    
    return uuid, fecha_timbrado, rfc_receptor
################################################################################################################################
def procesar_bbva(archivo):
    total = 0.0
    for linea in archivo:
        linea = linea.decode('utf-8').strip()
        # Encontrar la última secuencia de 15 dígitos seguidos de letras (nombre)
        import re
        match = re.search(r'(\d{15})([A-ZÁÉÍÓÚÑ\s]{5,})', linea)
        if match:
            sueldo_str = match.group(1)
            if sueldo_str.isdigit():
                total += int(sueldo_str) / 100.0
    return total

def procesar_ob(archivo):
    total = 0.0
    for linea in archivo:
        linea = linea.decode('utf-8').strip()
        if 'MXP' in linea:
            try:
                inicio = linea.index('MXP') + 3
                fin = inicio
                while fin < len(linea) and (linea[fin].isdigit() or linea[fin] == '.'):
                    fin += 1
                sueldo_str = linea[inicio:fin]
                total += float(sueldo_str)
            except:
                continue
    return total

def procesar_banorte(archivo):
    """
    Procesa un archivo de nómina y devuelve el total de sueldos.
    
    Parámetros:
        archivo: UploadedFile (por ejemplo request.FILES['archivo_nomina'])
    
    Retorna:
        float: total de la nómina
    """
    total = 0.0
    
    for linea in archivo:
        linea = linea.decode("utf-8", errors="ignore").strip()
        if linea.startswith("D"):
            try:
                entero_str = linea[107:112]   # parte entera (5 caracteres)
                decimales_str = linea[112:114]  # parte decimal (2 caracteres)
                sueldo = int(entero_str) + int(decimales_str) / 100
                total += sueldo
            except ValueError:
                continue
    
    return round(total, 2)

def procesar_pensiones(archivo):
    return procesar_ob(archivo)  # Mismo formato que OB


@perfil_seleccionado_required
def crear_gasto(request):
    colaborador = Profile.objects.all()
    articulos_gasto = Articulo_Gasto.objects.all()
    distritos = Distrito.objects.none()
    
    conceptos = Product.objects.all()
    pk = request.session.get('selected_profile_id')
    usuario = colaborador.get(id = pk)
    tipos = Tipo_Gasto.objects.filter(familia = 'usuario')
    proyectos = Proyecto.objects.filter(~Q(status_de_entrega__status = "INACTIVO"),activo=True, distrito = usuario.distritos)     
    empresas = Empresa.objects.all()

    if usuario.distritos.nombre == "MATRIZ":
         
        if usuario.tipo.tesoreria:
            tipos = Tipo_Gasto.objects.filter(familia__in = ['usuario','tesoreria'])
        if usuario.tipo.nombre == "CONTADOR" or usuario.tipo.nombre == "SUPERINTENDENCIA_CONTABILIDAD":
            tipos = Tipo_Gasto.objects.filter(familia__in = ['usuario','contabilidad'])


        if usuario.tipo.rh: #Se pone a lo útimo para que sea el último filtro que haga
            tipos = Tipo_Gasto.objects.filter(familia__in=['usuario', 'rh', 'rh_nomina'])
            distritos = Distrito.objects.filter().exclude(nombre__in=["BRASIL","MATRIZ ALTERNATIVO","ALTAMIRA ALTERNATIVO","VH SECTOR 6"])
        if usuario.tipo.subdirector:
            superintendentes = colaborador.filter(tipo__dg = True, distritos = usuario.distritos, st_activo =True, sustituto__isnull = True) 
      
        else:    
            superintendentes = colaborador.filter(tipo__subdirector = True, distritos = usuario.distritos, st_activo =True, sustituto__isnull = True) 
    elif usuario.distritos.nombre == "BRASIL":
        superintendentes = colaborador.filter(distritos=usuario.distritos, tipo__supervisor = True, st_activo = True).exclude(tipo__nombre="Admin")
    elif usuario.tipo.superintendente and not usuario.tipo.nombre == "Admin" and not usuario.tipo.nombre == "GERENCIA":
        superintendentes = colaborador.filter(tipo__superintendente = True, staff =  usuario.staff, distritos = usuario.distritos )  
    else:
        superintendentes = colaborador.filter(tipo__superintendente=True, distritos = usuario.distritos, st_activo =True, sustituto__isnull = True).exclude(tipo__nombre="Admin").exclude(tipo__nombre="GERENCIA")

   
    #subproyectos = Subproyecto.objects.all()
    proveedores = Proveedor_direcciones.objects.filter(nombre__familia__nombre = "IMPUESTOS")
    #tipos = Tipo_Gasto.objects.all().exclude(id = 6)
    #print('tipos:',tipos)
    colaboradores = colaborador.filter(distritos = usuario.distritos, )
    error_messages = {}

    
    tipos_para_select2 = [
        {
            'id': item.id,
            'text': str(item.tipo),
            'familia': str(item.familia),
        } for item in tipos
    ]


    proyectos_para_select2 = [
        {
            'id': item.id, 
            'text': str(item.nombre)
        } for item in proyectos
    ]

    superintendentes_para_select2 = [
        {
            'id': super.id, 
            'text': str(super.staff.staff.first_name) + (' ') + str(super.staff.staff.last_name)
        } for super in superintendentes
    ]

    #print(proveedores)
    proveedores_para_select2 = [
        {
            'id':proveedor.id,
            'text': str(proveedor.nombre.razon_social)
        } for proveedor in proveedores
    ] 
    print('distritos',distritos)
    distritos_para_select2 = [
        {
            'id': distrito.id,
            'text': str(distrito.nombre)
        } for distrito in distritos 
    ]

    empresa_para_salect2 = [
        {
            'id': empresa.id,
            'text': str(empresa.nombre)
        } for empresa in empresas
    ]
    
    gasto, created = Solicitud_Gasto.objects.get_or_create(complete= False, staff=usuario)
    if not gasto.inicio_form:
        gasto.inicio_form = timezone.now()  # Asigna la fecha y hora con zona horaria
        gasto.save()

    print('gasto', gasto.id)
    max_folio = Solicitud_Gasto.objects.filter(distrito = usuario.distritos, complete=True).aggregate(Max('folio'))['folio__max']
    
    
    articulo, created = articulos_gasto.get_or_create(completo = False, staff=usuario)

    productos = articulos_gasto.filter(gasto=gasto, completo = True) 

    
    colaboradores_para_select2 = [
        {
            'id': item.id,
            'text': str(item.staff.staff.first_name) + (' ') + str(item.staff.staff.last_name)
        } for item in colaboradores
    ]
    

    articulos_gasto = conceptos.filter(gasto = True, baja_item = False) #Cuando Alberto envie los conceptos se implementa
    #articulos_gasto = conceptos.filter(gasto = True)
    facturas = Factura.objects.filter(solicitud_gasto = gasto)
    form_product = Articulo_GastoForm()
    form_vale = Vale_Rosa_Form()
    form = Solicitud_GastoForm()
    factura_form = UploadFileForm()

    productos_para_select2 = [
        {
            'id': item.id,
            'text': str(item.nombre),
            'iva': str(item.iva)
        } for item in articulos_gasto
    ]

    archivos_nomina = ArchivoSoporte.objects.filter(solicitud = gasto)

    total_bbva = archivos_nomina.filter(tipo__nombre='BBVA').first()
    total_ob = archivos_nomina.filter(tipo__nombre='OB').first()
    total_pensiones = archivos_nomina.filter(tipo__nombre='PENSIONES').first()
    total_banorte = archivos_nomina.filter(tipo__nombre='BANORTE').first()
    # ⚠️ Al final, obtener todos los archivos RH cargados de este gasto
    archivos_rh = ArchivoSoporte.objects.filter(solicitud=gasto).exclude(tipo__nombre__in=['BANORTE', 'OB', 'BBVA', 'PENSIONES'])
    error_messages = {}

    if request.method =='POST': 
        if "btn_agregar" in request.POST:
            form = Solicitud_GastoForm(request.POST, instance=gasto)
            if form.is_valid():
                
                if max_folio == None:
                    max_folio = 0
                gasto = form.save(commit=False)
                gasto.folio = max_folio + 1
                
                gasto.complete = True
                gasto.created_at = datetime.now()
                gasto.staff =  usuario
                if gasto.tipo.familia == "rh_nomina" or gasto.tipo.familia == "rh":
                    gasto.distrito = form.cleaned_data['distrito']
                    gasto.transferencia_finanzas = True
                else:
                    gasto.distrito = usuario.distritos

                max_folio = Solicitud_Gasto.objects.filter(
                distrito=gasto.distrito, 
                complete=True
                ).aggregate(Max('folio'))['folio__max'] or 0
                
                gasto.folio = max_folio + 1

                gasto.save()
                #form.save()
                messages.success(request, f'La solicitud {gasto.folio} ha sido creada')
                return redirect('mis-gastos')
            else:
                for field, errors in form.errors.items():
                    error_messages[field] = errors.as_text()
    
        if "btn_producto" in request.POST:
            form_product = Articulo_GastoForm(request.POST, instance=articulo)
            if form_product.is_valid():
                articulo = form_product.save(commit=False)
                articulo.gasto = gasto
                articulo.completo = True
                articulo.save()
                messages.success(request, 'Haz agregado un artículo correctamente')
                return redirect('crear-gasto')
            else:
                for field, errors in form_product.errors.items():
                    error_messages[f'Líneas de Gasto| Campo:{field}'] = errors.as_text()
        if "btn_factura" in request.POST:
            factura_form = UploadFileForm(request.POST, request.FILES or None)
            if factura_form.is_valid():
                archivos_pdf = request.FILES.getlist('archivo_pdf')
                archivos_xml = request.FILES.getlist('archivo_xml')
                if not archivos_pdf and not archivos_xml:
                    messages.error(request, 'Debes subir al menos un archivo PDF o XML.')
                    return HttpResponse(status=204)
                
                # Iterar sobre el número máximo de archivos en cualquiera de las listas
                max_len = max(len(archivos_pdf), len(archivos_xml))
                facturas_registradas = []
                facturas_duplicadas = []

                for i in range(max_len):
                    archivo_pdf = archivos_pdf[i] if i < len(archivos_pdf) else None
                    archivo_xml = archivos_xml[i] if i < len(archivos_xml) else None
                    
                    if archivo_xml:
                        archivo_procesado = eliminar_caracteres_invalidos(archivo_xml)
                        
                        #factura_temp = Factura(archivo_xml=archivo_xml)
                        #factura_temp.archivo_xml.save(archivo_xml.name, archivo_procesado, save=False)

                        uuid_extraido, fecha_timbrado_extraida, rfc_receptor = extraer_datos_del_xml(archivo_procesado)
                        RFC_RECEPTOR_ESPERADO = "GVO020226811"
                        if rfc_receptor and rfc_receptor != RFC_RECEPTOR_ESPERADO:
                            messages.error(request, f"RFC receptor inválido ({rfc_receptor}). Se esperaba {RFC_RECEPTOR_ESPERADO}.")
                            break # Saltar al siguiente archivo si el RFC no coincide

                        # Verificar si ya existe una factura con el mismo UUID y fecha de timbrado en cualquiera de las tablas
                        existe_en_gastos = Factura.objects.filter(uuid=uuid_extraido, fecha_timbrado=fecha_timbrado_extraida).exists()
                        existe_en_compras = Facturas.objects.filter(uuid=uuid_extraido, fecha_timbrado=fecha_timbrado_extraida).exists()
                        existe_en_viaticos = Viaticos_Factura.objects.filter(uuid=uuid_extraido, fecha_timbrado=fecha_timbrado_extraida).exists()

                        if existe_en_gastos or existe_en_compras or existe_en_viaticos:
                            facturas_duplicadas.append(uuid_extraido)
                            continue

                        # ✅ Aquí ya es seguro crear TU factura para ESTA iteración
                        factura = Factura.objects.create(
                            solicitud_gasto=gasto,
                            hecho=False,
                            subido_por=usuario,
                        # si tienes campos uuid/fecha_timbrado en el modelo, puedes setearlos aquí también
                        )
                        
                        if existe_en_gastos or existe_en_compras or existe_en_viaticos:
                            # Si una factura existente se encuentra, verificamos si su solicitud no está aprobada
                            if existe_en_gastos and (existe_en_gastos.solicitud_gasto.autorizar is False or existe_en_gastos.solicitud_gasto.autorizar2 is False):
                                existe_en_gastos.delete()
                                guardar_factura(factura, archivo_xml, uuid_extraido, fecha_timbrado_extraida, usuario)

                            elif existe_en_compras and (existe_en_compras.oc.autorizado1 is False or existe_en_compras.oc.autorizado2 is False):
                                existe_en_compras.delete()
                                guardar_factura(factura, archivo_xml, uuid_extraido, fecha_timbrado_extraida, usuario)

                            elif existe_en_viaticos and (existe_en_viaticos.solicitud_viatico.autorizar is False or existe_en_viaticos.solicitud_viatico.autorizar2 is False):
                                existe_en_viaticos.delete()
                                guardar_factura(factura, archivo_xml, uuid_extraido, fecha_timbrado_extraida, usuario)

                            else:
                                # Si no cumple las condiciones de eliminación, consideramos la factura duplicada
                                facturas_duplicadas.append(uuid_extraido)
                                continue  # Saltar al siguiente archivo si se encuentra duplicado
                        else:
                            # Si no existe ninguna factura, guardar la nueva
                            guardar_factura(factura, archivo_xml, uuid_extraido, fecha_timbrado_extraida, usuario)
                            #messages.success(request, 'Las facturas se registraron de manera exitosa')
                    if archivo_pdf:
                        factura.archivo_pdf = archivo_pdf
                        factura.hecho = True
                        factura.fecha_subida = datetime.now()
                        factura.subido_por = usuario
                        factura.save()
                      
                        facturas_registradas.append(uuid_extraido if archivo_xml else f"Factura PDF {archivo_pdf.name}")
                    #messages.success(request, 'Los facturas se registraron de manera exitosa')
                     # Mensajes de éxito o duplicados
                #return HttpResponse(status=204)
                if facturas_registradas:
                    messages.success(request, f'Se han registrado las siguientes facturas: {", ".join(facturas_registradas)}')
                if facturas_duplicadas:
                    messages.warning(request, f'Las siguientes no se pudieron subir porque ya estaban registradas: {", ".join(facturas_duplicadas)}')

            else:
                messages.error(request,'No se pudo subir tu documento') 
        if "btn_valerosa" in request.POST:
            form_vale = Vale_Rosa_Form(request.POST, request.FILES or None)
            if form_vale.is_valid():
                vale = form_vale.save(commit=False)
                vale.creado_por = gasto.staff
                vale.gasto = gasto
                vale.save()
                return redirect('crear-gasto')
        if 'btn_documentos_rh' in request.POST:
            archivo = None
            tipo_documento = None
            
            # Iterar sobre los archivos enviados en el form
            for key, uploaded_file in request.FILES.items():
                if uploaded_file:
                    archivo = uploaded_file
                    tipo_documento = key  # ¡Aquí está el tipo de documento!
                    print(f'Tipo de documento: {tipo_documento}, Archivo: {archivo.name}')
                    break  # Solo tomamos el primero (en este caso es uno solo)

            if archivo and tipo_documento:
                try:
                    tipo = TipoArchivoSoporte.objects.get(nombre=tipo_documento)
                    ArchivoSoporte.objects.create(
                        solicitud=gasto,
                        tipo=tipo,
                        archivo=archivo,
                        total=0.0  # Asignar un total de 0.0 por defecto
                    )
                    messages.success(request, f'Archivo {tipo_documento} subido exitosamente.')
                    return redirect('crear-gasto', pk=gasto.pk)
                except TipoArchivoSoporte.DoesNotExist:
                    messages.error(request, f'Tipo de documento {tipo_documento} no encontrado.')
               
        if "btn_nomina" in request.POST:
            total_bbva = total_ob = total_pensiones = total_banorte = 0.0
            archivo_bbva = request.FILES.get('archivo_bbva')
            archivo_ob = request.FILES.get('archivo_ob')
            archivo_pensiones = request.FILES.get('archivo_pensiones')
            archivo_banorte = request.FILES.get('archivo_banorte')

            
            if archivo_bbva:
                total_bbva = procesar_bbva(archivo_bbva)
                tipo_bbva = TipoArchivoSoporte.objects.get(nombre='BBVA')
                ArchivoSoporte.objects.create(
                    solicitud=gasto,
                    tipo=tipo_bbva,
                    archivo=archivo_bbva,
                    total=total_bbva
                )
            if archivo_ob:
                total_ob = procesar_ob(archivo_ob)
                tipo_ob = TipoArchivoSoporte.objects.get(nombre='OB')
                ArchivoSoporte.objects.create(
                    solicitud=gasto,
                    tipo=tipo_ob,
                    archivo=archivo_ob,
                    total=total_ob
                )
            if archivo_pensiones:
                total_pensiones = procesar_pensiones(archivo_pensiones)
                tipo_pensiones = TipoArchivoSoporte.objects.get(nombre='PENSIONES')
                ArchivoSoporte.objects.create(
                    solicitud=gasto,
                    tipo=tipo_pensiones,
                    archivo=archivo_pensiones,
                    total=total_pensiones
                )
            if archivo_banorte:
                total_banorte = procesar_banorte(archivo_banorte)
                tipo_banorte = TipoArchivoSoporte.objects.get(nombre='BANORTE')
                ArchivoSoporte.objects.create(
                    solicitud=gasto,
                    tipo=tipo_banorte,
                    archivo=archivo_banorte,
                    total=total_banorte
                )



            messages.success(request, f'Total BBVA: ${total_bbva:,.2f}, OB: ${total_ob:,.2f}, Pensiones: ${total_pensiones:,.2f} Banorte: ${total_banorte:,.2f}')
            return redirect('crear-gasto')
            #except Exception as e:
            #    messages.error(request, f'Error al procesar archivos de nómina: {str(e)}')

    total_nomina = archivos_nomina.aggregate(suma=Sum('total'))['suma'] or 0.0
    #print(distritos_para_select2)


    context= {
        'error_messages':error_messages,
        'tipos_para_select2':tipos_para_select2,
        'colaboradores_para_select2':colaboradores_para_select2,
        'superintendentes_para_select2':superintendentes_para_select2,
        'proyectos_para_select2':proyectos_para_select2,
        'productos_para_select2':productos_para_select2,
        'proveedores_para_select2':proveedores_para_select2,
        'distritos_para_select2':distritos_para_select2,
        'facturas':facturas,
        'productos':productos,
        'archivo_bbva': total_bbva,
        'archivo_ob': total_ob,
        'archivo_pensiones': total_pensiones,
        'archivos_rh': archivos_rh,
        'archivo_banorte': total_banorte,
        'form':form,
        'total_nomina': total_nomina,
        'form_product': form_product,
        'empresa_para_salect2': empresa_para_salect2,
        'form_vale': form_vale,
        'gasto':gasto,
        #'superintendentes':superintendentes,
        #'proyectos':proyectos,
        #'subproyectos':subproyectos,
        'factura_form': factura_form,
    }
    return render(request, 'gasto/crear_gasto.html', context)

def eliminar_archivo_pago(request, archivo_id):
    if request.method == 'POST':
        archivo = get_object_or_404(ArchivoSoporte, id=archivo_id)  # cambia ArchivoPago por el modelo real
        archivo.archivo.delete()  # borra el archivo físico
        archivo.delete()          # borra el registro
        messages.success(request, 'Archivo eliminado correctamente.')
    return redirect(request.META.get('HTTP_REFERER', '/'))


def eliminar_archivo(request, archivo_id):
    archivo = get_object_or_404(ArchivoSoporte, id=archivo_id)
    if request.method == 'POST':
        archivo.delete()
        messages.success(request, 'Archivo eliminado correctamente.')
    return redirect('crear-gasto')  # Ajusta al nombre real de tu vista de gasto

@perfil_seleccionado_required
def agregar_vale_rosa(request, pk):
    tipo = request.GET.get('tipo')  # puede ser 'gasto' o 'viatico'
    print(tipo)
    if tipo == 'gasto':
        objeto = get_object_or_404(Solicitud_Gasto, id=pk)
    elif tipo == 'viatico':
        objeto = get_object_or_404(Solicitud_Viatico, id=pk)
    else:
        messages.error(request, 'Tipo de vale no reconocido.')
        return redirect('mis-gastos')  # o cualquier página segura
    
    form = Vale_Rosa_Form()
    
    if request.method == 'POST':
        form = Vale_Rosa_Form(request.POST, request.FILES or None)
        if form.is_valid():
            vale = form.save(commit=False)
            vale.creado_por = objeto.staff
            if tipo == 'gasto':
                vale.gasto = objeto
            elif tipo == 'viatico':
                vale.viatico = objeto
                print(objeto)
            vale.save()
            messages.success(request, 'Vale Rosa agregado exitosamente')
            next_url = request.GET.get('next') or 'mis-gastos'
            return redirect(next_url)
        else:
            messages.error(request, 'Por favor completa todos los campos.')
 
    
    context = {
        'objeto': objeto,
        'tipo': tipo,
        'form': form,
    }
    
    return render(request, 'gasto/crear_vale_rosa.html', context)



@perfil_seleccionado_required
def delete_gasto(request, pk):
    articulo = Articulo_Gasto.objects.get(id=pk)
    messages.success(request,f'El articulo {articulo.producto} ha sido eliminado exitosamente')
    articulo.delete()

    return redirect('crear-gasto')

def eliminar_vale_rosa(request, pk):
    # No se de donde proviene, por lo tanto me dirije al crear-gasto o crear-viatico, habría que mejorar esto
    print(pk)
    vale = get_object_or_404(ValeRosa, pk=pk)
    if vale.gasto:
        gasto_id = vale.gasto.id  # ValeRosa tiene FK a Gasto
        vale.delete()
        return redirect('crear-gasto') 
    elif vale.viatico:
        viatico_id = vale.viatico.id  # ValeRosa tiene FK a Viatico
        vale.delete()
        return redirect('crear-viatico') 

@perfil_seleccionado_required
def eliminar_factura(request, pk):
    articulo = Factura.objects.get(id=pk)
    messages.success(request,f'La factura {articulo.id} ha sido eliminada exitosamente')
    articulo.delete()

    return redirect('crear-gasto')

@perfil_seleccionado_required
def eliminar_factura_gasto(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk_perfil)
    factura = Factura.objects.get(id=pk)
    gasto = factura.solicitud_gasto
    comentario = request.POST.get('comentario')
    # Obtener el parámetro `next` de la URL
    next_url = request.GET.get('next', None)
    print('URLLLLLLLLLLLLLLLLLLL2')
    print(next_url)
    # Construir la URL usando reverse
    matriz_url = reverse('matriz-facturas-gasto', args=[gasto.id])
    static_path = settings.STATIC_ROOT
    img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
    img_path2 = os.path.join(static_path,'images','logo_vordcab.jpg')
    image_base64 = get_image_base64(img_path)
    logo_v_base64 = get_image_base64(img_path2)
    # Crear el mensaje HTML
    html_message = f"""
    <html>
        <head>
            <meta charset="UTF-8">
        </head>
        <body style="font-family: Arial, sans-serif; color: #333; background-color: #f4f4f4; margin: 0; padding: 0;">
            <table width="100%" cellspacing="0" cellpadding="0" style="background-color: #f4f4f4; padding: 20px;">
                <tr>
                    <td align="center">
                        <table width="600px" cellspacing="0" cellpadding="0" style="background-color: #ffffff; padding: 20px; border-radius: 10px;">
                            <tr>
                                <td align="center">
                                    <img src="data:image/jpeg;base64,{logo_v_base64}" alt="Logo" style="width: 100px; height: auto;" />
                                </td>
                            </tr>
                            <tr>
                                <td style="padding: 20px;">
                                    <p style="font-size: 18px; text-align: justify;">
                                        <p>Estimado {factura.solicitud_gasto.staff.staff.staff.first_name} {factura.solicitud_gasto.staff.staff.staff.last_name},</p>
                                    </p>
                                    <p style="font-size: 16px; text-align: justify;">
                                        Estás recibiendo este correo porque tu factura subida el: <strong>{factura.fecha_subida.date()}</strong> en el gasto <strong>{gasto.folio}</strong> ha sido eliminada.</p>
                                    <p>Comentario:</p>
                                    {comentario}
                               
                                    </p>
                                <p style="font-size: 16px; text-align: justify;">
                                    Att: {perfil.staff.staff.first_name} {perfil.staff.staff.last_name}
                                </p>
                                    <p style="text-align: center; margin: 20px 0;">
                                        <img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width: 50px; height: auto; border-radius: 50%;" />
                                    </p>
                                    <p style="font-size: 14px; color: #999; text-align: justify;">
                                        Este mensaje ha sido automáticamente generado por SAVIA 2.0
                                    </p>
                                </td>
                            </tr>
                        </table>
                    </td>
                </tr>
            </table>
        </body>
    </html>
    """
    try:
        email = EmailMessage(
            f'Factura eliminada',
            body=html_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[gasto.staff.staff.staff.email],
            headers={'Content-Type': 'text/html'}
            )
        email.content_subtype = "html " # Importante para que se interprete como HTML
        if factura.archivo_pdf:
            pdf_path = factura.archivo_pdf.path
            if os.path.exists(pdf_path):  # Verificar si el archivo realmente existe
                with open(pdf_path, 'rb') as pdf_file:
                    email.attach(factura.archivo_pdf.name, pdf_file.read(), 'application/pdf')
            else:
                print(f"El archivo PDF no se encuentra en la ruta: {pdf_path}")

        if factura.archivo_xml:
            xml_path = factura.archivo_xml.path
            if os.path.exists(xml_path):  # Verificar si el archivo realmente existe
                with open(xml_path, 'rb') as xml_file:
                    email.attach(factura.archivo_xml.name, xml_file.read(), 'application/xml')
            else:
                print(f"El archivo XML no se encuentra en la ruta: {xml_path}")

        email.send()
        messages.success(request, f'La factura {factura.id} ha sido eliminada exitosamente')
    except (BadHeaderError, SMTPException, socket.gaierror) as e:
        error_message = f'La factura {factura.id} ha sido eliminada, pero el correo no ha sido enviado debido a un error: {e}'
        messages.success(request, error_message)
    factura.delete()
    # Si next_url existe, redirigir agregando el parámetro `next`
    if next_url:
        return redirect(f'{matriz_url}?next={next_url}')
    else:
        return redirect(matriz_url)

def extraer_uuid_y_año(archivo_xml):
    try:
        tree = ET.parse(archivo_xml)
        root = tree.getroot()
        ns = {
            'cfdi': 'http://www.sat.gob.mx/cfd/4',
            'xsi': 'http://www.w3.org/2001/XMLSchema-instance',
            'tfd': 'http://www.sat.gob.mx/TimbreFiscalDigital'
        }

        # Buscar la fecha en el Comprobante
        # Buscar la fecha en el Comprobante
       
       # Buscar el UUID en el TimbreFiscalDigital
        uuid_element = root.find('.//tfd:TimbreFiscalDigital', ns)
        if uuid_element is None:
            print("No se encontró el elemento TimbreFiscalDigital")
        uuid = uuid_element.get('UUID') if uuid_element is not None else None
         
        #comprobante = root.find('.//cfdi:Comprobante', ns)
        #if comprobante is None:
        #    print("No se encontró el elemento Comprobante")
        fecha = uuid_element.get('FechaTimbrado') if uuid_element is not None else None
        ano = fecha[:4] if fecha else None
        
        
        print(f"UUID: {uuid}, Año: {ano}")
        return uuid, ano
    except ET.ParseError as e:
        print(f"Error al parsear el archivo XML: {e}")
        return None, None
    
def verificar_uuid_unico(uuid, ano):
    facturas = Factura.objects.filter(archivo_xml__icontains=ano) #[300:310]
    cont_facturas = facturas.count()
    print('conteo:',cont_facturas)
    for factura in facturas:
        if factura.archivo_xml:
            uuid_existente, año_existente = extraer_uuid_y_año(factura.archivo_xml.path)
            if uuid == uuid_existente and ano == año_existente:
                return False
    return True


def guardar_factura(factura, archivo_xml, uuid_extraido, fecha_timbrado_extraida, usuario):
    factura.archivo_xml = archivo_xml
    factura.uuid = uuid_extraido
    factura.fecha_timbrado = fecha_timbrado_extraida
    factura.hecho = True
    factura.fecha_subida = datetime.now()
    factura.subido_por = usuario
    factura.save()

@perfil_seleccionado_required
def factura_nueva_gasto(request, pk):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    gasto = Solicitud_Gasto.objects.get(id = pk)
    #facturas = Facturas.objects.filter(pago = pago, hecho=True)
    #facturas, created = Factura.objects.get_or_create(solicitud_gasto=gasto, hecho=False)
    

    form = UploadFileForm()

    if request.method == 'POST':
        if 'btn_registrar' in request.POST:
            form = UploadFileForm(request.POST, request.FILES or None)
            if form.is_valid():
                archivos_pdf = request.FILES.getlist('archivo_pdf')
                archivos_xml = request.FILES.getlist('archivo_xml')
                if not archivos_pdf and not archivos_xml:
                    messages.error(request, 'Debes subir al menos un archivo PDF o XML.')
                    return HttpResponse(status=204)
                
                # Iterar sobre el número máximo de archivos en cualquiera de las listas
                max_len = max(len(archivos_pdf), len(archivos_xml))
                facturas_registradas = []
                facturas_duplicadas = []

                for i in range(max_len):
                    archivo_pdf = archivos_pdf[i] if i < len(archivos_pdf) else None
                    archivo_xml = archivos_xml[i] if i < len(archivos_xml) else None
                    factura, created = Factura.objects.get_or_create(solicitud_gasto=gasto, hecho=False)
                    if archivo_xml:
                        archivo_procesado = eliminar_caracteres_invalidos(archivo_xml)

                        # Guardar temporalmente para extraer datos
                        #factura_temp = Factura(archivo_xml=archivo_xml)
                        #factura_temp.archivo_xml.save(archivo_xml.name, archivo_procesado, save=False)

                        
                        uuid_extraido, fecha_timbrado_extraida, rfc_receptor = extraer_datos_del_xml(archivo_procesado)
                        # Verificar si ya existe una factura con el mismo UUID y fecha de timbrado en cualquiera de las tablas
                        factura_existente = Factura.objects.filter(uuid=uuid_extraido, fecha_timbrado=fecha_timbrado_extraida).first()
                        facturas_existentes = Facturas.objects.filter(uuid=uuid_extraido, fecha_timbrado=fecha_timbrado_extraida).first()
                        viaticos_factura_existente = Viaticos_Factura.objects.filter(uuid=uuid_extraido, fecha_timbrado=fecha_timbrado_extraida).first()
                        RFC_RECEPTOR_ESPERADO = "GVO020226811"
                       
                        if rfc_receptor and rfc_receptor != RFC_RECEPTOR_ESPERADO:
                            messages.error(request, f"RFC receptor inválido ({rfc_receptor}). Se esperaba {RFC_RECEPTOR_ESPERADO}.")
                            break # Saltar al siguiente archivo si el RFC no coincide

                        if factura_existente or facturas_existentes or viaticos_factura_existente:
                            # Si una factura existente se encuentra, verificamos si su solicitud no está aprobada
                            if factura_existente and (factura_existente.solicitud_gasto.autorizar is False or factura_existente.solicitud_gasto.autorizar2 is False):
                                factura_existente.delete()
                                guardar_factura(factura, archivo_xml, uuid_extraido, fecha_timbrado_extraida, usuario)

                            elif facturas_existentes and (facturas_existentes.oc.autorizado1 is False or facturas_existentes.oc.autorizado2 is False):
                                facturas_existentes.delete()
                                guardar_factura(factura, archivo_xml, uuid_extraido, fecha_timbrado_extraida, usuario)

                            elif viaticos_factura_existente and (viaticos_factura_existente.solicitud_viatico.autorizar is False or viaticos_factura_existente.solicitud_viatico.autorizar2 is False):
                                viaticos_factura_existente.delete()
                                guardar_factura(factura, archivo_xml, uuid_extraido, fecha_timbrado_extraida, usuario)

                            else:
                                # Si no cumple las condiciones de eliminación, consideramos la factura duplicada
                                facturas_duplicadas.append(uuid_extraido)
                                continue  # Saltar al siguiente archivo si se encuentra duplicado
                        else:
                            # Si no existe ninguna factura, guardar la nueva
                            guardar_factura(factura, archivo_xml, uuid_extraido, fecha_timbrado_extraida, usuario)
                            #messages.success(request, 'Las facturas se registraron de manera exitosa')
                    if archivo_pdf:
                        factura.archivo_pdf = archivo_pdf
                        factura.hecho = True
                        factura.fecha_subida = datetime.now()
                        factura.subido_por = usuario
                        factura.save()
                      
                        facturas_registradas.append(uuid_extraido if archivo_xml else f"Factura PDF {archivo_pdf.name}")
                    #messages.success(request, 'Los facturas se registraron de manera exitosa')
                     # Mensajes de éxito o duplicados
                #return HttpResponse(status=204)
                if facturas_registradas:
                    messages.success(request, f'Se han registrado las siguientes facturas: {", ".join(facturas_registradas)}')
                if facturas_duplicadas:
                    messages.warning(request, f'Las siguientes no se pudieron subir porque ya estaban registradas: {", ".join(facturas_duplicadas)}')

            else:
                messages.error(request,'No se pudo subir tu documento')


    context={
        'form': form, 
        'gasto': gasto,
    }

    return render(request, 'gasto/registrar_nueva_factura_gasto.html', context)

def prellenar_formulario_gastos(request):
    print('prellenar_formulario_gastos')
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        pdf_content = request.FILES.get('comprobante_pago')
        
        if not pdf_content:
            return JsonResponse({'error': 'No file uploaded'}, status=400)
        
        pdf_content = pdf_content.read()
        
        
        texto_extraido = extraer_texto_de_pdf(pdf_content)
        datos_extraidos = encontrar_variables(texto_extraido)
            #divisa_cuenta_extraida = datos_extraidos.get('divisa_cuenta', '').strip()

     
        divisa_cuenta_extraida = datos_extraidos.get('divisa_cuenta', '').strip()

        
        fecha_str = datos_extraidos.get('fecha', '').strip()
        #print(fecha_str)
        fecha_formato_correcto = None  # Valor por defecto en caso de que no se pueda procesar la fecha
        
        if fecha_str:
            try:
                fecha_obj = datetime.strptime(fecha_str, '%d/%m/%Y')
                fecha_formato_correcto = fecha_obj.strftime('%Y-%m-%d')
            except ValueError:
                # Opcional: Agregar alguna forma de logging o notificación de que la fecha no es válida
                print('Se lo llevó madres')
                pass
        
        numero_cuenta_extraido = datos_extraidos.get('cuenta_retiro', '').strip().lstrip('0')
       
        cuenta_objeto = None
       
        #print('numero_cuenta_extraido',numero_cuenta_extraido)
        if numero_cuenta_extraido:
            try:
                
                cuenta_objeto = Cuenta.objects.get(cuenta__contains=numero_cuenta_extraido)
                print('cuenta_objeto:', cuenta_objeto)
            except Cuenta.DoesNotExist:
                print('Cuenta retiro no encontrada:', numero_cuenta_extraido)
                return JsonResponse({'error': 'Cuenta retiro no encontrada'}, status=404)
            except Exception as e:
                print('Error inesperado al buscar cuenta retiro:', e)
                print(traceback.format_exc())
                return JsonResponse({'error': 'Error interno'}, status=500)
            
        
        
        #print("destino_cuenta",datos_extraidos.get('cuenta_deposito', '').strip().lstrip('0') or None) 
        datos_para_formulario = {
            'monto': datos_extraidos.get('importe_operacion', '').replace('MXP', '').replace(',', '').strip() or None,
            'pagado_real': fecha_formato_correcto,  # Valor procesado o None
            'cuenta': cuenta_objeto.id if cuenta_objeto else None,
            'divisa_cuenta': divisa_cuenta_extraida or None,
            'hora_operacion': datos_extraidos.get('hora_operacion', '') or None,
        }
        
        return JsonResponse(datos_para_formulario)
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


def prellenar_formulario_transferencia(request):
    #print('prellenar_formulario_gastos')
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        pdf_content = request.FILES.get('comprobante_pago')
        print('pdf_content:',pdf_content)
        if not pdf_content:
            return JsonResponse({'error': 'No file uploaded'}, status=400)
        
        pdf_content = pdf_content.read()
        formato_detectado = detectar_formato_pdf(pdf_content)
        
        if formato_detectado == "formato_1":
            print('formato 1')
            texto_extraido = extraer_texto_de_pdf(pdf_content)
            datos_extraidos = encontrar_variables_transferencia(texto_extraido)
           
            #divisa_cuenta_extraida = datos_extraidos.get('divisa_cuenta', '').strip()

        elif formato_detectado == "formato_2":
            print('formato 2')
            bloques = extraer_bloques_formato_2(pdf_content)
            datos_extraidos = encontrar_variables_bloques(bloques)
            datos_extraidos = encontrar_variables(bloques)
        else:
            print('formato no detectado')
        divisa_cuenta_extraida = datos_extraidos.get('divisa_cuenta', '').strip()

        print("Formato detectado:", formato_detectado)
        #print("Datos extraídos iniciales:", datos_extraidos)
        #texto_extraido = extraer_texto_de_pdf(pdf_content)
        #print("Texto extraído:", texto_extraido)
        
        print("Datos extraídos:", datos_extraidos)
        titular_cuenta = datos_extraidos.get('titular_cuenta_2','').strip()
        fecha_str = datos_extraidos.get('fecha', '').strip()
        #titular_cuenta = datos_extraidos.get('titular_cuenta','').strip()
        #print('titular_cuenta:',titular_cuenta)
        fecha_formato_correcto = None  # Valor por defecto en caso de que no se pueda procesar la fecha
        
        if fecha_str:
            try:
                fecha_obj = datetime.strptime(fecha_str, '%d/%m/%Y')
                fecha_formato_correcto = fecha_obj.strftime('%Y-%m-%d')
            except ValueError:
                # Opcional: Agregar alguna forma de logging o notificación de que la fecha no es válida
                print('Se lo llevó madres')
                pass
        
        numero_cuenta_extraido = datos_extraidos.get('cuenta_retiro', '').strip().lstrip('0')
        numero_cuenta_deposito = datos_extraidos.get('cuenta_deposito', '').strip().lstrip('0') 
        cuenta_objeto = None
        cuenta_deposito = None
        #print('numero_cuenta_extraido',numero_cuenta_extraido)
        if numero_cuenta_extraido:
            try:
                if formato_detectado == "formato_1":
                    cuenta_objeto = Cuenta.objects.get(cuenta__contains=numero_cuenta_extraido)
                if formato_detectado == "formato_2":
                    cuenta_objeto = Cuenta.objects.get(cuenta__endswith=numero_cuenta_extraido)
                print('cuenta_objeto:', cuenta_objeto)
            except Cuenta.DoesNotExist:
                print('Cuenta retiro no encontrada:', numero_cuenta_extraido)
                return JsonResponse({'error': 'Cuenta retiro no encontrada'}, status=404)
            except Exception as e:
                print('Error inesperado al buscar cuenta retiro:', e)
                print(traceback.format_exc())
                return JsonResponse({'error': 'Error interno'}, status=500)
            
        if numero_cuenta_deposito:
            try:
                if formato_detectado == "formato_1":
                    cuenta_deposito = Cuenta.objects.get(cuenta__contains = numero_cuenta_deposito)
                if formato_detectado == "formato_2":
                    cuenta_deposito = Cuenta.objects.get(cuenta__endswith = numero_cuenta_deposito)
                print('cuenta_objeto:',cuenta_deposito)
            except Cuenta.DoesNotExist:
                # Manejar el caso donde la cuenta no existe
                print('Fallo cuenta deposito')
                return JsonResponse({'error': 'Account not found'}, status=404)
        
        
        #print("destino_cuenta",datos_extraidos.get('cuenta_deposito', '').strip().lstrip('0') or None) 
        datos_para_formulario = {
            'monto': datos_extraidos.get('importe_operacion', '').replace('MXP', '').replace(',', '').strip() or None,
            'pagado_real': fecha_formato_correcto,  # Valor procesado o None
            'cuenta': cuenta_objeto.id if cuenta_objeto else None,
            'divisa_cuenta': divisa_cuenta_extraida or None,
            'destino_cuenta': cuenta_deposito.id if cuenta_deposito else None,
            'titular_cuenta':titular_cuenta or None,
        }
        
        return JsonResponse(datos_para_formulario)
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


def encontrar_variables_transferencia(texto):
    datos = {}
    #print('si entra a encontrar_transferencia')
    #print('TEXTO EXTRAÍDO:\n', texto)

    if "Cuenta de retiro" in texto:  # Formato 1
        #print("Formato 1 detectado")

        retiro = re.search(r"Cuenta de retiro:\s*([\d\s]+)", texto)
        deposito = re.search(r"Cuenta de depósito:\s*([\d\s]+)", texto)
        importe = re.search(r"(?:Importe|Importe de la operación):\s?([\d,.]+)", texto)
        fecha = re.search(r"Fecha de creación:\s*([\d/]+)", texto)
        divisa = re.search(r"Divisa de la cuenta:\s*(\w+)", texto)
        

        datos['cuenta_retiro'] = retiro.group(1).replace(" ", "") if retiro else None
        datos['cuenta_deposito'] = deposito.group(1).replace(" ", "") if deposito else None
        datos['importe_operacion'] = importe.group(1) if importe else None
        datos['fecha'] = fecha.group(1) if fecha else None
        datos['divisa_cuenta'] = divisa.group(1) if divisa else None

    elif "Datos de la operación" in texto:  # Formato 2
        print("Formato 2 detectado")

        cuentas = re.findall(r"Cuenta:\s+•?(\d+)", texto)
        print("Cuentas encontradas:", cuentas)
        retiro = cuentas[0] if len(cuentas) > 0 else None
        deposito = cuentas[1] if len(cuentas) > 1 else None

        importe = re.search(r"Importe\s*\$?\s*([\d,\.]+)", texto)
        fecha = re.search(r"Fecha y hora de creación:\s*(\d{1,2} de \w+ de \d{4})", texto)

        datos['cuenta_retiro'] = retiro
        datos['cuenta_deposito'] = deposito
        datos['importe_operacion'] = importe.group(1) if importe else None
        datos['fecha'] = fecha.group(1) if fecha else None
        datos['divisa_cuenta'] = 'MXN'

    titulares = re.findall(r"Titular de la cuenta:\s*([^\n\r]+)", texto, re.DOTALL)
    if titulares:
        datos['titular_cuenta_1'] = titulares[0].strip()
        datos['titular_cuenta_2'] = titulares[1].strip() if len(titulares) > 1 else None
    else:
        datos['titular_cuenta_1'] = None
        datos['titular_cuenta_2'] = None

    #print('DATOS EXTRAÍDOS:', datos)
    return datos



@perfil_seleccionado_required
def editar_gasto(request, pk):
    producto = Articulo_Gasto.objects.get(id=pk)

    form = Articulo_Gasto_Edit_Form(instance=producto)

    if request.method =='POST':
        form = Articulo_Gasto_Edit_Form(request.POST, instance=producto)

        if form.is_valid():
            form.save()

            messages.success(request,f'Se ha guardado el artículo {producto} correctamente')
            return HttpResponse(status=204)
        #else:
            #messages.error(request,'Se lo llevo SPM')


    context= {
        'producto': producto,
        'form': form,
        }

    return render(request, 'gasto/editar_gasto.html', context)


@perfil_seleccionado_required
def solicitudes_gasto(request):
    pk_perfil = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk_perfil)

    # Preparar un Prefetch para articulos_gasto
    articulos_gasto_prefetch = Prefetch('articulo_gasto_set', queryset=Articulo_Gasto.objects.filter(completo=True, producto__isnull=False, proyecto__isnull=False), to_attr='articulos_filtrados')
    
    if perfil.tipo.nombre == "Admin":  
        gastos = Solicitud_Gasto.objects.filter(complete=True,  distrito = perfil.distritos).order_by('-created_at') #Temporalmente le metí el filtro de distrito
    elif perfil.tipo.nombre == "Gerente" or perfil.tipo.superintendente == True or perfil.tipo.nombre == "CONTADOR":
        gastos = Solicitud_Gasto.objects.filter(complete=True, distrito = perfil.distritos).order_by('-folio')
    elif perfil.tipo.rh == True and perfil.tipo.documentos == True:    
        gastos = Solicitud_Gasto.objects.filter(complete=True, distrito = perfil.distritos, tipo__tipo__in = ['APOYOS A EMPLEADOS'] ).order_by('-folio')
    else:
        gastos = Solicitud_Gasto.objects.filter(complete=True, staff = perfil).order_by('-folio')

 

    myfilter = Solicitud_Gasto_Filter(request.GET, queryset=gastos)
    gastos = myfilter.qs

    #Set up pagination
    p = Paginator(gastos, 10)
    page = request.GET.get('page')
    gastos_list = p.get_page(page)

    if request.method =='POST' and 'btnExcel' in request.POST:

        return convert_excel_gasto_matriz(gastos)
    
    usuario_view = False

    context= {
        'gastos':gastos,
        'myfilter':myfilter,
        'gastos_list': gastos_list,
        'usuario_view': usuario_view,
        }

    return render(request, 'tesoreria/mis_gastos.html',context)

@perfil_seleccionado_required
def detalle_gastos(request, pk):
    productos = Articulo_Gasto.objects.filter(gasto__id=pk)
    gasto = Solicitud_Gasto.objects.get(id=pk)
    facturas = Factura.objects.filter(solicitud_gasto = gasto )
    

    context= {
        'productos':productos,
        'facturas':facturas,
        'gasto':gasto,
        }

    return render(request, 'gasto/detalle_gasto.html', context)

@perfil_seleccionado_required
def gastos_pendientes_autorizar(request):
    pk = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk)
    
    if perfil.sustituto:
        perfil = Profile.objects.filter(staff=perfil.staff, tipo=perfil.tipo, distritos=perfil.distritos).first()

    solicitudes = Solicitud_Gasto.objects.filter(superintendente=perfil,complete=True, autorizar = None).order_by('-folio')
    ids_solicitudes_validadas = [solicitud.id for solicitud in solicitudes if solicitud.get_validado]

    solicitudes = Solicitud_Gasto.objects.filter(id__in=ids_solicitudes_validadas)

    myfilter=Solicitud_Gasto_Filter(request.GET, queryset=solicitudes)
    solicitudes = myfilter.qs

    for solicitud in solicitudes:
        articulos_gasto = Articulo_Gasto.objects.filter(gasto=solicitud)

        proyectos = set()
        subproyectos = set()

        for articulo in articulos_gasto:
            if articulo.proyecto:
                proyectos.add(str(articulo.proyecto.nombre))
            if articulo.subproyecto:
                subproyectos.add(str(articulo.subproyecto.nombre))

        solicitud.proyectos = ', '.join(proyectos)
        solicitud.subproyectos = ', '.join(subproyectos)
    #Set up pagination
    p = Paginator(solicitudes, 10)
    page = request.GET.get('page')
    ordenes_list = p.get_page(page)

    #Set up pagination
    p = Paginator(solicitudes, 10)
    page = request.GET.get('page')
    ordenes_list = p.get_page(page)

    #if request.method =='POST' and 'btnExcel' in request.POST:

        #return convert_excel_solicitud_matriz(solicitudes)

    context= {
        'perfil':perfil,
        'ordenes_list':ordenes_list,
        'myfilter':myfilter,
        }

    return render(request, 'gasto/pendientes_autorizar_gasto.html', context)

@perfil_seleccionado_required
def gastos_pendientes_autorizar2(request):

    #obtengo el id de usuario, lo paso como argumento a id de profiles para obtener el objeto profile que coindice con ese usuario_id
    pk = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk)    
    
    solicitudes = Solicitud_Gasto.objects.filter(complete=True, autorizar = True, autorizar2 = None, distrito = perfil.distritos).order_by('-folio')

    myfilter=Solicitud_Gasto_Filter(request.GET, queryset=solicitudes)
    solicitudes = myfilter.qs

    for solicitud in solicitudes:
        articulos_gasto = Articulo_Gasto.objects.filter(gasto=solicitud)

        proyectos = set()
        subproyectos = set()

        for articulo in articulos_gasto:
            if articulo.proyecto:
                proyectos.add(str(articulo.proyecto.nombre))
            if articulo.subproyecto:
                subproyectos.add(str(articulo.subproyecto.nombre))

        solicitud.proyectos = ', '.join(proyectos)
        solicitud.subproyectos = ', '.join(subproyectos)

    #Set up pagination
    p = Paginator(solicitudes, 10)
    page = request.GET.get('page')
    ordenes_list = p.get_page(page)

    #if request.method =='POST' and 'btnExcel' in request.POST:

        #return convert_excel_solicitud_matriz(solicitudes)

    context= {
        'ordenes_list':ordenes_list,
        'myfilter':myfilter,
        }

    return render(request, 'gasto/pendientes_autorizar_gasto2.html', context)

@perfil_seleccionado_required
def vales_rosa_pendientes_autorizar(request):

    #obtengo el id de usuario, lo paso como argumento a id de profiles para obtener el objeto profile que coindice con ese usuario_id
    pk = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk)    
    
    if perfil.sustituto:
        perfil = Profile.objects.filter(staff=perfil.staff, tipo=perfil.tipo, distritos=perfil.distritos).first()
    #vales_rosa = ValeRosa.objects.filter(esta_aprobado = None, gasto__superintendente = perfil).order_by('-gasto__folio')
    #print(perfil.tipo.nombre)
    #if perfil.tipo.nombre == "Admin": #Temporalmente hasta que este listo el desarrollo
    if perfil.distritos.nombre == 'MATRIZ' and perfil.tipo.nombre == "Admin":
        print('entra a matriz')
        vales_rosa = ValeRosa.objects.filter(
            Q(gasto__complete=True, gasto__autorizar2=True) |
            Q(viatico__autorizar2=True),
            esta_aprobado=None
        ).order_by(
            '-gasto__folio'  # o '-viatico__folio' si quieres alternar según tipo
        )
    elif perfil.distritos.nombre == 'MATRIZ':
        vales_rosa = ValeRosa.objects.filter(
        esta_aprobado=None
        ).filter(
            Q(
                gasto__isnull=False,
                gasto__complete=True,
                gasto__autorizar2=True,
                gasto__superintendente=perfil
            ) |
            Q(
                viatico__isnull=False,
                viatico__complete=True,
                viatico__autorizar2=True,
                viatico__superintendente=perfil
            )
        ).order_by('-gasto__folio', '-viatico__folio')
    else:
        vales_rosa = ValeRosa.objects.filter(
        esta_aprobado=None
        ).filter(
            Q(
                gasto__isnull=False,
                gasto__complete=True,
                gasto__autorizar2=True,
                gasto__autorizado_por2 = perfil
            ) |
            Q(
                viatico__isnull=False,
                viatico__complete=True,
                viatico__autorizar2=True,
                viatico__gerente = perfil
            )
        ).order_by('-gasto__folio', '-viatico__folio')
    #   vales_rosa = ValeRosa.objects.none()
    #myfilter=Solicitud_Gasto_Filter(request.GET, queryset=solicitudes)
    #solicitudes = myfilter.qs

    #for solicitud in solicitudes:
    #   articulos_gasto = Articulo_Gasto.objects.filter(gasto=solicitud)

    #    proyectos = set()
    #    subproyectos = set()

    #    for articulo in articulos_gasto:
    #        if articulo.proyecto:
    #            proyectos.add(str(articulo.proyecto.nombre))
    #        if articulo.subproyecto:
    #            subproyectos.add(str(articulo.subproyecto.nombre))

    #    solicitud.proyectos = ', '.join(proyectos)
    #    solicitud.subproyectos = ', '.join(subproyectos)


    #Set up pagination
    #p = Paginator(solicitudes, 10)
    #page = request.GET.get('page')
    #ordenes_list = p.get_page(page)

    #if request.method =='POST' and 'btnExcel' in request.POST:

        #return convert_excel_solicitud_matriz(solicitudes)

    context= {
        #'ordenes_list':ordenes_list,
        #'myfilter':myfilter,
        'vales_rosa':vales_rosa,
        }

    return render(request, 'gasto/pendientes_autorizar_vale_rosa.html', context)


@perfil_seleccionado_required
def autorizar_gasto(request, pk):
    #obtengo el id de usuario, lo paso como argumento a id de profiles para obtener el objeto profile que coindice con ese usuario_id
    pk_perfil = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk_perfil)    

    productos = Articulo_Gasto.objects.filter(gasto__id=pk)
    facturas = Factura.objects.filter(solicitud_gasto__id = pk)

    context= {
        'productos':productos,
        'facturas':facturas,
        'pk':pk,
        }

    gasto = Solicitud_Gasto.objects.get(id = pk)
    productos = Articulo_Gasto.objects.filter(gasto = gasto)

    if request.method =='POST' and 'btn_autorizar' in request.POST:
        gasto.autorizar = True
        gasto.approved_at = datetime.now()
        #gasto.approved_at_time = datetime.now().time()
        gasto.sol_autorizada_por = perfil
        if perfil.tipo.subdirector == True:
            gasto.autorizar2 = True
            gasto.approbado_fecha2 = datetime.now()
        for key, value in request.POST.items():
            if key.startswith("vale_"):
                try:
                    vale_id = int(key.split("_")[1])
                    vale = ValeRosa.objects.get(id=vale_id, gasto=gasto)
                    vale.aprobado_por = perfil
                    vale.aprobado_en = datetime.now()
                    if value == "aprobar":
                        vale.esta_aprobado = True
                        
                    elif value == "rechazar":
                        vale.esta_aprobado = False

                    vale.save()
                except (ValueError, ValeRosa.DoesNotExist):
                    continue
        gasto.save()
        messages.success(request, f'{perfil.staff.staff.first_name} {perfil.staff.staff.last_name} has autorizado la solicitud {gasto.folio}')
        return redirect ('gastos-pendientes-autorizar')


    context = {
        'gasto': gasto,
        'productos': productos,
    }

    return render(request,'gasto/autorizar_gasto.html', context)


@perfil_seleccionado_required
def cancelar_gasto(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk_perfil)    
    gasto = Solicitud_Gasto.objects.get(id = pk)
    productos = Articulo_Gasto.objects.filter(gasto = gasto)

    if request.method =='POST' and 'btn_cancelar' in request.POST:
        gasto.autorizar = False
        gasto.approved_at = datetime.now()
        #gasto.approved_at_time = datetime.now().time()
        gasto.sol_autorizada_por = perfil
        gasto.save()
        messages.info(request, f'{perfil.staff.staff.first_name} {perfil.staff.staff.last_name} has cancelado la solicitud {gasto.id}')
        return redirect ('gastos-pendientes-autorizar')

    context = {
        'gasto': gasto,
        'productos': productos,
    }

    return render(request,'gasto/cancelar_gasto.html', context)

@perfil_seleccionado_required
def autorizar_gasto2(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk_perfil)    
    gasto = Solicitud_Gasto.objects.get(id = pk)
    productos = Articulo_Gasto.objects.filter(gasto = gasto)

    if request.method =='POST' and 'btn_autorizar' in request.POST:
        gasto.autorizar2 = True
        gasto.approbado_fecha2 = datetime.now()
        gasto.autorizado_por2 = perfil
        gasto.save()
        messages.success(request, f'{perfil.staff.staff.first_name} {perfil.staff.staff.last_name} has autorizado el gasto {gasto.folio}')
        return redirect ('gastos-pendientes-autorizar2')


    context = {
        'gasto': gasto,
        'productos': productos,
    }

    return render(request,'gasto/autorizar_gasto2.html', context)


@perfil_seleccionado_required
def cancelar_gasto2(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk_perfil)    
    gasto = Solicitud_Gasto.objects.get(id = pk)
    productos = Articulo_Gasto.objects.filter(gasto = gasto)
    form = Autorizacion_Gasto_Form(instance = gasto)

    if request.method =='POST':
        form = Autorizacion_Gasto_Form(request.POST, instance = gasto)
        if form.is_valid():
            gasto = form.save(commit = False)
            gasto.autorizar2 = False
            gasto.autorizado_por2 = perfil
            gasto.approbado_fecha2 = datetime.now()
            #gasto.approved_at_time2 = datetime.now().time()
            gasto.save()
            messages.info(request, f'{perfil.staff.staff.first_name} {perfil.staff.staff.last_name} has cancelado la solicitud {gasto.folio}')
            return HttpResponse(status=204)

    context = {
        'form':form,
        'gasto': gasto,
        'productos': productos,
    }

    return render(request,'gasto/cancelar_gasto2.html', context)

@perfil_seleccionado_required
def autorizar_vale_rosa(request, pk):
    #obtengo el id de usuario, lo paso como argumento a id de profiles para obtener el objeto profile que coindice con ese usuario_id
    pk_perfil = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk_perfil)    

    #productos = Articulo_Gasto.objects.filter(gasto__id=pk)
    vale = ValeRosa.objects.get(id = pk)

    ## Determinar si es gasto o viático
    objeto = vale.gasto if vale.gasto else vale.viatico
    tipo_objeto = 'gasto' if vale.gasto else 'viatico'

    if request.method =='POST' and 'btn_autorizar' in request.POST:
        vale.esta_aprobado = True
        vale.aprobado_en = datetime.now()
        #gasto.approved_at_time = datetime.now().time()
        vale.aprobado_por = perfil
        #if perfil.tipo.subdirector == True:
        #    gasto.autorizar2 = True
        #    gasto.approbado_fecha2 = datetime.now()
        
        vale.save()
        messages.success(request, f'{perfil.staff.staff.first_name} {perfil.staff.staff.last_name} has autorizado la solicitud {vale.id}')
        return redirect ('vales-rosa-pendientes-autorizar')


    context = {
        'vale': vale,
        'objeto': objeto,
        'tipo': tipo_objeto,
    }

    return render(request,'gasto/autorizar_vale_rosa.html', context)

@perfil_seleccionado_required
def cancelar_vale_rosa(request, pk):
    #obtengo el id de usuario, lo paso como argumento a id de profiles para obtener el objeto profile que coindice con ese usuario_id
    pk_perfil = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk_perfil)    

    #productos = Articulo_Gasto.objects.filter(gasto__id=pk)
    vale = ValeRosa.objects.get(id = pk)

    ## Determinar si es gasto o viático
    objeto = vale.gasto if vale.gasto else vale.viatico
    tipo_objeto = 'gasto' if vale.gasto else 'viatico'

    if request.method =='POST' and 'btn_cancelar' in request.POST:
        vale.esta_aprobado = False
        vale.aprobado_en = datetime.now()
        #gasto.approved_at_time = datetime.now().time()
        vale.aprobado_por = perfil
        #if perfil.tipo.subdirector == True:
        #    gasto.autorizar2 = True
        #    gasto.approbado_fecha2 = datetime.now()
        
        vale.save()
        messages.success(request, f'{perfil.staff.staff.first_name} {perfil.staff.staff.last_name} has autorizado la solicitud {vale.id}')
        return redirect ('vales-rosa-pendientes-autorizar')


    context = {
        'vale': vale,
        'objeto': objeto,
        'tipo': tipo_objeto,
    }

    return render(request,'gasto/cancelar_vale_rosa.html', context)

def get_subproyectos(request):
    proyecto_id = request.GET.get('proyecto_id')
    if proyecto_id:
        subproyectos = Subproyecto.objects.filter(proyecto_id=proyecto_id)
        subproyecto_list = list(subproyectos.values('id', 'nombre'))  
        return JsonResponse(subproyecto_list, safe=False)
    return JsonResponse([], safe=False)

@login_required
def ajax_load_proyectos_por_distrito(request):
    distrito_id = request.GET.get('distrito_id')
    proyectos = Proyecto.objects.filter(
        ~Q(status_de_entrega__status="INACTIVO"),
        activo=True,
        distrito_id=distrito_id
    ).values('id', 'nombre')

    return JsonResponse(list(proyectos), safe=False)


# Create your views here.
#@login_required(login_url='user-login')
@perfil_seleccionado_required
def pago_gastos_autorizados(request):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)

    proyectos = Proyecto.objects.filter(activo=True, complete=True)
    subproyectos = Subproyecto.objects.filter(proyecto__in=proyectos)
    
    if usuario.tipo.tesoreria == True:
        if usuario.tipo.rh == True:
            gastos = Solicitud_Gasto.objects.filter( Q(tipo__tipo = "APOYOS A EMPLEADOS")|Q(tipo__tipo = "APOYO DE RENTA"),autorizar=True, pagada=False, distrito = usuario.distritos, autorizar2=True, cerrar_sin_pago_completo = False,para_pago= True,).annotate( 
                total_facturas=Count('facturas', filter=Q(facturas__solicitud_gasto__isnull=False)),autorizadas=Count(Case(When(Q(facturas__autorizada=True, facturas__solicitud_gasto__isnull=False), then=Value(1)))
                )).order_by('-approbado_fecha2')
        else:
            if usuario.distritos.nombre == 'MATRIZ':
                gastos = Solicitud_Gasto.objects.filter(
                     Q(distrito=usuario.distritos) & ~Q(tipo__familia="rh_nomina") | Q(tipo__familia="rh_nomina"),
                    autorizar=True, pagada=False, autorizar2=True, cerrar_sin_pago_completo = False, para_pago = True
                    ).annotate(
                        total_facturas=Count('facturas', filter=Q(facturas__solicitud_gasto__isnull=False)),autorizadas=Count(Case(When(Q(facturas__autorizada=True, facturas__solicitud_gasto__isnull=False), then=Value(1)))
                    )).order_by('-approbado_fecha2')
            else:
                gastos = Solicitud_Gasto.objects.filter(
                    autorizar=True, pagada=False, distrito = usuario.distritos, autorizar2=True, cerrar_sin_pago_completo = False, para_pago = True
                    ).exclude(
                        tipo__familia="rh_nomina"
                    ).annotate(
                        total_facturas=Count('facturas', filter=Q(facturas__solicitud_gasto__isnull=False)),autorizadas=Count(Case(When(Q(facturas__autorizada=True, facturas__solicitud_gasto__isnull=False), then=Value(1)))
                    )).order_by('-approbado_fecha2')
        myfilter = Solicitud_Gasto_Filter(request.GET, queryset=gastos)
        gastos = myfilter.qs

        for gasto in gastos:
            articulos_gasto = Articulo_Gasto.objects.filter(gasto=gasto)

            proyectos = set()
            subproyectos = set()

            for articulo in articulos_gasto:
                if articulo.proyecto:
                    proyectos.add(str(articulo.proyecto.nombre))
                if articulo.subproyecto:
                    subproyectos.add(str(articulo.subproyecto.nombre))

            gasto.proyectos = ', '.join(proyectos)
            gasto.subproyectos = ', '.join(subproyectos)

        p = Paginator(gastos, 50)
        page = request.GET.get('page')
        gastos_list = p.get_page(page)

        for gasto in gastos_list:
            # Determinar estado basado en total_facturas y autorizadas
            if gasto.total_facturas == 0:
                gasto.estado_facturas = 'sin_facturas'
            elif gasto.autorizadas == gasto.total_facturas:
                gasto.estado_facturas = 'todas_autorizadas'
            else:
                gasto.estado_facturas = 'pendientes'

        context= {
            'gastos_list':gastos_list,
            'gastos':gastos,
            'myfilter':myfilter,
            'proyectos': proyectos,
            'subproyectos': subproyectos,
            'selected_subproyecto': request.GET.get('subproyecto')
            }
    else:
        context= {

         }



    if request.method == 'POST' and 'btnReporte' in request.POST:
        return convert_excel_gasto_matriz(gastos)

        

    return render(request, 'gasto/pago_gastos_autorizados.html',context)


@perfil_seleccionado_required
def gastos_por_pagar(request):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    almacenes_distritos = set(usuario.almacen.values_list('distrito__id', flat=True))
    #tipos_prioridad = TipoPrioridad.objects.all()
    if usuario.tipo.cuentas_por_pagar:
        if usuario.distritos.nombre == 'MATRIZ':
            gastos = Solicitud_Gasto.objects.filter(
                Q(distrito=usuario.distritos) & ~Q(tipo__familia="rh_nomina") | Q(tipo__familia="rh_nomina"),
                autorizar=True, pagada=False, autorizar2=True, cerrar_sin_pago_completo = False, para_pago = False
                ).annotate(
                    total_facturas=Count('facturas', filter=Q(facturas__solicitud_gasto__isnull=False)),autorizadas=Count(Case(When(Q(facturas__autorizada=True, facturas__solicitud_gasto__isnull=False), then=Value(1)))
                )).order_by('-approbado_fecha2')
        else:
            gastos = Solicitud_Gasto.objects.filter(
                autorizar=True, pagada=False, distrito = usuario.distritos, autorizar2=True, cerrar_sin_pago_completo = False, para_pago = False
                ).exclude(
                    tipo__familia__in=["rh_nomina", "rh"]
                ).annotate(
                    total_facturas=Count('facturas', filter=Q(facturas__solicitud_gasto__isnull=False)),autorizadas=Count(Case(When(Q(facturas__autorizada=True, facturas__solicitud_gasto__isnull=False), then=Value(1)))
                )).order_by('-approbado_fecha2')
        
    myfilter = Solicitud_Gasto_Filter(request.GET, queryset=gastos)
    gastos = myfilter.qs
    
    for gasto in gastos:
        articulos_gasto = Articulo_Gasto.objects.filter(gasto=gasto)

        proyectos = set()
        subproyectos = set()

        for articulo in articulos_gasto:
            if articulo.proyecto:
                proyectos.add(str(articulo.proyecto.nombre))
            if articulo.subproyecto:
                subproyectos.add(str(articulo.subproyecto.nombre))

        gasto.proyectos = ', '.join(proyectos)
        gasto.subproyectos = ', '.join(subproyectos)

    p = Paginator(gastos, 50)
    page = request.GET.get('page')
    gastos_list = p.get_page(page)

    for gasto in gastos_list:
        # Determinar estado basado en total_facturas y autorizadas
        if gasto.total_facturas == 0:
            gasto.estado_facturas = 'sin_facturas'
        elif gasto.autorizadas == gasto.total_facturas:
            gasto.estado_facturas = 'todas_autorizadas'
        else:
            gasto.estado_facturas = 'pendientes'
    
    if request.method == 'POST' and 'btnReporte' in request.POST:
        #if usuario.tipo.tesoreria or usuario.tipo.finanzas:
        return convert_excel_gasto_matriz(gastos)
        #else:
        #    return convert_excel_matriz_compras_autorizadas(compras)
       
    
    if request.method == 'POST':
        gastos_ids = request.POST.getlist('gastos_ids')
        #print('gastos_ids:',gastos_ids)
        if gastos_ids:
            for gasto_id in gastos_ids:
                parcial = request.POST.get(f'parcial_{gasto_id}')
                #print('parcial',parcial)
                  # Asegurarte de que monto no sea None y que sea un número válido
                if parcial:
                    try:
                        parcial = float(parcial)
                    except ValueError:
                        parcial = 0  # O algún valor por defecto en caso de error
                Solicitud_Gasto.objects.filter(id=gasto_id).update(para_pago=True, manda_pago = usuario, parcial = parcial)
            # Después de la actualización, redirige para restablecer el conteo y sumatoria
            return redirect('gastos-por-pagar')

    context= {
        'usuario':usuario,
        'gastos':gastos,
        'myfilter':myfilter,
        'gastos_list':gastos_list,
        #'tipos_prioridad': tipos_prioridad,
        }

    return render(request, 'gasto/gastos_por_pagar.html',context)



@perfil_seleccionado_required
def finanzas_transferencia_gastos_autorizados(request):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)

    proyectos = Proyecto.objects.filter(activo=True, complete=True)
    subproyectos = Subproyecto.objects.filter(proyecto__in=proyectos)
    
    if usuario.tipo.finanzas == True:
       
        gastos = Solicitud_Gasto.objects.filter(tipo__tipo = "NOMINA",autorizar=True, autorizar2=True, transferencia_finanzas = True, pagada=False).annotate(
            total_facturas=Count('facturas', filter=Q(facturas__solicitud_gasto__isnull=False)),autorizadas=Count(Case(When(Q(facturas__autorizada=True, facturas__solicitud_gasto__isnull=False), then=Value(1)))
            )).order_by('-approbado_fecha2')
        #else:
        #    gastos = Solicitud_Gasto.objects.filter(autorizar=True, pagada=False, distrito = usuario.distritos, autorizar2=True).annotate(
        #        total_facturas=Count('facturas', filter=Q(facturas__solicitud_gasto__isnull=False)),autorizadas=Count(Case(When(Q(facturas__autorizada=True, facturas__solicitud_gasto__isnull=False), then=Value(1)))
        #        )).order_by('-approbado_fecha2')
        myfilter = Solicitud_Gasto_Filter(request.GET, queryset=gastos)
        gastos = myfilter.qs

        for gasto in gastos:
            articulos_gasto = Articulo_Gasto.objects.filter(gasto=gasto)

            proyectos = set()
            subproyectos = set()

            for articulo in articulos_gasto:
                if articulo.proyecto:
                    proyectos.add(str(articulo.proyecto.nombre))
                if articulo.subproyecto:
                    subproyectos.add(str(articulo.subproyecto.nombre))

            gasto.proyectos = ', '.join(proyectos)
            gasto.subproyectos = ', '.join(subproyectos)

        p = Paginator(gastos, 50)
        page = request.GET.get('page')
        gastos_list = p.get_page(page)

        for gasto in gastos_list:
            # Determinar estado basado en total_facturas y autorizadas
            if gasto.total_facturas == 0:
                gasto.estado_facturas = 'sin_facturas'
            elif gasto.autorizadas == gasto.total_facturas:
                gasto.estado_facturas = 'todas_autorizadas'
            else:
                gasto.estado_facturas = 'pendientes'

        context= {
            'gastos_list':gastos_list,
            'gastos':gastos,
            'myfilter':myfilter,
            'proyectos': proyectos,
            'subproyectos': subproyectos,
            'selected_subproyecto': request.GET.get('subproyecto')
            }
    else:
        context= {

         }



    if request.method == 'POST' and 'btnReporte' in request.POST:
        return convert_excel_gasto_matriz(gastos)

        

    return render(request, 'gasto/finanzas_pago_gastos_autorizados.html',context)

@perfil_seleccionado_required
def pago_gasto(request, pk):
    pk_usuario = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_usuario)
    gasto = Solicitud_Gasto.objects.get(id=pk)
    cargos = Tipo_Pago.objects.get(id = 1)
    pagos_alt = Pago.objects.filter(
        gasto=gasto,
        hecho=True
    ).filter(
    Q(tipo=cargos) | Q(tipo__isnull=True)
    )
    cuentas = Cuenta.objects.filter(moneda__nombre = 'PESOS')
    error_messages = []
    
    form = Pago_Gasto_Form()
    remanente = gasto.get_total_solicitud - gasto.monto_pagado
    
    cuentas_para_select2 = [
        {'id': cuenta.id,
         'text': str(cuenta.cuenta), 
        } for cuenta in cuentas]

    if request.method == 'POST':
        if "myBtn" in request.POST:
            pago, created = Pago.objects.get_or_create(tesorero = usuario, gasto__distrito = usuario.distritos, hecho=False, gasto=gasto)
            form = Pago_Gasto_Form(request.POST or None, request.FILES or None, instance = pago)
            if form.is_valid():
                pago = form.save(commit = False)
                 
                pago.pagado_date = datetime.now()
                #pago.pagado_hora = datetime.now().time()
                pago.hecho = True
                monto_actual = pago.monto
                total_pagado = gasto.monto_pagado  + monto_actual
                #gasto.parcial = gasto.parcial - monto_actual
                total_sol = gasto.get_total_solicitud
                TOLERANCIA = Decimal(0.2)
                #El bloque a continuación se generó para resolver los problemas de redondeo, se comparan las dos cantidades redondeadas en una variable y se activa una bandera (flag) que indica si son iguales o no!
                if monto_actual <= 0:
                    messages.error(request, f'El pago {monto_actual} debe ser mayor a 0')
                else:
                     # 1) resta parcialidad
                    nuevo_remanente = gasto.parcial - monto_actual
                    print('nuevo_remanente',nuevo_remanente)
                 # 2) si ya cubriste la parcialidad -> cierra "para pago"
                    if nuevo_remanente < -TOLERANCIA:
                        messages.error(request, "El pago excede la parcialidad más allá de la tolerancia")
                    elif abs(nuevo_remanente) <= TOLERANCIA:
                        gasto.parcial = Decimal("0")
                        gasto.para_pago = False
                    else:
                        gasto.parcial = nuevo_remanente

                    
                    # 3) si ya cubriste el total del gasto -> marca pagada
                    if abs(total_sol - total_pagado) <= TOLERANCIA:
                        gasto.pagada = True

                    gasto.save()  
                    
                    pago.save()
                    pagos = Pago.objects.filter(gasto=gasto, hecho=True)
                    archivo_gasto = attach_gasto_pdf(request, gasto.id)
                    email = EmailMessage(
                        f'Gasto Autorizado {gasto.id}',
                        f'Estimado(a) {gasto.staff.staff.staff.first_name} {gasto.staff.staff.staff.last_name}:\n\nEstás recibiendo este correo porque ha sido pagado el gasto con folio: {gasto.folio}.\n\n\nGrupo Vordcab S.A de C.V.\n\n Este mensaje ha sido automáticamente generado por SAVIA 2.0',
                        'savia@grupovordcab.com',
                        ['ulises_huesc@hotmail.com'], #,gasto.staff.staff.staff.email
                        )
                    email.attach(f'Gasto_folio_{gasto.id}.pdf',archivo_gasto,'application/pdf')
                    email.attach('Pago.pdf',pago.comprobante_pago.read(),'application/pdf')
                    
                    #if pagos.count() > 0:
                        #for item in pagos:
                        #    email.attach(f'Gasto{gasto.folio}_P{item.id}.pdf',item.comprobante_pago.read(),'application/pdf')
                    email.send()

                    messages.success(request,f'Gracias por registrar tu pago, {usuario.staff.staff.first_name}')
                    return redirect('pago-gastos-autorizados')
        if "cerrar_sin_pago" in request.POST:
            print('Está entrando')
            gasto.comentario_cierre = request.POST.get('comentario_cierre')
            gasto.cerrar_sin_pago_completo = True
            gasto.fecha_cierre = date.today()
            gasto.persona_cierre = usuario  # Asegúrate de tener esta variable ya disponible en tu vista
            gasto.save()
            messages.success(request, f'Gasto {gasto.folio} cerrada sin pago completo.')
            return redirect('pago-gastos-autorizados')
        else:
            print('No está entrando')
            for field, errors in form.errors.items():
                error_messages.append(f"{field}: {errors.as_text()}")
            
            print('Error messages:', error_messages)
         

    context= {
        'gasto':gasto,
        #'pago':pago,
        'cuentas_para_select2':cuentas_para_select2,
        'form':form,
        'pagos_alt':pagos_alt,
        'error_messages': error_messages,
        'remanente':remanente,
    }

    return render(request,'gasto/pago_gasto.html',context)

@perfil_seleccionado_required
def finanzas_transferencia(request, pk):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    gasto = Solicitud_Gasto.objects.get(id=pk)
    tipos_pago = Tipo_Pago.objects.all()
    transferencia = tipos_pago.get(id = 3)
    abono = tipos_pago.get(id = 2)
    transaccion, created = Pago.objects.get_or_create(tesorero = usuario, hecho=False, tipo = transferencia, gasto=gasto)
    transaccion2, created = Pago.objects.get_or_create(tesorero = usuario, hecho=False, tipo = abono, gasto=gasto)
    form = Cargo_Abono_Form(instance=transaccion)
    form_transferencia = Transferencia_Form(prefix='abono')
    pagos_alt = Pago.objects.filter(gasto=gasto, hecho=True, tipo = abono)
    remanente = gasto.get_total_solicitud - gasto.monto_pagado_transferencia
    error_messages = []

    #form.fields['tipo'].queryset = Tipo_Pago.objects.filter(id = 3)
    print(Tipo_Pago.objects.filter(id=3))
    cuentas = Cuenta.objects.filter(moneda__nombre = 'PESOS')
      
    cuentas_para_select2 = [
        {'id': cuenta.id,
         'text': str(cuenta.cuenta) +' '+ str(cuenta.moneda), 
         'moneda': str(cuenta.moneda),
        } for cuenta in cuentas]

    if request.method == 'POST':
        if "envio" in request.POST:
            form = Cargo_Abono_Form(request.POST, request.FILES, instance = transaccion)
            form_transferencia = Transferencia_Form(request.POST, instance = transaccion2, prefix='abono')
            
            if form.is_valid() and form_transferencia.is_valid():
                cargo = form.save(commit=False)
                cargo.pagado_date = date.today()
                #cargo.tipo = Tipo_Pago.objects.get(id = 1)
                cargo.pagado_hora = datetime.now().time() 
                cargo.hecho = True
                
                abono = form_transferencia.save(commit=False)
                abono.monto = cargo.monto
                #abono.tipo = Tipo_Pago.objects.get(id = 2)
                abono.comentario = f"{cargo.comentario} (Relacionado con cuenta {cargo.cuenta})"
                abono.pagado_real = cargo.pagado_real
                abono.pagado_date = date.today()
                abono.pagado_hora = datetime.now().time()
                abono.comprobante_pago = cargo.comprobante_pago
                abono.hecho = True
               

                cargo.comentario = f"{cargo.comentario} (Relacionado con cuenta {abono.cuenta})"
               
                
                #Aqui empieza el ciclo de validación de los montos
                total_pagado = round(gasto.monto_pagado_transferencia  + abono.monto,2)
                total_sol = round(gasto.get_total_solicitud,2)
                #El bloque a continuación se generó para resolver los problemas de redondeo, se comparan las dos cantidades redondeadas en una variable y se activa una bandera (flag) que indica si son iguales o no!
                if total_sol == total_pagado:
                    flag = True
                else:
                    flag = False
                if total_pagado > gasto.get_total_solicitud:
                    messages.error(request,f'{usuario.staff.staff.first_name}, el monto introducido más los pagos anteriores superan el monto total del gasto')
                else:
                    if flag:
                        gasto.transferencia_finanzas = False
                        gasto.save()
                    cargo.save()
                    abono.save()
                    messages.success(request,f'{usuario.staff.staff.first_name}, Has agregado correctamente la transferencia')
                    return redirect('transferencia-gastos-autorizados')
            else:
                for field, errors in form.errors.items():
                    error_messages.append(f"{field}: {errors.as_text()}")
                for field, errors in form_transferencia.errors.items():
                    error_messages.append(f"{field}: {errors.as_text()}")

    context= {
        'gasto':gasto,
        'pago':transaccion,
        'form':form,
        'pagos_alt':pagos_alt,
        'remanente':remanente,
        'form_transferencia': form_transferencia,
        'cuentas_para_select2': cuentas_para_select2,
        'error_messages': error_messages,
    }

    return render(request, 'gasto/transferencia_finanzas.html',context)

def generar_archivo_zip(facturas, gasto):
    nombre = gasto.folio if gasto.folio else ''
    zip_filename = f'facturas_compragasto-{nombre}.zip'
    
    # Crear un archivo zip en memoria
    in_memory_zip = io.BytesIO()

    with zipfile.ZipFile(in_memory_zip, 'w') as zip_file:
        for factura in facturas:
            if factura.archivo_pdf:
                pdf_path = factura.archivo_pdf.path
                zip_file.write(pdf_path, os.path.basename(pdf_path))
            if factura.archivo_xml:
                # Generar el PDFreader
                response = generar_cfdi_gasto(None, factura.id)
                pdf_filename = f"{factura.id}.pdf" if factura.id else f"factura_{factura.id}.pdf"
                # Añadir el contenido del PDF al ZIP
                zip_file.writestr(pdf_filename, response.content)
                #Añadir el xml
                xml_path = factura.archivo_xml.path
                zip_file.write(xml_path, os.path.basename(xml_path))

    # Resetear el puntero del archivo en memoria
    in_memory_zip.seek(0)

    return in_memory_zip, zip_filename

@perfil_seleccionado_required
def matriz_facturas_gasto(request, pk):
    pk_usuario = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_usuario)
    gasto = Solicitud_Gasto.objects.get(id = pk)
    articulos_gasto = Articulo_Gasto.objects.filter(gasto = gasto)
    facturas = Factura.objects.filter(solicitud_gasto = gasto, hecho=True)
    vales_rosa = ValeRosa.objects.filter(gasto = gasto)
    pagos = Pago.objects.filter(gasto = gasto)
    form =  Facturas_Gastos_Form(instance=gasto)
    next_url = request.GET.get('next','mis-gastos')


    #if gasto.approbado_fecha2:
        #fecha_aprobacion = gasto.approbado_fecha2.date()

        # último día del mes de aprobación
        #ultimo_dia_mes = calendar.monthrange(
        #    fecha_aprobacion.year,
        #    fecha_aprobacion.month
        #)[1]

        #fecha_corte = fecha_aprobacion.replace(day=ultimo_dia_mes)

        #hoy = timezone.localdate()

        #if hoy > fecha_corte:
        #    gasto.cerrar_facturacion = True
        #    gasto.save(update_fields=['cerrar_facturacion'])
    #factura_form = FacturaForm()
    #next_url = request.GET.get('next', 'mis-gastos') 
    #print(next_url)

    if request.method == 'POST':
        form = Facturas_Gastos_Form(request.POST, instance=gasto)
        if "btn_factura_completa" in request.POST:
            fecha_hora = datetime.today()
            for factura in facturas:
                checkbox_name = f'autorizar_factura_{factura.id}'
                #print("Nombre del checkbox esperado:", checkbox_name)  # Imprimir el nombre esperado
                if checkbox_name in request.POST:
                    print('PASO 1')
                    factura.autorizada = True
                    factura.autorizada_por = usuario
                    factura.autorizada_el = fecha_hora
                else:
                    print('No paso')
                    factura.autorizada = False
                factura.save()
            if form.is_valid():
                gasto = form.save(commit=False)
                gasto.verificacion_facturas = usuario
                gasto.save()
                messages.success(request,'Haz cambiado el status de facturas completas')
                return redirect(next_url) 
            else:
                messages.error(request,'No está validando')
        elif "btn_descargar_todo" in request.POST:
            in_memory_zip, zip_filename = generar_archivo_zip(facturas, gasto)
            response = HttpResponse(in_memory_zip, content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
            return response
        elif 'salir' in request.POST:
            return redirect(next_url)
        elif 'btn_deletevalerosa' in request.POST:
            print('borrando_vale')
            vale_id = request.POST.get('vale_id')
            vale = get_object_or_404(ValeRosa, pk=vale_id)
            #gasto_id = vale.gasto.id  # ValeRosa tiene FK a Gasto
            vale.delete()
            return redirect('matriz-facturas-gasto', gasto.id)

            
    context={
        'next_url':next_url,
        'form':form,
        'pagos':pagos,
        'articulos_gasto':articulos_gasto,
        'gasto':gasto,
        'facturas':facturas,
        'usuario':usuario,
        'vales_rosa':vales_rosa,
        }

    return render(request, 'gasto/matriz_factura_gasto.html', context)

@perfil_seleccionado_required
def facturas_gasto(request, pk):
    articulo = Articulo_Gasto.objects.get(id = pk)
    #facturas = Facturas.objects.filter(pago = pago, hecho=True)
    #factura, created = Facturas.objects.get_or_create(pago=pago, hecho=False)
    #form = Articulo_Gasto_Factura_Form(instance= articulo)

    #if request.method == 'POST':
    #    form = Articulo_Gasto_Factura_Form(request.POST or None, request.FILES or None, instance = articulo)
    #    if form.is_valid():
    #        form.save()
    #        messages.success(request,'Las facturas se subieron de manera exitosa')
    #        return redirect('matriz-compras')
    #    else:
    #        form = Articulo_Gasto_Factura_Form()
    #        messages.error(request,'No se pudo subir tu documento')

    context={
        'articulo':articulo,
        #'form':form,
        
        }

    return render(request, 'gasto/facturas_gasto.html', context)


@perfil_seleccionado_required
def matriz_gasto_entrada(request):
    pk_usuario = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_usuario)

    #articulos_gasto = Articulo_Gasto.objects.all()
    articulos_gasto = Articulo_Gasto.objects.filter(
        Q(producto__nombre = "Materiales (Almacén)")|Q(producto__nombre = "HERRAMIENTA"), 
        completo = True, 
        validacion = False, 
        gasto__autorizar = None, 
        gasto__tipo__tipo='REEMBOLSO',
        gasto__distrito = usuario.distritos
        )

    context={
        'articulos_gasto':articulos_gasto,
        #'form':form,
    }

    return render(request, 'gasto/matriz_entrada_almacen.html', context)

@perfil_seleccionado_required
def gasto_entrada(request, pk):
    pk_usuario = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_usuario)
    #Tengo que revisar primero si ya existe una orden pendiente del usuario
    articulo_gasto = Articulo_Gasto.objects.get(id=pk)
    facturas = Factura.objects.filter(solicitud_gasto=articulo_gasto.gasto)
    entrada, created = Entrada_Gasto_Ajuste.objects.get_or_create(completo= False, almacenista=usuario, gasto = articulo_gasto)
    
    last_order = Order.objects.filter(staff__distritos = usuario.distritos).order_by('-last_folio_number').first()
    productos = Conceptos_Entradas.objects.filter(entrada=entrada, completo = True)
    articulos = Inventario.objects.filter(producto__gasto = False,  distrito = usuario.distritos)
    form_product = Conceptos_EntradasForm()
    form = Entrada_Gasto_AjusteForm()
    
    productos_para_select2 = [
        {
            'id': item.id,
            'text': str(item.producto.nombre),
            'iva': str(item.producto.iva)
        } for item in articulos
    ]


    if request.method =='POST':
        if "input_agregar" in request.POST:
            form = Entrada_Gasto_AjusteForm(request.POST, instance = entrada)
            if form.is_valid():
                #El elemento entrada es el principal y es el objeto 
                entrada = form.save(commit=False)
                entrada.completo = True
                entrada.completado_fecha = datetime.now()
                entrada.save()
                articulo_gasto.validacion = True
                articulo_gasto.save()
                messages.success(request, f'La entrada del gasto {entrada.id} ha sido creada')
               
                #abrev= usuario.distritos.abreviado <----- el folio ya no lleva
                if last_order == None or last_order.last_folio_number==None:
                    #No hay órdenes para este distrito todavía
                    folio_number = 1
                else:
                    folio_number = last_order.last_folio_number + 1
                last_folio_number = folio_number

                #Se crea una solicitud para poder despachar los artículos
                tipo = Tipo_Orden.objects.get(tipo ='normal')
                folio = folio_number
                orden_producto, created = Order.objects.get_or_create(staff = articulo_gasto.staff, complete = None, distrito = articulo_gasto.staff.distritos)
                orden_producto.folio = folio
                orden_producto.tipo = tipo
                orden_producto.last_folio_number = last_folio_number
                orden_producto.created_at = datetime.now()
                orden_producto.approved_at = datetime.now()
                orden_producto.autorizar = True
                orden_producto.supervisor = articulo_gasto.staff
                orden_producto.superintendente = articulo_gasto.gasto.superintendente
                operacion = Operacion.objects.get(nombre="GASTO")
                orden_producto.operacion = operacion
                orden_producto.complete = True
                #Esta parte es un poco confusa, porque los articulos no siempre están dirigidos al mismo proyecto y subproyecto
                orden_producto.proyecto = articulo_gasto.proyecto
                orden_producto.subproyecto = articulo_gasto.subproyecto
                orden_producto.save()
                #----------------------------------------------------------------#
                #Los productos son cada uno de los items contenidos en la entrada por ajuste y son un objeto "inventario"
                #por cada uno de los productos se va a hacer lo siguiente
                for item_producto in productos:

                    producto_inventario = Inventario.objects.get(producto= item_producto.concepto_material.producto)
                    #productos_por_surtir = ArticulosparaSurtir.objects.filter(articulos__producto=producto_inventario, requisitar = True)
                    articulo_ordenado = ArticulosOrdenados.objects.create(producto=producto_inventario, orden = orden_producto, cantidad=item_producto.cantidad)
                    productos_por_surtir = ArticulosparaSurtir.objects.create(
                        articulos = articulo_ordenado,
                        cantidad=item_producto.cantidad,
                        precio = item_producto.precio_unitario,
                        surtir=True,
                        comentario="esta solicitud es proveniente de un gasto",
                        created_at=datetime.now(),
                        #created_at_time=datetime.now().time(),
                    )
                    #Calculo el precio  y agrega al inventario

                    if producto_inventario.price == 0:
                        producto_inventario.price = item_producto.precio_unitario
                    producto_inventario.price = ((item_producto.precio_unitario * item_producto.cantidad)+ ((producto_inventario.cantidad_apartada + producto_inventario.cantidad) * producto_inventario.price))/(producto_inventario.cantidad + item_producto.cantidad + producto_inventario.cantidad_apartada)
                    #La cantidad en inventario + la cantidad del producto en la entrada <-----esta parte es la que no veo sucediendo
                    producto_inventario.cantidad_apartada = producto_inventario.cantidad_apartada + item_producto.cantidad
                    #producto_inventario.save()
                    producto_inventario._change_reason = f'Esta es una entrada desde un gasto {item_producto.id}'
                    producto_inventario.save()

                static_path = settings.STATIC_ROOT
                img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
                img_path2 = os.path.join(static_path,'images','logo_vordcab.jpg')
       
                image_base64 = get_image_base64(img_path)
                logo_v_base64 = get_image_base64(img_path2)
                # Crear el mensaje HTML
                html_message = f"""
                <html>
                    <head>
                        <meta charset="UTF-8">
                    </head>
                    <body>
                        <p><img src="data:image/jpeg;base64,{logo_v_base64}" alt="Imagen" style="width:100px;height:auto;"/></p>
                        <p>Estimado {articulo_gasto.staff.staff.staff.first_name} {articulo_gasto.staff.staff.staff.last_name},</p>
                        <p>Estás recibiendo este correo porque tu gasto folio:{articulo_gasto.gasto} {articulo_gasto.producto.nombre} ha sido validado</p>
                        <p>y ha recibido entrada de almacén por {usuario.staff.staff.first_name} {usuario.staff.staff.last_name}.</p>
                        <p>Favor de pasar a firmar el vale de salida para terminar con este proceso.</p>
                        <p><img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width:50px;height:auto;border-radius:50%"/></p>
                        <p>Este mensaje ha sido automáticamente generado por SAVIA 2.0</p>
                    </body>
                </html>
                """
                email = EmailMessage(
                    f'Entrada de producto por gasto: {articulo_gasto.producto.nombre} |Gasto: {articulo_gasto.gasto.id}',
                    body=html_message,
                    from_email = 'savia@vordcab.com',
                    to = ['ulises_huesc@hotmail.com'],#articulo_gasto.staff.staff.email],
                    headers={'Content-Type': 'text/html'}
                    )
                email.content_subtype = "html " # Importante para que se interprete como HTML
                #email.send()
                
                return redirect('matriz-gasto-entrada')
        if "input_producto" in request.POST:
            articulo, created = Conceptos_Entradas.objects.get_or_create(completo = False, entrada = entrada)
            form_product = Conceptos_EntradasForm(request.POST, instance=articulo)
            if form_product.is_valid():
                articulo = form_product.save(commit=False)
                articulo.completo = True
                articulo.save()
                messages.success(request, 'Has guardado exitosamente un artículo')
                return redirect('gasto-entrada',pk= pk)

    context= {
        'facturas':facturas,
        'articulo_gasto':articulo_gasto,
        'productos_para_select2':productos_para_select2,
        'productos':productos,
        'form':form,
        'form_product': form_product,
        'articulos':articulos,
        'entrada':entrada,
    }

    return render(request, 'gasto/crear_entrada.html', context)


@perfil_seleccionado_required
def delete_articulo_entrada(request, pk):
   
    articulo = Conceptos_Entradas.objects.get(id=pk)
    gasto = articulo.entrada.gasto.id
    messages.success(request,f'El articulo {articulo.concepto_material} ha sido eliminado exitosamente')
    articulo.delete()

    return redirect('gasto-entrada',pk= gasto)

@perfil_seleccionado_required
def descargar_pdf_gasto(request, pk):
    gasto = get_object_or_404(Solicitud_Gasto, id=pk)
    buf = render_pdf_gasto(gasto.id)
    return FileResponse(buf, as_attachment=True, filename='gasto_' + str(gasto.folio) + '.pdf')

def attach_gasto_pdf(request, pk):
    gasto = get_object_or_404(Solicitud_Gasto, id=pk)
    buf = render_pdf_gasto(gasto.id)

    return buf.getvalue()

def render_pdf_gasto(pk):
    #Configuration of the PDF object
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    #Here ends conf.
    gasto = Solicitud_Gasto.objects.get(id=pk)
    productos = Articulo_Gasto.objects.filter(gasto=gasto, completo=True)
    facturas = Factura.objects.filter(solicitud_gasto = gasto, hecho = True)
    vales = ValeRosa.objects.filter(gasto = gasto)

    data_facturas = [['Datos de XML', 'Nombre', 'Monto']]  # Encabezados de la tabla de facturas
    total_facturas = 0
    total_vales = 0
    total_comprobado = 0
    suma_total = Decimal('0.00')

    styles = getSampleStyleSheet()
    normal_style = styles['Normal']
    custom_style = ParagraphStyle(
        name='CustomStyle',
        parent=normal_style,
        fontSize=10,
        alignment=1,  # Alineación centrada (0 = izquierda, 1 = centro, 2 = derecha)
    )
    for vale in vales:
        if vale.esta_aprobado:
            total_vales += float(vale.monto)
    # Iterar sobre cada factura y sumar el total
    for factura in facturas:
        try:
            if not factura.archivo_xml:  # Verifica si hay un archivo asociado
                continue  # Si no hay archivo, salta a la siguiente factura
            datos_emisor = factura.emisor  # Llamar a la propiedad 'emisor'
            # Acceder directamente a los datos de XML
            resultados = datos_emisor.get('resultados', [])
            nombre = datos_emisor.get('nombre', 'No disponible')
            total_xml_str = datos_emisor.get('total', '0.00')  # Obtener el total o usar '0.00' como predeterminado
            #Para el total de todas las factuas
            total = datos_emisor.get('total', 0.0)  # Obtener el total o usar 0.0 si no está disponible
            total_facturas += float(total)  # Sumar el total al total general
            try:
                total_factura = Decimal(total_xml_str)  # Convertir a Decimal
            except (InvalidOperation, ValueError):
                total_factura = Decimal('0.00')  # Si no es convertible, usar 0.00
            
            # Sumar al total acumulado
            suma_total += total_factura

            # Añadir los datos a la lista
            data_facturas.append([
                Paragraph(str(resultados), custom_style), 
                Paragraph(nombre, custom_style),
                Paragraph(f"${total_factura:,.2f}", custom_style)  # Formatear el total como una cadena de texto
            ])
        except (AttributeError, FileNotFoundError) as e:

            continue  # Si 'emisor' no existe, saltar a la siguiente iteración

    print('MONTOOOOO')
    print(total_facturas)

   #Azul Vordcab
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)
    caja_iso = 760
    #Elaborar caja
    #c.line(caja_iso,500,caja_iso,720)



    #Encabezado
    c.drawString(420,caja_iso,'Preparado por:')
    c.drawString(420,caja_iso-10,'SUP. ADMON')
    c.drawString(520,caja_iso,'Aprobación')
    c.drawString(520,caja_iso-10,'SUB ADM')
    c.drawString(150,caja_iso-20,'Número de documento')
    #c.drawString(160,caja_iso-30,'F-ADQ-N4-01.02')
    c.drawString(245,caja_iso-20,'Clasificación del documento')
    #c.drawString(275,caja_iso-30,'Controlado')
    c.drawString(355,caja_iso-20,'Nivel del documento')
    c.drawString(380,caja_iso-30, 'N5')
    c.drawString(440,caja_iso-20,'Revisión No.')
    c.drawString(452,caja_iso-30,'000')
    c.drawString(510,caja_iso-20,'Fecha de Emisión')
    c.drawString(525,caja_iso-30,'01/2024')

    caja_proveedor = caja_iso - 65
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(150,750,250,20, fill=True, stroke=False) #Barra azul superior Solicitud
    c.rect(20,caja_proveedor - 8,565,20, fill=True, stroke=False) #Barra azul superior Proveedor | Detalle
    c.rect(20,575,565,2, fill=True, stroke=False) #Linea posterior horizontal
    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(280,755,'Comprobación de Gastos')
    c.setLineWidth(.3) #Grosor
    c.line(20,caja_proveedor-8,20,575) #Eje Y donde empieza, Eje X donde empieza, donde termina eje y,donde termina eje x (LINEA 1 contorno)
    c.line(585,caja_proveedor-8,585,575) #Linea 2 contorno
    c.drawInlineImage('static/images/logo_vordcab.jpg',45,730, 3 * cm, 1.5 * cm) #Imagen vortec

    c.setFillColor(white)
    c.setFont('Helvetica-Bold',11)
    #c.drawString(120,caja_proveedor,'Infor')
    c.drawString(300,caja_proveedor, 'Detalles')
    inicio_central = 300
    #c.line(inicio_central,caja_proveedor-25,inicio_central,520) #Linea Central de caja Proveedor | Detalle
    c.setFillColor(black)
    c.setFont('Helvetica',9)
    c.drawString(30,caja_proveedor-20,'Solicitó:')
    c.drawString(30,caja_proveedor-40,'Distrito:')
    c.drawString(30,caja_proveedor-60,'Clase')
    c.drawString(30,caja_proveedor-80,'Banco:')
    c.drawString(30,caja_proveedor-100,'Fecha:')
    # Segunda columna del encabezado
    c.drawString(280,caja_proveedor-60,'Depositar a:')
    c.drawString(280,caja_proveedor-20,'Cuenta:')
    c.drawString(280,caja_proveedor-40,'Clabe:')
    c.drawString(280,caja_proveedor-80,'Empresa')


    
    c.setFont('Helvetica-Bold',12)
    c.drawString(500,caja_proveedor-20,'FOLIO:')
    c.setFillColor(rojo)
    c.setFont('Helvetica-Bold',12)
    c.drawString(540,caja_proveedor-20, str(gasto.folio))

    c.setFillColor(black)
    c.setFont('Helvetica',9)
    c.drawString(100,caja_proveedor-20, gasto.staff.staff.staff.first_name+' '+ gasto.staff.staff.staff.last_name)
    c.drawString(100,caja_proveedor-40, gasto.staff.distritos.nombre)
    if gasto.tipo:
        c.drawString(100,caja_proveedor-60, gasto.tipo.tipo)
    else: 
        c.drawString(100,caja_proveedor-60, "Sin registro")
    
    if gasto.approved_at:
        c.drawString(100,caja_proveedor-100, gasto.approved_at.strftime("%d/%m/%Y"))
    # Segunda Columna del encabezado
    if gasto.colaborador:
        print(gasto.colaborador)
        c.drawString(350,caja_proveedor-60,gasto.colaborador.staff.staff.first_name+' '+ gasto.colaborador.staff.staff.last_name)
        if gasto.colaborador.staff.cuenta_bancaria:
            c.drawString(350,caja_proveedor-20,str(gasto.colaborador.staff.cuenta_bancaria))
        else:
            c.drawString(350,caja_proveedor-20, "Sin registro")
        if gasto.colaborador.staff.clabe:
            c.drawString(350,caja_proveedor-40,str(gasto.colaborador.staff.clabe))
        else:
            c.drawString(350,caja_proveedor-40, "Sin registro")
        if gasto.colaborador.staff.banco:
            c.drawString(100,caja_proveedor-80, gasto.colaborador.staff.banco.nombre)
        else:
            c.drawString(100,caja_proveedor-80, "Sin registro")
        
    
        if gasto.colaborador.staff.empresa:
            c.drawString(350,caja_proveedor-80, gasto.colaborador.staff.empresa.nombre)
        else:
            c.drawString(350,caja_proveedor-80, "Sin registro")

    else:
        c.drawString(350,caja_proveedor-60,gasto.staff.staff.staff.first_name+' '+ gasto.staff.staff.staff.last_name)
        if gasto.staff.staff.cuenta_bancaria:
            c.drawString(350,caja_proveedor-20,str(gasto.staff.staff.cuenta_bancaria))
        else:
            c.drawString(350,caja_proveedor-20, "Sin registro")
        if gasto.staff.staff.clabe:
            c.drawString(350,caja_proveedor-40,str(gasto.staff.staff.clabe))
        else:
            c.drawString(350,caja_proveedor-40, "Sin registro")
        if gasto.staff.staff.banco:
            c.drawString(100,caja_proveedor-80, gasto.staff.staff.banco.nombre)
        else:
            c.drawString(100,caja_proveedor-80, "Sin registro")
        if gasto.staff.staff.empresa:
            c.drawString(350,caja_proveedor-80,gasto.staff.staff.empresa.nombre)
        else:
            c.drawString(350,caja_proveedor-80, "Sin registro")

    #Create blank list
    data =[]

    data.append(['''Código''', '''Nombre''', '''Cantidad''','''Precio''', '''Subtotal''', '''Total''','''Comentario'''])


    high = 540
    for producto in productos:
        # Convert to Decimal and round to two decimal places
        cantidad = producto.cantidad if producto.cantidad is not None else 0
        cantidad_redondeada = Decimal(cantidad).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        precio = producto.precio_unitario if producto.cantidad is not None else 0
        precio_unitario_redondeado = Decimal(precio).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        subtotal = Decimal(cantidad_redondeada * precio_unitario_redondeado).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        otros_impuestos = producto.otros_impuestos if producto.otros_impuestos is not None else 0
        total = Decimal(subtotal) + Decimal(otros_impuestos)
        data.append([
            producto.producto.codigo, 
            producto.producto.nombre,
            cantidad_redondeada, 
            precio_unitario_redondeado,
            subtotal, 
            total,
            producto.comentario,
            ])
        high = high - 18


    c.setFillColor(prussian_blue)
    c.rect(20,30,565,30, fill=True, stroke=False)
    c.setFillColor(white)
    #Primer renglón
    c.drawCentredString(70,48,'Clasificación:')
    c.drawCentredString(140,48,'Nivel:')
    c.drawCentredString(240,48,'Preparado por:')
    c.drawCentredString(350,48,'Aprobado:')
    c.drawCentredString(450,48,'Fecha emisión:')
    c.drawCentredString(550,48,'Rev:')
    #Segundo renglón
    c.drawCentredString(70,34,'Controlado')
    c.drawCentredString(140,34,'N5')
    c.drawCentredString(240,34,'SEOV-ALM-N4-01-01')
    c.drawCentredString(350,34,'SUB ADM')
    c.drawCentredString(450,34,'24/Oct/2018')
    c.drawCentredString(550,34,'001')

    c.setFillColor(black)
    width, height = letter
    styles = getSampleStyleSheet()
    styleN = styles["BodyText"]

    if gasto.comentario is not None:
        comentario = gasto.comentario
    else:
        comentario = "No hay comentarios"

   
    c.setFillColor(prussian_blue)
    c.rect(20,30,565,30, fill=True, stroke=False)
    c.setFillColor(white)
    # Personalizar el estilo de los párrafos
    custom_style = ParagraphStyle(
    'CustomStyle',
        parent=styles['BodyText'],
        fontSize=6,  # Reducir el tamaño de la fuente a 6
        leading=8,   # Aumentar el espacio entre líneas para asegurar que el texto no se superponga
        alignment=TA_LEFT,  # Alineación del texto
        # Puedes añadir más ajustes si es necesario
    )
    for i, row in enumerate(data):
        for j, item in enumerate(row):
            if i!=0 and j == 6:
                text = '' if item is None else str(item)
                data[i][j] = Paragraph(text, custom_style)

    table = Table(data, colWidths=[1.2 * cm, 6 * cm, 1.5 * cm, 1.5 * cm, 1.5 * cm, 1.5* cm, 6 * cm,])
    table_style = TableStyle([ #estilos de la tabla
        ('INNERGRID',(0,0),(-1,-1), 0.25, colors.white),
        ('BOX',(0,0),(-1,-1), 0.25, colors.black),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        #ENCABEZADO
        ('TEXTCOLOR',(0,0),(-1,0), white),
        ('FONTSIZE',(0,0),(-1,0), 8),
        ('BACKGROUND',(0,0),(-1,0), prussian_blue),
        #CUERPO
        ('TEXTCOLOR',(0,1),(-1,-1), colors.black),
        ('FONTSIZE',(0,1),(-1,-1), 6),
        ])
    table.setStyle(table_style)

    #pdf size
    pos_table1 = high-20
    table.wrapOn(c, width, height)
    table.drawOn(c, 20, pos_table1)
    # Crear una lista de datos para la tabla secundaria
    data_secundaria = []
    data_secundaria.append(['Proyecto', 'Subproyecto'])  # Encabezados de la tabla secundaria

    # Añadir filas de proyectos y subproyectos
    for producto in productos:
        if producto.proyecto is None:
            data_secundaria.append(['None','None'])
        else:
            data_secundaria.append([producto.proyecto.nombre, producto.subproyecto.nombre])

    # Crear la tabla secundaria
    table_secundaria = Table(data_secundaria, colWidths=[7 * cm, 7 * cm])  # Ajusta las medidas según necesites

    # Estilo para la tabla secundaria
    table_secundaria_style = TableStyle([
        ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
        ('BOX', (0,0), (-1,-1), 0.25, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BACKGROUND', (0,0), (1,0), colors.grey),  # Fondo gris para los encabezados
        ('TEXTCOLOR', (0,0), (1,0), colors.whitesmoke),  # Texto blanco para los encabezados
        ('TEXTCOLOR', (0,1), (-1,-1), colors.black),  # Texto negro para el cuerpo
        ('FONTSIZE', (0,0), (-1,-1), 8),  # Tamaño de fuente para toda la tabla
        # Añade aquí más estilos si lo necesitas
    ])

    table_secundaria.setStyle(table_secundaria_style)

    # Posición de la tabla secundaria en el PDF
    x_pos = 20  # Ajusta la posición X como sea necesario
    y_pos = pos_table1 - (len(data) * 18) - 20  # Ajusta la posición Y según el espacio ocupado por la primera tabla y cualquier otro contenido

    # Dibujar la tabla secundaria en el canvas
    table_secundaria.wrapOn(c, width, height)
    table_secundaria.drawOn(c, x_pos, y_pos)

    """
    # Crear la tabla de facturas
    table_factura = Table(data_facturas, colWidths=[11 * cm, 6 * cm, 2 * cm])

   # Estilo para la tabla secundaria
    table_facturas_style = TableStyle([ #estilos de la tabla
        ('INNERGRID',(0,0),(-1,-1), 0.25, colors.white),
        ('BOX',(0,0),(-1,-1), 0.25, colors.black),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        #ENCABEZADO
        ('TEXTCOLOR',(0,0),(-1,0), white),
        ('FONTSIZE',(0,0),(-1,0), 8),
        ('BACKGROUND',(0,0),(-1,0), prussian_blue),
        #CUERPO
        ('TEXTCOLOR',(0,1),(-1,-1), colors.black),
        ('FONTSIZE',(0,1),(-1,-1), 6),
        ])
    

    table_factura.setStyle(table_facturas_style)
    """
    #Parrafó de totales
    data_totales = []
    total_comprobado = total_facturas + total_vales
    diferencia_totales = total_comprobado - float(gasto.get_total_solicitud)
    if diferencia_totales > 0:
        color_diferencia = colors.green
    elif diferencia_totales < 0:
        color_diferencia = colors.red
    else:
        color_diferencia = colors.black 
    total_str = "${:,.2f}".format(total_comprobado)  # Convierte Decimal a string y formatea
    # 4. Posición de la tabla de facturas en el PDF
    # Asumiendo que 'y_pos' es la posición Y después de dibujar la tabla secundaria y cualquier otro contenido
    

    data_totales = [
    ['Total solicitado', 'Total Facturado', 'Total No deducible', 'Total Comprobado', 'Saldo A cargo/Favor en Pesos'],  # Encabezados
    ['$' + str(gasto.get_total_solicitud), f"${total_facturas:,.2f}", f"${total_vales:,.2f}",f"${total_comprobado:,.2f}", Paragraph(f'${diferencia_totales:,.2f}', ParagraphStyle('CustomStyle', textColor=color_diferencia))]
    ]

    #data_totales.append(['Total solicitado', 'Total comprobado', 'Saldo A cargo/Favor en Pesos'])  # Encabezados de la tabla secundaria
    #data_totales.append(['$' + str(gasto.get_total_solicitud), total_str, '$' + str(diferencia_totales) ])
    table_totales = Table(data_totales, colWidths=[3 * cm, 3 * cm, 3 * cm, 3 * cm, 5 * cm])  # Ajusta las medidas según necesites
    table_totales.setStyle(table_secundaria_style)
    # Añadir filas de proyectos y subproyectos
   
    table_totales.wrapOn(c, width, height)
    y_totales_pos = y_pos - (len(data_totales) * 15 + 30) 
    table_totales.drawOn(c, 20, y_totales_pos)

    c.setFillColor(prussian_blue)
    c.rect(20, y_totales_pos-50,565,25, fill=True, stroke=False)
    c.setFillColor(white)
    c.drawCentredString(320, y_totales_pos-45,'Observaciones')
    c.setFillColor(black)
    options_conditions_paragraph = Paragraph(comentario, styleN)
    # Crear un marco (frame) en la posición específica

    frame = Frame(50, 0, width, y_totales_pos-45, id='normal')

    # Agregar el párrafo al marco
    frame.addFromList([options_conditions_paragraph], c)
    c.drawCentredString(230, y_totales_pos-190, gasto.staff.staff.staff.first_name +' '+ gasto.staff.staff.staff.last_name)
    c.line(180, y_totales_pos-195,280,  y_totales_pos-195)
    c.drawCentredString(230,  y_totales_pos-205, 'Solicitado')
   
    
    if gasto.staff.distritos.nombre == "MATRIZ":  
        if gasto.autorizar:
            c.drawCentredString(410, y_totales_pos-190, gasto.superintendente.staff.staff.first_name +' '+ gasto.superintendente.staff.staff.last_name)
        if gasto.autorizar is None:
            c.setFillColor(colors.orange) 
            c.drawCentredString(410, y_totales_pos-190, 'No autorizado aún')
        if gasto.autorizar == False:
            c.setFillColor(colors.red)
            c.drawCentredString(410, y_totales_pos-190, 'Cancelado')
    elif gasto.autorizar2 and gasto.autorizado_por2:
        c.drawCentredString(410, y_totales_pos-190, gasto.autorizado_por2.staff.staff.first_name +' '+ gasto.autorizado_por2.staff.staff.last_name)
    elif gasto.autorizar2:
        c.drawCentredString(410, y_totales_pos-190, gasto.superintendente.staff.staff.first_name +' '+ gasto.superintendente.staff.staff.last_name)
    elif gasto.autorizar2 is None:
        c.setFillColor(colors.orange) 
        c.drawCentredString(410, y_totales_pos-190, 'No autorizado aún')
    elif gasto.autorizar2 == False:
        c.setFillColor(colors.red)
        c.drawCentredString(410, y_totales_pos-190, 'Cancelado')
    c.setFillColor(black)
    c.line(360,  y_totales_pos-195,460,  y_totales_pos-195)
    c.drawCentredString(410, y_totales_pos-205,'Aprobado por')


    c.showPage()
    #y_facturas_pos =height - (len(data_facturas) * 60)  # Ajusta según sea necesario
    #table_factura.wrapOn(c, width, height)
    #table_factura.drawOn(c, 20, y_facturas_pos)
    # Crear la tabla de facturas
    rows_per_page = 11
    total_rows = len(data_facturas)
    first = True

    for page in range((total_rows - 1) // rows_per_page + 1):
        start_row = page * rows_per_page
        end_row = start_row + rows_per_page + 1  # +1 para incluir el encabezado
        page_data = data_facturas[start_row:end_row]  # Datos de la página actual

        table_factura = Table(page_data, colWidths=[11 * cm, 6 * cm, 3 * cm])
        
        # Estilo para la tabla
        table_facturas_style = TableStyle([
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.white),
            ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])
        
        if first:
            table_facturas_style.add('TEXTCOLOR', (0, 0), (-1, 0), colors.white)
            table_facturas_style.add('FONTSIZE', (0, 0), (-1, 0), 8)
            table_facturas_style.add('BACKGROUND', (0, 0), (-1, 0), prussian_blue)
            first = False  # Cambiar el estado del booleano para las siguientes páginas
        
        # Estilo del cuerpo
        table_facturas_style.add('TEXTCOLOR', (0, 1), (-1, -1), colors.black)
        table_facturas_style.add('FONTSIZE', (0, 1), (-1, -1), 6)
        table_factura.setStyle(table_facturas_style)

        # Obtener la altura de la página
        page_height = c._pagesize[1]

        # Calcular la altura total que ocupará la tabla
        table_height = len(page_data) * 60  # Ajusta el valor 20 según el tamaño de cada fila

        # Calcular la posición vertical para dibujar la tabla
        y_position = page_height - table_height - 10  # Deja espacio en la parte superior
        table_factura.wrapOn(c, c._pagesize[0], c._pagesize[1])
        table_factura.drawOn(c, 20, y_position)

        # Si no es la última página, añadir una nueva página
        if page < (total_rows - 1) // rows_per_page:
            c.showPage()  # Crear una nueva página
    
    

    c.save()
    buf.seek(0)

    return buf


def convert_excel_gasto_matriz(gastos):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Gastos_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Gastos')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')
    percent_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(percent_style)

    columns = ['Folio','Fecha Autorización','Distrito','Proyectos','Subproyectos','Comentarios','Colaborador','Solicitado para',
               'Importe','Fecha Creación','Status','Autorizado por','Tiene Facturas','Status de Pago']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16
        if col_num == 5: #Columna del proveedor
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        if col_num == 2:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 20

    columna_max = len(columns)+2

    # Agregar los mensajes
    ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por SAVIA 2.0. UH}').style = messages_style
    ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}').style = messages_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 30

    # Agregar los encabezados de las nuevas columnas debajo de los mensajes
    ws.cell(row=3, column = columna_max, value="Total de Gastos").style = head_style
    ws.cell(row=4, column = columna_max, value="Sumatoria de Pagos Pendientes").style = head_style
   

    # Asumiendo que las filas de datos comienzan en la fila 2 y terminan en row_num
    ws.cell(row=3, column=columna_max + 1, value=f"=COUNTA(A:A)-1").style = body_style
    ws.cell(row=4, column=columna_max + 1, value=f"=SUM(I:I)").style = money_resumen_style
  

   
    
    for gasto in gastos:
        row_num = row_num + 1    
        
        # Manejar autorizado_at_2
        if gasto.approbado_fecha2 and isinstance(gasto.approbado_fecha2, datetime):
        # Si autorizado_at_2 es timezone-aware, conviértelo a timezone-naive
            autorizado_at_2_naive = gasto.approbado_fecha2.astimezone(pytz.utc).replace(tzinfo=None)
        else:
            autorizado_at_2_naive = ''
        
        # Manejar created_at
        if gasto.created_at and isinstance(gasto.created_at, datetime):
        # Si created_at es timezone-aware, conviértelo a timezone-naive
           created_at_naive = gasto.created_at.astimezone(pytz.utc).replace(tzinfo=None)
        else:
            created_at_naive = ''

        if gasto.pagada:
            pagada = "Tiene Pago"
        else: 
            pagada ="No tiene pago"

        if gasto.facturas.filter(archivo_xml__isnull=False).exists():
            tiene_facturas = 'Sí'
        else:
            tiene_facturas = 'No'
        
        if gasto.autorizar2:
            status = "Autorizado"
            if gasto.distrito:
                if gasto.distrito.nombre == "MATRIZ":
                    autorizado_por = str(gasto.superintendente.staff.staff.first_name) + ' ' + str(gasto.superintendente.staff.staff.last_name)
            else:
                if gasto.autorizado_por2:
                    autorizado_por = str(gasto.autorizado_por2.staff.staff.first_name) + ' ' + str(gasto.autorizado_por2.staff.staff.last_name)
                else:
                    autorizado_por ="NR"
        elif gasto.autorizar2 == False:
            status = "Cancelado"
            if gasto.distrito.nombre == "MATRIZ":
                autorizado_por = str(gasto.superintendente.staff.staff.first_name) + ' ' + str(gasto.superintendente.staff.staff.last_name)
            else:
                autorizado_por = str(gasto.autorizado_por2.staff.staff.first_name) + ' ' + str(gasto.autorizado_por2.staff.staff.last_name)
        elif gasto.autorizar:
            autorizado_por =str(gasto.superintendente.staff.staff.first_name) + ' ' + str(gasto.superintendente.staff.staff.last_name)
            status = "Autorizado | Falta una autorización"
        elif gasto.autorizar == False:
            status = "Cancelado"
            if gasto.superintendente:
                autorizado_por = str(gasto.superintendente.staff.staff.first_name) + ' ' + str(gasto.superintendente.staff.staff.last_name)
            else:
                autorizado_por = "No identificado"
        else:
            autorizado_por = "Faltan autorizaciones"
            status = "Faltan autorizaciones"

        proyectos = set()
        subproyectos = set()
        comentarios = set()
        articulos_gasto = Articulo_Gasto.objects.filter(gasto=gasto)
        for articulo in articulos_gasto:
            if articulo.proyecto:
                proyectos.add(str(articulo.proyecto.nombre))
            if articulo.subproyecto:
                subproyectos.add(str(articulo.subproyecto.nombre))
            if articulo.comentario:
                comentarios.add(str(articulo.comentario))

        proyectos_str = ', '.join(proyectos)
        subproyectos_str = ', '.join(subproyectos)
        comentarios_str = ', '.join(comentarios)

        row = [
            gasto.folio,
            autorizado_at_2_naive,
            gasto.distrito.nombre if gasto.distrito else '',
            proyectos_str,
            subproyectos_str,
            comentarios_str,
            gasto.staff.staff.staff.first_name + ' ' + gasto.staff.staff.staff.last_name,
            gasto.colaborador.staff.staff.first_name + ' '  + gasto.colaborador.staff.staff.last_name if gasto.colaborador else '',
            gasto.get_total_solicitud,
            created_at_naive,
            status,
            autorizado_por,
            tiene_facturas,
            pagada,
            #f'=IF(I{row_num}="",G{row_num},I{row_num}*G{row_num})',  # Calcula total en pesos usando la fórmula de Excel
            #created_at_naive,
        ]

    
        for col_num in range(len(row)):
            (ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style
            if col_num ==1 or col_num == 9:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style
            if col_num == 8:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = money_style
       
    
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)

def convert_excel_matriz_gastos(articulos_comprados):
    #print('si entra a la función')
    # Crea un objeto BytesIO para guardar el archivo Excel
    output = BytesIO()

    # Crea un libro de trabajo y añade una hoja
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet("Producto_pendientes")

     
    date_format = workbook.add_format({'num_format': 'dd/mm/yyyy'})
    # Define los estilos
    head_style = workbook.add_format({'bold': True, 'font_color': 'FFFFFF', 'bg_color': '333366', 'font_name': 'Arial', 'font_size': 11})
    body_style = workbook.add_format({'font_name': 'Calibri', 'font_size': 10})
    money_style = workbook.add_format({'num_format': '$ #,##0.00', 'font_name': 'Calibri', 'font_size': 10})
    date_style = workbook.add_format({'num_format': 'dd/mm/yyyy', 'font_name': 'Calibri', 'font_size': 10})
    percent_style = workbook.add_format({'num_format': '0.00%', 'font_name': 'Calibri', 'font_size': 10})
    messages_style = workbook.add_format({'font_name':'Arial Narrow', 'font_size':11})

    columns = ['Compra', 'Requisición','Solicitud','Sector', 'Codigo', 'Producto', 'Cantidad Pendiente', 'Unidad','Proveedor',
               'Usuario Solicitante']

    columna_max = len(columns)+2

    worksheet.write(0, columna_max - 1, 'Reporte Creado Automáticamente por SAVIA Vordcab. UH', messages_style)
    worksheet.write(1, columna_max - 1, 'Software desarrollado por Grupo Vordcab S.A. de C.V.', messages_style)
    worksheet.set_column(columna_max - 1, columna_max, 30)  # Ajusta el ancho de las columnas nuevas
    
   
    for i, column in enumerate(columns):
        worksheet.write(0, i, column, head_style)
        worksheet.set_column(i, i, 15)  # Ajusta el ancho de las columnas

    worksheet.set_column('L:L', 12,  money_style)
    worksheet.set_column('M:M', 12, money_style) 
    # Asumiendo que ya tienes tus datos de compras
    row_num = 0
    for articulo in articulos_comprados:
        row_num += 1
        # Aquí asumimos que ya hiciste el procesamiento necesario de cada compra
        pagos = Pago.objects.filter(oc=articulo.oc)
        
        #tipo_de_cambio_promedio_pagos = pagos.aggregate(Avg('tipo_de_cambio'))['tipo_de_cambio__avg']

        # Usar el tipo de cambio de los pagos, si existe. De lo contrario, usar el tipo de cambio de la compra
        if articulo.oc.req.orden.sector:
            sector = f"{articulo.oc.req.orden.sector.nombre}"
        else:
            sector = ' '
        #tipo = tipo_de_cambio_promedio_pagos or compra_list.tipo_de_cambio
        #tipo_de_cambio = '' if tipo == 0 else tipo
        #created_at = compra_list.created_at.replace(tzinfo=None)
        #approved_at = compra_list.req.approved_at

        row = [
            articulo.oc.folio,
            articulo.oc.req.folio,
            articulo.oc.req.orden.folio,
            sector,
            articulo.producto.producto.articulos.producto.producto.codigo,
            articulo.producto.producto.articulos.producto.producto.nombre,
            articulo.cantidad_pendiente if articulo.cantidad_pendiente != None else articulo.cantidad,
            articulo.producto.producto.articulos.producto.producto.unidad.nombre,
            articulo.oc.proveedor.nombre.razon_social,
            f"{articulo.oc.req.orden.staff.staff.staff.first_name} {articulo.oc.req.orden.staff.staff.staff.last_name}",
        ]
        
        for col_num, cell_value in enumerate(row):
        # Define el formato por defecto
            cell_format = body_style

            # Aplica el formato de fecha para las columnas con fechas
            if col_num in [7, 8]:  # Asume que estas son tus columnas de fechas
                cell_format = date_style
        
            # Aplica el formato de dinero para las columnas con valores monetarios
            elif col_num in [11, 12]:  # Asume que estas son tus columnas de dinero
                cell_format = money_style

            # Finalmente, escribe la celda con el valor y el formato correspondiente
            worksheet.write(row_num, col_num, cell_value, cell_format)

      
        #worksheet.write_formula(row_num, 19, f'=IF(ISBLANK(R{row_num+1}), L{row_num+1}, L{row_num+1}*R{row_num+1})', money_style)
    
   
    workbook.close()

    # Construye la respuesta
    output.seek(0)

    response = HttpResponse(
        output.read(), 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    response['Content-Disposition'] = f'attachment; filename=Producto_pendientes_entrada_{dt.date.today()}.xlsx'
      # Establecer una cookie para indicar que la descarga ha iniciado
    response.set_cookie('descarga_iniciada', 'true', max_age=20)  # La cookie expira en 20 segundos
    output.close()
    return response

def entradas_por_gasto(request):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    entradas = Conceptos_Entradas.objects.filter(entrada__completo=True, concepto_material__producto__servicio=False, entrada__gasto__gasto__distrito = usuario.distritos).order_by('-id')
    myfilter = Conceptos_EntradasFilter(request.GET, queryset=entradas)
    entradas = myfilter.qs
   
    entradas_data = list(entradas.values())

    #Set up pagination
    p = Paginator(entradas, 50)
    page = request.GET.get('page')
    entradas_list = p.get_page(page)

    #if request.method == "POST" and 'btnExcel' in request.POST:
        #print(entradas)
    #    return convert_entradas_to_xls2(entradas)

    context = {
        'entradas_list':entradas_list,
        'entradas':entradas,
        'myfilter':myfilter,
        }
    #task_id_entradas =   request.session.get('task_id_entradas')

    #if request.method == "POST" and 'btnExcel' in request.POST:
        #if not task_id_entradas:
            #task =  convert_entradas_to_xls_task.delay(entradas_data)
            #task_id = task.id
            #request.session['task_id_entradas'] = task_id
            #context['task_id_entradas'] = task_id 

    return render(request,'gasto/reporte_entradas_gasto.html', context)

def crear_pdf_cfdi_buffer(factura):
    #factura = Factura.objects.get(id=pk)
    data = factura.emisor
    # Verificar y asignar un valor predeterminado para impuestos si es None
    if data['impuestos'] is None:
        data['impuestos'] = 0.0

    # Verificar y asignar un valor predeterminado para total si es None
    if data['total'] is None:
        data['total'] = 0.0

    # Verificar y asignar un valor predeterminado para subtotal si es None
    if data['subtotal'] is None:
        data['subtotal'] = 0.0
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    if not data:
        return HttpResponse("Error al parsear el archivo XML", status=400)

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    
    # Generar código QR
    qr_data = f"https://verificacfdi.facturaelectronica.sat.gob.mx/default.aspx?id={data['uuid']}&re={data['rfc_emisor']}&rr={data['rfc_receptor']}&tt={data['total']}&fe={data['sello_cfd'][-8:]}"
    qr_img = qrcode.make(qr_data)
    
    with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as temp_file:
        qr_img.save(temp_file)
        temp_file.seek(0)
        qr_x = 500
        qr_y = height - 700
        qr_size = 2.75 * cm
        c.drawImage(temp_file.name, qr_x, qr_y, qr_size, qr_size)

    # Título
    c.setFillColor(prussian_blue)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(30, height - 40, "FACTURA GENERADA POR SAVIA 2.0")

    # Datos del Emisor
    c.setFillColor(black)
    c.setFont("Helvetica-Bold", 12)
    c.drawString(30, height - 80, "Datos del Emisor:")
    
    c.setFont("Helvetica", 8)
    alineado_x = 30
    alineado_y = height - 100
    alineado_y2 = alineado_y
    line_height = 12

    c.drawString(alineado_x, alineado_y, f"RFC: {data['rfc_emisor']}")
    alineado_y -= line_height
    c.drawString(alineado_x, alineado_y, f"Nombre: {data['nombre_emisor']}")
    alineado_y -= line_height
    c.drawString(alineado_x, alineado_y, f"Régimen Fiscal: {data['regimen_fiscal_emisor']}")
    alineado_y -= line_height
    c.drawString(alineado_x, alineado_y, f"Lugar de Expedición: {data['lugar_expedicion']}")
    alineado_y -= line_height
    c.drawString(alineado_x, alineado_y, f"Fecha y hora de expedición: {data['fecha']}")
    alineado_y -= line_height
    c.drawString(alineado_x, alineado_y, f"Moneda: {data['moneda']}")
    alineado_y -= line_height
    c.drawString(alineado_x, alineado_y, f"Forma de Pago: {data['forma_pago']}")

    # Datos del Receptor
    alineado_y -= 2 * line_height
    c.setFont("Helvetica-Bold", 12)
    c.drawString(alineado_x + 350, height - 80, "Datos del Receptor:")
    
    c.setFont("Helvetica", 8)
    alineado_y -= line_height
    c.drawString(alineado_x + 350, alineado_y2, f"RFC: {data['rfc_receptor']}")
    alineado_y2 -= line_height
    c.drawString(alineado_x + 350, alineado_y2, f"Nombre: {data['nombre_receptor']}")
    alineado_y2 -= line_height
    c.drawString(alineado_x + 350, alineado_y2, f"Régimen Fiscal: {data['regimen_fiscal_receptor']}")
    alineado_y2 -= line_height
    c.drawString(alineado_x + 350, alineado_y2, f"Régimen Fiscal: {data['codigo_postal']}")
    alineado_y2 -= line_height
    c.drawString(alineado_x + 350, alineado_y2, f"Uso del CFDI: {data['uso_cfdi']}")

    # Conceptos (Tabla)
    alineado_y -= line_height
    # Configuración del estilo para los párrafos
    styles = getSampleStyleSheet()
    styleN = styles['Normal']
    styleN.wordWrap = 'CJK'  # Ajusta automáticamente el texto
    # Crear un estilo personalizado
    custom_style = ParagraphStyle(
        'CustomStyle',
        parent=styleN,
        fontSize=6,  # Ajusta el tamaño del texto aquí
        leading=7,   # Ajusta el interlineado aquí si es necesario
    )

    # Preparamos los datos de la tabla
    table_data = [["CANT", "CLAVE", "CONCEPTO", "U DE M", "P.U.", "IMPORTE", "IMPUESTO", "TIPO TASA"]]
    for item in data['resultados']:
        descripcion = item['descripcion']
        cantidad = float(item['cantidad'])
        unidad = item['unidad']
        valor_unitario = float(item['precio'])
        importe = float(item['importe'])
        # Verificar y convertir solo si el valor no es 'N/A'
         # Inicializar las variables impuesto y tasa
        impuesto = item['impuesto']
        tasa = item['tasa_cuota']
       # Convertir impuesto a float, o 0.0 si es 'N/A' o None
        try:
            impuesto = float(impuesto)
        except (TypeError, ValueError):
            impuesto = 0.0

        # Convertir tasa a float, o 0.0 si es 'N/A' o None
        try:
            tasa = float(tasa)
        except (TypeError, ValueError):
            tasa = 0.0

        clave = item.get('clave') or item.get('clave_prod_serv', 'SIN_CLAVE')
         # Crear un párrafo para la descripción
        descripcion_paragraph = Paragraph(descripcion, custom_style)
        unidad_paragraph = Paragraph(unidad, custom_style)
        table_data.append([
            f"{cantidad:.2f}",
            clave,
            descripcion_paragraph,
            unidad_paragraph,
            f"{valor_unitario:,.2f}",
            f"{importe:,.2f}",
            f"{impuesto:,.2f}",
            f"{tasa:.2f}",
        ])

    # Crear la tabla
    table = Table(table_data, colWidths=[1.0 * cm, 1.5 * cm, 8.5 * cm, 1.5 * cm, 2 * cm, 2 * cm, 1.5 * cm, 1.5 * cm, 1.5 * cm])
    table.setStyle(TableStyle([
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 0), (-1, 0), prussian_blue),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
    ]))

    # Guardar la tabla en el PDF
    table.wrapOn(c, width, height)
    table.drawOn(c, alineado_x, alineado_y - len(table_data) * line_height)

    # Ajustar el alineado_y para seguir escribiendo debajo de la tabla
    alineado_y -= len(table_data) * line_height + 2 * line_height

    # Totales
    c.setFont("Helvetica-Bold", 12)
   
    c.setFont("Helvetica", 10)
    alineado_y -= line_height

     # Importe con letra
    alineado_y -= 2 * line_height
    c.drawString(alineado_x, alineado_y, "Importe con Letra:")
    total_letras = num2words(float(data['total']), lang='es', to='currency', currency='MXN')
    c.drawString(alineado_x, alineado_y - 10, total_letras)
    #c.drawRightString(alineado_x, alineado_y , f"{data['importe_con_letra']}")
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.setFillColor(prussian_blue)
    c.rect(alineado_x + 390 ,alineado_y - 50,110,62, fill=True, stroke=False) #Barra azul superior | Subtotal
    c.setFillColor(white)
    c.drawRightString(alineado_x + 500, alineado_y , f"Subtotal:")
    c.setFillColor(black)
    c.drawRightString(alineado_x + 555, alineado_y, f"{float(data['subtotal']):,.2f}")
    alineado_y -= line_height
    c.setFillColor(white)
    c.drawRightString(alineado_x + 500, alineado_y, f"Impuestos trasladados:")
    c.setFillColor(black)
    c.drawRightString(alineado_x + 555, alineado_y, f"{float(data['impuestos']):,.2f}")
    alineado_y -= line_height
    if data['iva_retenido'] > 0:
        c.setFillColor(white)
        c.drawRightString(alineado_x + 500, alineado_y, f"Impuestos retenidos:")
        c.setFillColor(black)
        c.drawRightString(alineado_x + 555, alineado_y, f"{float(data['iva_retenido']):,.2f}")
        alineado_y -= line_height
    if data['isr_retenido'] > 0:
        c.setFillColor(white)
        c.drawRightString(alineado_x + 500, alineado_y, f"ISR:")
        c.setFillColor(black)
        c.drawRightString(alineado_x + 555, alineado_y, f"{float(data['isr_retenido']):,.2f}")
        alineado_y -= line_height
    c.setFillColor(white)
    c.drawRightString(alineado_x + 500, alineado_y, f"Total:")
    c.setFillColor(black)
    c.drawRightString(alineado_x + 555, alineado_y, f"{float(data['total']):,.2f}")
    # Otros detalles
    

    otros_detalles = [
        ["Folio Fiscal", "Fecha y Hora de Certificación", "No. Certificado Digital", "Método de Pago"],
        [data['uuid'], data['fecha_timbrado'], data['no_certificado'], data['metodo_pago']]
    ]
    detalles_table = Table(otros_detalles, colWidths=[5 * cm, 5 * cm, 4.5 * cm, 4.5 * cm])
    detalles_table.setStyle(TableStyle([
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, 0), (-1, 0), prussian_blue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
    ]))

    # Guardar la tabla de detalles en el PDF
    detalles_table.wrapOn(c, width, height)
    detalles_table.drawOn(c, alineado_x, 180)
    alineado_y -= 4 * line_height
     # Utilizar Paragraph para las líneas largas
    styles = getSampleStyleSheet()
    styleN = styles["BodyText"]
    styleN.fontSize = 6
    c.setFont("Helvetica", 6)
    c.line(30,177,580,177)
    c.drawString(alineado_x, 170, f"ESTE DOCUMENTO ES UNA REPRESENTACIÓN IMPRESA DE UN CFDI v4.0")
    
    # Reducir el ancho de los párrafos
    reduced_width = width * 0.7  # Ajusta este valor según sea necesario

    sello_cfd_paragraph = Paragraph(f"Sello Digital del CFDI: {data['sello_cfd']}", styleN)
    sello_cfd_paragraph.wrapOn(c,  reduced_width, line_height * 4)
    sello_cfd_paragraph.drawOn(c, alineado_x, 130)
    alineado_y -= line_height * 5
    
    sello_sat_paragraph = Paragraph(f"Sello del SAT: {data['sello_sat']}", styleN)
    sello_sat_paragraph.wrapOn(c,  reduced_width, line_height * 4)
    sello_sat_paragraph.drawOn(c, alineado_x, 90)
    alineado_y -= line_height * 3
    c.drawString(alineado_x, 40, f"No. serie CSD SAT {data['no_certificadoSAT']}")

    sello_cfd_paragraph = Paragraph(f"Cadena Original del complemento de certificación digital del SAT: {data['cadena_original']}", styleN)
    sello_cfd_paragraph.wrapOn(c,  reduced_width, line_height * 4)
    sello_cfd_paragraph.drawOn(c, alineado_x, 50)
    alineado_y -= line_height * 5
    
   

    c.showPage()
    c.save()

    buffer.seek(0)

    return buffer

def descargar_vale_rosa_pdf(request, vale_id):
    buffer = generar_pdf_vale_rosa(vale_id)
    return FileResponse(buffer, as_attachment=True, filename=f"vale_rosa_{vale_id}.pdf")
    
def generar_cfdi_gasto(request, pk):
    factura = Factura.objects.get(id=pk)
    buffer = crear_pdf_cfdi_buffer(factura)
    # Crear la respuesta HTTP con el PDF
    folio_fiscal = factura.emisor.get('uuid', f'factura_{factura.id}')
    return HttpResponse(buffer, content_type='application/pdf', headers={
        'Content-Disposition': f'attachment; filename="{folio_fiscal}.pdf"'
    })

def crear_pdf_cfdi_gasto(factura):
    buffer = crear_pdf_cfdi_buffer(factura)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_file:
        tmp_file.write(buffer.read())
        return tmp_file.name

def generar_pdf_vale_rosa(vale_id):
    vale = ValeRosa.objects.select_related('gasto', 'creado_por').get(id=vale_id)
    
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Establecer fondo rosa
    c.setFillColorRGB(1, 0.9, 0.9)
    c.rect(0, 0, width, height, stroke=0, fill=1)

    # Colores y fuentes
    c.setFont("Helvetica-Bold", 14)
    c.setFillColor(red)
    c.drawString(40, height - 50, "Vale Provisional de Caja")

    c.setFont("Helvetica", 10)
    c.setFillColor(black)
    # Recuadro superior derecho para importe en número
    c.rect(450, height - 60, 90, 20)  # (x, y, width, height)
    # Escribir el monto en número dentro del recuadro
    c.setFont("Helvetica-Bold", 10)
    c.drawRightString(500, height - 55, f"${vale.monto:.2f}")
    # Rectángulo de importe en letra
    c.rect(40, height - 80, 500, 20)
    c.drawString(45, height - 75, "IMPORTE (EN LETRA):")

    monto_entero = int(vale.monto)
    centavos = int(round((Decimal(vale.monto) - monto_entero) * 100))
    monto_letras = f"{num2words(monto_entero, lang='es').upper()} PESOS {centavos:02d}/100 M.N."
    c.drawString(200, height - 75, monto_letras)


       # Recuadro de concepto
    c.setFont("Helvetica-Bold", 11)
    c.drawString(45, height - 110, "CONCEPTO")

    c.setLineWidth(1)
    c.rect(40, height - 270, 520, 150)

    #c.setFont("Helvetica", 10)
    #text_object = c.beginText(45, height - 150)
    #text_object.setLeading(14)
    
    #for line in vale.motivo.splitlines():
    #    text_object.textLine(line)
    #c.drawText(text_object)
    styles = getSampleStyleSheet()
    styleN = styles['Normal']
    styleN.fontName = 'Helvetica'
    styleN.fontSize = 10
    styleN.leading = 14  # Espaciado entre líneas

    texto = (vale.motivo or "").replace('\n', '<br/>')  # Reemplaza saltos de línea por etiquetas HTML
    parrafo = Paragraph(texto, styleN)
    pad = 6
    frame = Frame( 40 + pad, height - 270 + pad, 520 - 2 * pad, 150 - 2 * pad, showBoundary=0)
    frame.addFromList([parrafo], c)

    # Sección inferior (fecha, autorizado por, recibido por)
    c.rect(40, 80, 520, 60)

    # Divisiones internas
    c.line(210, 80, 210, 140)
    c.line(370, 80, 370, 140)

    c.setFont("Helvetica", 10)
    c.drawString(60, 130, "FECHA")
    c.drawString(230, 130, "AUTORIZADO POR")
    c.drawString(410, 130, "RECIBIDO POR")
   
    fecha_vale = vale.creado_en.strftime('%d/%m/%Y')
    if vale.esta_aprobado:
        if vale.aprobado_por:
            autorizado = str(vale.aprobado_por.staff.staff.first_name+' '+ vale.aprobado_por.staff.staff.last_name)
        else:
            autorizado = "AUTORIZADO"
    else:
        autorizado = "PENDIENTE"
    #autorizado = str(vale.aprobado_por.staff.staff.first_name+' '+ vale.aprobado_por.staff.staff.last_name) if vale.aprobado_por else "PENDIENTE"
    if vale.gasto:
        if vale.gasto.colaborador:
            recibido = str(vale.gasto.colaborador.staff.staff.first_name + ' ' + vale.gasto.colaborador.staff.staff.last_name)
        else:
            recibido = str(vale.gasto.staff.staff.staff.first_name + ' ' + vale.gasto.staff.staff.staff.last_name)
    elif vale.viatico:
        if vale.viatico.colaborador:
            recibido = str(vale.viatico.colaborador.staff.staff.first_name + ' ' + vale.viatico.colaborador.staff.staff.last_name)
        else:
            recibido = str(vale.viatico.staff.staff.staff.first_name + ' ' + vale.viatico.staff.staff.staff.last_name)
    else:
        recibido = "No asignado"


    c.setFont("Helvetica-Bold", 10)
    c.drawString(60, 110, fecha_vale)
    c.drawString(230, 110, autorizado)
    c.drawString(410, 110, recibido)

    c.setFont("Helvetica", 7)
    c.drawString(50, 50, "Generado automáticamente por SAVIA")

    c.showPage()
    c.save()
    buffer.seek(0)

    return buffer  # Puedes devolverlo como HttpResponse en una vista

@perfil_seleccionado_required
def matriz_gasto_rh(request):
    pk_usuario = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_usuario)
    solicitudes = Solicitud_Gasto.objects.filter(
        tipo__familia = 'rh_nomina',
        dispersion = False,
    ).order_by('-id').distinct()
    myfilter=Solicitud_Gasto_Filter(request.GET, queryset=solicitudes)
       #Set up pagination
    p = Paginator(solicitudes, 10)
    page = request.GET.get('page')
    ordenes_list = p.get_page(page)

    #Set up pagination
    p = Paginator(solicitudes, 10)
    page = request.GET.get('page')
    ordenes_list = p.get_page(page)

    context={
        'ordenes_list': ordenes_list,
        'myfilter': myfilter,
        
    }

    return render(request, 'gasto/matriz_gasto_rh.html', context)

@perfil_seleccionado_required
def editar_gasto_rh(request, pk):
    articulos_gasto = Articulo_Gasto.objects.all()
    #distritos = Distrito.objects.none()
    
    conceptos = Product.objects.all()
    pk_usuario = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_usuario)
    gasto = Solicitud_Gasto.objects.get(id=pk)
    proyectos = Proyecto.objects.filter(~Q(status_de_entrega__status = "INACTIVO"),activo=True, distrito = gasto.distrito) 
    productos = articulos_gasto.filter(gasto=gasto, completo = True) 
    
    error_messages = {}
    archivos_nomina = ArchivoSoporte.objects.filter(solicitud=gasto)

    total_bbva = archivos_nomina.filter(tipo__nombre='BBVA').first()
    total_ob = archivos_nomina.filter(tipo__nombre='OB').first()
    total_pensiones = archivos_nomina.filter(tipo__nombre='PENSIONES').first()
    # ⚠️ Al final, obtener todos los archivos RH cargados de este gasto
    archivos_rh = ArchivoSoporte.objects.filter(solicitud=gasto).exclude(tipo__nombre__in=['BANORTE', 'OB', 'BBVA', 'PENSIONES'])
    
    proyectos_para_select2 = [
        {
            'id': item.id, 
            'text': str(item.nombre)
        } for item in proyectos
    ]

    articulos_gasto_select = conceptos.filter(gasto = True, baja_item = False)
   

    productos_para_select2 = [
        {
            'id': item.id,
            'text': str(item.nombre),
            'iva': str(item.iva)
        } for item in articulos_gasto_select
    ]

    form_product = Articulo_GastoForm()

    if request.method =='POST': 
        
        if "btn_agregar" in request.POST:
            gasto.dispersion = True
            gasto.save()
               
            messages.success(request, f'La solicitud {gasto.folio} ha sido editada')
            return redirect('matriz-gasto-rh')
        else:
            messages.error(request, f'La solicitud {gasto.folio} no ha sido guardada correctamente, verifica los campos')
    
        if "btn_producto" in request.POST:
            articulo, created = articulos_gasto.get_or_create(completo = False, staff=usuario)
            form_product = Articulo_GastoForm(request.POST, instance=articulo)
            if form_product.is_valid():
                articulo = form_product.save(commit=False)
                articulo.gasto = gasto
                articulo.completo = True
                articulo.save()
                messages.success(request, 'Haz agregado un artículo correctamente')
                return redirect('editar-gasto-rh', pk = gasto.id)
            else:
                for field, errors in form_product.errors.items():
                    error_messages[f'Líneas de Gasto| Campo:{field}'] = errors.as_text()
 
    total_nomina = archivos_nomina.aggregate(suma=Sum('total'))['suma'] or 0.0
    context={
        'proyectos_para_select2': proyectos_para_select2,
        'productos_para_select2': productos_para_select2,
        'gasto': gasto,
        'productos': productos,
        'form_product': form_product,
        'archivo_bbva': total_bbva,
        'archivo_ob': total_ob,
        'archivo_pensiones': total_pensiones,
        'archivos_rh': archivos_rh,
        'total_nomina': total_nomina,

    }

    return render(request, 'gasto/editar_gasto_rh.html', context)

@perfil_seleccionado_required
def delete_gasto_rh(request, pk):
    articulo = Articulo_Gasto.objects.get(id=pk)
    gasto = Solicitud_Gasto.objects.get(id=articulo.gasto.id)
    messages.success(request,f'El articulo {articulo.producto} ha sido eliminado exitosamente')
    articulo.delete()

    return redirect('editar-gasto-rh', pk = gasto.id)