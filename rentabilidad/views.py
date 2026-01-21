from django.shortcuts import render, redirect
from django.db import transaction
from django.db.models import Sum, Q
from django.db.models.functions import TruncMonth
from django.conf import settings
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import HttpResponse
from calendar import month_name,  monthrange
from collections import defaultdict

from .models import Costos, Solicitud_Costos, Tipo_Costo, Solicitud_Ingresos, Ingresos, Depreciaciones, Concepto
from user.models import Profile, Distrito
from solicitudes.models import Contrato
from compras.models import Moneda
from user.decorators import perfil_seleccionado_required
from .forms import Costo_Form, Solicitud_Costo_Form, Solicitud_Ingreso_Form, Ingreso_Form, Depreciacion_Form, Solicitud_Costo_Indirecto_Form, Solicitud_Costo_Indirecto_Central_Form, UploadExcelForm
from .filters import Costos_Form
from datetime import date, datetime, timedelta
from dateutil.relativedelta import relativedelta
import re
import unicodedata


from decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl.styles import NamedStyle, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl import load_workbook


# Create your views here.
@perfil_seleccionado_required
def costos(request):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    costos = Solicitud_Costos.objects.filter(complete = True)
    tipos = Tipo_Costo.objects.all()
    distritos = Distrito.objects.exclude(id__in = [7,8,16]).exclude(status=False)
    myfilter= Costos_Form(request.GET, queryset=costos)

    #Set up pagination
    p = Paginator(costos, 20)
    page = request.GET.get('page')
    costos_list = p.get_page(page)

    context = {
        'costos':costos,
        'myfilter': myfilter,
        'costos_list': costos_list,
        'tipos': tipos,
        'distritos':distritos,
         }

    return render(request,'rentabilidad/costos.html', context)

@perfil_seleccionado_required
def conceptos_costos(request, pk):
    costo = Solicitud_Costos.objects.get(id=pk)
    conceptos = Costos.objects.filter(solicitud = costo)


    context = {
        'costo':costo,
        'conceptos':conceptos,
    }

    return render(request,'rentabilidad/conceptos_costos.html',context)

@perfil_seleccionado_required
def add_costo(request, tipo):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    distritos = Distrito.objects.exclude(id__in = [7,8,16]).exclude(status=False) #7 MATRIZ ALTERNATIVO, 8 ALTAMIRA ALTERNATIVO,16 BRASIL
    solicitud, created =  Solicitud_Costos.objects.get_or_create(created_by=usuario, complete = False)
    costos = Costos.objects.filter(solicitud = solicitud)
    if tipo == "directo":
        form = Solicitud_Costo_Form()
        tipos = Tipo_Costo.objects.filter(id__in = [2])
        form.fields['tipo'].queryset = tipos
        form.fields['distrito'].queryset = distritos
    elif tipo == "indirecto":
        form = Solicitud_Costo_Indirecto_Form()
        tipos = Tipo_Costo.objects.filter(id__in = [3,4])
        form.fields['tipo'].queryset = tipos
        form.fields['distrito'].queryset = distritos
    elif tipo == "central":
        form = Solicitud_Costo_Indirecto_Central_Form()
        tipos = Tipo_Costo.objects.filter(id__in = [1])
        print('siii')
    costo_form = Costo_Form()

    #El problema reciente es que aparecen 5 tipos de costos, pero solo tengo definidos:
    #  1 central
    #  2 directo
    #  3 operativo < indirecto
    #  4 administrativo < indirecto


    if request.method =='POST':
        if "btn_agregar" in request.POST:
            if tipo == "directo":
                form = Solicitud_Costo_Form(request.POST, instance = solicitud)
            elif tipo == "indirecto":
                form = Solicitud_Costo_Indirecto_Form(request.POST, instance = solicitud)
            elif tipo == "central":
                form = Solicitud_Costo_Indirecto_Central_Form(request.POST, instance = solicitud)
            print('estou aqui')
            if form.is_valid():
                
                solicitud = form.save(commit=False)
                solicitud.created_at = date.today()
                if tipo == "central":
                    distrito = Distrito.objects.get(nombre = "MATRIZ")
                    solicitud.distrito = distrito
                    solicitud.tipo = Tipo_Costo.objects.get(nombre = "Indirecto Central")
                solicitud.complete = True
                solicitud.save()
                messages.success(request,'Has agregado correctamente la Solicitud')
                return redirect('rentabilidad-costos')  
            else:
                print('Nao é valido')
                print(form.errors)  
        if "btn_costo" in request.POST:
            costo, created = Costos.objects.get_or_create(complete = False, solicitud = solicitud)
            form = Costo_Form(request.POST, instance = costo)
            if form.is_valid():
                costo = form.save(commit=False)
                costo.complete = True
                costo.save()
                messages.success(request,'Has agregado correctamente un costo')
                return redirect('add-costo', tipo=tipo)
                
 

    context = {
        'tipo': tipo,
        'form': form,
        'costo_form': costo_form,
        'costos':costos,
        'solicitud':solicitud,
        }

    return render(request,'rentabilidad/add_costo.html',context)

def _to_decimal(value):
    """
    Acepta 1234.56, "1,234.56", 1234, etc.
    """
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    s = str(value).strip().replace(",", "")
    if s == "":
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None
    

def limpiar_espacios(s):
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip())

def normalizar_clave(valor):
    if valor is None:
        return ""
    s = re.sub(r"\s+", " ", str(valor).strip())
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s.upper()

@perfil_seleccionado_required
def carga_costos_excel(request):
    pk_perfil = request.session.get("selected_profile_id")
    usuario = Profile.objects.get(id=pk_perfil)

    solicitud, _ = Solicitud_Costos.objects.get_or_create(created_by=usuario, complete=False)

    if request.method == "POST":
        form = UploadExcelForm(request.POST, request.FILES)
        if not form.is_valid():
            return render(request, "rentabilidad/carga_costos_excel.html", {"form": form})

        f = form.cleaned_data["file"]

        try:
            wb = load_workbook(filename=f, data_only=True)
            ws = wb.active
        except Exception as e:
            messages.error(request, f"No se pudo leer el Excel: {e}")
            return redirect(request.path)

        COL_CONCEPTO = 2  # B
        COL_MONTO = 5     # E
        MIN_ROW = 2       # ajusta si tu reporte tiene encabezados

        filas = []
        errores = []

        # 1) Leer filas y validar monto básico
        for row_idx, row in enumerate(ws.iter_rows(min_row=MIN_ROW, values_only=True), start=MIN_ROW):
            concepto_raw = row[COL_CONCEPTO - 1] if len(row) >= COL_CONCEPTO else None
            monto_raw = row[COL_MONTO - 1] if len(row) >= COL_MONTO else None

            concepto_txt = (str(concepto_raw).strip() if concepto_raw is not None else "")
            monto = _to_decimal(monto_raw)
            concepto_txt = normalizar_clave(concepto_txt)
            # saltar filas vacías en B y E
            if concepto_txt == "" and (monto is None or monto == 0):
                continue

            # opcional: ignorar encabezados de reporte
            upper = concepto_txt.upper()
            if upper.startswith("CONTRATO") or upper in {"CUENTA", "NOMBRE"}:
                continue

            if concepto_txt == "":
                errores.append(f"Fila {row_idx}: columna B vacía (concepto).")
                continue

            if monto is None:
                errores.append(f"Fila {row_idx}: columna E inválida ({monto_raw}).")
                continue

            filas.append((limpiar_espacios(concepto_txt), monto))

        if errores:
            return render(request, "rentabilidad/carga_costos_excel.html", {"form": form, "errores": errores})

        if not filas:
            messages.warning(request, "No se encontraron filas válidas para cargar.")
            return redirect("add-costo", tipo="directo")

        # 2) Resolver conceptos existentes (sin crear nuevos)
        nombres = sorted({c for c, _ in filas})
        
        existentes = Concepto.objects.filter(nombre__in=nombres)  # <-- ajusta 'nombre' si tu campo se llama distinto

        mapa = {normalizar_clave(c.nombre): c for c in existentes}

        faltantes = [n for n in nombres if n not in mapa]

        # 3) Crear solo los Costos que sí tengan Concepto
        costos_a_crear = []
        for concepto_txt, monto in filas:
            concepto_obj = mapa.get(concepto_txt)
            if not concepto_obj:
                continue

            costos_a_crear.append(Costos(
                solicitud=solicitud,
                concepto=concepto_obj,  # FK
                monto=monto,
                complete=True
            ))

        with transaction.atomic():
            Costos.objects.bulk_create(costos_a_crear, batch_size=500)

        # 4) Mensajes finales
        if costos_a_crear:
            messages.success(request, f"Se crearon {len(costos_a_crear)} costos correctamente.")
        else:
            messages.warning(request, "No se creó ningún costo porque ninguno de los conceptos existe en el catálogo.")

        if faltantes:
            # evita mensajes gigantes (opcional)
            max_show = 25
            lista = ", ".join(faltantes[:max_show])
            extra = "" if len(faltantes) <= max_show else f" ... (+{len(faltantes)-max_show} más)"
            messages.warning(request, f"No se encontraron en el catálogo: {lista}{extra}")

        return redirect("add-costo", tipo="directo")

    else:
        form = UploadExcelForm()

    return render(request, "rentabilidad/carga_costos_excel.html", {"form": form, "solicitud": solicitud})


@perfil_seleccionado_required
def delete_costo(request, tipo, pk):
    costo = Costos.objects.get(id=pk)
    messages.success(request,f'El costo {costo.concepto} ha sido eliminado exitosamente')
    costo.delete()

    return redirect('add-costo', tipo=tipo)


def reporte_costos(request):
    tipo_id = request.GET.get("tipo_id")  # <- capturamos el valor del select
    distrito_id = request.GET.get("distrito_id", 1)  # puedes hacerlo dinámico también
    print(tipo_id)
    if int(tipo_id) == 1:
        distrito_id = 6
        
    costos = Costos.objects.filter(solicitud__distrito_id=distrito_id, solicitud__tipo_id=tipo_id)
    fecha_inicio = request.GET.get("fecha_inicio")  # viene como YYYY-MM
    fecha_fin = request.GET.get("fecha_fin")        # viene como YYYY-MM

    tabla, meses = get_tabla_costos(tipo_id, distrito_id, fecha_inicio, fecha_fin)
    distrito_nombre = Distrito.objects.get(id=distrito_id).nombre
    tipo_nombre = Tipo_Costo.objects.get(id=tipo_id).nombre
    print(tabla)

    # 🚩 Si el usuario presiona el botón de Excel
    if request.method == "POST" and "btnReporte" in request.POST:
        return generar_costos_excel(tabla, meses, distrito_id, tipo_id, fecha_inicio, fecha_fin)

    context = { 
        "tabla": tabla,
        "meses": meses,
        "tipo_id": tipo_id,
        "distrito_id": distrito_id,
        "distrito_nombre": distrito_nombre,
        "tipo_nombre": tipo_nombre,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
    }

    return render(request, "rentabilidad/reportes_costos.html", context)



def get_tabla_costos(tipo_id=None, distrito_id=None, fecha_inicio=None, fecha_fin=None):
    costos = Costos.objects.all()
    print('tipo_id',tipo_id)
    print('distrito_id',distrito_id)
    print('fecha_inicio',fecha_inicio)
    print('fecha_fin',fecha_fin)

    if distrito_id:
        costos = costos.filter(solicitud__distrito_id=distrito_id)
    #print(costos)
    tipo_nombre = None
    if tipo_id:
        costos = costos.filter(solicitud__tipo_id=tipo_id)
        tipo_nombre = Tipo_Costo.objects.get(id=tipo_id).nombre
    
    if fecha_inicio and fecha_fin:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio, "%Y-%m").date()
            y, m = [int(x) for x in fecha_fin.split("-")]
            last_day = monthrange(y, m)[1]
            fecha_fin = date(y, m, last_day)
            costos = costos.filter(solicitud__fecha__range=[fecha_inicio, fecha_fin])
            #print(costos)
        except ValueError:
            pass

    # 🚩 Agrupamos por concepto y mes
    data = (
        costos.annotate(mes=TruncMonth("solicitud__fecha"))
        .values("mes")
        .annotate(total=Sum("monto"))
        .order_by("concepto__nombre", "mes")
    )

    tabla = {}
    meses = set()

    for row in data:
        mes_value = row["mes"]
        mes = mes_value.strftime("%B %Y")
        meses.add(mes)
        total_mes = row["total"]

    # 🚩 Caso especial: Indirecto Central
    if tipo_nombre == "Indirecto Central":
        # 🚩 Caso especial: Indirecto Central
        tabla_ingresos, meses_ingresos = get_tabla_ingresos_contrato(
            fecha_inicio.strftime("%Y-%m"),
            fecha_fin.strftime("%Y-%m")
        )

        tabla_distribuida = {}
        for mes in meses:
            total_indirecto_mes = sum(
                row["total"] for row in data if row["mes"].strftime("%B %Y") == mes
            )

            for contrato, meses_dict in tabla_ingresos.items():
                if mes in meses_dict:
                    prorrateo = meses_dict[mes]["prorrateo"] / Decimal("100.00")
                    costo_asignado = total_indirecto_mes * prorrateo

                    if contrato not in tabla_distribuida:
                        tabla_distribuida[contrato] = {}
                    tabla_distribuida[contrato][mes] = {
                        "monto": costo_asignado,
                        "porcentaje": meses_dict[mes]["prorrateo"]  # en %
                    }

        tabla = tabla_distribuida
     # 🚩 Caso: indirectos operativos o administrativos → prorrateo por distrito
    elif tipo_nombre in ["Indirectos Operativos", "Indirectos Administrativos"]:
        print('aqui indirectos')
        tabla_ingresos, _ = get_tabla_ingresos_contrato(
            fecha_inicio.strftime("%Y-%m"), fecha_fin.strftime("%Y-%m"), distrito_id
        )

        costos_mes = (
            costos.annotate(mes=TruncMonth("solicitud__fecha"))
            .values("mes")
            .annotate(total=Sum("monto"))
        )
        print('costos_mes',costos_mes)
        for row in costos_mes:
            mes = row["mes"].strftime("%B %Y")
            meses.add(mes)
            total_costos_mes = row["total"]

            for contrato, meses_dict in tabla_ingresos.items():
                print('meses_dict',meses_dict)
                if mes in meses_dict:
                    
                    prorrateo = meses_dict[mes]["prorrateo"]

                    if contrato not in tabla:
                        tabla[contrato] = {}
                    if mes not in tabla[contrato]:
                        tabla[contrato][mes] = {"monto": 0, "porcentaje": 0}

                    tabla[contrato][mes]["monto"] += total_costos_mes * (prorrateo / 100)
                    tabla[contrato][mes]["porcentaje"] = prorrateo

    # 🚩 Caso: cualquier otro tipo → costos normales
    else:
        data = (
            costos.annotate(mes=TruncMonth("solicitud__fecha"))
            .values("concepto__nombre", "mes")
            .annotate(total=Sum("monto"))
            .order_by("concepto__nombre", "mes")
        )

        for row in data:
            concepto = row["concepto__nombre"]
            mes = row["mes"].strftime("%B %Y")
            meses.add(mes)

            if concepto not in tabla:
                tabla[concepto] = {}
            tabla[concepto][mes] = {"monto": row["total"], "porcentaje": 0}
    meses = sorted(meses, key=lambda m: (m.split()[1], list(month_name).index(m.split()[0])))

    return tabla, meses

def generar_costos_excel(tabla, meses, distrito_id=None, tipo_id=None, fecha_inicio=None, fecha_fin=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Costos"

     # Logo en A1
    img_path = "static/images/logo_vordcab.jpg"  # cambia según tu ruta real
    try:
        img = Image(img_path)
        img.width = img.width * 1
        img.height = img.height * 1
        ws.add_image(img, "A1")
        ws.row_dimensions[1].height = 70  # ajustar alto de fila 1
    except FileNotFoundError:
        pass  # si no hay logo, no truena

    # Encabezado de filtros
    fila = 2
    
    distrito_nombre = Distrito.objects.get(id=distrito_id).nombre
    tipo_nombre = Tipo_Costo.objects.get(id=tipo_id).nombre
    ws.cell(row=fila, column=1, value=f"Distrito: {distrito_nombre}")
    fila += 1
    ws.cell(row=fila,column=1, value=f"Tipo de Costo: {tipo_nombre}")
    fila += 1
    if fecha_inicio and fecha_fin:
        ws.cell(row=fila, column=1, value=f"Rango de meses: {fecha_inicio} → {fecha_fin}")
        fila += 2  # dejar espacio antes de la tabla

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)

     # Estilo moneda
    money_style = NamedStyle(name="money_style")
    money_style.number_format = u'"$"#,##0.00'
    money_style.font = Font(name="Calibri", size=10)

    # Encabezados
    cell = ws.cell(row=fila, column=1, value="Concepto")
    cell.style = head_style

    col_offset = 2   # 👈 inicializar antes del loop
    for mes in meses:
        ws.cell(row=fila, column=col_offset, value=f"{mes} (Monto)").style = head_style
        ws.cell(row=fila, column=col_offset+1, value=f"{mes} (%)").style = head_style
        col_offset += 2

    # Filas
    for row_idx, (concepto, valores) in enumerate(tabla.items(), start=fila+1):
        ws.cell(row=row_idx, column=1, value=concepto)

        col_offset = 2
        for mes in meses:
            val = valores.get(mes, {"monto": 0, "porcentaje": 0})
            monto = float(val.get("monto", 0))
            porcentaje = float(val.get("porcentaje", 0))

            # monto
            cell = ws.cell(row=row_idx, column=col_offset, value=monto)
            cell.style = money_style

            # porcentaje
            ws.cell(row=row_idx, column=col_offset+1, value=porcentaje)

            col_offset += 2

    # Ajustar ancho columnas
    for i, col in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = max(
            (len(str(cell.value)) for cell in col if cell.value), default=10
        ) + 2

    # Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_costos.xlsx"'
    wb.save(response)
    return response

@perfil_seleccionado_required
def ingresos(request):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    ingresos = Solicitud_Ingresos.objects.filter(complete = True)
    distritos = Distrito.objects.exclude(id__in = [7,8,16]).exclude(status=False)
    #myfilter= ContratoFilter(request.GET, queryset=contratos)

    #Set up pagination
    #p = Paginator(contratos, 10)
    #page = request.GET.get('page')
    #contratos_list = p.get_page(page)

    context = {
        'ingresos':ingresos,
        #'myfilter': myfilter,
        'distritos':distritos,
         }

    return render(request,'rentabilidad/ingresos.html', context)

@perfil_seleccionado_required
def add_ingresos(request):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    distritos = Distrito.objects.exclude(id__in = [7,8,16]).exclude(status=False) #7 MATRIZ ALTERNATIVO, 8 ALTAMIRA ALTERNATIVO,16 BRASIL
    monedas = Moneda.objects.exclude(id__in = [3] )
    solicitud, created =  Solicitud_Ingresos.objects.get_or_create(created_by=usuario, complete = False)
    ingresos = Ingresos.objects.filter(solicitud = solicitud)
    form = Solicitud_Ingreso_Form()
    form.fields['distrito'].queryset = distritos
    ingreso_form = Ingreso_Form()
    ingreso_form.fields['moneda'].queryset = monedas

    if request.method =='POST':
        
       
        if "btn_agregar" in request.POST:
            form = Solicitud_Ingreso_Form(request.POST, instance = solicitud)
            print('estou aqui')
            if form.is_valid():
                solicitud = form.save(commit=False)
                solicitud.created_at = date.today()
                solicitud.complete = True
                solicitud.save()
                messages.success(request,'Has agregado correctamente la Solicitud')
                return redirect('rentabilidad-ingresos')  
            else:
                print('Nao é valido')
                print(form.errors)  
        if "btn_ingreso" in request.POST:
            ingreso, created = Ingresos.objects.get_or_create(complete = False, solicitud = solicitud)
            form = Ingreso_Form(request.POST, instance = ingreso)
            if form.is_valid():
                ingreso = form.save(commit=False)
                ingreso.complete = True
                ingreso.save()
                messages.success(request,'Has agregado correctamente un ingreso')
                return redirect('add-ingreso')
                
 

    context = {
        'form': form,
        'ingreso_form': ingreso_form,
        'ingresos':ingresos,
        'solicitud':solicitud,
        }

    return render(request,'rentabilidad/add_ingreso.html',context)

@perfil_seleccionado_required
def delete_ingreso(request, pk):
    ingreso = Ingresos.objects.get(id=pk)
    messages.success(request,f'El ingreso {ingreso.concepto} ha sido eliminado exitosamente')
    ingreso.delete()

    return redirect('add-ingreso')

def get_tabla_ingresos_distrito(distrito_id, fecha_inicio=None, fecha_fin=None):
    ingresos = Ingresos.objects.filter(solicitud__complete=True, solicitud__distrito_id=distrito_id)

    if fecha_inicio and fecha_fin:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio, "%Y-%m").date()
            y, m = [int(x) for x in fecha_fin.split("-")]
            last_day = monthrange(y, m)[1]
            fecha_fin = date(y, m, last_day)
            ingresos = ingresos.filter(solicitud__fecha__range=[fecha_inicio, fecha_fin])
        except ValueError:
            pass

    tabla = {}
    meses = set()
    total_distrito = Decimal("0.00")

    # Recorremos ingresos con conversión de moneda
    for ingreso in ingresos:
        monto = ingreso.monto
        if ingreso.moneda and ingreso.moneda.nombre.upper() == "DOLARES":
            if ingreso.tipo_cambio:
                monto = monto * ingreso.tipo_cambio
            else:
                monto = Decimal("0.00")

        mes_value = ingreso.solicitud.fecha.replace(day=1)
        mes = mes_value.strftime("%B %Y")
        meses.add(mes)

        contrato_nombre = ingreso.contrato.nombre

        if contrato_nombre not in tabla:
            tabla[contrato_nombre] = {}

        if mes not in tabla[contrato_nombre]:
            tabla[contrato_nombre][mes] = {"monto": Decimal("0.00")}

        tabla[contrato_nombre][mes]["monto"] += monto
        total_distrito += monto

    # 👉 calcular % participación por contrato
    participacion = {}
    for contrato, meses_data in tabla.items():
        total_contrato = sum([mes_data["monto"] for mes_data in meses_data.values()])
        if total_distrito > 0:
            participacion[contrato] = round((total_contrato / total_distrito) * 100, 2)
        else:
            participacion[contrato] = 0

    meses = sorted(meses, key=lambda m: datetime.strptime(m, "%B %Y"))

    return tabla, meses, participacion

from decimal import Decimal

def get_tabla_ingresos_contrato(fecha_inicio=None, fecha_fin=None, distrito_id=None):
    ingresos = Ingresos.objects.filter(solicitud__complete=True)

    if distrito_id:
        ingresos = ingresos.filter(solicitud__distrito_id=distrito_id)

    if fecha_inicio and fecha_fin:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio, "%Y-%m").date()
            y, m = [int(x) for x in fecha_fin.split("-")]
            last_day = monthrange(y, m)[1]
            fecha_fin = date(y, m, last_day)
            ingresos = ingresos.filter(solicitud__fecha__range=[fecha_inicio, fecha_fin])
        except ValueError:
            pass

    # Agrupamos totales por contrato y mes
    data = (
        ingresos.annotate(mes=TruncMonth("solicitud__fecha"))
        .values("contrato_id", "contrato__nombre", "mes")
        .annotate(total=Sum("monto"))
        .order_by("contrato__nombre", "mes")
    )

    tabla = {}
    meses = set()

    # 🔑 necesitamos acumular totales por mes para calcular prorrateos
    totales_por_mes = {}

    for row in data:
        contrato_id = row["contrato_id"]
        contrato_nombre = row["contrato__nombre"]
        mes_value = row["mes"]

        if not mes_value:
            continue

        mes = mes_value.strftime("%B %Y")
        meses.add(mes)

        # Normalizamos montos considerando tipo de cambio
        ingresos_mes = Ingresos.objects.filter(
            contrato_id=contrato_id,
            solicitud__complete=True,
            solicitud__fecha__year=mes_value.year,
            solicitud__fecha__month=mes_value.month,
        )
        if distrito_id:
            ingresos_mes = ingresos_mes.filter(solicitud__distrito_id=distrito_id)

        monto_total = Decimal("0.00")
        for ingreso in ingresos_mes:
            monto = ingreso.monto
            if ingreso.moneda and ingreso.moneda.nombre.upper() == "DOLARES":
                monto = monto * ingreso.tipo_cambio if ingreso.tipo_cambio else Decimal("0.00")
            monto_total += monto

        # Guardamos en tabla
        if contrato_nombre not in tabla:
            tabla[contrato_nombre] = {}
        tabla[contrato_nombre][mes] = {"monto": monto_total}  # prorrateo lo calculamos después

        # Acumulamos totales por mes (para prorrateo)
        totales_por_mes[mes] = totales_por_mes.get(mes, Decimal("0.00")) + monto_total

    # 🔑 Segundo paso: calcular prorrateo (%)
    for contrato, meses_dict in tabla.items():
        for mes, valores in meses_dict.items():
            total_mes = totales_por_mes.get(mes, Decimal("0.00"))
            if total_mes > 0:
                valores["prorrateo"] = round(valores["monto"] / total_mes * 100, 2)
            else:
                valores["prorrateo"] = 0

    meses = sorted(meses, key=lambda m: datetime.strptime(m, "%B %Y"))
    return tabla, meses

    

def calcular_prorrateo_contrato(contrato_id, year, month):
    total_mes = Decimal("0.00")
    total_general = Decimal("0.00")

    # Todos los ingresos de ese contrato
    ingresos = Ingresos.objects.filter(contrato_id=contrato_id, solicitud__complete = True)

    for ingreso in ingresos:
        monto = ingreso.monto
        if ingreso.moneda and ingreso.moneda.nombre.upper() == "DOLARES":
            if ingreso.tipo_cambio:  # evitar None
                monto = monto * ingreso.tipo_cambio
            else:
                monto = Decimal("0.00")  # o puedes decidir que cuente como monto directo

        total_general += monto

        if ingreso.solicitud.fecha.year == year and ingreso.solicitud.fecha.month == month:
            total_mes += monto

    if total_general == 0:
        return Decimal("0.00")

    return (total_mes / total_general).quantize(Decimal("0.0001"))


def reporte_ingresos(request):
    distrito_id = request.GET.get("distrito_id")
    #print(distrito_id)
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")
    if distrito_id:
        distrito_nombre = Distrito.objects.get(id = distrito_id).nombre
    else:
        distrito_nombre = ""
    print('distrito',distrito_nombre)
    tabla, meses = get_tabla_ingresos_contrato(fecha_inicio, fecha_fin, distrito_id)
    
    if request.method == "POST" and "btnReporte" in request.POST:
        return generar_ingresos_excel(tabla, meses, distrito_id, fecha_inicio, fecha_fin)

    context = {
        "tabla": tabla,
        "meses": meses,
        "distrito_id": distrito_id,
        "distrito_nombre": distrito_nombre,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
    }
    return render(request, "rentabilidad/reporte_ingresos.html", context)

def reporte_ingresos_contrato(request):
    distrito_id = request.GET.get("distrito_id")
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")

    tabla, meses = get_tabla_ingresos_contrato(fecha_inicio, fecha_fin, distrito_id)

    if request.method == "POST" and "btnReporte" in request.POST:
        return generar_ingresos_excel(tabla, meses, None, fecha_inicio, fecha_fin)

    context = {
        "tabla": tabla,
        "meses": meses,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
    }
    return render(request, "rentabilidad/reporte_ingresos_contrato.html", context)



def generar_ingresos_excel(tabla, meses, distrito_id=None, fecha_inicio=None, fecha_fin=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ingresos"
    img_path = "static/images/logo_vordcab.jpg"
    img = Image(img_path)
    img.width = img.width * 1
    img.height = img.height * 1
    ws.add_image(img, "A1")
    ws.row_dimensions[1].height = 70 
    # Encabezado de filtros
    fila = 2
    if distrito_id:
        try:
            distrito_nombre = Distrito.objects.get(id=distrito_id).nombre
        except Distrito.DoesNotExist:
            distrito_nombre = distrito_id
        ws.cell(row=fila, column=1, value=f"Distrito: {distrito_nombre}")
        fila += 1
    if fecha_inicio and fecha_fin:
        ws.cell(row=fila, column=1, value=f"Rango de meses: {fecha_inicio} → {fecha_fin}")
        fila += 2  # dejar un espacio antes de la tabla

    # Crear estilo del encabezado
    head_style = NamedStyle(name="head_style")
    head_style.font = Font(name='Arial', color='00FFFFFF', bold=True, size=11)
    head_style.fill = PatternFill("solid", fgColor='00003366')
    if "head_style" not in wb.named_styles:
        wb.add_named_style(head_style)

    # Encabezados de tabla
    cell = ws.cell(row=fila, column=1, value="Contrato")
    cell.style = head_style
    for col_idx, mes in enumerate(meses, start=2):
        cell = ws.cell(row=fila, column=col_idx, value=mes)
        cell.style = head_style

    # Filas de datos
    for row_idx, (contrato, valores) in enumerate(tabla.items(), start=fila+1):
        ws.cell(row=row_idx, column=1, value=contrato)
        for col_idx, mes in enumerate(meses, start=2):
            ws.cell(row=row_idx, column=col_idx, value=valores.get(mes, 0))

    # Ajustar ancho columnas
    for i, col in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = max(
            (len(str(cell.value)) for cell in col if cell.value), default=10
        ) + 2

    # Respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_ingresos.xlsx"'
    wb.save(response)
    return response

# Create your views here.
@perfil_seleccionado_required
def depreciaciones(request):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    depreciaciones = Depreciaciones.objects.filter(complete = True)
    #tipos = Tipo_Costo.objects.all()
    distritos = Distrito.objects.exclude(id__in = [7,8,16]).exclude(status=False)
    #myfilter= ContratoFilter(request.GET, queryset=contratos)

    #Set up pagination
    #p = Paginator(contratos, 10)
    #page = request.GET.get('page')
    #contratos_list = p.get_page(page)

    context = {
        'depreciaciones': depreciaciones,
        #'myfilter': myfilter,
        #'tipos': tipos,
        'distritos':distritos,
         }

    return render(request,'rentabilidad/depreciaciones.html', context)


def reporte_depreciaciones(request):
    distrito_id = request.GET.get("distrito_id")
    fecha_inicio = request.GET.get("fecha_inicio")   # "YYYY-MM"
    fecha_fin = request.GET.get("fecha_fin")         # "YYYY-MM"
    # fecha de "hoy" (puede ser corte contable o date.today())
    hoy = date.today()
    mes_corte = date(hoy.year, hoy.month, 1)

    # 1) Parseo de fechas como en tus otros reportes
    if fecha_inicio and fecha_fin:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio, "%Y-%m").date()
            y, m = [int(x) for x in fecha_fin.split("-")]
            last_day = monthrange(y, m)[1]
            fecha_fin = date(y, m, last_day)
        except ValueError:
            fecha_inicio = fecha_fin = None

    # 2) Query de depreciaciones del distrito
    depreciaciones = Depreciaciones.objects.filter(
        distrito__id=distrito_id, complete=True
    )

    # 3) Meses (claves internas y etiquetas visibles)
    meses = []
    meses_formateados = {}
    if fecha_inicio and fecha_fin:
        actual = fecha_inicio
        while actual <= fecha_fin:
            key = actual.strftime("%Y-%m")       # clave interna (estable)
            label = actual.strftime("%b %Y")     # etiqueta visible (cabecera)
            meses.append(key)
            meses_formateados[key] = label
            # avanzar un mes completo
            days_in_month = monthrange(actual.year, actual.month)[1]
            actual += timedelta(days=days_in_month)


    # 4) Tabla: { "Contrato" : { "Concepto" : { "YYYY-MM": "$1,234.00" } } }
    tabla = defaultdict(lambda: defaultdict(dict))
    remanentes = defaultdict(dict)

    for dep in depreciaciones:
        meses_dep = dep.meses_a_depreciar or 1
        monto_mensual = dep.monto / meses_dep
        monto_total = dep.monto


        inicio = date(dep.mes_inicial.year, dep.mes_inicial.month, 1)
        # ✅ meses transcurridos INCLUYENDO el mes actual
        meses_trans = _months_inclusive(inicio, mes_corte)
        # no puede exceder los meses a depreciar
        meses_trans = min(meses_trans, meses_dep)
        remanente = monto_total - (meses_trans * monto_mensual)
        contrato_nombre = getattr(dep.contrato, "nombre", str(dep.contrato))
        concepto = dep.concepto or "Sin concepto"
        remanentes[contrato_nombre][concepto] = f"${remanente:,.2f}"

        for i in range(meses_dep):
            year = inicio.year + (inicio.month - 1 + i) // 12
            month = (inicio.month - 1 + i) % 12 + 1
            fecha_mes = date(year, month, 1)
            key = fecha_mes.strftime("%Y-%m")

            if fecha_inicio and fecha_fin and fecha_inicio <= fecha_mes <= fecha_fin:
                contrato_nombre = getattr(dep.contrato, "nombre", str(dep.contrato))
                concepto = dep.concepto or "Sin concepto"

                tabla[contrato_nombre][concepto][key] = f"${monto_mensual:,.2f}"

    # 5) Contexto
    distrito_nombre = ""
    first_dep = depreciaciones.first()
    if first_dep and first_dep.distrito:
        distrito_nombre = first_dep.distrito.nombre

    tabla_dict = {
    str(contrato): {
        str(concepto): dict(valores)         # dict() para el tercer nivel
        for concepto, valores in sorted(conceptos.items(), key=lambda x: x[0])
    }
    for contrato, conceptos in sorted(tabla.items(), key=lambda x: x[0])
    }

    totales = {mes: 0 for mes in meses}

    for contrato, conceptos in tabla_dict.items():
        for concepto, valores in conceptos.items():
            for mes in meses:
                if mes in valores:
                    # quitar $ y comas, convertir a float
                    monto = float(valores[mes].replace("$", "").replace(",", ""))
                    totales[mes] += monto

    # Convertir de nuevo a string formateado
    totales_fmt = {mes: f"${totales[mes]:,.2f}" for mes in meses}

      # 🔑 calcular remanente
    #diferencia = relativedelta(mes_actual, inicio)
    #meses_transcurridos = diferencia.years * 12 + diferencia.months
    #if meses_transcurridos > meses_dep:
    #    meses_transcurridos = meses_dep

    #remanente = monto_total - (meses_transcurridos * monto_mensual)
    #contrato_nombre = getattr(dep.contrato, "nombre", str(dep.contrato))
    #concepto = dep.concepto or "Sin concepto"
    

    if request.method == "POST" and "btnReporte" in request.POST:
        return generar_depreciaciones_excel(tabla, meses, remanentes, distrito_id, fecha_inicio, fecha_fin)
    
    context = {
        "distrito_nombre": distrito_nombre,
        "fecha_inicio": fecha_inicio.strftime("%b %Y") if fecha_inicio else "",
        "fecha_fin": fecha_fin.strftime("%b %Y") if fecha_fin else "",
        "meses": meses,
        "meses_formateados": meses_formateados,
        "tabla": tabla_dict,   # 👈 ya no uses dict(sorted(...)) sobre el defaultdict original
        "totales": totales_fmt,   # 👈 aquí van los totales
        "remanentes": dict(remanentes),
    }
    return render(request, "rentabilidad/reporte_depreciaciones.html", context)


def _months_inclusive(start_month: date, current_month: date) -> int:
    """Meses entre start y current INCLUYENDO el mes actual.
       Si current < start, regresa 0."""
    if current_month < start_month:
        return 0
    return (current_month.year - start_month.year) * 12 + (current_month.month - start_month.month) + 1

@perfil_seleccionado_required
def add_depreciacion(request):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    distritos = Distrito.objects.exclude(id__in = [7,8,16]).exclude(status=False) #7 MATRIZ ALTERNATIVO, 8 ALTAMIRA ALTERNATIVO,16 BRASIL
    #monedas = Moneda.objects.exclude(id__in = [3] )
    depreciacion, created =  Depreciaciones.objects.get_or_create(created_by=usuario, complete = False)
    #depreciaciones = Ingresos.objects.filter(solicitud = solicitud)
    form = Depreciacion_Form()
    form.fields['distrito'].queryset = distritos
    #ingreso_form = Ingreso_Form()
    #ingreso_form.fields['moneda'].queryset = monedas

    if request.method =='POST':
        if "btn_agregar" in request.POST:
            form = Depreciacion_Form(request.POST, instance = depreciacion)
            print('estou aqui')
            if form.is_valid():
                depreciacion = form.save(commit=False)
                depreciacion.created_at = date.today()
                depreciacion.complete = True
                depreciacion.save()
                messages.success(request,'Has agregado correctamente la Solicitud')
                return redirect('rentabilidad-depreciaciones')  
            else:
                print('Nao é valido')
                print(form.errors)  
        
                
 

    context = {
        'form': form,
        }

    return render(request,'rentabilidad/add_depreciaciones.html',context)


def generar_depreciaciones_excel(tabla, meses, remanentes, distrito_id=None, fecha_inicio=None, fecha_fin=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Depreciaciones"

    money_style = NamedStyle(name="money_style")
    money_style.number_format = u'"$"#,##0.00'   # 👈 formato moneda
    money_style.font = Font(name="Calibri", size=10)
    wb.add_named_style(money_style)

    # Logo
    img_path = "static/images/logo_vordcab.jpg"
    img = Image(img_path)
    img.width = img.width * 1
    img.height = img.height * 1
    ws.add_image(img, "A1")
    ws.row_dimensions[1].height = 70 

    # Encabezado de filtros
    fila = 2
    if distrito_id:
        try:
            distrito_nombre = Distrito.objects.get(id=distrito_id).nombre
        except Distrito.DoesNotExist:
            distrito_nombre = distrito_id
        ws.cell(row=fila, column=1, value=f"Distrito: {distrito_nombre}")
        fila += 1
    if fecha_inicio and fecha_fin:
        ws.cell(row=fila, column=1, value=f"Rango de meses: {fecha_inicio} → {fecha_fin}")
        fila += 2  # dejar un espacio antes de la tabla

    # Estilo encabezado
    head_style = NamedStyle(name="head_style")
    head_style.font = Font(name='Arial', color='00FFFFFF', bold=True, size=11)
    head_style.fill = PatternFill("solid", fgColor='00003366')
    if "head_style" not in wb.named_styles:
        wb.add_named_style(head_style)

    # Encabezados de tabla
    ws.cell(row=fila, column=1, value="Contrato / Concepto").style = head_style
    for col_idx, mes in enumerate(meses, start=2):
        ws.cell(row=fila, column=col_idx, value=mes).style = head_style
    ws.cell(row=fila, column=len(meses)+2, value="Remanente").style = head_style
    # Filas de datos
    row_idx = fila + 1
    for contrato, conceptos in tabla.items():
        # Fila contrato
        ws.cell(row=row_idx, column=1, value=contrato).font = Font(bold=True)
        row_idx += 1

        for concepto, valores in conceptos.items():
            ws.cell(row=row_idx, column=1, value=concepto)
            for col_idx, mes in enumerate(meses, start=2):
                monto_str = valores.get(mes, 0)
                try:
                    # quitar $ y comas si es string
                    monto = float(str(monto_str).replace("$", "").replace(",", ""))
                except:
                    monto = 0
                cell = ws.cell(row=row_idx, column=col_idx, value=monto)
                cell.style = money_style   # 👈 formato moneda

            # 👉 Columna de remanente
            rem_str = remanentes.get(contrato, {}).get(concepto, 0)
            try:
                rem = float(str(rem_str).replace("$", "").replace(",", ""))
            except:
                rem = 0
            rem_cell = ws.cell(row=row_idx, column=len(meses)+2, value=rem)
            rem_cell.style = money_style
            row_idx += 1

    ws.cell(row=row_idx, column=1, value="TOTAL").font = Font(bold=True)

    for col_idx, mes in enumerate(meses, start=2):
        # Calculamos el rango desde la primera fila de datos hasta la fila anterior
        col_letter = get_column_letter(col_idx)
        start_row = fila + 1   # justo después de los encabezados
        end_row = row_idx - 1  # última fila antes del total
        formula = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"

        cell = ws.cell(row=row_idx, column=col_idx, value=formula)
        cell.style = money_style
        cell.font = Font(bold=True)

    # 👉 Columna de remanente total (puede ser suma también)
    col_letter = get_column_letter(len(meses)+2)
    start_row = fila + 1
    end_row = row_idx - 1
    formula = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
    cell = ws.cell(row=row_idx, column=len(meses)+2, value=formula)
    cell.style = money_style
    cell.font = Font(bold=True)

     # Ajustar anchos
    for i, col in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = max(
            (len(str(cell.value)) for cell in col if cell.value), default=10
        ) + 2

    # Respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_depreciaciones.xlsx"'
    wb.save(response)
    return response



def reporte_rentabilidad_mensual(request):
    distrito_id = request.GET.get("distrito_id")
    mes_anio = request.GET.get("mes_anio")  # formato YYYY-MM

    contratos_data = []
    totales = {
        "ingresos": 0,
        "depreciaciones": 0,
        "rentabilidad": 0,
        "directos": 0,
        "ind_adm": 0,
        "ind_oper": 0,
        "ind_central": 0,
    }

    tipos_costos_totales = {}  # acumulados por tipo de costo

    distrito_nombre = ""
    fecha_label = ""

    if distrito_id and mes_anio:
        try:
            y, m = [int(x) for x in mes_anio.split("-")]
            fecha_mes = date(y, m, 1)
            fecha_label = fecha_mes.strftime("%B %Y")
        except ValueError:
            fecha_mes = None

        if fecha_mes:
            try:
                distrito_nombre = Distrito.objects.get(id=distrito_id).nombre
            except Distrito.DoesNotExist:
                distrito_nombre = "?"
            
            
            # 👉 Traemos los ingresos prorrateados (por contrato y mes)
            tabla_ingresos, _ = get_tabla_ingresos_contrato(
                fecha_inicio=mes_anio, fecha_fin=mes_anio
            )
            # total nacional de ingresos (para Indirecto Central)
            ingreso_total_nacional = sum(
                v[fecha_label]["monto"]
                for v in tabla_ingresos.values()
                if fecha_label in v
            )

            print('ingreso total nacional', ingreso_total_nacional)
            # Obtener todos los contratos que tienen algo en ese mes

            # total de ingresos del distrito (ya en pesos)
            solicitudes_ingresos_distrito = Solicitud_Ingresos.objects.filter(
                distrito_id = distrito_id,
                fecha__year=y,
                fecha__month=m,
            )
            ingresos_distrito = sum(s.get_total for s in solicitudes_ingresos_distrito)
            print('ingresos distrito', ingresos_distrito)

            contratos = (
                Contrato.objects.filter(
                    Q(sc_contratos__distrito__id=int(distrito_id),
                    sc_contratos__fecha__year=y,
                    sc_contratos__fecha__month=m)
                    |
                    Q(i_contratos__solicitud__distrito__id=int(distrito_id),
                    i_contratos__solicitud__fecha__year=y,
                    i_contratos__solicitud__fecha__month=m)
                )
                .distinct()
            )
            print('contratos', contratos)

            for contrato in contratos:
                row = {
                    "contrato": contrato.nombre or str(contrato),
                    "ingresos": 0,
                    "depreciaciones": 0,
                    "directos": 0,  # cada tipo de costo → monto
                    "ind_oper": 0,
                    "ind_adm": 0,
                    "ind_central": 0,
                    "rentabilidad": 0,
                }

                # ------------------------
                # Ingresos desde tabla_ingresos
                # ------------------------
                # usar str() porque en tabla_ingresos la llave es texto
                contrato_key = str(contrato)

                ingreso_contrato = (
                    tabla_ingresos.get(contrato_key, {})
                    .get(fecha_label, {})
                    .get("monto", 0)
                )
                #print('ingreso contrato', ingreso_contrato)

                #prorrateo_distrito = (
                #    tabla_ingresos.get(contrato_key, {})
                #    .get(fecha_label, {})
                #    .get("prorrateo", 0)
                #)
                prorrateo_distrito = ingreso_contrato / ingresos_distrito * 100 if ingresos_distrito else 0
                print('prorrateo distrito', prorrateo_distrito)

                row["ingresos"] = ingreso_contrato

                # ------------------------
                # Directos (suma directa)
                # ------------------------
                row["directos"] = (
                    Costos.objects.filter(
                        solicitud__distrito__id= distrito_id,
                        solicitud__contrato = contrato,
                        #solicitud__tipo__nombre="Directo",
                        solicitud__fecha__year=y,
                        solicitud__fecha__month=m,
                    ).aggregate(total=Sum("monto"))["total"]
                    or  Decimal("0.00") 
                )

                #print('directos', row["directos"])
                       # ------------------------
                # Indirectos Operativos / Administrativos (prorrateados en el distrito)
                # ------------------------
                total_ind_op = (
                    Costos.objects.filter(
                        solicitud__distrito__id=distrito_id,
                        solicitud__tipo__nombre="Indirectos Operativos",
                        solicitud__fecha__year=y,
                        solicitud__fecha__month=m,
                    ).aggregate(total=Sum("monto"))["total"]
                    or 0
                )

                total_ind_adm = (
                    Costos.objects.filter(
                        solicitud__distrito__id=distrito_id,
                        solicitud__tipo__nombre="Indirectos Administrativos",
                        solicitud__fecha__year=y,
                        solicitud__fecha__month=m,
                    ).aggregate(total=Sum("monto"))["total"]
                    or 0
                )

                #print('ind adm',total_ind_adm)

                row["ind_oper"] = total_ind_op * (prorrateo_distrito / 100)
                row["ind_adm"] = total_ind_adm * (prorrateo_distrito / 100)

                # ------------------------
                # Indirecto Central (prorrateado a nivel nacional)
                # ------------------------
                total_ind_central = (
                    Costos.objects.filter(
                        solicitud__tipo__nombre="Indirecto Central",
                        solicitud__fecha__year=y,
                        solicitud__fecha__month=m,
                    ).aggregate(total=Sum("monto"))["total"]
                    or 0
                )


                prorrateo_nacional = (
                    (ingreso_contrato / ingreso_total_nacional * 100)
                    if ingreso_total_nacional
                    else 0
                )
                #print('ingreso contrato', ingreso_contrato)
                #print('ingreso total nacional', ingreso_total_nacional)
                #print('prorrateo nacional', prorrateo_nacional) >>> verificado, parece estar correcto
                row["ind_central"] = total_ind_central * (prorrateo_nacional / 100)

               
                

                 # ------------------------
                # Depreciaciones (ya lo tenías)
                # ------------------------
                depreciaciones_total = 0
                depreciaciones = Depreciaciones.objects.filter(
                    contrato=contrato, distrito__id=distrito_id, complete=True
                )
                for dep in depreciaciones:
                    meses_dep = dep.meses_a_depreciar or 1
                    monto_mensual = dep.monto / meses_dep
                    inicio = date(dep.mes_inicial.year, dep.mes_inicial.month, 1)
                    for i in range(meses_dep):
                        year = inicio.year + (inicio.month - 1 + i) // 12
                        month = (inicio.month - 1 + i) % 12 + 1
                        fecha_iter = date(year, month, 1)
                        if fecha_iter == fecha_mes:
                            depreciaciones_total += monto_mensual
                            break
                row["depreciaciones"] = depreciaciones_total

                # ------------------------
                # Rentabilidad
                # ------------------------
                row["rentabilidad"] = (
                    row["ingresos"]
                    - row["directos"]
                    - row["ind_oper"]
                    - row["ind_adm"]
                    - row["ind_central"]
                    - row["depreciaciones"]
                )

                contratos_data.append(row)

                     # ------------------------
                # Totales
                # ------------------------
                totales["ingresos"] += row["ingresos"]
                totales["directos"] += row["directos"]
                totales["ind_oper"] += row["ind_oper"]
                totales["ind_adm"] += row["ind_adm"]
                totales["ind_central"] += row["ind_central"]
                totales["depreciaciones"] += row["depreciaciones"]
                totales["rentabilidad"] += row["rentabilidad"]


    
        # 👉 Si el usuario pidió Excel
    if request.method == "POST" and "btnReporte" in request.POST:
        return generar_rentabilidad_excel(
            contratos_data=contratos_data, 
            tipos_costos_totales=tipos_costos_totales,
            distrito_id=distrito_id, 
            mes_anio=mes_anio, 
            fecha_label=fecha_label
            )


    context = {
        "distritos": Distrito.objects.exclude(id__in=[7, 8, 16]).exclude(status=False),
        "distrito_id": distrito_id,
        "distrito_nombre": distrito_nombre,
        "mes_anio": mes_anio,
        "fecha_label": fecha_label,
        "contratos_data": contratos_data,
        "totales": totales,
        "tipos_costos_totales": tipos_costos_totales,  # para la fila final
    }
    return render(request, "rentabilidad/rentabilidad.html", context)

def generar_rentabilidad_excel(contratos_data, tipos_costos_totales, distrito_id=None, mes_anio=None, fecha_label=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rentabilidad"

    # Estilo moneda
    money_style = NamedStyle(name="money_style")
    money_style.number_format = u'"$"#,##0.00'
    money_style.font = Font(name="Calibri", size=10)
    if "money_style" not in wb.named_styles:
        wb.add_named_style(money_style)

    # Logo
    img_path = "static/images/logo_vordcab.jpg"
    img = Image(img_path)
    img.width = img.width * 1
    img.height = img.height * 1
    ws.add_image(img, "A1")
    ws.row_dimensions[1].height = 70

    # Encabezado de filtros
    fila = 2
    if distrito_id:
        try:
            distrito_nombre = Distrito.objects.get(id=distrito_id).nombre
        except Distrito.DoesNotExist:
            distrito_nombre = distrito_id
        ws.cell(row=fila, column=1, value=f"Distrito: {distrito_nombre}")
        fila += 1
    if fecha_label:
        ws.cell(row=fila, column=1, value=f"Mes: {fecha_label}")
        fila += 2

    # Estilo encabezado
    head_style = NamedStyle(name="head_style")
    head_style.font = Font(name='Arial', color='00FFFFFF', bold=True, size=11)
    head_style.fill = PatternFill("solid", fgColor='00003366')
    if "head_style" not in wb.named_styles:
        wb.add_named_style(head_style)

    # 👉 Encabezados dinámicos
    col_idx = 1
    ws.cell(row=fila, column=col_idx, value="Contrato").style = head_style
    col_idx += 1
    ws.cell(row=fila, column=col_idx, value="Ingresos").style = head_style
    col_idx += 1

    # Tipos de costos en columnas
    tipos_order = list(tipos_costos_totales.keys())
    for tipo in tipos_order:
        ws.cell(row=fila, column=col_idx, value=tipo).style = head_style
        col_idx += 1

    ws.cell(row=fila, column=col_idx, value="Amortización").style = head_style
    col_idx += 1
    ws.cell(row=fila, column=col_idx, value="Rentabilidad").style = head_style

    # 👉 Filas de datos
    row_idx = fila + 1
    for row in contratos_data:
        col = 1
        ws.cell(row=row_idx, column=col, value=row["contrato"])
        col += 1

        # Ingresos
        cell = ws.cell(row=row_idx, column=col, value=row["ingresos"])
        cell.style = money_style
        col += 1

        # Costos por tipo (si no existe en contrato → 0)
        for tipo in tipos_order:
            monto = row["tipos_costos"].get(tipo, 0)
            cell = ws.cell(row=row_idx, column=col, value=monto)
            cell.style = money_style
            col += 1

        # Depreciaciones
        cell = ws.cell(row=row_idx, column=col, value=row["depreciaciones"])
        cell.style = money_style
        col += 1

        # Rentabilidad
        cell = ws.cell(row=row_idx, column=col, value=row["rentabilidad"])
        cell.style = money_style
        col += 1

        row_idx += 1

    # 👉 Totales con fórmula
    ws.cell(row=row_idx, column=1, value="TOTAL").font = Font(bold=True)
    for col in range(2, len(tipos_order) + 5):  # ingresos + tipos + amortización + rentabilidad
        col_letter = get_column_letter(col)
        start_row = fila + 1
        end_row = row_idx - 1
        formula = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
        cell = ws.cell(row=row_idx, column=col, value=formula)
        cell.style = money_style
        cell.font = Font(bold=True)

    # Ajustar anchos
    for i, col in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = max(
            (len(str(cell.value)) for cell in col if cell.value), default=10
        ) + 2

    # Respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_rentabilidad.xlsx"'
    wb.save(response)
    return response


