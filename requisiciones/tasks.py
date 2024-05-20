# tasks.py
from django.http import JsonResponse, HttpResponse 
from django.db.models import F, Avg, Value, Case, When 
from django.db.models.functions import Concat
from django.conf import settings
from django.core.files.storage import FileSystemStorage

from compras.models import Compra

from celery import shared_task
# Import Excel Stuff
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill
from openpyxl.utils import get_column_letter
import datetime as dt
import os
from datetime import date
from entradas.models import Entrada, EntradaArticulo
from tesoreria.models import Pago
from requisiciones.models import Salidas
from io import BytesIO
import xlsxwriter


@shared_task
def convert_entradas_to_xls_task(entradas):
    #response= HttpResponse(content_type = "application/ms-excel")
    #response['Content-Disposition'] = 'attachment; filename = Entradas_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Entradas')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    number_style = NamedStyle(name='number_style', number_format='#,##0.00')
    number_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(number_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)


    columns = ['Vale','Folio Solicitud','Folio Compra','Folio Req','Fecha','Solicitante','Proveedor','Proyecto','Subproyecto','Area','Código','Articulo','Cantidad','Moneda','Tipo de Cambio','Precio']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16

    columna_max = len(columns)+2

    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por SAVIA 2.0 UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style

    rows = []
    
    for item in entradas:
        # Obtén todos los pagos relacionados con esta entrada
        pk = item['id']
        entrada = EntradaArticulo.objects.get(id=pk)
        pagos = Pago.objects.filter(oc= entrada.entrada.oc)
        # Calcula el tipo de cambio promedio de estos pagos
        tipo_de_cambio_promedio_pagos = pagos.aggregate(Avg('tipo_de_cambio'))['tipo_de_cambio__avg']

        # Usar el tipo de cambio de los pagos, si existe. De lo contrario, usar el tipo de cambio de la entrada
        tipo_de_cambio = tipo_de_cambio_promedio_pagos or entrada.entrada.oc.tipo_de_cambio

        row = [
            entrada.entrada.folio,
            entrada.entrada.oc.req.orden.folio,
            entrada.entrada.oc.folio,
            #entrada.articulo_comprado.oc.folio,
            entrada.entrada.oc.req.folio,
            entrada.created_at.date(),
            f"{entrada.entrada.oc.req.orden.staff.staff.staff.first_name} {entrada.entrada.oc.req.orden.staff.staff.staff.last_name}",
            entrada.entrada.oc.proveedor.nombre.razon_social,
            entrada.entrada.oc.req.orden.proyecto.nombre,
            entrada.entrada.oc.req.orden.subproyecto.nombre,
            entrada.entrada.oc.req.orden.operacion.nombre if entrada.entrada.oc.req.orden.operacion else "Sin operación",
            entrada.articulo_comprado.producto.producto.articulos.producto.producto.codigo,
            entrada.articulo_comprado.producto.producto.articulos.producto.producto.nombre,
            entrada.cantidad,
            entrada.entrada.oc.moneda.nombre,
            tipo_de_cambio,
            entrada.articulo_comprado.precio_unitario,
        ]
        if row[9] == "DOLARES":
            if row[10] is None or row[10] < 15:
                row[10] = 17  # O cualquier valor predeterminado que desees
        elif row[10] is None:
                row[10] = ""

        rows.append(row)

    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            (ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style
            if col_num == 4:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style
            if col_num == 12:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = number_style
            if col_num == 14 or col_num == 15:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style =  money_style

    file_name='Matriz_Entradas_' + str(date.today()) + '.xlsx'
    file_storage_location = os.path.join(settings.MEDIA_ROOT,'reportes',file_name)

    sheet = wb['Sheet']
    wb.remove(sheet)
    #wb.save(response)
    wb.save(file_storage_location)
    fs = FileSystemStorage()
    with open(file_storage_location, 'rb') as excel_file:
        filename = fs.save('reportes/' + file_name, excel_file)
        file_url = fs.url(filename)
    # Puedes devolver el identificador único (nombre del archivo) o la URL para descargar.
    # return JsonResponse({'report_id': filename}) # Devolver el identificador único.
    return {'file_url': file_url}  # Devolver la URL para descargar.
    #return(response)


@shared_task
def convert_salidas_to_xls_task(salidas):
    #response= HttpResponse(content_type = "application/ms-excel")
    #response['Content-Disposition'] = 'attachment; filename = Salidas_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Salidas')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)
    number_style = NamedStyle(name='number_style', number_format='#,##0.00')
    number_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(number_style)

    columns = ['Vale Salida','Folio Solicitud','Fecha','Solicitante','Proyecto','Subproyecto','Área','Código','Articulo','Comentario','Cantidad','Precio','Total']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        if col_num == 3 or col_num == 8 or col_num == 9:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        else:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 16


    columna_max = len(columns)+2

    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por SAVIA 2.0. UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Grupo Vordcab S.A. de C.V.}')).style = messages_style
    
    rows = []
    for  salidas_list in salidas:
        pk = salidas_list['id']
        salida = Salidas.objects.get(id=pk)
        if salida.precio > 0:
            precio_condicional = salida.precio
        elif salida.producto.precio > 0:
            precio_condicional = salida.producto.precio
        else:
            precio_condicional = salida.producto.articulos.producto.price

        if salida.vale_salida.solicitud.proyecto:
            proyecto = f"{salida.vale_salida.solicitud.proyecto.nombre}"
        else:
            proyecto = " "

        if salida.vale_salida.solicitud.proyecto:
            subproyecto = f"{salida.vale_salida.solicitud.subproyecto.nombre}"
        else:
            subproyecto = " "

        if salida.vale_salida.comentario:
            comentario = f"{salida.vale_salida.comentario}"
        else:
            comentario = " "

        row = [
            salida.vale_salida.folio,
            salida.vale_salida.solicitud.folio,
            salida.created_at,
            f"{salida.producto.articulos.orden.staff.staff.staff.first_name} {salida.producto.articulos.orden.staff.staff.staff.last_name}",
            proyecto,
            subproyecto,
            salida.vale_salida.solicitud.operacion if salida.vale_salida.solicitud.operacion else "Sin operación",
            salida.producto.articulos.producto.producto.codigo,
            salida.producto.articulos.producto.producto.nombre,
            comentario,
            #f"{salida.vale_salida.material_recibido_por.staff.staff.first_name} {salida.vale_salida.material_recibido_por.staff.staff.last_name}",
            salida.cantidad,
            precio_condicional,
            
        ]

        rows.append(row)
    
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            (ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style
            if col_num == 2:
                value = (row[col_num]).date()
                (ws.cell(row = row_num, column = col_num+1, value=value)).style = date_style
            if col_num == 10:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = number_style
            if col_num == 11:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = money_style
        ws.cell(row=row_num, column=len(row) + 1, value=f'=K{row_num} * L{row_num}').style = money_style
    
    (ws.cell(column = columna_max , row = 3, value=f'=SUM(M2:M{row_num})')).style = money_resumen_style

    sheet = wb['Sheet']
    wb.remove(sheet)
    file_name='Matriz_Salidas_' + str(date.today()) + '.xlsx'
    file_storage_location = os.path.join(settings.MEDIA_ROOT,'reportes',file_name)
    #wb.save(response)
    wb.save(file_storage_location)
    fs = FileSystemStorage()
    with open(file_storage_location, 'rb') as excel_file:
        filename = fs.save('reportes/' + file_name, excel_file)
        file_url = fs.url(filename)
    # Puedes devolver el identificador único (nombre del archivo) o la URL para descargar.
    # return JsonResponse({'report_id': filename}) # Devolver el identificador único.
    return {'file_url': file_url}  # Devolver la URL para descargar.
    #return(response)
#Aquí termina la implementación del XLSX



