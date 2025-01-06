from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.core.files.base import ContentFile
from django.core.mail import EmailMessage, BadHeaderError
from smtplib import SMTPException
from django.core.paginator import Paginator
from django.db.models.functions import Concat
from django.db.models import Value, Sum, Case, When, F, Value, Q, Avg, Max
from django.contrib import messages
from django.http import JsonResponse, HttpResponse, FileResponse
from django.conf import settings
from django.utils import timezone
from celery.result import AsyncResult
import xlsxwriter
from xlsxwriter.utility import xl_col_to_name

from io import BytesIO

import os
import base64
import json
import csv
import ast # Para leer el csr many to many
import decimal
import io
import qrcode

from decimal import Decimal
from PIL import Image
from solicitudes.models import Proyecto, Subproyecto
from solicitudes.filters import SolicitudesFilter
from dashboard.models import Inventario, Order, ArticulosparaSurtir, ArticulosOrdenados, Inventario_Batch, Product, Marca
from dashboard.forms import  Inventario_BatchForm
from entradas.models import Entrada, EntradaArticulo
from requisiciones.models import Salidas, ValeSalidas
from tesoreria.models import Pago
from user.models import Profile, User
from user.decorators import perfil_seleccionado_required
from compras.models import Compra
from .models import ArticulosRequisitados, Requis, Devolucion, Devolucion_Articulos, Tipo_Devolucion
from .tasks import convert_entradas_to_xls_task, convert_salidas_to_xls_task
from .filters import ArticulosparaSurtirFilter, SalidasFilter, EntradasFilter, DevolucionFilter, RequisFilter, RequisProductosFilter, HistoricalSalidasFilter, Historical_articulos_surtir_filter
from .forms import SalidasForm, ArticulosRequisitadosForm, ValeSalidasForm, ValeSalidasProyForm, RequisForm, Rechazo_Requi_Form, DevolucionArticulosForm, DevolucionForm
#from compras.views import clear_task_id, verificar_estado
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill
from openpyxl.utils import get_column_letter
import datetime as dt
from datetime import date, datetime, timedelta
from pathlib import Path

from pyexcelerate import Workbook, Color as PXColor, Style, Font, Fill, Alignment, Format
#PDF generator
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.colors import Color, black, blue, red, white
from reportlab.lib.units import cm
from reportlab.lib.pagesizes import letter, portrait
from reportlab.rl_config import defaultPageSize 
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Frame
from bs4 import BeautifulSoup

import urllib.request, urllib.parse, urllib.error


@perfil_seleccionado_required
def requisiciones_status(request):
    pk = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk)
    
    if perfil.tipo.nombre == "PROVEEDORES" or perfil.tipo.nombre == "VIS_ADQ":
        requis = Requis.objects.filter(complete = True).order_by('-folio')
    else:
        requis = Requis.objects.filter(orden__distrito = perfil.distritos, complete = True).order_by('-folio')
   
    #requis = Requis.objects.filter(autorizar=True, colocada=False)

    myfilter = RequisFilter(request.GET, queryset=requis)
    requis = myfilter.qs

     #Set up pagination
    p = Paginator(requis, 50)
    page = request.GET.get('page')
    requis_list = p.get_page(page)

    if request.method == 'POST' and 'btnExcel' in request.POST:
        return convert_excel_matriz_requis(requis)

    context= {
        'myfilter': myfilter,
        'requis':requis,
        'requis_list':requis_list,
        }

    return render(request, 'requisiciones/requisiciones.html',context)


@perfil_seleccionado_required
def requisiciones_productos(request):
    pk = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk)
    
    if perfil.tipo.nombre == "PROVEEDORES" or perfil.tipo.nombre == "VIS_ADQ":
        articulos_requisitados = ArticulosRequisitados.objects.filter(req__complete = True).order_by('-req__folio')
    else:
        articulos_requisitados = ArticulosRequisitados.objects.filter(req__orden__distrito = perfil.distritos, req__complete = True).order_by('-req__folio')
   
    #requis = Requis.objects.filter(autorizar=True, colocada=False)

    myfilter = RequisProductosFilter(request.GET, queryset= articulos_requisitados)
    articulos_requisitados = myfilter.qs

     #Set up pagination
    p = Paginator(articulos_requisitados, 50)
    page = request.GET.get('page')
    articulos_requisitados_list = p.get_page(page)

    if request.method == 'POST' and 'btnExcel' in request.POST:
        return convert_excel_matriz_requis_productos(articulos_requisitados)

    context= {
        'myfilter': myfilter,
        'articulos_requisitados': articulos_requisitados,
        'articulos_requisitados_list': articulos_requisitados_list,
        }

    return render(request, 'requisiciones/requisiciones_productos.html',context)


# Create your views here.
@perfil_seleccionado_required
def liberar_stock(request, pk):
    usuario = Profile.objects.get(staff__id=request.user.id)
    orden = Order.objects.get(id = pk)
    productos= ArticulosparaSurtir.objects.filter(articulos__orden = orden, surtir=True)
    vale_salida, created = ValeSalidas.objects.get_or_create(almacenista = usuario,complete = False,solicitud=orden)
    salidas = Salidas.objects.filter(vale_salida = vale_salida)
    cantidad_items = salidas.count()
    proyectos = Proyecto.objects.filter(activo=True)
    subproyectos = Subproyecto.objects.all()


    formVale = ValeSalidasProyForm()
    form = SalidasForm()
    users = Profile.objects.all()

    if request.method == 'POST':
        formVale = ValeSalidasProyForm(request.POST, instance=vale_salida)
        vale = formVale.save(commit=False)
        vale.complete = True
        for producto in productos:
            if producto.cantidad == 0:
                producto.salida = True
                producto.surtir = False
                producto.save()
        if formVale.is_valid():
            formVale.save()
            messages.success(request,'La salida se ha generado de manera exitosa')
            return redirect('solicitud-autorizada')

    context= {
        'proyectos':proyectos,
        'subproyectos':subproyectos,
        'productos':productos,
        'orden':orden,
        'form':form,
        'formVale':formVale,
        'users': users,
        'vale_salida':vale_salida,
        'cantidad_items':cantidad_items,
        'salidas':salidas,
        }
    return render(request,'requisiciones/liberar_stock.html',context)



@perfil_seleccionado_required
def solicitud_autorizada(request):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)

    if usuario.tipo.almacen == True:
        #productos= ArticulosparaSurtir.objects.filter(Q(salida=False) | Q(surtir=True), articulos__orden__autorizar = True)
        #productos= ArticulosparaSurtir.objects.filter(Q(salida=False) | Q(surtir=True), articulos__orden__autorizar = True, articulos__orden__tipo__tipo = "normal")
        #productos= ArticulosparaSurtir.objects.filter(surtir=True, articulos__orden__autorizar = True, articulos__orden__tipo__tipo = "normal", articulos__orden__distrito = usuario.distritos, articulos__orden__complete = True).order_by('-articulos__orden__id')
        productos = ArticulosparaSurtir.objects.filter(surtir=True, articulos__orden__autorizar=True, articulos__orden__tipo__tipo="normal", articulos__orden__distrito=usuario.distritos, articulos__orden__complete=True).order_by('-id')

    #else:
        #productos = Requis.objects.filter(complete=None)
    myfilter = ArticulosparaSurtirFilter(request.GET, queryset=productos)
    productos = myfilter.qs
    #Here is where call a function to generate XLSX, using Openpyxl library

    #Set up pagination
    p = Paginator(productos, 15)
    page = request.GET.get('page')
    productos_list = p.get_page(page)


    if request.method == 'POST' and 'btnExcel' in request.POST:
        return convert_solicitud_autorizada_to_xls(productos)


    context= {
        'productos':productos,
        'productos_list':productos_list,
        'myfilter':myfilter,
        'usuario':usuario,
        }
    return render(request, 'requisiciones/solicitudes_autorizadas.html',context)

@perfil_seleccionado_required
def solicitudes_autorizadas_pendientes(request):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    
    if usuario.tipo.almacenista == True:
        productos= ArticulosparaSurtir.objects.filter(salida=False, surtir=False, articulos__orden__autorizar = True, articulos__orden__tipo__tipo = "normal", articulos__orden__distrito = usuario.distritos).order_by('-articulos__orden__approved_at')

    myfilter = ArticulosparaSurtirFilter(request.GET, queryset=productos)
    productos = myfilter.qs

    #Set up pagination
    p = Paginator(productos, 20)
    page = request.GET.get('page')
    productos_list = p.get_page(page)

    #Here is where call a function to generate XLSX, using Openpyxl library

    if request.method == 'POST' and 'btnExcel' in request.POST:
        return convert_solicitud_autorizada_to_xls(productos)


    context= {
        'productos_list':productos_list,
        #'productos':productos,
        'myfilter':myfilter,
        'usuario':usuario,
        }
    return render(request, 'requisiciones/solicitudes_autorizadas_no_surtidas.html',context)


def update_devolucion(request):
    data= json.loads(request.body)
    action = data["action"]
    cantidad = decimal.Decimal(data["val_cantidad"])
    devolucion = data["devolucion"]
    producto_id = data["id"]
    comentario = data["comentario"]
    devolucion = Devolucion.objects.get(id = devolucion)
    
    
    if devolucion.tipo.nombre == "SALIDA":
        producto = Salidas.objects.get(vale_salida=devolucion.salida.vale_salida, producto__id = producto_id)
        inv_del_producto = Inventario.objects.get(producto = producto.producto.articulos.producto.producto, distrito =producto.vale_salida.solicitud.distrito)
    else:
        producto = ArticulosparaSurtir.objects.get(id = producto_id)
        inv_del_producto = Inventario.objects.get(producto = producto.articulos.producto.producto, distrito =producto.articulos.orden.distrito)
        


    if action == "add":
        cantidad_total = producto.cantidad - cantidad
        if cantidad_total < 0:
            messages.error(request,f'La cantidad que se quiere ingresar sobrepasa la cantidad disponible. {cantidad_total} mayor que {producto.cantidad}')
        else:
            if devolucion.tipo.nombre == "SALIDA":
                devolucion_articulos, created = Devolucion_Articulos.objects.get_or_create(producto= producto.producto, vale_devolucion = devolucion, complete=False)
            else:
                devolucion_articulos, created = Devolucion_Articulos.objects.get_or_create(producto=producto, vale_devolucion = devolucion, complete=False)
            
            producto.seleccionado = True
            #Se le resta a la cantidad de artículos para surtir
            producto.cantidad = producto.cantidad - cantidad
            producto.cancelada_at = datetime.now()
            #La cantidad de la devolución es igual a la cantidad que se marcó en la devolución (daaa)
            devolucion_articulos.cantidad = cantidad
            devolucion_articulos.comentario = comentario
            devolucion_articulos.precio = producto.precio
            devolucion_articulos.complete = True
            if producto.cantidad == 0: #Si la cantidad de artículos para surtir es igual a 0, si la cantidad a devolver es 0 entonces ya no se puede surtir
                producto.surtir = False
            messages.success(request,'Has agregado producto para devolución de manera exitosa')
            producto.save()
            devolucion_articulos.save()
    if action == "remove":
        if devolucion.tipo.nombre == "SALIDA":
            item = Devolucion_Articulos.objects.get(producto=producto.producto, vale_devolucion = devolucion, complete = True)
        else:
            item = Devolucion_Articulos.objects.get(producto=producto, vale_devolucion = devolucion, complete = True)
        producto.cantidad = producto.cantidad + item.cantidad
        producto.seleccionado = False
        messages.success(request,'Has eliminado un producto de tu listado')
        producto.save()
        item.delete()

    return JsonResponse('Item updated, action executed: '+data["action"], safe=False)

@perfil_seleccionado_required
def autorizar_devolucion(request, pk):
    print("Estoy en autorizar_devolucion")
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    devolucion= Devolucion.objects.get(id=pk)
    productos = Devolucion_Articulos.objects.filter(vale_devolucion = devolucion)
    if devolucion.tipo.nombre == "SALIDA":
        salida = Salidas.objects.get(id=devolucion.salida.id)
    
    if request.method == 'POST' and 'btnAutorizar' in request.POST:
        for producto in productos:
            if devolucion.tipo.nombre == "SALIDA":
                producto_surtir = Salidas.objects.get(id=devolucion.salida.id)
                inv_del_producto = Inventario.objects.get(producto = producto.producto.articulos.producto.producto, distrito = usuario.distritos ) 
                inv_del_producto._change_reason = f'Esta es una devolucion desde un salida {devolucion.id}'
            else:
                producto_surtir = ArticulosparaSurtir.objects.get(articulos = producto.producto.articulos)
                inv_del_producto = Inventario.objects.get(producto = producto_surtir.articulos.producto.producto, distrito = usuario.distritos )
                inv_del_producto._change_reason = f'Esta es una devolucion desde un surtimiento de inventario {devolucion.id}'
                #try: Por el momento es un poco confuso lo que se quiere lograr acá, por ello lo voy a comentar, al parecer es un intento de regresar a la entrada ese material pero puede generar más errores que beneficios
                    # y en realidad lo único requerido es devolver el material al almacén  
                    #entrada = EntradaArticulo.objects.get(articulo_comprado__producto__producto=producto_surtir, entrada__oc__req__orden=producto_surtir.articulos.orden, agotado = False)
                    
                    # Verificar si la cantidad en la entrada es suficiente
                    #if entrada.cantidad_por_surtir >= producto.cantidad:
                    #    print(entrada)
                        # Reducir la cantidad de la entrada según la cantidad de la devolución
                    #    entrada.cantidad_por_surtir -= producto.cantidad 
                    #   entrada.save()
                    #else:
                        # Manejar el caso en que no hay suficiente cantidad en la entrada (opcional)
                    #    entrada.cantidad_por_surtir = 0
                    #    entrada.agotado = True
                    #    entrada.save()
                #except EntradaArticulo.DoesNotExist:
                    # Manejar el caso en que no hay una entrada asociada (opcional)
                    #messages.error(request, 'No se encontró una entrada asociada para el producto.')
                    
            inv_del_producto.cantidad = inv_del_producto.cantidad + producto.cantidad
            inv_del_producto.save()
            messages.success(request,'Has autorizado exitosamente una devolución')
        devolucion.autorizada = True
        if devolucion.tipo.nombre == "SALIDA":
            salida.cancelada = True
            salida.save()
        devolucion.save()
        return redirect('matriz-autorizar-devolucion')
    elif request.method == 'POST' and 'btnCancelar' in request.POST:
        for producto in productos:
            if devolucion.tipo.nombre == "SALIDA":
                producto_surtir = Salidas.objects.get(id=devolucion.salida.id)
            else:
                producto_surtir = ArticulosparaSurtir.objects.get(articulos = producto.producto.articulos)
            producto_surtir.cantidad = producto_surtir.cantidad + producto.cantidad
            producto_surtir.surtir = True
            producto_surtir.save()
            #inv_del_producto.save()
        devolucion.autorizada = False
        devolucion.save()
        return redirect('matriz-autorizar-devolucion')
    
    
    context= {
        'productos':productos,
        'devolucion':devolucion,
        }

    return render(request, 'requisiciones/autorizar_devolucion.html',context)

@perfil_seleccionado_required
def cancelar_devolucion(request, pk):
    devolucion= Devolucion.objects.get(id=pk)
    productos = Devolucion_Articulos.objects.filter(vale_devolucion = devolucion)

    if request.method == 'POST' and 'btnCancelar' in request.POST:
        for producto in productos:
            if devolucion.tipo.nombre == "SALIDA":
                producto_surtir = Salidas.objects.get(salida=devolucion.salida)
            else:
                producto_surtir = ArticulosparaSurtir.objects.get(articulos = producto.producto.articulos)
            producto_surtir.cantidad = producto_surtir.cantidad + producto.cantidad
            producto_surtir.surtir = True
            producto_surtir.save()
            #inv_del_producto.save()
        devolucion.autorizada = False
        devolucion.save()
        return redirect('matriz-autorizar-devolucion')

    context= {
        'productos':productos,
        'devolucion':devolucion,
        }

    return render(request, 'requisiciones/cancelar_devolucion.html',context)


@perfil_seleccionado_required
def matriz_autorizar_devolucion(request):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    devoluciones= Devolucion.objects.filter(complete=True, autorizada=None, solicitud__distrito = usuario.distritos)
    #print(devoluciones)

    
    myfilter = DevolucionFilter(request.GET, queryset = devoluciones)
    devoluciones = myfilter.qs

    #Set up pagination
    p = Paginator(devoluciones, 20)
    page = request.GET.get('page')
    devoluciones_list = p.get_page(page)

    #Here is where call a function to generate XLSX, using Openpyxl library

    #if request.method == 'POST' and 'btnExcel' in request.POST:
    #    return convert_solicitud_autorizada_to_xls(productos)


    context= {
        'devoluciones_list':devoluciones_list,
        'devoluciones':devoluciones,
        'myfilter':myfilter,
        'usuario':usuario,
        }
    return render(request, 'requisiciones/matriz_devoluciones_autorizar.html',context)

@perfil_seleccionado_required
def salida_material(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    orden = Order.objects.get(id = pk)
    print(orden)
    productos = ArticulosparaSurtir.objects.filter(articulos__orden = orden, surtir=True)
    productos_no_seleccionados = productos.filter(seleccionado_salida = False)
    vale_salidas = ValeSalidas.objects.filter(solicitud__distrito = usuario.distritos)
    vale_salida, created = vale_salidas.get_or_create(almacenista = usuario,complete = False,solicitud=orden)
    salidas = Salidas.objects.filter(vale_salida = vale_salida)
    cantidad_items = salidas.count()
    # Obtener el valor máximo actual de folio para los objetos ValeSalidas del distrito del usuario
    max_folio = vale_salidas.aggregate(Max('folio'))['folio__max']
    # Calcular el nuevo folio como el consecutivo del máximo actual
    nuevo_folio = (max_folio or 0) + 1

    formVale = ValeSalidasForm()
    form = SalidasForm()
    users = Profile.objects.filter(distritos = usuario.distritos, st_activo = True )

    material_recibido_por = [
        {'id': user.id, 
         'text': str(user.staff.staff.first_name) + (' ') + str(user.staff.staff.last_name)
         #'distrito': proveedor.
        } for user in users]

    if request.method == 'POST':
        formVale = ValeSalidasForm(request.POST, instance=vale_salida)
        
        if formVale.is_valid():
            #formVale.save()
            vale = formVale.save(commit=False)
            cantidad_salidas = 0
            cantidad_productos = productos.count()
            for producto in productos:
                producto.seleccionado_salida = False
                print(producto,"cantidad:", producto.cantidad)
                if producto.cantidad <= 0:
                    producto.salida=True
                    producto.surtir=False
                    cantidad_salidas = cantidad_salidas + 1
                    print(producto)
                producto.save()
            if cantidad_productos == cantidad_salidas:
                orden.requisitado == True #Esta variable creo que podría ser una variable estúpida
                orden.save()
            vale.created_at = date.today()
            vale.complete = True
            max_folio = vale_salidas.aggregate(Max('folio'))['folio__max']
            nuevo_folio = (max_folio or 0) + 1
            vale.folio = nuevo_folio
            vale.save()
            messages.success(request,'La salida se ha generado de manera exitosa')
            return redirect('reporte-salidas')
        if not formVale.is_valid():
            messages.error(request,'No capturaste el usuario')

    context= {
        'productos':productos_no_seleccionados,
        'form':form,
        'formVale':formVale,
        'users': users,
        'material_recibido_por':material_recibido_por,
        'nuevo_folio':nuevo_folio,
        'vale_salida':vale_salida,
        'cantidad_items':cantidad_items,
        'salidas':salidas,
        }

    return render(request, 'requisiciones/salida_material.html',context)

@perfil_seleccionado_required
def devolucion_material(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    orden = Order.objects.get(id = pk)
    productos_sel = ArticulosparaSurtir.objects.filter(articulos__orden = orden, surtir=True)
    tipo = Tipo_Devolucion.objects.get(nombre ="APARTADO" )
    devolucion, created = Devolucion.objects.get_or_create(almacenista = usuario,complete = False,solicitud=orden, tipo=tipo)
    productos = Devolucion_Articulos.objects.filter(vale_devolucion = devolucion)
    cantidad_items = productos.count()
    form = DevolucionArticulosForm()
    form2 = DevolucionForm()

    form.fields['producto'].queryset = productos_sel

    if request.method == 'POST':
        if "agregar_devolucion" in request.POST:
            form2 = DevolucionForm(request.POST, instance=devolucion)
            if form2.is_valid():
                devolucion = form2.save(commit=False)
                devolucion.complete= True
                devolucion.hora = datetime.now().time()
                devolucion.fecha = date.today()
                devolucion.tipo.nombre = "SIN SALIDA" 
                devolucion.save()
                for producto in productos_sel:
                    producto.seleccionado = False
                    producto.save()
                messages.success(request,f'{usuario.staff.staff.first_name}, Has hecho la devolución de manera exitosa')
                email = EmailMessage(
                    f'Cancelación de solicitud: {orden.folio}',
                    f'Estimado {orden.staff.staff.staff.first_name} {orden.staff.staff.staff.last_name},\n Estás recibiendo este correo porque tu solicitud: {orden.folio} ha sido devuelta al almacén por {usuario.staff.staff.first_name} {usuario.staff.staff.last_name}, con el siguiente comentario {devolucion.comentario} para más información comunicarse al almacén.\n\n Este mensaje ha sido automáticamente generado por SAVIA 2.0',
                    settings.DEFAULT_FROM_EMAIL,
                    ['ulises_huesc@hotmail.com'],#orden.staff.staff.email],
                    )
                email.send()
                return redirect('solicitud-autorizada')

    context= {
        'orden':orden,
        'productos':productos,
        'form':form,
        'form2':form2,
        'devolucion': devolucion,
        'cantidad_items':cantidad_items,
        'productos_sel': productos_sel,
        }

    return render(request, 'requisiciones/devolucion_material_no_salida.html',context)

@perfil_seleccionado_required
def devolucion_material_salida(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    salidas = Salidas.objects.all()
    salida = salidas.get(id=pk)
    vale_salida = ValeSalidas.objects.get(id=salida.vale_salida.id)
    orden = Order.objects.get(id = vale_salida.solicitud.id)
    #Esta es la parte que varía de devolución de material, aquí los productos deben ser salida = True
    #productos_sel = ArticulosparaSurtir.objects.filter(articulos__orden = orden, salida=True)
    productos_sel = salida
    #print(productos_sel)
   
    tipo = Tipo_Devolucion.objects.get(nombre ="SALIDA" )
    devolucion, created = Devolucion.objects.get_or_create(almacenista = usuario,complete = False,solicitud=orden,tipo =tipo, salida =salida)
    productos = Devolucion_Articulos.objects.filter(vale_devolucion = devolucion)
    cantidad_items = productos.count()
    form = DevolucionArticulosForm()
    form2 = DevolucionForm()


    if request.method == 'POST':
        if "agregar_devolucion" in request.POST:
            form2 = DevolucionForm(request.POST, instance=devolucion)
            if form2.is_valid():
                devolucion = form2.save(commit=False)
                devolucion.complete= True
                devolucion.hora = datetime.now().time()
                devolucion.fecha = date.today()
                devolucion.save()
                #for producto in productos_sel:
                productos_sel.seleccionado = False
                productos_sel.save()
                messages.success(request,f'{usuario.staff.staff.first_name}, Has hecho la devolución de manera exitosa')
                email = EmailMessage(
                    f'Cancelación de solicitud: {orden.folio}',
                    f'Estimado {orden.staff.staff.staff.first_name} {orden.staff.staff.staff.last_name},\n Estás recibiendo este correo porque tu solicitud: {orden.folio} ha sido devuelta al almacén por {usuario.staff.staff.first_name} {usuario.staff.staff.last_name}, con el siguiente comentario {devolucion.comentario} para más información comunicarse al almacén.\n\n Este mensaje ha sido automáticamente generado por SAVIA VORDTEC',
                    settings.DEFAULT_FROM_EMAIL,
                    ['ulises_huesc@hotmail.com'],#orden.staff.staff.email],
                    )
                email.send()
                return redirect('reporte-salidas')

    context= {
        'orden':orden,
        'productos':productos,
        'form':form,
        'form2':form2,
        'devolucion': devolucion,
        'cantidad_items':cantidad_items,
        'productos_sel': productos_sel,
        }

    return render(request, 'requisiciones/devolucion_material.html',context)

@perfil_seleccionado_required
def solicitud_autorizada_firma(request):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    #Aquí aparecen todas las ordenes, es decir sería el filtro para administrador
    productos= Salidas.objects.filter(producto__articulos__orden__autorizar = True, salida_firmada=False)
    myfilter = SalidasFilter(request.GET, queryset=productos)
    productos = myfilter.qs

    #Here is where XLSX is generated, using Openpyxl library | Aquí es donde se genera el XLSX
    if request.method == "POST" and 'btnExcel' in request.POST:

        return convert_solicitud_autorizada_orden_to_xls(productos)

    context= {
        'productos':productos,
        'myfilter':myfilter,
        'usuario':usuario,
        }
    return render(request, 'requisiciones/solicitudes_autorizadas_firma.html',context)


def update_salida(request):
    data= json.loads(request.body)
    action = data["action"]
    cantidad = decimal.Decimal (data["val_cantidad"])
    vale_salida_id = data["salida"]
    producto_id = data["id"]
    id_salida =data["id_salida"]
    producto = ArticulosparaSurtir.objects.get(id = producto_id)
    print(producto_id, producto, 'cantidad_sal', cantidad)
    vale_salida = ValeSalidas.objects.get(id = vale_salida_id)
    inv_del_producto = Inventario.objects.get(producto = producto.articulos.producto.producto, distrito = producto.articulos.orden.distrito)
    entradas = EntradaArticulo.objects.filter(articulo_comprado__producto__producto = producto, agotado=False, entrada__oc__req__orden= producto.articulos.orden).aggregate(cantidad_surtir=Sum('cantidad_por_surtir'))
    suma_entradas = entradas['cantidad_surtir']
    #Si no existen entradas la suma_entradas es igual a None, lo convierto en 0 para que pueda pasar la condicional #Definitoria
    if suma_entradas == None:
        suma_entradas = 0

    if action == "add":
        #con cantidad total establezco si la "cantidad" no sobrepasa lo que tengo que surtir(producto.cantidad)     
        cantidad_total = producto.cantidad - cantidad
        producto.seleccionado_salida = True
        entradas_dir = EntradaArticulo.objects.filter(articulo_comprado__producto__producto=producto, agotado=False, entrada__oc__req__orden=producto.articulos.orden, articulo_comprado__producto__producto__articulos__orden__tipo__tipo = 'normal').order_by('id')

        try:
            EntradaArticulo.objects.filter(articulo_comprado__producto__producto__articulos__producto = inv_del_producto, articulo_comprado__producto__producto__articulos__orden__tipo__tipo = 'resurtimiento', agotado = False)
            
        except EntradaArticulo.DoesNotExist:
            entrada_res = None
        else:
            entrada_res = EntradaArticulo.objects.filter(articulo_comprado__producto__producto__articulos__producto = inv_del_producto, articulo_comprado__producto__producto__articulos__orden__tipo__tipo = 'resurtimiento', agotado = False).order_by('id')
        print('2',entrada_res)
        if entradas_dir.exists():
            entradas = EntradaArticulo.objects.filter(articulo_comprado__producto__producto = producto, agotado=False, entrada__oc__req__orden= producto.articulos.orden)
            for entrada in entradas:
                if producto.cantidad > 0:
                    salida, created = Salidas.objects.get_or_create(producto=producto, vale_salida = vale_salida, complete=False)
                    salida.precio = entrada.articulo_comprado.precio_unitario
                    if entrada.cantidad_por_surtir >= cantidad:
                        salida.cantidad = cantidad
                        cantidad = 0 #la cantidad se vuelve 0 porque si la condición se cumple indica que la cantidad por surtir es capaz de abastecer toda la cantidad
                        producto.cantidad = producto.cantidad - salida.cantidad
                        salida.entrada = entrada.id
                        entrada.cantidad_por_surtir = entrada.cantidad_por_surtir - salida.cantidad
                        salida.complete = True
                        if entrada.cantidad_por_surtir <= 0:
                            entrada.agotado = True
                        producto.save()
                        entrada.save()
                        salida.save()
                    elif entrada.cantidad_por_surtir < cantidad and cantidad > 0: #Le meto la condicional para que no se repita el proceso si la cantidad es igual o menor que 0 
                        salida.cantidad = entrada.cantidad_por_surtir #No puedo surtir mas que la cantidad que tengo disponible en la entrada
                        cantidad = cantidad - salida.cantidad #La nueva cantidad a surtir es la cantidad menos lo que ya salió
                        producto.cantidad = producto.cantidad - salida.cantidad
                        salida.entrada = entrada.id
                        salida.complete = True
                        entrada.agotado = True
                        entrada.cantidad_por_surtir = 0
                        #producto.salida =
                        #True si vuelvo la entrada de resurtimiento verdadera anulo la posibilidad de realizar más salidas
                        producto.save()
                        entrada.save()
                        salida.save()
                    inv_del_producto.cantidad_entradas = inv_del_producto.cantidad_entradas - salida.cantidad
                    #inv_del_producto.cantidad = inv_del_producto.cantidad - salida.cantidad si hago una salida que proviene de entradas voy a obtener un inv_del_producto negativo
                    inv_del_producto.save()
        elif entrada_res.exists():   #si hay resurtimiento
            for entrada in entrada_res:
                if cantidad > 0: #Se cambia producto.cantidad, se tiene que comparar con la cantidad de la salida no contra la cantidad disponible
                    salida, created = Salidas.objects.get_or_create(producto=producto, vale_salida = vale_salida, complete=False)
                    #Que hace el código a continuación la cantidad de la salida se compara contra la cantidad por surtir de la entrada
                    #L1 si es mayor la se guarda la cantidad_ant 
                    #L2 se le resta a la cantidad lo que queda en la entrada, es decir la nueva cantidad es lo que no se pudo surtir con esa entrada(cantidad) 
                    #L3 la cantidad de la salida es igual a la cantidad original menos la cantidad que no se pudo surtir con esa entrada
                    #L4 se vacía la entrada y por lo tanto se marca como agotada.
                    # Y si no la cantidad de la salida es igual a la cantidad(puede ser modificada por el bucle anterior o no) y 
                    # entrada por surtir es igual a la cantidad por surtir menos la cantidad de la salida y la cantidad se agota 04/12/2024 
                    if cantidad >= entrada.cantidad_por_surtir:
                        cantidad_ant = cantidad
                        cantidad = cantidad - entrada.cantidad_por_surtir
                        salida.cantidad = cantidad_ant - cantidad
                        entrada.cantidad_por_surtir = 0
                        entrada.agotado = True
                    else:
                        salida.cantidad = cantidad
                        entrada.cantidad_por_surtir = entrada.cantidad_por_surtir - salida.cantidad
                        cantidad = 0
                    producto.cantidad = producto.cantidad - salida.cantidad
                    salida.entrada = entrada.id
                    salida.complete = True
                    #if producto.cantidad_requisitar <= 0: #Esta línea se considera errónea 04/12/2024
                    #    producto.requisitar = False  #Esta línea se considera errónea 04/12/2024
                    if producto.cantidad <= 0:
                        producto.surtir = False
                    print(salida)
                    entrada.save()
                    producto.save()
                    inv_del_producto.cantidad_entradas = inv_del_producto.cantidad_entradas - salida.cantidad
                    inv_del_producto._change_reason = f'Esta es la salida de un artículo desde un resurtimiento de inventario {salida.id}'
                    salida.precio = entrada.articulo_comprado.precio_unitario
                    salida.save()
        else:    #si no hay resurtimiento
             # Verificar si ya existe un registro similar creado en el último segundo
            now = timezone.now()
            similar_entries = Salidas.objects.filter(
                producto=producto,
                vale_salida=vale_salida,
                cantidad=cantidad,
                complete=False,
                created_at__gte=now - timedelta(seconds=1)
            )
            if not similar_entries.exists():
                salida, created = Salidas.objects.get_or_create(producto=producto, vale_salida = vale_salida, complete=False)
                salida.cantidad = cantidad
                salida.entrada = 0
                salida.complete = True
                producto.cantidad = producto.cantidad - salida.cantidad 
                if producto.cantidad_requisitar <= 0:
                    producto.requisitar = False
                salida.precio = inv_del_producto.price
                inv_del_producto._change_reason = f'Esta es la salida de inventario de un artículo'   
                salida.save()
        producto.save()
        inv_del_producto.save()
        
    if action == "remove":
        item = Salidas.objects.get(vale_salida = vale_salida, id = id_salida)
        id_entrada = item.entrada
        if id_entrada != None:
            if id_entrada != 0:
                entrada = EntradaArticulo.objects.get(id=item.entrada)
                inv_del_producto.cantidad_entradas = inv_del_producto.cantidad_entradas + item.cantidad
                entrada.cantidad_por_surtir = entrada.cantidad_por_surtir + item.cantidad
                entrada.agotado = False
                entrada.save()
            #if entrada.entrada.oc.req.orden.tipo.tipo == "normal":
            #    inv_del_producto.cantidad_apartada = inv_del_producto.cantidad_apartada + item.cantidad
        if vale_salida.solicitud.tipo.tipo == "normal":
            inv_del_producto.cantidad_apartada = inv_del_producto.cantidad_apartada + item.cantidad
        #inv_del_producto.cantidad = inv_del_producto.cantidad + item.cantidad
        producto.seleccionado_salida = False
        producto.salida= False
        producto.cantidad = producto.cantidad + item.cantidad
        producto.surtir = True
        #producto.cantidad_requisitar = producto.cantidad_requisitar + producto.cantidad
        producto._change_reason = f'Esto es una eliminación de un artículo en una salida'
        inv_del_producto._change_reason = f'Esta es una eliminación de un artìculo en una salida {item.id}'
        producto.save()
        inv_del_producto.save()
        item.delete()

    return JsonResponse('Item updated, action executed: '+data["action"], safe=False)


@perfil_seleccionado_required
def salida_material_usuario(request, pk):
    producto= Salidas.objects.get(id = pk)
    producto_surtir = ArticulosparaSurtir.objects.get(articulos = producto.producto.articulos)

    if request.method == 'POST':
        producto.salida_firmada = True
        producto_surtir.salida = True
        producto_surtir.firma = True
        producto_surtir.save()
        producto.save()

        messages.success(request,f'Has realizado la salida del producto {producto.producto.articulos.producto.producto} con éxito')
        return redirect('solicitud-autorizada-firma')

    context= {
        'productos':producto,
    }

    return render(request, 'requisiciones/salida_material_usuario.html',context)

@perfil_seleccionado_required
def matriz_salida_activos(request):
    productos = Salidas.objects.filter(validacion_activos = False, producto__articulos__producto__producto__activo = True)
    #producto_surtir = ArticulosparaSurtir.objects.get(articulos = producto.producto.articulos)
    #activo = Activo.objects.filter(activo = productos.producto.producto)


    context= {
        'productos':productos,
    }

    return render(request, 'requisiciones/matriz_salida_activos.html',context)

@perfil_seleccionado_required
def solicitud_autorizada_orden(request):
    pk_perfil = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk_perfil)
    ordenes = Order.objects.filter(requisitar = True, complete=True, autorizar=True, staff__distritos=perfil.distritos, requisitado = False)


    if perfil.tipo.almacenista == True:
        ordenes = Order.objects.filter(requisitar = True, requisitado=False, distrito = perfil.distritos).order_by('-folio')

    #Este es un filtro por perfil supervisor o superintendente, es decir puede ver todo lo del distrito
    #productos= ArticulosparaSurtir.objects.filter(Q(salida=False) | Q(requisitar=True), articulos__orden__autorizar = True )
    myfilter=SolicitudesFilter(request.GET, queryset=ordenes)
    ordenes = myfilter.qs

    p = Paginator(ordenes, 50)
    page = request.GET.get('page')
    ordenes = p.get_page(page)


    if request.method == "POST" and 'btnExcel' in request.POST:

        return convert_solicitud_autorizada_orden_to_xls(ordenes)

    context= {
        'ordenes':ordenes,
        'myfilter':myfilter,
        }

    return render(request, 'requisiciones/solicitudes_autorizadas_orden.html',context)

@perfil_seleccionado_required
def detalle_orden(request, pk):
    orden = Order.objects.get(id=pk)
    productos = ArticulosparaSurtir.objects.filter(articulos__orden__id = pk, requisitar= True)

    context = {
        'productos': productos,
        'orden': orden,
     }
    return render(request,'requisiciones/orden_detail.html', context)


@perfil_seleccionado_required
def requisicion_autorizacion(request):
    pk_perfil = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk_perfil)
    #perfil = Profile.objects.get(staff__id=request.user.id)
    #obtengo el id de usuario, lo paso como argumento a id de profiles para obtener el objeto profile que coindice con ese usuario_id

    #Este es un filtro por perfil supervisor o superintendente, es decir puede ver todo lo del distrito
    
    #ordenes = Order.objects.filter(complete=True, autorizar=True, staff__distrito=perfil.distrito)
    if perfil.distritos.nombre == "MATRIZ" or perfil.distritos.nombre == "BRASIL" and perfil.tipo.supervisor:   
        requis = Requis.objects.filter(autorizar=None, orden__supervisor = perfil, complete =True)
    elif perfil.tipo.superintendente == True and perfil.tipo.nombre != "Admin":
        requis = Requis.objects.filter(autorizar=None, orden__superintendente=perfil, complete =True)
    elif perfil.tipo.nombre == "Admin":
        requis = Requis.objects.filter(autorizar=None, complete = True, orden__distrito = perfil.distritos)
    #else:
        #requis = Requis.objects.filter(complete=None)
    #requis = Requis.objects.filter(autorizar=None)


    context= {
        'requis':requis,
        }

    return render(request, 'requisiciones/requisiciones_autorizacion.html',context)

@perfil_seleccionado_required
def requisicion_creada_detalle(request, pk):
    productos = ArticulosRequisitados.objects.filter(req = pk)
    requis = Requis.objects.get(id = pk)

    context = {
        'productos': productos,
        'requis': requis,
     }

    return render(request,'requisiciones/requisicion_creada_detalle.html', context)

def update_requisicion(request):
    data= json.loads(request.body)
    action = data["action"]
    producto_id = data["id"]
    pk = data["requi"]
    cantidad = decimal.Decimal(data["cantidad"])
    
    requi = Requis.objects.get(id=pk)
    orden = Order.objects.get(id=requi.orden.id)

    producto = ArticulosparaSurtir.objects.get(id = producto_id)
    ordenado = ArticulosOrdenados.objects.filter(orden = orden, producto = producto.articulos.id)
   
    if action == "add":
        item, created = ArticulosRequisitados.objects.get_or_create(req=requi, producto = producto, cantidad = cantidad)  
        producto.requisitar = False
        producto.seleccionado = True
        producto.save()
        item.save()
    if action == "remove":
        item = ArticulosRequisitados.objects.get(req = requi, producto = producto)
        articulo_requisitado = ArticulosparaSurtir.objects.get(id =producto_id)
        articulo_requisitado.requisitar = True
        articulo_requisitado.seleccionado = False
        articulo_requisitado.save()
        item.delete()

    return JsonResponse('Item updated, action executed: '+data["action"], safe=False)

def obtener_consecutivo(distrito, requis):
    # Obtener la última requisición del distrito basado en la fecha de creación
    ultima_requisicion = requis.filter(orden__staff__distrito=distrito, complete=True).order_by('-created_at').first()

    if not ultima_requisicion:
        # Si no hay ninguna requisición previa, devolver 1 (será el primer folio)
        return 1

    # Extraer el número de folio (después de la abreviatura del distrito)
    ultimo_numero_folio = int(ultima_requisicion.folio.replace(distrito.abreviado, ''))

    # Devolver el siguiente número
    return ultimo_numero_folio + 1


@perfil_seleccionado_required
def requisicion_detalle(request, pk):
    pk_perfil = request.session.get('selected_profile_id') 
    usuario = Profile.objects.get(id = pk_perfil)
    #Vista de creación de requisición
    productos = ArticulosparaSurtir.objects.filter(articulos__orden__id = pk, requisitar= True)
    orden = Order.objects.get(id = pk)
    pk = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk)
    requisiciones = Requis.objects.all()
    requi, created = requisiciones.get_or_create(complete=False, orden=orden)
    requis = requisiciones.filter(orden__distrito = orden.distrito, complete = True)
    # Extraemos solo los números de los folios del distrito
    last_requi = requis.order_by('-folio').first()
    if last_requi:
        folio = last_requi.folio + 1
    else:
        folio = 1
    
    abrev = usuario.distritos.abreviado
    folio_preview = str(folio)
    #for producto in productos:
    productos_requisitados = ArticulosRequisitados.objects.filter(req = requi)
    error_messages = {}
    form = RequisForm()
    


    if request.method == 'POST':
        form = RequisForm(request.POST, instance = requi)
        if form.is_valid():
            requi = form.save(commit=False)
            requi.created_by = usuario
            requi.complete = True
            orden.requisitado = True
            conteo_pendientes_requisitar = productos.filter(requisitar = True).count()
            if conteo_pendientes_requisitar > 0: #cuento cuantos productos están pendientes por requisitar 
                orden.requisitado = False
            else:
                orden.requisitado = True
            for producto in productos:
                producto.seleccionado = False
                producto.save()
                #if producto.requisitar == False:
                #    orden.requisitado = False
                #    orden.save()
            if productos_requisitados:
                #folio_consecutivo = obtener_consecutivo(usuario.distrito, requisiciones)
                requi.folio = folio
                requi.save()
                form.save()
                orden.save()
                messages.success(request,f'Has realizado la requisición {requi.folio} con éxito')
                return redirect('solicitud-autorizada-orden')
            else:
                messages.error(request,'No se puede crear la requisición debido a que no hay productos agregados')
        else:
            for field, errors in form.errors.items():
                error_messages[field] = errors.as_text()
            

    #print(orden)
    context = {
        'error_messages':error_messages,
        'productos': productos,
        'productos_requisitados':productos_requisitados,
        'orden': orden,
        'folio':folio_preview,
        'requi':requi,
        'form':form,
        }

    return render(request,'requisiciones/detalle_requisitar_editar.html', context)

# Convertir la imagen a base64
def get_image_base64(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode()


@perfil_seleccionado_required
def requisicion_autorizar(request, pk):
    
    pk_perfil = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk_perfil)
    
    requi = Requis.objects.get(id = pk)
    productos = ArticulosRequisitados.objects.filter(req = pk)
    costo_aprox = 0
    for producto in productos:
        #producto = producto.producto.articulos.producto.price
        costo_aprox = costo_aprox + producto.cantidad * producto.producto.articulos.producto.price

    try:
        presupuesto = requi.orden.subproyecto.presupuesto or 0  # Default to 0 if None
        gastado = requi.orden.subproyecto.gastado or 0  # Default to 0 if None
        porcentaje = "{0:.2f}%".format((gastado/presupuesto)*100)
    except ZeroDivisionError:
        porcentaje = " 0%"
    resta = presupuesto - gastado - costo_aprox

    if request.method == 'POST':
        requi.requi_autorizada_por = perfil
        requi.approved_at_time = datetime.now().time()
        requi.approved_at = date.today()
        requi.autorizar = True
        requi.save()
        static_path = settings.STATIC_ROOT
        img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
        img_path2 = os.path.join(static_path,'images','logo_vordcab.jpg')
       
        
        
        image_base64 = get_image_base64(img_path)
        logo_v_base64 = get_image_base64(img_path2)
        # Crear el mensaje HTML
        html_message = f"""
        <html>
            <head>
                <meta charset="UTF-8">
            </head>
            <body style="font-family: Arial, sans-serif; color: #333; background-color: #f4f4f4; margin: 0; padding: 0;">
                <table width="100%" cellspacing="0" cellpadding="0" style="background-color: #f4f4f4; padding: 20px;">
                    <tr>
                        <td align="center">
                            <table width="600px" cellspacing="0" cellpadding="0" style="background-color: #ffffff; padding: 20px; border-radius: 10px;">
                                <tr>
                                    <td align="center">
                                        <img src="data:image/jpeg;base64,{logo_v_base64}" alt="Logo" style="width: 100px; height: auto;" />
                                    </td>
                                </tr>
                                <tr>
                                    <td style="padding: 20px;">
                                        <p style="font-size: 18px; text-align: justify;">
                                            <p>Estimado {requi.orden.staff.staff.staff.first_name} {requi.orden.staff.staff.staff.last_name},</p>
                                        </p>
                                        <p style="font-size: 16px; text-align: justify;">
                                            Estás recibiendo este correo porque tu sol: {requi.orden.folio}| Req: {requi.folio} ha sido autorizada, por {requi.requi_autorizada_por.staff.staff.first_name} {requi.requi_autorizada_por.staff.staff.last_name}.</p>
                                    <p style="font-size: 16px; text-align: justify;">
                                        <p>El siguiente paso del sistema: Generación de OC</p>
                                    </p>
                                        <p style="text-align: center; margin: 20px 0;">
                                            <img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width: 50px; height: auto; border-radius: 50%;" />
                                        </p>
                                        <p style="font-size: 14px; color: #999; text-align: justify;">
                                            Este mensaje ha sido automáticamente generado por SAVIA 2.0
                                        </p>
                                    </td>
                                </tr>
                            </table>
                        </td>
                    </tr>
                </table>
            </body>
        </html>
        """
        #Crear y enviar el correo
        try:
            email = EmailMessage(
                f'Requisición Autorizada {requi.folio}',
                body=html_message,
                from_email = settings.DEFAULT_FROM_EMAIL,
                to= ['ulises_huesc@hotmail.com',requi.orden.staff.staff.staff.email],
                headers={'Content-Type': 'text/html'}
                )
            email.content_subtype = "html " # Importante para que se interprete como HTML
            email.send()
            messages.success(request,f'Has autorizado la requisición {requi.folio} con éxito')
        except (BadHeaderError, SMTPException) as e:
            error_message = f'Has autorizado la requisición {requi.folio} con éxito pero el correo de notificación no ha sido enviado debido a un error: {e}'
            messages.success(request, error_message)
        return redirect('requisicion-autorizacion')

    context = {
        'productos': productos,
        'requis': requi,
        'costo_aprox': costo_aprox,
        'porcentaje': porcentaje,
        'resta': resta,
     }

    return render(request,'requisiciones/requisiciones_autorizar.html', context)

@perfil_seleccionado_required
def requisicion_cancelar_compras(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    colaborador = Profile.objects.all()
    perfil = colaborador.get(id = pk_perfil)
    
    #perfil = Profile.objects.get(staff=usuario)
    requis = Requis.objects.get(id = pk)
    productos = ArticulosRequisitados.objects.filter(req = pk)

    if request.method == 'POST':
        form= Rechazo_Requi_Form(request.POST,instance=requis)
        if form.is_valid():
            requis.autorizada_por = perfil
            requis.autorizar = False
            requis.save()
            comentario_rechazo = requis.comentario_compras if requis.comentario_compras else requis.comentario_rechazo 
            try:
                email = EmailMessage(
                    f'Requisición Rechazada {requis.folio}',
                    f'Estimado {requis.orden.staff.staff.staff.first_name} {requis.orden.staff.staff.staff.last_name},\n Estás recibiendo este correo porque tu solicitud: {requis.orden.folio}| Req: {requis.folio} ha sido rechazada,\n por {requis.autorizada_por.staff.staff.first_name} {requis.autorizada_por.staff.staff.last_name} por el siguiente motivo: \n " {comentario_rechazo} ".\n\n Este mensaje ha sido automáticamente generado por SAVIA 2.0',
                    settings.DEFAULT_FROM_EMAIL,
                    ['ulises_huesc@hotmail.com',requis.orden.staff.staff.staff.email],
                    )
                email.send()
                messages.success(request,f'Has cancelado la requisición {requis.folio}')
            except (BadHeaderError, SMTPException) as e:
                error_message = f'Has cancelado la requisición {requis.folio} con éxito, pero el correo de notificación no ha sido enviado debido a un error: {e}'
                messages.success(request, error_message)
            return redirect('requisicion-autorizada')
    else:
        form = Rechazo_Requi_Form(instance=requis)


    context = {
        'productos': productos,
        'requis': requis,
        'form':form,
     }
    return render(request,'requisiciones/requisiciones_cancelar_compra.html', context)

@perfil_seleccionado_required
def requisicion_cancelar(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    colaborador = Profile.objects.all()
    perfil = colaborador.get(id = pk_perfil)
    
    #perfil = Profile.objects.get(staff=usuario)
    requis = Requis.objects.get(id = pk)
    productos = ArticulosRequisitados.objects.filter(req = pk)

    if request.method == 'POST':
        form= Rechazo_Requi_Form(request.POST,instance=requis)
        if form.is_valid():
            requis.autorizada_por = perfil
            requis.autorizar = False
            requis.save()
            try:
                email = EmailMessage(
                    f'Requisición Rechazada {requis.folio}',
                    f'Estimado {requis.orden.staff.staff.staff.first_name} {requis.orden.staff.staff.staff.last_name},\n Estás recibiendo este correo porque tu solicitud: {requis.orden.folio}| Req: {requis.folio} ha sido rechazada,\n por {requis.autorizada_por.staff.staff.first_name} {requis.autorizada_por.staff.staff.last_name} por el siguiente motivo: \n " {requis.comentario_compras} ".\n\n Este mensaje ha sido automáticamente generado por SAVIA 2.0',
                    settings.DEFAULT_FROM_EMAIL,
                    ['ulises_huesc@hotmail.com',requis.orden.staff.staff.staff.email],
                    )
                email.send()
                messages.success(request,f'Has cancelado la requisición {requis.folio}')
            except (BadHeaderError, SMTPException) as e:
                error_message = f'Has cancelado la requisición {requis.folio} con éxito, pero el correo de notificación no ha sido enviado debido a un error: {e}'
                messages.success(request, error_message)
            return redirect('requisicion-autorizacion')
    else:
        form = Rechazo_Requi_Form(instance=requis)


    context = {
        'productos': productos,
        'requis': requis,
        'form':form,
     }
    return render(request,'requisiciones/requisiciones_cancelar.html', context)

def render_pdf_view(request, pk):
    #Configuration of the PDF object
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    #Here ends conf.
    orden = Order.objects.get(id=pk)
    productos = ArticulosOrdenados.objects.filter(orden=pk)
    #salidas = Salidas.objects.filter(producto__articulos__orden__id=pk)


   #Azul Vordcab
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)
    caja_iso = 760
    #Elaborar caja
    #c.line(caja_iso,500,caja_iso,720)



    #Encabezado
    c.drawString(420,caja_iso,'Preparado por:')
    c.drawString(420,caja_iso-10,'SUP. ADMON')
    c.drawString(520,caja_iso,'Aprobación')
    c.drawString(520,caja_iso-10,'SUB ADM')
    c.drawString(150,caja_iso-20,'Número de documento')
    c.drawString(160,caja_iso-30,'F-ADQ-N4-01.02')
    c.drawString(245,caja_iso-20,'Clasificación del documento')
    c.drawString(275,caja_iso-30,'Controlado')
    c.drawString(355,caja_iso-20,'Nivel del documento')
    c.drawString(380,caja_iso-30, 'N5')
    c.drawString(440,caja_iso-20,'Revisión No.')
    c.drawString(452,caja_iso-30,'000')
    c.drawString(510,caja_iso-20,'Fecha de Emisión')
    c.drawString(525,caja_iso-30,'1-Sep.-18')

    caja_proveedor = caja_iso - 65
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(150,750,250,20, fill=True, stroke=False) #Barra azul superior Solicitud
    c.rect(20,caja_proveedor - 8,565,20, fill=True, stroke=False) #Barra azul superior Proveedor | Detalle
    c.rect(20,575,565,2, fill=True, stroke=False) #Linea posterior horizontal
    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(280,755,'Solicitud')
    c.setLineWidth(.3) #Grosor
    c.line(20,caja_proveedor-8,20,575) #Eje Y donde empieza, Eje X donde empieza, donde termina eje y,donde termina eje x (LINEA 1 contorno)
    c.line(585,caja_proveedor-8,585,575) #Linea 2 contorno
    c.drawInlineImage('static/images/logo_vordcab.jpg',45,730, 3 * cm, 1.5 * cm) #Imagen vortec

    c.setFillColor(white)
    c.setFont('Helvetica-Bold',11)
    #c.drawString(120,caja_proveedor,'Infor')
    c.drawString(300,caja_proveedor, 'Detalles')
    inicio_central = 300
    #c.line(inicio_central,caja_proveedor-25,inicio_central,520) #Linea Central de caja Proveedor | Detalle
    c.setFillColor(black)
    c.setFont('Helvetica',9)
    c.drawString(30,caja_proveedor-20,'Solicitó:')
    c.drawString(30,caja_proveedor-40,'Distrito:')
    c.drawString(30,caja_proveedor-60,'Proyecto')
    c.drawString(30,caja_proveedor-80,'Subproyecto:')
    c.drawString(30,caja_proveedor-100,'Fecha de Aprobación:')
    
    c.setFont('Helvetica-Bold',12)
    c.drawString(500,caja_proveedor-20,'FOLIO:')
    c.setFillColor(rojo)
    c.setFont('Helvetica-Bold',12)
    c.drawString(540,caja_proveedor-20, str(orden.folio))

    c.setFillColor(black)
    c.setFont('Helvetica',9)
    c.drawString(130,caja_proveedor-20, orden.staff.staff.staff.first_name+' '+ orden.staff.staff.staff.last_name)
    c.drawString(130,caja_proveedor-40, orden.staff.distritos.nombre)
    c.drawString(130,caja_proveedor-60, orden.proyecto.nombre)
    c.drawString(130,caja_proveedor-80, orden.subproyecto.nombre)
    if orden.approved_at:
        c.drawString(130,caja_proveedor-100, orden.approved_at.strftime("%d/%m/%Y"))
    else:
        c.setFillColor(rojo)
        c.drawString(130,caja_proveedor-100, "No Aprobado aún")
    #Create blank list
    data =[]
    styles = getSampleStyleSheet()
    paragraph_style = styles["BodyText"]
    compact_style = ParagraphStyle(
        name="CompactStyle",
        fontName="Helvetica",
        fontSize=6,  # Tamaño de fuente más pequeño
        leading=7,  # Espaciado entre líneas
        textColor=colors.black,
        alignment=0,  # Alineación a la izquierda (puedes cambiar a 1=centrado o 2=derecha si es necesario)
        spaceBefore=0,  # Sin espacio antes del párrafo
        spaceAfter=0,   # Sin espacio después del párrafo
    )
    encabezado = [['''Código''', '''Nombre''', '''Cantidad''','''Comentario''']]


    high = 480
    for i in range(1):
        for producto in productos:
            if producto.comentario:
                comentario = Paragraph(producto.comentario, compact_style)
            else:
                comentario = ''
            data.append([producto.producto.producto.codigo, producto.producto.producto.nombre,producto.cantidad,comentario])
            high = high - 15
    if high <= 480-(15*15):
        high= 480-(15*15)

    c.setFillColor(prussian_blue)
    c.rect(20,30,565,30, fill=True, stroke=False)
    c.setFillColor(white)
    #Primer renglón
    c.drawCentredString(70,48,'Clasificación:')
    c.drawCentredString(140,48,'Nivel:')
    c.drawCentredString(240,48,'Preparado por:')
    c.drawCentredString(350,48,'Aprobado:')
    c.drawCentredString(450,48,'Fecha emisión:')
    c.drawCentredString(550,48,'Rev:')
    #Segundo renglón
    c.drawCentredString(70,34,'Controlado')
    c.drawCentredString(140,34,'N5')
    c.drawCentredString(240,34,'SEOV-ALM-N4-01-01')
    c.drawCentredString(350,34,'SUB ADM')
    c.drawCentredString(450,34,'24/Oct/2018')
    c.drawCentredString(550,34,'001')

    c.setFillColor(black)
    width, height = letter
    styles = getSampleStyleSheet()
    styleN = styles["BodyText"]

    if orden.comentario is not None:
        comentario = orden.comentario
    else:
        comentario = "No hay comentarios"

    c.setFillColor(prussian_blue)
    c.rect(20,230,565,25, fill=True, stroke=False)
    c.setFillColor(white)
    c.drawCentredString(320,235,'Observaciones')
    options_conditions_paragraph = Paragraph(comentario, styleN)
    # Crear un marco (frame) en la posición específica
    frame = Frame(20, 30, 570, 200, id='normal')
    # Agregar el párrafo al marco
    frame.addFromList([options_conditions_paragraph], c)
    c.setFillColor(prussian_blue)
    c.rect(20,30,565,30, fill=True, stroke=False)

    c.setFillColor(black)
    c.drawCentredString(180,140, orden.staff.staff.staff.first_name +' '+ orden.staff.staff.staff.last_name)
    c.line(140,139,220,139)
    c.drawCentredString(180,130, 'Solicitado')
    if orden.autorizar == False:
        c.setFillColor(rojo)
        c.drawCentredString(410,140, '{Esta orden ha sido Cancelada}')
        c.setFont('Helvetica-Bold',14)
        c.drawString(370,670, 'CANCELADA')
    elif orden.autorizar:
        c.drawCentredString(410,140, orden.supervisor.staff.staff.first_name+' '+ orden.supervisor.staff.staff.last_name)
        c.setFillColor(prussian_blue)
        c.setFont('Helvetica-Bold',14)
        c.drawString(370,670, 'APROBADA')
    else:
        c.drawCentredString(410,140, orden.supervisor.staff.staff.first_name+' '+ orden.supervisor.staff.staff.last_name)
        c.setFillColor(rojo)
        c.setFont('Helvetica-Bold',14)
        c.drawString(370,670, 'NO AUTORIZADA AÚN')
    c.setFillColor(black)
    c.setFont('Helvetica',9)
    c.line(360,139,460,139)
    c.drawCentredString(410,130,'Supervisor')

    #table = Table(data, colWidths=[1.2 * cm, 12 * cm, 1.5 * cm, 5.2 * cm,])
    table_style = TableStyle([ #estilos de la tabla
        ('INNERGRID',(0,0),(-1,-1), 0.25, colors.white),
        ('BOX',(0,0),(-1,-1), 0.25, colors.black),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        #ENCABEZADO
        ('TEXTCOLOR',(0,0),(-1,0), white),
        ('FONTSIZE',(0,0),(-1,0), 8),
        ('BACKGROUND',(0,0),(-1,0), prussian_blue),
        #CUERPO
        ('TEXTCOLOR',(0,1),(-1,-1), colors.black),
        ('FONTSIZE',(0,1),(-1,-1), 6),
        ])
    #table.setStyle(table_style)
    # Configuración inicial
    rows_first_page = 10  # Filas para la primera página
    rows_other_pages = 17  # Filas para las demás páginas
    data_len = len(data) 

    page_start = 0
    first_page = True  # Bandera para determinar si es la primera página

    while page_start < data_len:
        # Determinar el número de filas en esta página
        rows_per_page = rows_first_page if first_page else rows_other_pages
        
        page_end = min(page_start + rows_per_page, data_len)  # Fin de la página actual
        page_data = encabezado + data[page_start:page_end]  # Datos para esta página
        table = Table(page_data, colWidths=[1.2 * cm, 10 * cm, 1.5 * cm, 7.2 * cm])
        table.setStyle(table_style)
        
        # Ajustar la posición según si es la primera página o no
        if first_page:
            table_y_position = high  # Posición más alta para la primera página
        else:
            table_y_position = height - 520  # Posición estándar para las demás páginas
        
        table.wrapOn(c, width, height)  # Preparar la tabla
        table.drawOn(c, 20, table_y_position)  # Dibujar la tabla en la posición calculada
        
        # Actualizar el inicio de la siguiente página
        page_start = page_end
        first_page = False  # Después de la primera iteración, cambia la bandera
        
        if page_start < data_len:  # Si hay más datos, agregar una nueva página
            c.showPage()
    
   
    
    #pdf size
    #table.wrapOn(c, width, height)
    #table.drawOn(c, 20, high)

    #c.showPage()
    c.save()
    buf.seek(0)

    return FileResponse(buf, as_attachment=True, filename='reporte_' + str(orden.folio) +'.pdf')

@perfil_seleccionado_required
def reporte_entradas(request):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    entradas = EntradaArticulo.objects.filter(entrada__completo = True, articulo_comprado__producto__producto__articulos__producto__producto__servicio = False, entrada__oc__req__orden__distrito = usuario.distritos ).order_by('-entrada__entrada_date')
    myfilter = EntradasFilter(request.GET, queryset=entradas)
    entradas = myfilter.qs
   
    entradas_data = list(entradas.values())

    #Set up pagination
    p = Paginator(entradas, 50)
    page = request.GET.get('page')
    entradas_list = p.get_page(page)

    if request.method == "POST" and 'btnExcel' in request.POST:
        #print(entradas)
        return convert_entradas_to_xls2(entradas)

    context = {
        'entradas_list':entradas_list,
        'entradas':entradas,
        'myfilter':myfilter,
        }
    #task_id_entradas =   request.session.get('task_id_entradas')

    #if request.method == "POST" and 'btnExcel' in request.POST:
        #if not task_id_entradas:
            #task =  convert_entradas_to_xls_task.delay(entradas_data)
            #task_id = task.id
            #request.session['task_id_entradas'] = task_id
            #context['task_id_entradas'] = task_id 

    return render(request,'requisiciones/reporte_entradas.html', context)

@perfil_seleccionado_required
def reporte_devoluciones(request):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    if usuario.tipo.almacen == True:
        entradas = Devolucion.objects.filter(solicitud__distrito = usuario.distritos).order_by('-fecha').select_related('solicitud')
    else:
        entradas = Devolucion.objects.none()
    myfilter = DevolucionFilter(request.GET, queryset=entradas)
    entradas = myfilter.qs

    if request.method == "POST" and 'btnExcel' in request.POST:
        #print(entradas)
        return convert_devoluciones_to_xls2(entradas)
    
    #Set up pagination
    p = Paginator(entradas, 50)
    page = request.GET.get('page')
    ordenes_list = p.get_page(page)

    context = {
        'ordenes_list':ordenes_list,
        'entradas':entradas,
        'myfilter':myfilter,
        }
    
    return render(request,'requisiciones/reporte_devoluciones.html', context)

@perfil_seleccionado_required
def reporte_entradas_servicios(request):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    if usuario.tipo.nombre == "Admin":
        entradas = EntradaArticulo.objects.filter(entrada__completo = True, articulo_comprado__producto__producto__articulos__producto__producto__servicio = True,).order_by('-entrada__entrada_date')
    elif usuario.tipo.nombre == "ALMACEN" or usuario.tipo.nombre == "SUPERIN_ADM" or usuario.tipo.comprador == True:
        entradas = EntradaArticulo.objects.filter(entrada__completo = True,articulo_comprado__producto__producto__articulos__producto__producto__servicio = True, entrada__oc__req__orden__distrito = usuario.distritos).order_by('-entrada__entrada_date')
    else:
        entradas = EntradaArticulo.objects.filter(entrada__completo = True, entrada__almacenista__distritos = usuario.distritos,articulo_comprado__producto__producto__articulos__producto__producto__servicio = True, entrada__oc__req__orden__distrito = usuario.distritos, entrada__oc__req__orden__staff = usuario).order_by('-entrada__entrada_date')

    
    myfilter = EntradasFilter(request.GET, queryset=entradas)
    entradas = myfilter.qs

    #entradas_data = list(entradas.values())

    #Set up pagination
    p = Paginator(entradas, 50)
    page = request.GET.get('page')
    entradas_list = p.get_page(page)

    if request.method == "POST" and 'btnExcel' in request.POST:
        return convert_entradas_to_xls2(entradas)

    context = {
        'entradas_list':entradas_list,
        'entradas':entradas,
        'myfilter':myfilter,
        }

    return render(request,'requisiciones/reporte_entradas.html', context)

@perfil_seleccionado_required
def reporte_salidas(request):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    salidas = Salidas.objects.filter(vale_salida__solicitud__distrito = usuario.distritos, vale_salida__complete = True).order_by('-vale_salida__folio')
    myfilter = SalidasFilter(request.GET, queryset=salidas)
    salidas = myfilter.qs
    salidas_filtradas = salidas.filter(producto__articulos__producto__producto__servicio = False)
    salidas_data =  list(salidas_filtradas.values())
    #Set up pagination
    p = Paginator(salidas, 50)
    page = request.GET.get('page')
    salidas_list = p.get_page(page)

    if request.method == "POST" and 'btnExcel' in request.POST:
        return generate_excel_report2(salidas_filtradas)

    context = {
        'salidas':salidas,
        'salidas_list':salidas_list,
        'myfilter':myfilter,
        }

    #task_id_salidas = request.session.get('task_id_salidas')
    
    #if request.method == "POST" and 'btnExcel' in request.POST:
    #    if not task_id_salidas:
    #        task =  convert_salidas_to_xls_task.delay(salidas_data)
    #        task_id = task.id
    #        request.session['task_id_salidas'] = task_id
    #        context['task_id_salidas'] = task_id 

    return render(request,'requisiciones/reporte_salidas.html', context)

def verificar_estado_salidas(request):
    task_id = request.session.get('task_id_salidas')  # Asumiendo que el task_id se pasa como parámetro GET

    if not task_id:
        return JsonResponse({'error': 'No se proporcionó task_id'}, status=400)


    task_result = AsyncResult(task_id)

    if task_result.ready():
        if task_result.successful():
            result = task_result.result
            response_data = {'task_id': task_id, 'status': 'SUCCESS', 'result': result}
        elif task_result.failed():
            response_data = {'task_id': task_id, 'status': 'FAILURE', 'result': str(task_result.result)}
        else:
            response_data = {'task_id': task_id, 'status': task_result.status}
    else:
        response_data = {'task_id': task_id, 'status': 'PENDING'}

    return JsonResponse(response_data)

def clear_task_id_salidas(request):
    if 'task_id_salidas' in request.session:
        del request.session['task_id_salidas']
    return JsonResponse({'status': 'success'})

def verificar_estado_entradas(request):
    task_id = request.session.get('task_id_entradas')  # Asumiendo que el task_id se pasa como parámetro GET

    if not task_id:
        return JsonResponse({'error': 'No se proporcionó task_id'}, status=400)


    task_result = AsyncResult(task_id)

    if task_result.ready():
        if task_result.successful():
            result = task_result.result
            response_data = {'task_id': task_id, 'status': 'SUCCESS', 'result': result}
        elif task_result.failed():
            response_data = {'task_id': task_id, 'status': 'FAILURE', 'result': str(task_result.result)}
        else:
            response_data = {'task_id': task_id, 'status': task_result.status}
    else:
        response_data = {'task_id': task_id, 'status': 'PENDING'}

    return JsonResponse(response_data)

def clear_task_id_entradas(request):
    if 'task_id_entradas' in request.session:
        del request.session['task_id_entradas']
    return JsonResponse({'status': 'success'})

@perfil_seleccionado_required
def historico_articulos_para_surtir(request):
    registros = ArticulosparaSurtir.history.all()

    myfilter = Historical_articulos_surtir_filter(request.GET, queryset=registros)
    registros = myfilter.qs

    #Set up pagination
    p = Paginator(registros, 30)
    page = request.GET.get('page')
    registros_list = p.get_page(page)

    context = {
        'myfilter': myfilter,
        'registros_list':registros_list,
        }

    return render(request,'requisiciones/historicos_articulos_para_surtir.html',context)

@perfil_seleccionado_required
def historico_salidas(request):
    registros = Salidas.history.all()

    myfilter = HistoricalSalidasFilter(request.GET, queryset=registros)
    registros = myfilter.qs

    #Set up pagination
    p = Paginator(registros, 30)
    page = request.GET.get('page')
    registros_list = p.get_page(page)

    context = {
        'myfilter':myfilter,
        'registros_list':registros_list,
        }

    return render(request,'requisiciones/historico_salidas.html',context)


def convert_solicitud_autorizada_to_xls(productos):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Solicitudes_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Solicitudes')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    number_style = NamedStyle(name='number_style', number_format='#,##0.00')
    number_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(number_style)

    columns = ['Folio','Solicitante','Proyecto','Subproyecto','Código','Artículo','Creado','Cantidad']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16

    columna_max = len(columns)+2

    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por Savia V2. UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style

    rows = productos.values_list(
        'articulos__orden__folio',
        Concat('articulos__orden__staff__staff__first_name',Value(' '),'articulos__orden__staff__staff__last_name'),
        'articulos__orden__proyecto__nombre',
        'articulos__orden__subproyecto__nombre',
        'articulos__producto__producto__codigo',
        'articulos__producto__producto__nombre',
        'articulos__orden__approved_at',
        'cantidad')

    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            if col_num == 6:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style
            if col_num == 7 or col_num == 4:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = number_style
            else:
                (ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style
    
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)
    #Aquí termina la implementación del XLSX

def convert_solicitud_autorizada_orden_to_xls(ordenes):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Solicitudes_pend_requisicion' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Solicitudes')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)

    columns = ['Folio','Solicitante','Proyecto','Subproyecto','Creado']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16

    columna_max = len(columns)+2

    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por Savia V2. UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style

    rows = ordenes.values_list('id',Concat('staff__staff__first_name',Value(' '),'staff__staff__last_name'),
                            'proyecto__nombre','subproyecto__nombre','created_at')

    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            (ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style
            if col_num == 4:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style

    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)
#Aquí termina la implementación del XLSXid="btnBuscar"


def render_salida_pdf(request, pk):
    #Configuration of the PDF object
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=portrait(letter))
    #Here ends conf.
    articulo = Salidas.objects.get(id=pk)
    vale = ValeSalidas.objects.get(id = articulo.vale_salida.id)
    productos = Salidas.objects.filter(vale_salida = vale)
    styles = getSampleStyleSheet()
    styles['BodyText'].fontSize = 6

    #Azul Vordcab
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)
    caja_iso = 770
    #Elaborar caja
    #c.line(caja_iso,500,caja_iso,720)


    c.drawString(420,caja_iso,'Preparado por:')
    c.drawString(420,caja_iso-10,'SUP. ADMON')
    c.drawString(520,caja_iso,'Aprobación')
    c.drawString(520,caja_iso-10,'SUB ADM')
    c.drawString(150,caja_iso-25,'Número de documento')
    c.drawString(160,caja_iso-35,'SEOV-ALM-N4-01-03')
    c.drawString(245,caja_iso-25,'Clasificación del documento')
    c.drawString(275,caja_iso-35,'Controlado')
    c.drawString(355,caja_iso-25,'Nivel del documento')
    c.drawString(380,caja_iso-35, 'N5')
    c.drawString(440,caja_iso-25,'Revisión No.')
    c.drawString(452,caja_iso-35,'001')
    c.drawString(510,caja_iso-25,'Fecha de Emisión')
    c.drawString(525,caja_iso-35,'24-Oct.-18')


    c.drawString(510,caja_iso-50,'Folio: ')
    #c.drawString(530,caja_iso-50, str(vale.folio))
    c.drawString(510,caja_iso-60,'Fecha:')
    c.drawString(540,caja_iso-60,vale.created_at.strftime("%d/%m/%Y"))

    c.setFillColor(rojo)
    c.setFont('Helvetica-Bold',12)
    c.drawString(530,caja_iso-50, str(vale.folio))
    

    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(150,caja_iso-15,250,20, fill=True, stroke=False) #Barra azul superior Orden de Compra

    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(280,caja_iso-10,'Vale de Salida Almacén')
    c.setLineWidth(.3) #Grosor

    c.drawInlineImage('static/images/logo_vordcab.jpg',45,caja_iso-40, 3 * cm, 1.5 * cm) #Imagen vortec
   

    data =[]
    productos_data = []
    high = 670
    data.append(['''Código''','''Producto''', '''Cantidad''', '''Unidad''']) #,'''P.Unitario''', '''Importe'''

    for producto in productos:
        producto_nombre = Paragraph(producto.producto.articulos.producto.producto.nombre, styles["BodyText"])
        data.append([producto.producto.articulos.producto.producto.codigo, producto_nombre, producto.cantidad, producto.producto.articulos.producto.producto.unidad])
        high = high - 18
        #Lo vuelvo a captura de otra manera para el código QR
        nombre_producto = producto.producto.articulos.producto.producto.nombre
        codigo_producto = producto.producto.articulos.producto.producto.codigo
        producto_info = {
            'codigo': codigo_producto,
            'nombre': nombre_producto,
            'cantidad': str(producto.cantidad),
            'unidad': str(producto.producto.articulos.producto.producto.unidad),
            #'precio_unitario': str(producto.precio),
            #'importe': str(producto.precio * producto.cantidad)
        }
        productos_data.append(producto_info)
    
    # Variables de paginación
    width, height = letter
    high = 680  # Posición inicial en la primera página
    rows_per_page_first = 21
    rows_per_page_subsequent = 24
    row_height = 18  # Altura por fila en puntos

    remaining_data = data[1:]  # Excluir encabezado para procesamiento

    # Dibujar primera página
    c.setFillColor(black)
    c.setFont('Helvetica', 8)
    current_page_data = [data[0]] + remaining_data[:rows_per_page_first]
    remaining_data = remaining_data[rows_per_page_first:]
    table = Table(current_page_data, colWidths=[2.0 * cm, 12 * cm, 3.0 * cm, 3.0 * cm])
    table.setStyle(TableStyle([
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.white),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TEXTCOLOR', (0, 0), (-1, 0), white),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 0), (-1, 0), prussian_blue),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 6)
    ]))
    table.wrapOn(c, width, height)
    table.drawOn(c, 20, high - row_height * len(current_page_data))  # Ajuste de posición de acuerdo a la altura de filas

    c.setFillColor(black)
    c.setFont('Helvetica',8)
    proyecto_y = 285 if high > 500 else high - 30
    proyecto_y -= 60
    # Generar el código QR
    qr = qrcode.QRCode(
        error_correction=qrcode.constants.ERROR_CORRECT_Q,  # Aumenta el nivel de corrección
        box_size=10,
        border=6,  # Aumentar el borde
    )

    folio = str(vale.folio)
    fecha = vale.created_at.strftime("%d/%m/%Y")

    # Limitar productos_data a 5 elementos
    if len(productos_data) > 5:
        productos_data = productos_data[:5] + [{"nombre": "etc..."}]  # Tomar solo los primeros 5 y añadir "etc."

    # Crear una lista de nombres de productos
    productos_string = ', '.join(producto['nombre'] for producto in productos_data)  # Asegúrate de que 'nombre' es la clave correcta

    # Crear el string QR
    qr_info = f'Folio: {folio}, Fecha: {fecha}, Productos: {productos_string}'

    # Añadir datos al QR
    try:
        qr.add_data(qr_info)
        qr.make(fit=True)  # Intentar ajustar automáticamente
    except ValueError as e:
        print(f"Error: {e}")

    # Generar la imagen del QR
    qr_image = qr.make_image(fill_color="black", back_color="white")

    # Guardar la imagen del QR en un archivo temporal
    qr_image_path = 'temp_qr.png'
    qr_image.save(qr_image_path)
    c.drawImage(qr_image_path, 440, proyecto_y-50, 100, 100)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(20,proyecto_y - 5 ,350,20, fill=True, stroke=False) #3ra linea azul
    c.setFillColor(black)
    c.setFont('Helvetica',7)


    c.setFillColor(white)
    c.setLineWidth(.1)
    c.setFont('Helvetica-Bold',10)
    c.drawCentredString(120,proyecto_y,'Proyecto')
    c.drawCentredString(300,proyecto_y,'Subproyecto')
   

    c.setFont('Helvetica',8)
    c.setFillColor(black)
    c.drawCentredString(120,proyecto_y - 15, str(vale.solicitud.proyecto.nombre))
    c.drawCentredString(300,proyecto_y - 15, str(vale.solicitud.subproyecto.nombre))
    c.drawString(20,proyecto_y-65,'Comentarios:')
    if vale.comentario:
        c.drawString(80,proyecto_y - 65, str(vale.comentario))
    else:
        c.drawString(80,proyecto_y - 65, 'Sin comentarios')

    c.setFillColor(black)
    c.setFont('Helvetica',8)
    #c.line(135,high-200,215, high-200) #Linea de Autorizacion
    c.drawCentredString(150,proyecto_y - 180,'Entregó')
    c.drawCentredString(150,proyecto_y - 190, vale.almacenista.staff.staff.first_name +' '+vale.almacenista.staff.staff.last_name)

    #c.line(370,proyecto_y - 20,430, proyecto_y - 20)
    c.drawCentredString(450,proyecto_y - 180,'Recibió')
    c.drawCentredString(450,proyecto_y - 190, vale.material_recibido_por.staff.staff.first_name +' '+vale.material_recibido_por.staff.staff.last_name)


    #c.line(240, high-200, 310, high-200)
    c.drawCentredString(280,proyecto_y - 180,'Autorizó')
    c.drawCentredString(280,proyecto_y - 190, vale.solicitud.staff.staff.staff.first_name + ' ' + vale.solicitud.staff.staff.staff.last_name)

    c.setFont('Helvetica',10)
    c.setFillColor(prussian_blue)
    c.setFont('Helvetica', 9)
    c.setFillColor(black)

    c.setFillColor(prussian_blue)
    c.rect(20,proyecto_y - 215,565,20, fill=True, stroke=False)
    c.setFillColor(white)
    if remaining_data:
        c.showPage()

    # Dibujar páginas subsiguientes
    while remaining_data:
        c.setFillColor(black)
        c.setFont('Helvetica', 8)
        current_page_data = [data[0]] + remaining_data[:rows_per_page_subsequent]
        remaining_data = remaining_data[rows_per_page_subsequent:]

        # Altura acumulativa para cada página
        high_subsequent = 720  # Ajusta esta altura si deseas comenzar desde más abajo en cada página
        num_rows_current_page = len(current_page_data) - 1  # Excluye el encabezado
        total_table_height = row_height * num_rows_current_page

        table = Table(current_page_data, colWidths=[2.0 * cm, 12 * cm, 3.0 * cm, 3.0 * cm])
        table.setStyle(TableStyle([
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.white),
            ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TEXTCOLOR', (0, 0), (-1, 0), white),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 0), (-1, 0), prussian_blue),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 6)
        ]))
        table.wrapOn(c, width, height)
        table.drawOn(c, 20, high_subsequent - total_table_height)  # Ajusta la posición en cada página
        
        if remaining_data:
            c.showPage()

    c.save()
    c.showPage()
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename='vale_salida_'+str(vale.folio) +'.pdf')

def render_entrada_pdf(request, pk):
    #Configuration of the PDF object
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=portrait(letter))
    #Here ends conf.
    articulo = EntradaArticulo.objects.get(id=pk)
    vale = Entrada.objects.get(id = articulo.entrada.id)
    productos = EntradaArticulo.objects.filter(entrada= vale)
    styles = getSampleStyleSheet()
    styles['BodyText'].fontSize = 6

    #Azul Vordcab
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)
    caja_iso = 770
    #Elaborar caja
    #c.line(caja_iso,500,caja_iso,720)


    c.drawString(420,caja_iso,'Preparado por:')
    c.drawString(420,caja_iso-10,'SUP. ADMON')
    c.drawString(520,caja_iso,'Aprobación')
    c.drawString(520,caja_iso-10,'SUB ADM')
    #Segundo renglón
    c.drawString(150,caja_iso-25,'Número de documento')
    c.drawString(160,caja_iso-35,'SEOV-ALM-N4-01-02')
    c.drawString(245,caja_iso-25,'Clasificación del documento')
    c.drawString(275,caja_iso-35,'Controlado')
    c.drawString(355,caja_iso-25,'Nivel del documento')
    c.drawString(380,caja_iso-35, 'N5')
    c.drawString(440,caja_iso-25,'Revisión No.')
    c.drawString(452,caja_iso-35,'001')
    c.drawString(510,caja_iso-25,'Fecha de Emisión')
    c.drawString(525,caja_iso-35,'24-Oct.-18')


    c.drawString(510,caja_iso-50,'Folio: ')
    #c.drawString(530,caja_iso-50, str(vale.folio))
    c.drawString(510,caja_iso-60,'Fecha:')
    c.drawString(540,caja_iso-60,vale.entrada_date.strftime("%d/%m/%Y"))

    c.setFillColor(rojo)
    c.setFont('Helvetica-Bold',12)
    c.drawString(540,caja_iso-50, str(vale.folio))
    

    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(150,caja_iso-15,250,20, fill=True, stroke=False) #Barra azul superior Orden de Compra

    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(280,caja_iso-10,'Vale de Entrada Almacén')
    c.setLineWidth(.3) #Grosor

    c.drawInlineImage('static/images/logo_vordcab.jpg',45,caja_iso-40, 3 * cm, 1.5 * cm) #Imagen vortec
   

    data =[]
    productos_data = []
    high = 670
    data.append(['''Código''','''Producto''', '''Cantidad''', '''Unidad'''])
    for producto in productos:
        producto_nombre = Paragraph(producto.articulo_comprado.producto.producto.articulos.producto.producto.nombre, styles["BodyText"])
        data.append([producto.articulo_comprado.producto.producto.articulos.producto.producto.codigo, producto_nombre, producto.cantidad, producto.articulo_comprado.producto.producto.articulos.producto.producto.unidad])
        high = high - 18
        #Lo vuelvo a captura de otra manera para el código QR
        nombre_producto = producto.articulo_comprado.producto.producto.articulos.producto.producto.nombre
        codigo_producto = producto.articulo_comprado.producto.producto.articulos.producto.producto.codigo
        producto_info = {
            'codigo': codigo_producto,
            'nombre': nombre_producto,
            'cantidad': str(producto.cantidad),
            'unidad': str(producto.articulo_comprado.producto.producto.articulos.producto.producto.unidad),
        }
        productos_data.append(producto_info)
    
    
    # Generar el código QR
    #qr = qrcode.QRCode(
    #    version=1,
    #    error_correction=qrcode.constants.ERROR_CORRECT_L,
    #    box_size=10,
    #    border=4,
    #)
    #folio = str(vale.folio)
    #fecha = vale.created_at.strftime("%d/%m/%Y")
    #qr_info = {
    #    'folio': folio,
    #    'fecha': fecha,
    #    'productos': productos_data
    #}
    #qr_data = json.dumps(qr_info)
    #qr.add_data(qr_data)
    #qr.make(fit=True)

    # Generar la imagen del QR y guardarla
    #qr_image = qr.make_image(fill_color="black", back_color="white")
    #qr_image_path = '/tmp/temp_qr.png'
    #qr_image.save(qr_image_path)
    #c.drawInlineImage(qr_image_path, 500, 440, 100, 100)  # Reemplaza x, y, width, height con tus valores


    c.setFillColor(black)
    c.setFont('Helvetica',8)
    proyecto_y = 485 if high > 500 else high - 30

    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(20,proyecto_y - 5 ,350,20, fill=True, stroke=False) #3ra linea azul
    c.setFillColor(black)
    c.setFont('Helvetica',7)


    c.setFillColor(white)
    c.setLineWidth(.1)
    c.setFont('Helvetica-Bold',10)
    c.drawCentredString(120,proyecto_y,'Proyecto')
    c.drawCentredString(300,proyecto_y,'Subproyecto')

    c.setFont('Helvetica',8)
    c.setFillColor(black)
    c.drawCentredString(120,proyecto_y - 15, str(vale.oc.req.orden.proyecto))
    c.drawCentredString(300,proyecto_y - 15, str(vale.oc.req.orden.subproyecto))


    c.setFillColor(black)
    c.setFont('Helvetica',8)
    #c.line(135,high-200,215, high-200) #Linea de Autorizacion
    c.drawCentredString(200,proyecto_y - 40,'Recibió')
    if vale.almacenista:
        c.drawCentredString(200,proyecto_y - 50, vale.almacenista.staff.staff.first_name +' '+vale.almacenista.staff.staff.last_name)

    #c.line(370,proyecto_y - 20,430, proyecto_y - 20)
    #c.drawCentredString(400,proyecto_y - 30,'Recibió')
    #c.drawCentredString(400,proyecto_y - 40, vale.material_recibido_por.staff.staff.first_name +' '+vale.material_recibido_por.staff.staff.last_name)


    #c.line(240, high-200, 310, high-200)
    c.drawCentredString(425,proyecto_y - 40,'Proveedor')
    c.drawCentredString(425,proyecto_y - 50, vale.oc.proveedor.nombre.razon_social)

    c.setFont('Helvetica',10)
    c.setFillColor(prussian_blue)
    c.setFont('Helvetica', 9)
    c.setFillColor(black)

    c.setFillColor(prussian_blue)
    c.rect(20,proyecto_y - 75,565,20, fill=True, stroke=False)
    c.setFillColor(white)

    width, height = letter
    table = Table(data, colWidths=[2 * cm, 12.5 * cm, 2.5 * cm, 2.5 * cm,])
    table.setStyle(TableStyle([ #estilos de la tabla
        ('INNERGRID',(0,0),(-1,-1), 0.25, colors.white),
        ('BOX',(0,0),(-1,-1), 0.25, colors.black),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        #ENCABEZADO
        ('TEXTCOLOR',(0,0),(-1,0), white),
        ('FONTSIZE',(0,0),(-1,0), 10),
        ('BACKGROUND',(0,0),(-1,0), prussian_blue),
        #CUERPO
        ('TEXTCOLOR',(0,1),(-1,-1), colors.black),
        ('FONTSIZE',(0,1),(-1,-1), 6),
        ]))
    table.wrapOn(c, width, height)
    table.drawOn(c, 20, high)
    c.save()
    c.showPage()
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename='vale_entrada_'+str(vale.folio) +'.pdf')

def convert_excel_matriz_requis(requis):
      #print('si entra a la función')
    # Crea un objeto BytesIO para guardar el archivo Excel
    output = BytesIO()

    # Crea un libro de trabajo y añade una hoja
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet("Matriz_Requisiciones")

     
    date_format = workbook.add_format({'num_format': 'dd/mm/yyyy'})
    # Define los estilos
    head_style = workbook.add_format({'bold': True, 'font_color': 'FFFFFF', 'bg_color': '333366', 'font_name': 'Arial', 'font_size': 11})
    body_style = workbook.add_format({'font_name': 'Calibri', 'font_size': 10})
    money_style = workbook.add_format({'num_format': '$ #,##0.00', 'font_name': 'Calibri', 'font_size': 10})
    date_style = workbook.add_format({'num_format': 'dd/mm/yyyy', 'font_name': 'Calibri', 'font_size': 10})
    percent_style = workbook.add_format({'num_format': '0.00%', 'font_name': 'Calibri', 'font_size': 10})
    messages_style = workbook.add_format({'font_name':'Arial Narrow', 'font_size':11})

    columns = ['Requisición', 'Solicitud', 'Distrito', 'Proyecto', 'Subproyecto', 'Area', 'Solicitante', 'Creado','Autorización',   'Status']

    columna_max = len(columns)+2

    worksheet.write(0, columna_max - 1, 'Reporte Creado Automáticamente por SAVIA 2.0 Vordcab. UH', messages_style)
    worksheet.write(1, columna_max - 1, 'Software desarrollado por Grupo Vordcab S.A. de C.V.', messages_style)
    worksheet.set_column(columna_max - 1, columna_max, 30)  # Ajusta el ancho de las columnas nuevas
    
    # Escribir encabezados debajo de los mensajes
    #worksheet.write(2, columna_max - 1, "Fecha Inicial", head_style)
    #worksheet.write(3, columna_max - 1, "Fecha Final", head_style)
    #worksheet.write(4, columna_max - 1, "Total de OC's", head_style)
    #worksheet.write(5, columna_max - 1, "Requisiciones Aprobadas", head_style)
    #worksheet.write(6, columna_max - 1, "Requisiciones Atendidas", head_style)
    #worksheet.write(7, columna_max - 1, "KPI Colocadas/Aprobadas", head_style)
    #worksheet.write(8, columna_max - 1, "OC Entregadas", head_style)
    #worksheet.write(9, columna_max - 1, "OC Autorizadas", head_style)
    #worksheet.write(10, columna_max - 1, "KPI OC Entregadas/Total de OC", head_style)
    
    #indicador = num_requis_atendidas/num_approved_requis
    #letra_columna = xl_col_to_name(columna_max)
    #formula = f"={letra_columna}9/{letra_columna}10"
    # Escribir datos y fórmulas
    #worksheet.write(2, columna_max, start_date, date_style)  # Ejemplo de escritura de fecha
    #worksheet.write(3, columna_max, end_date, date_style)
    #worksheet.write_formula(4, columna_max, '=COUNTA(A:A)-1', body_style)  # Ejemplo de fórmula
    #worksheet.write(5, columna_max, num_approved_requis, body_style)
    #worksheet.write(6, columna_max, num_requis_atendidas, body_style)
    #worksheet.write(7, columna_max, indicador, percent_style)  # Ajuste del índice de fila y columna para xlsxwriter
    #worksheet.write_formula(8, columna_max, '=COUNTIF(S:S, "Entregada")', body_style)
    # Escribir otra fórmula COUNTIF, también con el estilo corporal
    #worksheet.write_formula(9, columna_max, '=COUNTIF(O:O, "Autorizado")', body_style)
    #worksheet.write_formula(10, columna_max, formula, percent_style)

    for i, column in enumerate(columns):
        worksheet.write(0, i, column, head_style)
        worksheet.set_column(i, i, 15)  # Ajusta el ancho de las columnas

    #worksheet.set_column('L:L', 12,  money_style)
    #worksheet.set_column('M:M', 12, money_style) 
    # Asumiendo que ya tienes tus datos de compras
    row_num = 0
    for req in requis:
        row_num += 1

        if req.colocada:
            status = 'Colocada'
        elif req.autorizar:
            status= 'Autorizada'
        elif req.autorizar == False: 
            status= 'Cancelada'
        else:
            status= 'No Autorizado Aún'
        # Aquí asumimos que ya hiciste el procesamiento necesario de cada compra
        #pagos = Pago.objects.filter(oc=compra_list)
        #tipo_de_cambio_promedio_pagos = pagos.aggregate(Avg('tipo_de_cambio'))['tipo_de_cambio__avg']

        # Usar el tipo de cambio de los pagos, si existe. De lo contrario, usar el tipo de cambio de la compra
        #tipo = tipo_de_cambio_promedio_pagos or compra_list.tipo_de_cambio
        #tipo_de_cambio = '' if tipo == 0 else tipo
        #created_at = compra_list.created_at.replace(tzinfo=None)
        #approved_at = compra_list.req.approved_at

        row = [
            req.folio,
            req.orden.folio,
            req.orden.distrito.nombre,
            req.orden.proyecto.nombre if req.orden.proyecto else '',
            req.orden.subproyecto.nombre if req.orden.subproyecto else '',
            req.orden.operacion.nombre if req.orden.operacion else '',
            f"{req.orden.staff.staff.staff.first_name} {req.orden.staff.staff.staff.last_name}",
            req.created_at if req.created_at else ' ',  # Convertir a 'naive'
            req.approved_at if req.approved_at else ' ',  # Convertir a 'naive'
            status,
        ]
        
        for col_num, cell_value in enumerate(row):
        # Define el formato por defecto
            cell_format = body_style

            # Aplica el formato de fecha para las columnas con fechas
            if col_num in [6, 7]:  # Índices 6 y 7 corresponden a las columnas 7 y 8
                cell_format = date_style
                if isinstance(cell_value, datetime):  # Verificar que el valor sea un objeto datetime
                    if cell_value.tzinfo is not None:  # Verificar si el objeto datetime tiene información de zona horaria
                        cell_value = cell_value.astimezone()  # Convertir a zona horaria local o UTC
                        cell_value = cell_value.replace(tzinfo=None)  # Eliminar la información de zona horaria
            # Aplica el formato de dinero para las columnas con valores monetarios
            #elif col_num in [11, 12]:  # Asume que estas son tus columnas de dinero
                #cell_format = money_style

            # Finalmente, escribe la celda con el valor y el formato correspondiente
            worksheet.write(row_num, col_num, cell_value, cell_format)

      
        #worksheet.write_formula(row_num, 19, f'=IF(ISBLANK(R{row_num+1}), L{row_num+1}, L{row_num+1}*R{row_num+1})', money_style)
    
   
    workbook.close()

    # Construye la respuesta
    output.seek(0)

    response = HttpResponse(
        output.read(), 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    response['Content-Disposition'] = f'attachment; filename=Matriz_requisiciones_{dt.date.today()}.xlsx'
      # Establecer una cookie para indicar que la descarga ha iniciado
    response.set_cookie('descarga_iniciada', 'true', max_age=20)  # La cookie expira en 20 segundos
    output.close()
    return response


def convert_excel_matriz_requis_productos(requis):
    # Crea un objeto BytesIO para guardar el archivo Excel
    output = BytesIO()

    # Crea un libro de trabajo y añade una hoja
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet("Matriz_Requisiciones_Productos")

     
    date_format = workbook.add_format({'num_format': 'dd/mm/yyyy'})
    # Define los estilos
    head_style = workbook.add_format({'bold': True, 'font_color': 'FFFFFF', 'bg_color': '333366', 'font_name': 'Arial', 'font_size': 11})
    body_style = workbook.add_format({'font_name': 'Calibri', 'font_size': 10})
    money_style = workbook.add_format({'num_format': '$ #,##0.00', 'font_name': 'Calibri', 'font_size': 10})
    date_style = workbook.add_format({'num_format': 'dd/mm/yyyy', 'font_name': 'Calibri', 'font_size': 10})
    percent_style = workbook.add_format({'num_format': '0.00%', 'font_name': 'Calibri', 'font_size': 10})
    messages_style = workbook.add_format({'font_name':'Arial Narrow', 'font_size':11})

    columns = ['Requisición', 'Solicitud', 'Solicitante', 'Proyecto', 'Subproyecto','Código', 'Producto','Unidad', 'Cantidad','Autorización','Status']

    columna_max = len(columns)+2

    worksheet.write(0, columna_max - 1, 'Reporte Creado Automáticamente por SAVIA 2.0 Vordcab. UH', messages_style)
    worksheet.write(1, columna_max - 1, 'Software desarrollado por Grupo Vordcab S.A. de C.V.', messages_style)
    worksheet.set_column(columna_max - 1, columna_max, 30)  # Ajusta el ancho de las columnas nuevas

    for i, column in enumerate(columns):
        worksheet.write(0, i, column, head_style)
        worksheet.set_column(i, i, 15)  # Ajusta el ancho de las columnas

    row_num = 0
    for req in requis:
        row_num += 1

        if req.req.colocada:
            status = 'Colocada'
        elif req.req.autorizar:
            status= 'Autorizada'
        elif req.req.autorizar == False: 
            status= 'Cancelada'
        else:
            status= 'No Autorizado Aún'

        row = [
            req.req.folio,
            req.req.orden.folio,
            f"{req.req.orden.staff.staff.staff.first_name} {req.req.orden.staff.staff.staff.last_name}",
            req.req.orden.proyecto.nombre if req.req.orden.proyecto else '',
            req.req.orden.subproyecto.nombre if req.req.orden.subproyecto else '',
            req.producto.articulos.producto.producto.codigo if req.producto.articulos.producto else '',
            str(req.producto.articulos.producto.producto.nombre) if req.producto.articulos.producto else '',
            str(req.producto.articulos.producto.producto.unidad) if req.producto.articulos.producto else '',

            req.cantidad,
            (str(req.req.approved_at) + str(req.req.approved_at_time)) if req.req.autorizar else '',
            status,
        ]
        
        for col_num, cell_value in enumerate(row):
        # Define el formato por defecto
            cell_format = body_style

            # Finalmente, escribe la celda con el valor y el formato correspondiente
            worksheet.write(row_num, col_num, cell_value, cell_format)

      
        #worksheet.write_formula(row_num, 19, f'=IF(ISBLANK(R{row_num+1}), L{row_num+1}, L{row_num+1}*R{row_num+1})', money_style)
    
   
    workbook.close()

    # Construye la respuesta
    output.seek(0)

    response = HttpResponse(
        output.read(), 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    response['Content-Disposition'] = f'attachment; filename=Matriz_requisiciones_{dt.date.today()}.xlsx'
      # Establecer una cookie para indicar que la descarga ha iniciado
    response.set_cookie('descarga_iniciada', 'true', max_age=20)  # La cookie expira en 20 segundos
    output.close()
    return response

def convert_entradas_to_xls2(entradas):

    output = io.BytesIO()

    columns = ['Vale', 'Folio Solicitud', 'Folio Compra', 'Folio Req', 'Fecha', 'Solicitante', 'Proveedor', 'Proyecto', 'Subproyecto', 'Área', 'Código', 'Articulo', 'Cantidad', 'Familia','Moneda', 'Tipo de Cambio', 'Precio', 'Total']
    data = [columns]

    for item in entradas:
        #pk = item['id']
        entrada = EntradaArticulo.objects.get(id = item.id)
        pagos = Pago.objects.filter(oc=entrada.entrada.oc)
        tipo_de_cambio_promedio_pagos = pagos.aggregate(Avg('tipo_de_cambio'))['tipo_de_cambio__avg']
        tipo_de_cambio = tipo_de_cambio_promedio_pagos or entrada.entrada.oc.tipo_de_cambio
        print(tipo_de_cambio)
        print(item)

        row = [
            entrada.entrada.folio,
            entrada.entrada.oc.req.orden.folio,
            entrada.entrada.oc.folio,
            entrada.entrada.oc.req.folio,
            entrada.created_at.date(),#.strftime('%Y-%m-%d'),  # Formatea la fecha para la celda
            f"{entrada.entrada.oc.req.orden.staff.staff.staff.first_name} {entrada.entrada.oc.req.orden.staff.staff.staff.last_name}",
            entrada.entrada.oc.proveedor.nombre.razon_social,
            entrada.entrada.oc.req.orden.proyecto.nombre if entrada.entrada.oc.req.orden.proyecto else "Sin Proyecto",
            entrada.entrada.oc.req.orden.subproyecto.nombre if entrada.entrada.oc.req.orden.subproyecto else "Sin Subproyecto",
            entrada.entrada.oc.req.orden.operacion.nombre if entrada.entrada.oc.req.orden.operacion else "Sin operación",
            entrada.articulo_comprado.producto.producto.articulos.producto.producto.codigo,
            entrada.articulo_comprado.producto.producto.articulos.producto.producto.nombre,
            entrada.cantidad,
            entrada.articulo_comprado.producto.producto.articulos.producto.producto.familia,
            entrada.entrada.oc.moneda.nombre,
            tipo_de_cambio,
            entrada.articulo_comprado.precio_unitario,
        ]
        if row[9] == "DOLARES":
            if row[10] is None or row[10] < 15:
                row[10] = 17  # O cualquier valor predeterminado que desees
        elif row[10] is None:
            row[10] = ""

        data.append(row)

    wb = Workbook()
    ws = wb.new_sheet("Entradas", data=data)

    # Aplicar estilos a los encabezados
    header_style = Style(
        font=Font(bold=True, color=PXColor(255, 255, 255)),
        fill=Fill(background=PXColor(51, 51, 102)),
        alignment=Alignment(horizontal='center', vertical='center')
    )

    # Aplicar estilos a las celdas de datos
    date_style = Style(
        format=Format('dd/mm/yyyy'),
        alignment=Alignment(horizontal='left')
    )
    number_style = Style(
        format=Format('#,##0.00'),
        alignment=Alignment(horizontal='right')
    )
    money_style = Style(
        format=Format('$#,##0.00'),
        alignment=Alignment(horizontal='right')
    )
    body_style = Style(
        alignment=Alignment(horizontal='left')
    )

    
    for row_num in range(2, len(data) + 1):
        ws[row_num][18].value = f'=IF(P{row_num} = 0, Q{row_num}*M{row_num}, Q{row_num}*P{row_num}*M{row_num})'

    for col_num in range(1, len(columns) + 1):
        if col_num == 5:  # Fecha
            ws.set_col_style(col_num, date_style)
        elif col_num in [15, 16, 17, 18]:  # Dinero
            ws.set_col_style(col_num, money_style)
        else:
            ws.set_col_style(col_num, body_style)


    for col_num in range(1, len(columns) + 1):
        ws[1][col_num].style = header_style


    wb.save(output)  # Guardar el libro de trabajo en el objeto BytesIO

    # Configurar la respuesta para descargar el archivo
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    file_name = 'Matriz_Entradas_' + str(date.today()) + '.xlsx'
    response['Content-Disposition'] = f'attachment; filename={file_name}'
    response.set_cookie('descarga_iniciada', 'true', max_age=20)

    # Cerrar el objeto BytesIO
    output.close()
    return response

def convert_devoluciones_to_xls2(entradas):
    # Crea un objeto BytesIO para guardar el archivo Excel
    output = BytesIO()

    # Crea un libro de trabajo y añade una hoja
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet("Matriz_Devoluciones")

     
    date_format = workbook.add_format({'num_format': 'dd/mm/yyyy'})
    # Define los estilos
    head_style = workbook.add_format({'bold': True, 'font_color': 'FFFFFF', 'bg_color': '333366', 'font_name': 'Arial', 'font_size': 11})
    body_style = workbook.add_format({'font_name': 'Calibri', 'font_size': 10})
    money_style = workbook.add_format({'num_format': '$ #,##0.00', 'font_name': 'Calibri', 'font_size': 10})
    date_style = workbook.add_format({'num_format': 'dd/mm/yyyy', 'font_name': 'Calibri', 'font_size': 10})
    percent_style = workbook.add_format({'num_format': '0.00%', 'font_name': 'Calibri', 'font_size': 10})
    messages_style = workbook.add_format({'font_name':'Arial Narrow', 'font_size':11})

    #columns = ['Folio Solicitud', 'Solicitante', 'Almacenista','Proyecto', 'Subproyecto', 'Fecha creación','Productos','Tipo','Autorizada','Fecha autorización','Comentario']
    columns = ['Folio Solicitud', 'Solicitante', 'Almacenista','Proyecto', 'Subproyecto', 'Fecha creación','Tipo','Autorizada','Fecha autorización','Comentario']

    columna_max = len(columns)+2

    worksheet.write(0, columna_max - 1, 'Reporte Creado Automáticamente por SAVIA 2.0 Vordcab. UH', messages_style)
    worksheet.write(1, columna_max - 1, 'Software desarrollado por Grupo Vordcab S.A. de C.V.', messages_style)
    worksheet.set_column(columna_max - 1, columna_max, 30)  # Ajusta el ancho de las columnas nuevas

    for i, column in enumerate(columns):
        worksheet.write(0, i, column, head_style)
        worksheet.set_column(i, i, 15)  # Ajusta el ancho de las columnas

    row_num = 0
    for dev in entradas:
        if dev.autorizada is True:
            autorizado = 'Autorizado'
        elif dev.autorizada is False:
            autorizado = 'No Autorizado'
        else:
            autorizado = 'Pendiente'
        row_num += 1
        # Crear la lista de productos con nombre y cantidad
        #productos_lista = [
        #    f"{producto['producto__producto__nombre']} (Cantidad: {producto['cantidad']})"
        #    for producto in dev.solicitud.productos.values('producto__producto__nombre', 'cantidad')
        #]
        # Unir la lista en una cadena
        #productos_str = ", ".join(productos_lista)

        row = [
            dev.solicitud.folio,
            f"{dev.solicitud.staff.staff.staff.first_name} {dev.solicitud.staff.staff.staff.last_name}",
            f"{dev.almacenista.staff.staff.first_name} {dev.almacenista.staff.staff.last_name}",
            dev.solicitud.proyecto.nombre,
            dev.solicitud.subproyecto.nombre,
            str(dev.created_at),
            #productos_str,  # Productos concatenados
            dev.tipo.nombre,
            autorizado,
            str(dev.fecha),
            dev.comentario,
        ]
        
        for col_num, cell_value in enumerate(row):
        # Define el formato por defecto
            cell_format = body_style

            # Finalmente, escribe la celda con el valor y el formato correspondiente
            worksheet.write(row_num, col_num, cell_value, cell_format)

      
        #worksheet.write_formula(row_num, 19, f'=IF(ISBLANK(R{row_num+1}), L{row_num+1}, L{row_num+1}*R{row_num+1})', money_style)
    
   
    workbook.close()

    # Construye la respuesta
    output.seek(0)

    response = HttpResponse(
        output.read(), 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    response['Content-Disposition'] = f'attachment; filename=Matriz_requisiciones_{dt.date.today()}.xlsx'
      # Establecer una cookie para indicar que la descarga ha iniciado
    response.set_cookie('descarga_iniciada', 'true', max_age=20)  # La cookie expira en 20 segundos
    output.close()
    return response


def generate_excel_report(salidas):
    #print(salidas)
    output = BytesIO()

    # Crea un libro de trabajo y añade una hoja
    workbook = xlsxwriter.Workbook(output, {'constant_memory': True})
    worksheet = workbook.add_worksheet("Matriz_Salidas")

    # Define los formatos que necesitas
    head_style = workbook.add_format({'bold': True, 'font_color': 'FFFFFF', 'bg_color': '333366', 'font_name': 'Arial', 'font_size': 11})
    body_style = workbook.add_format({'font_name': 'Calibri', 'font_size': 10})
    money_style = workbook.add_format({'num_format': '$ #,##0.00', 'font_name': 'Calibri', 'font_size': 10})
    date_style = workbook.add_format({'num_format': 'dd/mm/yyyy', 'font_name': 'Calibri', 'font_size': 10})

    # Escribe los encabezados
    columns = ['Vale Salida','Folio Solicitud','Fecha','Solicitante','Proyecto','Subproyecto','Área','Código','Articulo','Material recibido por',
               'Cantidad','Precio','Moneda','TC','Total']
    
    for i, column in enumerate(columns):
        worksheet.write(0, i, column, head_style)
        worksheet.set_column(i, i, 15)  # Ajusta el ancho de las columnas

    # Preparar los datos
    row_num = 0
    for salida in salidas:
        row_num += 1
        if salida.precio > 0:
            precio_condicional = salida.precio
        elif salida.producto.precio > 0:
            precio_condicional = salida.producto.precio
        else:
            precio_condicional = salida.producto.articulos.producto.price

        if salida.entrada:
            entrada = Entrada.objects.get(id = entrada)
            moneda = str(entrada.oc.moneda.nombre)
            tc = entrada.oc.tipo_de_cambio
        else:
            moneda = "PESOS"
            tc = " "

        rows = [
            salida.vale_salida.folio,
            salida.vale_salida.solicitud.folio,
            salida.created_at.strftime('%Y-%m-%d'),  # Formatea la fecha para la celda
            f"{salida.producto.articulos.orden.staff.staff.staff.first_name} {salida.producto.articulos.orden.staff.staff.staff.last_name}",
            salida.producto.articulos.orden.proyecto.nombre if salida.producto.articulos.orden.proyecto else " ",
            salida.producto.articulos.orden.subproyecto.nombre if salida.producto.articulos.orden.subproyecto else " ",
            salida.producto.articulos.orden.operacion.nombre if salida.producto.articulos.orden.operacion else "Sin operación",
            salida.producto.articulos.producto.producto.codigo,
            salida.producto.articulos.producto.producto.nombre,
            f"{salida.vale_salida.material_recibido_por.staff.staff.first_name} {salida.vale_salida.material_recibido_por.staff.staff.last_name}",
            salida.cantidad,
            precio_condicional,
            moneda,
            tc
        ]

         # Escribe la fila en el archivo
        for col_num, cell_value in enumerate(rows):
        # Define el formato por defecto
            cell_format = body_style

            # Aplica el formato de fecha para las columnas con fechas
            if col_num in [3]:  # Asume que estas son tus columnas de fechas
                cell_format = date_style
        
            # Aplica el formato de dinero para las columnas con valores monetarios
            elif col_num in [11]:  # Asume que estas son tus columnas de dinero
                cell_format = money_style

            # Finalmente, escribe la celda con el valor y el formato correspondiente
            worksheet.write(row_num, col_num, cell_value, cell_format)

        worksheet.write_formula(row_num, 14, f'=K{row_num + 1}*L{row_num + 1}', money_style)
    
    workbook.close()

    # Construye la respuesta
    output.seek(0)

    response = HttpResponse(
        output.read(), 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    response['Content-Disposition'] = f'attachment; filename=Matriz_Salidas_{dt.date.today()}.xlsx'
      # Establecer una cookie para indicar que la descarga ha iniciado
    response.set_cookie('descarga_iniciada', 'true', max_age=20)  # La cookie expira en 20 segundos
    output.close()
    return response


def generate_excel_report2(salidas):
    output = io.BytesIO()

    columns = ['Vale Salida', 'Folio Solicitud', 'Fecha', 'Solicitante', 'Proyecto', 'Subproyecto', 'Área', 'Código', 'Articulo', 'Material recibido por', 'Comentario','Cantidad', 'Precio', 'Moneda','TC','Total']
    data = [columns]

    for salida in salidas:
        if salida.precio > 0:
            precio_condicional = salida.precio
        elif salida.producto.precio > 0:
            precio_condicional = salida.producto.precio
        else:
            precio_condicional = salida.producto.articulos.producto.price

        if salida.vale_salida.material_recibido_por:
            recibido = f"{salida.vale_salida.material_recibido_por.staff.staff.first_name} {salida.vale_salida.material_recibido_por.staff.staff.last_name}"
        else:
            recibido = "NR"

        if salida.vale_salida.comentario:
            comentario = f"{salida.vale_salida.comentario}"
        else:
            comentario = " "

        if salida.entrada:
            id = salida.entrada
            try:
                entrada_articulo = EntradaArticulo.objects.get(id = id)
                moneda = str(entrada_articulo.entrada.oc.moneda.nombre)
                tc = entrada_articulo.entrada.oc.tipo_de_cambio
            except EntradaArticulo.DoesNotExist:
                moneda = "PESOS"
                tc = 0       
        else:
            moneda = "PESOS"
            tc = 0

        solicitante = f"{salida.vale_salida.solicitud.staff.staff.staff.first_name} {salida.vale_salida.solicitud.staff.staff.staff.last_name}"

        rows = [
            salida.vale_salida.folio,
            salida.vale_salida.solicitud.folio,
            salida.created_at.date(),#.strftime('%Y-%m-%d'),  # Formatea la fecha para la celda
            solicitante,
            salida.vale_salida.solicitud.proyecto.nombre if salida.vale_salida.solicitud.proyecto else " ",
            salida.vale_salida.solicitud.subproyecto.nombre if salida.vale_salida.solicitud.subproyecto else " ",
            salida.producto.articulos.orden.operacion.nombre if salida.producto.articulos.orden.operacion else "Sin operación",
            salida.producto.articulos.producto.producto.codigo,
            salida.producto.articulos.producto.producto.nombre,
            recibido,
            comentario,
            salida.cantidad,
            precio_condicional,
            moneda,
            tc,
            None  # Placeholder for the total formula
        ]
        data.append(rows)

    # Crear el archivo Excel usando pyexcelerate
    wb = Workbook()
    ws = wb.new_sheet("Matriz_Salidas", data=data)

    # Aplicar estilos a los encabezados
    header_style = Style(
        font=Font(bold=True, color=PXColor(255, 255, 255)),
        fill=Fill(background=PXColor(51, 51, 102)),
        alignment=Alignment(horizontal='center', vertical='center')
    )

    # Aplicar estilos a las celdas de datos
    date_style = Style(
        format=Format('dd/mm/yyyy'),
        alignment=Alignment(horizontal='left')
    )
    money_style = Style(
        format=Format('$#,##0.00'),
        alignment=Alignment(horizontal='right')
    )
    body_style = Style(
        alignment=Alignment(horizontal='left')
    )

   
    for row_num in range(2, len(data) + 1):
        formula = f'=IF(O{row_num} = 0, M{row_num}*L{row_num}, M{row_num}*L{row_num}*O{row_num})'
        ws[row_num][16].value = formula

    for col_num in range(1, len(columns) + 1):
        if col_num == 3:  # Fecha
            ws.set_col_style(col_num, date_style)
        elif col_num in [13, 15, 16]:  # Dinero
            ws.set_col_style(col_num, money_style)
        else:
            ws.set_col_style(col_num, body_style)


    for col_num in range(1, len(columns) + 1):
        ws[1][col_num].style = header_style

    wb.save(output)  # Guardar el libro de trabajo en el objeto BytesIO

    # Configurar la respuesta para descargar el archivo
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    file_name = 'Matriz_Salidas_' + str(date.today()) + '.xlsx'
    response['Content-Disposition'] = f'attachment; filename={file_name}'
    response.set_cookie('descarga_iniciada', 'true', max_age=20)

    # Cerrar el objeto BytesIO
    output.close()
    return response

def render_requisicion_pdf_view(request, pk):
    #Configuration of the PDF object
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    #Here ends conf.
    requisicion = Requis.objects.get(id=pk)
    productos = ArticulosRequisitados.objects.filter(req=pk)
    #salidas = Salidas.objects.filter(producto__articulos__orden__id=pk)


   #Azul Vordcab
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)
    caja_iso = 760
    #Elaborar caja
    #c.line(caja_iso,500,caja_iso,720)



    #Encabezado
    c.drawString(420,caja_iso,'Preparado por:')
    c.drawString(420,caja_iso-10,'SUP. ADMON')
    c.drawString(520,caja_iso,'Aprobación')
    c.drawString(520,caja_iso-10,'SUB ADM')
    c.drawString(150,caja_iso-20,'Número de documento')
    c.drawString(160,caja_iso-30,'SEOV-ADQ-N4-01.01')
    c.drawString(245,caja_iso-20,'Clasificación del documento')
    c.drawString(275,caja_iso-30,'Controlado')
    c.drawString(355,caja_iso-20,'Nivel del documento')
    c.drawString(380,caja_iso-30, 'N5')
    c.drawString(440,caja_iso-20,'Revisión No.')
    c.drawString(452,caja_iso-30,'000')
    c.drawString(510,caja_iso-20,'Fecha de Emisión')
    c.drawString(525,caja_iso-30,'22-Nov.-17')

    caja_proveedor = caja_iso - 65
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(150,750,250,20, fill=True, stroke=False) #Barra azul superior Solicitud
    c.rect(20,caja_proveedor - 8,565,20, fill=True, stroke=False) #Barra azul superior Proveedor | Detalle
    c.rect(20,575,565,2, fill=True, stroke=False) #Linea posterior horizontal
    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(280,755,'Requisición')
    c.setLineWidth(.3) #Grosor
    c.line(20,caja_proveedor-8,20,575) #Eje Y donde empieza, Eje X donde empieza, donde termina eje y,donde termina eje x (LINEA 1 contorno)
    c.line(585,caja_proveedor-8,585,575) #Linea 2 contorno
    c.drawInlineImage('static/images/logo_vordcab.jpg',45,730, 3 * cm, 1.5 * cm) #Imagen vortec

    c.setFillColor(white)
    c.setFont('Helvetica-Bold',11)
    #c.drawString(120,caja_proveedor,'Infor')
    c.drawString(300,caja_proveedor, 'Detalles')
    inicio_central = 300
    #c.line(inicio_central,caja_proveedor-25,inicio_central,520) #Linea Central de caja Proveedor | Detalle
    c.setFillColor(black)
    c.setFont('Helvetica',9)
    c.drawString(30,caja_proveedor-20,'Solicitó:')
    c.drawString(30,caja_proveedor-40,'Distrito:')
    c.drawString(30,caja_proveedor-60,'Proyecto')
    c.drawString(30,caja_proveedor-80,'Subproyecto:')
    c.drawString(30,caja_proveedor-100,'Fecha de Aprobación:')
    
    c.setFont('Helvetica-Bold',12)
    c.drawString(500,caja_proveedor-20,'FOLIO:')
    c.setFillColor(rojo)
    c.setFont('Helvetica-Bold',12)
    c.drawString(540,caja_proveedor-20, str(requisicion.folio))

    c.setFillColor(black)
    c.setFont('Helvetica',9)
    almacenista = Profile.objects.filter(tipo__almacenista = True).first()
    if requisicion.orden.staff:
        c.drawString(130,caja_proveedor-20, requisicion.orden.staff.staff.staff.first_name+' '+ requisicion.orden.staff.staff.staff.last_name)
    else:    
        c.drawString(130,caja_proveedor-20, almacenista.staff.staff.last_name+' '+ almacenista.staff.staff.last_name)
    c.drawString(130,caja_proveedor-40, requisicion.orden.staff.distritos.nombre)
    c.drawString(130,caja_proveedor-60, requisicion.orden.proyecto.nombre)
    c.drawString(130,caja_proveedor-80, requisicion.orden.subproyecto.nombre)
    if requisicion.approved_at:
        c.drawString(130,caja_proveedor-100, requisicion.approved_at.strftime("%d/%m/%Y"))
    else:
        c.setFillColor(rojo)
        c.drawString(130,caja_proveedor-100, "No Aprobado aún")
    #Create blank list
    data =[]

    encabezado = [['''Código''', '''Nombre''', '''Cantidad''','''Comentario''']]


    high = 540
    for producto in productos:
        data.append([producto.producto.articulos.producto.producto.codigo, producto.producto.articulos.producto.producto.nombre,producto.cantidad, producto.producto.articulos.comentario])
        high = high - 18


    c.setFillColor(prussian_blue)
    c.rect(20,30,565,30, fill=True, stroke=False)
    c.setFillColor(white)
    #Primer renglón
    c.drawCentredString(70,48,'Clasificación:')
    c.drawCentredString(140,48,'Nivel:')
    c.drawCentredString(240,48,'Preparado por:')
    c.drawCentredString(350,48,'Aprobado:')
    c.drawCentredString(450,48,'Fecha emisión:')
    c.drawCentredString(550,48,'Rev:')
    #Segundo renglón
    c.drawCentredString(70,34,'Controlado')
    c.drawCentredString(140,34,'N5')
    c.drawCentredString(240,34,'SEOV-ALM-N4-01-01')
    c.drawCentredString(350,34,'SUB ADM')
    c.drawCentredString(450,34,'24/Oct/2018')
    c.drawCentredString(550,34,'001')

    c.setFillColor(black)
    width, height = letter
    styles = getSampleStyleSheet()
    styleN = styles["BodyText"]

    if requisicion.comentario_compras is not None:
        comentario = requisicion.comentario_compras
    else:
        comentario = "No hay comentarios"

    c.setFillColor(prussian_blue)
    c.rect(20,230,565,25, fill=True, stroke=False)
    c.setFillColor(white)
    c.drawCentredString(320,235,'Observaciones')
    options_conditions_paragraph = Paragraph(comentario, styleN)
    # Crear un marco (frame) en la posición específica
    frame = Frame(20, -110, width-40, high-50, id='normal')
    # Agregar el párrafo al marco
    frame.addFromList([options_conditions_paragraph], c)
    c.setFillColor(prussian_blue)
    c.rect(20,30,565,30, fill=True, stroke=False)
    c.setFillColor(white)

    c.setFillColor(black)
    if requisicion.orden:
        c.drawCentredString(180,high-240, requisicion.orden.staff.staff.staff.first_name +' '+ requisicion.orden.staff.staff.staff.last_name)
    else:    
         c.drawCentredString(180,high-240, almacenista.staff.staff.last_name+' '+ almacenista.staff.staff.last_name)
    #c.drawCentredString(180,high-240, requisicion.created_by.staff.staff.first_name +' '+ requisicion.created_by.staff.staff.last_name)
    c.line(140,high-241,220,high-241)
    c.drawCentredString(180,high-250, 'Solicitado')
    if requisicion.autorizar == False:
        c.setFillColor(rojo)
        c.drawCentredString(410, high-240, '{Esta requisicion ha sido Cancelada}')
        c.setFont('Helvetica-Bold',14)
        c.drawString(370,670, 'CANCELADA')
    elif requisicion.autorizar:
        c.setFillColor(prussian_blue)
        c.drawCentredString(410,high-240, requisicion.orden.superintendente.staff.staff.first_name+' '+ requisicion.orden.superintendente.staff.staff.last_name)
        c.setFont('Helvetica-Bold',14)
        c.drawString(370,670, 'APROBADA')
    else:
        c.setFillColor(rojo)
        c.drawCentredString(410,high-240, requisicion.orden.superintendente.staff.staff.first_name+' '+ requisicion.orden.superintendente.staff.staff.last_name)
        c.setFont('Helvetica-Bold',14)
        c.drawString(370,670, 'NO AUTORIZADA AÚN')
    c.setFillColor(black)
    c.setFont('Helvetica',12)
    c.line(360,high-241,460,high-241)
    c.drawCentredString(410,high-250,'Superintendente')

    #table = Table(data, colWidths=[1.2 * cm, 12 * cm, 1.5 * cm, 5.2 * cm,])
    table_style = TableStyle([ #estilos de la tabla
        ('INNERGRID',(0,0),(-1,-1), 0.25, colors.white),
        ('BOX',(0,0),(-1,-1), 0.25, colors.black),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        #ENCABEZADO
        ('TEXTCOLOR',(0,0),(-1,0), white),
        ('FONTSIZE',(0,0),(-1,0), 8),
        ('BACKGROUND',(0,0),(-1,0), prussian_blue),
        #CUERPO
        ('TEXTCOLOR',(0,1),(-1,-1), colors.black),
        ('FONTSIZE',(0,1),(-1,-1), 6),
        ])
    #table.setStyle(table_style)
    rows_per_page = 15
    data_len = len(data) 
    for page_start in range(0, data_len, rows_per_page):
        page_end = min(page_start + rows_per_page, data_len)
        #page_data = data[page_start:page_end + 1]  # +1 para incluir el encabezado en cada página
        page_data = encabezado + data[page_start:page_end] 
        table = Table(page_data, colWidths=[1.2 * cm, 12 * cm, 1.5 * cm, 5.2 * cm])
        table.setStyle(table_style)
         # Calcular el alto de la tabla para la página actual
        table_height = data_len * 18 #espacio_por_fila
        # Calcular la posición 'y' inicial para la tabla basada en el alto de la tabla
        table_y_position = height - table_height - 30 - (210 if page_start == 0 else 0)  # Ajustar el margen superior
        #table_y_position = height - 30 - (210 if page_start == 0 else 0)  # Ajustar el margen superior
        

        table.wrapOn(c, width, height)  # Preparar la tabla
        table.drawOn(c, 20, table_y_position)  # Dibujar la tabla en la posición calculada

        if page_end < data_len:  # Si hay más páginas, preparar una nueva página
            c.showPage()
    
   
    
    #pdf size
    #table.wrapOn(c, width, height)
    #table.drawOn(c, 20, high)

    #c.showPage()
    c.save()
    buf.seek(0)

    return FileResponse(buf, as_attachment=True, filename='Requisición_' + str(requisicion.folio) +'.pdf')
