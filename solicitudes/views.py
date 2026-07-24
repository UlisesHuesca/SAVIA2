from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.conf import settings

import os
import socket
from django.contrib import messages
from django.db.models.functions import Concat
from django.core.mail import EmailMessage, BadHeaderError
from smtplib import SMTPException
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Sum, Value, F, Sum, When, Case, DecimalField, Max, Q, ExpressionWrapper
from django.utils import timezone
from django.utils.text import slugify
from dashboard.models import Activo, Inventario, Order, ArticulosOrdenados, ArticulosparaSurtir, Inventario_Batch, Marca, Product, Tipo_Orden, Plantilla, ArticuloPlantilla, Unidad, Producto_Calidad
from requisiciones.models import Requis, ArticulosRequisitados, ValeSalidas, Salidas
from requisiciones.views import get_image_base64
from compras.models import Compra
from tesoreria.models import Pago
from solicitudes.models import Subproyecto, Operacion, Proyecto, Sector
from entradas.models import EntradaArticulo, Entrada
from gastos.models import Entrada_Gasto_Ajuste, Conceptos_Entradas
from .forms import InventarioForm, OrderForm, Inv_UpdateForm, Inv_UpdateForm_almacenista, ArticulosOrdenadosForm, Conceptos_EntradasForm, Entrada_Gasto_AjusteForm, Order_Resurtimiento_Form, ArticulosOrdenadosComentForm, Plantilla_Form, ArticuloPlantilla_Form
from dashboard.forms import Inventario_BatchForm
from user.models import Profile, Distrito, Almacen
from user.decorators import perfil_seleccionado_required
import json
from .filters import InventoryFilter, SolicitudesFilter, SolicitudesProdFilter, InventarioFilter, HistoricalInventarioFilter, HistoricalProductoFilter
import decimal


import xlsxwriter
from django.http import HttpResponse
from io import BytesIO
# Import Pagination Stuff
from datetime import date, datetime, timedelta, time
# Import Excel Stuff

import xlsxwriter
#from django.http import HttpResponse
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import datetime as dt
import csv
import ast
import pandas as pd
# Create your views here.


def updateItem(request):
    data= json.loads(request.body)
    productId = data['productId']
    action = data['action']
    tipoId = data['type']
    pk = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk)
    #usuario = Profile.objects.get(staff__id=request.user.id)
    producto = Inventario.objects.get(id=productId)

    tipo = Tipo_Orden.objects.get(id = tipoId)
    order, created = Order.objects.get_or_create(staff=usuario, complete=False, tipo = tipo, distrito = usuario.distritos)

    orderItem, created = ArticulosOrdenados.objects.get_or_create(orden = order, producto = producto)

    if action == 'add':
        orderItem.cantidad = (orderItem.cantidad + 1)
        message = f"Item was added: {orderItem}"
        orderItem.save()
    elif action == 'remove':
        orderItem.delete()
        message = f"Item was removed: {orderItem}"

    return JsonResponse(message, safe=False)

def updateItemRes(request):
    data= json.loads(request.body)
    productId = data['productId']
    action = data['action']

    pk = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk)
    producto = Inventario.objects.get(id=productId)
    tipo = Tipo_Orden.objects.get(tipo ='resurtimiento')
    order, created = Order.objects.get_or_create(staff = usuario, complete = False, tipo=tipo, distrito = usuario.distritos)
    orderItem, created = ArticulosOrdenados.objects.get_or_create(orden = order, producto= producto)

    if action == 'add':
        orderItem.cantidad = (orderItem.cantidad + 1)
        message = f"Item was added: {orderItem}"
        orderItem.save()
    elif action == 'remove':
        orderItem.delete()
        message = f"Item was removed: {orderItem}"

    return JsonResponse(message, safe=False)

#Vista de seleccion de productos, requiere login
@login_required(login_url='user-login')
@perfil_seleccionado_required
def product_selection_resurtimiento(request):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    tipo = Tipo_Orden.objects.get(tipo ='resurtimiento')
    order, created = Order.objects.get_or_create(staff = usuario, complete = False, tipo=tipo, distrito = usuario.distritos)
    productos = Inventario.objects.filter(cantidad__lt =F('minimo'), distrito = usuario.distritos)
    cartItems = order.get_cart_quantity
    myfilter=InventoryFilter(request.GET, queryset=productos)
    productos = myfilter.qs

    #Set up pagination
    p = Paginator(productos, 30)
    page = request.GET.get('page')
    productos_list = p.get_page(page)

    context= {
        'orden': order,
        'myfilter': myfilter,
        'productos_list':productos_list,
        'productosordenadosres':cartItems,
        }
    return render(request, 'solicitud/product_selection_resurtimiento.html', context)

@login_required(login_url='user-login')
@perfil_seleccionado_required
def crear_plantilla(request):
    pk_perfil = request.session.get('selected_profile_id')
    colaboradores = Profile.objects.all()
    usuario = colaboradores.get(id = pk_perfil)
    productos = Inventario.objects.filter(distrito = usuario.distritos)
    
    plantilla, created = Plantilla.objects.get_or_create(creador = usuario, complete = False)
    productos_plantilla = ArticuloPlantilla.objects.filter(plantilla = plantilla)
    form = Plantilla_Form()
    form_producto = ArticuloPlantilla_Form()
    error_messages = {}

    if request.method =='POST' and "CrearBtn" in request.POST:
        form = Plantilla_Form(request.POST, instance=plantilla)
        if form.is_valid():
            plantilla = form.save(commit=False)
            plantilla.complete = True
            plantilla.save()
            messages.success(request, 'Has creado exitósamente la plantilla')
            return redirect('matriz-plantillas')
        else:
            for field, errors in form.errors.items():
                error_messages[field] = errors.as_text()
    else:
        messages.error('No está entrando al POST')
        
           


    context = {
        'error_messages':error_messages,
        'plantilla':plantilla,
        'productos':productos,
        'form':form,
        'form_producto':form_producto,
        'productos_plantilla':productos_plantilla,
    }

    return render(request, 'solicitud/crear_plantilla.html', context)

def update_item_plantilla(request):
    data= json.loads(request.body)
    plantilla_id = int(data['plantilla_id'])
    id_producto = int(data['id_producto'])
    cantidad = decimal.Decimal(data['cantidad'])
    comentario = str(data['comentario'])
    comentario_plantilla = str(data['comentario_plantilla'])
    print(comentario)
    action = data['action']

    usuario = Profile.objects.get(staff__id=request.user.id)
    producto = Inventario.objects.get(id=id_producto)
    
    plantilla = Plantilla.objects.get(id=plantilla_id)

    item, created = ArticuloPlantilla.objects.get_or_create(plantilla = plantilla, producto= producto)

    if action == 'add':
        item.cantidad = cantidad
        item.modified_at = date.today()
        item.modified_by = usuario
        item.comentario_articulo = comentario
        item.comentario_plantilla = comentario_plantilla
        messages.success(request, f'El producto {item.producto.producto.nombre} ha sido creado')
        item.save()
    elif action == 'remove':
        messages.success(request, f'El producto {item.producto.producto.nombre} ha sido eliminado')
        item.delete()
        

    return JsonResponse('Item updated, action executed: '+data["action"], safe=False)

#Vista de seleccion de productos, requiere login
@login_required(login_url='user-login')
@perfil_seleccionado_required
def product_selection(request):
    pk = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk)
    #usuario = Profile.objects.get(staff__id=request.user.id)
    tipo = Tipo_Orden.objects.get(tipo ='normal')
    #order, created = Order.objects.get_or_create(staff = usuario, complete = False, tipo = tipo)
    order, created = Order.objects.get_or_create(staff = usuario, complete = False, tipo=tipo, distrito = usuario.distritos)
    #Traer todos los productos no criticos y solo los criticos con rev_Calidad
    #if usuario.tipo.activos == True: #Solo el personal de activos puede solicitar activos
    productos = Inventario.objects.filter(complete=True, distrito=usuario.distritos,)
    #else:
        #productos = Inventario.objects.filter(complete=True, distrito=usuario.distritos,producto__activo=False).filter(Q(producto__critico=False) | Q(producto__critico=True, producto__rev_calidad=True))
    
    cartItems = order.get_cart_quantity

    myfilter=InventoryFilter(request.GET, queryset=productos)
    productos = myfilter.qs

    #Set up pagination
    p = Paginator(productos, 30)
    page = request.GET.get('page')
    productos_list = p.get_page(page)


    context= {
        'orden':order,
        'myfilter': myfilter,
        'productos_list':productos_list,
        'productos':productos,
        'productosordenados':cartItems,
        }
    return render(request, 'solicitud/product_selection.html', context)

#Vista para crear solicitud
@login_required(login_url='user-login')
@perfil_seleccionado_required
def checkout(request):
    usuarios = Profile.objects.all()
    pk = request.session.get('selected_profile_id')
    usuario = usuarios.get(id = pk)
    error_messages = {}
    ordenes = Order.objects.all()
    #Tengo que revisar primero si ya existe una orden pendiente del usuario
    orders = ordenes.filter(staff__distritos = usuario.distritos)
    #Se genera el folio trayendo primero el ultimo folio del distrito
    last_order = ordenes.filter(distrito=usuario.distritos).order_by('-folio').first()
    if last_order.folio is not None:
        #folio = last_order.folio
        folio_number = last_order.folio + 1
    else:
        # No hay órdenes para este distrito todavía
        folio_number = 1
    abrev = usuario.distritos.abreviado
    folio_preview = folio_number
    #***************************************************
    proyectos = Proyecto.objects.filter(~Q(status_de_entrega__status = "INACTIVO"),activo=True, distrito=usuario.distritos )
    subproyectos = Subproyecto.objects.all()
    activos = Activo.objects.filter(responsable__distritos = usuario.distritos)
    tipo = Tipo_Orden.objects.get(tipo ='normal')
    sectores = Sector.objects.all()
    operaciones = Operacion.objects.exclude(nombre='GASTO')

    order, created = ordenes.get_or_create(staff = usuario, complete = False, tipo=tipo, distrito = usuario.distritos)
    if not order.inicio_form:
        order.inicio_form = timezone.now()  # Asigna la fecha y hora con zona horaria
        order.save()
    if usuario.tipo.supervisor:
        supervisores = usuarios.filter(id = pk)
        order.supervisor = usuario
    else:
        supervisores = usuarios.filter(distritos=usuario.distritos, tipo__supervisor = True, st_activo = True).exclude(tipo__nombre="Admin")

    #print(usuario.distritos)
    if usuario.distritos.nombre == "MATRIZ": 
        #print("Quev")
        superintendentes = usuarios.filter(tipo__subdirector = True, sustituto__isnull = True, st_activo =True,distritos=usuario.distritos)
    elif usuario.distritos.nombre == "BRASIL":
        superintendentes = usuarios.filter(tipo__oc_gerencia = True, sustituto__isnull = True, st_activo =True,distritos=usuario.distritos).exclude(tipo__nombre="Admin")
    elif usuario.tipo.autorizacion == True and usuario.tipo.requisiciones == True and usuario.tipo.nombre != "Admin":
        superintendentes = usuarios.filter(staff=usuario.staff, distritos=usuario.distritos)
        #order.superintendente = superintendentes
        #print("Ques")
    else:
        superintendentes = usuarios.filter(distritos=usuario.distritos, tipo__autorizacion = True, tipo__requisiciones = True, st_activo = True).exclude(tipo__nombre="Admin")
        #print("Quee")

    proyectos_para_select2 = [
        {
            'id': item.id, 
            'text': str(item.nombre)
        } for item in proyectos
    ]

    sectores_para_select2 = [
        {
            'id': item.id, 
            'text': str(item.nombre)
        } for item in sectores
    ]

    activos_para_select2 = [
        {
            'id': item.id, 
            'text': str(item.eco_unidad) + ('|') + str(item.descripcion)
        } for item in activos
    ]
    

    operaciones_para_select2 = [
        {
            'id': item.id, 
            'text': str(item.nombre)
        } for item in operaciones
    ]

    superintendentes_para_select2 = [
        {
            'id': super.id, 
            'text': str(super.staff.staff.first_name) + (' ') + str(super.staff.staff.last_name)
        } for super in superintendentes
    ]
    supervisores_para_select2 = [
        {
            'id': super.id, 
            'text': str(super.staff.staff.first_name) + (' ') + str(super.staff.staff.last_name)
        } for super in supervisores
    ]
    form = OrderForm(instance = order)
    form_comentario = ArticulosOrdenadosComentForm(prefix='form_comentario')
    
    if order.staff != usuario:
        productos = None
        cartItems = 0
    else:
        productos = order.productos.all()
        cartItems = order.get_cart_quantity

    if request.method =='POST':
        form = OrderForm(request.POST, request.FILES, instance=order)
        if form.is_valid():
            order = form.save(commit=False)
            order.created_at = timezone.now()
            #order.created_at_time = datetime.now().time() 
            lineas_productos = []
            order.folio = folio_number
            productos_html = '<ul>'
            if usuario.tipo.supervisor == True or usuario.distritos.nombre == "BRASIL": #Si el usuario es supervisor
               
                for producto in productos:
                    productos_html += f'<li>{producto.producto.producto.nombre}: {producto.cantidad}.</li>'
                    # We fetch inventory product corresponding to product (that's why we use product.id)
                    # We create a new product line in a new database to control the ArticlestoDeliver (ArticulosparaSurtir)
                    prod_inventario = Inventario.objects.get(id = producto.producto.id)
                    ordensurtir , created = ArticulosparaSurtir.objects.get_or_create(articulos = producto)
                   
                    if not ordensurtir.procesado:
                        if producto.producto.producto.servicio == True: #or producto.producto.producto.activo == True: Se comenta para evitar que se creen requisiciones de activos
                            ordensurtir.requisitar = True
                            ordensurtir.cantidad_requisitar = producto.cantidad
                            ordensurtir.procesado = True
                            print(producto.producto.producto.servicio)
                            #if producto.producto.producto.servicio == True: #or producto.producto.producto.activo == True:
                            requi, created = Requis.objects.get_or_create(complete = True, orden = order)
                            requitem, created = ArticulosRequisitados.objects.get_or_create(req = requi, producto = ordensurtir, cantidad = producto.cantidad, almacenista = usuario)
                            #requis = Requis.objects.filter(orden__distrito = usuario.distritos, complete = True)
                            #last_requi = requis.order_by('-folio').first()
                            max_folio = Requis.objects.filter(orden__distrito=usuario.distritos, complete=True).aggregate(Max('folio'))['folio__max']
                            requi.folio = max_folio + 1
                            numero_servicios = productos.filter(producto = producto.producto.producto.servicio).count()
                            if productos.count() == numero_servicios: 
                                order.requisitar=False
                                order.requisitado = True
                            ordensurtir.requisitar = False
                            requi.save()
                            requitem.save()
                            ordensurtir.save()
                            #order.fin = datetime.now()
                            order.save()
                        #cond:1 evalua si la cantidad en inventario es mayor que lo solicitado
                        elif prod_inventario.cantidad >= producto.cantidad and order.tipo.tipo == "normal":  #si la cantidad solicitada es menor que la cantidad en inventario
                            prod_inventario.cantidad = prod_inventario.cantidad - producto.cantidad
                            ordensurtir.cantidad = producto.cantidad
                            ordensurtir.precio = prod_inventario.price
                            ordensurtir.procesado = True
                            ordensurtir.surtir = True
                            ordensurtir.requisitar = False
                            prod_inventario._change_reason = f'Se modifica el inventario en view: autorizada_sol:{order.id}|{order.folio} | S{ordensurtir.cantidad} cond:1'
                            ordensurtir.save()
                            prod_inventario.cantidad_apartada = prod_inventario.apartada()
                            prod_inventario.save()
                        elif prod_inventario.cantidad < producto.cantidad and producto.cantidad > 0 and order.tipo.tipo == "normal": #si la cantidad solicitada es mayor que la cantidad en inventario
                            ordensurtir.cantidad = prod_inventario.cantidad #lo que puedes surtir es igual a lo que tienes en el inventario
                            ordensurtir.precio = prod_inventario.price
                            ordensurtir.cantidad_requisitar = producto.cantidad - ordensurtir.cantidad #lo que falta por surtir
                            #if prod_inventario.cantidad_apartada == None: #Esto es solo para evitar Nulls
                            #    prod_inventario.cantidad_apartada = 0
                            prod_inventario.cantidad = 0
                            if ordensurtir.cantidad > 0: #si lo que se puede surtir es mayor que 0
                                ordensurtir.surtir = True
                            ordensurtir.requisitar = True
                            order.requisitar = True
                            ordensurtir.procesado = True
                            prod_inventario._change_reason = f'Se modifica el inventario en view: autorizada_sol:{order.id}|{order.folio} | S{ordensurtir.cantidad} R{ordensurtir.cantidad_requisitar} cond:2'
                            ordensurtir.save()
                            prod_inventario.cantidad_apartada = prod_inventario.apartada()
                            prod_inventario.save()
                            order.save()
                        elif prod_inventario.cantidad + prod_inventario.cantidad_entradas == 0:
                            ordensurtir.requisitar = True
                            ordensurtir.cantidad_requisitar = producto.cantidad
                            ordensurtir.save()
                            order.save()
                order.autorizar = True
                order.approved_at = date.today()
                order.approved_at_time = datetime.now().time()
                static_path = settings.STATIC_ROOT
                img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
                img_path2 = os.path.join(static_path,'images','logo_vordcab.jpg')
                productos_html += '</ul>'
                image_base64 = get_image_base64(img_path)
                logo_v_base64 = get_image_base64(img_path2)
                # Crear el mensaje HTML
                html_message = f"""
                <html>
                    <head>
                        <meta charset="UTF-8">
                    </head>
                    <body style="font-family: Arial, sans-serif; color: #333; background-color: #f4f4f4; margin: 0; padding: 0;">
                        <table width="100%" cellspacing="0" cellpadding="0" style="background-color: #f4f4f4; padding: 20px;">
                            <tr>
                                <td align="center">
                                    <table width="600px" cellspacing="0" cellpadding="0" style="background-color: #ffffff; padding: 20px; border-radius: 10px;">
                                        <tr>
                                            <td align="center">
                                                <img src="data:image/jpeg;base64,{logo_v_base64}" alt="Logo" style="width: 100px; height: auto;" />
                                            </td>
                                        </tr>
                                        <tr>
                                            <td style="padding: 20px;">
                                                <p style="font-size: 18px; text-align: justify;">
                                                    <p>Estimado {order.staff.staff.staff.first_name} {order.staff.staff.staff.last_name},</p>
                                                </p>
                                                <p style="font-size: 16px; text-align: justify;">
                                                    Estás recibiendo este correo porque tu solicitud folio: <strong>{order.folio}</strong> ha sido aprobada.</p>
                                                <p>Con los productos siguientes</p>
                                                {productos_html}
                                                </p>
                                            <p style="font-size: 16px; text-align: justify;">
                                                El siguiente paso del sistema: Requisitar los productos.
                                            </p>
                                                <p style="text-align: center; margin: 20px 0;">
                                                    <img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width: 50px; height: auto; border-radius: 50%;" />
                                                </p>
                                                <p style="font-size: 14px; color: #999; text-align: justify;">
                                                    Este mensaje ha sido automáticamente generado por SAVIA 2.0
                                                </p>
                                            </td>
                                        </tr>
                                    </table>
                                </td>
                            </tr>
                        </table>
                    </body>
                </html>
                """
                try:
                    email = EmailMessage(
                        f'Solicitud Autorizada {order.folio}',
                        body=html_message,
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        to=[order.staff.staff.staff.email],
                        headers={'Content-Type': 'text/html'}
                        )
                    email.content_subtype = "html " # Importante para que se interprete como HTML
                    email.send()
                    messages.success(request, f'La solicitud {order.folio} ha sido creada')
                except (BadHeaderError, SMTPException, socket.gaierror) as e:
                    error_message = f'La solicitud {order.folio} ha sido creada, pero el correo no ha sido enviado debido a un error: {e}'
                    messages.success(request, error_message)
                order.sol_autorizada_por = Profile.objects.get(id=usuario.id)    
                cartItems = '0'
            else:
                for producto in productos:
                    productos_html += f'<li>{producto.producto.producto.nombre}: {producto.cantidad}.</li>'
                static_path = settings.STATIC_ROOT
                img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
                img_path2 = os.path.join(static_path,'images','logo_vordcab.jpg')
                productos_html += '</ul>'
                image_base64 = get_image_base64(img_path)
                logo_v_base64 = get_image_base64(img_path2)
                # Crear el mensaje HTML
                html_message = f"""
                <html>
                    <head>
                        <meta charset="UTF-8">
                    </head>
                    <body>
                        <p><img src="data:image/jpeg;base64,{logo_v_base64}" alt="Imagen" style="width:100px;height:auto;"/></p>
                        <p>Estimado {order.staff.staff.staff.first_name} {order.staff.staff.staff.last_name},</p>
                        <p>Estás recibiendo este correo porque tu solicitud folio:{order.folio}  se ha generado</p>
                        <p>Con los productos siguientes</p>
                        {productos_html}
                        <p><img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width:50px;height:auto;border-radius:50%"/></p>
                        <p>Este mensaje ha sido automáticamente generado por SAVIA 2.0</p>
                    </body>
                </html>
                """
                try:
                    email = EmailMessage(
                        f'Solicitud Autorizada {order.folio}',
                        body=html_message,
                        from_email= settings.DEFAULT_FROM_EMAIL,
                        to=[order.staff.staff.staff.email],
                        headers={'Content-Type': 'text/html'}
                        )
                    email.content_subtype = "html " # Importante para que se interprete como HTML
                    email.send()
                    messages.success(request, f'La solicitud {order.folio} ha sido creada')
                except (BadHeaderError, SMTPException, socket.gaierror) as e:
                    error_message = f'La solicitud {order.folio} ha sido creada, pero el correo no ha sido enviado debido a un error: {e}'
                    messages.success(request, error_message)
            order.complete = True
            order.save()
            #print(order.inicio_form)
            #print(order.created_at)
            #inicio_form_naive = order.inicio_form.replace(tzinfo=None)
            #created_at_naive = order.created_at.replace(tzinfo=None)
            #delta = order.created_at - order.inicio_form

            # Obtenemos los segundos totales de la diferencia
            #segundos_totales = delta.total_seconds()
            #print(segundos_totales)
            #print(inicio_form_naive)
            #print(created_at_naive)
            return redirect('solicitud-matriz')
        else:
            for field, errors in form.errors.items():
                error_messages[field] = errors.as_text()


    context= {
        'error_messages':error_messages,
        'proyectos_para_select2': proyectos_para_select2,
        'sectores_para_select2':sectores_para_select2,
        'operaciones_para_select2':operaciones_para_select2,
        'activos_para_select2':activos_para_select2,
        'superintendentes_para_select2':superintendentes_para_select2,
        'supervisores_para_select2':supervisores_para_select2,
        'form':form,
        'form_comentario': form_comentario,
        'productos':productos,
        'usuario_distrito': usuario.distritos.nombre,
        'orden':order,
        #'activos':activos,
        #'sectores': sectores,
        'productosordenados':cartItems,
        'folio_preview': folio_preview, 
        #'supervisores':supervisores,
        #'superintendentes':superintendentes,
        #'proyectos':proyectos,
        'subproyectos':subproyectos,
    }
    return render(request, 'solicitud/checkout.html', context)

@login_required(login_url='user-login')
def select_product_quality(request, pk):
    producto = Product.objects.get(id = pk)
    calidad_producto = Producto_Calidad.objects.get(producto = producto)
    requerimientos = calidad_producto.requerimientos_calidad.all()

    context = {
        'calidad_producto': calidad_producto,
        'requerimientos': requerimientos,
    }
    return render(request, 'solicitud/select_product_quality.html', context)

@login_required(login_url='user-login')
@perfil_seleccionado_required
def product_quantity_edit(request, pk):
    item = ArticulosOrdenados.objects.get(id= pk)
    form= ArticulosOrdenadosForm(instance = item)

    if request.method == 'POST':
        form = ArticulosOrdenadosForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return HttpResponse(status=204)

    context = {
        'form': form,
        'item':item,
        }

    return render(request, 'solicitud/product_quantity_edit.html', context)


def update_comentario(request):
    data= json.loads(request.body)
    pk = data["solicitud_id"]
    dato = data["dato"]
    tipo = data["tipo"]
    producto = ArticulosOrdenados.objects.get(id=pk)
    if tipo == "comentario": 
        producto.comentario = dato
    if tipo == "cantidad":
        producto.cantidad = dato
    producto.save()
    # Construye un objeto de respuesta que incluya el dato y el tipo.
    response_data = {
        'dato': dato,
        'tipo': tipo
    }

    return JsonResponse(response_data, safe=False)
    
@login_required(login_url='user-login')
@perfil_seleccionado_required
def product_comment_add(request, pk):
    item = ArticulosOrdenados.objects.get(id= pk)
    form= ArticulosOrdenadosComentForm(instance = item)
    item = get_object_or_404(ArticulosOrdenados, id=pk)

    if request.method == 'POST':
        form = ArticulosOrdenadosComentForm(request.POST, instance=item)
        form.save()
        return HttpResponse(status=204)

    context = {
        'form': form,
        'item':item,
        }

    return render(request, 'solicitud/product_comment_add.html', context)

#Vista para crear solicitud de resurtimiento
@login_required(login_url='user-login')
@perfil_seleccionado_required
def checkout_resurtimiento(request):
    pk = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk)
    #usuario = Profile.objects.get(staff=request.user)
    #Tengo que revisar primero si ya existe una orden pendiente del usuario
    if usuario.distritos.nombre == "MATRIZ":
        superintendentes = Profile.objects.filter(distritos=usuario.distritos, tipo__subdirector = True, st_activo = True).exclude(tipo__nombre="Admin")
    else:    
        superintendentes = Profile.objects.filter(distritos=usuario.distritos, tipo__autorizacion = True, tipo__requisiciones = True, st_activo = True).exclude(tipo__nombre="Admin")
    proyectos = Proyecto.objects.filter(activo=True, distrito=usuario.distritos )
    #subproyectos = Subproyecto.objects.all()
    ordenes = Order.objects.filter(distrito = usuario.distritos)
    last_order = ordenes.filter(distrito=usuario.distritos).order_by('-folio').first()
    if last_order:
        #folio = last_order.folio
        folio_number = last_order.folio + 1
    else:
        # No hay órdenes para este distrito todavía
        folio_number = 1
    #consecutivo = orders.count()+1
    error_messages = {}


    tipo = Tipo_Orden.objects.get(tipo ='resurtimiento')
    order, created = Order.objects.get_or_create(staff = usuario, complete = False, tipo=tipo, distrito = usuario.distritos)
    almacen = Operacion.objects.get(nombre = "ALMACEN")

    superintendentes_para_select2 = [
        {
            'id': super.id, 
            'text': str(super.staff.staff.first_name) + (' ') + str(super.staff.staff.last_name)
        } for super in superintendentes
    ]

    proyectos_para_select2 = [
        {
            'id': item.id, 
            'text': str(item.nombre)
        } for item in proyectos
    ]


    
    productos = order.productos.all()
    cartItems = order.get_cart_quantity

    form = OrderForm(instance = order)


    if request.method =='POST':
        form = Order_Resurtimiento_Form(request.POST, instance=order)
        if form.is_valid():
            order = form.save(commit=False)
            order.supervisor = usuario
            order.created_at = datetime.now()
            #order.created_at_time = datetime.now().time()
            order.complete = True
            order.area = almacen
            order.folio = folio_number
            requi, created = Requis.objects.get_or_create(complete = True, orden = order)
            max_folio = Requis.objects.filter(orden__distrito=usuario.distritos, complete=True).aggregate(Max('folio'))['folio__max']
            requi.folio = max_folio + 1
            requi.save()
            for producto in productos:
                ordensurtir , created = ArticulosparaSurtir.objects.get_or_create(articulos = producto)
                requitem, created = ArticulosRequisitados.objects.get_or_create(req = requi, producto= ordensurtir, cantidad = producto.cantidad)
                ordensurtir.requisitar = True
                ordensurtir.cantidad_requisitar = producto.cantidad
                ordensurtir.save()
                requitem.save()
            order.requisitado = True
            order.autorizar = True
            order.approved_at = date.today()
            #rder.approved_at_time = datetime.now().time()
            requi.save()
            order.save()
            #abrev= usuario.distrito.abreviado
            #order.folio = str(abrev) + str(order.id).zfill(4)
            messages.success(request, f'La solicitud {order.folio} junto con la requisición {requi.folio} ha sido creada')
            cartItems = '0'
            return redirect('solicitud-matriz')
        else:
            for field, errors in form.errors.items():
                error_messages[field] = errors.as_text()
            # Si quisieras hacer algo más con cada error, este es el lugar
    else:
        #messages.errors('Error')
        form = OrderForm(request.POST)

    context= {
        'superintendentes_para_select2':superintendentes_para_select2,
        'proyectos_para_select2': proyectos_para_select2,
        'error_messages':error_messages,
        #'proyectos':proyectos,
        'form':form,
        'productos':productos,
        'orden':order,
        'productosordenadosres':cartItems,
        'superintendentes':superintendentes,
        #'subproyectos':subproyectos,
    }
    return render(request, 'solicitud/checkout_resurtimiento.html', context)


#Vista para crear solicitud
@login_required(login_url='user-login')
@perfil_seleccionado_required
def checkout_editar(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    order = Order.objects.get(id=pk)

    #usuario = Profile.objects.get(id=request.user.id)

    productos = order.productos.all()
    cartItems = order.get_cart_quantity
    form = OrderForm(instance=order, distrito = usuario.distritos)


    if request.method =='POST':
        form = OrderForm(request.POST, instance=order, distrito = usuario.distritos)
        order.complete = True
        if form.is_valid():
            form.save()
            cartItems = '0'
            return redirect('solicitud-matriz')

    context= {
        'form':form,
        'productos':productos,
        'orden':order,
        'productosordenados':cartItems,
    }
    return render(request, 'solicitud/checkout.html', context)

@login_required(login_url='user-login')
@perfil_seleccionado_required
def solicitud_pendiente(request):

    #obtengo el id de usuario, lo paso como argumento a id de profiles para obtener el objeto profile que coindice con ese usuario_id
    pk = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk)

    
    #Este es un filtro por perfil supervisor o superintendente, es decir puede ver todo lo del distrito
    if perfil.tipo.nombre == "Admin":
        ordenes = Order.objects.filter(complete=True).order_by('-created_at','-folio')
    elif perfil.tipo.gerente == True or perfil.tipo.superintendente == True or perfil.tipo.nombre == "Control" or perfil.tipo.almacenista == True:
        ordenes = Order.objects.filter(complete=True, distrito=perfil.distritos).order_by('-created_at','-folio')
    elif perfil.tipo.supervisor == True:
        ordenes = Order.objects.filter(Q(supervisor=perfil)|Q(staff = perfil), complete=True, distrito=perfil.distritos).order_by('-created_at','-folio')
    else:
        ordenes = Order.objects.filter(complete=True, staff = perfil).order_by('-created_at','-folio')

    myfilter=SolicitudesFilter(request.GET, queryset=ordenes)
    ordenes = myfilter.qs

    #Set up pagination
    p = Paginator(ordenes, 10)
    page = request.GET.get('page')
    ordenes_list = p.get_page(page)

    if request.method =='POST' and 'btnExcel' in request.POST:

        return convert_excel_solicitud_matriz(ordenes)

    context= {
        'perfil':perfil,
        'ordenes_list':ordenes_list,
        'myfilter':myfilter,
        }

    return render(request, 'solicitud/solicitudes_pendientes.html',context)

@perfil_seleccionado_required
@login_required(login_url='user-login')
def solicitud_matriz(request):
    pk = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk)
    #obtengo el id de usuario, lo paso como argumento a id de profiles para obtener el objeto profile que coindice con ese usuario_id
    


     #Este es un filtro por perfil supervisor o superintendente, es decir puede ver todo lo del distrito
    if perfil.tipo.nombre == "Admin" or perfil.tipo.nombre == "Control":
        almacenes_distritos = set(perfil.almacen.values_list('distrito__id', flat=True))
        ordenes = Order.objects.filter(complete=True, distrito__in=almacenes_distritos).order_by('-created_at','-folio')
    elif perfil.tipo.superintendente == True:
        ordenes = Order.objects.filter(complete=True, distrito=perfil.distritos).order_by('-created_at','-folio')
    elif perfil.tipo.supervisor == True:
        ordenes = Order.objects.filter(Q(supervisor = perfil) | Q(staff = perfil), complete=True, distrito=perfil.distritos).order_by('-created_at','-folio')
    else:
        ordenes = Order.objects.filter(complete=True, staff = perfil).order_by('-created_at','-folio')


    myfilter=SolicitudesFilter(request.GET, queryset=ordenes)
    ordenes = myfilter.qs

    #Set up pagination
    p = Paginator(ordenes, 10)
    page = request.GET.get('page')
    ordenes_list = p.get_page(page)

    if request.method =='POST' and 'btnExcel' in request.POST:
        return convert_excel_solicitud_matriz(ordenes)
  
    context= {
        'ordenes_list':ordenes_list,
        'myfilter':myfilter,
        }

    return render(request, 'solicitud/solicitudes_pendientes.html',context)

@perfil_seleccionado_required
@login_required(login_url='user-login')
def matriz_plantillas(request):
    pk = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk)
    #obtengo el id de usuario, lo paso como argumento a id de profiles para obtener el objeto profile que coindice con ese usuario_id
    plantillas_list= Plantilla.objects.filter(complete=True)

    #myfilter=SolicitudesFilter(request.GET, queryset=ordenes)
    #ordenes = myfilter.qs

    #Set up pagination
    #p = Paginator(ordenes, 10)
    #page = request.GET.get('page')
    #ordenes_list = p.get_page(page)

    #if request.method =='POST' and 'btnExcel' in request.POST:

    #    return convert_excel_solicitud_matriz(ordenes)

    context= {
        'plantillas_list':plantillas_list,
        #'myfilter':myfilter,
        }

    return render(request, 'solicitud/matriz_plantillas.html',context)

@login_required(login_url='user-login')
@perfil_seleccionado_required
def productos_plantilla(request, pk):
    plantilla = Plantilla.objects.get(id=pk)
    productos = ArticuloPlantilla.objects.filter(plantilla=plantilla)

    context = {
        'productos':productos,
    }

    return render(request,'solicitud/productos_plantilla.html', context)

@perfil_seleccionado_required
@login_required(login_url='user-login')
def editar_plantilla(request, pk):
    plantilla = Plantilla.objects.get(id=pk)
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
   
    productos = Inventario.objects.filter(distrito = usuario.distritos)
    error_messages = {}
    productos_plantilla = ArticuloPlantilla.objects.filter(plantilla = plantilla)
    form = Plantilla_Form(instance = plantilla)
    form_producto = ArticuloPlantilla_Form()

    if request.method =='POST' and "CrearBtn" in request.POST:
        form = Plantilla_Form(request.POST, instance=plantilla)
        if form.is_valid():
            plantilla = form.save(commit=False)
            plantilla.complete = True
            plantilla.modified_at = date.today()
            plantilla.modified_by = usuario
            plantilla.save()
            messages.success(request, 'Has modificado exitósamente la plantilla')
            return redirect('matriz-plantillas')
        else:
            for field, errors in form.errors.items():
                error_messages[field] = errors.as_text()
    else:
        messages.error(request, 'No está entrando')
       
           


    context = {
        'error_messages':error_messages,
        'plantilla':plantilla,
        'productos':productos,
        'form':form,
        'form_producto':form_producto,
        'productos_plantilla':productos_plantilla,
    }
    
    return render(request, 'solicitud/editar_plantilla.html', context)

@perfil_seleccionado_required
@login_required(login_url='user-login')
def crear_solicitud_plantilla(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    tipo = Tipo_Orden.objects.get(tipo ='normal')
    order, created = Order.objects.get_or_create(staff = usuario, complete = False, tipo=tipo, distrito = usuario.distritos)
    
    # Obtiene la plantilla por su ID
    plantilla = Plantilla.objects.get(id=pk)
    
    # Añade productos de la plantilla a la orden
    for articulo in plantilla.articuloplantilla_set.all():
        # Aquí asumo que tienes un modelo que conecta un producto con una orden (quizás se llame "ArticulosOrdenados" o algo similar).
        # Si ese modelo no existe, deberás adaptar este código.
        articulo_orden, created = ArticulosOrdenados.objects.get_or_create(orden=order, producto=articulo.producto)
        articulo_orden.cantidad += articulo.cantidad  # Aumenta la cantidad basada en la plantilla
         # Copia el comentario del artículo de la plantilla al artículo ordenado.
        articulo_orden.comentario = articulo.comentario_articulo

        articulo_orden.save()

    return redirect('solicitud-checkout')  # Redirige al usuario a la selección de productos, donde ahora verá los productos de la plantilla añadidos

@perfil_seleccionado_required
@login_required(login_url='user-login')
def solicitud_matriz_productos(request):
    pk = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk)
   
     #Este es un filtro por perfil supervisor o superintendente, es decir puede ver todo lo del distrito
    if perfil.tipo.superintendente == True or perfil.tipo.nombre == "Control":
        productos = ArticulosOrdenados.objects.filter(orden__complete=True, orden__distrito=perfil.distritos).order_by('-orden__folio')
    elif perfil.tipo.supervisor == True:
        productos = ArticulosOrdenados.objects.filter(orden__complete=True, orden__distrito=perfil.distritos, orden__supervisor=perfil).order_by('-orden__folio')
    else:
        productos = ArticulosOrdenados.objects.filter(orden__complete=True, orden__staff = perfil).order_by('-orden__folio')

    myfilter=SolicitudesProdFilter(request.GET, queryset=productos)
    productos = myfilter.qs
    perfil = Profile.objects.get(id = pk)


    #Set up pagination
    p = Paginator(productos, 15)
    page = request.GET.get('page')
    productos_list = p.get_page(page)

    if request.method =='POST' and 'btnExcel' in request.POST:

        return convert_excel_solicitud_matriz_productos(productos)

    context= {
        'productos':productos_list,
        'myfilter':myfilter,
        }
    return render(request, 'solicitud/solicitudes_creadas_productos.html',context)



@login_required(login_url='user-login')
@perfil_seleccionado_required
def inventario(request):
    pk = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk)
      
    sql_apartadas = """SELECT 
    dashboard_inventario.id AS id,
    SUM(dashboard_articulosparasurtir.cantidad) AS total_cantidad_por_surtir
    FROM
        dashboard_inventario
    JOIN
        dashboard_articulosordenados ON dashboard_inventario.id = dashboard_articulosordenados.producto_id
    JOIN
        dashboard_articulosparasurtir ON dashboard_articulosordenados.id = dashboard_articulosparasurtir.articulos_id
    WHERE
        dashboard_articulosparasurtir.surtir = TRUE
    GROUP BY
        dashboard_inventario.id; """

    sql_entradas = """SELECT 
    dashboard_inventario.id AS id,
        SUM(entradas_entradaarticulo.cantidad_por_surtir) AS total_entradas_por_surtir
    FROM
        dashboard_inventario
    JOIN
        dashboard_product ON dashboard_inventario.producto_id = dashboard_product.id
    JOIN
        dashboard_articulosordenados ON dashboard_inventario.id = dashboard_articulosordenados.producto_id
    JOIN
        dashboard_articulosparasurtir ON dashboard_articulosordenados.id = dashboard_articulosparasurtir.articulos_id
    JOIN
        requisiciones_articulosrequisitados ON dashboard_articulosparasurtir.id = requisiciones_articulosrequisitados.producto_id
    JOIN
        compras_articulocomprado ON requisiciones_articulosrequisitados.id = compras_articulocomprado.producto_id
    JOIN
        entradas_entradaarticulo ON compras_articulocomprado.id = entradas_entradaarticulo.articulo_comprado_id
    WHERE 
        dashboard_inventario.complete = TRUE AND
        dashboard_product.servicio = FALSE AND
        dashboard_product.gasto = FALSE
    GROUP BY
        dashboard_inventario.id;
    """

    resultados_sql_apartadas = Inventario.objects.raw(sql_apartadas)
    resultados_sql_entradas = Inventario.objects.raw(sql_entradas)
    dict_resultados = {r.id: r.total_cantidad_por_surtir for r in resultados_sql_apartadas}
    dict_entradas = {r.id: r.total_entradas_por_surtir for r in resultados_sql_entradas}

    #perfil = Profile.objects.get(staff__staff__id=request.user.id)
    inventario = Inventario.objects.all()
    conteo = inventario.count()


    existencia = Inventario.objects.filter(
        complete=True,
        producto__servicio = False, 
        producto__gasto = False,
        distrito = perfil.distritos
        ).order_by('producto__codigo')

    
    
    if perfil.tipo.nombre == 'Admin' or perfil.tipo.nombre == 'SuperAdm':
        perfil_flag = True
    else:
        perfil_flag = False

    valor_inv = 0
    objetos_a_actualizar = []
    for inv in existencia:
        #inv.total_entradas = dict_entradas.get(inv.id,0)
        inv.total_apartado = dict_resultados.get(inv.id,0) #2 ciclos for uno para calcular el valor del inventario
        inv.cantidad_apartada = inv.total_apartado
        inv.valor_producto = (inv.cantidad + inv.total_apartado) * inv.price # y otro para calcular los apartados
        valor_inv += inv.valor_producto
        objetos_a_actualizar.append(inv)

    with transaction.atomic():
        Inventario.objects.bulk_update(
            objetos_a_actualizar,
            ['cantidad_apartada']  # 👈 solo este campo se actualiza en la tabla
        )
   
    #qs = existencia.annotate(
    #    total_valor=ExpressionWrapper(
    #        (F('cantidad') + F('cantidad_apartada')) * F('price'),
    #        output_field=DecimalField(max_digits=20, decimal_places=2)
    #    )
    #)

    #valor_inv = qs.aggregate(suma=Sum('total_valor'))['suma'] or 0

    myfilter = InventarioFilter(request.GET, queryset=existencia)
    existencia = myfilter.qs

    #Set up pagination
    p = Paginator(existencia, 50)
    page = request.GET.get('page')
    existencia_list = p.get_page(page)

    #for inventario in existencia_list:
    #    inventario.cantidad_por_surtir = dict_resultados.get(inventario.id,0)

    cuenta_productos = existencia.count()

    if request.method =='POST' and 'btnExcel' in request.POST:
        #return convert_excel_inventario(existencia, valor_inv, dict_entradas, dict_resultados)
        return convert_excel_inventario_xlsxwriter(existencia, valor_inv, dict_entradas, dict_resultados)

    context = {
        'conteo':conteo,
        'cuenta_productos':cuenta_productos,
        'perfil_flag':perfil_flag,
        'existencia': existencia,
        'myfilter': myfilter,
        'existencia_list':existencia_list,
        #'perfil':perfil,
        #'entradas':entradas,
        'valor_inv': valor_inv,
        }

    return render(request,'dashboard/inventario.html', context)

@login_required(login_url='user-login')
@perfil_seleccionado_required
def ajuste_inventario(request):
    pk = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk)
    productos_sel = Inventario.objects.filter(complete=True, producto__servicio = False, producto__gasto = False)
    ajuste, created = Entrada_Gasto_Ajuste.objects.get_or_create(almacenista = usuario, completo = False, gasto = None)
    productos_ajuste = Conceptos_Entradas.objects.filter(entrada = ajuste)
    cantidad_items = productos_ajuste.count()
    form = Conceptos_EntradasForm()
    form2 = Entrada_Gasto_AjusteForm()

    form.fields['concepto_material'].queryset = productos_sel

    if request.method == 'POST':
        if "agregar_ajuste" in request.POST:
            form2 = Entrada_Gasto_AjusteForm(request.POST, instance=ajuste)
            if form2.is_valid():
                ajuste.completo= True
                ajuste.completado_hora = datetime.now().time()
                ajuste.completado_fecha = date.today()
                messages.success(request,f'{usuario.staff.first_name},Has hecho un ajuste de manera exitosa')
                #ajuste.save()
                for item_producto in productos_ajuste:
                    producto_inventario = Inventario.objects.get(producto= item_producto.concepto_material.producto)
                    productos_por_surtir = ArticulosparaSurtir.objects.filter(articulos__producto=producto_inventario, requisitar = True)
                    #Calculo el precio 
                    producto_inventario.price = ((item_producto.precio_unitario * item_producto.cantidad)+ ((producto_inventario.cantidad_apartada + producto_inventario.cantidad) * producto_inventario.price))/(producto_inventario.cantidad + item_producto.cantidad + producto_inventario.cantidad_apartada)
                    #La cantidad en inventario + la cantidad del producto en la entrada
                    producto_inventario.cantidad = producto_inventario.cantidad + item_producto.cantidad
                    for item in productos_por_surtir:
                        orden_producto = Order.objects.get(id = item.articulos.orden.id)                
                        #Si la cantidad en inventario es mayor que la cantidad requisitada
                        if producto_inventario.cantidad >= item.cantidad_requisitar:
                            cantidad = item.cantidad_requisitar
                        else:
                            cantidad = producto_inventario.cantidad
                        item.requisitar = False
                        item.cantidad = item.cantidad + cantidad
                        item.cantidad_requisitar = item.cantidad_requisitar - cantidad
                        if item.cantidad_requisitar == 0:
                            item.surtir = True
                        #Se reduce la cantidad de inventario y se aumenta la apartada
                        producto_inventario.cantidad = producto_inventario.cantidad - cantidad
                        #producto_inventario.cantidad_apartada = producto_inventario.cantidad_apartada + cantidad
                        producto_inventario.save()
                        item.save()
                        articulos_por_surtir = ArticulosparaSurtir.objects.filter(articulos__orden=orden_producto)
                        #Se cuentan los articulos por surtir de esa orden, se cuentan los articulos que ya no requieren requisición
                        numero_articulos = articulos_por_surtir.count()
                        numero_articulos_requisitados = articulos_por_surtir.filter(requisitar = False).count()
                        #si el numero total de articulos por surtir ya no requieren requisición
                        if numero_articulos == numero_articulos_requisitados:
                            orden_producto.requisitar = False   # entonces ya no se requiere que la Orden se requisite
                            orden_producto.save()
                    producto_inventario._change_reason = f'Esta es una ajuste desde un ajuste de inventario {ajuste.id}'
                    producto_inventario.save()
                ajuste.save()
                #email = EmailMessage(
                #    f'Ajuste de producto: {ajuste.id}',
                #    f'Estimado {usuario.staff.first_name} {usuario.staff.last_name},\n Estás recibiendo este correo porque tu solicitud: {orden.folio} ha sido devuelta al almacén por {usuario.staff.first_name} {usuario.staff.last_name}, con el siguiente comentario {devolucion.comentario} para más información comunicarse al almacén.\n\n Este mensaje ha sido automáticamente generado por SAVIA VORDTEC',
                #    'savia@vordtec.com',
                #    ['ulises_huesc@hotmail.com'],#orden.staff.staff.email],
                #    )
                #email.send()

                return redirect('solicitud-inventario')

    context= {
        'productos_ajuste':productos_ajuste,
        'form':form,
        'form2':form2,
        'ajuste': ajuste,
        'cantidad_items':cantidad_items,
        'productos_sel': productos_sel,
        }

    return render(request, 'dashboard/ajuste_inventario.html',context)

def update_ajuste(request):
    data= json.loads(request.body)
    action = data["action"]
    cantidad = decimal.Decimal(data["cantidad"])
    ajuste = data["ajuste"]
    producto_id = int(data["id"])
    precio = decimal.Decimal(data["precio"])
    producto = Inventario.objects.get(id=producto_id)
    ajuste = Entrada_Gasto_Ajuste.objects.get(id = ajuste)
    if action == "add":
        articulo, created = Conceptos_Entradas.objects.get_or_create(concepto_material=producto, entrada = ajuste)
        articulo.precio_unitario = precio
        articulo.cantidad = cantidad
        articulo.save()
        messages.success(request,'Has agregado producto de manera exitosa')
        ajuste.save()
    if action == "remove":
        articulo = Conceptos_Entradas.objects.get(concepto_material = producto, entrada = ajuste)
        messages.success(request,'Has eliminado un producto de tu listado')
        articulo.delete()
    return JsonResponse('Item updated, action executed: '+data["action"], safe=False)


@login_required(login_url='user-login')
@perfil_seleccionado_required
def upload_batch_inventario(request):

    form = Inventario_BatchForm(request.POST or None, request.FILES or None)


    if form.is_valid():
        form.save()
        form = Inventario_BatchForm()
        inventario_list = Inventario_Batch.objects.get(activated = False)

        f = open(inventario_list.file_name.path, 'r')
        reader = csv.reader(f)
        next(reader)

        for row in reader:
            if Product.objects.filter(codigo=row[0]):
                producto = Product.objects.get(codigo=row[0])
                if Distrito.objects.filter(nombre = row[1]):
                    distrito = Distrito.objects.get(nombre = row[1])
                    if Almacen.objects.filter(nombre = row[2]):
                        almacen = Almacen.objects.get(nombre = row[2])
                        inventario = Inventario(producto=producto,distrito=distrito, almacen=almacen, ubicacion=row[3], estante=row[4], cantidad=row[5], price=row[6], minimo=row[7],comentario=row[8],complete=True)
                        inventario.save()
                        #marcas_str = row[2]
                        #marcas_names = ast.literal_eval(marcas_str)
                        #marcas_names = map(str.lower, marcas_names) # normalize them, all lowercase
                        #marcas_names = list(set(marcas_names)) # remove duplicates

                        #for marca in marcas_names:
                        #    marca, _ = Marca.objects.get_or_create(nombre=marca)
                        #    inventario.marca.add(marca)
                        #    inventario.save()
                    else:
                        messages.error(request,'El almacén no existe en la base de datos')
                else:
                     messages.error(request,'El distrito no existe en la base de datos')
            else:
                messages.error(request,f'El producto código:{row[0]} ya existe dentro de la base de datos')

        inventario_list.activated = True
        inventario_list.save()
    elif request.FILES:
        messages.error(request,'El formato no es CSV')




    context = {
        'form': form,
        }

    return render(request,'dashboard/upload_batch_inventario.html', context)

@login_required(login_url='user-login')
def upload_batch_inventario_actualizacion(request):
    form = Inventario_BatchForm(request.POST or None, request.FILES or None)

    if request.method == 'POST' and form.is_valid():
        batch = form.save()
        file_path = batch.file_name.path
        errores = []
        try:
            df = pd.read_csv(file_path, encoding='latin1')
        except Exception as e:
            return HttpResponse(f"Error al leer el archivo CSV: {e}", status=400)

        for index, row in df.iterrows():
            codigo_producto = row.iloc[0]
            ubicacion = row.iloc[1]
            cantidad = row.iloc[3]
            almacen_nombre = row.iloc[2]
            precio = row.iloc[4]
            nombre_producto = row.iloc[5] 
            unidad_nombre = row.iloc[6] 
            distrito_nombre = row.iloc[7] 
            servicio_str = row.iloc[8]
            print(codigo_producto)
            print(nombre_producto)
            print(almacen_nombre)
            try:
                producto = Product.objects.get(codigo=codigo_producto)
                print(producto)
                distrito = Distrito.objects.get(nombre=almacen_nombre)
                print(distrito)
                almacen = Almacen.objects.get(nombre=almacen_nombre)
                inventario = Inventario.objects.filter(producto=producto).first()
                print(inventario)
                
                if inventario:
                    inventario.ubicacion = ubicacion
                    inventario.cantidad = cantidad
                    inventario.almacen = almacen
                    inventario.distrito = distrito
                    inventario.price = precio
                    inventario.complete = True
                    inventario.updated_at = timezone.now()
                    inventario.save()
                    # Actualizar producto
                    if nombre_producto:
                        producto.nombre = nombre_producto
                    if servicio_str:
                        producto.servicio = True if servicio_str.strip().upper() == 'SI' else False
                    if unidad_nombre:
                        try:
                            unidad = Unidad.objects.get(nombre=unidad_nombre)
                            producto.unidad = unidad
                        except Unidad.DoesNotExist:
                            errores.append({codigo_producto, f"Unidad '{unidad_nombre}' no encontrada"})
                    producto.updated_at = timezone.now()
                    producto.save()
                    
                    #Este es el masivo que elimina los articulos por surtir #No utilizar a menos que sea nesario
                    #articulos_surtir = ArticulosparaSurtir.objects.filter(
                    #    surtir=True, 
                    #    articulos__producto= inventario
                    #)
                    #print(articulos_surtir)
                    #for articulo in articulos_surtir:
                    #    articulo.surtir = False
                    #    articulo.cantidad = 0
                    #    articulo.save()

            except Product.DoesNotExist:
                errores.append({codigo_producto, f"Producto no encontrado"})
            except Almacen.DoesNotExist:
                errores.append({codigo_producto, f"Almacen '{almacen_nombre}' no encontrado"})
            except Distrito.DoesNotExist:
                errores.append({codigo_producto, f"Distrito '{distrito_nombre}' no encontrado"})
            except Unidad.DoesNotExist:
                errores.append({codigo_producto, f"Unidad '{unidad_nombre}' no encontrada"})

        batch.activated = True
        batch.save()
        # Si hay errores, guardar archivo y devolver enlace
        # Si hay errores, guardar archivo CSV
        # Si hay errores, generar y devolver el archivo como descarga directa
        if errores:
            print('Estoy entrando en errores')
            errores_df = pd.DataFrame(errores)
            root_dir = settings.BASE_DIR
            error_dir = os.path.join(root_dir, 'errores_batch')
            os.makedirs(error_dir, exist_ok=True)

            error_file_name = f'errores_batch_{batch.id}.csv'
            error_file_path = os.path.join(error_dir, error_file_name)

            errores_df.to_csv(error_file_path, index=False, encoding='utf-8')

            # Si lo quieres servir vía navegador local, solo devuelves la ruta relativa o absoluta:
            error_file_url = f"/errores_batch/{error_file_name}"
            return JsonResponse({'status': 'partial', 'error_file': error_file_url}, status=207)
        return HttpResponse(status=204)

    return render(request, 'dashboard/upload_batch_inventario.html', {'form': form}) 

@login_required(login_url='user-login')
def upload_batch_inventario_nuevos(request):
    form = Inventario_BatchForm(request.POST or None, request.FILES or None)

    if request.method == 'POST' and form.is_valid():
        batch = form.save()
        file_path = batch.file_name.path
        errores = []

        try:
            df = pd.read_csv(file_path, encoding='latin1')
        except Exception as e:
            return HttpResponse(f"Error al leer el archivo CSV: {e}", status=400)

        for index, row in df.iterrows():
            codigo_producto = row.iloc[0]
            ubicacion = row.iloc[1]
            cantidad = row.iloc[3]
            almacen_nombre = row.iloc[2]
            precio = row.iloc[4]
            nombre_producto = row.iloc[5]
            unidad_nombre = row.iloc[6]
            distrito_nombre = row.iloc[7]
            servicio_str = row.iloc[8]
            #Para cuando NO tienes el producto y también quieres agregar el inventario (Cargar ambos nuevos)
            #Inica Bloque #0
            # Valida que no exista producto con ese código o nombre
            if Product.objects.filter(codigo=codigo_producto).exists():
                errores.append({'Código': codigo_producto, 'Razón': "El código ya existe"})
                continue

            if Product.objects.filter(nombre=nombre_producto).exists():
                errores.append({'Código': codigo_producto, 'Razón': f"El nombre '{nombre_producto}' ya existe"})
                continue

          
            

            try:
                producto = Product(
                    codigo=codigo_producto,
                    nombre=nombre_producto,
                    updated_at=timezone.now()
                    
                )
            #Termina Bloque #0

            #Para cuando tienes el producto pero quieres agregar el inventario
            #Inicia Bloque #1 
            #try:
            #    producto = Product.objects.get(codigo=codigo_producto)

                # Actualizar nombre si es diferente
            #    if nombre_producto and producto.nombre != nombre_producto:
            #        producto.nombre = nombre_producto

            #Acaba Bloque #1
                if servicio_str:       
                    producto.servicio = True if servicio_str.strip().upper() == 'SI' else False
                    

                if unidad_nombre:
                    try:
                        unidad = Unidad.objects.get(nombre=unidad_nombre)
                        producto.unidad = unidad
                        
                    except Unidad.DoesNotExist:
                        errores.append({'Código': codigo_producto, 'Razón': f"Unidad '{unidad_nombre}' no encontrada"})
                        continue


                producto.save()

                if distrito_nombre:
                    try:
                        distrito = Distrito.objects.get(nombre=distrito_nombre)
                    except Distrito.DoesNotExist:
                        errores.append({'Código': codigo_producto, 'Razón': f"Distrito '{distrito_nombre}' no encontrado"})
                        continue

                try:
                    almacen = Almacen.objects.get(nombre=almacen_nombre)
                except Almacen.DoesNotExist:
                    errores.append({'Código': codigo_producto, 'Razón': f"Almacén '{almacen_nombre}' no encontrado"})
                    continue

                Inventario.objects.create(
                    producto=producto,
                    ubicacion=ubicacion,
                    cantidad=cantidad,
                    price=precio,
                    complete=True,
                    distrito=distrito,
                    almacen=almacen,
                )

            #except Product.DoesNotExist:
            #    errores.append({'Código': codigo_producto, 'Razón': "Producto con ese código no existe"})
            #    continue

            except Exception as e:
                errores.append({'Código': codigo_producto, 'Razón': f"Error inesperado: {str(e)}"})

        batch.activated = True
        batch.save()

        if errores:
            errores_df = pd.DataFrame(errores)
            root_dir = settings.BASE_DIR
            error_dir = os.path.join(root_dir, 'errores_batch')
            os.makedirs(error_dir, exist_ok=True)

            error_file_name = f'errores_nuevos_batch_{batch.id}.csv'
            error_file_path = os.path.join(error_dir, error_file_name)

            errores_df.to_csv(error_file_path, index=False, encoding='utf-8', sep=';')

            error_file_url = f"/errores_batch/{error_file_name}"
            return JsonResponse({'status': 'partial', 'error_file': error_file_url}, status=207)

        return HttpResponse(status=204)

    return render(request, 'dashboard/upload_batch_inventario.html', {'form': form})

@login_required(login_url='user-login')
@perfil_seleccionado_required
def inventario_add(request):
    #usuario = request.user.id
    pk = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk)

    #productos = Product.objects.exclude(id__in=existing)
    form = InventarioForm(distrito = perfil.distritos)
    #form.fields['producto'].queryset = productos


    if request.method =='POST':
        form = InventarioForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.complete = True
            item._change_reason = 'Se agrega producto el inventario en view: inventario_add'
            item.distrito = perfil.distritos
            item.save()
            messages.success(request, f'El artículo {item.producto.codigo}:{item.producto.nombre} se ha agregado exitosamente')
            return redirect('solicitud-inventario')
    #else:
        #form = InventarioForm()

    context = {
        'form': form,
        #'productos':productos,
        }

    return render(request,'dashboard/inventario_add.html',context)

@login_required(login_url='user-login')
@perfil_seleccionado_required
def inventario_update_modal(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk_perfil)
    item = Inventario.objects.get(id=pk)



    if perfil.tipo.nombre == 'SuperAdm' or perfil.tipo.nombre == 'Admin':
        flag_perfil = True
    else:
        flag_perfil = False


    if request.method =='POST':
        if perfil.tipo.nombre == 'SuperAdm' or perfil.tipo.nombre == 'Admin':
            form = Inv_UpdateForm(request.POST, instance=item)
        else:
            form = Inv_UpdateForm_almacenista(request.POST, instance= item)
        if request.POST['comentario'] and 'btnUpdate' in request.POST:
            if form.is_valid():
                item = form.save(commit=False)
                item._change_reason = item.comentario +'. Se modifica inventario en view: inventario_update_modal'
                item.save()
                messages.success(request, f'El artículo {item.producto.codigo}:{item.producto.nombre} se ha actualizado exitosamente')
                return HttpResponse(status=204)
        else:
            messages.error(request, 'Debes agregar un comentario con respecto al cambio realizado')
    else:
        if perfil.tipo.nombre == 'SuperAdm' or perfil.tipo.nombre == 'Admin' or perfil.tipo.nombre == "Almacen":
            form = Inv_UpdateForm(instance=item)
        else:
            form = Inv_UpdateForm_almacenista(instance= item)


    context = {
        'flag_perfil':flag_perfil,
        'form': form,
        'item':item,
        }

    return render(request,'dashboard/inventario_update_modal.html',context)


@login_required(login_url='user-login')
@perfil_seleccionado_required
def historico_inventario(request):
    registros = Inventario.history.all()

    myfilter = HistoricalInventarioFilter(request.GET, queryset=registros)
    registros = myfilter.qs

    #Set up pagination
    p = Paginator(registros, 30)
    page = request.GET.get('page')
    registros_list = p.get_page(page)

    context = {
        'registros_list':registros_list,
        'myfilter':myfilter,
        }

    return render(request,'dashboard/historico_inventario.html',context)


@login_required(login_url='user-login')
@perfil_seleccionado_required
def historico_producto(request):
    registros = Product.history.all()

    myfilter = HistoricalProductoFilter(request.GET, queryset=registros)
    registros = myfilter.qs

    #Set up pagination
    p = Paginator(registros, 30)
    page = request.GET.get('page')
    registros_list = p.get_page(page)

    context = {
        'registros_list':registros_list,
        'myfilter':myfilter,
        }

    return render(request,'dashboard/historico_producto.html',context)


@login_required(login_url='user-login')
@perfil_seleccionado_required
def inventario_delete(request, pk):
    item = Inventario.objects.get(id=pk)

    if request.method == 'POST':
        item.delete()
        return redirect('solicitud-inventario')

    return render(request,'dashboard/inventario_delete.html')

@login_required(login_url='user-login')
@perfil_seleccionado_required
def solicitud_autorizacion(request):
    pk = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk)
    if usuario.sustituto:
        usuario= Profile.objects.filter(staff=usuario.staff, tipo=usuario.tipo, distritos=usuario.distritos).first()
        
    
    #Estoy obteniendo todos los perfiles que coinciden con el mismo staff, tipo y distrito que el usuario actual
    usuarios = Profile.objects.filter(staff=usuario.staff, tipo=usuario.tipo, distritos=usuario.distritos)
    
    
    
   
    

    #Este es un filtro por perfil supervisor o superintendente, es decir puede ver todo lo del distrito
    ordenes = Order.objects.filter(complete=True, autorizar=None, distrito =usuario.distritos, supervisor__in= usuarios).order_by('-folio')
    #ordenes = ordenes.filter(supervisor=perfil)
    myfilter=SolicitudesFilter(request.GET, queryset=ordenes)
    ordenes = myfilter.qs


    context= {
        'myfilter':myfilter,
        'ordenes':ordenes,
        #'perfil':perfil,
        }

    return render(request, 'autorizacion/solicitudes_pendientes_autorizacion.html',context)

@login_required(login_url='user-login')
@perfil_seleccionado_required
def detalle_autorizar(request, pk):
    orden = Order.objects.get(id=pk)
    productos = ArticulosOrdenados.objects.filter(orden=pk)

    context = {
        'orden':orden,
        'productos': productos,
     }
    return render(request,'autorizacion/detail.html', context)

@login_required(login_url='user-login')
@perfil_seleccionado_required
def autorizada_sol(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk_perfil)
    #perfil = Profile.objects.get(id=usuario)
    order = Order.objects.get(id = pk)
    productos = ArticulosOrdenados.objects.filter(orden = pk)
    requis = Requis.objects.filter(orden__staff__distritos = perfil.distritos)
    consecutivo = requis.count() + 1

    if request.method =='POST':
        #We go trough all the products one by one contained in the order
        for producto in productos:
            # We fetch inventory product corresponding to product (that's why we use product.id)
            # We create a new product line in a new database to control the ArticlestoDeliver (ArticulosparaSurtir)
            prod_inventario = Inventario.objects.get(id = producto.producto.id)
            ordensurtir , created = ArticulosparaSurtir.objects.get_or_create(articulos = producto)
            #cond:1 evalua si la cantidad en inventario es mayor que lo solicitado
            if order.tipo.tipo == "resurtimiento" or  producto.producto.producto.servicio == True: # or  producto.producto.producto.activo == True: se comenta para crear la creación de requisiciones automáticas de activos #prod_inventario.cantidad + prod_inventario.cantidad_entradas == 0 or 
                ordensurtir.requisitar = True
                ordensurtir.cantidad_requisitar = producto.cantidad
                if producto.producto.producto.servicio == True or producto.producto.producto.activo == True:
                    requi, created = Requis.objects.get_or_create(complete = True, orden = order)
                    requitem, created = ArticulosRequisitados.objects.get_or_create(req = requi, producto= ordensurtir, cantidad = producto.cantidad)
                    requis = Requis.objects.filter(orden__distrito = perfil.distritos, complete = True)
                    max_folio = Requis.objects.filter(orden__distrito=perfil.distritos, complete=True).aggregate(Max('folio'))['folio__max']
                    requi.folio = max_folio + 1
                    numero_servicios = productos.filter(producto = producto.producto.producto.servicio).count()
                    if productos.count() == numero_servicios: #No tengo claridad que es lo que pretendo contar acá
                        order.requisitar=False
                        order.requisitado = True
                    ordensurtir.requisitar = False
                    requi.save()
                    requitem.save()
                ordensurtir.save()
                order.save()
            elif producto.cantidad >= prod_inventario.cantidad and producto.cantidad > 0 and order.tipo.tipo == "normal" and producto.producto.producto.servicio == False: #and producto.producto.producto.activo == False: #si la cantidad solicitada es mayor que la cantidad en inventario 
                ordensurtir.cantidad = prod_inventario.cantidad #lo que puedes surtir es igual a lo que tienes en el inventario
                ordensurtir.precio = prod_inventario.price
                ordensurtir.cantidad_requisitar = producto.cantidad - ordensurtir.cantidad #lo que falta por surtir
                
                prod_inventario.cantidad_apartada = prod_inventario.apartada()
                prod_inventario.cantidad = 0
                if ordensurtir.cantidad > 0: #si lo que se puede surtir es mayor que 0
                    ordensurtir.surtir = True
                ordensurtir.requisitar = True
                order.requisitar = True
                prod_inventario.save()
                ordensurtir.save()
                order.save()
            elif prod_inventario.cantidad >= producto.cantidad and order.tipo.tipo == "normal":
                prod_inventario.cantidad = prod_inventario.cantidad - producto.cantidad
                prod_inventario.cantidad_apartada = prod_inventario.apartada()
                prod_inventario._change_reason = f'Se modifica el inventario en view: autorizada_sol:{order.id} Autorización de solicitudes cond:1'
                ordensurtir.cantidad = producto.cantidad
                ordensurtir.precio = prod_inventario.price
                ordensurtir.surtir = True
                ordensurtir.requisitar = False
                ordensurtir.save()
                prod_inventario.save()
            
          
        order.autorizar = True
        order.approved_at = date.today()
        order.approved_at_time = datetime.now().time()
        #send_mail(
        #    f'Solicitud Autorizada {order.folio}',
        #    f'{order.staff.staff.first_name}, la solicitud {order.folio} ha sido autorizada. Este mensaje ha sido automáticamente generado por SAVIA X',
        #    'saviax.vordcab@gmail.com',
        #    [order.staff.staff.email],
        #    )
        order.sol_autorizada_por = Profile.objects.get(id = pk_perfil)
        order.save()

        messages.success(request, f'{perfil.staff.staff.first_name} has autorizado la solicitud {order.folio}')
        return redirect('solicitud-pendientes-autorizacion')


    context = {
        'orden': order,
        'productos': productos,
    }

    return render(request,'autorizacion/autorizada.html', context)

@login_required(login_url='user-login')
@perfil_seleccionado_required
def cancelada_sol(request, pk):
    order = Order.objects.get(id = pk)
    productos = ArticulosOrdenados.objects.filter(orden = pk)

    if request.method =='POST':
        order.autorizar = False
        order.save()
        messages.success(request, f'La orden {order} ha sido cancelada')
        return redirect('solicitud-pendientes-autorizacion')


    context = {
        'orden': order,
        'productos': productos,
    }

    return render(request,'autorizacion/cancelada.html', context)

@login_required(login_url='user-login')
@perfil_seleccionado_required
def status_sol_old(request, pk):
    solicitud = Order.objects.get(id = pk)
    product_solicitudes = ArticulosOrdenados.objects.filter(orden=pk)
    #product_surtir = ArticulosparaSurtir.objects.filter(articulos__orden = pk)
    
    context = {
        'productos_solicitados': product_solicitudes,
        'solicitud':solicitud,
    }

    return render(request,'solicitud/detalle.html', context)



@login_required(login_url='user-login')
@perfil_seleccionado_required
def status_sol(request, pk):

    solicitud = get_object_or_404(
        Order.objects
        .select_related(
            'distrito',
            'supervisor__staff__staff',
            'superintendente__staff__staff',
            'superintendente__tipo',
        )
        .prefetch_related(
            'productos',
            'requis__productos',
            'requis__compras__pagos',
            'requis__compras__vale_entrada',
            'requis__compras__articulocomprado_set',
            'vale_salida',
        ),
        pk=pk
    )

    trazabilidad = []

    # =====================================================
    # FUNCIONES AUXILIARES
    # =====================================================

    def nombre_perfil(perfil):
        if not perfil:
            return 'No asignado'

        try:
            nombre = perfil.staff.staff.first_name
            apellido = perfil.staff.staff.last_name

            return f'{nombre} {apellido}'.strip()

        except AttributeError:
            return str(perfil)

    def agregar_siguiente_etapa(
        tipo,
        descripcion,
        icono,
        responsable_accion=None,
        responsable_autorizacion=None,
        autorizador=None,
    ):
        trazabilidad.append({
            'tipo': tipo,
            'folio': 'Siguiente etapa',
            'estado': 'pendiente',
            'descripcion': descripcion,
            'fecha': None,
            'icono': icono,
            'detalle': None,
            'objeto': None,
            'es_siguiente_etapa': True,

            'responsable_accion': responsable_accion,

            'rol_autorizador': responsable_autorizacion,
            'autorizador': autorizador,
        })

    # =====================================================
    # DATOS GENERALES
    # =====================================================

    es_matriz = (
        solicitud.distrito.nombre.strip().upper() == 'MATRIZ'
    )

    nombre_supervisor = nombre_perfil(solicitud.supervisor)
    nombre_superintendente = nombre_perfil(
        solicitud.superintendente
    )

    # -----------------------------------------------------
    # RUTAS POSIBLES
    # -----------------------------------------------------
    requisiciones = list(solicitud.requis.all())

    ruta_almacen = (
        solicitud.complete is True
        and solicitud.autorizar is True
        and not solicitud.requisitar
        and not solicitud.requisitado
        and not requisiciones
    )

    ruta_compra = (
        solicitud.requisitar is True
        or solicitud.requisitado is True
        or bool(requisiciones)
    )

    if ruta_almacen:
        ruta_solicitud = 'Surtido directo por almacén'

    elif ruta_compra:
        ruta_solicitud = 'Requisición y compra'

    else:
        ruta_solicitud = 'Ruta pendiente de definir'

    # =====================================================
    # VARIABLES DE CONTROL DEL AVANCE
    # =====================================================

    hay_requisiciones = False
    hay_requisicion_autorizada = False

    hay_compras = False
    hay_compra_autorizada = False
    hay_compra_con_servicios = False

    hay_pagos = False
    hay_pago_realizado = False

    hay_entradas = False
    hay_entrada_completa = False

    hay_salidas = False
    hay_salida_completa = False



    # =====================================================
    # 1. SOLICITUD
    # =====================================================

    if solicitud.autorizar is True:
        estado = 'completado'
        descripcion = 'Solicitud autorizada'
        fecha = solicitud.approved_at

    elif solicitud.autorizar is False:
        estado = 'cancelado'
        descripcion = 'Solicitud cancelada'
        fecha = None

    else:
        estado = 'proceso'
        descripcion = 'Solicitud pendiente de autorización'
        fecha = None

    trazabilidad.append({
        'tipo': 'Solicitud',
        'folio': solicitud.folio,
        'estado': estado,
        'descripcion': descripcion,
        'fecha': fecha,
        'icono': 'fa-file-circle-check',
        'detalle': 'solicitud',
        'objeto': solicitud,
        'rol_autorizador': 'Supervisor',
        'autorizador': nombre_supervisor,
    })

    # =====================================================
    # RUTA DE REQUISICIÓN Y COMPRA
    # =====================================================

    if ruta_compra:

        requisiciones = list(solicitud.requis.all())
        hay_requisiciones = bool(requisiciones)

        # =================================================
        # 2. REQUISICIONES
        # =================================================

        for requisicion in requisiciones:

            if requisicion.autorizar is True:
                estado = 'completado'
                descripcion = 'Requisición autorizada'
                fecha = requisicion.approved_at
                hay_requisicion_autorizada = True

            elif requisicion.autorizar is False:
                estado = 'cancelado'
                descripcion = 'Requisición cancelada'
                fecha = None

            elif requisicion.complete is False:
                estado = 'proceso'
                descripcion = 'Requisición en elaboración'
                fecha = None

            else:
                estado = 'proceso'
                descripcion = 'Requisición pendiente de autorización'
                fecha = None

            if es_matriz:
                rol_autorizador = 'Supervisor'
                autorizador = nombre_supervisor
            else:
                rol_autorizador = 'Superintendente'
                autorizador = nombre_superintendente

            trazabilidad.append({
                'tipo': 'Requisición',
                'folio': requisicion.folio or 'Sin folio',
                'estado': estado,
                'descripcion': descripcion,
                'fecha': fecha,
                'icono': 'fa-clipboard-list',
                'detalle': 'requisicion',
                'objeto': requisicion,
                'rol_autorizador': rol_autorizador,
                'autorizador': autorizador,
            })

            compras = list(requisicion.compras.all())

            if compras:
                hay_compras = True

            # =============================================
            # 3. ÓRDENES DE COMPRA
            # =============================================

            for compra in compras:

                if es_matriz:
                    rol_autorizador_compra = 'Subdirector'
                    autorizador_compra = nombre_superintendente
                else:
                    rol_autorizador_compra = 'Gerente del distrito'
                    autorizador_compra = 'No asignado'

                if compra.solo_servicios:
                    hay_compra_con_servicios = True

                if compra.autorizado2 is True:
                    estado = 'completado'
                    descripcion = 'Orden de compra autorizada'
                    fecha = compra.autorizado_at_2
                    hay_compra_autorizada = True

                elif compra.autorizado1 is False:
                    estado = 'cancelado'
                    if es_matriz:
                        descripcion = (
                            'Orden de compra cancelada por Subdirector'
                        )
                    else:
                        descripcion = (
                            'Orden de compra cancelada por Superintendente'
                           
                        )

                    fecha = compra.autorizado_at_2


                elif compra.autorizado2 is False:
                    estado = 'cancelado'
                    descripcion = (
                        'Orden de compra cancelada por Gerencia'
                    )
                    fecha = compra.autorizado_at_2
                
               

                elif compra.autorizado1 is True:
                    estado = 'proceso'

                    if es_matriz:
                        descripcion = (
                            'Pendiente de autorización del Subdirector'
                        )
                    else:
                        descripcion = (
                            'Pendiente de autorización del Gerente '
                            'del distrito'
                        )

                    fecha = compra.autorizado_at

                elif compra.autorizado1 is False:
                    estado = 'cancelado'
                    descripcion = (
                        'Orden de compra cancelada en la '
                        'primera autorización'
                    )
                    fecha = compra.autorizado_at

                elif compra.complete:
                    estado = 'proceso'
                    descripcion = (
                        'Orden de compra pendiente de autorización'
                    )
                    fecha = compra.created_at

                else:
                    estado = 'proceso'
                    descripcion = 'Orden de compra en elaboración'
                    fecha = compra.created_at

                trazabilidad.append({
                    'tipo': 'Orden de compra',
                    'folio': compra.folio or 'Sin folio',
                    'estado': estado,
                    'descripcion': descripcion,
                    'fecha': fecha,
                    'icono': 'fa-cart-shopping',
                    'detalle': 'compra',
                    'objeto': compra,
                    'rol_autorizador': rol_autorizador_compra,
                    'autorizador': autorizador_compra,
                    'compra_con_servicios': compra.solo_servicios, 
                })

                # =========================================
                # 4. PAGOS
                # =========================================

                pagos = list(compra.pagos.all())

                if pagos:
                    hay_pagos = True

                for pago in pagos:

                    if pago.hecho:
                        estado = 'completado'
                        descripcion = 'Pago realizado'
                        fecha = (
                            pago.pagado_real
                            or pago.pagado_date
                        )
                        hay_pago_realizado = True

                    else:
                        estado = 'proceso'
                        descripcion = 'Pago pendiente'
                        fecha = None

                    trazabilidad.append({
                        'tipo': 'Pago',
                        'folio': f'Pago #{pago.id}',
                        'estado': estado,
                        'descripcion': descripcion,
                        'fecha': fecha,
                        'icono': 'fa-money-check-dollar',
                        'objeto': pago,
                    })

                # =========================================
                # 5. ENTRADAS
                # =========================================

                entradas = list(compra.vale_entrada.all())

                if entradas:
                    hay_entradas = True

                for entrada in entradas:

                    if entrada.completo:
                        estado = 'completado'
                        descripcion = (
                            'Material recibido en almacén'
                        )
                        fecha = getattr(
                            entrada,
                            'entrada_date',
                            getattr(entrada, 'fecha', None)
                        )
                        hay_entrada_completa = True

                    else:
                        estado = 'proceso'
                        descripcion = (
                            'Recepción de material en proceso'
                        )
                        fecha = None

                    trazabilidad.append({
                        'tipo': 'Entrada de almacén',
                        'folio': entrada.folio or 'Sin folio',
                        'estado': estado,
                        'descripcion': descripcion,
                        'fecha': fecha,
                        'icono': 'fa-warehouse',
                        'detalle': 'entrada',
                        'objeto': entrada,
                    })

    # =====================================================
    # 6. SALIDAS
    # Aplica para las dos rutas
    # =====================================================

    salidas = list(solicitud.vale_salida.all())
    hay_salidas = bool(salidas)

    for vale in salidas:

        if vale.complete:
            estado = 'completado'
            descripcion = 'Material entregado al solicitante'
            fecha = vale.created_at
            hay_salida_completa = True

        else:
            estado = 'proceso'
            descripcion = 'Surtido de almacén en proceso'
            fecha = vale.created_at

        trazabilidad.append({
            'tipo': 'Salida de almacén',
            'folio': vale.folio or 'Sin folio',
            'estado': estado,
            'descripcion': descripcion,
            'fecha': fecha,
            'icono': 'fa-truck-ramp-box',
            'detalle': 'salida',
            'objeto': vale,
        })

    # =====================================================
    # SIGUIENTE ETAPA
    # Se agrega después de analizar todo el proceso
    # =====================================================

    if solicitud.autorizar is True:

        # -------------------------------------------------
        # RUTA 1: SURTIDO DIRECTO DESDE ALMACÉN
        # -------------------------------------------------

        if ruta_almacen:

            if not hay_salidas:
                agregar_siguiente_etapa(
                    tipo='Surtido de almacén',
                    descripcion=(
                        'La solicitud está autorizada y está '
                        'pendiente de ser surtida por el almacén'
                    ),
                    icono='fa-boxes-stacked',
                    responsable_accion='Almacén',
                )

            # Si ya existe salida en proceso, su tarjeta
            # real ya muestra el estado.

        # -------------------------------------------------
        # RUTA 2: REQUISICIÓN Y COMPRA
        # -------------------------------------------------

        elif ruta_compra:

            if not hay_requisiciones:
                agregar_siguiente_etapa(
                    tipo='Requisición',
                    descripcion='Pendiente de elaboración de la requisición',
                    icono='fa-clipboard-list',
                    responsable_accion='Almacén',
                    responsable_autorizacion=(
                        'Supervisor'
                        if es_matriz
                        else 'Superintendente'
                    ),
                    autorizador=(
                        nombre_supervisor
                        if es_matriz
                        else nombre_superintendente
                    ),
                )
            elif not hay_requisicion_autorizada:
                # La tarjeta de requisición ya informa
                # quién debe autorizar.
                pass

            elif not hay_compras:
                agregar_siguiente_etapa(
                    tipo='Orden de compra',
                    descripcion=(
                        'Pendiente de elaboración de la '
                        'Orden de Compra'
                    ),
                    icono='fa-cart-shopping',
                    responsable_accion='Área de Compras',
             
                )

            elif not hay_compra_autorizada:
                # La tarjeta de la OC ya muestra
                # el autorizador correspondiente.
                pass

            elif hay_compra_autorizada:
                if not hay_pagos:
                    agregar_siguiente_etapa(
                        tipo='Pago',
                        descripcion='Pendiente de programación del pago',
                        icono='fa-money-check-dollar',
                        responsable_accion='Tesorería',
                    )

                elif not hay_pago_realizado:
                    # Ya existe un pago pendiente.
                    pass

                            
                # -----------------------------------------
                # RECEPCIÓN / ENTRADA / SALIDA
                # -----------------------------------------

                if not hay_entradas:
                    agregar_siguiente_etapa(
                        tipo='Entrada de almacén',
                        descripcion=(
                            'Pendiente de recepción del material '
                            'en almacén'
                        ),
                        icono='fa-warehouse',
                        responsable_accion='Almacén',
                    )

                elif not hay_entrada_completa:
                    # Ya existe una entrada en proceso.
                    pass

                elif not hay_compra_con_servicios and not hay_salidas:
                    agregar_siguiente_etapa(
                        tipo='Salida de almacén',
                        descripcion=(
                            'Pendiente de entrega del material '
                            'al solicitante'
                        ),
                        icono='fa-truck-ramp-box',
                        responsable_accion='Almacén',
                    )

        # -------------------------------------------------
        # RUTA NO DEFINIDA
        # -------------------------------------------------

        else:
            agregar_siguiente_etapa(
                tipo='Definición de surtimiento',
                descripcion=(
                    'La solicitud está autorizada, pero todavía '
                    'no se ha definido si será surtida directamente '
                    'por almacén o mediante requisición'
                ),
                icono='fa-code-branch',
            )

    context = {
        'solicitud': solicitud,
        'productos_solicitados': solicitud.productos.all(),
        'trazabilidad': trazabilidad,
        'ruta_solicitud': ruta_solicitud,
    }

    return render(
        request,
        'solicitud/trazabilidad.html',
        context
    )


# AJAX
def load_subproyectos(request):
    #term = request.GET.get('term')
    proyecto_id = request.GET.get('proyecto_id')
    
    
    subproyectos =Subproyecto.objects.filter(proyecto__id = proyecto_id, status__nombre = "Activo" ).values('id','nombre')
    data = list(subproyectos)
    return JsonResponse(data, safe=False)
    #return render(request, 'solicitud/subproyecto_dropdown_list_options.html',{'subproyectos': subproyectos})

def convert_excel_inventario(existencia, valor_inventario, dict_entradas, dict_resultados):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Inventario_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Inventario')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)

    columns = ['Código','Producto','Distrito','Unidad','Cantidad','Cantidad Apartada','Cantidad Entradas','Minimos','Ubicación','Estante','Precio','Total']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        if col_num == 0:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 10
        if col_num== 1:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        else:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 15


    columna_max = len(columns)+3

    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por SAVIA 2.0. UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style
    (ws.cell(column = columna_max, row = 3, value='Inventario Costo Total:')).style = messages_style
    (ws.cell(column = columna_max +1, row=3, value = valor_inventario)).style = money_resumen_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 20
    ws.column_dimensions[get_column_letter(columna_max + 1)].width = 20

    #rows = existencia.values_list('producto__codigo','producto__nombre','distrito__nombre','producto__unidad__nombre','cantidad','cantidad_apartada','ubicacion','estante','price')
    
    rows = existencia.all()

    

    for inventario in rows:
        row_num += 1
        
        inventario.total_entradas = dict_entradas.get(inventario.id, 0)
        inventario.total_apartado = dict_resultados.get(inventario.id, 0)                      
        # Aquí estás creando una lista manualmente con los valores que necesitas
        row = [
            inventario.producto.codigo,
            inventario.producto.nombre,
            inventario.distrito.nombre,
            inventario.producto.unidad.nombre,
            inventario.cantidad,
            #inventario.apartada,  # Aquí utilizas la propiedad apartada
            inventario.total_apartado,
            inventario.total_entradas,
            inventario.minimo,
            inventario.ubicacion,
            inventario.estante,
            inventario.price
        ]

        for col_num in range(len(row)):
            cell = ws.cell(row=row_num, column=col_num +1, value=row[col_num])
            if col_num > 2 and col_num != 8:
                cell.style = body_style #(ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = body_style
            elif col_num == 8:
                cell.style = money_style #(ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = money_style
            else:
                cell.style = body_style#(ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style

        total_value = inventario.cantidad * inventario.price + inventario.apartada * inventario.price
        total_cell = ws.cell(row=row_num, column=len(row)+1, value=total_value)
        total_cell.style = money_style

    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)


def convert_excel_solicitud_matriz_productos(productos):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Solicitudes_por_producto_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Solicitudes')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)

    columns = ['Folio','Solicitante','Proyecto','Subproyecto','Operación','Cantidad','Código', 'Producto','Creado']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16
        if col_num == 4 or col_num == 7:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 25



    columna_max = len(columns)+2

    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por Savia 2.0 UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Grupo Vordcab S.A. de C.V.}')).style = messages_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 20

    rows = productos.values_list('orden__id',Concat('orden__staff__staff__staff__first_name',Value(' '),'orden__staff__staff__staff__last_name'),'orden__proyecto__nombre','orden__subproyecto__nombre',
                                'orden__operacion__nombre','cantidad','producto__producto__codigo','producto__producto__nombre','orden__created_at')

    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            cell_value = row[col_num]
            if col_num == 8 and isinstance(cell_value, datetime):  # Asumiendo que 'orden__created_at' es un objeto datetime
                cell_value = cell_value.replace(tzinfo=None)  # Convertir a 'naive'
            (ws.cell(row = row_num, column = col_num+1, value=str(cell_value))).style = body_style
            if col_num == 8:
                (ws.cell(row = row_num, column = col_num+1, value=cell_value)).style = date_style
    # Continuación del código...
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)

def convert_excel_solicitud_matriz(ordenes):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Solicitudes_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Solicitudes')

    #Create heading style and adding to workbook | Crear estilos y agregarlos al Workbook
    #Head 
    head1_style = NamedStyle(name = "head1_style")
    head1_style.font = Font(name = 'Arial', color = '00003366', bold = True, size = 18)
    #Head table
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)
    # Construir la ruta completa a la imagen
    static_path = settings.STATIC_ROOT
    img_path = os.path.join(static_path,'images','logo_vordcab.jpg')
    # Añadir imagen
    img = Image(img_path)
    # Ajustar tamaño de la imagen
    img.width = 120  # Ajusta el ancho como sea necesario
    img.height = 120  # Ajusta el alto como sea necesario

    ws.add_image(img, 'A1')
    
     # Añadir título y fechas
    (ws.cell(column = 3, row = 1, value = 'MATRIZ DE SOLICITUDES')).style = head1_style
    (ws.cell(column = 3, row = 4, value='Reporte Creado Automáticamente por SAVIA 2.0. UH')).style = messages_style
    (ws.cell(column = 3, row = 5, value='Software desarrollado por Grupo Vordcab S.A. de C.V.')).style = messages_style
    #fecha_min = min((orden.created_at for orden in ordenes if orden.created_at is not None), default=None)  # Suponiendo que 'fecha' es un atributo de tus 'ordenes'
    #fecha_max = max((orden.created_at for orden in ordenes if orden.created_at is not None), default=None)
    #ws['B6'] = f'Fecha desde: {fecha_min} hasta: {fecha_max}'

    #ws.column_dimensions[get_column_letter(columna_max)].width = 20

    #Comenzar en la fila 6
    row_num = 6
    

    columns = ['Folio','Solicitante','Distrito','Proyecto','Subproyecto','Operación','Creado']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 25
        if col_num == 5:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 20

    columna_max = len(columns)+2

    
    rows = ordenes.values_list('folio',Concat('staff__staff__staff__first_name',Value(' '),'staff__staff__staff__last_name'),'distrito__nombre','proyecto__nombre','subproyecto__nombre',
                                'operacion__nombre','created_at')

    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            (ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style
            cell_value = row[col_num]
            if col_num == 6 and isinstance(cell_value, datetime):
                cell_value = cell_value.replace(tzinfo=None)  # Remover información de zona horaria
            (ws.cell(row = row_num, column = col_num+1, value=str(cell_value))).style = body_style
            if col_num == 6:
                (ws.cell(row = row_num, column = col_num+1, value=cell_value)).style = date_style
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)


def convert_excel_inventario_xlsxwriter(existencia, valor_inventario, dict_entradas, dict_resultados):
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet('Inventario')

    # Definir los estilos antes de usarlos
    head_style = workbook.add_format({'bold': True, 'font_color': 'white', 'bg_color': '333366', 'font_name': 'Arial', 'font_size': 11})
    body_style = workbook.add_format({'font_name': 'Calibri', 'font_size': 10})
    money_style = workbook.add_format({'num_format': '$ #,##0.00', 'font_name': 'Calibri', 'font_size': 10})
    money_resumen_style = workbook.add_format({'num_format': '$ #,##0.00', 'font_name': 'Calibri', 'font_size': 14, 'bold': True})
    date_style = workbook.add_format({'num_format': 'dd/mm/yyyy', 'font_name': 'Calibri', 'font_size': 10})

    # Definir las columnas antes de utilizar la variable `columns`
    columns = ['Código', 'Producto', 'Distrito', 'Unidad', 'Cantidad', 'Cantidad Apartada', 'Minimos', 'Ubicación', 'Estante', 'Precio', 'Total']
    
    # Escribir el encabezado con los estilos definidos
    #worksheet.write_row('A1', columns, head_style)

    # Establecer los anchos de las columnas después de definir `columns`
    worksheet.set_column('A:A', 10)
    worksheet.set_column('B:B', 30)
    for i, column in enumerate(columns):
        worksheet.write(0, i, column, head_style)
        worksheet.set_column(i, i, 15)  # Ajusta el ancho de las columnas

    # Escribir los datos
    row_num = 0
    for inventario in existencia:
        row_num += 1
        #inventario.total_entradas = dict_entradas.get(inventario.id, 0)
        inventario.total_apartado = dict_resultados.get(inventario.id, 0)
        #total_value = (inventario.cantidad + inventario.total_apartado) * inventario.price
    
        row = [
            inventario.producto.codigo, #A
            inventario.producto.nombre, #B
            inventario.distrito.nombre, #C
            inventario.producto.unidad.nombre, #D
            inventario.cantidad, #E
            inventario.total_apartado, #F
            #inventario.total_entradas,
            inventario.minimo, #G
            inventario.ubicacion, #H
            inventario.estante, #I
        ]
    
        for col_num, item in enumerate(row, start=1):  # Enumerate empieza con 1 para A1, ajusta según sea necesario
            worksheet.write(row_num, col_num - 1, item, body_style)
    
        # Ahora escribe los valores con formateo especial directamente
        worksheet.write(row_num, 9, inventario.price, money_style)  # Columna 10 (J) para el precio
        worksheet.write_formula(row_num, 10, f"=(E{row_num+1}+F{row_num+1})*J{row_num+1}", money_style)
       #worksheet.write(row_num, 10, total_value, money_style)  # Columna 11 (K) para el valor total

    # Escribir el total del inventario
    worksheet.set_column('N:N', 30)
    worksheet.set_column('O:O', 30)
    worksheet.write('N2', 'Inventario Costo Total:', head_style)
    worksheet.write('O2', valor_inventario, money_resumen_style)

    workbook.close()
    output.seek(0)

    response = HttpResponse(
        output.read(), 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
      # Establecer una cookie para indicar que la descarga ha iniciado
    response.set_cookie('iniciada', 'true', max_age=20)  # La cookie expira en 20 segundos
    

    
    response['Content-Disposition'] = f'attachment; filename=Inventario_{dt.date.today()}.xlsx'
    output.close()
    return response



@login_required(login_url='user-login')
@perfil_seleccionado_required
def analisis_rotacion(request):
    pk = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk)

    hoy = timezone.localdate()

    fecha_inicio_texto = request.GET.get('fecha_inicio',hoy.replace(day=1).isoformat(),)
    fecha_final_texto = request.GET.get('fecha_final',hoy.isoformat(),)
    distrito_id = request.GET.get('distrito')

    resultados = []
    error = None
    distrito_seleccionado = None
    almacenes_distritos = (
        perfil.almacen
            .values_list(
            'distrito_id',
            flat=True,
        )
        .distinct()
    )
    # Sustituir este queryset por el filtro de distritos
    # que el usuario tenga autorizado.
    distritos =  (
        Distrito.objects
        .filter(id__in=almacenes_distritos)
        .order_by('nombre')
    )

    try:
        fecha_inicio = date.fromisoformat(fecha_inicio_texto)
        fecha_final = date.fromisoformat(fecha_final_texto)
    except ValueError:
        fecha_inicio = hoy.replace(day=1)
        fecha_final = hoy
        error = 'Las fechas proporcionadas no son válidas.'

    if fecha_inicio > fecha_final:
        error = 'La fecha final no puede ser anterior a la inicial.'

    if fecha_final > hoy:
        error = 'La fecha final no puede ser posterior al día actual.'

    if distrito_id and not error:
        try:
            distrito_seleccionado = distritos.get(id=distrito_id)
        except Distrito.DoesNotExist:
            error = 'El distrito seleccionado no es válido.'

    if distrito_seleccionado and not error:
        inicio_periodo = timezone.make_aware(
            datetime.combine(fecha_inicio, time.min)
        )

        fin_exclusivo = timezone.make_aware(datetime.combine(fecha_final + timedelta(days=1),time.min,))

        ahora = timezone.now()

        # Si el reporte termina hoy, el corte final será
        # la hora actual y no el final de un día aún incompleto.
        corte_final = min(fin_exclusivo, ahora)

        inventarios = list(
            Inventario.objects
            .filter(
                distrito=distrito_seleccionado,
                complete=True,
                producto__servicio = False
            )
            .select_related('producto', 'distrito')
            .order_by('producto')
        )

        inventario_ids = [
            inventario.id
            for inventario in inventarios
        ]

        # --------------------------------------------------
        # Cantidad apartada actual
        # --------------------------------------------------

        apartados_query = (
            ArticulosparaSurtir.objects
            .filter(
                articulos__producto_id__in=inventario_ids,
                surtir=True,
            )
            .values(
                inventario_id=F(
                    'articulos__producto_id'
                )
            )
            .annotate(
                total_apartado=Sum('cantidad')
            )
        )

        apartados = {
            fila['inventario_id']:
                fila['total_apartado'] or decimal.Decimal('0')
            for fila in apartados_query
        }

        # --------------------------------------------------
        # Entradas posteriores a los cortes
        # --------------------------------------------------

        ruta_inventario_entrada = (
            'articulo_comprado__producto__producto'
            '__articulos__producto_id'
        )

        entradas_query = (
            EntradaArticulo.objects
            .filter(
                **{
                    f'{ruta_inventario_entrada}__in':
                        inventario_ids,
                },
                entrada__completo=True,
                entrada__cancelada=False,
                entrada__entrada_date__gte=inicio_periodo,
                entrada__entrada_date__lte=ahora,
            )
            .values(
                inventario_id=F(
                    ruta_inventario_entrada
                )
            )
            .annotate(
                posteriores_inicio=Sum('cantidad'),
                posteriores_final=Sum(
                    'cantidad',
                    filter=Q(
                        entrada__entrada_date__gte=corte_final
                    ),
                ),
            )
        )

        entradas = {
            fila['inventario_id']: {
                'posteriores_inicio':
                    fila['posteriores_inicio']
                    or decimal.Decimal('0'),
                'posteriores_final':
                    fila['posteriores_final']
                    or decimal.Decimal('0'),
            }
            for fila in entradas_query
        }

        # --------------------------------------------------
        # Salidas del periodo y posteriores a los cortes
        # --------------------------------------------------

        salidas_query = (
            Salidas.objects
            .filter(
                producto__articulos__producto_id__in=
                    inventario_ids,
                complete=True,
                vale_salida__cancelada=False,
                created_at__gte=inicio_periodo,
                created_at__lte=ahora,
            )
            .values(
                inventario_id=F(
                    'producto__articulos__producto_id'
                )
            )
            .annotate(
                posteriores_inicio=Sum('cantidad'),
                posteriores_final=Sum(
                    'cantidad',
                    filter=Q(
                        created_at__gte=corte_final
                    ),
                ),
                salidas_periodo=Sum(
                    'cantidad',
                    filter=Q(
                        created_at__lt=corte_final
                    ),
                ),
            )
        )

        salidas = {
            fila['inventario_id']: {
                'posteriores_inicio':
                    fila['posteriores_inicio']
                    or decimal.Decimal('0'),
                'posteriores_final':
                    fila['posteriores_final']
                    or decimal.Decimal('0'),
                'salidas_periodo':
                    fila['salidas_periodo']
                    or decimal.Decimal('0'),
            }
            for fila in salidas_query
        }

        # --------------------------------------------------
        # Construcción de resultados
        # --------------------------------------------------

        for inventario in inventarios:
            apartado_actual = apartados.get(
                inventario.id,
                decimal.Decimal('0'),
            )

            inventario_fisico_actual = (
                inventario.cantidad
                + apartado_actual
            )

            datos_entradas = entradas.get(
                inventario.id,
                {},
            )
            datos_salidas = salidas.get(
                inventario.id,
                {},
            )

            entradas_posteriores_inicio = (
                datos_entradas.get(
                    'posteriores_inicio',
                    decimal.Decimal('0'),
                )
            )
            entradas_posteriores_final = (
                datos_entradas.get(
                    'posteriores_final',
                    decimal.Decimal('0'),
                )
            )

            salidas_posteriores_inicio = (
                datos_salidas.get(
                    'posteriores_inicio',
                    decimal.Decimal('0'),
                )
            )
            salidas_posteriores_final = (
                datos_salidas.get(
                    'posteriores_final',
                    decimal.Decimal('0'),
                )
            )

            cantidad_salidas_periodo = (
                datos_salidas.get(
                    'salidas_periodo',
                    decimal.Decimal('0'),
                )
            )
           
           
            inventario_inicial = (
                inventario_fisico_actual
                - entradas_posteriores_inicio
                + salidas_posteriores_inicio
            )

            inventario_final = (
                inventario_fisico_actual
                - entradas_posteriores_final
                + salidas_posteriores_final
            )

            inventario_promedio = (
                inventario_inicial
                + inventario_final
            ) / decimal.Decimal('2')

            if inventario_promedio > 0:
                rotacion = (
                    cantidad_salidas_periodo
                    / inventario_promedio
                )
            else:
                rotacion = None

            precio_unitario = (
                inventario.price
                or decimal.Decimal('0')
            )

            importe_promedio = (
                inventario_promedio
                * precio_unitario
            )

            resultados.append({
                'inventario': inventario,
                'cantidad_disponible':
                    inventario.cantidad,
                'cantidad_apartada':
                    apartado_actual,
                'inventario_fisico_actual':
                    inventario_fisico_actual,

                'entradas_posteriores':
                    entradas_posteriores_inicio,
                'salidas_posteriores':
                    salidas_posteriores_inicio,

                'inventario_inicial':
                    inventario_inicial,
                'inventario_final':
                    inventario_final,
                'inventario_promedio':
                    inventario_promedio,
                'cantidad_salidas':
                    cantidad_salidas_periodo,
                'rotacion': rotacion,

                'precio_unitario': precio_unitario,
                'importe_promedio': importe_promedio,
            })

        # Mayor rotación primero; los productos sin base quedan al final.
        resultados.sort(
            key=lambda fila: (
                fila['rotacion'] is not None,
                fila['rotacion'] or decimal.Decimal('0'),
            ),
            reverse=True,
        )

    if request.GET.get('exportar') == 'excel':
        return exportar_rotacion_excel(resultados=resultados, distrito=distrito_seleccionado, fecha_inicio=fecha_inicio, fecha_final=fecha_final,)

    paginator = Paginator(resultados, 50)
    pagina = paginator.get_page(request.GET.get('page'))

    parametros = request.GET.copy()
    parametros.pop('page', None)

    context = {
        'distritos': distritos,
        'distrito_seleccionado':
            distrito_seleccionado,
        'fecha_inicio': fecha_inicio,
        'fecha_final': fecha_final,
        'pagina': pagina,
        'error': error,
        'querystring': parametros.urlencode(),
    }

    return render(request,'dashboard/analisis_rotacion.html',context,)


def exportar_rotacion_excel(resultados,distrito,fecha_inicio,fecha_final,):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Rotación'

    color_encabezado = '1F4E78'
    color_texto = 'FFFFFF'

    # Título
    worksheet.merge_cells('A1:M1')
    titulo = worksheet['A1']
    titulo.value = 'Análisis de rotación de inventario'
    titulo.font = Font(
        bold=True,
        size=14,
        color=color_texto,
    )
    titulo.fill = PatternFill(
        fill_type='solid',
        fgColor=color_encabezado,
    )
    titulo.alignment = Alignment(
        horizontal='center',
        vertical='center',
    )

    worksheet.row_dimensions[1].height = 24

    # Datos del reporte
    worksheet['A2'] = 'Almacén'
    worksheet['B2'] = str(distrito)

    worksheet['A3'] = 'Periodo'
    worksheet['B3'] = fecha_inicio
    worksheet['C3'] = fecha_final

    worksheet['B3'].number_format = 'dd/mm/yyyy'
    worksheet['C3'].number_format = 'dd/mm/yyyy'

    encabezados = [
        'Producto',
        'Precio unitario',
        'Disponible actual',
        'Apartado actual',
        'Físico actual',
        'Entradas posteriores',
        'Salidas posteriores',
        'Inventario inicial',
        'Inventario final',
        'Inventario promedio',
        'Valor inventario promedio',
        'Salidas del periodo',
        'Rotación',
    ]

    fila_encabezado = 5

    for numero_columna, encabezado in enumerate(
        encabezados,
        start=1,
    ):
        celda = worksheet.cell(
            row=fila_encabezado,
            column=numero_columna,
            value=encabezado,
        )

        celda.font = Font(
            bold=True,
            color=color_texto,
        )
        celda.fill = PatternFill(
            fill_type='solid',
            fgColor=color_encabezado,
        )
        celda.alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True,
        )

    worksheet.row_dimensions[fila_encabezado].height = 35

    # Resultados
    fila_inicial = fila_encabezado + 1

    for numero_fila, resultado in enumerate(
        resultados,
        start=fila_inicial,
    ):
        valores = [
            str(resultado['inventario'].producto),
            resultado['precio_unitario'],
            resultado['cantidad_disponible'],
            resultado['cantidad_apartada'],
            resultado['inventario_fisico_actual'],
            resultado['entradas_posteriores'],
            resultado['salidas_posteriores'],
            resultado['inventario_inicial'],
            resultado['inventario_final'],
            resultado['inventario_promedio'],
            resultado['importe_promedio'],
            resultado['cantidad_salidas'],
            resultado['rotacion'],
        ]

        for numero_columna, valor in enumerate(
            valores,
            start=1,
        ):
            celda = worksheet.cell(
                row=numero_fila,
                column=numero_columna,
                value=valor,
            )

            if numero_columna > 1:
                celda.alignment = Alignment(
                    horizontal='right'
                )

    ultima_fila = worksheet.max_row

    formato_cantidad = '#,##0.00'
    formato_moneda = '"$"#,##0.00'
    formato_rotacion = '0.00'

    columnas_cantidad = [
        'C', 'D', 'E', 'F', 'G',
        'H', 'I', 'J', 'L',
    ]

    for columna in columnas_cantidad:
        for fila in range(fila_inicial, ultima_fila + 1):
            worksheet[f'{columna}{fila}'].number_format = (
                formato_cantidad
            )

    for columna in ['B', 'K']:
        for fila in range(fila_inicial, ultima_fila + 1):
            worksheet[f'{columna}{fila}'].number_format = (
                formato_moneda
            )

    for fila in range(fila_inicial, ultima_fila + 1):
        worksheet[f'M{fila}'].number_format = formato_rotacion

    # Congelar encabezado y activar filtros
    worksheet.freeze_panes = 'A6'
    worksheet.auto_filter.ref = (
        f'A5:M{ultima_fila}'
    )

    anchos = {
        'A': 55,
        'B': 16,
        'C': 16,
        'D': 16,
        'E': 16,
        'F': 18,
        'G': 18,
        'H': 18,
        'I': 18,
        'J': 18,
        'K': 22,
        'L': 18,
        'M': 12,
    }

    for columna, ancho in anchos.items():
        worksheet.column_dimensions[columna].width = ancho

    nombre_distrito = (
        slugify(str(distrito))
        or 'almacen'
    )

    nombre_archivo = (
        f'rotacion_{nombre_distrito}_'
        f'{fecha_inicio:%Y%m%d}_'
        f'{fecha_final:%Y%m%d}.xlsx'
    )

    response = HttpResponse(
        content_type=(
            'application/vnd.openxmlformats-'
            'officedocument.spreadsheetml.sheet'
        )
    )

    response['Content-Disposition'] = (
        f'attachment; filename="{nombre_archivo}"'
    )

    workbook.save(response)

    return response