from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, Http404, JsonResponse, FileResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.utils import translation
from django.urls import reverse
from django.conf import settings
from django.forms import inlineformset_factory
from django.utils.http import urlencode
from django.db import IntegrityError
from django.db.models import Sum, Q, Prefetch, Avg, FloatField, Case, When, F,DecimalField, ExpressionWrapper, Max
from .models import Product, Subfamilia, Order, Products_Batch, Familia, Unidad, Inventario, Producto_Calidad, Requerimiento_Calidad, PriceRefChange
from compras.models import Proveedor, Proveedor_Batch, Proveedor_Direcciones_Batch, Proveedor_direcciones, Estatus_proveedor, Estado, DocumentosProveedor, Debida_Diligencia
from solicitudes.models import Subproyecto, Proyecto, Contrato, Status_Contrato
from requisiciones.models import Salidas, ValeSalidas
from user.models import Profile, Distrito, Banco
from .forms import ProductForm, Products_BatchForm, AddProduct_Form, Proyectos_Form, ProveedoresForm, Proyectos_Add_Form, Proveedores_BatchForm, ProveedoresDireccionesForm, Proveedores_Direcciones_BatchForm, Subproyectos_Add_Form, ProveedoresExistDireccionesForm, Add_ProveedoresDireccionesForm, DireccionComparativoForm, Profile_Form, PrecioRef_Form
from .forms import RequerimientoCalidadForm, Add_Product_CriticoForm, Add_ProveedoresDir_Alt_Form, Comentario_Proveedor_Doc_Form, Contrato_form
from user.decorators import perfil_seleccionado_required, tipo_usuario_requerido
from .filters import ProductFilter, ProyectoFilter, ProveedorFilter, SubproyectoFilter, ProductCalidadFilter, ContratoFilter, PriceRefChangeFilter
from user.filters import ProfileFilter
from proveedores_externos.views import extraer_tipo_contribuyente
import csv
from django.core.paginator import Paginator
from datetime import date, datetime
import plotly.express as px
from plotly.subplots import make_subplots
import plotly.graph_objects as go
import pandas as pd
import io
import os
#import decimal
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill
from openpyxl.utils import get_column_letter
import datetime as dt
import json
import csv
from charset_normalizer import detect

#PDF generator
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.colors import Color, black, blue, red, white
from reportlab.lib.units import cm
from reportlab.lib.pagesizes import letter
from reportlab.rl_config import defaultPageSize

from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Frame, PageBreak
from bs4 import BeautifulSoup

# Create your views here.
@login_required(login_url='user-login')
@perfil_seleccionado_required
def index(request):
    pk = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk)
    inventarios = Inventario.objects.all()
    proyectos = Proyecto.objects.all()
    proveedor = usuario.proveedor
    mostrar_modal = True
    prealta = False
    if proveedor:
        mostrar_modal = (
            not proveedor.acepto_politica or
            not proveedor.acepto_politica_proveedor or
            not proveedor.acepto_codigo_etica or
            not proveedor.acepto_aviso_privacidad
        )
        if proveedor.direcciones.filter(estatus__nombre = "PREALTA").exists():
            prealta = True

    #print(prealta)
    # Obtener los proyectos y calcular el total
    #proyectos_total = [(proyecto, proyecto.get_projects_gastado) for proyecto in proyectos]

    # Obtener los inventarios y calcular el costo de salidas
    #inventarios_costo_salidas = [(inventario, inventario.costo_salidas) for inventario in inventarios]

    # Ordenar los inventarios por el costo de salidas en orden descendente
    #inventarios_costo_salidas_sorted = sorted(inventarios_costo_salidas, key=lambda x: x[1], reverse=True)
    # Ordenar los proyectos por el total en orden descendente
    #proyectos_total_sorted = sorted(proyectos_total, key=lambda x: x[1], reverse=True)

    # Tomar solo los primeros 50 inventarios ordenados
    #inventarios_top_50 = inventarios_costo_salidas_sorted[:50]


    # Preparar los datos para el gráfico
    #x = [proyecto.nombre for proyecto, _ in proyectos_total_sorted]
    #y = [total for _, total in proyectos_total_sorted]
    #x2 = [inventario.producto.nombre[:15] + '...' if len(inventario.producto.nombre) > 10 else inventario.producto.nombre for inventario,_ in inventarios_top_50]
    #y2 = [costo_salidas for _, costo_salidas in inventarios_top_50]

   # Crear el gráfico de barras
    #fig = make_subplots()
    #fig.add_trace(go.Bar(x=x, y=y, marker=dict(color='#3E92CC')),1,1)
    # Crear el gráfico de barras
    #fig2 = make_subplots()
    #fig2.add_trace(go.Bar(x=x2, y=y2, marker=dict(color='#3E92CC')),1,1)

    #fig.update_layout(
    #    plot_bgcolor='#9a9b9d',
    #    paper_bgcolor='white',
    #    font_color= '#3E92CC',
    #    )

    #fig2.update_layout(
    #    plot_bgcolor='#9a9b9d',
    #    paper_bgcolor='white',
    #    font_color= '#3E92CC',
    #    )

    #Convertir el gráfico en HTML para pasar a la plantilla
    #graph_proyectos = fig.to_html(full_html=False)
    #graph_inventarios = fig2.to_html(full_html=False)

    context = {
        'mostrar_modal': mostrar_modal,
        'prealta': prealta,
        #'select_profile':selected_profil
        #'graph_proyectos': graph_proyectos,
        #'graph_inventarios':graph_inventarios,
        }
    
    return render(request,'dashboard/index.html',context)


@login_required(login_url='user-login')
def select_profile(request):
    user = request.user.id

    profiles = Profile.objects.filter( Q(staff__staff__id=user) & Q(sustituto__isnull = True) & Q(st_activo = True)| Q(sustituto__staff__id=user))
    
    if request.method == 'POST':
        profile_id = request.POST.get('profile')
        try:
            profile = Profile.objects.get(id=profile_id)
            request.session['selected_profile_id'] = profile.id
            # **Cambiar idioma según el perfil seleccionado**
            if profile.distritos.nombre == "BRASIL":
               
                translation.activate('pt')
                request.session[settings.LANGUAGE_SESSION_KEY] = 'pt-br'
                print(request.session[settings.LANGUAGE_SESSION_KEY])
            else:
                translation.activate('es')
                request.session[settings.LANGUAGE_SESSION_KEY] = 'es-MX'
            return redirect('dashboard-index')   
        except Profile.DoesNotExist:
            messages.error(request, 'El perfil seleccionado no es válido')
    else:
        # En el GET, muestra el formulario con los perfiles disponibles
        form = Profile_Form()
        form.fields['profile'].queryset = profiles

    request.LANGUAGE_CODE = translation.get_language()  # 🔹 Forzar el idioma actual
    print(f"Idioma activado: {request.LANGUAGE_CODE}")  # Verificar en la consola
        
    context = {
        'form': form,
    }
    return render(request, 'dashboard/select_profile.html', context)


@perfil_seleccionado_required
def proyectos(request):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)

    sql_salidas = """SELECT 
    solicitudes_proyecto.id AS id,
    solicitudes_proyecto.nombre AS nombre,
    SUM(requisiciones_salidas.cantidad * requisiciones_salidas.precio) AS total_salidas
    FROM
        solicitudes_proyecto
    JOIN
        dashboard_order ON solicitudes_proyecto.id = dashboard_order.proyecto_id
    JOIN
        dashboard_articulosordenados ON dashboard_order.id = dashboard_articulosordenados.orden_id
    JOIN
        dashboard_articulosparasurtir ON dashboard_articulosordenados.id = dashboard_articulosparasurtir.articulos_id
    JOIN
        requisiciones_salidas ON dashboard_articulosparasurtir.id = requisiciones_salidas.producto_id
    GROUP BY
        id, nombre
    ORDER BY
        id;
    """


    sql_gastos_pagados ="""SELECT 
    solicitudes_proyecto.id AS id,
    solicitudes_proyecto.nombre AS nombre,
    SUM(
        (gastos_articulo_gasto.cantidad * gastos_articulo_gasto.precio_unitario * 1.16) + 
        COALESCE(gastos_articulo_gasto.otros_impuestos, 0) - COALESCE(gastos_articulo_gasto.impuestos_retenidos, 0)
    ) AS total_pagado
    FROM
        solicitudes_proyecto
    JOIN
        gastos_articulo_gasto ON solicitudes_proyecto.id = gastos_articulo_gasto.proyecto_id
    LEFT JOIN
        tesoreria_pago ON gastos_articulo_gasto.id = tesoreria_pago.gasto_id
    WHERE
        tesoreria_pago.hecho = true
    GROUP BY
        id, nombre
    ORDER BY
        id;
    """


    sql_compras_pagos = """SELECT 
    solicitudes_proyecto.id  AS id,
    solicitudes_proyecto.nombre AS nombre,
        SUM((CASE 
                    -- Cuando la moneda es PESOS
                    WHEN cuenta_moneda.nombre = 'PESOS' THEN tesoreria_pago.monto
                    -- Cuando la moneda es DÓLARES
                  
					WHEN cuenta_moneda.nombre = 'DOLARES' THEN
						CASE 
							-- Si tiene tipo de cambio en tesoreria_pago
							WHEN tesoreria_pago.tipo_de_cambio IS NOT NULL THEN tesoreria_pago.monto * tesoreria_pago.tipo_de_cambio
							-- Usar 17 como valor predeterminado si no tiene tipo de cambio
							WHEN tesoreria_pago.tipo_de_cambio IS NULL AND compras_compra.tipo_de_cambio IS NOT NULL THEN tesoreria_pago.monto * compras_compra.tipo_de_cambio 
							ELSE tesoreria_pago.monto * 17
						END
			END
    )) AS total_pagado
    FROM
        solicitudes_proyecto
    JOIN
        dashboard_order ON solicitudes_proyecto.id = dashboard_order.proyecto_id
    JOIN 
        requisiciones_requis ON dashboard_order.id = requisiciones_requis.orden_id
    JOIN
        compras_compra ON requisiciones_requis.id = compras_compra.req_id
    JOIN
        compras_moneda ON compras_compra.moneda_id = compras_moneda.id
    LEFT JOIN 
        tesoreria_pago ON compras_compra.id = tesoreria_pago.oc_id
    LEFT JOIN
        tesoreria_cuenta ON tesoreria_pago.cuenta_id = tesoreria_cuenta.id
    LEFT JOIN
        compras_moneda AS cuenta_moneda ON tesoreria_cuenta.moneda_id = cuenta_moneda.id -- Utilizando el alias cuenta_moneda
    WHERE
        tesoreria_pago.hecho = True
    GROUP BY
        id, nombre
    ORDER BY
        proyecto_id;"""

    sql_compras = """SELECT 
	solicitudes_proyecto.id  AS id,
    solicitudes_proyecto.nombre AS nombre,
    SUM(
        CASE 
            WHEN compras_moneda.nombre = 'DOLARES' AND pagos_promedio.avg_tipo_de_cambio IS NOT NULL THEN compras_compra.costo_oc * pagos_promedio.avg_tipo_de_cambio
            WHEN compras_moneda.nombre = 'DOLARES' AND pagos_promedio.avg_tipo_de_cambio IS NULL AND compras_compra.tipo_de_cambio IS NOT NULL THEN compras_compra.costo_oc * compras_compra.tipo_de_cambio
            WHEN compras_moneda.nombre = 'DOLARES' AND pagos_promedio.avg_tipo_de_cambio IS NULL AND compras_compra.tipo_de_cambio IS NULL THEN compras_compra.costo_oc * 17
            ELSE compras_compra.costo_oc
        END
    ) AS total_costo_oc
    FROM
	    solicitudes_proyecto
    JOIN
	    dashboard_order ON solicitudes_proyecto.id = dashboard_order.proyecto_id
    JOIN 
	    requisiciones_requis ON dashboard_order.id = requisiciones_requis.orden_id
    JOIN
	    compras_compra ON requisiciones_requis.id = compras_compra.req_id
    JOIN
	    compras_moneda ON compras_compra.moneda_id = compras_moneda.id
    LEFT JOIN (
	    SELECT oc_id, AVG(tipo_de_cambio) AS avg_tipo_de_cambio
        FROM tesoreria_pago
	    group by oc_id
    ) AS pagos_promedio ON compras_compra.id = pagos_promedio.oc_id
    group by
	    id, nombre
    ORDER BY
	    id;
       """
   
     # Prefetching related data
    proyectos = Proyecto.objects.filter(distrito = usuario.distritos)
    proyecto_compras_total = proyectos.raw(sql_compras)
    proyecto_pagos_total = proyectos.raw(sql_compras_pagos)
    proyecto_gastos_total = proyectos.raw(sql_gastos_pagados)
    proyectos_salidas = proyectos.raw(sql_salidas)
    dict_compras = {r.id: r.total_costo_oc for r in proyecto_compras_total}
    dict_pagos = {r.id: r.total_pagado for r in proyecto_pagos_total}
    dict_gastos = {r.id: r.total_pagado for r in proyecto_gastos_total}
    dict_salidas = {r.id: r.total_salidas for r in proyectos_salidas}

    myfilter=ProyectoFilter(request.GET, queryset=proyectos)
    proyectos = myfilter.qs

    if request.method == 'POST' and 'btnReporte' in request.POST:
        proyectos_completos = asignar_totales(proyectos, dict_compras, dict_pagos, dict_gastos, dict_salidas)
        return convert_excel_matriz_proyectos(proyectos_completos)

    #Set up pagination
    p = Paginator(proyectos, 10)
    page = request.GET.get('page')
    proyectos_list = p.get_page(page)
    
    proyectos_paginados = asignar_totales(proyectos_list, dict_compras, dict_pagos, dict_gastos, dict_salidas)

    
    
    context = {
        'proyectos':proyectos,
        'proyectos_list':proyectos_list,
        'myfilter':myfilter,
        }
    
    return render(request,'dashboard/proyectos.html',context)

@perfil_seleccionado_required
def contratos(request):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    contratos = Contrato.objects.all()

    myfilter= ContratoFilter(request.GET, queryset=contratos)

    #Set up pagination
    p = Paginator(contratos, 10)
    page = request.GET.get('page')
    contratos_list = p.get_page(page)

    context = {
        'contratos':contratos,
        'myfilter': myfilter,
        'contratos_list': contratos_list,
         }

    return render(request,'dashboard/contratos.html', context)

@perfil_seleccionado_required
def proyectos_contrato(request, pk):
    pk_profile = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_profile)
    contrato = Contrato.objects.get(id = pk)
    proyectos = Proyecto.objects.filter(contrato = contrato)
    myfilter=ProyectoFilter(request.GET, queryset=proyectos)
    proyectos = myfilter.qs

    #Set up pagination
    p = Paginator(proyectos, 10)
    page = request.GET.get('page')
    proyectos_list = p.get_page(page)

    context = {
        'myfilter': myfilter,
        'proyectos_list': proyectos_list,
        'proyectos':proyectos,
        }

    return render(request,'dashboard/proyectos.html', context)

@login_required(login_url='user-login')
@perfil_seleccionado_required
def add_contratos(request):
    #usuario = Profile.objects.get(staff=request.user
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    #distrito = usuario.distritos

    form = Contrato_form()

    if request.method =='POST':
        contrato, created = Contrato.objects.get_or_create(complete = False)
        form = Contrato_form(request.POST, instance = contrato)
        if form.is_valid():
            contrato = form.save(commit=False)
            tipo_activo = Status_Contrato.objects.get(nombre = "Activo")
            contrato.status_contrato = tipo_activo
            contrato.created_at = date.today()
            contrato.created_by = usuario
            contrato.complete = True
            contrato.save()
            messages.success(request,'Has agregado correctamente el Contrato')
            return redirect('configuracion-contratos')
    else:
        form = Contrato_form()

    context = {
        'form': form,
        }

    return render(request,'dashboard/add_contratos.html',context)


def asignar_totales(proyectos_queryset, dict_compras, dict_pagos, dict_gastos, dict_salidas):
    for proyecto in proyectos_queryset:
        proyecto.total_compras = dict_compras.get(proyecto.id, 0)
        proyecto.total_pagos = dict_pagos.get(proyecto.id, 0)
        proyecto.total_gastos = dict_gastos.get(proyecto.id, 0)
        proyecto.total_salidas = dict_salidas.get(proyecto.id, 0)
    return proyectos_queryset

@login_required(login_url='user-login')
@perfil_seleccionado_required
def subproyectos(request, pk):
    proyecto = Proyecto.objects.get(id=pk)
    subproyectos = Subproyecto.objects.filter(proyecto=proyecto)

    myfilter=SubproyectoFilter(request.GET, queryset=subproyectos)
    subproyectos = myfilter.qs

    #Set up pagination
    p = Paginator(subproyectos, 50)
    page = request.GET.get('page')
    subproyectos_list = p.get_page(page)

    context = {
        'proyecto':proyecto,
        'subproyectos':subproyectos,
        'subproyectos_list':subproyectos_list,
        'myfilter':myfilter,
        }

    return render(request,'dashboard/subproyectos.html',context)



@login_required(login_url='user-login')
@perfil_seleccionado_required
def proyectos_edit(request, pk):

    proyecto = Proyecto.objects.get(id=pk)

    if request.method =='POST':
        form = Proyectos_Form(request.POST, instance=proyecto)
        if form.is_valid():
            form.save()
            messages.success(request,f'Has actualizado correctamente el proyecto {proyecto.nombre}')
            return redirect('configuracion-proyectos')
    else:
        form = Proyectos_Form(instance=proyecto)


    context = {
        'form': form,
        'proyecto':proyecto,
        }
    return render(request,'dashboard/proyectos_edit.html', context)

@login_required(login_url='user-login')
@perfil_seleccionado_required
def proveedor_direcciones(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    almacenes_distritos = set(usuario.almacen.values_list('distrito__id', flat=True))
    razon = request.GET.get('razon_social', '')
    rfc = request.GET.get('rfc', '')
    next_url = request.GET.get('next') or request.POST.get('next')
    #print(almacenes_distritos)
    if usuario.tipo.proveedores:
        proveedor = Proveedor.objects.get(id=pk)
        if usuario.tipo.nombre == "Subdirector_Alt":
            direcciones = Proveedor_direcciones.objects.filter(nombre__id=pk, completo = True, distrito__id= 8)
        else:
            direcciones = Proveedor_direcciones.objects.filter(nombre__id=pk, completo = True, distrito__id__in = almacenes_distritos)
    else:
        raise Http404("No tienes permiso para ver esta vista")
    context = {
        'next': next_url,
        'proveedor':proveedor,
        'direcciones':direcciones,
        'razon': razon,
        'rfc': rfc,
        }
    return render(request,'dashboard/direcciones_proveedor.html', context)

@login_required(login_url='user-login')
@perfil_seleccionado_required
def proyectos_add(request):
    #usuario = Profile.objects.get(staff=request.user
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    distrito = usuario.distritos

    form = Proyectos_Add_Form()

    if request.method =='POST':
        proyecto, created = Proyecto.objects.get_or_create(distrito = distrito, complete = False)
        form = Proyectos_Add_Form(request.POST, instance = proyecto)
        if form.is_valid():
            proyecto = form.save(commit=False)
            proyecto.activo = True
            proyecto.complete = True
            proyecto.save()
            messages.success(request,'Has agregado correctamente el proyecto')
            return redirect('configuracion-proyectos')
    else:
        form = Proyectos_Add_Form()

    context = {
        'form': form,
        }

    return render(request,'dashboard/proyectos_add.html',context)

@login_required(login_url='user-login')
@perfil_seleccionado_required
def subproyectos_add(request, pk):
    proyecto = Proyecto.objects.get(id=pk)
    form = Subproyectos_Add_Form()

    if request.method =='POST':
        form = Subproyectos_Add_Form(request.POST)
        if form.is_valid():
            subproyecto = form.save(commit=False)
            subproyecto.proyecto = proyecto
            subproyecto.save()
            messages.success(request,'Has agregado correctamente el subproyecto')
            return redirect('subproyectos', pk=proyecto.id)
    else:
        form = Subproyectos_Add_Form()

    context = {
        'form': form,
        'proyecto':proyecto,
        }

    return render(request,'dashboard/subproyectos_add.html',context)

@login_required(login_url='user-login')
@perfil_seleccionado_required
def subproyectos_edit(request, pk):
    subproyecto = Subproyecto.objects.get(id=pk)
    proyecto = Proyecto.objects.get(id=subproyecto.proyecto.id)
    form = Subproyectos_Add_Form(instance=subproyecto)

    if request.method =='POST':
        form = Subproyectos_Add_Form(request.POST, instance=subproyecto)
        if form.is_valid():
            form.save()
            messages.success(request,'Has editado correctamente el subproyecto')
            return redirect('subproyectos', pk=subproyecto.proyecto.id)
    else:
        form = Subproyectos_Add_Form(instance=subproyecto)

    context = {
        'form': form,
        'proyecto':proyecto,
        }

    return render(request,'dashboard/subproyectos_add.html',context)


@login_required(login_url='user-login')
@perfil_seleccionado_required
def staff(request):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)

    perfiles = Profile.objects.filter(staff__staff__is_active = True, sustituto__isnull=True, distritos = usuario.distritos)
    cuenta_perfiles = perfiles.count()

    myfilter = ProfileFilter(request.GET, queryset=perfiles)
    perfiles = myfilter.qs
    cuenta_filtrados = perfiles.count()

    #Set up pagination
    p = Paginator(perfiles, 30)
    page = request.GET.get('page')
    registros_list = p.get_page(page)

    context = {
        'registros_list':registros_list,
        'myfilter':myfilter,
        'cuenta_perfiles':cuenta_perfiles,
        'cuenta_filtrados':cuenta_filtrados,
        }
    return render(request,'dashboard/staff.html', context)

@login_required(login_url='user-login')
@perfil_seleccionado_required
def product(request):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    items = Product.objects.filter(completado = True).order_by('codigo')

    myfilter=ProductFilter(request.GET, queryset=items)
    items = myfilter.qs

    #Set up pagination
    p = Paginator(items, 50)
    page = request.GET.get('page')
    items_list = p.get_page(page)

    context = {
        'usuario':usuario,
        'items': items,
        'myfilter':myfilter,
        'items_list':items_list,
        }


    return render(request,'dashboard/product.html', context)


@perfil_seleccionado_required
def proveedores(request):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    almacenes_distritos = set(usuario.almacen.values_list('distrito__id', flat=True))
    # Obtén los IDs de los proveedores que cumplan con las condiciones deseadas
    proveedores_dir = Proveedor_direcciones.objects.filter(distrito__id__in = almacenes_distritos)
    proveedores_ids = proveedores_dir.values_list('nombre', flat=True).distinct()
    
    almacenes_distritos = set(usuario.almacen.values_list('distrito__id', flat=True))
    if usuario.tipo.proveedores:
        proveedores = Proveedor.objects.filter(
            id__in=proveedores_ids, 
            completo=True, 
            direcciones__distrito__in = almacenes_distritos
            ).exclude(familia__nombre="IMPUESTOS").distinct()
    else:
        proveedores = Proveedor.objects.none()
    total_prov = proveedores.count()

    myfilter=ProveedorFilter(request.GET, queryset=proveedores)
    proveedores = myfilter.qs

    if request.method == 'POST' and 'btnExcel' in request.POST:
        return convert_excel_proveedores(proveedores_dir)

    #Set up pagination
    p = Paginator(proveedores, 50)
    page = request.GET.get('page')
    proveedores_list = p.get_page(page)
    # Añadir datos de proveedor_direcciones
    for proveedor in proveedores_list:
        direccion = Proveedor_direcciones.objects.filter(
            nombre=proveedor,
            distrito=usuario.distritos
        ).last()  # Obtener la ultima dirección que coincida (más actual)
        if direccion:
            proveedor.telefono = direccion.telefono
            proveedor.contacto = direccion.contacto
            proveedor.distrito = direccion.distrito
            proveedor.domicilio = direccion.domicilio
       

    context = {
        'usuario':usuario,
        'proveedores': proveedores,
        'myfilter':myfilter,
        'proveedores_list':proveedores_list,
        'total_prov':total_prov,
        }


    return render(request,'dashboard/proveedores.html', context)

@perfil_seleccionado_required
def proveedores_altas(request):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
   
   
    if usuario.tipo.proveedores_edicion:
        proveedores = Proveedor.objects.filter(
                completo=True, 
                direcciones__estatus__nombre = "PREALTA",
                ).exclude(familia__nombre="IMPUESTOS").distinct()
    else:
        proveedores = Proveedor.objects.filter(
                completo=True, 
                direcciones__estatus__nombre = "PREALTA",
                direcciones__distrito__id = usuario.distritos.id,
                ).exclude(familia__nombre="IMPUESTOS").distinct()
        
    for proveedor in proveedores:
        proveedor.politicas_no_autorizadas = (
            not proveedor.acepto_politica or
            not proveedor.acepto_politica_proveedor or
            not proveedor.acepto_codigo_etica or
            not proveedor.acepto_aviso_privacidad
        )
    total_prov = proveedores.count()

    myfilter=ProveedorFilter(request.GET, queryset=proveedores)
    proveedores = myfilter.qs

    #if request.method == 'POST' and 'btnExcel' in request.POST:
        #return convert_excel_proveedores(proveedores_dir)

    #Set up pagination
    p = Paginator(proveedores, 50)
    page = request.GET.get('page')
    proveedores_list = p.get_page(page)
    # Añadir datos de proveedor_direcciones
    for proveedor in proveedores_list:
        direccion = Proveedor_direcciones.objects.filter(
            nombre=proveedor,
            #distrito=usuario.distritos
        ).last()  # Obtener la ultima dirección que coincida (más actual)
        if direccion:
            proveedor.telefono = direccion.telefono
            proveedor.contacto = direccion.contacto
            proveedor.distrito = direccion.distrito
            proveedor.domicilio = direccion.domicilio

    context = {
        'usuario':usuario,
        'proveedores': proveedores,
        'myfilter':myfilter,
        'proveedores_list':proveedores_list,
        'total_prov':total_prov,
        }


    return render(request,'dashboard/proveedores_altas.html', context)

def autorizar_alta_proveedor(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    proveedor = Proveedor.objects.get(id=pk)
    proveedor_direcciones = Proveedor_direcciones.objects.filter(nombre=proveedor, estatus__nombre="PREALTA").first()
    
    if request.method =='POST':
        #print('si entra al ciclo')
        status = Estatus_proveedor.objects.get(nombre="NUEVO")
        
        
        proveedor_direcciones.estatus = status
        proveedor_direcciones.save()
         # Asignar folio automáticamente
        # Tomamos el país desde la primera dirección asociada
        
        #pais = proveedor_direcciones.estado.pais.nombre
        #print(f"Pais: {pais}")
        # Obtener el último folio consecutivo para ese país
        #ultimo_folio = Proveedor.objects.filter(
        #    direcciones__estado__pais__nombre__iexact=pais,
        #    folio_consecutivo__isnull=False
        #).aggregate(Max('folio_consecutivo'))['folio_consecutivo__max'] or 0

        #proveedor.folio_consecutivo = ultimo_folio + 1
        #proveedor.save()
        messages.success(request,f'Has autorizado correctamente el alta del proveedor {proveedor.razon_social}')
        return redirect('proveedores-altas')
    
    context = {
        'proveedor':proveedor,
    }
    return render(request,'dashboard/autorizar_alta_proveedor.html', context)

def cancelar_alta_proveedor(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    proveedor = Proveedor.objects.get(id=pk)
    proveedor_direcciones = Proveedor_direcciones.objects.filter(nombre=proveedor, estatus__nombre="PREALTA").first()
    
    if request.method =='POST':
        #print('si entra al ciclo')
        status = Estatus_proveedor.objects.get(nombre="RECHAZADO")
        
        
        proveedor_direcciones.estatus = status
        proveedor_direcciones.save()
         # Asignar folio automáticamente
        # Tomamos el país desde la primera dirección asociada
        
        #pais = proveedor_direcciones.estado.pais.nombre
        #print(f"Pais: {pais}")
        # Obtener el último folio consecutivo para ese país
        #ultimo_folio = Proveedor.objects.filter(
        #    direcciones__estado__pais__nombre__iexact=pais,
        #    folio_consecutivo__isnull=False
        #).aggregate(Max('folio_consecutivo'))['folio_consecutivo__max'] or 0

        #proveedor.folio_consecutivo = ultimo_folio + 1
        #proveedor.save()
        messages.success(request,f'Has autorizado correctamente el alta del proveedor {proveedor.razon_social}')
        return redirect('proveedores-altas')
    
    context = {
        'proveedor':proveedor,
    }
    return render(request,'dashboard/cancelar_alta_proveedor.html', context)
    

@login_required(login_url='user-login')
@perfil_seleccionado_required
def matriz_revision_proveedor(request):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    proveedores = Proveedor_direcciones.objects.filter(estatus__nombre = "REVISION")

    total_prov = proveedores.count()

    myfilter=ProveedorFilter(request.GET, queryset=proveedores)
    proveedores = myfilter.qs

    #Set up pagination
    p = Paginator(proveedores, 50)
    page = request.GET.get('page')
    proveedores_list = p.get_page(page)

    context = {
        'usuario':usuario,
        'proveedores': proveedores,
        'myfilter':myfilter,
        'proveedores_list':proveedores_list,
        'total_prov':total_prov,
        }


    return render(request,'dashboard/matriz_revision_proveedor.html', context)


@login_required(login_url='user-login')
@perfil_seleccionado_required
def proveedores_update(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    razon = request.GET.get('razon_social', '')
    rfc = request.GET.get('rfc', '')
    next_url = request.GET.get('next') or request.POST.get('next')

    if usuario.tipo.proveedores == True:
        proveedores = Proveedor.objects.get(id=pk)
        error_messages = {}
        if request.method =='POST':
            form = ProveedoresForm(request.POST, instance=proveedores)
            if form.is_valid():
                form.save()
                messages.success(request,f'Has actualizado correctamente el proyecto {proveedores.razon_social}')
                #return redirect(f"{reverse('dashboard-proveedores')}?razon_social={razon}&rfc={rfc}")
                return redirect(next_url)
            else:
                for field, errors in form.errors.items():
                    error_messages[field] = errors.as_text()

        else:
            form = ProveedoresForm(instance=proveedores)
    else:
        raise Http404("No tienes permiso para ver esta vista")
    context = {
        'next': next_url,
        'error_messages': error_messages,
        'form': form,
        'proveedores':proveedores,
        'razon': razon,
        'rfc': rfc,
        }

    return render(request,'dashboard/proveedores_update.html', context)



@login_required(login_url='user-login')
@perfil_seleccionado_required
def add_proveedores_old(request):
    usuario = Profile.objects.get(staff=request.user)
    item, created = Proveedor.objects.get_or_create(creado_por=usuario, completo = False)
   

    if request.method =='POST':
        form = ProveedoresForm(request.POST, request.FILES or None, instance = item)
        if form.is_valid():
            item = form.save(commit=False)
            item.completo = True
            item.save()
            # Recuperas los filtros que venían en el POST
            messages.success(request,f'Has agregado correctamente el proveedor {item.razon_social}')
           
            return redirect('dashboard-proveedores')
    else:
        form = ProveedoresForm(instance=item)


    context = {
        'form': form,
        'item':item,
      
        }
    return render(request,'dashboard/add_proveedores.html', context)

@login_required(login_url='user-login')
@perfil_seleccionado_required
def add_proveedor_direccion(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    if usuario.tipo.proveedores == True:
        proveedor = Proveedor.objects.get(id=pk)
        form = ProveedoresDireccionesForm()
        error_messages = {}

        if request.method =='POST':
            item, created = Proveedor_direcciones.objects.get_or_create(nombre = proveedor, creado_por = usuario, completo = False)
            form = ProveedoresDireccionesForm(request.POST, instance = item)
            if form.is_valid():
                item = form.save(commit=False)
                item.disitrito = usuario.distritos
                item.created_at = datetime.now()
                item.completo = True
                item.save()
                messages.success(request,f'Has agregado correctamente la direccion del proveedor {item.nombre.razon_social}')
                return redirect('dashboard-proveedores')
            else:
                for field, errors in form.errors.items():
                    error_messages[field] = errors.as_text()
        else:
            form = ProveedoresDireccionesForm()
    else:
        raise Http404("No tienes permiso para ver esta vista")

    context = {
        'form': form,
        #'item':item,
        'proveedor':proveedor,
        'error_messages': error_messages,
        }
    return render(request,'dashboard/add_proveedor_direccion.html', context)

@login_required(login_url='user-login')
@perfil_seleccionado_required
def add_proveedores2(request, pk=None):
    pk_perfil = request.session.get('selected_profile_id')
    colaborador_sel = Profile.objects.all()
    usuario = colaborador_sel.get(id = pk_perfil)
    if usuario.tipo.nombre == "Subdirector_Alt":
        proveedor, created = Proveedor.objects.get_or_create(creado_por=usuario, completo=False)
        proveedores_dir_ids = Proveedor_direcciones.objects.filter(~Q(estatus__nombre ="REVISION"),~Q(distrito = usuario.distritos)).values_list('id', flat=True)
    elif usuario.tipo.proveedores == True:
        proveedor, created = Proveedor.objects.get_or_create(creado_por=usuario, completo=False)
        proveedores_dir_ids = Proveedor_direcciones.objects.filter(~Q(estatus__nombre ="REVISION"),~Q(distrito = usuario.distritos)).values_list('id', flat=True)
    
    proveedores = Proveedor.objects.filter(direcciones__id__in=proveedores_dir_ids)
    print('proveedores:',proveedores.count())

    if usuario.tipo.nombre == "Subdirector_Alt":
        ProveedorDireccionesFormSet = inlineformset_factory(Proveedor, Proveedor_direcciones, form=Add_ProveedoresDir_Alt_Form, extra=1)
    elif usuario.tipo.proveedores == True:
        ProveedorDireccionesFormSet = inlineformset_factory(Proveedor, Proveedor_direcciones, form=Add_ProveedoresDireccionesForm, extra=1)
    else:
        ProveedorDireccionesFormSet = inlineformset_factory(Proveedor, Proveedor_direcciones, form=ProveedoresDireccionesForm, extra=1)
        
    error_messages = {}
    if request.method == 'POST':
        form = ProveedoresForm(request.POST, instance = proveedor)
        formset = ProveedorDireccionesFormSet(request.POST, instance=proveedor)
        if form.is_valid() and formset.is_valid():
            proveedor = form.save(commit=False)
            proveedor.completo = True
            proveedor.save()
            direcciones = formset.save(commit=False)
            direccion = direcciones[0]
            #direccion.distrito = usuario.distritos
            if usuario.tipo.proveedores == False:
                estatus = Estatus_proveedor.objects.get(nombre ="REVISION")
                direccion.estatus = estatus
            direccion.creado_por = usuario
            direccion.enviado_fecha = date.today()
            direccion.completo = True
            direccion.save()
            messages.success(request, f'Has agregado correctamente el proveedor {proveedor.razon_social} y sus direcciones')
            return redirect('dashboard-proveedores')
        else:
            for field, errors in form.errors.items():
                error_messages[field] = errors.as_text()
            for form in formset.forms:
                for field, errors in form.errors.items():
                    error_messages[field] = errors.as_text()
                
    else:
        form = ProveedoresForm(instance=proveedor)
        formset = ProveedorDireccionesFormSet(instance=proveedor)

    #else:
        #raise Http404("No tienes permiso para ver esta vista")
    context = {
        'form': form,
        'formset': formset,
        'error_messages': error_messages,
    }
    return render(request, 'dashboard/add_proveedores_&_direccion.html', context)

@login_required(login_url='user-login')
@perfil_seleccionado_required
def add_proveedores_comparativo(request, pk=None):
    pk_perfil = request.session.get('selected_profile_id')
    colaborador = Profile.objects.all()
    usuario = colaborador.get(id = pk_perfil)
    proveedor, created = Proveedor.objects.get_or_create(creado_por=usuario, completo=False)
    #proveedores_dir_ids = Proveedor_direcciones.objects.filter(~Q(estatus__nombre ="REVISION"),~Q(distrito = usuario.distrito)).values_list('id', flat=True)
    error_messages = {}
    #proveedores = Proveedor.objects.filter(proveedor_direcciones__id__in=proveedores_dir_ids)
    print('usuario_tipo:',usuario.tipo.proveedores)

    #if usuario.tipo.proveedores == True:
    #    ProveedorDireccionesFormSet = inlineformset_factory(Proveedor, Proveedor_direcciones, form=Add_ProveedoresDireccionesForm, extra=1)
    #else:
    ProveedorDireccionesFormSet = inlineformset_factory(Proveedor, Proveedor_direcciones, form=DireccionComparativoForm, extra=1)
    

    if request.method == 'POST':
        form = ProveedoresForm(request.POST, instance = proveedor)
        formset = ProveedorDireccionesFormSet(request.POST, instance=proveedor)
        if form.is_valid() and formset.is_valid():
            proveedor = form.save(commit=False)
            proveedor.completo = True
            proveedor.save()
            direcciones = formset.save(commit=False)
            direccion = direcciones[0]
            direccion.distrito = usuario.distritos
            if usuario.tipo.proveedores == False:
                estatus = Estatus_proveedor.objects.get(nombre ="COTIZACION")
                direccion.estatus = estatus
            direccion.creado_por = usuario
            direccion.enviado_fecha = date.today()
            direccion.completo = True
            direccion.save()
            messages.success(request, f'Has agregado correctamente el proveedor {proveedor.razon_social} y sus direcciones')
            return redirect('comparativos')
        else:
            for field, errors in form.errors.items():
                error_messages[field] = errors.as_text()
    else:
        form = ProveedoresForm(instance=proveedor)
        formset = ProveedorDireccionesFormSet(instance=proveedor)

    context = {
        'error_messages':error_messages,
        'form': form,
        'formset': formset,
    }
    return render(request, 'dashboard/add_proveedor_direccion_cotizacion.html', context)

@login_required(login_url='user-login')
@perfil_seleccionado_required
def add_proveedores(request, pk=None):
    pk_perfil = request.session.get('selected_profile_id')
    colaborador = Profile.objects.all()
    usuario = colaborador.get(id = pk_perfil)
    #proveedor, created = Proveedor.objects.get_or_create(creado_por=usuario, completo=False)
    proveedores_dir_ids = Proveedor_direcciones.objects.filter(~Q(estatus__nombre ="REVISION"),~Q(distrito = usuario.distritos)).values_list('id', flat=True)
    
    proveedores = Proveedor.objects.filter(proveedor_direcciones__id__in=proveedores_dir_ids)
    print('proveedores:',proveedores.count())

    
    form = ProveedoresExistDireccionesForm()
    form.fields['nombre'].queryset = proveedores

    if request.method == 'POST':
        form = ProveedoresExistDireccionesForm(request.POST)
        if form.is_valid():
            proveedor = form.save(commit=False)
            proveedor.completo = True
            proveedor.save()
            #direcciones = formset.save(commit=False)
            #direccion = direcciones[0]
            #direccion.distrito = usuario.distrito
            #estatus = Estatus_proveedor.objects.get(nombre ="REVISION")
            #direccion.creado_por = usuario
            #direccion.estatus = estatus
            #direccion.completo = True
            #direccion.save()
            messages.success(request, f'Has agregado correctamente el proveedor {proveedor.razon_social} y sus direcciones')
            return redirect('dashboard-proveedores')
    
    context = {
        'proveedores':proveedores,
        'form': form,
    }
    return render(request, 'dashboard/proveedor_exist_&_direccion.html', context)

@login_required(login_url='user-login')
@perfil_seleccionado_required
def edit_proveedores(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    colaborador = Profile.objects.all()
    usuario = colaborador.get(id = pk_perfil)
    proveedor_direccion = Proveedor_direcciones.objects.get(id=pk)
    proveedor = Proveedor.objects.get(id = proveedor_direccion.nombre.id)
    #romper

    ProveedorDireccionesFormSet = inlineformset_factory(Proveedor, Proveedor_direcciones, form =Edit_ProveedoresDireccionesForm, extra=0)
    form = ProveedoresForm(instance=proveedor)
    formset = ProveedorDireccionesFormSet(instance=proveedor)

    if request.method == 'POST':
        form = ProveedoresForm(request.POST or None, instance =proveedor)
        formset = ProveedorDireccionesFormSet(request.POST or None, instance=proveedor)
        if form.is_valid and formset.is_valid():
            
            form.save()
            direcciones = formset.save(commit=False)
            for item in direcciones:
                item.actualizado_por = usuario
                item.modificado_fecha = date.today()
                item.save()
            messages.success(request, 'Has agregado correctamente el proveedor y sus direcciones')
            return redirect('dashboard-proveedores')
        else:
            print(form.errors) 
            print(formset.errors) 
            messages.success(request, 'No está validando')

    context = {
        'form': form,
        'proveedor':proveedor,
        'formset': formset,
    }
    
    return render(request,'dashboard/add_proveedor_direccion.html', context)


@login_required(login_url='user-login')
@perfil_seleccionado_required
def edit_proveedor_direccion(request, pk):

    pk_perfil = request.session.get('selected_profile_id')
    colaborador = Profile.objects.all()
    usuario = colaborador.get(id = pk_perfil)
    next_param = request.POST.get('next') or request.GET.get('next')
    print('next_param:', next_param)
    if usuario.tipo.proveedores == True:
        direccion = Proveedor_direcciones.objects.get(id = pk)
        proveedor = Proveedor.objects.get(id = direccion.nombre.id)

        if request.method =='POST':
            form = ProveedoresDireccionesForm(request.POST, instance = direccion, profile = usuario)
            if form.is_valid():
                direccion = form.save(commit=False)
                direccion.actualizado_por = usuario
                direccion.completo = True
                direccion.save()
                base_url = reverse('proveedor-direcciones', kwargs={'pk': proveedor.id})
                messages.success(request,'Has actualizado correctamente la direccion del proveedor')
                  # Agregamos el `next` a la URL si existe
                if next_param:
                    print('next_param:', next_param)
                    query_params = {'next': next_param}
                    url = f"{base_url}?{urlencode(query_params)}"
                else:
                    print('no hay next_param')
                    url = base_url

                # Redirigimos a la URL final
                return redirect(url)
                #return redirect('proveedor-direcciones', pk= proveedor.id)
        else:
            form = ProveedoresDireccionesForm(instance = direccion, profile = usuario)
    else:
        raise Http404("No tienes permiso para ver esta vista")
    context = {
        'proveedor':proveedor,
        'form': form,
        'direccion':direccion,
       
        }
    return render(request,'dashboard/edit_direcciones_proveedores.html', context)


@perfil_seleccionado_required
def upload_batch_proveedores(request):

    form = Proveedores_BatchForm(request.POST or None, request.FILES or None)


    if form.is_valid():
        form.save()
        form = Proveedores_BatchForm()
        proveedores_list = Proveedor_Batch.objects.get(activated = False)
        print(0)
        f = open(proveedores_list.file_name.path, 'r')
        reader = csv.reader(f)
        next(reader)

        for row in reader:
            razon_social = row[0].strip()
            
            rfc = row[1].strip()
            print(rfc)
            creado_por_id = row[2].strip()
            familia_nombre = row[3].strip()
            extranjero = True if row[4].strip().upper() == 'SI' else False
            visita = True if row[5].strip().upper() == 'SI' else False

            proveedor = Proveedor.objects.filter(razon_social=razon_social).first()

            
            if not proveedor:
                print(2)
                creado_por = Profile.objects.filter(id=creado_por_id).first() if creado_por_id else None
                familia = Familia.objects.filter(nombre=familia_nombre).first() if familia_nombre else None
                if not creado_por:
                    messages.error(request, f'El perfil "{creado_por}" no existe en la base de datos.')
               
                # Crear y guardar la dirección del proveedor
                proveedor = Proveedor(
                    razon_social=razon_social,
                    rfc = rfc,
                    creado_por=creado_por,
                    familia=familia,
                    extranjero= extranjero,
                    visita=visita,
                )
                proveedor.save()
                print(proveedor)

        proveedores_list.activated = True
        proveedores_list.save()
      
    elif request.FILES:
        messages.error(request,'El formato no es CSV')

    context = {
        'form': form,
        }

    return render(request,'dashboard/upload_batch_proveedor.html', context)

@login_required(login_url='user-login')
@perfil_seleccionado_required
def upload_batch_proveedores_direcciones(request):

    form = Proveedores_Direcciones_BatchForm(request.POST or None, request.FILES or None)


    if form.is_valid():
        form.save()
        form = Proveedores_Direcciones_BatchForm()
        proveedores_list = Proveedor_Direcciones_Batch.objects.get(activated=False)

        f = open(proveedores_list.file_name.path, 'r')
        reader = csv.reader(f)
        next(reader)

        for row in reader:
            if Proveedor.objects.filter(razon_social=row[0]):
                nombre = Proveedor.objects.get(razon_social=row[0])
                if Distrito.objects.filter(nombre = row[1]):
                    distrito = Distrito.objects.get(nombre = row[1])
                    if Banco.objects.filter(nombre= row[7]):
                        banco = Banco.objects.get(nombre = row[7])
                        if Estatus_proveedor.objects.filter(nombre = row[12]):
                            estatus = Estatus_proveedor.objects.get(nombre = row[12])
                            if Estado.objects.filter(nombre = row[3]):
                                financiamiento = row[10].strip().upper() == 'SI' if len(row) > 1 else False
                                estado = Estado.objects.get(nombre = row[3])
                                proveedor_direccion = Proveedor_direcciones(nombre=nombre, distrito=distrito,domicilio=row[2],estado=estado,contacto=row[4],email=row[5], telefono= row[6], banco=banco, clabe=row[8], cuenta=row[9], financiamiento=financiamiento,dias_credito=row[11],estatus=estatus)
                                proveedor_direccion.save()
                            else:
                                messages.error(request,f'El estado:{row[3]} no existe dentro de la base de datos')
                        else:
                             messages.error(request,f'El estatus:{row[11]} no existe dentro de la base de datos')
                    else:
                         messages.error(request,f'El banco:{row[7]} no existe dentro de la base de datos')
                else:
                    messages.error(request,f'El distrito:{row[1]} no existe dentro de la base de datos')
            else:
                messages.error(request,f'El proveedor código:{row[0]} no existe dentro de la base de datos')

        proveedores_list.activated = True
        proveedores_list.save()
    elif request.FILES:
        messages.error(request,'El formato no es CSV')

    context = {
        'form': form,
        }

    return render(request,'dashboard/upload_batch_proveedor_direcciones.html', context)

@perfil_seleccionado_required
@tipo_usuario_requerido('proveedores_edicion')
def documentacion_proveedores(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    proveedor = Proveedor.objects.get(id=pk)       
    razon = request.GET.get('razon_social', '')
    rfc = request.GET.get('rfc', '')   
    next_url = request.GET.get('next') or request.POST.get('next')  

    direcciones = Proveedor_direcciones.objects.filter(nombre= proveedor, completo = True).exclude(estatus__nombre__in=["NO REGISTR","RECHAZADO"])
    tiene_servicio = proveedor.direcciones.filter(servicio=True).exists()
    tiene_arrendamiento = proveedor.direcciones.filter(arrendamiento=True).exists()
    tiene_producto = proveedor.direcciones.filter(producto=True).exists()

    # Obtener todos los documentos del proveedor
    documentos = DocumentosProveedor.objects.filter(proveedor=proveedor)
    proveedor.ultimo_cuestionario = proveedor.cuestionarios.order_by('-id').first()
    
   # Definir los tipos de documentos requeridos
    # Lista de los 8 tipos de documentos
    tipos_documentos = [
        'csf',
        'comprobante_domicilio',
        'opinion_cumplimiento',
        'credencial_acta_constitutiva',
        'calificacion',
        'curriculum',
        'competencias',
        'contrato',
        'factura_predial',
        'calidad',
        'otros',
        'visita',
        'carta_credito',
        'cotizacion',
        'busqueda_mediatica',
        'repse',
        'cumplimiento_imss',
    ]

    documentos_count = {tipo: 0 for tipo in tipos_documentos}
    documentos_validados_count = {tipo: 0 for tipo in tipos_documentos}

    for documento in documentos:
        tipo = documento.tipo_documento
        documentos_count[tipo] += 1  # Contar cuántos documentos hay de cada tipo
        if documento.validada:
            documentos_validados_count[tipo] += 1  # Contar cuántos están validados


    if request.method == 'POST':
        #form =  Comentario_Proveedor_Doc_Form(request.POST, instance=proveedor)
        if "btn_validacion" in request.POST:
            fecha_hora = datetime.today()
            for documento in documentos:
                #print(request.POST)
                checkbox_name = f'validar_documento_{documento.id}'
                print("Nombre del checkbox esperado:", checkbox_name)  # Imprimir el nombre esperado
                if checkbox_name in request.POST:
                    print('PASO 1')
                    documento.validada = True
                    documento.validada_por = usuario
                    documento.validada_fecha = fecha_hora
                else:
                    print('No paso')
                    documento.validada = False
                documento.save()
            
        if "btn_eliminar_docto" in request.POST:
            for documento in documentos:
                eliminar_checkbox_name = f'eliminar_documento_{documento.id}'
                print(documento)
                if documento.archivo: 
                    if eliminar_checkbox_name in request.POST: # Verificar que tenga archivo
                        print(documento.archivo)
                        #ruta_archivo = os.path.join(settings.MEDIA_ROOT, str(documento.archivo))
                        #if os.path.exists(ruta_archivo):
                        #   os.remove(ruta_archivo)  # Eliminar archivo del servidor
                        documento.delete()  # Eliminar el registro de la base de datos
                        #documento.save()
            messages.success(request, f"Documentos eliminados correctamente.")
        return redirect(request.path) 
        #else:
        #    messages.error(request,'No está validando')
    print(documentos_count)
    print(documentos_validados_count)
    context = {
        'proveedor':proveedor,
        'direcciones':direcciones,
        'tiene_servicio': tiene_servicio,
        'tiene_arrendamiento': tiene_arrendamiento,
        'tiene_producto': tiene_producto,
        'documentos_count': documentos_count,  # Dict con el total de documentos por tipo
        'documentos_validados_count': documentos_validados_count,  # Dict con validados por tipo
        'documentos': documentos,
        'razon': razon,
        'rfc': rfc,
        'next': next_url,
        }
    
    return render(request,'dashboard/documentacion_proveedor.html', context)

def update_comentario(request):
    data= json.loads(request.body)
    pk = data["pk"]
    dato = data["data"]
    tipo = data["tipo"]
    proveedor = Proveedor.objects.get(id=pk)
    
    if tipo == "acta": 
        proveedor.comentario_acta = dato
        indice = 1
    if tipo == "csf":
        proveedor.comentario_csf = dato
        indice = 2
    if tipo == "domicilio":
        proveedor.comentario_comprobante_domicilio = dato
        indice = 3
    if tipo == "opinion":
        proveedor.comentario_opinion_cumplimiento = dato
        indice = 4
    if tipo == "cv":
        proveedor.comentario_curriculum = dato
        indice = 5
    if tipo == "competencias":
        proveedor.comentario_competencias = dato
        indice = 6
    if tipo == "contrato":
        proveedor.comentario_contrato = dato
        indice = 7
    if tipo == "factura":
        proveedor.comentario_factura = dato
        indice = 8
    if tipo == "calificacion":
        proveedor.comentario_calificacion = dato
        indice = 9
    if tipo == "otros":
        proveedor.comentario_otros = dato
        indice = 10
    if tipo == "calidad":
        proveedor.comentario_calidad = dato
        indice = 11
    if tipo == "visita":
        proveedor.comentario_visita = dato
        indice = 12
    if tipo == 'cotizacion':
        proveedor.comentario_cotizacion = dato
        indice = 13
    if tipo == 'carta_credito':
        proveedor.comentario_carta_credito = dato
        indice = 14
    #if tipo == 'busqueda_mediatica':
    #    proveedor.comentario_busqueda_mediatica = dato
    #    indice = 15

    proveedor.save()
    # Construye un objeto de respuesta que incluya el dato y el tipo.
    response_data = {
        'dato': dato,
        'tipo': tipo,
        'proveedor_id':pk,
        'indice': indice, 
    }

    return JsonResponse(response_data, safe=False)

@login_required(login_url='user-login')
@perfil_seleccionado_required
def upload_batch_products(request):
    form = Products_BatchForm(request.POST or None, request.FILES or None)

    if form.is_valid():
        form.save()
        form = Products_BatchForm()
        product_list = Products_Batch.objects.get(activated=False)

        try:
            # Detectar la codificación del archivo
            with open(product_list.file_name.path, 'rb') as raw_file:
                result = detect(raw_file.read())
                encoding = result['encoding']

            # Abrir el archivo con la codificación detectada
            with open(product_list.file_name.path, 'r', encoding=encoding) as f:
                reader = csv.reader(f)
                next(reader)  # Omitir la primera fila (cabecera)

                for row in reader:
                    if not Product.objects.filter(codigo=row[0]).exists():
                        if Unidad.objects.filter(nombre=row[2]):
                            unidad = Unidad.objects.get(nombre=row[2])
                            if Familia.objects.filter(nombre=row[3]):
                                familia = Familia.objects.get(nombre=row[3])
                                critico = row[5].strip().upper() == 'SI'
                                iva = row[6].strip().upper() == 'SI'
                                activo = row[7].strip().upper() == 'SI'
                                servicio = row[8].strip().upper() == 'SI'

                                if Subfamilia.objects.filter(nombre=row[4], familia=familia):
                                    subfamilia = Subfamilia.objects.get(nombre=row[4], familia=familia)
                                    producto = Product(
                                        codigo=row[0],
                                        nombre=row[1],
                                        unidad=unidad,
                                        familia=familia,
                                        subfamilia=subfamilia,
                                        critico=critico,
                                        iva=iva,
                                        activo=activo,
                                        servicio=servicio,
                                        baja_item=False,
                                        completado=True
                                    )
                                    producto.save()
                                else:
                                    producto = Product(
                                        codigo=row[0],
                                        nombre=row[1],
                                        unidad=unidad,
                                        familia=familia,
                                        critico=critico,
                                        iva=iva,
                                        activo=activo,
                                        servicio=servicio,
                                        baja_item=False,
                                        completado=True
                                    )
                                    producto.save()
                            else:
                                messages.error(request, f'La familia no existe dentro de la base de datos, producto:{row[0]}')
                        else:
                            messages.error(request, f'La unidad no existe dentro de la base de datos, producto:{row[0]}')
                    else:
                        messages.error(request, f'El producto código:{row[0]} ya existe dentro de la base de datos')

            product_list.activated = True
            product_list.save()

        except UnicodeDecodeError as e:
            messages.error(request, f'Error de codificación: {str(e)}')
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')

    elif request.FILES:
        messages.error(request, 'El formato no es CSV')

    context = {
        'form': form,
    }

    return render(request, 'dashboard/upload_batch_products.html', context)


#@login_required(login_url='user-login')
#def product_delete(request, pk):
#    item = Product.objects.get(id=pk)
#    if request.method == 'POST':
#        item.delete()
#        return redirect('dashboard-product')

#    return render(request,'dashboard/product_delete.html')


@login_required(login_url='user-login')
@perfil_seleccionado_required
def add_product(request):
    item, created = Product.objects.get_or_create(completado=False)

    if request.method =='POST':
        form = AddProduct_Form(request.POST, request.FILES or None, instance = item)
        #form.save(commit=False)
        item.completado = True
        if form.is_valid():
            form.save()
            item.save()
            messages.success(request,f'Has agregado correctamente el producto {item.nombre}')
            return redirect('dashboard-product')
    else:
        form = AddProduct_Form()


    context = {
        'form': form,
        'item':item,
        }
    return render(request,'dashboard/add_product.html', context)

@login_required(login_url='user-login')
@perfil_seleccionado_required
def product_update(request, pk):

    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    item = Product.objects.get(id=pk)

    if request.method =='POST':
        form = AddProduct_Form(request.POST, request.FILES or None, instance=item, )
        if form.is_valid():
            form.save()
            messages.success(request,f'Has actualizado correctamente el producto {item.nombre}')
            return redirect('dashboard-product')
    else:
        form = AddProduct_Form(instance=item)


    context = {
        'usuario': usuario,
        'form': form,
        'item':item,
        }
    return render(request,'dashboard/product_update.html', context)


@login_required(login_url='user-login')
@perfil_seleccionado_required
def precio_referencia(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    item = get_object_or_404(Product, id=pk)
    error_messages = {}

    if request.method == 'POST':
        form = PrecioRef_Form(request.POST)
        if form.is_valid():
            # 1) Bloqueo en aplicación
            hay_pendiente = PriceRefChange.objects.filter(
                product=item,
                autorizado__isnull=True
            ).exists()

            
            if hay_pendiente:
                messages.error(request,f"Ya existe una solicitud pendiente de autorización para el producto {item.nombre}.")
                return redirect('dashboard-product')
            cambio = form.save(commit=False)
            cambio.product = item
            cambio.solicitado_por = usuario
            cambio.solicitado_en = datetime.now()
            cambio.save()
            messages.success(request, f"Se generó una solicitud de cambio para {item.nombre}.")
            return redirect('dashboard-product')
    else:
        form = PrecioRef_Form()

    context = {
        'error_messages': error_messages,
        'form': form,
        'item': item,
    }
    return render(request, 'dashboard/precio_referencia.html', context)


@login_required(login_url='user-login')
def price_ref_pending(request):
    """Vista principal para listar y filtrar solicitudes pendientes"""
    cambios = PriceRefChange.objects.filter(autorizado__isnull=True).select_related('product', 'solicitado_por')
    myfilter = PriceRefChangeFilter(request.GET, queryset=cambios)
    cambios = myfilter.qs

    context = {
        'solicitudes': cambios,
        'myfilter': myfilter,
    }
    return render(request, 'dashboard/price_ref_pending.html', context)

@login_required(login_url='user-login')
def price_ref_authorize(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    cambio = get_object_or_404(PriceRefChange, pk=pk)
    producto = cambio.product

    # Validación de estado previo
    if cambio.autorizado is not None:
        messages.error(request, "Esta solicitud ya fue procesada.")
        return redirect('price_ref_pending')

    # Actualiza el cambio
    cambio.autorizado = True
    cambio.autorizado_por = usuario
    cambio.autorizado_en = datetime.now()
    cambio.save()

    # Actualiza el producto
    producto.precioref = cambio.new_value
    producto.porcentaje = cambio.new_porcentaje
    #producto.actualizacion_precioref = cambio.created_at
    producto.save()

    messages.success(request, f'Solicitud de cambio de precio para {producto.nombre} autorizada correctamente.')
    return redirect('price_ref_pending')

def price_ref_reject(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    cambio = get_object_or_404(PriceRefChange, pk=pk)

   
    # Marcar como cancelada (reutilizamos autorizado=False)
    cambio.autorizado = False
    # Opcional: deja traza en motivo
    nota_cancel = f"Cancelada por {usuario} el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    cambio.motivo = (cambio.motivo or "")
    cambio.motivo = (cambio.motivo + ("\n" if cambio.motivo else "") + nota_cancel).strip()
    # Si quieres registrar quién “aprobó” el cambio de estado, reutiliza campos:
    cambio.autorizado_por = usuario
    cambio.autorizado_en = datetime.now()
    cambio.save(update_fields=['autorizado', 'motivo', 'autorizado_por', 'autorizado_en'])

    messages.success(request, "La solicitud fue cancelada correctamente.")
    return redirect('price_ref_pending')

@login_required(login_url='user-login')
def price_ref_history(request):
    # Solo historial: ya procesadas (no pendientes)
    # Si usas Django < 5: usa autorizado__isnull=False
    solicitudes = PriceRefChange.objects.filter(autorizado__isnull=False).select_related('product', 'solicitado_por', 'autorizado_por').order_by('-autorizado_en', '-solicitado_en')
    print('solicitudes:', solicitudes)
    # Con django-filter (opcional)
    #myfilter = PriceRefHistoryFilter(request.GET, queryset=base) if 'PriceRefHistoryFilter' in globals() else None
    #qs = (myfilter.qs if myfilter else base).order_by('-autorizado_en', '-solicitado_en')
    context = {
        'solicitudes': solicitudes,
        #'myfilter': myfilter,
    }

    return render(request, 'dashboard/price_ref_history.html', context)


def load_subfamilias(request):

    familia_id = request.GET.get('familia_id')
    subfamilias = Subfamilia.objects.filter(familia_id = familia_id)

    return render(request, 'dashboard/subfamilia_dropdown_list_options.html',{'subfamilias': subfamilias})


@login_required(login_url='user-login')
@perfil_seleccionado_required
def staff_detail(request, pk):
    worker = User.objects.get(id=pk)
    context={
        'worker': worker,
        }
    return render(request,'dashboard/staff_detail.html', context)

def convert_excel_matriz_proyectos(proyectos):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Matriz_proyectos_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Proyectos')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')
    percent_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(percent_style)


    #Se quita la columna gastado salidas por el momento. 'Suma de Compras','Cliente',
    columns = ['ID','Proyectos','Descripción','Status de Entrega','Monto',
              'Pagado Compras','Pagado Gastos','Creado']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16
        if col_num == 2:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 40

    columna_max = len(columns)+2

    # Agregar los mensajes
    ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por SAVIA 2.0. UH}').style = messages_style
    ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}').style = messages_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 30
   
    rows = [
        (
            p.id,
            p.nombre,
            p.descripcion,
            p.status_de_entrega if p.status_de_entrega is not None else "ND", 
            p.get_projects_total if p.get_projects_total is not None else 0, 
            #p.suma_salidas if p.suma_salidas is not None else 0,
            p.get_total_comprado if p.get_total_comprado is not None else 0, 
            p.get_total_gastado if p.get_total_gastado is not None else 0, 
            #p.suma_gastos if p.suma_gastos is not None else 0, 
            p.created_at
        ) 
        for p in proyectos
    ]

    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            (ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style
            if col_num == 7:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style
            if col_num in [5,6]:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = money_style
    
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)

def convert_excel_proveedores(proveedores):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Proveedores_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Proveedores')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    number_style = NamedStyle(name='number_style', number_format='#,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(number_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)

    columns = ['Distrito','Razón Social','RFC','Domicilio','Teléfono','Estado','Contacto','Email','Email Opción',
               'Banco','Clabe','Cuenta','Financiamiento','Días Crédito','Estatus']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16
        if col_num == 5 or col_num == 8:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 25
        if col_num == 1:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 50

    proveedores_ids = proveedores.values_list('nombre', flat=True).distinct()
    proveedores_unicos = Proveedor.objects.filter(id__in=proveedores_ids, completo=True).count()

    columna_max = len(columns)+2

    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por SAVIA 2.0. UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 20
    ws.cell(row=3, column= columna_max, value="Número de proveedores:")
    ws.cell(row=3, column = columna_max + 1, value=proveedores_unicos).style = number_style

    rows = proveedores.values_list('distrito__nombre','nombre__razon_social','nombre__rfc','domicilio','telefono','estado__nombre',
                                   'contacto','email','email_opt','banco__nombre','clabe','cuenta','financiamiento','dias_credito',
                                   'estatus__nombre'
                              )

    

    #for row, subtotal, iva, total in zip(rows,subtotales, ivas, totales):
    for row in rows:
        row_num += 1
        #row_with_additional_columns = list(row) + [subtotal, iva, total]  # Agrega el subtotal a la fila existente
        for col_num in range(len(row)):
            (ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style
            if col_num == 5:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = body_style
            if col_num == 13:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = number_style
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)

    return(response)

@login_required(login_url='user-login')
@perfil_seleccionado_required
def product_calidad(request):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    if usuario.tipo.calidad == True:    
        items = Product.objects.filter(critico = True, completado = True).order_by('codigo')

        myfilter=ProductCalidadFilter(request.GET, queryset=items)
        items = myfilter.qs

        #Set up pagination
        p = Paginator(items, 50)
        page = request.GET.get('page')
        items_list = p.get_page(page)

        context = {
            'usuario':usuario,
            'items': items,
            'myfilter':myfilter,
            'items_list':items_list,
            }


        return render(request,'dashboard/product_calidad.html', context)
    else:
        raise Http404("No tienes permiso para ver esta vista")
    
#def add_requerimiento_calidad(request, pk):
    #pk_perfil = request.session.get('selected_profile_id')
    #usuario = Profile.objects.get(id = pk_perfil)
    #producto_calidad = get_object_or_404(Producto_Calidad, producto__id=pk)
    #if request.method == 'POST':
    #    req_form = RequerimientoCalidadForm(request.POST)
    #    if req_form.is_valid():
    #        requerimiento = req_form.save(commit=False)
    #        requerimiento.solicitud = producto_calidad
    #        requerimiento.updated_by = usuario
    #        requerimiento.save()
    #        return JsonResponse({'success': True, 'id': requerimiento.id, 'nombre': requerimiento.nombre, 'fecha': requerimiento.fecha.strftime('%Y-%m-%d')})
    #    else:
    #        errors = req_form.errors.as_json()
    #        return JsonResponse({'success': False, 'errors': errors})
    #return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)

def eliminar_requerimiento_calidad(request, pk):
    try:
        requerimiento = Requerimiento_Calidad.objects.get(id=pk)
        requerimiento.delete()
        return JsonResponse({'success': True})
    except Requerimiento_Calidad.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Requerimiento no encontrado'})
    
@login_required(login_url='user-login')
@perfil_seleccionado_required
def product_calidad_update(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id = pk_perfil)
    if usuario.tipo.calidad == True:  
        item = get_object_or_404(Product, id=pk)

        error_messages = {}
        
        req_form = RequerimientoCalidadForm()

        # Obtener o crear Producto_Calidad asociado
        producto_calidad, created = Producto_Calidad.objects.get_or_create(producto=item)
        #form = ProductCalidadForm(instance=producto_calidad)
        #requisitos = producto_calidad.requisitos
        
        
        #if requisitos is None:
        #    requisitos = ''
        if request.method == 'POST':
            print('estoy acá')
            print(request.POST)
            if "actualizar" in request.POST:
                #print('estoy acá2')
                #form = ProductCalidadForm(request.POST, instance=item)
                #if form.is_valid():
                #producto_calidad = form.save(commit=False)
                producto_calidad.updated_by = usuario
                producto_calidad.updated_at = date.today()
                producto_calidad.save()
                messages.success(request, f'Se ha actualizado el producto {item.nombre}')
                return redirect('product_calidad')
                
            if "requirement" in request.POST:
                req_form = RequerimientoCalidadForm(request.POST)
                print('estoy acá3')
                if req_form.is_valid():
                    print('estoy acá4')
                    requerimiento = req_form.save(commit=False)
                    requerimiento.solicitud = producto_calidad
                    requerimiento.updated_by = usuario
                    requerimiento.save()
                    return redirect('product_calidad_update',pk=pk)
                else:
                    # Manejo de errores en formularios
                    for field, errors in req_form.errors.items():
                        error_messages[field] = errors.as_text()
                    for field, errors in req_form.errors.items():
                        error_messages[field] = errors.as_text()

            else:
                #form = ProductCalidadForm(instance=item)
                req_form = RequerimientoCalidadForm()

        context = {

            'error_messages': error_messages,
            #'form': form,
            'req_form': req_form,
            'item': item,
            'producto_calidad': producto_calidad,
            #'requisitos': requisitos,  # Aquí pasas el campo
        }
        return render(request, 'dashboard/product_calidad_update.html', context)
    else:
        raise Http404("No tienes permiso para ver esta vista")
    
def Add_Product_Critico(request):
    # Obtén el perfil y distrito
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id=pk_perfil)
    distrito = usuario.distritos
    if usuario.tipo.calidad == True:
        # Filtra los productos disponibles
        productos_filtrados = Product.objects.filter(critico=False)

          # Respuesta AJAX para Select2
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        term = (request.GET.get('term') or '').strip()
        qs = productos_filtrados
        if term:
            qs = qs.filter(Q(nombre__icontains=term) | Q(codigo__icontains=term))[:20]

        results = []
        for p in qs:
            results.append({
                "id": p.id,
                "text": f"{p.nombre} ({p.codigo or ''})",     # lo que muestra Select2
                "nombre": p.nombre,
                "codigo": p.codigo or "",
                "unidad": getattr(p.unidad, "nombre", "") or "",
                "familia": getattr(p.familia, "nombre", "") or "",
                "subfamilia": getattr(p.subfamilia, "nombre", "") or "",
                # si 'servicio' es booleano, conviértelo a texto legible
                "servicio": ("Sí" if getattr(p, "servicio", False) else "No")
                         if isinstance(getattr(p, "servicio", None), bool)
                         else (p.servicio or ""),
            })
        return JsonResponse({"results": results})

    # --- resto igual ---
    form = Add_Product_CriticoForm()
    form.fields['product'].queryset = productos_filtrados.none()

    if request.method == 'POST':
        form = Add_Product_CriticoForm(request.POST)
        form.fields['product'].queryset = productos_filtrados  # Reasigna el queryset para asegurar que esté actualizado

        if form.is_valid():
            product = form.cleaned_data['product']
            product.critico = True
            product.save()
            messages.success(request, f'El producto {product.nombre} ha sido marcado como crítico.')
            return redirect('product_calidad')
        else:
            # Mostrar los errores de validación
            error_message = "Hubo un error con el formulario. Los siguientes campos no son válidos:\n"
            for field, errors in form.errors.items():
                for error in errors:
                    error_message += f" - {field}: {error}\n"
            messages.error(request, error_message)

    context = {
        'form': form,
    }
    return render(request, 'dashboard/add_product_critico.html', context)
#else:
#    raise Http404("No tienes permiso para ver esta vista")
    

import fitz

def extraer_domicilio_fiscal(pdf_path: str) -> str:
    doc = fitz.open(pdf_path)
    texto = []
    for page in doc:
        texto.append(page.get_text() or "")
    doc.close()
    texto = "\n".join(texto)

    # Guardar los pares clave: valor
    campos = {}
    for line in texto.splitlines():
        if ":" in line:
            k, v = line.split(":", 1)
            campos[k.strip()] = v.strip()

    # Extraer campos básicos
    tipo_vialidad = campos.get("Tipo de Vialidad", "CALLE")
    vialidad      = campos.get("Nombre de Vialidad", "")
    num_ext       = campos.get("Número Exterior", "")
    cp            = campos.get("Código Postal", "")
    colonia       = campos.get("Nombre de la Colonia", "")
    localidad     = campos.get("Nombre de la Localidad", "")
    entidad       = campos.get("Nombre de la Entidad Federativa", "")

    # Construir dirección
    domicilio = f"{tipo_vialidad} {vialidad}"
    if num_ext:
        domicilio += f" #{num_ext}"
    if cp:
        domicilio += f" C.P. {cp}"
    if colonia:
        domicilio += f", Colonia {colonia}"
    if localidad:
        domicilio += f", {localidad}"
    if entidad:
        domicilio += f", {entidad}"

    return domicilio




def descargar_pdf_dd(request, pk):
    proveedor = get_object_or_404(Proveedor, id=pk)
    buf = generar_pdf_dd(proveedor, request)
    return FileResponse(buf, as_attachment=True, filename='cuestionario_debida_diligencia_' + str(proveedor.razon_social) + '.pdf')
    
def generar_pdf_dd(proveedor, request):
    #Configuration of the PDF object
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    # ====== Variables de extracción de datos ======
    dd = Debida_Diligencia.objects.filter(proveedor = proveedor).first()
    pk_perfil = request.session.get('selected_profile_id')
    usuario = Profile.objects.get(id=pk_perfil)

    try:
        documento_csf = DocumentosProveedor.objects.get(
                proveedor = proveedor,
                tipo_documento="csf",
                activo=True
            ) 
    except DocumentosProveedor.DoesNotExist:
        documento_csf = None

    if documento_csf == None:
        print("No se ha subido el documento CSF.")
        domicilio_fiscal = None
    else:  
        domicilio_fiscal = extraer_domicilio_fiscal(documento_csf.archivo.path)
        print(f"El contribuyente es: {domicilio_fiscal}")

    if documento_csf == None:
        print("No se ha subido el documento CSF.")
        tipo = None
    else:  
        tipo = extraer_tipo_contribuyente(documento_csf.archivo.path)
        print(f"El contribuyente es: {tipo}")

    width, height = letter
    # ===== Colores =====
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    azul_claro = Color(0.2421875,0.5703125,0.796875)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    # ===== Encabezado =====
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(150,750,300,20, fill=True, stroke=False) #Barra azul superior Cuestionario de Debida Diligencia
    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(300,755,'Cuestionario de Debida Diligencia')
    c.drawInlineImage('static/images/logo_vordcab.jpg',45,730, 3 * cm, 1.5 * cm) #Imagen vortec

    c.setFillColor(black)
    c.setFont('Helvetica-Bold', 12)
    c.drawString(40, 710,"Información General")
    # ===== Estilos de texto =====
    styles = getSampleStyleSheet()
    label = ParagraphStyle('label', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9)
    value = ParagraphStyle('value', parent=styles['Normal'], fontName='Helvetica', fontSize=9)

    def S(txt): return '' if txt is None else str(txt)
    #elaborado_perfil = proveedor.perfiles.first()
    proveedor_direcciones = proveedor.direcciones.first()
    perfil_proveedor = Profile.objects.filter(proveedor=proveedor).first()
    # === Datos de ejemplo ===
    elaborado_por = perfil_proveedor.staff.staff.first_name + ' ' + perfil_proveedor.staff.staff.last_name
    cargo         = dd.cargo
    fecha         = dd.fecha.strftime('%d/%m/%Y') if dd and dd.fecha else ''
    tel           = proveedor_direcciones.telefono
    correo        = proveedor_direcciones.email
    representante = dd.representante_nombre
    razon_social  = S(getattr(proveedor, 'razon_social', ''))
    rfc           = S(getattr(proveedor, 'rfc', ''))
    if proveedor_direcciones.estado:
        if proveedor_direcciones.estado.pais.nombre == "MEXICO":
            nacionalidad = "MEXICANA"
        else:
            nacionalidad = "EXTRANJERA"
    else:
        nacionalidad = "No estado definido"

    #nacionalidad  = "MEXICANA" if proveedor_direcciones.estado.pais.nombre == "MEXICO" else "EXTRANJERA"
    domicilio     =  proveedor_direcciones.domicilio if proveedor_direcciones else ''
    sitio_web     = dd.sitio_web
    # === Data para tabla ===
    data = [
        [Paragraph('Elaborado por (Nombre):', label), Paragraph(elaborado_por, value)],
        [Paragraph('Cargo:', label), Paragraph(cargo, value)],
        [Paragraph('Fecha:', label), Paragraph(fecha, value)],
        [Paragraph('Teléfono:', label), Paragraph(tel, value)],
        [Paragraph('Correo electrónico:', label), Paragraph(correo, value)],
        [Paragraph('Representante, apoderado legal u oficial de cumplimiento:', label), Paragraph(representante, value)],
        [Paragraph('Razón Social:', label), Paragraph(razon_social, value)],
        [Paragraph('Nacionalidad:', label), Paragraph(nacionalidad, value)],
        [Paragraph('Domicilio:', label), Paragraph(domicilio, value)],
        [Paragraph('Domicilio Fiscal:', label), Paragraph(domicilio_fiscal if documento_csf else 'No disponible', value)],
        [Paragraph('RFC:', label), Paragraph(rfc, value)],
        [Paragraph('Sitio web:', label), Paragraph(sitio_web, value)]
    ]

    # === Estilo de tabla ===
    table = Table(data, colWidths=[6*cm, 12*cm], hAlign='LEFT')
    ts = TableStyle([
        ('BOX', (0,0), (-1,-1), 0.25, colors.black),
        ('INNERGRID', (0,0), (-1,-1), 0.25, colors.grey),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
    ])
    # Subrayado en la columna de valores
    for row in range(len(data)):
        ts.add('LINEBELOW', (1,row), (1,row), 0.6, colors.black)
    table.setStyle(ts)

    # === Dibuja tabla en el canvas ===
    # wrapOn calcula ancho/alto que ocupa la tabla
    w, h = table.wrapOn(c, width, height)
    table.drawOn(c, 40, 700-h)  
        # === Sección 2: Accionistas (solo si Persona Física) ===
    # usa el 'tipo' que ya calculaste arriba
    y_actual = 700 - h - 20  # siguiente posición vertical bajo la tabla 1
    min_y = 120              # margen mínimo para decidir si salto de página

     # Garantizar color y fuente visibles
    c.setFillColor(black)
    c.setFont('Helvetica-Bold', 12)
    c.drawString(40, y_actual,"SECCION I — Partes Relacionadas.")


    if tipo == "Persona Moral":
        accionistas = list(dd.accionistas.all()) if dd else []
        if accionistas:
            data_acc = [["Nombre", "% Participación", "Nacionalidad"]]
            for a in accionistas:
                pct = f"{a.porcentaje_participacion}%" if a.porcentaje_participacion is not None else ""
                data_acc.append([a.nombre or "", pct, a.nacionalidad or ""])

            table_acc = Table(data_acc, colWidths=[9*cm, 3.5*cm, 5.5*cm], hAlign='LEFT')
            ts_acc = TableStyle([
                ('BOX', (0,0), (-1,-1), 0.25, colors.black),
                ('INNERGRID', (0,0), (-1,-1), 0.25, colors.grey),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('FONTSIZE', (0,0), (-1,-1), 9),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                ('ALIGN', (1,1), (1,-1), 'RIGHT'),
            ])
            table_acc.setStyle(ts_acc)

            w_acc, h_acc = table_acc.wrapOn(c, width, height)
            if y_actual - h_acc < min_y:
                c.showPage()
                y_actual = 740

            table_acc.drawOn(c, 40, y_actual - h_acc)
            y_actual -= (h_acc + 12)
        else:
            c.setFont('Helvetica-Bold', 10)
            y_actual = y_actual - 10
            c.setFillColor(prussian_blue)     # azul prusia
            c.drawString(60, y_actual, "2. No se registraron accionistas.")
            y_actual -= 14

    elif tipo == "Persona Física":
        c.setFont('Helvetica', 10)
        y_actual = y_actual - 10
        c.setFillColor(azul_claro)     # azul prusia
        c.drawString(60, y_actual ,"2. Este proveedor no tiene accionistas porque es una persona física.")
        y_actual -= 14

    # === Sección 3: Estructura de Alta Dirección ===
   
    respuesta3 = "Sí" if dd.tiene_alta_direccion else "No"

    c.setFillColor(black)
    pregunta3 = "3. ¿Existe en su compañía una estructura jerárquica definida para los miembros de la alta dirección?"

    # Título + pregunta (con salto de página si hace falta)
    if y_actual < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    c.setFont('Helvetica-Bold', 12)
    if y_actual < min_y:
        c.showPage(); c.setFillColor(azul_claro); y_actual = 750
        #c.setFont('Helvetica-Bold', 12); c.drawString(40, y_actual-10, titulo3); y_actual -= 18
        c.setFont('Helvetica', 10)
    
    c.setFont('Helvetica-Bold', 10)
    c.drawString(40, y_actual-10, pregunta3); y_actual -= 16
    c.setFont("Helvetica", 10)
    c.setFillColor(azul_claro)
    #c.drawString(60, y_actual-10, respuesta3); y_actual -= 16

    if dd and dd.tiene_alta_direccion:
        # Renderizar tabla de Miembro_Alta_Direccion
        miembros = list(dd.miembros_direccion.all())
        if miembros:
            data_dir = [["Nombre", "Años de servicio", "Cargo", "Nacionalidad"]]
            for m in miembros:
                data_dir.append([
                    m.nombre or "",
                    str(m.anios_servicio) if m.anios_servicio is not None else "",
                    m.cargo or "",
                    m.nacionalidad or ""
                ])

            table_dir = Table(data_dir, colWidths=[7*cm, 3*cm, 5*cm, 3*cm], hAlign='LEFT')
            table_dir.setStyle(TableStyle([
                ('BOX', (0,0), (-1,-1), 0.25, colors.black),
                ('INNERGRID', (0,0), (-1,-1), 0.25, colors.grey),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('FONTSIZE', (0,0), (-1,-1), 9),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                ('ALIGN', (1,1), (1,-1), 'RIGHT'),  # Años a la derecha
            ]))
            w_dir, h_dir = table_dir.wrapOn(c, width, height)
            if y_actual - h_dir < min_y:
                c.showPage();  c.setFillColor(prussian_blue) ; y_actual = 750
                #c.setFont('Helvetica-Bold', 12); c.drawString(40, y_actual, titulo3); y_actual -= 18
                c.setFont('Helvetica', 10); c.drawString(40, y_actual, pregunta3); y_actual -= 16
                w_dir, h_dir = table_dir.wrapOn(c, width, height)
            table_dir.drawOn(c, 40, y_actual - h_dir)
            y_actual -= (h_dir + 12)
        else:
            # Sí, pero no cargaron miembros
            if y_actual < min_y:
                c.showPage();  c.setFillColor(prussian_blue) ; y_actual = 750
                #c.setFont('Helvetica-Bold', 12); c.drawString(40, y_actual, titulo3); y_actual -= 18
                c.setFont('Helvetica', 10); c.drawString(40, y_actual, pregunta3); y_actual -= 16
            c.drawString(60, y_actual, "Sí. No se registraron miembros de la alta dirección.")
            y_actual -= 14
    else:
        # No (o None)
        if y_actual < min_y:
            c.showPage(); c.setFillColor(black); y_actual = 750
            #c.setFont('Helvetica-Bold', 12); c.drawString(40, y_actual, titulo3); y_actual -= 18
            c.setFont('Helvetica', 10); c.drawString(40, y_actual, pregunta3); y_actual -= 16
        c.drawString(60, y_actual - 10, respuesta3)
        y_actual -= 14
    
        # === Pregunta 4 ===
    # 4. ¿Alguno(s) de sus empleados clave, accionista(s) o Miembro(s) de la alta dirección de su empresa es (son)
    # servidor(es) o funcionario(s) público(s)?  Sí ______   No ______
    # En caso afirmativo: NOMBRE | CARGO | PUESTO GUBERNAMENTAL | PERIODO EN FUNCIONES
    # Estilo para preguntas (puedes declararlo una sola vez junto a tus estilos)
    pstyle_pregunta = ParagraphStyle(
        "pstyle_pregunta",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=12,          # espacio entre líneas
    )
    pstyle_respuesta = ParagraphStyle(
        "pstyle_respuesta",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        textColor = azul_claro,
        fontSize=10,
        leading=12,          # espacio entre líneas
    )
    pregunta4 = ("4. ¿Alguno(s) de sus empleados clave, accionista(s) o Miembro(s) de la alta dirección de su empresa "
                 "es (son) servidor(es) o funcionario(s) público(s)?")

    c.setFillColor(black)
    c.setFont('Helvetica', 10)
    if y_actual < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    #c.drawString(40, y_actual-10, pregunta4); y_actual -= 16
    max_width = width - 80  # márgenes izq/der de 40
    p4 = Paragraph(pregunta4, pstyle_pregunta)
    w4, h4 = p4.wrapOn(c, max_width, height)
    if y_actual - h4 < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750  # nueva página

    # dibuja el párrafo; ojo: Paragraph se coloca con la esquina inferior izquierda
    p4.drawOn(c, 40, y_actual - h4)
    y_actual -= (h4 + 6)  # avanza el cursor vertical
    if dd and dd.empleado_funcionarios_publicos:
        funcionarios = dd.funcionarios_publicos.all()
        if funcionarios:
            data_f = [["NOMBRE", "CARGO", "PUESTO GUBERNAMENTAL", "PERIODO EN FUNCIONES"]]
            for funcionario in funcionarios:
                nombre = funcionario.nombre
                cargo_empresa = funcionario.cargo
                puesto_gob   = funcionario.puesto_gubernamental
                periodo      = funcionario.periodo_funciones    

                data_f.append([nombre, cargo_empresa, puesto_gob, str(periodo)])

            table_f = Table(data_f, colWidths=[6*cm, 5*cm, 6*cm, 3*cm], hAlign='LEFT')
            table_f.setStyle(TableStyle([
                ('BOX', (0,0), (-1,-1), 0.25, colors.black),
                ('INNERGRID', (0,0), (-1,-1), 0.25, colors.grey),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('FONTSIZE', (0,0), (-1,-1), 9),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ]))

            w_f, h_f = table_f.wrapOn(c, width, height)
            if y_actual - h_f < min_y:
                c.showPage(); c.setFillColor(black); y_actual = 750
                c.setFont('Helvetica', 10); c.drawString(40, y_actual, pregunta4); y_actual -= 16
                w_f, h_f = table_f.wrapOn(c, width, height)

            table_f.drawOn(c, 40, y_actual - h_f)
            y_actual -= (h_f + 12)
        else:
            if y_actual < min_y:
                c.showPage(); c.setFillColor(azul_claro); y_actual = 750
                c.setFont('Helvetica', 10); c.drawString(40, y_actual, pregunta4); y_actual -= 16
            c.setFillColor(azul_claro)
            c.setFont('Helvetica',10)
            c.drawString(40, y_actual - 5, "Sí. No se registraron funcionarios/servidores públicos relacionados.")
            y_actual -= 14
    else:
        # Respuesta NO: solo la leyenda
        if y_actual < min_y:
            c.showPage(); c.setFillColor(azul_claro); y_actual = 750
            c.setFont('Helvetica', 10); c.drawString(40, y_actual, pregunta4); y_actual -= 16
        c.setFillColor(azul_claro)
        c.setFont('Helvetica',10)
        c.drawString(60, y_actual - 5, "No.")
        y_actual -= 14

    c.setFillColor(black)
    pregunta5 = (
        "5. ¿La empresa, pertenece parcial o totalmente a un “servidor público” o una persona "
        "relacionada a uno de ellos, ya sea como accionista, inversionista, socio de la organización, "
        "o de cualquier otra manera?"
    )

    # Párrafo de la pregunta (ajuste de línea)
    p5 = Paragraph(pregunta5, pstyle_pregunta)
    max_w = width - 80
    w5, h5 = p5.wrapOn(c, max_w, height)
    if y_actual - h5 < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    p5.drawOn(c, 40, y_actual - h5)
    y_actual -= (h5 + 6)

        # Sí / No marcado según el modelo
    si = bool(dd and dd.pertenece_funcionario_publico)
    c.setFont("Helvetica", 10)
    if y_actual < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    #c.drawString(60, y_actual, "Sí" if si else "No")
    #y_actual -= 14

    if si:
        # Tabla: relaciones con servidores públicos (Relacion_Servidor_Publico)
        c.drawString(60, y_actual, "Sí" if si else "No")
        y_actual -= 14
        relaciones = dd.relaciones_servidores.all()
        if relaciones:
            data_rel = [[
                "NOMBRE DEL SERVIDOR PÚBLICO",
                "TIPO DE RELACIÓN (ACCIONISTA, INVERSIONISTA, SOCIO, ETC.)",
                "PORCENTAJE DE PARTICIPACIÓN"
            ]]
            for r in relaciones:
                pct = f"{r.porcentaje_participacion}%" if r.porcentaje_participacion is not None else ""
                data_rel.append([r.nombre_servidor or "", r.tipos_relacion or "", pct])

            table_rel = Table(data_rel, colWidths=[7*cm, 7*cm, 4*cm], hAlign="LEFT")
            table_rel.setStyle(TableStyle([
                ('BOX', (0,0), (-1,-1), 0.25, colors.black),
                ('INNERGRID', (0,0), (-1,-1), 0.25, colors.grey),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('FONTSIZE', (0,0), (-1,-1), 9),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                ('ALIGN', (2,1), (2,-1), 'RIGHT'),
            ]))

            w_rel, h_rel = table_rel.wrapOn(c, width, height)
            if y_actual - h_rel < min_y:
                c.showPage(); c.setFillColor(prussian_blue); y_actual = 750
                # reimprime la pregunta y la línea Sí/No para contexto mínimo (opcional)
                p5.drawOn(c, 40, 750 - h5); y_actual = 750 - h5 - 6
                c.setFont("Helvetica", 10)
                c.drawString(60, y_actual, "Sí" if si else "No")
                y_actual -= 14
                w_rel, h_rel = table_rel.wrapOn(c, width, height)

            table_rel.drawOn(c, 40, y_actual - h_rel)
            y_actual -= (h_rel + 12)
        else:
            if y_actual < min_y:
                c.showPage(); c.setFillColor(prussian_blue); y_actual = 750
                p5.drawOn(c, 40, 750 - h5); y_actual = 750 - h5 - 6
                c.setFont("Helvetica", 10)
                c.drawString(60, y_actual, "Sí" if si else "No")
                y_actual -= 14
            c.setFillColor(azul_claro)
            c.drawString(40, y_actual - 5, "Sí. No se registraron relaciones con servidores públicos.")
            y_actual -= 14
    else:
        if y_actual < min_y:
            c.showPage(); c.setFillColor(black); y_actual = 750
            p5.drawOn(c, 40, 750 - h5); y_actual = 750 - h5 - 6
            c.setFont("Helvetica", 10)
            c.drawString(60, y_actual, "Sí" if si else "No")
            y_actual -= 14
        c.setFillColor(azul_claro)
        c.drawString(60, y_actual-5, "No")
        y_actual -= 14


    c.showPage()

    # --- NUEVA HOJA: Pregunta A ---
    # forzar segunda hoja
    c.setFillColor(black)
    y_actual = 750
    min_y = 80

    preguntaA = (
        "A. ¿Hará del conocimiento a Grupo Vordcab en caso de existir, relaciones familiares de cualquiera "
        "de sus empleados, socios, funcionarios, directores o accionistas y algún empleado de Grupo Vordcab "
        "u otro servidor público del gobierno mexicano (Personas Políticamente Expuestas) y notificará si "
        "surge tal situación durante la vigencia del Acuerdo Comercial?"
    )

    # Párrafo con ajuste de línea
    pA = Paragraph(preguntaA, pstyle_pregunta)
    max_w = width - 80  # márgenes de 40 izq/der
    wA, hA = pA.wrapOn(c, max_w, height)
    pA.drawOn(c, 40, y_actual - hA)
    y_actual -= (hA + 8)

    # Respuesta: Sí / No según dd.notificar_relacion_familiar
    respA = "Sí" if dd.notificar_relacion_familiar else "No"
    c.setFillColor(azul_claro)
    c.setFont("Helvetica", 10)
    c.drawString(60, y_actual, f"{respA}")
    y_actual -= 14
    max_w = width - 80  # márgenes de 40 izq/der


    c.setFillColor(black)
    c.setFont('Helvetica-Bold', 12)
    y_actual -= 10
    c.drawString(40, y_actual,"SECCIÓN II. Fuentes de Financiamiento")
    # ---------- P6 ----------
    preg6 = (
        "6. ¿Alguna vez le ha sido suspendida la realización de actos, operaciones o servicios por las "
        "instituciones bancarias o de crédito (bloqueo de cuentas bancarias y/o activos) por orden de la "
        "Unidad de Inteligencia Financiera en México o autoridad equivalente en el extranjero?"
    )
    p6 = Paragraph(preg6, pstyle_pregunta)
    w6, h6 = p6.wrapOn(c, max_w, height)
    if y_actual - h6 < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    p6.drawOn(c, 40, y_actual - h6 - 10)
    y_actual -= (h6 + 16)

    # Respuesta P6: si es True, muestra el detalle; si no, "No"
    bloq = dd.cuentas_bloqueadas
    detalle = dd.detalle_cuentas_bloqueadas

    # Línea de respuesta (Sí/No) en azul prusia
    if y_actual < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    c.setFont("Helvetica", 10)
    c.setFillColor(azul_claro)
    c.drawString(60, y_actual, "Sí" if bloq else "No")
    y_actual -= 14
    c.setFillColor(black)

    # Detalle como párrafo, SOLO si es “Sí” y hay texto
    if bloq and detalle:
        p6_det = Paragraph(detalle, pstyle_respuesta)  # usa tu estilo de párrafo
        w6d, h6d = p6_det.wrapOn(c, max_w, height)
        if y_actual - h6d < min_y:
            c.showPage(); c.setFillColor(black); y_actual = 750
        c.setFillColor(azul_claro)
        p6_det.drawOn(c, 60, y_actual - h6d)  # un poco más adentro que la respuesta
        y_actual -= (h6d + 8)
    # ---------- P7 ----------
    preg7 = (
        "7. ¿Cuenta con fuentes de financiamientos externas para alcanzar la capacidad económica requerida "
        "para la proveeduría del producto o servicio?"
    )
    p7 = Paragraph(preg7, pstyle_pregunta)
    w7, h7 = p7.wrapOn(c, max_w, height)
    if y_actual - h7 < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    c.setFillColor(azul_claro)
    p7.drawOn(c, 40, y_actual - h7)
    y_actual -= (h7 + 6)

    fin_ext = dd.financiamiento_externo  # <- asegúrate del nombre exacto
    fuentes = dd.fuentes_financiamiento
    # Línea de respuesta (Sí/No) en azul prusia
    if y_actual < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    c.setFont("Helvetica", 10)
    c.setFillColor(azul_claro)
    c.drawString(60, y_actual, "Sí" if fin_ext else "No")
    y_actual -= 14
    c.setFillColor(black)

    # Detalle como párrafo, SOLO si es “Sí” y hay texto en fuentes
    if fin_ext and fuentes:
        p7_det = Paragraph(fuentes, pstyle_respuesta)
        w7d, h7d = p7_det.wrapOn(c, max_w, height)
        if y_actual - h7d < min_y:
            c.showPage(); c.setFillColor(azul_claro); y_actual = 750
        p7_det.drawOn(c, 60, y_actual - h7d)
        y_actual -= (h7d + 8)

    # ---------- P8 ----------
    preg8 = (
        "8. ¿Existen en su empresa controles, políticas y vigilancia de cumplimiento a las disposiciones "
        "nacionales e internacionales que prohíben el Lavado de Dinero y el financiamiento al terrorismo?"
    )
    p8 = Paragraph(preg8, pstyle_pregunta)
    w8, h8 = p8.wrapOn(c, max_w, height)
    if y_actual - h8 < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    p8.drawOn(c, 40, y_actual - h8)
    y_actual -= (h8 + 6)

    val8 = dd.controles_antilavado
    resp8 = ("Sí" if val8 else "No") if isinstance(val8, bool) else (str(val8) if val8 is not None else "No")

    if y_actual < min_y:
        c.showPage(); c.setFillColor(azul_claro); y_actual = 750
    c.setFont("Helvetica", 10)
    c.setFillColor(azul_claro)
    c.drawString(60, y_actual, resp8)
    y_actual -= 14
    #   ============= SECCION III ==============
    c.setFillColor(black)
    c.setFont('Helvetica-Bold', 12)
    y_actual -= 10
    c.drawString(40, y_actual,"SECCIÓN III. Conflicto de intereses")

    max_w = width - 80
    preg10 = (
        "10. ¿Alguna de las personas que serán responsables de interactuar con Grupo Vordcab o sus empresas "
        "¿fueron en el pasado trabajadores de Grupo Vordcab?"
    )
    p10 = Paragraph(preg10, pstyle_pregunta)

    w10, h10 = p10.wrapOn(c, max_w, height)
    if y_actual - h10 < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    p10.drawOn(c, 40, y_actual - h10)
    y_actual -= (h10 + 6)

    # Respuesta (azul prusia)
    resp10 = "Sí" if dd.responsables_interactuar else "No"
    if y_actual < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    c.setFont("Helvetica", 10)
    c.setFillColor(azul_claro)
    c.drawString(60, y_actual, resp10)
    y_actual -= 14
    c.setFillColor(black)

    # Nota/indicaciones (opcional, si quieres el texto largo previo a la tabla)
    nota10 = (
        "Personas que serán responsables de interactuar con Grupo Vordcab (únicamente las 3 principales), "
        "informando si ha sido empleado de Grupo Vordcab, en su caso el año en que causó baja y el puesto que desempeñó:"
    )
    p10b = Paragraph(nota10, pstyle_pregunta)
    w10b, h10b = p10b.wrapOn(c, max_w, height)
    if y_actual - h10b < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    p10b.drawOn(c, 40, y_actual - h10b)
    y_actual -= (h10b + 8)

    # --- Tabla de responsables_interaccion (SIEMPRE se muestra) ---
    data_ri = [["Nombre", "¿Trabajó para Grupo Vordcab?", "Año", "Puesto que ocupó"]]

    # Puedes iterar directo o usar values_list para eficiencia
    responsables = dd.responsables_interaccion.all()
    for responsable in responsables:
        data_ri.append([
            responsable.nombre,
            "Sí" if responsable.trabajo_previo_vordcab else "No",
            str(responsable.anio_baja),
            responsable.puesto_ocupado,
        ])

    
    table_ri = Table(data_ri, colWidths=[7*cm, 5*cm, 2*cm, 5*cm], hAlign="LEFT")
    table_ri.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 0.25, colors.black),
        ('INNERGRID', (0,0), (-1,-1), 0.25, colors.grey),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('ALIGN', (2,1), (2,-1), 'CENTER'),
    ]))

    w_ri, h_ri = table_ri.wrapOn(c, width, height)
    if y_actual - h_ri < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
        # Reimprime la pregunta y respuesta para contexto mínimo (opcional)
        p10.drawOn(c, 40, 750 - h10); y_actual = 750 - h10 - 6
        c.setFont("Helvetica", 10); c.setFillColor(prussian_blue)
        c.drawString(60, y_actual, resp10); y_actual -= 14
        c.setFillColor(black)
        p10b.drawOn(c, 40, y_actual - h10b); y_actual -= (h10b + 8)
        w_ri, h_ri = table_ri.wrapOn(c, width, height)

    table_ri.drawOn(c, 40, y_actual - h_ri)
    y_actual -= (h_ri + 12)

   # ============= SECCIÓN IV. Ética =============
    c.setFillColor(black)
    c.setFont('Helvetica-Bold', 12)
    y_actual -= 10
    c.drawString(40, y_actual, "SECCIÓN IV. Ética")
    y_actual -= 18
    max_w = width - 80  # márgenes 40–40

    # ---------- 11 ----------
    preg11 = ("11. ¿Su empresa apoya y respeta la protección de los derechos humanos fundamentales, "
            "reconocidos internacionalmente dentro de su ámbito de influencia?")
    p11 = Paragraph(preg11, pstyle_pregunta)
    w11, h11 = p11.wrapOn(c, max_w, height)
    if y_actual - h11 < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    p11.drawOn(c, 40, y_actual - h11)
    y_actual -= (h11 + 6)

    c.setFont("Helvetica", 10)
    c.setFillColor(azul_claro)
    c.drawString(60, y_actual, "Sí" if bool(getattr(dd, "respeta_derechos_humanos", False)) else "No")
    y_actual -= 14
    c.setFillColor(black)

    # ---------- 12 ----------
    preg12 = "12. ¿Su empresa promueve y aplica la eliminación de toda forma de trabajo forzoso o realizado bajo coacción?"
    p12 = Paragraph(preg12, pstyle_pregunta)
    w12, h12 = p12.wrapOn(c, max_w, height)
    if y_actual - h12 < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    p12.drawOn(c, 40, y_actual - h12)
    y_actual -= (h12 + 6)

    c.setFont("Helvetica", 10)
    c.setFillColor(azul_claro)
    c.drawString(60, y_actual, "Sí" if  dd.elimina_trabajo_forzoso else "No")
    y_actual -= 14
    c.setFillColor(black)

    # ---------- 13 ----------
    preg13 = ("13. ¿Todos sus empleados, sin importar su nivel de responsabilidad, trabajan con contrato y "
            "reciben prestaciones acordes a la ley?")
    p13 = Paragraph(preg13, pstyle_pregunta)
    w13, h13 = p13.wrapOn(c, max_w, height)
    if y_actual - h13 < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    p13.drawOn(c, 40, y_actual - h13)
    y_actual -= (h13 + 6)

    val13 = bool(getattr(dd, "empleados_contrato_prestaciones", False))
    c.setFont("Helvetica", 10)
    c.setFillColor(azul_claro)
    c.drawString(60, y_actual, "Sí" if val13 else "No")
    y_actual -= 14
    c.setFillColor(black)

    if not val13:
        texto_exp = (getattr(dd, "explicacion_sin_contrato", "") or "").strip()
        if texto_exp:
            p13_det = Paragraph(texto_exp, pstyle_respuesta)
            w13d, h13d = p13_det.wrapOn(c, max_w, height)
            if y_actual - h13d < min_y:
                c.showPage(); c.setFillColor(black); y_actual = 750
            p13_det.drawOn(c, 60, y_actual - h13d)
            y_actual -= (h13d + 8)

    # ---------- 14 ----------
    preg14 = "14. ¿Su empresa promueve y aplica la erradicación del trabajo infantil?"
    p14 = Paragraph(preg14, pstyle_pregunta)
    w14, h14 = p14.wrapOn(c, max_w, height)
    if y_actual - h14 < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    p14.drawOn(c, 40, y_actual - h14)
    y_actual -= (h14 + 6)

    c.setFont("Helvetica", 10)
    c.setFillColor(azul_claro)
    c.drawString(60, y_actual, "Sí" if bool(getattr(dd, "erradica_trabajo_infantil", False)) else "No")
    y_actual -= 14
    c.setFillColor(black)

    # ---------- 15 ----------
    preg15 = ("15. ¿Su empresa promueve y aplica la abolición de prácticas de discriminación en el empleo y la ocupación?")
    p15 = Paragraph(preg15, pstyle_pregunta)
    w15, h15 = p15.wrapOn(c, max_w, height)
    if y_actual - h15 < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    p15.drawOn(c, 40, y_actual - h15)
    y_actual -= (h15 + 6)

    c.setFont("Helvetica", 10)
    c.setFillColor(azul_claro)
    c.drawString(60, y_actual, "Sí" if bool(getattr(dd, "elimina_discriminacion", False)) else "No")
    y_actual -= 14
    c.setFillColor(black)

    # ---------- 16 ----------
    preg16 = "16. ¿Su empresa asegura un enfoque preventivo que favorezca el medio ambiente?"
    p16 = Paragraph(preg16, pstyle_pregunta)
    w16, h16 = p16.wrapOn(c, max_w, height)
    if y_actual - h16 < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    p16.drawOn(c, 40, y_actual - h16)
    y_actual -= (h16 + 6)

    c.setFont("Helvetica", 10)
    c.setFillColor(azul_claro)
    c.drawString(60, y_actual, "Sí" if bool(getattr(dd, "enfoque_medio_ambiente", False)) else "No")
    y_actual -= 14
    c.setFillColor(black)

    # ---------- NUEVA HOJA: SECCIÓN V ----------
    c.showPage()
    c.setFillColor(black)
    y_actual = 750
    min_y = 80

    max_w = width - 80  # márgenes 40–40

    # Encabezado sección
    c.setFont('Helvetica-Bold', 12)
    c.drawString(40, y_actual, "SECCIÓN V: Características de su programa de integridad o cumplimiento")
    y_actual -= 18
    c.setFont('Helvetica', 10)
    #c.drawString(40, y_actual, "Responda: Sí; No; En proceso")
    #y_actual -= 14

    # ========== A ==========
    pregA = "A. ¿Cuenta con un código de ética?"
    pA = Paragraph(pregA, pstyle_pregunta)
    wA, hA = pA.wrapOn(c, max_w, height)
    need = hA + 6 + 14
    if y_actual - need < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    pA.drawOn(c, 40, y_actual - hA); y_actual -= (hA + 6)
    c.setFont("Helvetica", 10); c.setFillColor(azul_claro)
    c.drawString(60, y_actual, dd.codigo_etica)
    y_actual -= 14; c.setFillColor(black)

    # ========== B ==========
    pregB = "B. ¿Cuenta con un código de conducta?"
    pB = Paragraph(pregB, pstyle_pregunta)
    wB, hB = pB.wrapOn(c, max_w, height)
    need = hB + 6 + 14
    if y_actual - need < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    pB.drawOn(c, 40, y_actual - hB); y_actual -= (hB + 6)
    c.setFont("Helvetica", 10); c.setFillColor(azul_claro)
    c.drawString(60, y_actual, "Sí" if bool(getattr(dd, "codigo_conducta", False)) else "No")
    y_actual -= 14; c.setFillColor(black)

    # ========== C ==========
    pregC = "C. ¿Cuenta con una política anticorrupción?"
    pC = Paragraph(pregC, pstyle_pregunta)
    wC, hC = pC.wrapOn(c, max_w, height)
    need = hC + 6 + 14
    if y_actual - need < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    pC.drawOn(c, 40, y_actual - hC); y_actual -= (hC + 6)
    c.setFont("Helvetica", 10); c.setFillColor(azul_claro)
    c.drawString(60, y_actual, "Sí" if bool(getattr(dd, "politica_anticorrupcion", False)) else "No")
    y_actual -= 14; c.setFillColor(black)

    # ========== D ==========
    pregD = "D. ¿Tiene algún otro documento que regule el comportamiento ético?"
    pD = Paragraph(pregD, pstyle_pregunta)
    wD, hD = pD.wrapOn(c, max_w, height)
    need = hD + 6 + 14
    if y_actual - need < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    pD.drawOn(c, 40, y_actual - hD); y_actual -= (hD + 6)
    c.setFont("Helvetica", 10); c.setFillColor(azul_claro)
    c.drawString(60, y_actual, "Sí" if bool(getattr(dd, "otro_documento_etico", False)) else "No")
    y_actual -= 14; c.setFillColor(black)

    # (E omitida)

    # ========== F (sub-preguntas) ==========
    pF = Paragraph("F. El documento/política/códigos indicado(s) incluye(n) las siguientes características:", pstyle_pregunta)
    wF, hF = pF.wrapOn(c, max_w, height)
    need = hF + 6
    if y_actual - need < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    pF.drawOn(c, 40, y_actual - hF); y_actual -= (hF + 6)

    # F.1
    pF1 = Paragraph("   F.1 Transparencia en cuanto a donativos", pstyle_pregunta)
    w, h = pF1.wrapOn(c, max_w, height); need = h + 6 + 14
    if y_actual - need < min_y: c.showPage(); c.setFillColor(black); y_actual = 750
    pF1.drawOn(c, 40, y_actual - h); y_actual -= (h + 6)
    c.setFont("Helvetica",10); c.setFillColor(azul_claro)
    c.drawString(60, y_actual, "Sí" if bool(getattr(dd,"transparencia_donativos",False)) else "No")
    y_actual -= 14; c.setFillColor(black)

    # F.2
    pF2 = Paragraph("   F.2 Son de conocimiento público", pstyle_pregunta)
    w, h = pF2.wrapOn(c, max_w, height); need = h + 6 + 14
    if y_actual - need < min_y: c.showPage(); c.setFillColor(black); y_actual = 750
    pF2.drawOn(c, 40, y_actual - h); y_actual -= (h + 6)
    c.setFont("Helvetica",10); c.setFillColor(azul_claro)
    c.drawString(60, y_actual, "Sí" if bool(getattr(dd,"conocimiento_publico",False)) else "No")
    y_actual -= 14; c.setFillColor(black)

    # F.3
    pF3 = Paragraph("   F.3 Son extensivos a grupos de interés, partes relacionadas, proveedores y otros", pstyle_pregunta)
    w, h = pF3.wrapOn(c, max_w, height); need = h + 6 + 14
    if y_actual - need < min_y: c.showPage(); c.setFillColor(black); y_actual = 750
    pF3.drawOn(c, 40, y_actual - h); y_actual -= (h + 6)
    c.setFont("Helvetica",10); c.setFillColor(azul_claro)
    c.drawString(60, y_actual, "Sí" if bool(getattr(dd,"extensivo_grupos_interes",False)) else "No")
    y_actual -= 14; c.setFillColor(black)

    # F.4
    pF4 = Paragraph("   F.4 Transparencia en aportaciones y contribuciones políticas", pstyle_pregunta)
    w, h = pF4.wrapOn(c, max_w, height); need = h + 6 + 14
    if y_actual - need < min_y: c.showPage(); c.setFillColor(black); y_actual = 750
    pF4.drawOn(c, 40, y_actual - h); y_actual -= (h + 6)
    c.setFont("Helvetica",10); c.setFillColor(azul_claro)
    c.drawString(60, y_actual, "Sí" if bool(getattr(dd,"transparencia_contribuciones_politicas",False)) else "No")
    y_actual -= 14; c.setFillColor(black)

    # F.5
    pF5 = Paragraph("   F.5 Prohibición expresa de ofrecer o recibir sobornos", pstyle_pregunta)
    w, h = pF5.wrapOn(c, max_w, height); need = h + 6 + 14
    if y_actual - need < min_y: c.showPage(); c.setFillColor(black); y_actual = 750
    pF5.drawOn(c, 40, y_actual - h); y_actual -= (h + 6)
    c.setFont("Helvetica",10); c.setFillColor(azul_claro)
    c.drawString(60, y_actual, "Sí" if bool(getattr(dd,"prohibicion_sobornos",False)) else "No")
    y_actual -= 14; c.setFillColor(black)

    # F.6
    pF6 = Paragraph("   F.6 Prohibición de ofrecer o recibir incentivos de cualquier tipo", pstyle_pregunta)
    w, h = pF6.wrapOn(c, max_w, height); need = h + 6 + 14
    if y_actual - need < min_y: c.showPage(); c.setFillColor(black); y_actual = 750
    pF6.drawOn(c, 40, y_actual - h); y_actual -= (h + 6)
    c.setFont("Helvetica",10); c.setFillColor(azul_claro)
    c.drawString(60, y_actual, "Sí" if bool(getattr(dd,"prohibicion_incentivos",False)) else "No")
    y_actual -= 14; c.setFillColor(black)

    # F.7
    pF7 = Paragraph("   F.7 Prohibición de operaciones con recursos de procedencia ilícita (LD/FT)", pstyle_pregunta)
    w, h = pF7.wrapOn(c, max_w, height); need = h + 6 + 14
    if y_actual - need < min_y: c.showPage(); c.setFillColor(black); y_actual = 750
    pF7.drawOn(c, 40, y_actual - h); y_actual -= (h + 6)
    c.setFont("Helvetica",10); c.setFillColor(azul_claro)
    c.drawString(60, y_actual, "Sí" if bool(getattr(dd,"prohibicion_lavado_dinero",False)) else "No")
    y_actual -= 14; c.setFillColor(black)

    # ========== G ==========
    pregG = ("G. ¿Cuenta con un manual de organización y procedimientos que describan las funciones "
            "y responsabilidades de las áreas?")
    pG = Paragraph(pregG, pstyle_pregunta)
    wG, hG = pG.wrapOn(c, max_w, height)
    need = hG + 6 + 14
    if y_actual - need < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    pG.drawOn(c, 40, y_actual - hG); y_actual -= (hG + 6)
    c.setFont("Helvetica", 10); c.setFillColor(azul_claro)
    c.drawString(60, y_actual, "Sí" if bool(getattr(dd, "manual_organizacion", getattr(dd, "manual_organización", False))) else "No")
    y_actual -= 14; c.setFillColor(black)

    # ========== H ==========
    pregH = "H. ¿Verifica el perfil ético de sus empleados antes de su contratación? explique cómo"
    pH = Paragraph(pregH, pstyle_pregunta)
    wH, hH = pH.wrapOn(c, max_w, height)
    need = hH + 6 + 14
    if y_actual - need < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    pH.drawOn(c, 40, y_actual - hH); y_actual -= (hH + 6)

    val_h = bool(getattr(dd, "verifica_perfil_etico", False))
    c.setFont("Helvetica", 10); c.setFillColor(azul_claro)
    c.drawString(60, y_actual, "Sí" if val_h else "No")
    y_actual -= 14; c.setFillColor(black)

    if val_h:
        desc_h = (getattr(dd, "descripcion_verificacion", "") or "").strip()
        if desc_h:
            par_h = Paragraph(desc_h, pstyle_respuesta)
            w, h = par_h.wrapOn(c, max_w - 20, height)
            if y_actual - h < min_y:
                c.showPage(); c.setFillColor(black); y_actual = 750
            par_h.drawOn(c, 60, y_actual - h)
            y_actual -= (h + 8)

    # ========== I ==========
    pregI = "I. ¿Brinda capacitación a su personal en temas relacionados al combate a la corrupción?"
    pI = Paragraph(pregI, pstyle_pregunta)
    wI, hI = pI.wrapOn(c, max_w, height)
    need = hI + 6 + 14
    if y_actual - need < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    pI.drawOn(c, 40, y_actual - hI); y_actual -= (hI + 6)
    c.setFont("Helvetica", 10); c.setFillColor(azul_claro)
    c.drawString(60, y_actual, "Sí" if bool(getattr(dd, "capacitacion_anticorrupcion", getattr(dd, "capacitacion_anticorrupción", False))) else "No")
    y_actual -= 14; c.setFillColor(black)

    # ========== J ==========
    pregJ = "J. ¿Cuenta con algún medio o mecanismo institucional para denunciar Hechos de Corrupción?"
    pJ = Paragraph(pregJ, pstyle_pregunta)
    wJ, hJ = pJ.wrapOn(c, max_w, height)
    need = hJ + 6 + 14
    if y_actual - need < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    pJ.drawOn(c, 40, y_actual - hJ); y_actual -= (hJ + 6)

    val_j = getattr(dd, "medio_denuncia", None)
    if isinstance(val_j, bool):
        c.setFont("Helvetica", 10); c.setFillColor(azul_claro)
        c.drawString(60, y_actual, "Sí" if val_j else "No")
        y_actual -= 14; c.setFillColor(black)
    else:
        par_j = Paragraph(str(val_j or ""), pstyle_respuesta)
        w, h = par_j.wrapOn(c, max_w - 20, height)
        if y_actual - h < min_y:
            c.showPage(); c.setFillColor(black); y_actual = 750
        par_j.drawOn(c, 60, y_actual - h)
        y_actual -= (h + 8)

    # ========== K ==========
    pregK = ("K. ¿En caso de recibirse una denuncia, se le da seguimiento hasta obtener resultados? "
            "¿Por qué medio se da seguimiento?")
    pK = Paragraph(pregK, pstyle_pregunta)
    wK, hK = pK.wrapOn(c, max_w, height)
    need = hK + 6 + 14
    if y_actual - need < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    pK.drawOn(c, 40, y_actual - hK); y_actual -= (hK + 6)

    val_k = bool(getattr(dd, "seguimiento_denuncia", getattr(dd, "guimiento_denuncia", False)))
    c.setFont("Helvetica", 10); c.setFillColor(azul_claro)
    c.drawString(60, y_actual, "Sí" if val_k else "No")
    y_actual -= 14; c.setFillColor(black)

    if val_k:
        desc_k = (getattr(dd, "descripcion_seguimiento", "") or "").strip()
        if desc_k:
            par_k = Paragraph(desc_k, pstyle_respuesta)
            w, h = par_k.wrapOn(c, max_w - 20, height)
            if y_actual - h < min_y:
                c.showPage(); c.setFillColor(black); y_actual = 750
            par_k.drawOn(c, 60, y_actual - h)
            y_actual -= (h + 8)

    # ========== L ==========
    pregL = ("L. ¿Sus directivos hablan en forma clara sobre temas de corrupción y la responsabilidad "
            "de los empleados?")
    pL = Paragraph(pregL, pstyle_pregunta)
    wL, hL = pL.wrapOn(c, max_w, height)
    need = hL + 6 + 8  # 8 para el párrafo de texto
    if y_actual - need < min_y:
        c.showPage(); c.setFillColor(black); y_actual = 750
    pL.drawOn(c, 40, y_actual - hL); y_actual -= (hL + 6)

    directivos_corrupcion = dd.directivos_hablan_de_corrupcion
    if directivos_corrupcion:
        respuesta_directivos_corrupcion = "Sí"
        par_l = Paragraph(respuesta_directivos_corrupcion, pstyle_respuesta)
        w, h = par_l.wrapOn(c, max_w - 20, height)
        if y_actual - h < min_y:
            c.showPage(); c.setFillColor(black); y_actual = 750
        par_l.drawOn(c, 60, y_actual - h)
        y_actual -= (h + 8)
    else:
        c.setFont("Helvetica", 10); c.setFillColor(azul_claro)
        c.drawString(60, y_actual, "No")
        y_actual -= 14; c.setFillColor(black)

    
    c.showPage()   
    c.save()
    buf.seek(0)
    return buf


# -------- dibujar encabezado empresarial simple --------
def draw_header(c, cambio):
    width, height = letter
    # barra título
      #Azul Vordcab
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)
    caja_iso = 760
    
    #Elaborar caja
    #c.line(caja_iso,500,caja_iso,720)
    c.drawString(430,caja_iso,'Preparado por:')
    c.drawString(410,caja_iso-10,'SUPT. DE ADQUISIONES')
    c.drawString(530,caja_iso,'Aprobación')
    c.drawString(525,caja_iso-10,'SUBD ADTVO')

    c.drawString(110,caja_iso-20,'Número de documento')
    c.drawString(112,caja_iso-30,'SEOV-ADQ-N4-03.01')
    c.drawString(205,caja_iso-20,'Clasificación del documento')
    c.drawString(235,caja_iso-30,'Controlado')
    c.drawString(315,caja_iso-20,'Nivel del documento')
    c.drawString(340,caja_iso-30, 'N5')
    c.drawString(400,caja_iso-20,'Revisión No.')
    c.drawString(412,caja_iso-30,'003')
    c.drawString(450,caja_iso-20,'Fecha de Emisión')
    c.drawString(460,caja_iso-30,'22/02/2023')
    c.drawString(520,caja_iso-20,'Fecha de Revisión')
    c.drawString(530,caja_iso-30,'22/02/2023')

    caja_proveedor = caja_iso - 65
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(150,750,250,20, fill=True, stroke=False) #Barra azul superior Orden de Compra
    
    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(280,755,'Tabla de Precios de Referencia')
    c.drawInlineImage('static/images/logo_vordcab.jpg',20,730, 3 * cm, 1.5 * cm) #Imagen vortec

    fecha = (cambio.autorizado_en.date() if cambio.autorizado is not None and cambio.autorizado_en
             else cambio.solicitado_en.date() if cambio.solicitado_en
             else date.today())
    c.drawString(80, 705, fecha.strftime("%d/%m/%Y"))
    
    # ===== Badge de estatus (verde/rojo/blanco) =====
    status_txt = "APROBADO" if cambio.autorizado is True else \
                 "RECHAZADO" if cambio.autorizado is False else "PENDIENTE"
    bg = colors.HexColor('#16a34a') if cambio.autorizado is True else \
         colors.HexColor('#dc2626') if cambio.autorizado is False else colors.HexColor('#9ca3af')
    x_badge, y_badge, w_badge, h_badge = (width-160, 705-4, 120, 18)
    c.setFillColor(bg)
    c.roundRect(x_badge, y_badge, w_badge, h_badge, 3, fill=True, stroke=0)
    c.setFillColor(white)
    c.setFont('Helvetica-Bold',10)
    c.drawCentredString(x_badge + w_badge/2, y_badge + 4, status_txt)
    c.setFillColor(black)

# -------- dibujar la matriz (encabezados + 1 fila) --------
def draw_matrix(c, x_left, y_top, producto):
    styles = getSampleStyleSheet()
    H = ParagraphStyle('H', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, alignment=1)
    C = ParagraphStyle('C', parent=styles['Normal'], fontName='Helvetica', fontSize=8, alignment=1)
    L = ParagraphStyle('L', parent=styles['Normal'], fontName='Helvetica', fontSize=8, leading=10, alignment=0)

    # columnas
    col_widths = [
        1.1*cm,   # Item
        7*cm,   # Descripción
        2*cm,   # Servicio o Material
        2.6*cm,   # Marcar si es un Activo
        2*cm,   # Precio MXN
        2*cm,   # Precio USD (No aplica)
        2.5*cm,   # % tolerancia
    ]

    # encabezado (1 fila)
    head = [[
        Paragraph("Ítem", H),
        Paragraph("Descripción del producto", H),
        Paragraph("Servicio o<br/>Material", H),
        Paragraph("Activo", H),
        Paragraph("P.U. <br/> Pesos", H),
        Paragraph("P.U. dólares", H),
        Paragraph("Porcentaje de<br/>tolerancia", H),
    ]]

    t_head = Table(head, colWidths=col_widths, rowHeights=[1.3*cm])
    t_head.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.6, colors.black),
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F5F5F5')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))

    w, h_head = t_head.wrapOn(c, 0, 0)
    t_head.drawOn(c, x_left, y_top - h_head)

    # datos del producto → UNA FILA
    # Ajusta descripción como necesites; aquí uso nombre y un ejemplo de nota
    descripcion = producto.product.nombre or ""
    serv_mat = 'Servicio' if producto.product.servicio else "Material"  # si lo necesitas dinámico, cambia por tu lógica
    activo_mark = "Sí" if producto.product.activo else "No"     # idem: si es activo o no
    precio_mxn = f"${producto.new_value:,.2f}" if producto.new_value is not None else ""
    precio_usd = "No aplica"
    pct_tol = f"{producto.new_porcentaje:.2f} %" if producto.new_porcentaje is not None else ""

    row = [[
        Paragraph("1", C),
        Paragraph(descripcion, L),
        Paragraph(serv_mat, C),
        Paragraph(activo_mark, C),
        Paragraph(precio_mxn, C),
        Paragraph(precio_usd, C),
        Paragraph(pct_tol, C),
    ]]

    t_row = Table(row, colWidths=col_widths, rowHeights=[1.0*cm])
    t_row.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('LEFTPADDING', (1,0), (1,0), 4),
        ('ALIGN', (1,0), (1,0), 'LEFT'),
    ]))
    w, h_row = t_row.wrapOn(c, 0, 0)
    t_row.drawOn(c, x_left, y_top - h_head - 6 - h_row)

    # devuelve la altura consumida, por si quieres continuar
    return h_head + 6 + h_row

def draw_signatures(c, cambio, y_bottom=140):
    width, _ = letter
    col_w = (width - 80) / 2.0
    x1 = 40
    x2 = 40 + col_w + 20
    y = y_bottom

    # línea + nombre solicitante
    c.setStrokeColor(colors.black)
    #c.line(x1, y, x1 + col_w, y)
    c.setFont('Helvetica-Bold',9)
    solicitante = f"{cambio.solicitado_por.staff.staff.first_name} {cambio.solicitado_por.staff.staff.last_name}"
    c.drawCentredString(x1 + col_w/2, y - 12, solicitante)
    c.setFont('Helvetica',8)
    c.drawCentredString(x1 + col_w/2, y - 26, "Solicita")

    fecha_sol = cambio.solicitado_en.strftime("%d/%m/%Y %H:%M")
    c.drawCentredString(170, 100, fecha_sol)

    # línea + nombre autorizador (puede estar vacío si está pendiente)
    #c.line(x2, y, x2 + col_w, y)
    c.setFont('Helvetica-Bold',9)
    autorizador = f"{cambio.autorizado_por.staff.staff.first_name} {cambio.autorizado_por.staff.staff.last_name}"
    c.drawCentredString(x2 + col_w/2, y - 12, autorizador or " ")
    c.setFont('Helvetica',8)
    c.drawCentredString(x2 + col_w/2, y - 26, "Autoriza")

    fecha_aut = cambio.autorizado_en.strftime("%d/%m/%Y %H:%M")
    c.drawCentredString(455, 100, fecha_aut)


def pdf_precio_referencia(request, product_id):
  
    producto = PriceRefChange.objects.get(pk=product_id)
    

    # ------------- crear PDF -------------
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    width, height = letter

    # encabezado simple
    draw_header(c, producto)

    # matriz
    top_y = height - 120  # posición vertical de inicio de la tabla
    draw_matrix(c, x_left=20, y_top=top_y, producto=producto)

    
    # Firmas (abajo)
    draw_signatures(c, producto, y_bottom=140)

    c.showPage()
    c.save()
    buf.seek(0)
    filename = f"precio_ref_{producto.product.codigo or producto.id}.pdf"
    return FileResponse(buf, as_attachment=False, filename=filename)