from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect, FileResponse
from django.views.decorators.cache import cache_page
from django.db.models import F, Avg, Value, ExpressionWrapper, fields, Sum, Q, DateField, Count, Case, When, Value, DecimalField
from django.db.models.functions import Concat, Coalesce
from django.utils import timezone
from django.contrib import messages
from django.core.files.base import ContentFile
from django.core.mail import EmailMessage, BadHeaderError
from django.core.exceptions import ObjectDoesNotExist
from smtplib import SMTPException
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from django.urls import reverse
from .tasks import convert_excel_matriz_compras_task, convert_excel_solicitud_matriz_productos_task, convert_excel_solicitud_matriz_productos_task2
from dashboard.models import Inventario, Activo, Order, ArticulosOrdenados, ArticulosparaSurtir, Producto_Calidad
from requisiciones.models import Requis, ArticulosRequisitados
from user.models import Profile
from tesoreria.models import Pago, Facturas
from user.decorators import perfil_seleccionado_required, tipo_usuario_requerido
from .filters import CompraFilter, ArticulosRequisitadosFilter,  ArticuloCompradoFilter, HistoricalArticuloCompradoFilter, HistoricalCompraFilter, ComparativoFilter
from .models import ArticuloComprado, Compra, Proveedor_direcciones, Cond_pago, Uso_cfdi, Moneda, Comparativo, Item_Comparativo, Proveedor
from .forms import CompraForm, ArticuloCompradoForm, ArticulosRequisitadosForm, ComparativoForm, Item_ComparativoForm, Compra_ComentarioForm, UploadFileForm, Compra_ComentarioGerForm, RequisDevolucionForm
from requisiciones.forms import Articulo_Cancelado_Form
from requisiciones.filters import RequisFilter
from tesoreria.forms import Facturas_Form
from requisiciones.views import get_image_base64
from django.utils.timezone import make_aware, is_aware
import pytz
import socket

import json
import time
import os
import io
import ssl
import decimal
from io import BytesIO
from datetime import date, datetime, timedelta
from num2words import num2words
import time

#PDF generator
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.colors import Color, black, blue, red, white
from reportlab.lib.units import cm
from reportlab.lib.pagesizes import letter
from reportlab.rl_config import defaultPageSize

from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Frame, PageBreak
from bs4 import BeautifulSoup

import urllib.request, urllib.parse, urllib.error

# Import Excel Stuff
import xlsxwriter
from xlsxwriter.utility import xl_col_to_name


from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill
from openpyxl.utils import get_column_letter
import datetime as dt
from user.logger_config import get_custom_logger

from pyexcelerate import Workbook, Color as PXColor, Style, Font, Fill, Alignment, Format


logger = get_custom_logger(__name__)
#from urllib.parse import (
#    ParseResult,
#    SplitResult,
#    _coerce_args,
#    _splitnetloc,
#    _splitparams,
#    scheme_chars,
#)
#from urllib.parse import urlencode as original_urlencode
#from urllib.parse import uses_params

# Create your views here.

@login_required(login_url='user-login')
@perfil_seleccionado_required
def requisiciones_autorizadas(request):
    pk = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk)
    if perfil.tipo.compras == True:
        requis = Requis.objects.filter(orden__distrito = perfil.distritos, autorizar=True, colocada=False, complete = True, devuelta = False).order_by('-approved_at')
    else:
        requis = Requis.objects.filter(autorizar=True, colocada=False, complete =True)
    #requis = Requis.objects.filter(autorizar=True, colocada=False)

    myfilter = RequisFilter(request.GET, queryset=requis)
    requis = myfilter.qs

    tag = dof()

     #Set up pagination
    p = Paginator(requis, 50)
    page = request.GET.get('page')
    requis_list = p.get_page(page)

    context= {
        'perfil':perfil,
        'myfilter': myfilter,
        'requis':requis,
        'tags':tag,
        'requis_list':requis_list,
        }

    return render(request, 'compras/requisiciones_autorizadas.html',context)


@perfil_seleccionado_required
def requis_devolver(request, pk):
    requis = get_object_or_404(Requis, pk=pk)
    form = RequisDevolucionForm(request.POST, instance=requis)
    if request.method == 'POST':
        if 'btn_devolver' in request.POST: 
            if form.is_valid():
                obj = form.save(commit=False)
                obj.devuelta = True
                if not obj.fecha_devolucion:
                    obj.fecha_devolucion = timezone.now().date()
                obj.save()
                messages.success(request, f"Requis #{requis.folio} devuelta.")
            else:
                messages.error(request, "Revisa el formulario de devolución.")
            return redirect('requisicion-autorizada')
    
    context= {
        'requis':requis,
        'form':form,
    }
    
    return render(request, 'compras/requis_devolver.html',context)


@login_required(login_url='user-login')
@perfil_seleccionado_required
def productos_pendientes(request):
    pk_perfil = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id = pk_perfil)
    if perfil.tipo.compras == True:
        requis = Requis.objects.filter(autorizar=True, colocada=False)
    else:
        requis = Requis.objects.filter(complete=None)

    articulos = ArticulosRequisitados.objects.filter(req__autorizar = True, req__colocada=False, cancelado = False).order_by('-req__orden__folio')
    myfilter = ArticulosRequisitadosFilter(request.GET, queryset=articulos)
    articulos = myfilter.qs

    #Set up pagination
    p = Paginator(articulos, 50)
    page = request.GET.get('page')
    articulos_list = p.get_page(page)

    context= {
        'requis':requis,
        'articulos':articulos,
        'articulos_list':articulos_list,
        'myfilter':myfilter,
        }

    return render(request, 'compras/productos_pendientes.html',context)


@login_required(login_url='user-login')
@perfil_seleccionado_required
def eliminar_articulos(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id=pk_perfil)
    requis = Requis.objects.get(id=pk)
    productos = ArticulosRequisitados.objects.filter(req=pk, cantidad_comprada__lt=F("cantidad"), cancelado=False).annotate(
        cantidad_restante=ExpressionWrapper(F("cantidad") - F("cantidad_comprada"),output_field=DecimalField(max_digits=14, decimal_places=2)))
    form = Articulo_Cancelado_Form()

    if request.method == 'POST' and 'btn_eliminar' in request.POST:
        # Procesar los checkboxes de los productos seleccionados
        fecha_hora = datetime.now()
        productos_eliminados = []
        for producto in productos:
            checkbox_name = f'seleccionar_producto_{producto.id}'
            if checkbox_name in request.POST:
                producto.cancelado = True
                producto.comentario_cancelacion = request.POST.get(f'comentario_producto_{producto.id}', '')
                productos_eliminados.append({
                    'nombre': producto.producto.articulos.producto.producto.nombre,
                    'cantidad': producto.cantidad,
                    'comentario': producto.comentario_cancelacion,
                })
            else:
                producto.cancelado = False
            producto.save()
        if productos_eliminados:
            productos_eliminados_html = "".join(
                f"<tr>"
                f"<td>{producto['nombre']}</td>"
                f"<td>{producto['cantidad']}</td>"
                f"<td>{producto['comentario']}</td>"
                f"</tr>"
                for producto in productos_eliminados
            )
        else:
            productos_eliminados_html = "<tr><td colspan='3'>Ningún producto eliminado.</td></tr>"
        #static_path = settings.STATIC_ROOT
        #img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
        #img_path2 = os.path.join(static_path,'images','logo_vordcab.jpg')
        #image_base64 = get_image_base64(img_path)
        #logo_v_base64 = get_image_base64(img_path2)
        # Actualizar el estado de la requisición
        productos_actualizados = ArticulosRequisitados.objects.filter(req=pk, cantidad_comprada__lt=F("cantidad"))
        productos_cancelados = productos_actualizados.filter(cancelado=True).count()
        productos_requisitados = productos_actualizados.count()
        productos_comprados = productos_actualizados.filter(art_surtido=True).count() + productos_cancelados
        if productos_requisitados == productos_comprados:
            requis.colocada = True
            requis.save()
        # Enviar correo electrónico con formato HTML
        correo_html = f"""
        <p>Estimado(a) {requis.orden.staff.staff.staff.first_name}:</p>
        <p>Estás recibiendo este correo porque los siguientes productos han sido eliminados de la solicitud:</p>
        <p><b>Folio de orden:</b> {requis.orden.folio} <br>
        <b>Folio de requisición:</b> {requis.folio}</p>
        <table border="1" cellpadding="5" cellspacing="0" style="border-collapse: collapse; width: 100%;">
            <thead>
                <tr>
                    <th>Producto</th>
                    <th>Cantidad</th>
                    <th>Comentario</th>
                </tr>
            </thead>
            <tbody>
                {productos_eliminados_html}
            </tbody>
        </table>
        <p>Atte.<br>
        {perfil.staff.staff.first_name} {perfil.staff.staff.last_name}<br>
        GRUPO VORDCAB S.A. de C.V.</p>
        <p><i>Este mensaje ha sido automáticamente generado por SAVIA VORDCAB.</i></p>
        """
        try:
            email = EmailMessage(
                f'Producto Eliminado',
                correo_html,
                settings.DEFAULT_FROM_EMAIL,
                ['ulises_huesc@hotmail.com', requis.orden.staff.staff.staff.email],
            )
            email.content_subtype = "html"  # Indicar que el contenido es HTML
            email.send()
            messages.success(request,f' Has eliminado el producto correctamente')
        except (BadHeaderError, SMTPException, socket.gaierror) as e:
            error_message = f'{perfil.staff.staff.first_name}, Has eliminado el producto correctamente pero el correo de notificación no ha sido enviado debido a un error: {e}'
            messages.success(request, error_message)
        return redirect('requisicion-autorizada')
    context = {
        'form':form,
        'productos': productos,
        'requis': requis,
        }

    return render(request,'compras/eliminar_articulos.html', context)

@perfil_seleccionado_required
def articulos_restantes(request, pk):
    productos = ArticulosRequisitados.objects.filter(req = pk, cantidad_comprada__lt = F("cantidad"), cancelado=False)
    #productos = ArticulosRequisitados.objects.filter(req = pk, cantidad_comprada__lt = F("cantidad"))
    requis = Requis.objects.get(id = pk)
    # Leer parámetro ?next o usar HTTP_REFERER
    url_origen = request.GET.get('next') or request.META.get('HTTP_REFERER') or '/'

    context = {
        'productos': productos,
        'requis': requis,
        'url_origen': url_origen,
        }

    return render(request,'compras/articulos_restantes.html', context)

def dof():
#Trying to fetch DOF
    try:
        # Configurar el tiempo máximo de espera (en segundos)
        timeout = 2  # Ajusta el tiempo de espera según tus necesidades
        socket.setdefaulttimeout(timeout)
    
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE

        url = 'https://www.dof.gob.mx/#gsc.tab=0'
        html = urllib.request.urlopen(url, context=ctx).read()
        soup = BeautifulSoup(html,'html.parser')
        #tags = soup.find_all('p')

        tags = []
        for tag in soup.find_all('p'):
        #for anchor in tag.find_all('span'):
            tags.append(tag.contents)

        #substr = 'DOLAR'
        #if any(substr in str for str in tags):
        #   tag = tags[str][1]


        tag = tags[4][3]

        return tag
    except socket.timeout:
        return "Error: El tiempo de espera para la consulta ha sido superado."
    except Exception as e:
        # Manejo de la excepción - log, mensaje de error, etc.
        return f"Error al obtener datos: {e}"

@perfil_seleccionado_required
def oc(request, pk):
    productos = ArticulosRequisitados.objects.filter(req = pk)
    req = Requis.objects.get(id = pk)
    usuario = Profile.objects.get(id=request.user.id)
    oc, created = Compra.objects.get_or_create(complete = False, req = req, creada_por = usuario)
    form_product = ArticuloCompradoForm()
    form = CompraForm(instance=oc)


    context= {
        'req':req,
        'form':form,
        'oc':oc,
        'productos':productos,
        'form_product':form_product,
        }

    return render(request, 'compras/oc.html',context)

@perfil_seleccionado_required
def compras_devueltas(request):
    pk_perfil = request.session.get('selected_profile_id')
    colaborador = Profile.objects.all()
    usuario = colaborador.get(id = pk_perfil)
    #usuario = Profile.objects.get(staff__id=request.user.id)
    compras = Compra.objects.filter(regresar_oc = True, req__orden__distrito = usuario.distritos)
    myfilter = CompraFilter(request.GET, queryset=compras)
    compras = myfilter.qs
    
    #form_product = ArticuloCompradoForm()
    #form = CompraForm(instance=oc)
   


    context= {
        'myfilter':myfilter,
        'compras_list':compras,
        }

    return render(request, 'compras/compras_devueltas.html',context)

@perfil_seleccionado_required
def compra_edicion(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    colaborador = Profile.objects.all()
    usuario = colaborador.get(id = pk_perfil)
    oc = get_object_or_404(Compra, id=pk)
    #colaborador_sel = Profile.objects.all()
    productos_comp = ArticuloComprado.objects.filter(oc = oc)
    productos = ArticulosRequisitados.objects.filter(req = oc.req, sel_comp = False, cancelado = False)
    req = Requis.objects.get(id = oc.req.id)
    comparativos = Comparativo.objects.filter(creada_por__distritos = usuario.distritos, completo =True,)
    if oc.comparativo_model:
        try:
            comparativo_inicial = oc.comparativo_model.id
        except Comparativo.DoesNotExist:
            comparativo_inicial = 'null' 
        except Comparativo.MultipleObjectsReturned:
            # Manejo opcional si hay múltiples objetos (esto no debería suceder si la lógica está bien diseñada)
            comparativo_inicial = 'null' 
    else:
        comparativo_inicial = 'null'
    #proveedores = Proveedor_direcciones.objects.filter(
    #    Q(estatus__nombre='NUEVO') | Q(estatus__nombre='APROBADO'))
    if not (oc.complete == False and oc.regresar_oc == True):
        logger.warning(f"Intento acceso no autorizado a compra edición por usuario  {request.user.first_name} {request.user.last_name}")
        return render(request,'partials/acceso_denegado.html') 
    else:
        proveedores = Proveedor_direcciones.objects.filter(Q(estatus__nombre="NUEVO") | Q(estatus__nombre="APROBADO"),distrito = usuario.distritos,)
        if proveedores.filter(id = oc.proveedor.id):
            proveedor_obj = proveedores.get(id=oc.proveedor.id)
            proveedor_inicial = {
                'id': proveedor_obj.id,
                'text': proveedor_obj.nombre.razon_social,
                'distrito': proveedor_obj.distrito.nombre,
                'status': proveedor_obj.estatus.nombre,
                'domicilio': proveedor_obj.domicilio
            }
        else:
            proveedor_inicial = 'null'
        form_product = ArticuloCompradoForm()
        form = CompraForm(instance=oc)
        error_messages = {}
        #'distrito__nombre','domicilio','estatus__nombre'
        proveedor_para_select2 = [
            {'id': proveedor.id, 
            'text': proveedor.nombre.razon_social,
            'distrito': proveedor.distrito.nombre,
            'status': proveedor.estatus.nombre,
            'domicilio': proveedor.domicilio
            #'distrito': proveedor.
            } for proveedor in proveedores]


        productos_para_select2 = [
            {'id': producto.id,
            'text': str(producto), 
            'cantidad': str(producto.cantidad), 
            'cantidad_pendiente': str(producto.cantidad_comprada),
            'precioref': str(producto.producto.articulos.producto.producto.precioref),
            'porcentaje': str(producto.producto.articulos.producto.producto.porcentaje)
            } for producto in productos]
        
        productos_comp_to_function = [
            {
                'id': producto.id,
                'precio': str(producto.precio_unitario),
                'precio_ref': str(producto.producto.producto.articulos.producto.producto.precioref),
                'porcentaje': str(producto.producto.producto.articulos.producto.producto.porcentaje)
            } for producto in productos_comp
        ] 


        comparativos_para_select2 = [
            {
                'id': comparativo.id, 
                'text': str(comparativo.nombre)
            } for comparativo in comparativos
        ]



        tag = dof()
        subtotal = 0
        iva = 0
        total = 0
        dif_cant = 0
        #form.fields['deposito_comprador'].queryset = colaborador_sel
        for item in productos_comp:
            subtotal = decimal.Decimal(subtotal + item.cantidad * item.precio_unitario)
            if item.producto.producto.articulos.producto.producto.iva == True:
                iva = round(subtotal * decimal.Decimal(0.16),2)
            total = decimal.Decimal(subtotal + decimal.Decimal(iva))

        if request.method == 'POST' and  "crear" in request.POST:
            form = CompraForm(request.POST, instance=oc)
            costo_oc = 0
            costo_iva = 0
            articulos = ArticuloComprado.objects.filter(oc=oc)
            requisitados = ArticulosRequisitados.objects.filter(req = oc.req)
            cuenta_art_comprados = requisitados.filter(art_surtido = True).count()
            cuenta_art_totales = requisitados.count()
           
           
            for producto in requisitados:
                dif_cant = dif_cant + producto.cantidad - producto.cantidad_comprada
                if producto.art_surtido == False:
                    producto.sel_comp = False
                    producto.save()
            if form.is_valid():
                #validación que comprueba si los art_comprados son igual a los articulos a los requisitados
                if cuenta_art_totales == cuenta_art_comprados and cuenta_art_comprados > 0:
                    req.colocada = True
                else:
                    req.colocada = False
                #for articulo in articulos:
                #    costo_oc = costo_oc + articulo.precio_unitario * articulo.cantidad
                #    if articulo.producto.producto.articulos.producto.producto.iva == True:
                #        costo_iva = decimal.Decimal(costo_oc * decimal.Decimal(0.16))
                oc = form.save(commit = False)
                oc.complete = True
                oc.costo_iva = iva
                oc.costo_oc = total
                oc.regresar_oc = False
                oc.save()
                req.save()
                messages.success(request,f'{usuario.staff.staff.first_name}, Has modificado la OC {oc.folio} correctamente')
                return redirect('compras-devueltas')
            else:
                for field, errors in form.errors.items():
                    error_messages[field] = errors.as_text()



        context= {
            'comparativos_para_select2': comparativos_para_select2,
            'comparativo_inicial':comparativo_inicial,
            'proveedor_para_select2': proveedor_para_select2,
            'proveedor_inicial':proveedor_inicial,
            'productos_comp_to_function': productos_comp_to_function,
            'error_messages': error_messages,
            'req':req,
            'form':form,
            'form_product':form_product,
            'productos_para_select2':productos_para_select2,
            #'proveedores':proveedores,
            'productos':productos,
            'oc':oc,
            'productos_comp':productos_comp,
            'subtotal':subtotal,
            'iva':iva,
            'total':total,
            }

        return render(request, 'compras/compra_edicion.html',context)


@perfil_seleccionado_required
def update_oc(request):
    data= json.loads(request.body)
    action = data["action"]
    cantidad = data["val_cantidad"]
    producto_id = data["id"]
    pk = data["oc"]
    productos = ArticulosRequisitados.objects.get(id=producto_id)
    precio = data["val_precio"]
    oc = Compra.objects.get(id=pk)
    if action == "add":
        cantidad_total = productos.cantidad_comprada + decimal.Decimal(cantidad)
        if cantidad_total > productos.cantidad:
            messages.error(request,f'La cantidad que se quiere comprar sobrepasa la cantidad requisitada {cantidad_total} mayor que {productos.cantidad}')
        else:
            comp_item, created = ArticuloComprado.objects.get_or_create(oc=oc, producto=productos)
            productos.cantidad_comprada = productos.cantidad_comprada + decimal.Decimal(cantidad)
            #messages.success(request,f'Estos son los productos comprados ahora {productos.cantidad_comprada}')
            if productos.cantidad_comprada == productos.cantidad:
                productos.art_surtido = True
            if comp_item.cantidad == None:
                comp_item.cantidad = 0
            comp_item.cantidad = comp_item.cantidad + decimal.Decimal(cantidad)
            comp_item.precio_unitario = precio
            productos.sel_comp = True
            comp_item.save()
            productos.save()
            response_data = {
             'codigo': comp_item.producto.producto.articulos.producto.producto.nombre,
             'producto': comp_item.producto.producto.articulos.producto.producto.codigo,
             'cantidad': comp_item.cantidad,
             'precio': comp_item.precio_unitario,
            }

    if action == "remove":
        comp_item = ArticuloComprado.objects.get(oc = oc, producto = productos)
        productos.art_surtido = False
        productos.sel_comp = False
        productos.cantidad_comprada = productos.cantidad_comprada - comp_item.cantidad
        productos.save()
        comp_item.delete()
        response_data = {
            'item':action,
        }

    return JsonResponse(response_data)

#@cache_page(60 * 60)  # Cache la vista durante 1 hora
@perfil_seleccionado_required
def oc_modal(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    colaborador = Profile.objects.all()
    usuario = colaborador.get(id = pk_perfil)
    req = Requis.objects.get(id = pk)
    
    productos = ArticulosRequisitados.objects.filter(req = pk, cantidad_comprada__lt = F("cantidad"), cancelado=False, sel_comp = False)
    compras = Compra.objects.all()

    #Traigo el folio de la ultima compra con el folio más grande, no importa este complete o no  
    last_oc = compras.filter(req__orden__distrito = req.orden.distrito).order_by('-folio').first()
    new_folio = last_oc.folio + 1 if last_oc is not None else 1
    #Hago un get or create para traer el ultimo dato a llenar y si este es un create entonces se le asigna el nuevo valor para el folio 
    oc, created = compras.get_or_create(complete = False, req = req, creada_por = usuario, regresar_oc = False, defaults={'folio': new_folio})
    # Esta lógica es para aquellos que ya se crearon pero no tienen folio, se les asigna el nuevo folio
    if not oc.folio:
        oc.folio = new_folio
        oc.save()
    folio_preview = oc.folio
    
    comparativos = Comparativo.objects.filter(creada_por__distritos = usuario.distritos, completo =True)
    productos_comp = ArticuloComprado.objects.filter(oc=oc)
    form = CompraForm(instance=oc)
    form_product = ArticuloCompradoForm()
    if (req.colocada != False):
        logger.warning(f"Intento acceso no autorizado a compra cuya requisición {req.folio} por usuario  {usuario.staff.staff.first_name} {usuario.staff.staff.last_name}")
        return render(request,'partials/acceso_denegado.html') 
    else:
        tag = dof()
        subtotal = 0
        iva = 0
        total = 0
        dif_cant = 0

        #last_oc = compras.filter(complete = True, req__orden__distrito = req.orden.distrito).order_by('-folio').first()
        #if last_oc:
        #    folio = last_oc.folio + 1
        #else:
        #    folio = 1
        #folio_preview = folio
            
        #abrev = req.orden.distrito.abreviado
        error_messages = {}

        productos_para_select2 = [
            {'id': producto.id,
            'text': str(producto), 
            'cantidad': str(producto.cantidad), 
            'cantidad_pendiente': str(producto.cantidad_comprada),
            'precioref': str(producto.producto.articulos.producto.producto.precioref),
            'porcentaje': str(producto.producto.articulos.producto.producto.porcentaje),
            'comentario': str(producto.producto.articulos.comentario)
            } for producto in productos]
        
        productos_comp_to_function = [
            {
                'id': producto.id,
                'precio': str(producto.precio_unitario),
                'precio_ref': str(producto.producto.producto.articulos.producto.producto.precioref),
                'porcentaje': str(producto.producto.producto.articulos.producto.producto.porcentaje)
            } for producto in productos_comp
        ] 
        comparativos_para_select2 = [
            {
                'id': comparativo.id, 
                'text': str(comparativo.nombre)
            } for comparativo in comparativos
        ]

        for item in productos_comp:
            subtotal = decimal.Decimal(subtotal + item.cantidad * item.precio_unitario)
            if item.producto.producto.articulos.producto.producto.iva == True:
                iva = round(subtotal * decimal.Decimal(0.16),2)
            total = decimal.Decimal(subtotal + decimal.Decimal(iva))

        if request.method == 'POST' and  "crear" in request.POST:
            form = CompraForm(request.POST, instance=oc)
            
            if form.is_valid():
                costo_oc = 0
                costo_iva = 0
                articulos = ArticuloComprado.objects.filter(oc=oc)
                requisitados = ArticulosRequisitados.objects.filter(req = oc.req)
                borrados = requisitados.filter(cancelado = True).count()
                cuenta_art_comprados = requisitados.filter(art_surtido = True).count() #Añadir aquí que sume los articulos que hayan sido cancelados con el borrador para que la saque de la vista
                cuenta_art_comprados += borrados
                cuenta_art_totales = requisitados.count()
                if cuenta_art_totales == cuenta_art_comprados and cuenta_art_comprados > 0: #Compara los artículos comprados vs artículos requisitados
                    req.colocada = True
                else:
                    req.colocada = False
                for articulo in articulos:
                    costo_oc = costo_oc + articulo.precio_unitario * articulo.cantidad
                    if articulo.producto.producto.articulos.producto.producto.iva == True:
                        costo_iva = decimal.Decimal(costo_oc * decimal.Decimal(0.16))
                for producto in requisitados:
                    dif_cant = dif_cant + producto.cantidad - producto.cantidad_comprada
                    if producto.art_surtido == False:
                        producto.sel_comp = False
                        producto.save()
            
                if oc.tipo_de_cambio != None and oc.tipo_de_cambio > 0:
                    oc.costo_iva = decimal.Decimal(costo_iva)
                    oc.costo_oc = decimal.Decimal(costo_oc + costo_iva)
                else:
                    oc.costo_iva = decimal.Decimal(costo_iva)
                    oc.costo_oc = decimal.Decimal(costo_oc + costo_iva)

                #last_oc = Compra.objects.filter(complete = True, req__orden__distrito = req.orden.distrito).order_by('-folio').first()
                #if last_oc:
                #    folio = last_oc.folio + 1
                #else:
                #    folio = 1
                oc = form.save(commit = False)
                oc.complete = True
                #oc.folio = folio ############
                oc.created_at = date.today()
                #form.save()
                oc.save()
                req.save()
                static_path = settings.STATIC_ROOT
                img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
                img_path2 = os.path.join(static_path,'images','logo_vordcab.jpg')
        
                image_base64 = get_image_base64(img_path)
                logo_v_base64 = get_image_base64(img_path2)
                # Construir el HTML para la lista de artículos
                articulos_html = """
                <table border="1" style="border-collapse: collapse; width: 100%;">
                    <thead>
                        <tr>
                            <th>Artículo</th>
                            <th>Cantidad</th>
                            <th>Observación</th>
                        </tr>
                    </thead>
                    <tbody>
                """

                for articulo in articulos:
                    articulos_html += f"""
                        <tr>
                            <td>{articulo.producto.producto.articulos.producto.producto.nombre}</td>
                            <td>{articulo.cantidad}</td>
                            <td>{articulo.producto.producto.articulos.comentario}
                        </tr>
                    """

                articulos_html += """
                    </tbody>
                </table>
                """


                # Crear el mensaje HTML
                html_message = f"""
                <html>
                    <head>
                        <meta charset="UTF-8">
                    </head>
                    <body style="font-family: Arial, sans-serif; color: #333; background-color: #f4f4f4; margin: 0; padding: 0;">
                        <table width="100%" cellspacing="0" cellpadding="0" style="background-color: #f4f4f4; padding: 20px;">
                            <tr>
                                <td align="center">
                                    <table width="600px" cellspacing="0" cellpadding="0" style="background-color: #ffffff; padding: 20px; border-radius: 10px;">
                                        <tr>
                                            <td align="center">
                                                <img src="data:image/jpeg;base64,{logo_v_base64}" alt="Logo" style="width: 100px; height: auto;" />
                                            </td>
                                        </tr>
                                        <tr>
                                            <td style="padding: 20px;">
                                                <p style="font-size: 18px; text-align: justify;">
                                                    <p>Estimado {oc.req.orden.staff.staff.staff.first_name} {oc.req.orden.staff.staff.staff.last_name},</p>
                                                </p>
                                                <p style="font-size: 16px; text-align: justify;">
                                                    Estás recibiendo este correo porque tu solicitud: {oc.req.orden.folio}| Req: {oc.req.folio} se ha convertido en la OC: {oc.folio}, creada por {oc.creada_por.staff.staff.first_name} {oc.creada_por.staff.staff.last_name}.</p>
                                                </p>
                                            <p style="font-size: 16px; text-align: justify;">
                                                El siguiente paso del sistema: Autorización de OC por Superintedencia Administrativa.
                                            </p>
                                            {articulos_html}
                                                <p style="text-align: center; margin: 20px 0;">
                                                    <img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width: 50px; height: auto; border-radius: 50%;" />
                                                </p>
                                                <p style="font-size: 14px; color: #999; text-align: justify;">
                                                    Este mensaje ha sido automáticamente generado por SAVIA 2.0
                                                </p>
                                            </td>
                                        </tr>
                                    </table>
                                </td>
                            </tr>
                        </table>
                    </body>
                </html>
                """
                try:
                    email = EmailMessage(
                        f'OC Elaborada {oc.folio}',
                        body=html_message,
                        #f'Estimado {requi.orden.staff.staff.staff.first_name} {requi.orden.staff.staff.staff.last_name},\n Estás recibiendo este correo porque tu solicitud: {requi.orden.folio}| Req: {requi.folio} ha sido autorizada,\n por {requi.requi_autorizada_por.staff.staff.first_name} {requi.requi_autorizada_por.staff.staff.last_name}.\n El siguiente paso del sistema: Generación de OC \n\n Este mensaje ha sido automáticamente generado por SAVIA VORDTEC',
                        from_email = settings.DEFAULT_FROM_EMAIL,
                        to= [oc.req.orden.staff.staff.staff.email,],
                        headers={'Content-Type': 'text/html'}
                        )
                    email.content_subtype = "html " # Importante para que se interprete como HTML
                    email.send()
                    messages.success(request,f'{usuario.staff.staff.first_name}, Has generado la OC {oc.folio} correctamente')
                except (BadHeaderError, SMTPException, socket.gaierror) as e:
                    error_message = f'{usuario.staff.staff.first_name}, Has generado la OC {oc.folio} correctamente pero el correo de notificación no ha sido enviado debido a un error: {e}'
                    messages.success(request, error_message)
                return redirect('requisicion-autorizada')
            else:
                for field, errors in form.errors.items():
                    error_messages[field] = errors.as_text()
        


        
        context= {
            'comparativos_para_select2': comparativos_para_select2,
            'productos_comp_to_function': productos_comp_to_function,
            'error_messages': error_messages,
            'req':req,
            'form':form,
            'form_product':form_product,
            'productos_para_select2':productos_para_select2,
            'oc':oc,
            'folio':folio_preview,
            'productos':productos,
            'tag':tag,
            'productos_comp':productos_comp,
            'subtotal':subtotal,
            'iva':iva,
            'total':total,
            }
        
        return render(request, 'compras/oc.html', context)

@perfil_seleccionado_required
def mostrar_comparativo(request, pk):
    compra = Compra.objects.get(id=pk)
    comparativo = Comparativo.objects.get(id=compra.comparativo_model.id)
    productos = Item_Comparativo.objects.filter(comparativo = comparativo)
    
    context= {
        'comparativo':comparativo,
        'productos':productos,
        'compra':compra,
        }

    return render(request, 'compras/mostrar_comparativo.html',context)


@perfil_seleccionado_required
def matriz_oc(request):
    pk_perfil = request.session.get('selected_profile_id')
    colaborador_sel = Profile.objects.all()
    usuario = colaborador_sel.get(id = pk_perfil)
    almacenes_distritos = set(usuario.almacen.values_list('distrito__id', flat=True))
    if usuario.tipo.proveedores or usuario.tipo.nombre == "VIS_ADQ":
        print("Usuario con permiso de ver todas las OC")
        compras = Compra.objects.filter(complete = True, req__orden__distrito__id__in = almacenes_distritos).annotate(
            total_facturas=Count('facturas', filter=Q(facturas__hecho=True)),
            autorizadas=Count(Case(When(Q(facturas__autorizada=True, facturas__hecho=True), then=Value(1))))
            ).order_by('-folio')
    else:
        print("Usuario sin permiso de ver todas las OC")
        compras = Compra.objects.filter(complete=True, req__orden__distrito = usuario.distritos).annotate(
            total_facturas=Count('facturas', filter=Q(facturas__hecho=True)),
            autorizadas=Count(Case(When(Q(facturas__autorizada=True, facturas__hecho=True), then=Value(1))))
            ).order_by('-folio')
    
    myfilter = CompraFilter(request.GET, queryset=compras)
    compras = myfilter.qs
    print(compras)


    compras_data = list(compras.values())
    # Obtienes las fechas de inicio y finalización del filtro
    start_date = myfilter.form.cleaned_data.get('start_date')
    end_date = myfilter.form.cleaned_data.get('end_date')
   
    num_approved_requis = 0
    num_requis_atendidas = 0
    # Asegúrate de que start_date y end_date son objetos datetime "aware"
    if start_date is not None and end_date is not None:
    # Si las fechas no tienen información de la zona horaria, hazlas "aware"
        # Filtrar las requisiciones aprobadas dentro del rango de fechas
        if usuario.tipo.proveedores or usuario.tipo.nombre == "VIS_ADQ":
            approved_requis = Requis.objects.filter(approved_at__gte=start_date, approved_at__lte=end_date, autorizar = True)
        else:
            approved_requis = Requis.objects.filter(approved_at__gte=start_date, approved_at__lte=end_date, autorizar = True, orden__distrito = usuario.distritos)
        approved_requis_ids = approved_requis.values_list('id', flat=True)
        num_approved_requis = approved_requis.count() 

        # Contar el número de requisiciones aprobadas
        compras_colocadas_ids = Compra.objects.filter(
            created_at__gte=start_date, 
            created_at__lte=end_date, 
            req__colocada=True,
            req_id__in=approved_requis_ids,
            req__orden__distrito = usuario.distritos
        ).values_list('req', flat=True).distinct()

        num_requis_atendidas = len(set(compras_colocadas_ids))


    # Calcular el total de órdenes de compra
    total_de_oc = compras.count()
     # Calcular el número de OC que cumplen el criterio (created_at - approved_at <= 3)
    time_difference = ExpressionWrapper(F('created_at') - F('req__approved_at'), output_field=fields.DurationField())
    compras_con_criterio = compras.annotate(time_difference=time_difference).filter(time_difference__lte=timedelta(days=3))
    oc_cumplen = compras_con_criterio.count()

     # Calcular el indicador de cumplimiento (oc_cumplen / total_de_oc)
    #if total_de_oc > 0:
    #    cumplimiento = (oc_cumplen / total_de_oc)*100
    #else:
    #    cumplimiento = 0

     #Set up pagination
    p = Paginator(compras, 50)
    page = request.GET.get('page')
    compras_list = p.get_page(page)

    # Proceso para asignar estados y calcular la suma total de facturas
    for compra in compras_list:
        if compra.total_facturas == 0:
            compra.estado_facturas = 'sin_facturas'
        elif compra.autorizadas == compra.total_facturas:
            compra.estado_facturas = 'todas_autorizadas'
        else:
            compra.estado_facturas = 'pendientes'
        
        # Sumar totales de facturas relacionadas que cumplan con las condiciones
        compra.suma_total_facturas = sum(
            decimal.Decimal(factura.emisor['total'])
            for factura in compra.facturas.all()
            if factura.factura_xml and factura.hecho and factura.autorizada and factura.emisor is not None
        )
        
    context= {
        'usuario':usuario,
        #'num_approved_requis': num_approved_requis,
        'compras_list':compras_list,
        'compras':compras,
        'myfilter':myfilter,
        #'cumplimiento': cumplimiento,
        }
    
    
    #task_id = request.session.get('task_id')

    if request.method == 'POST' and 'btnExcel' in request.POST:
        #if compras.count() > 2400:
        #    if not task_id:
        #        task = convert_excel_matriz_compras_task.delay(compras_data, num_requis_atendidas, num_approved_requis, start_date, end_date)
        #        task_id = task.id
        #        request.session['task_id'] = task_id
        #        context['task_id'] = task_id
        #        cantidad = compras.count()
        #        context['cantidad'] = cantidad
        #        messages.success(request, f'Tu reporte se está generando {task_id}')
        #else:
        return convert_excel_matriz_compras(compras, num_requis_atendidas, num_approved_requis, start_date, end_date)
        
        

    

    return render(request, 'compras/matriz_compras.html',context)


#def generar_reporte(request):
    

from celery.result import AsyncResult
from celery.exceptions import CeleryError

def verificar_estado_productos(request):
    task_id = request.session.get('task_id_producto')  # Asumiendo que el task_id se pasa como parámetro GET

    if not task_id:
        return JsonResponse({'error': 'No se proporcionó task_id'}, status=400)

    try:
        task_result = AsyncResult(task_id)

        if task_result.ready():
            if task_result.successful():
                result = task_result.result
                response_data = {'task_id': task_id, 'status': 'SUCCESS', 'result': result}
            elif task_result.failed():
                response_data = {'task_id': task_id, 'status': 'FAILURE', 'result': str(task_result.result)}
            else:
                response_data = {'task_id': task_id, 'status': task_result.status}
        else:
            response_data = {'task_id': task_id, 'status': 'PENDING'}
    except CeleryError as e:
        response_data = {'error': str(e), 'status': 'ERROR'}

    return JsonResponse(response_data)

def clear_task_id_productos(request):
    if 'task_id_producto' in request.session:
        del request.session['task_id_producto']
    return JsonResponse({'status': 'success'})


def verificar_estado(request):
    task_id = request.session.get('task_id')  # Asumiendo que el task_id se pasa como parámetro GET

    if not task_id:
        return JsonResponse({'error': 'No se proporcionó task_id'}, status=400)


    task_result = AsyncResult(task_id)

    if task_result.ready():
        if task_result.successful():
            result = task_result.result
            response_data = {'task_id': task_id, 'status': 'SUCCESS', 'result': result}
        elif task_result.failed():
            response_data = {'task_id': task_id, 'status': 'FAILURE', 'result': str(task_result.result)}
        else:
            response_data = {'task_id': task_id, 'status': task_result.status}
    else:
        response_data = {'task_id': task_id, 'status': 'PENDING'}

    return JsonResponse(response_data)

def clear_task_id(request):
    if 'task_id' in request.session:
        del request.session['task_id']
    return JsonResponse({'status': 'success'})



@perfil_seleccionado_required
def matriz_oc_productos(request):
    pk_perfil = request.session.get('selected_profile_id')
    colaborador_sel = Profile.objects.all()
    usuario = colaborador_sel.get(id = pk_perfil)
    compras = Compra.objects.filter(complete=True)
    almacenes_distritos = set(usuario.almacen.values_list('distrito__id', flat=True))
    if usuario.tipo.proveedores or usuario.tipo.nombre == "VIS_ADQ":
        articulos = ArticuloComprado.objects.filter(
            oc__complete = True,
            oc__req__orden__distrito__id__in = almacenes_distritos
            ).order_by('-oc__created_at')
    else:
        articulos = ArticuloComprado.objects.filter(oc__complete = True, oc__req__orden__distrito = usuario.distritos).order_by('-oc__created_at')
    
    #articulos = ArticuloComprado.objects.filter(oc__complete = True, oc__req__orden__distrito = usuario.distritos).order_by('-oc__created_at')
    myfilter = ArticuloCompradoFilter(request.GET, queryset=articulos)
    articulos = myfilter.qs
    articulos_data = list(articulos.values())
    
    #Set up pagination
    p = Paginator(articulos, 50)
    page = request.GET.get('page')
    articulos_list = p.get_page(page)

    context= {
        'articulos_list':articulos_list,
        'articulos':articulos,
        'compras':compras,
        'myfilter':myfilter,
        }
    
    #task_id_producto = request.session.get('task_id_producto')

    

    if request.method == 'POST' and 'btnExcel' in request.POST:
        #if articulos.count() > 10000:
        #return convert_excel_solicitud_matriz_productos_quick(articulos)
        #else:
        return convert_excel_solicitud_matriz_productos_prov2(articulos)
        #    if not task_id_producto:
        #        task = convert_excel_solicitud_matriz_productos_task2.delay(articulos_data)
        #        task_id_producto = task.id
        #        request.session['task_id_producto'] = task_id_producto
        #        context['task_id_producto'] = task_id_producto
        #        cantidad = articulos.count()
        #        context['cantidad'] = cantidad
        #        messages.success(request, f'Tu reporte se está generando {task_id_producto}')
        #elif usuario.tipo.nombre == "PROVEEDORES":
            #print(articulos.count())
        
        #else:
        #    return convert_excel_solicitud_matriz_productos(articulos)
        
    return render(request, 'compras/matriz_oc_productos.html',context)

@perfil_seleccionado_required
def productos_oc(request, pk):
    compra = Compra.objects.get(id=pk)
    productos = ArticuloComprado.objects.filter(oc=compra)


    context = {
        'compra':compra,
        'productos':productos,
    }

    return render(request,'compras/oc_producto.html',context)

@perfil_seleccionado_required
def upload_facturas(request, pk):
    pago = Pago.objects.get(id = pk)
    facturas = Facturas.objects.filter(pago = pago, hecho=True)
    factura, created = Facturas.objects.get_or_create(pago=pago, hecho=False)
    form = Facturas_Form()

    if request.method == 'POST':
        form = Facturas_Form(request.POST or None, request.FILES or None, instance = factura)
        factura = form.save(commit=False)
        factura.fecha_subido = date.today()
        factura.hora_subido = datetime.now().time()
        factura.hecho = True
        if form.is_valid():
            form.save()
            factura.save()
            messages.success(request,'Las facturas se subieron de manera exitosa')
            return redirect('matriz-compras')
        else:
            form = Facturas_Form()
            messages.error(request,'No se pudo subir tu documento')

    context={
        'facturas':facturas,
        'form':form,
        }

    return render(request, 'compras/upload.html', context)

@perfil_seleccionado_required
def upload_xml(request, pk):
    compra = Compra.objects.get(id = pk)
    form = CompraFactForm()

    if request.method == 'POST':
        form = CompraFactForm(request.POST or None, request.FILES or None, instance = compra)
        if form.is_valid():
            form.save()
            return redirect('matriz-compras')
        else:
            form = CompraFactForm()
            messages.error(request,'No se pudo subir tu documento')

    context={
        'compra':compra,
        'form': form,
        }

    return render(request, 'compras/upload_xml.html', context)

@perfil_seleccionado_required
@tipo_usuario_requerido('oc_superintendencia')
def autorizacion_oc1(request):
    pk_perfil = request.session.get('selected_profile_id') 
    usuario = Profile.objects.get(id = pk_perfil)
    print(usuario)
    
    if usuario.sustituto:
        usuario= Profile.objects.filter(staff=usuario.staff, tipo=usuario.tipo, distritos=usuario.distritos).first()


    if usuario.tipo.subdirector == True:
        #if usuario.distritos.nombre == "MATRIZ": #Esto lo comenté para que sirvan los cambios para Brasil
        compras = Compra.objects.filter(complete=True, autorizado1= None, req__orden__superintendente = usuario).order_by('-folio')
    #elif usuario.tipo.subdirector == True and usuario.tipo.rh = True:
    #    compras = 
        #else:
        #    compras = Compra.objects.filter(complete=True, autorizado1= None, req__orden__distrito = usuario.distritos).order_by('-folio')
    elif usuario.tipo.oc_superintendencia == True:
        compras = Compra.objects.filter(complete=True, autorizado1= None, req__orden__distrito = usuario.distritos).order_by('-folio')
    else:
        compras = Compra.objects.none()
    #compras = Compra.objects.filter(complete=True, autorizado1= None).order_by('-folio')
    myfilter = CompraFilter(request.GET, queryset=compras)
    compras = myfilter.qs
    form = Compra_ComentarioForm()

    p = Paginator(compras, 50)
    page = request.GET.get('page')
    compras_list = p.get_page(page)

    context= {
        'form':form,
        'compras':compras,
        'myfilter':myfilter,
        'compras_list':compras_list,
        }

    return render(request, 'compras/autorizacion_oc1.html',context)

@perfil_seleccionado_required
def cancelar_oc1(request, pk):
    pk_perfil = request.session.get('selected_profile_id') 
    usuario = Profile.objects.get(id = pk_perfil)
    compra = Compra.objects.get(id = pk)
    productos = ArticuloComprado.objects.filter(oc = pk)

    if compra.costo_fletes == None:
        costo_fletes = 0
    
    #Si hay tipo de cambio es porque la compra fue en dólares entonces multiplico por tipo de cambio la cantidad
    #Escenario con dólares
    if compra.tipo_de_cambio:
        costo_oc = compra.costo_oc * compra.tipo_de_cambio
        if compra.costo_fletes:
            costo_fletes = compra.costo_fletes * compra.tipo_de_cambio
    #Escenario con pesos
    else:
        costo_oc = compra.costo_oc
        if compra.costo_fletes:
            costo_fletes = compra.costo_fletes
        else:
            costo_fletes = 0
    costo_total = costo_fletes + costo_oc
    resta = compra.req.orden.subproyecto.presupuesto - costo_total - compra.req.orden.subproyecto.gastado
    try:
        porcentaje = "{0:.2f}%".format((costo_oc/compra.req.orden.subproyecto.presupuesto)*100)
    except ZeroDivisionError:
        porcentaje = "0"

    if request.method == 'POST':
        compra.oc_autorizada_por = usuario
        compra.autorizado1 = False
        compra.autorizado_date1 = date.today()
        compra.autorizado_hora1 = datetime.now().time()
        compra.save()

        static_path = settings.STATIC_ROOT
        img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
        img_path2 = os.path.join(static_path,'images','logo_vordcab.jpg')
        image_base64 = get_image_base64(img_path)
        logo_v_base64 = get_image_base64(img_path2)
        # Crear el mensaje HTML
        html_message = f"""
        <html>
            <head>
                <meta charset="UTF-8">
            </head>
            <body style="font-family: Arial, sans-serif; color: #333; background-color: #f4f4f4; margin: 0; padding: 0;">
                <table width="100%" cellspacing="0" cellpadding="0" style="background-color: #f4f4f4; padding: 20px;">
                    <tr>
                        <td align="center">
                            <table width="600px" cellspacing="0" cellpadding="0" style="background-color: #ffffff; padding: 20px; border-radius: 10px;">
                                <tr>
                                    <td align="center">
                                        <img src="data:image/jpeg;base64,{logo_v_base64}" alt="Logo" style="width: 100px; height: auto;" />
                                    </td>
                                </tr>
                                <tr>
                                    <td style="padding: 20px;">
                                        <p style="font-size: 18px; text-align: justify;">
                                            <p>Estimado {compra.req.orden.staff.staff.staff.first_name} {compra.req.orden.staff.staff.staff.last_name},</p>
                                        </p>
                                        <p style="font-size: 16px; text-align: justify;">
                                            Estás recibiendo este correo porque tu OC con folio: <strong>{compra.folio}</strong> solicitud folio: <strong>{compra.req.orden.folio}</strong> ha sido cancelada.</p>
                                        </p>
                                    <p style="font-size: 16px; text-align: justify;">
                                        Att: {usuario.staff.staff.first_name} {usuario.staff.staff.last_name}
                                    </p>
                                        <p style="text-align: center; margin: 20px 0;">
                                            <img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width: 50px; height: auto; border-radius: 50%;" />
                                        </p>
                                        <p style="font-size: 14px; color: #999; text-align: justify;">
                                            Este mensaje ha sido automáticamente generado por SAVIA 2.0
                                        </p>
                                    </td>
                                </tr>
                            </table>
                        </td>
                    </tr>
                </table>
            </body>
        </html>
        """
        try:
            email = EmailMessage(
                f'OC Cancelada',
                body=html_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[compra.req.orden.staff.staff.staff.email],
                headers={'Content-Type': 'text/html'}
                )
            email.content_subtype = "html " # Importante para que se interprete como HTML
            email.send()
            messages.success(request,f'Has cancelado la compra con FOLIO: {compra.folio}')
        except (BadHeaderError, SMTPException, socket.gaierror) as e:
            error_message = f'La compra {compra.folio} ha sido cancelada, pero el correo no ha sido enviado debido a un error: {e}'
            messages.success(request, error_message)
        return redirect('autorizacion-oc1')

    context = {
        'compra':compra,
        'productos': productos,
        'costo_oc':costo_oc,
        'productos':productos,
        'tipo_cambio':compra.tipo_de_cambio,
        'resta':resta,
        'porcentaje':porcentaje,
        'costo_total':costo_total,
     }
    return render(request,'compras/cancelar_oc1.html', context)

@perfil_seleccionado_required
def cancelar_oc2(request, pk):
    pk_perfil = request.session.get('selected_profile_id') 
    usuario = Profile.objects.get(id = pk_perfil)
    compra = Compra.objects.get(id = pk)
    productos = ArticuloComprado.objects.filter(oc = pk)
    form = Compra_ComentarioForm(instance = compra)

    costo_fletes = 0
    if compra.costo_fletes == None:
        costo_fletes = 0
    #Si hay tipo de cambio es porque la compra fue en dólares entonces multiplico por tipo de cambio la cantidad
    #Escenario con dólares
    if compra.tipo_de_cambio:
        costo_oc = compra.costo_oc * compra.tipo_de_cambio
        if compra.costo_fletes:
            costo_fletes = compra.costo_fletes * compra.tipo_de_cambio
    #Escenario con pesos
    else:
        costo_oc = compra.costo_oc
        if compra.costo_fletes:
            costo_fletes = compra.costo_fletes
    costo_total = costo_fletes + costo_oc
    resta = compra.req.orden.subproyecto.presupuesto - costo_total - compra.req.orden.subproyecto.gastado
    
    try:
        porcentaje = "{0:.2f}%".format((costo_oc/compra.req.orden.subproyecto.presupuesto)*100)
    except ZeroDivisionError:
        porcentaje = "0"


    if request.method == 'POST':
        form = Compra_ComentarioForm(request.POST, instance=compra)
        if form.is_valid():
            compra = form.save(commit = False)
            compra.oc_autorizada_por2 = usuario
            compra.autorizado2 = False
            compra.autorizado_date2 = date.today()
            compra.autorizado_hora2 = datetime.now().time()
            compra.save()

            static_path = settings.STATIC_ROOT
            img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
            img_path2 = os.path.join(static_path,'images','logo_vordcab.jpg')
            image_base64 = get_image_base64(img_path)
            logo_v_base64 = get_image_base64(img_path2)
            # Crear el mensaje HTML
            html_message = f"""
            <html>
                <head>
                    <meta charset="UTF-8">
                </head>
                <body style="font-family: Arial, sans-serif; color: #333; background-color: #f4f4f4; margin: 0; padding: 0;">
                    <table width="100%" cellspacing="0" cellpadding="0" style="background-color: #f4f4f4; padding: 20px;">
                        <tr>
                            <td align="center">
                                <table width="600px" cellspacing="0" cellpadding="0" style="background-color: #ffffff; padding: 20px; border-radius: 10px;">
                                    <tr>
                                        <td align="center">
                                            <img src="data:image/jpeg;base64,{logo_v_base64}" alt="Logo" style="width: 100px; height: auto;" />
                                        </td>
                                    </tr>
                                    <tr>
                                        <td style="padding: 20px;">
                                            <p style="font-size: 18px; text-align: justify;">
                                                <p>Estimado {compra.req.orden.staff.staff.staff.first_name} {compra.req.orden.staff.staff.staff.last_name},</p>
                                            </p>
                                            <p style="font-size: 16px; text-align: justify;">
                                                Estás recibiendo este correo porque tu OC con folio: <strong>{compra.folio}</strong> solicitud con folio: <strong>{compra.req.orden.folio} ha sido cancelada.</p>
                                            </p>
                                            <p style="font-size: 16px; text-align: justify;">Comentario:</p>
                                            <p style="font-size: 16px; text-align: justify;">
                                                {compra.comentarios}
                                            </p>
                                        <p style="font-size: 16px; text-align: justify;">
                                            Att: {usuario.staff.staff.first_name} {usuario.staff.staff.last_name}
                                        </p>
                                            <p style="text-align: center; margin: 20px 0;">
                                                <img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width: 50px; height: auto; border-radius: 50%;" />
                                            </p>
                                            <p style="font-size: 14px; color: #999; text-align: justify;">
                                                Este mensaje ha sido automáticamente generado por SAVIA 2.0
                                            </p>
                                        </td>
                                    </tr>
                                </table>
                            </td>
                        </tr>
                    </table>
                </body>
            </html>
            """
            try:
                email = EmailMessage(
                    f'OC Cancelada',
                    body=html_message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[compra.req.orden.staff.staff.staff.email],
                    headers={'Content-Type': 'text/html'}
                    )
                email.content_subtype = "html " # Importante para que se interprete como HTML
                email.send()
                messages.success(request,f'Has cancelado la compra con FOLIO: {compra.folio}')
            except (BadHeaderError, SMTPException, socket.gaierror) as e:
                error_message = f'La compra {compra.folio} ha sido cancelada, pero el correo no ha sido enviado debido a un error: {e}'
                messages.success(request, error_message)
            return HttpResponse(status=204)
    
    context = {
        'form':form,
        'compra':compra,
        'productos': productos,
        'costo_oc':costo_oc,
        'productos':productos,
        'tipo_cambio':compra.tipo_de_cambio,
        'resta':resta,
        'porcentaje':porcentaje,
        'costo_total':costo_total,
     }
    return render(request,'compras/cancelar_oc2.html', context)


@perfil_seleccionado_required
def back_oc(request, pk):
    pk_perfil = request.session.get('selected_profile_id') 
    perfil = Profile.objects.get(id = pk_perfil)
    compra = Compra.objects.get(id = pk)
    productos = ArticuloComprado.objects.filter(oc = pk)
    #Traigo la requisición para poderla activar de nuevo, aunque esto ya no es necesario porque no se reactiva propiamente la requi
    #Crearé la cancelación de la OC?
    #Esto está afectando de alguna forma a los productos pendientes cuando se eliminan partidas de la OC?
    requi = Requis.objects.get(id=compra.req.id)

    if compra.costo_fletes == None or compra.costo_fletes == 0:
        costo_fletes = 0
    #Si hay tipo de cambio es porque la compra fue en dólares entonces multiplico por tipo de cambio la cantidad
    #Escenario con dólares
    if compra.tipo_de_cambio:
        costo_oc = compra.costo_oc * compra.tipo_de_cambio
        if compra.costo_fletes:
            costo_fletes = compra.costo_fletes * compra.tipo_de_cambio
    #Escenario con pesos
    else:
        costo_oc = compra.costo_oc
        if compra.costo_fletes:
            costo_fletes = compra.costo_fletes
    costo_total = costo_fletes + costo_oc
    resta = compra.req.orden.subproyecto.presupuesto - costo_total - compra.req.orden.subproyecto.gastado
   
    presupuesto = compra.req.orden.subproyecto.presupuesto or 0  # Default to 0 if None
    if presupuesto > 0:
        porcentaje = "{0:.2f}%".format((costo_oc / presupuesto) * 100)
    else:
        porcentaje = "0%"  # Default value when presupuesto is 0 or invalid

    form = Compra_ComentarioForm()

    if request.method == 'POST':
        form = Compra_ComentarioForm(request.POST, instance=compra)
        if form.is_valid():
            compra = form.save(commit = False)
            if not compra.autorizado1:
                compra.oc_autorizada_por = perfil
                compra.autorizado1 = None
                compra.complete = False
                compra.autorizado_at = datetime.now()
                #compra.autorizado_hora1 = datetime.now().time()
                compra.regresar_oc = True
            else:
                compra.oc_autorizada_por2 = perfil
                compra.autorizado2 = None
                compra.autorizado1 = None
                compra.complete = False
                compra.autorizado_at_2 = datetime.now()
                #compra.autorizado_hora2 = datetime.now().time()
                compra.regresar_oc = True
            #Esta línea es la que activa a la requi
            #requi.colocada = False
            compra.save()
            #requi.save()
            messages.success(request,f'Has regresado la compra con FOLIO: {compra.folio} y ahora podrás encontrar esos productos en el apartado devolución')
            return redirect('compras-devueltas')

    context = {
        'form':form,
        'compra':compra,
        'productos': productos,
        'costo_oc':costo_oc,
        'productos':productos,
        'tipo_cambio':compra.tipo_de_cambio,
        'resta':resta,
        'porcentaje':porcentaje,
        'costo_total':costo_total,
     }

    return render(request,'compras/back_oc.html', context)



@perfil_seleccionado_required
def autorizar_oc1(request, pk):
    pk_perfil = request.session.get('selected_profile_id') 
    usuario = Profile.objects.get(id = pk_perfil)
   
    compra = Compra.objects.get(id = pk)
    productos = ArticuloComprado.objects.filter(oc=pk)
    productos_criticos = productos.filter(producto__producto__articulos__producto__producto__critico=True)
    form = Compra_ComentarioForm()

    if compra.costo_fletes == None:
        costo_fletes = 0
    #Si hay tipo de cambio es porque la compra fue en dólares entonces multiplico por tipo de cambio la cantidad
    #Escenario con dólares
    if compra.tipo_de_cambio:
        costo_oc = compra.costo_oc * compra.tipo_de_cambio
        if compra.costo_fletes:
            costo_fletes = compra.costo_fletes * compra.tipo_de_cambio
        else:
            costo_fletes = 0
    #Escenario con pesos
    else:
        costo_oc = compra.costo_oc
        if compra.costo_fletes:
            costo_fletes = compra.costo_fletes
        else:
            costo_fletes = 0
    costo_total = costo_fletes + costo_oc
    resta = compra.req.orden.subproyecto.presupuesto - costo_oc - costo_fletes - compra.req.orden.subproyecto.gastado
    try:
        porcentaje = "{0:.2f}%".format((costo_oc/compra.req.orden.subproyecto.presupuesto)*100)
    except ZeroDivisionError:
        porcentaje = "0"


    if request.method == 'POST':
        form = Compra_ComentarioForm(request.POST, instance=compra)
        if form.is_valid():
            compra = form.save(commit = False)
            compra.autorizado1 = True
            compra.oc_autorizada_por = usuario
            compra.autorizado_at = datetime.now()
            #compra.autorizado_hora1 = datetime.now().time()
            if usuario.tipo.subdirector == True:
                compra = form.save(commit = False)
                compra.autorizado2 = True
                compra.oc_autorizada_por2 = usuario
                compra.autorizado_at_2 = datetime.now()
            compra.save()
            archivo_oc = attach_oc_pdf(request, compra.id)
            pdf_antisoborno = attach_antisoborno_pdf(request)
            pdf_privacidad = attach_aviso_privacidad_pdf(request)
            pdf_etica = attach_codigo_etica_pdf(request)
            pdf_politica_proveedor = attach_politica_proveedor(request)
            static_path = settings.STATIC_ROOT
            img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
            img_path2 = os.path.join(static_path,'images','logo_vordcab.jpg')
        
            image_base64 = get_image_base64(img_path)
            logo_v_base64 = get_image_base64(img_path2)

            articulos_html = """
            <table border="1" style="border-collapse: collapse; width: 100%;">
                <thead>
                    <tr>
                        <th>Producto Crítico</th>
                        <th>Requerimiento</th>
                        <th>Comentarios</th>
                    </tr>
                </thead>
                <tbody>
            """
            productos_criticos = productos_criticos
            for articulo in productos_criticos:
                producto = articulo.producto.producto.articulos.producto.producto
                # Si no existe producto_calidad, capturamos la excepción y seguimos
                try:
                    requerimientos = producto.producto_calidad.requerimientos_calidad.all()
                except ObjectDoesNotExist:
                    requerimientos = []  # o ProductoCalidad.objects.none()

                if requerimientos:
                    for requerimiento in requerimientos:
                        #print(requerimiento.requerimiento.nombre)
                        articulos_html += f"""
                            <tr>
                                <td>{producto.nombre}</td>
                                <td>{requerimiento.requerimiento.nombre}</td>
                                <td>{requerimiento.comentarios}</td>

                            </tr>
                        """
                else:
                    articulos_html += f"""
                        <tr>
                            <td>{producto.codigo}</td>
                            <td>Sin requerimiento</td>
                        </tr>
                    """
            articulos_html += """
                </tbody>
            </table>
            """
            # Crear el mensaje HTML
            if usuario.tipo.subdirector == True:
                html_message = f"""
                <html>
                    <head>
                        <meta charset="UTF-8">
                    </head>
                    <body style="font-family: Arial, sans-serif; color: #333; background-color: #f4f4f4; margin: 0; padding: 0;">
                        <table width="100%" cellspacing="0" cellpadding="0" style="background-color: #f4f4f4; padding: 20px;">
                            <tr>
                                <td align="center">
                                    <table width="600px" cellspacing="0" cellpadding="0" style="background-color: #ffffff; padding: 20px; border-radius: 10px;">
                                        <tr>
                                            <td align="center">
                                                <img src="data:image/jpeg;base64,{logo_v_base64}" alt="Logo" style="width: 100px; height: auto;" />
                                            </td>
                                        </tr>
                                        <tr>
                                            <td style="padding: 20px;">
                                                <p style="font-size: 18px; text-align: justify;">
                                                    <p>Estimado {compra.req.orden.staff.staff.staff.first_name} {compra.req.orden.staff.staff.staff.last_name},</p>
                                                </p>
                                                <p style="font-size: 16px; text-align: justify;">
                                                    Estás recibiendo este correo porque tu OC {compra.folio} | RQ: {compra.req.folio} |Sol: {compra.req.orden.folio} ha sido autorizada por {compra.oc_autorizada_por.staff.staff.first_name} {compra.oc_autorizada_por.staff.staff.last_name}.</p>
                                                </p>
                                            <p style="font-size: 16px; text-align: justify;">
                                                El siguiente paso del sistema: Pago por parte de tesorería.
                                            </p>
                                                <p style="text-align: center; margin: 20px 0;">
                                                    <img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width: 50px; height: auto; border-radius: 50%;" />
                                                </p>
                                                <p style="font-size: 14px; color: #999; text-align: justify;">
                                                    Este mensaje ha sido automáticamente generado por SAVIA 2.0
                                                </p>
                                            </td>
                                        </tr>
                                    </table>
                                </td>
                            </tr>
                        </table>
                    </body>
                </html>
                """
                if compra.cond_de_pago.nombre == "CREDITO":
                    html_message2 = f"""
                    <html>
                        <head>
                            <meta charset="UTF-8">
                        </head>
                        <body style="font-family: Arial, sans-serif; color: #333; background-color: #f4f4f4; margin: 0; padding: 0;">
                            <table width="100%" cellspacing="0" cellpadding="0" style="background-color: #f4f4f4; padding: 20px;">
                                <tr>
                                    <td align="center">
                                        <table width="600px" cellspacing="0" cellpadding="0" style="background-color: #ffffff; padding: 20px; border-radius: 10px;">
                                            <tr>
                                                <td align="center">
                                                    <img src="data:image/jpeg;base64,{logo_v_base64}" alt="Logo" style="width: 100px; height: auto;" />
                                                </td>
                                            </tr>
                                            <tr>
                                                <td style="padding: 20px;">
                                                    <p style="font-size: 18px; text-align: justify;">
                                                        <p>Estimado(a) {compra.proveedor.contacto}| Proveedor {compra.proveedor.nombre}:,</p>
                                                    </p>
                                                    <p style="font-size: 16px; text-align: justify;">
                                                        Estás recibiendo este correo porque has sido seleccionado para surtirnos la OC adjunta con folio: {compra.folio}.</p>
                                                    </p>
                                                <p style="font-size: 16px; text-align: justify;">
                                                    <p>&nbsp;</p>
                                                    Atte. {compra.creada_por.staff.staff.first_name} {compra.creada_por.staff.staff.last_name}.
                                                    <p>GRUPO VORDCAB S.A. de C.V.</p>
                                                    {f"{articulos_html}" if productos_criticos.exists() else ""}
                                                </p>
                                                    <p style="text-align: center; margin: 20px 0;">
                                                        <img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width: 50px; height: auto; border-radius: 50%;" />
                                                    </p>
                                                    <p style="font-size: 14px; color: #999; text-align: justify;">
                                                        Este mensaje ha sido automáticamente generado por SAVIA 2.0
                                                    </p>
                                                </td>
                                            </tr>
                                        </table>
                                    </td>
                                </tr>
                            </table>
                        </body>
                    </html>
                    """    
                    try:
                        email = EmailMessage(
                        f'Compra Autorizada {compra.folio}|SAVIA',
                        body=html_message2,
                        from_email =settings.DEFAULT_FROM_EMAIL,
                        to= [compra.creada_por.staff.staff.email, compra.proveedor.email,],
                        headers={'Content-Type': 'text/html'}
                        )
                        email.content_subtype = "html " # Importante para que se interprete como HTML
                        email.attach(f'OC_folio_{compra.folio}.pdf',archivo_oc,'application/pdf')
                        email.attach(f'Politica_antisoborno.pdf', pdf_antisoborno, 'application/pdf')
                        email.attach(f'Aviso_de_privacidad.pdf', pdf_privacidad, 'application/pdf')
                        email.attach(f'Codigo_de_etica.pdf', pdf_etica, 'application/pdf')
                        email.attach(f'Politica_proveedor.pdf', pdf_politica_proveedor, 'application/pdf')
                        # Adjuntar los archivos con nombres personalizados
                        for articulo in productos:
                            producto = articulo.producto.producto.articulos.producto.producto
                            if producto.critico:
                                requerimientos = producto.producto_calidad.requerimientos_calidad.all()
                                contador = 1  # Contador para evitar nombres duplicados
                                for requerimiento in requerimientos:
                                    archivo_path = requerimiento.url.path
                                    nombre_archivo = f"{producto.codigo}_requerimiento_{contador}{os.path.splitext(archivo_path)[1]}"
                                    
                                    # Abrir el archivo en modo binario y adjuntarlo directamente
                                    with open(archivo_path, 'rb') as archivo:
                                        email.attach(nombre_archivo, archivo.read())

                                    contador += 1  # Incrementar el contador para el siguiente archivo
                        email.send()
                    except (BadHeaderError, SMTPException, socket.gaierror) as e:
                        error_message = f'correo de notificación no ha sido enviado debido a un error: {e}'  
                else:
                    html_message = f"""
                    <html>
                        <head>
                            <meta charset="UTF-8">
                        </head>
                        <body style="font-family: Arial, sans-serif; color: #333; background-color: #f4f4f4; margin: 0; padding: 0;">
                            <table width="100%" cellspacing="0" cellpadding="0" style="background-color: #f4f4f4; padding: 20px;">
                                <tr>
                                    <td align="center">
                                        <table width="600px" cellspacing="0" cellpadding="0" style="background-color: #ffffff; padding: 20px; border-radius: 10px;">
                                            <tr>
                                                <td align="center">
                                                    <img src="data:image/jpeg;base64,{logo_v_base64}" alt="Logo" style="width: 100px; height: auto;" />
                                                </td>
                                            </tr>
                                            <tr>
                                                <td style="padding: 20px;">
                                                    <p style="font-size: 18px; text-align: justify;">
                                                        <p>Estimado {compra.req.orden.staff.staff.staff.first_name} {compra.req.orden.staff.staff.staff.last_name},</p>
                                                    </p>
                                                    <p style="font-size: 16px; text-align: justify;">
                                                        Estás recibiendo este correo porque tu OC {compra.folio} | RQ: {compra.req.folio} |Sol: {compra.req.orden.folio} ha sido autorizada por {compra.oc_autorizada_por.staff.staff.first_name} {compra.oc_autorizada_por.staff.staff.last_name}.</p>
                                                    </p>
                                                <p style="font-size: 16px; text-align: justify;">
                                                    El siguiente paso del sistema: Pago por parte de tesorería.
                                                </p>
                                                    <p style="text-align: center; margin: 20px 0;">
                                                        <img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width: 50px; height: auto; border-radius: 50%;" />
                                                    </p>
                                                    <p style="font-size: 14px; color: #999; text-align: justify;">
                                                        Este mensaje ha sido automáticamente generado por SAVIA 2.0
                                                    </p>
                                                </td>
                                            </tr>
                                        </table>
                                    </td>
                                </tr>
                            </table>
                        </body>
                    </html>
                    """    
                    try:
                        email = EmailMessage(
                        f'OC Autorizada Gerencia {compra.folio}|RQ: {compra.req.folio} |Sol: {compra.req.orden.folio}',
                        body=html_message,
                        from_email = settings.DEFAULT_FROM_EMAIL,
                        to= [compra.creada_por.staff.staff.email],
                        headers={'Content-Type': 'text/html'}
                        )
                        email.content_subtype = "html " # Importante para que se interprete como HTML
                        email.send()
                        messages.success(request, f'{usuario.staff.staff.first_name} has autorizado la compra {compra.folio}')
                    except (BadHeaderError, SMTPException, socket.gaierror) as e:
                        error_message = f'{usuario.staff.staff.first_name} has autorizado la compra {compra.folio} pero el correo de notificación no ha sido enviado debido a un error: {e}'
                        messages.success(request, error_message)    
                    return redirect('autorizacion-oc1')
            else:
                html_message = f"""
                <html>
                    <head>
                        <meta charset="UTF-8">
                    </head>
                    <body style="font-family: Arial, sans-serif; color: #333; background-color: #f4f4f4; margin: 0; padding: 0;">
                        <table width="100%" cellspacing="0" cellpadding="0" style="background-color: #f4f4f4; padding: 20px;">
                            <tr>
                                <td align="center">
                                    <table width="600px" cellspacing="0" cellpadding="0" style="background-color: #ffffff; padding: 20px; border-radius: 10px;">
                                        <tr>
                                            <td align="center">
                                                <img src="data:image/jpeg;base64,{logo_v_base64}" alt="Logo" style="width: 100px; height: auto;" />
                                            </td>
                                        </tr>
                                        <tr>
                                            <td style="padding: 20px;">
                                                <p style="font-size: 18px; text-align: justify;">
                                                    <p>Estimado {compra.req.orden.staff.staff.staff.first_name} {compra.req.orden.staff.staff.staff.last_name},</p>
                                                </p>
                                                <p style="font-size: 16px; text-align: justify;">
                                                    Estás recibiendo este correo porque tu OC {compra.folio} | RQ: {compra.req.folio} |Sol: {compra.req.orden.folio} ha sido autorizada por {compra.oc_autorizada_por.staff.staff.first_name} {compra.oc_autorizada_por.staff.staff.last_name}.</p>
                                                </p>
                                            <p style="font-size: 16px; text-align: justify;">
                                                El siguiente paso del sistema: Autorización de OC por Gerencia de Distrito.
                                            </p>
                                                <p style="text-align: center; margin: 20px 0;">
                                                    <img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width: 50px; height: auto; border-radius: 50%;" />
                                                </p>
                                                <p style="font-size: 14px; color: #999; text-align: justify;">
                                                    Este mensaje ha sido automáticamente generado por SAVIA 2.0
                                                </p>
                                            </td>
                                        </tr>
                                    </table>
                                </td>
                            </tr>
                        </table>
                    </body>
                </html>
                """    
            try:
                email = EmailMessage(
                    f'OC Autorizada {compra.folio}|RQ: {compra.req.folio} |Sol: {compra.req.orden.folio}',
                    body=html_message,
                    from_email = settings.DEFAULT_FROM_EMAIL,
                    to= [compra.req.orden.staff.staff.staff.email],
                    headers={'Content-Type': 'text/html'}
                )
                email.content_subtype = "html " # Importante para que se interprete como HTML
                email.send()
                messages.success(request, f'{usuario.staff.staff.first_name} has autorizado la compra {compra.folio}')
            except (BadHeaderError, SMTPException, socket.gaierror) as e:
                error_message = f'{usuario.staff.staff.first_name} has autorizado la compra {compra.folio} pero el correo de notificación no ha sido enviado debido a un error: {e}'
                messages.success(request, error_message)    
                
            return redirect('autorizacion-oc1')

    context={
        'form':form,
        'compra':compra,
        'costo_oc':costo_oc,
        'productos':productos,
        'tipo_cambio':compra.tipo_de_cambio,
        'resta':resta,
        'porcentaje':porcentaje,
        'costo_total':costo_total,
        }
    

    return render(request, 'compras/autorizar_oc1.html',context)

@perfil_seleccionado_required
@tipo_usuario_requerido('oc_gerencia')
def autorizacion_oc2(request):
    pk_perfil = request.session.get('selected_profile_id') 
    usuario = Profile.objects.get(id = pk_perfil)
    if usuario.tipo.oc_gerencia == True:
        compras = Compra.objects.filter(complete = True, autorizado1 = True, autorizado2= None, req__orden__distrito = usuario.distritos).order_by('-folio')
    else:
        compras = Compra.objects.none()
    
    myfilter = CompraFilter(request.GET, queryset=compras)
    compras = myfilter.qs
    
    context= {
        'compras':compras,
        'myfilter':myfilter,
        }

    return render(request, 'compras/autorizacion_oc2.html',context)

@perfil_seleccionado_required
def autorizar_oc2(request, pk):
    pk_perfil = request.session.get('selected_profile_id') 
    usuario = Profile.objects.get(id = pk_perfil)
    compra = Compra.objects.get(id = pk)
    productos = ArticuloComprado.objects.filter(oc=pk)
    productos_criticos = productos.filter(producto__producto__articulos__producto__producto__critico=True)
    if compra.costo_fletes == None:
        costo_fletes = 0
    #Si hay tipo de cambio es porque la compra fue en dólares entonces multiplico por tipo de cambio la cantidad

    if compra.tipo_de_cambio:
        costo_oc = compra.costo_oc * compra.tipo_de_cambio
        if compra.costo_fletes:
            costo_fletes = compra.costo_fletes * compra.tipo_de_cambio
        else:
            costo_fletes = 0
    #Escenario con pesos
    else:
        costo_oc = compra.costo_oc
        if compra.costo_fletes:
            costo_fletes = compra.costo_fletes
        else:
            costo_fletes = 0

    costo_total = costo_fletes + costo_oc
    resta = compra.req.orden.subproyecto.presupuesto - costo_oc - costo_fletes - compra.req.orden.subproyecto.gastado
    
    presupuesto = compra.req.orden.subproyecto.presupuesto or 0 
    
    if presupuesto > 0:
        porcentaje = "{0:.2f}%".format((costo_oc / presupuesto) * 100)
    else:
        porcentaje = "0%"  # Default value when presupuesto is 0 or invalid

    form = Compra_ComentarioGerForm()

   
    if request.method == 'POST':
        form = Compra_ComentarioGerForm(request.POST, instance=compra)
        if form.is_valid():
            compra = form.save(commit = False)
            compra.autorizado2 = True
            compra.oc_autorizada_por2 = usuario
            compra.autorizado_at_2 = datetime.now()
            #compra.autorizado_hora2 = datetime.now().time()
            compra.save()
            static_path = settings.STATIC_ROOT
            img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
            img_path2 = os.path.join(static_path,'images','logo_vordcab.jpg')
            image_base64 = get_image_base64(img_path)
            logo_v_base64 = get_image_base64(img_path2)
            articulos_html = """
            <table border="1" style="border-collapse: collapse; width: 100%;">
                <thead>
                    <tr>
                        <th>Producto</th>
                        <th>Requerimiento</th>
                        <th>Comentarios</th>
                    </tr>
                </thead>
                <tbody>
            """
            productos_criticos = productos_criticos
            for articulo in productos_criticos:
                producto = articulo.producto.producto.articulos.producto.producto
                try:
                    requerimientos = producto.producto_calidad.requerimientos_calidad.all()
                except ObjectDoesNotExist:
                    requerimientos = []  # o ProductoCalidad.objects.none()

                if requerimientos:
                    for requerimiento in requerimientos:
                        articulos_html += f"""
                            <tr>
                                <td>{producto.nombre}</td>
                                <td>{requerimiento.requerimiento.nombre}</td>
                                <td>{requerimiento.comentarios}</td>
                            </tr>
                        """
                else:
                    articulos_html += f"""
                        <tr>
                            <td>{producto.nombre}</td>
                            <td>Sin requerimientos</td>
                            <td>Sin comentarios</td>
                        </tr>
                    """
            articulos_html += """
                </tbody>
            </table>
            """
            # Crear el mensaje HTML
            if compra.cond_de_pago.nombre == "CREDITO":
                archivo_oc = attach_oc_pdf(request, compra.id)
                pdf_antisoborno = attach_antisoborno_pdf(request)
                pdf_privacidad = attach_aviso_privacidad_pdf(request)
                pdf_etica = attach_codigo_etica_pdf(request)
                pdf_politica_proveedor = attach_politica_proveedor(request)
                avisar_calidad_oc(compra)
                html_message2 = f"""
                <html>
                    <head>
                        <meta charset="UTF-8">
                    </head>
                    <body style="font-family: Arial, sans-serif; color: #333; background-color: #f4f4f4; margin: 0; padding: 0;">
                        <table width="100%" cellspacing="0" cellpadding="0" style="background-color: #f4f4f4; padding: 20px;">
                            <tr>
                                <td align="center">
                                    <table width="600px" cellspacing="0" cellpadding="0" style="background-color: #ffffff; padding: 20px; border-radius: 10px;">
                                        <tr>
                                            <td align="center">
                                                <img src="data:image/jpeg;base64,{logo_v_base64}" alt="Logo" style="width: 100px; height: auto;" />
                                            </td>
                                        </tr>
                                        <tr>
                                            <td style="padding: 20px;">
                                                <p style="font-size: 18px; text-align: justify;">
                                                    <p>Estimado(a) {compra.proveedor.contacto}| Proveedor {compra.proveedor.nombre}:,</p>
                                                </p>
                                                <p style="font-size: 16px; text-align: justify;">
                                                    Estás recibiendo este correo porque has sido seleccionado para surtirnos la OC adjunta con folio: {compra.folio}.</p>
                                                    <p>&nbsp;</p>
                                                    <p> Atte. {compra.creada_por.staff.staff.first_name} {compra.creada_por.staff.staff.last_name}</p> 
                                                    <p>GRUPO VORDCAB S.A. de C.V.</p>
                                                    {f"{articulos_html}" if productos_criticos.exists() else ""}
                                                </p>
                                                <p style="text-align: center; margin: 20px 0;">
                                                    <img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width: 50px; height: auto; border-radius: 50%;" />
                                                </p>
                                                <p style="font-size: 14px; color: #999; text-align: justify;">
                                                    Este mensaje ha sido automáticamente generado por SAVIA 2.0
                                                </p>
                                            </td>
                                        </tr>
                                    </table>
                                </td>
                            </tr>
                        </table>
                    </body>
                </html>
                """
                try:
                    email = EmailMessage(
                    f'Compra Autorizada {compra.folio}|SAVIA',
                    body=html_message2,
                    from_email =settings.DEFAULT_FROM_EMAIL,
                    to= [compra.creada_por.staff.staff.email, compra.proveedor.email,],
                    headers={'Content-Type': 'text/html'}
                    )
                    email.content_subtype = "html " # Importante para que se interprete como HTML
                    email.attach(f'folio:{compra.folio}.pdf',archivo_oc,'application/pdf')
                    email.attach(f'Politica_antisoborno.pdf', pdf_antisoborno, 'application/pdf')
                    email.attach(f'Aviso_de_privacidad.pdf', pdf_privacidad, 'application/pdf')
                    email.attach(f'Codigo_de_etica.pdf', pdf_etica, 'application/pdf')
                    email.attach(f'Politica_proveedor.pdf', pdf_politica_proveedor, 'application/pdf')
                    # Adjuntar los archivos con nombres personalizados
                    #articulos = ArticuloComprado.objects.filter(oc=compra)
                    #for articulo in articulos:
                        #producto = articulo.producto.producto.articulos.producto.producto
                        #if producto.critico:
                            #requerimientos = producto.producto_calidad.requerimientos_calidad.all()
                            #contador = 1  # Contador para evitar nombres duplicados
                            #for requerimiento in requerimientos:
                            #    archivo_path = requerimiento.url.path
                            #    nombre_archivo = f"{producto.codigo}_requerimiento_{contador}{os.path.splitext(archivo_path)[1]}"
                                
                                # Abrir el archivo en modo binario y adjuntarlo directamente
                            #    with open(archivo_path, 'rb') as archivo:
                            #        email.attach(nombre_archivo, archivo.read())

                            #    contador += 1  # Incrementar el contador para el siguiente archivo
                    email.send()
                except (BadHeaderError, SMTPException, socket.gaierror) as e:
                    error_message = f'correo de notificación no ha sido enviado debido a un error: {e}'  
                html_message = f"""
                <html>
                    <head>
                        <meta charset="UTF-8">
                    </head>
                    <body style="font-family: Arial, sans-serif; color: #333; background-color: #f4f4f4; margin: 0; padding: 0;">
                        <table width="100%" cellspacing="0" cellpadding="0" style="background-color: #f4f4f4; padding: 20px;">
                            <tr>
                                <td align="center">
                                    <table width="600px" cellspacing="0" cellpadding="0" style="background-color: #ffffff; padding: 20px; border-radius: 10px;">
                                        <tr>
                                            <td align="center">
                                                <img src="data:image/jpeg;base64,{logo_v_base64}" alt="Logo" style="width: 100px; height: auto;" />
                                            </td>
                                        </tr>
                                        <tr>
                                            <td style="padding: 20px;">
                                                <p style="font-size: 18px; text-align: justify;">
                                                    <p>Estimado {compra.req.orden.staff.staff.staff.first_name} {compra.req.orden.staff.staff.staff.last_name},</p>
                                                </p>
                                                <p style="font-size: 16px; text-align: justify;">
                                                    Estás recibiendo este correo porque tu OC {compra.folio} | RQ: {compra.req.folio} |Sol: {compra.req.orden.folio} ha sido autorizada por {compra.oc_autorizada_por2.staff.staff.first_name} {compra.oc_autorizada_por2.staff.staff.last_name}.</p>
                                                    <p>El siguiente paso del sistema: Recepción por parte de Almacén |Compra a crédito</p>
                                                </p>
                                                <p style="text-align: center; margin: 20px 0;">
                                                    <img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width: 50px; height: auto; border-radius: 50%;" />
                                                </p>
                                                <p style="font-size: 14px; color: #999; text-align: justify;">
                                                    Este mensaje ha sido automáticamente generado por SAVIA 2.0
                                                </p>
                                            </td>
                                        </tr>
                                    </table>
                                </td>
                            </tr>
                        </table>
                    </body>
                </html>
                """
                try:
                    email = EmailMessage(
                        f'OC Autorizada Gerencia {compra.folio}|RQ: {compra.req.folio} |Sol: {compra.req.orden.folio}',
                        body=html_message,
                        #f'Estimado {requi.orden.staff.staff.staff.first_name} {requi.orden.staff.staff.staff.last_name},\n Estás recibiendo este correo porque tu solicitud: {requi.orden.folio}| Req: {requi.folio} ha sido autorizada,\n por {requi.requi_autorizada_por.staff.staff.first_name} {requi.requi_autorizada_por.staff.staff.last_name}.\n El siguiente paso del sistema: Generación de OC \n\n Este mensaje ha sido automáticamente generado por SAVIA VORDTEC',
                        from_email = settings.DEFAULT_FROM_EMAIL,
                        to= [compra.creada_por.staff.staff.email,],
                        headers={'Content-Type': 'text/html'}
                        )
                    email.content_subtype = "html " # Importante para que se interprete como HTML
                    email.send()
                    
                    #for producto in productos:
                    #    if producto.producto.producto.articulos.producto.producto.especialista == True:
                    #        archivo_oc = attach_oc_pdf(request, compra.id)
                    #        email = EmailMessage(
                    #            f'Compra Autorizada {compra.folio}',
                    #            f'Estimado Nombre de Calidad,\n Estás recibiendo este correo porque ha sido aprobada una OC que contiene el producto código:{producto.producto.producto.articulos.producto.producto.codigo} descripción:{producto.producto.producto.articulos.producto.producto.nombre} el cual requiere la liberación de calidad\n Este mensaje ha sido automáticamente generado por SAVIA 2.0',
                    #            settings.DEFAULT_FROM_EMAIL,
                    #            ['ulises_huesc@hotmail.com'],
                    #            )
                    #        email.attach(f'folio:{compra.folio}.pdf',archivo_oc,'application/pdf')
                    #        email.send()
                    messages.success(request, f'{usuario.staff.staff.first_name} has autorizado la compra {compra.folio}')
                except (BadHeaderError, SMTPException, socket.gaierror) as e:
                    error_message = f'{usuario.staff.staff.first_name} has autorizado la compra {compra.folio} pero el correo de notificación no ha sido enviado debido a un error: {e}'
                    messages.warning(request, error_message)    
                
            else:
                html_message = f"""
                <html>
                    <head>
                        <meta charset="UTF-8">
                    </head>
                    <body style="font-family: Arial, sans-serif; color: #333; background-color: #f4f4f4; margin: 0; padding: 0;">
                        <table width="100%" cellspacing="0" cellpadding="0" style="background-color: #f4f4f4; padding: 20px;">
                            <tr>
                                <td align="center">
                                    <table width="600px" cellspacing="0" cellpadding="0" style="background-color: #ffffff; padding: 20px; border-radius: 10px;">
                                        <tr>
                                            <td align="center">
                                                <img src="data:image/jpeg;base64,{logo_v_base64}" alt="Logo" style="width: 100px; height: auto;" />
                                            </td>
                                        </tr>
                                        <tr>
                                            <td style="padding: 20px;">
                                                <p style="font-size: 18px; text-align: justify;">
                                                    <p>Estimado {compra.req.orden.staff.staff.staff.first_name} {compra.req.orden.staff.staff.staff.last_name}.</p>
                                                </p>
                                                <p style="font-size: 16px; text-align: justify;">
                                                    Estás recibiendo este correo porque tu OC {compra.folio} | RQ: {compra.req.folio} |Sol: {compra.req.orden.folio} ha sido autorizada por {compra.oc_autorizada_por2.staff.staff.first_name} {compra.oc_autorizada_por2.staff.staff.last_name}.</p>
                                                    <p>El siguiente paso del sistema: Pago por parte de tesorería</p>
                                                </p>
                                                <p style="text-align: center; margin: 20px 0;">
                                                    <img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width: 50px; height: auto; border-radius: 50%;" />
                                                </p>
                                                <p style="font-size: 14px; color: #999; text-align: justify;">
                                                    Este mensaje ha sido automáticamente generado por SAVIA 2.0
                                                </p>
                                            </td>
                                        </tr>
                                    </table>
                                </td>
                            </tr>
                        </table>
                    </body>
                </html>
                """
                try:
                    email = EmailMessage(
                        f'OC Autorizada Gerencia {compra.folio}|RQ: {compra.req.folio} |Sol: {compra.req.orden.folio}',
                        body=html_message,
                        #f'Estimado {requi.orden.staff.staff.staff.first_name} {requi.orden.staff.staff.staff.last_name},\n Estás recibiendo este correo porque tu solicitud: {requi.orden.folio}| Req: {requi.folio} ha sido autorizada,\n por {requi.requi_autorizada_por.staff.staff.first_name} {requi.requi_autorizada_por.staff.staff.last_name}.\n El siguiente paso del sistema: Generación de OC \n\n Este mensaje ha sido automáticamente generado por SAVIA VORDTEC',
                        from_email = settings.DEFAULT_FROM_EMAIL,
                        to= [compra.creada_por.staff.staff.email],
                        headers={'Content-Type': 'text/html'}
                        )
                    email.content_subtype = "html " # Importante para que se interprete como HTML
                    email.send()
                    messages.success(request, f'{usuario.staff.staff.first_name} has autorizado la compra {compra.folio}')
                except (BadHeaderError, SMTPException, socket.gaierror) as e:
                    error_message = f'{usuario.staff.staff.first_name} has autorizado la compra {compra.folio} pero el correo de notificación no ha sido enviado debido a un error: {e}'
                    messages.success(request, error_message)    
            return redirect('autorizacion-oc2')

    context={
        'form':form,
        'compra':compra,
        'costo_oc':costo_oc,
        'productos':productos,
        'tipo_cambio':compra.tipo_de_cambio,
        'resta':resta,
        'porcentaje':porcentaje,
        'costo_total':costo_total,
        }

    return render(request, 'compras/autorizar_oc2.html',context)

def avisar_calidad_oc(compra):
    """
    Envía un correo al encargado de calidad del distrito avisando que
    una OC próxima a llegar contiene productos críticos.
    """
    distrito = compra.req.orden.proyecto.distrito  # ajusta según tu modelo
     # 🔍 Buscar artículos de la OC cuyos productos sean críticos
    productos_criticos = ArticuloComprado.objects.filter(
        oc=compra,
        producto__producto__articulos__producto__producto__critico=True
    ).distinct()
    if not productos_criticos.exists():
        return  # nada que avisar

    encargados = Profile.objects.filter(tipo__calidad=True, distritos=distrito)
    if not encargados.exists():
        return

    # Tabla con los productos críticos
    articulos_html = """
        <table width="100%" border="1" cellspacing="0" cellpadding="5"
               style="border-collapse: collapse; font-size: 14px;">
            <tr style="background-color: #eee;">
                <th>Nombre</th>
                <th>Código</th>
                <th>Unidad</th>
                <th>Familia</th>
                <th>Subfamilia</th>
                <th>Cantidad</th>
                <th>Requerimientos</th>
                <th>Comentarios</th>
            </tr>
    """
    for art in productos_criticos:
        p = art.producto.producto.articulos.producto.producto
        requerimientos = []
        if hasattr(p, 'producto_calidad'):  # Verifica que tenga criticidad cargada
            requerimientos = p.producto_calidad.requerimientos_calidad.all()

        if requerimientos:
            first_row = True
            for req in requerimientos:
                articulos_html += f"""
                    <tr>
                        <td>{p.nombre if first_row else ''}</td>
                        <td>{p.codigo if first_row else ''}</td>
                        <td>{p.unidad.nombre if first_row and p.unidad else ''}</td>
                        <td>{p.familia.nombre if first_row and p.familia else ''}</td>
                        <td>{p.subfamilia.nombre if first_row and p.subfamilia else ''}</td>
                        <td>{art.cantidad if first_row else ''}</td>
                        <td>{req.requerimiento.nombre}</td>
                        <td>{req.comentarios or ''}</td>
                    </tr>
                """
                first_row = False
        else:
            articulos_html += f"""
                <tr>
                    <td>{p.nombre}</td>
                    <td>{p.codigo}</td>
                    <td>{p.unidad.nombre if p.unidad else ''}</td>
                    <td>{p.familia.nombre if p.familia else ''}</td>
                    <td>{p.subfamilia.nombre if p.subfamilia else ''}</td>
                    <td>{art.cantidad}</td>
                    <td>Sin requerimientos</td>
                    <td>Sin comentarios</td>
                </tr>
            """

    articulos_html += "</table>"

    # Cargar imágenes base64
    static_path = settings.STATIC_ROOT
    img_path = os.path.join(static_path,'images','SAVIA_Logo.png')
    img_path2 = os.path.join(static_path,'images','logo_vordcab.jpg')
    image_base64 = get_image_base64(img_path)
    logo_v_base64 = get_image_base64(img_path2)

    # HTML del correo
    html_message = f"""
    <html>
        <head><meta charset="UTF-8"></head>
        <body style="font-family: Arial, sans-serif; color: #333; background-color: #f4f4f4; margin: 0; padding: 0;">
            <table width="100%" cellspacing="0" cellpadding="0" style="background-color: #f4f4f4; padding: 20px;">
                <tr>
                    <td align="center">
                        <table width="600px" cellspacing="0" cellpadding="0"
                               style="background-color: #ffffff; padding: 20px; border-radius: 10px;">
                            <tr>
                                <td align="center">
                                    <img src="data:image/jpeg;base64,{logo_v_base64}" alt="Logo" style="width: 100px; height: auto;" />
                                </td>
                            </tr>
                            <tr>
                                <td style="padding: 20px;">
                                    <p style="font-size: 16px; text-align: justify;">
                                        Estimado(a) encargado(a) de calidad del distrito <strong>{distrito.nombre}</strong>,
                                    </p>
                                    <p style="font-size: 16px; text-align: justify;">
                                        Está próxima a llegar a almacén una <strong>Orden de Compra</strong> que contiene el siguiente o siguientes productos marcados como <strong>críticos</strong>:
                                    </p>
                                    <p style="font-size: 15px; text-align: justify;">
                                        Folio de OC: <strong>{compra.folio}</strong><br>
                                        Proveedor: <strong>{compra.proveedor.nombre}</strong>
                                    </p>
                                    {articulos_html}
                                    <p style="text-align: center; margin: 20px 0;">
                                        <img src="data:image/png;base64,{image_base64}" alt="Imagen" style="width: 50px; height: auto; border-radius: 50%;" />
                                    </p>
                                    <p style="font-size: 13px; color: #777; text-align: justify;">
                                        Este mensaje ha sido generado automáticamente por <strong>SAVIA 2.0</strong>.
                                    </p>
                                </td>
                            </tr>
                        </table>
                    </td>
                </tr>
            </table>
        </body>
    </html>
    """

    # Enviar correo a cada encargado de calidad
    for encargado in encargados:
        try:
            email = EmailMessage(
                subject=f"OC próxima a llegar con productos críticos | {compra.folio}",
                body=html_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[encargado.staff.staff.email],
                headers={'Content-Type': 'text/html'}
            )
            email.content_subtype = "html"  # importante para interpretar como HTML
            email.send()
        except Exception as e:
            print(f"Error enviando correo a {encargado.staff.staff.email}: {e}")

def handle_uploaded_file(file, model_instance, field_name):
    setattr(model_instance, field_name, file)
    model_instance.save()

@perfil_seleccionado_required
def comparativos(request):
    #creada_por 
    pk_perfil = request.session.get('selected_profile_id') 
    usuario = Profile.objects.get(id = pk_perfil)
    comparativos = Comparativo.objects.filter(completo = True, creada_por__distritos = usuario.distritos)
    form = UploadFileForm()
    error_messages = {}

    myfilter = ComparativoFilter(request.GET, queryset=comparativos)
    comparativos = myfilter.qs

     #Set up pagination
    p = Paginator(comparativos, 50)
    page = request.GET.get('page')
    comparativos_list = p.get_page(page)

    if request.method == "POST":
        mi_id = request.POST.get('mi_id')
        form = UploadFileForm(request.POST or None, request.FILES or None)
        files = request.FILES.getlist('file')
        cont = 0
        if form.is_valid():
            comparativo = comparativos.get(id=mi_id)
            for f in files:
                if cont == 0:
                    comparativo.cotizacion = files[cont]
                elif cont == 1:
                    comparativo.cotizacion2 = files[cont]
                elif cont == 2:
                    comparativo.cotizacion3 = files[cont]
                elif cont == 3:
                    comparativo.cotizacion4 = files[cont]
                elif cont == 4:
                    comparativo.cotizacion5 = files[cont]
                cont+=1
                comparativo.save()
            messages.success(request,f'Entrando {files}')
            return redirect('comparativos')
        else:
            for field, errors in form.errors.items():
                error_messages[field] = errors.as_text()

    
    context= {
        'myfilter':myfilter,
        'error_messages': error_messages,
        'comparativos':comparativos,
        'form':form,
        'comparativos_list':comparativos_list,
    }
    return render(request,'compras/comparativos.html', context)



@perfil_seleccionado_required
def crear_comparativo(request):
    start_time = time.time()  # Tiempo de inicio
    pk_perfil = request.session.get('selected_profile_id')
    colaborador_sel = Profile.objects.all()
    usuario = colaborador_sel.get(id = pk_perfil)
    
    comparativo, created = Comparativo.objects.get_or_create(completo= False, creada_por=usuario)
    productos = Item_Comparativo.objects.filter(comparativo = comparativo, completo = True)
    error_messages = {}
    form_item = Item_ComparativoForm()
    form = ComparativoForm()

    if request.method =='POST':
        if "btn_creacion" in request.POST:
            form = ComparativoForm(request.POST, instance=comparativo)
            if form.is_valid():
                comparativo = form.save(commit=False)
                comparativo.completo = True
                comparativo.created_at = date.today()
                comparativo.creado_por =  usuario
                comparativo.save()
                messages.success(request, f'El comparativo {comparativo.id} ha sido creado')
                return redirect('comparativos')
            else:
                for field, errors in form_item.errors.items():
                    error_messages[field] = errors.as_text()
                messages.error(request,f'No está validando {error_messages}' )
        if "btn_producto" in request.POST:
            articulo, created = Item_Comparativo.objects.get_or_create(completo = False, comparativo = comparativo)
            form_item = Item_ComparativoForm(request.POST, instance= articulo)
            if form_item.is_valid():
                articulo = form_item.save(commit=False)
                articulo.completo = True
                articulo.save()
                messages.success(request, 'Se ha agregado el artículo exitosamente')
                return redirect('crear_comparativo')
            else:
                for field, errors in form_item.errors.items():
                    error_messages[field] = errors.as_text()
                messages.error(request,f'No está validando {error_messages}' )
        if "btn_files" in request.POST:
            #mi_id = request.POST.get('mi_id')
            form = UploadFileForm(request.POST or None, request.FILES or None)
            files = request.FILES.getlist('file')
            cont = 0
            if form.is_valid():
                #comparativo = comparativos.get(id=mi_id)
                #comparativo = form.save(commit = False)
                for f in files:
                    if cont == 0:
                        comparativo.cotizacion = files[cont]
                    elif cont == 1:
                        comparativo.cotizacion2 = files[cont]
                    elif cont == 2:
                        comparativo.cotizacion3 = files[cont]
                    elif cont == 3:
                        comparativo.cotizacion4 = files[cont]
                    elif cont == 4:
                        comparativo.cotizacion5 = files[cont]
                    cont+=1
                comparativo.save()
                messages.success(request,f'Entrando {files}')
                return redirect('crear_comparativo')
        else:
            post_data = request.POST
            post_items = [f"{key}: {value}" for key, value in post_data.items()]
            post_content = ", ".join(post_items)
            messages.error(request, f"Datos recibidos en POST: {post_content}")
        

    end_time = time.time()  # Tiempo de finalización
    total_time = end_time - start_time  # Tiempo total de ejecución    
    
    context= {
        'error_messages': error_messages,
        'productos':productos,
        'form':form,
        'form_item':form_item,
        'comparativo':comparativo,
        'total_time': total_time,
    }

    return render(request, 'compras/crear_comparativo.html', context)



#Ajax Select2
def carga_proveedor(request):
    pk_perfil = request.session.get('selected_profile_id')
    colaborador_sel = Profile.objects.all()
    usuario = colaborador_sel.get(id = pk_perfil)
    term = request.GET.get('term')
    proveedores = Proveedor_direcciones.objects.filter(
         Q(estatus__nombre="NUEVO") | Q(estatus__nombre="APROBADO"),
         distrito = usuario.distritos, 
         nombre__razon_social__icontains = term
    ).values('id','nombre__razon_social','distrito__nombre','domicilio','estatus__nombre','financiamiento','dias_credito','moneda__nombre')
    data = list(proveedores)
    print(proveedores)
    return JsonResponse(data, safe=False)

def carga_proveedor_comparativo(request):
    pk_perfil = request.session.get('selected_profile_id')
    colaborador_sel = Profile.objects.all()
    usuario = colaborador_sel.get(id = pk_perfil)
    term = request.GET.get('term')
    proveedores = Proveedor.objects.filter(razon_social__icontains = term).values('id','razon_social')
    data = list(proveedores)
        
    return JsonResponse(data, safe=False)
    
#Ajax Select2
def carga_productos(request):
    pk_perfil = request.session.get('selected_profile_id')
    colaborador_sel = Profile.objects.all()
    usuario = colaborador_sel.get(id = pk_perfil)
    term = request.GET.get('term')
    articulos = Inventario.objects.filter(distrito = usuario.distritos, producto__nombre__icontains = term).values('id','producto__nombre')
    
    #data = [{"id": item['id'], "text": item['producto__nombre']} for item in articulos]
    data = list(articulos)
        
    return JsonResponse(data, safe=False)

@perfil_seleccionado_required
def editar_comparativo(request, pk):
    pk_perfil = request.session.get('selected_profile_id')
    #colaborador_sel = Profile.objects.all()
    usuario = Profile.objects.get(id = pk_perfil)
    #usuario = Profile.objects.get(staff__id=request.user.id)
    comparativo =Comparativo.objects.get(id = pk)
    productos = Item_Comparativo.objects.filter(comparativo = comparativo, completo = True)
    proveedores = Proveedor_direcciones.objects.all()
    articulos = Inventario.objects.all()
    form_item = Item_ComparativoForm()
    form = ComparativoForm(instance = comparativo)

    if request.method =='POST':
        if "btn_agregar" in request.POST:
            form = ComparativoForm(request.POST, request.FILES, instance = comparativo)
            #abrev= usuario.distrito.abreviado
            if form.is_valid():
                comparativo = form.save(commit=False)
                comparativo.completo = True
                comparativo.created_at = date.today()
                #comparativo.created_at_time = datetime.now().time()
                comparativo.creado_por =  usuario
                comparativo.save()
                #form.save()
                messages.success(request, f'El comparativo {comparativo.id} ha sido modificado')
                return redirect('comparativos')
        if "btn_producto" in request.POST:
            articulo, created = Item_Comparativo.objects.get_or_create(completo = False, comparativo = comparativo)
            #producto_id = request.POST.get('producto')
            #producto = Inventario.objects.filter(id = producto_id)
            #form_item.fields['producto'].queryset = producto
            form_item = Item_ComparativoForm(request.POST, instance=articulo)
            if form_item.is_valid():
                articulo = form_item.save(commit=False)
                articulo.completo = True
                articulo.save()
                messages.success(request, 'Se ha agregado el artículo exitosamente')
                return redirect('editar-comparativo')
        
    context= {
        'productos':productos,
        'form':form,
        'form_item':form_item,
        'articulos':articulos,
        'comparativo':comparativo,
        'proveedores':proveedores,
    }

    return render(request, 'compras/actualizar_comparativo.html', context)


@perfil_seleccionado_required
def articulos_comparativo(request, pk):
    articulos = Item_Comparativo.objects.filter(comparativo__id = pk , completo = True)

    context= {
        'articulos':articulos,
    }
    return render(request, 'compras/articulos_comparativo.html', context)

def articulo_comparativo_delete(request, pk):
   
    articulo = Item_Comparativo.objects.get(id=pk)
    comparativo = articulo.comparativo.id
   
    messages.success(request,f'El articulo ha sido eliminado exitosamente')
    articulo.delete()

    return redirect('crear_comparativo')

@perfil_seleccionado_required
def historico_articulos_compras(request):
    registros = ArticuloComprado.history.all()

    myfilter = HistoricalArticuloCompradoFilter(request.GET, queryset=registros)
    registros = myfilter.qs

    #Set up pagination
    p = Paginator(registros, 30)
    page = request.GET.get('page')
    registros_list = p.get_page(page)

    context = {
        'registros_list':registros_list,
        'myfilter':myfilter,
        }

    return render(request,'compras/historico_articulos_comprados.html',context)


@perfil_seleccionado_required
def historico_compras(request):
    registros = Compra.history.all()
    myfilter = HistoricalCompraFilter(request.GET, queryset=registros)
    registros = myfilter.qs
    #Set up pagination
    p = Paginator(registros, 30)
    page = request.GET.get('page')
    registros_list = p.get_page(page)
    
    context = {
        'registros_list':registros_list,
        'myfilter':myfilter,
        }
    return render(request,'compras/historico_compras.html',context)


def descargar_pdf(request, pk):
    compra = get_object_or_404(Compra, id=pk)
    buf = generar_pdf_nueva(compra)
    return FileResponse(buf, as_attachment=True, filename='oc_' + str(compra.folio) + '.pdf')

def attach_oc_pdf(request, pk):
    compra = get_object_or_404(Compra, id=pk)
    buf = generar_pdf(compra)

    return buf.getvalue()

def attach_politica_proveedor(request):
    #compra = get_object_or_404(Compra, id=pk)
    buf = generar_politica_proveedores()

    return buf.getvalue()

def descargar_antisoborno_pdf(request):
    buf = generar_politica_antisoborno()
    return FileResponse(buf, as_attachment=True, filename='Política_Antisoborno' +'.pdf')

def descargar_proveedores_pdf(request):
    buf = generar_politica_proveedores()
    return FileResponse(buf, as_attachment=True, filename='Política_Proveedores' +'.pdf')

def attach_antisoborno_pdf(request):
    buf = generar_politica_antisoborno()
    return buf.getvalue()

def descargar_codigo_etica_pdf(request):
    buf = generar_codigo_etica()
    return FileResponse(buf, as_attachment=True, filename='Código_ética' +'.pdf')

def attach_codigo_etica_pdf(request):
    buf = generar_codigo_etica()
    return buf.getvalue()

def descargar_aviso_privacidad_pdf(request):
    buf = generar_aviso_privacidad()
    return FileResponse(buf, as_attachment=True, filename='Aviso_de_Privacidad' +'.pdf')

def attach_aviso_privacidad_pdf(request):
    buf = generar_aviso_privacidad()
    return buf.getvalue()

from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import Paragraph
from reportlab.lib.units import cm

def wrap_text_to_fit(text, width, style):
    # Determinar la anchura máxima del texto que cabe en la columna
    words = text.split()
    lines = []
    line = []

    for word in words:
        # Comprueba si agregar la palabra excedería la longitud de la línea
        test_line = ' '.join(line + [word])
        if stringWidth(test_line, style.fontName, style.fontSize) > width:
            # Si la línea es demasiado larga, guarda la línea actual y comienza una nueva
            lines.append(' '.join(line))
            line = [word]
        else:
            line.append(word)
    lines.append(' '.join(line))  # Añade la última línea
    wrapped_text = '\n'.join(lines)
    return wrapped_text


def generar_pdf(compra):
    #Configuration of the PDF object
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    #doc = SimpleDocTemplate(buf, pagesize=letter)
    #Here ends conf.
    #compra = Compra.objects.get(id=pk)
    productos = ArticuloComprado.objects.filter(oc=compra.id)

    #Azul Vordcab
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)
    caja_iso = 760
    #Elaborar caja
    #c.line(caja_iso,500,caja_iso,720)
    c.drawString(430,caja_iso,'Preparado por:')
    c.drawString(405,caja_iso-10,'SUPT. DE ADQUISIONES')
    c.drawString(520,caja_iso,'Aprobación')
    c.drawString(515,caja_iso-10,'SUBD ADTVO')
    c.drawString(150,caja_iso-20,'Número de documento')
    c.drawString(160,caja_iso-30,'SEOV-ADQ-N4-01.02')
    c.drawString(245,caja_iso-20,'Clasificación del documento')
    c.drawString(275,caja_iso-30,'Controlado')
    c.drawString(355,caja_iso-20,'Nivel del documento')
    c.drawString(380,caja_iso-30, 'N5')
    c.drawString(440,caja_iso-20,'Revisión No.')
    c.drawString(452,caja_iso-30,'003')
    c.drawString(510,caja_iso-20,'Fecha de Emisión')
    c.drawString(525,caja_iso-30,'13/11/2017')

    caja_proveedor = caja_iso - 65
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(150,750,250,20, fill=True, stroke=False) #Barra azul superior Orden de Compra
    c.rect(20,caja_proveedor - 8,565,20, fill=True, stroke=False) #Barra azul superior Proveedor | Detalle
    c.rect(20,520,565,2, fill=True, stroke=False) #Linea posterior horizontal
    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(280,755,'Orden de compra')
    c.setLineWidth(.3) #Grosor
    c.line(20,caja_proveedor-8,20,520) #Eje Y donde empieza, Eje X donde empieza, donde termina eje y,donde termina eje x (LINEA 1 contorno)
    c.line(585,caja_proveedor-8,585,520) #Linea 2 contorno
    c.drawInlineImage('static/images/logo_vordcab.jpg',45,730, 3 * cm, 1.5 * cm) #Imagen vordcab

    c.setFillColor(white)
    c.setFont('Helvetica-Bold',11)
    c.drawString(120,caja_proveedor,'Proveedor')
    c.drawString(400,caja_proveedor, 'Detalles')
    inicio_central = 300
    c.line(inicio_central,caja_proveedor-25,inicio_central,520) #Linea Central de caja Proveedor | Detalle
    c.setFillColor(black)
    c.setFont('Helvetica',9)
    c.drawString(30,caja_proveedor-20,'Nombre:')
    c.drawString(30,caja_proveedor-40,'RFC:')
    c.drawString(30,caja_proveedor-60,'Uso del CFDI:')
    c.drawString(30,caja_proveedor-80,'Solicitó:')
    c.drawString(30,caja_proveedor-100,'Fecha:')
    c.drawString(30,caja_proveedor-120,'Proveedor Calif:')
    c.drawString(30,caja_proveedor-140,'Tiempo de Entrega:')
    c.drawString(30,caja_proveedor-160,'A.F.:')

    c.setFont('Helvetica-Bold',12)
    c.drawString(500,caja_proveedor-20,'FOLIO:')

    c.setFillColor(rojo)
    c.setFont('Helvetica-Bold',12)
    c.drawString(540,caja_proveedor-20, str(compra.folio))

    c.setFillColor(black)
    c.setFont('Helvetica',9)
    c.drawString(inicio_central + 10,caja_proveedor-35,'No. Requisición:')
    c.drawString(inicio_central + 10,caja_proveedor-55,'Método de pago:')
    c.drawString(inicio_central + 10,caja_proveedor-75,'Condiciones de pago:')
    c.drawString(inicio_central + 10,caja_proveedor-95,'Enviar Factura a:')
    c.drawString(inicio_central + 10,caja_proveedor-115,'Banco:')
    c.drawString(inicio_central + 10,caja_proveedor-135,'Cuenta:')
    c.drawString(inicio_central + 10,caja_proveedor-155,'Clabe:')

    c.setFillColor(black)
    c.setFont('Helvetica',9)
    if compra.proveedor.nombre.razon_social == 'COLABORADOR':
        c.drawString(100,caja_proveedor-20, compra.deposito_comprador.staff.staff.first_name+' '+compra.deposito_comprador.staff.staff.last_name)
    else:
        c.drawString(100,caja_proveedor-20, compra.proveedor.nombre.razon_social)
    c.drawString(100,caja_proveedor-40, compra.proveedor.nombre.rfc)
    c.drawString(100,caja_proveedor-60, compra.uso_del_cfdi.descripcion)
    c.drawString(100,caja_proveedor-80, compra.req.orden.staff.staff.staff.first_name +' '+ compra.req.orden.staff.staff.staff.last_name)
    c.drawString(100,caja_proveedor-100, compra.created_at.strftime("%d/%m/%Y"))
    c.drawString(100,caja_proveedor-120, compra.estatus_original)
    if compra.dias_de_entrega:
        c.drawString(110,caja_proveedor-140, str(compra.dias_de_entrega)+' '+'días hábiles')
    
    
    try:
        if compra.req.orden.activo is not None:
            eco_unidad = compra.req.orden.activo.eco_unidad
            descripcion = compra.req.orden.activo.descripcion
            c.drawString(60, caja_proveedor-160, f'{eco_unidad} {descripcion}')
        else:
            c.drawString(60, caja_proveedor-160, 'NA')
    except Activo.DoesNotExist:
        c.drawString(60, caja_proveedor-160, 'NA')


    c.drawString(inicio_central + 90,caja_proveedor-35, str(compra.req.folio))
    c.drawString(inicio_central + 90,caja_proveedor-55, 'Transferencia Electrónica')
    c.drawString(inicio_central + 90,caja_proveedor-95, compra.creada_por.staff.staff.email) #Esta parte hay que configurarla para que cambie de acuerdo al distrito
    if compra.proveedor.nombre.razon_social == 'COLABORADOR':
        c.drawString(inicio_central + 90,caja_proveedor-115, compra.deposito_comprador.banco.nombre)
        c.drawString(inicio_central + 90,caja_proveedor-135, compra.deposito_comprador.cuenta_bancaria)
        c.drawString(inicio_central + 90,caja_proveedor-155, compra.deposito_comprador.clabe)
    else:
        c.drawString(inicio_central + 90,caja_proveedor-115, compra.proveedor.banco.nombre)
        c.drawString(inicio_central + 90,caja_proveedor-135, compra.proveedor.cuenta)
        c.drawString(inicio_central + 90,caja_proveedor-155, compra.proveedor.clabe)




    if compra.cond_de_pago.nombre == "CREDITO":
        c.drawString(inicio_central + 100,caja_proveedor-75, compra.cond_de_pago.nombre + '  ' + str(compra.dias_de_credito) + ' días')
    else:
        c.drawString(inicio_central + 100,caja_proveedor-75, compra.cond_de_pago.nombre )

    
    data =[]
    cont = 0
    high = 495
    styles = getSampleStyleSheet()
    style_desc = styles["BodyText"]
    style_desc.wordWrap = 'CJK'
    style_desc.fontSize = 6
    style_desc.leading = 8

    data.append(['''Código''','''Producto''', '''Cantidad''', '''Unidad''', '''P.Unitario''', '''Importe'''])
    for producto in productos:
        importe = producto.precio_unitario * producto.cantidad
        importe_rounded = round(importe, 4)
        descripcion = Paragraph(producto.producto.producto.articulos.producto.producto.nombre, style_desc)
        precio_unitario = f"{producto.precio_unitario:,.4f}"
        data.append([
            producto.producto.producto.articulos.producto.producto.codigo,
            descripcion,
            producto.cantidad, 
            producto.producto.producto.articulos.producto.producto.unidad,
            precio_unitario,
            f"{importe_rounded:,.4f}"
        ])
        cont = cont + 1
        if cont < 16:
            high = high - 18

    c.setFillColor(black)
    c.setFont('Helvetica',8)

    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(20,200,340,20, fill=True, stroke=False) #2ra linea azul, donde esta el proyecto y el subproyecto, se coloca altura de 150
    c.setFillColor(black)
    c.setFont('Helvetica',7)

    c.setFillColor(white)
    c.setLineWidth(.1)
    c.setFont('Helvetica-Bold',10)
    c.drawString(25,205,'Proyecto:')
    c.drawString(100,205,compra.req.orden.proyecto.nombre)
    c.setFillColor(black)
    c.drawString(25,190,'Subproyecto:')
    c.drawString(25,175,'Elaboró:')
    c.drawString(25,160,'Moneda:')
    c.setFont('Helvetica',8)
    
    
    c.drawString(100,190,compra.req.orden.subproyecto.nombre)
    c.drawString(100,175,compra.creada_por.staff.staff.first_name + ' ' +compra.creada_por.staff.staff.last_name)
    c.drawString(100,160,compra.moneda.nombre)

    c.setLineWidth(.3)
    c.line(370,220,370,160) #Eje Y donde empieza, Eje X donde empieza, donde termina eje y,donde termina eje x (LINEA 1 contorno)
    c.line(370,160,580,160)

    c.setFillColor(black)
    c.setFont('Helvetica-Bold',9)

    montos_align = 480
    c.drawRightString(montos_align,210,'Sub Total:')
    c.drawRightString(montos_align,200,'IVA 16%:')
    c.drawRightString(montos_align,190,'Importe Neto:')
    
    c.setFillColor(prussian_blue)
    c.setFillColor(black)
    c.drawString(20,130,'Opciones y condiciones:')
    c.drawString(20,105,'Comentario Solicitud:')
    c.setFont('Helvetica',8)
    letras = 320
    c.drawString(20,140,'Total con letra:')
    #c.line(135,90,215,90 ) #Linea de Autorizacion
    #c.line(350,90,430,90)
    c.drawCentredString(175,70,'Autorización')
    c.drawCentredString(390,70,'Autorización')

    c.drawCentredString(175,80,'Superintendente Administrativo')
    c.drawCentredString(390,80,'Gerencia Zona')
    c.setFont('Helvetica-Bold',8)
    if compra.autorizado1 ==  True:
        c.drawCentredString(175,90,compra.oc_autorizada_por.staff.staff.first_name + ' ' +compra.oc_autorizada_por.staff.staff.last_name)
    if compra.autorizado2 == True:
        c.drawCentredString(390,90,compra.oc_autorizada_por2.staff.staff.first_name + ' ' + compra.oc_autorizada_por2.staff.staff.last_name)

    c.setFont('Helvetica',10)
    subtotal = compra.costo_oc - compra.costo_iva
    
    importe_neto = compra.costo_oc
    if compra.impuestos:
        subtotal = subtotal #- compra.impuestos
        c.setFillColor(black)
        c.setFont('Helvetica-Bold',9)
        #c.drawRightString(montos_align,170,'Impuestos Adicionales:')
        c.setFont('Helvetica',10)
        costo_impuestos = format(float(compra.impuestos), ',.2f')
        c.drawRightString(montos_align + 90, 180, '$' + str(costo_impuestos))
        c.drawRightString(montos_align, 180, 'Impuestos:')
        #importe_neto = importe_neto + compra.impuestos
    if compra.impuestos and compra.retencion:
        subtotal = subtotal #+ compra.retencion
        costo_retencion = format(float(compra.retencion), ',.2f')
        c.drawRightString(montos_align + 90, 170, '$' + str(costo_retencion))
        c.drawRightString(montos_align, 170, 'Retención:')
    elif compra.retencion:
        subtotal = subtotal #+ compra.retencion
        costo_retencion = format(float(compra.retencion), ',.2f')
        c.drawRightString(montos_align + 90, 180, '$' + str(costo_retencion))
        c.drawRightString(montos_align, 180, 'Retención:')
        #importe_neto = importe_neto - compra.retencion
    costo_subtotal = format(float(subtotal), ',.2f')
    c.drawRightString(montos_align + 90,210,'$ ' + str(costo_subtotal))
    costo_con_iva = format(float(compra.costo_iva), ',.2f')
    c.drawRightString(montos_align + 90,200,'$ ' + str(costo_con_iva))
    costo_oc =  format(float(compra.costo_oc), ',.2f')
    c.drawRightString(montos_align + 90,190,'$ ' + str(costo_oc))
    
   
    #if compra.costo_fletes is None:
    #c.setFillColor(prussian_blue)
       
    total =  format(float(compra.costo_plus_adicionales), ',.2f')
    if compra.impuestos and compra.retencion and compra.costo_fletes:
        c.drawRightString(montos_align,150,'Total:')
        c.drawRightString(montos_align + 90,150,'$ ' + str(total))
    elif compra.impuestos and compra.retencion:
        c.drawRightString(montos_align,160,'Total:')
        c.drawRightString(montos_align + 90,160,'$ ' + str(total))
    elif compra.impuestos or compra.retencion:
        c.drawRightString(montos_align,170,'Total:')
        c.drawRightString(montos_align + 90,170,'$ ' + str(total))

    if compra.costo_fletes:
        importe_neto = importe_neto + compra.costo_fletes
        c.drawRightString(montos_align,170,'Total:')
        c.drawRightString(montos_align,180,'Costo fletes:')
        c.drawRightString(montos_align + 90,180,'$ ' + str(compra.costo_fletes))
        c.drawRightString(montos_align + 90,170,'$ ' + str(total))
    
    
    c.setFillColor(prussian_blue)
    c.setFont('Helvetica', 9)
    
    def convertir_a_reales(valor):
        partes = f"{valor:.2f}".split(".")
        reales = num2words(int(partes[0]), lang='pt_BR')
        centavos = num2words(int(partes[1]), lang='pt_BR')
        return f"{reales} reais e {centavos} centavos"

    if compra.moneda.nombre == "PESOS":
        c.drawString(80,140, num2words(compra.costo_plus_adicionales, lang='es', to='currency', currency='MXN'))
    if compra.moneda.nombre == "DOLARES":
        c.drawString(80,140, num2words(compra.costo_plus_adicionales, lang='es', to='currency',currency='USD'))
    if compra.moneda.nombre == "REAIS":
        c.drawString(80,140, convertir_a_reales(compra.costo_plus_adicionales))

    c.setFillColor(black)
    width, height = letter
   
    styleN = styles["BodyText"]
    styleN.fontSize = 6

    #Comentario de opciones y condiciones
    if compra.opciones_condiciones is not None:
        options_conditions = compra.opciones_condiciones
    else:
        options_conditions = "NA"

    options_conditions_paragraph = Paragraph(options_conditions, styleN)


    # Crear un marco (frame) en la posición específica
    frame = Frame(135, 0, width-155, height-648, id='normal')

    # Agregar el párrafo al marco
    frame.addFromList([options_conditions_paragraph], c)
    c.setFillColor(prussian_blue)
    c.rect(20,30,565,30, fill=True, stroke=False)
    c.setFillColor(white)

    #Comentario de solicitud
    if compra.comentario_solicitud:
        paragraph_content = compra.req.orden.comentario
    else:
        paragraph_content = "NA"  # O cualquier valor por defecto que prefieras

    # Crear el párrafo con el contenido basado en la condición
    if paragraph_content is None:
        paragraph_content = " "    
    conditional_paragraph = Paragraph(paragraph_content, styleN)

    # Crear un nuevo frame similar al anterior pero ajustando la posición y/o tamaño si es necesario
    # Asumiendo 'width' y 'height' ya están definidos como antes
    new_frame = Frame(120, 0, width-155, height-675, id='conditional_frame')

    # Agregar el párrafo al nuevo marco
    new_frame.addFromList([conditional_paragraph], c)
    
    

    table = Table(data, colWidths=[1.2 * cm, 12.5 * cm, 1.5 * cm, 1.2 * cm, 1.5 * cm, 1.5 * cm,])
    table_style = TableStyle([ #estilos de la tabla
        ('INNERGRID',(0,0),(-1,-1), 0.25, colors.white),
        ('BOX',(0,0),(-1,-1), 0.25, colors.black),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        #ENCABEZADO
        ('TEXTCOLOR',(0,0),(-1,0), white),
        ('FONTSIZE',(0,0),(-1,0), 8),
        ('BACKGROUND',(0,0),(-1,0), prussian_blue),
        #CUERPO
        ('TEXTCOLOR',(0,1),(-1,-1), colors.black),
        ('FONTSIZE',(0,1),(-1,-1), 6),
        ])
    table_style2 = TableStyle([ #estilos de la tabla
        ('INNERGRID',(0,0),(-1,-1), 0.25, colors.white),
        ('BOX',(0,0),(-1,-1), 0.25, colors.black),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        #ENCABEZADO
        ('TEXTCOLOR',(0,0),(-1,0), colors.black),
        ('FONTSIZE',(0,0),(-1,0), 6),
        #('BACKGROUND',(0,0),(-1,0), prussian_blue),
        #CUERPO
        ('TEXTCOLOR',(0,1),(-1,-1), colors.black),
        ('FONTSIZE',(0,1),(-1,-1), 6),
        ])
    table.setStyle(table_style)

    rows_per_page = 15
    total_rows = len(data) - 1  # Excluye el encabezado
    remaining_rows = total_rows - rows_per_page

    if remaining_rows <= 0:
        # Si no hay suficientes filas para una segunda página, dibujar la tabla completa en la primera página
        table.wrapOn(c, c._pagesize[0], c._pagesize[1])
        table.drawOn(c, 20, high)  # Posición en la primera página
    else:
        # Dibujar las primeras 15 filas en la primera página
        first_page_data = data[:rows_per_page + 1]  # Incluye el encabezado
        first_page_table = Table(first_page_data, colWidths=[1.2 * cm, 13 * cm, 1.5 * cm, 1.2 * cm, 1.5 * cm, 1.5 * cm])
        first_page_table.setStyle(table_style)
        first_page_table.wrapOn(c, c._pagesize[0], c._pagesize[1])
        first_page_table.drawOn(c, 20, high)  # Posición en la primera página

        # Agregar una nueva página y dibujar las filas restantes en la segunda página
        c.showPage()
        remaining_data = data[rows_per_page + 1:]
        remaining_table = Table(remaining_data, colWidths=[1.2 * cm, 13 * cm, 1.5 * cm, 1.2 * cm, 1.5 * cm, 1.5 * cm])
        remaining_table.setStyle(table_style2)
        remaining_table.wrapOn(c, c._pagesize[0], c._pagesize[1])
        remaining_table_height = len(remaining_data) * 18
        remaining_table_y = c._pagesize[1] - 70 - remaining_table_height - 10  # Espacio para el encabezado
        remaining_table.drawOn(c, 20, remaining_table_y)  # Posición en la segunda página

        # Agregar el encabezado en la segunda página
        c.setFont('Helvetica', 8)
        c.drawString(420, caja_iso, 'Preparado por:')
        c.drawString(420, caja_iso - 10, 'SUP. ADMON')
        c.drawString(520, caja_iso, 'Aprobación')
        c.drawString(520, caja_iso - 10, 'SUB ADM')
        c.drawString(150, caja_iso - 20, 'Número de documento')
        c.drawString(160, caja_iso - 30, 'F-ADQ-N4-01.02')
        c.drawString(245, caja_iso - 20, 'Clasificación del documento')
        c.drawString(275, caja_iso - 30, 'Controlado')
        c.drawString(355, caja_iso - 20, 'Nivel del documento')
        c.drawString(380, caja_iso - 30, 'N5')
        c.drawString(440, caja_iso - 20, 'Revisión No.')
        c.drawString(452, caja_iso - 30, '000')
        c.drawString(510, caja_iso - 20, 'Fecha de Emisión')
        c.drawString(525, caja_iso - 30, '1-Sep.-18')

        caja_proveedor = caja_iso - 65
        c.setFont('Helvetica', 12)
        c.setFillColor(prussian_blue)
        c.rect(150, 750, 250, 20, fill=True, stroke=False)  # Barra azul superior Orden de Compra
        c.setFillColor(colors.white)
        c.setFont('Helvetica-Bold', 14)
        c.drawCentredString(280, 755, 'Orden de compra')
        c.drawInlineImage('static/images/logo_vordcab.jpg', 45, 730, 3 * cm, 1.5 * cm)  # Imagen vortec
    
    if compra.autorizado2 != True:
        c.setFont('Helvetica-Bold', 50)  # Tamaño de la fuente
        # Configurar color del texto (rojo claro, semitransparente)
        c.setFillColorRGB(1, 0, 0, alpha=0.2)
        # Guardar el estado del lienzo antes de rotar
        c.saveState()
        # Trasladar el origen de coordenadas al centro del lienzo
        c.translate(width / 2, height / 2)
        # Rotar el lienzo 45 grados
        c.rotate(45)
        # Dibujar el texto "NO AUTORIZADA" centrado
        c.drawCentredString(0, 0, "NO AUTORIZADA")

        # Restaurar el estado original del lienzo
        c.restoreState()
    #c.showPage()

    # --- NUEVA HOJA: Requerimientos de Calidad por producto ---
    rows_cal = []   # solo filas con requerimientos

    # (opcional) deduplicar por producto si la OC puede traer el mismo producto varias veces
    vistos = set()

    for ac in productos:
        try:
            base_product = ac.producto.producto.articulos.producto.producto
        except Exception:
            continue
        if not base_product:
            continue

        # (opcional) evita duplicados por Product
        if getattr(base_product, 'id', None) in vistos:
            continue
        vistos.add(base_product.id)

        # intenta tomar el OneToOne
        try:
            producto_calidad = base_product.producto_calidad
        except Producto_Calidad.DoesNotExist:
            continue

        if not producto_calidad:
            continue

        reqs = producto_calidad.requerimientos_calidad.select_related('requerimiento').all()
        if not reqs:
            continue

        codigo = base_product.codigo
        desc_par = Paragraph(base_product.nombre or '', style_desc)

        for req in reqs:
            req_nombre = req.requerimiento.nombre if req.requerimiento else ''
            comentario = req.comentarios or ''
            rows_cal.append([codigo, desc_par, req_nombre, comentario])

    # Si no hay nada que mostrar, NO crear la hoja
    if rows_cal:
        c.showPage()
        # encabezado hoja
        c.setFont('Helvetica', 12)
        c.setFillColor(prussian_blue)
        c.rect(20, 750, 565, 24, fill=True, stroke=False)
        c.setFillColor(white)
        c.setFont('Helvetica-Bold', 13)
        c.drawCentredString(300, 756, 'Requerimientos de Calidad por Producto')

        data_cal = [['Código', 'Producto', 'Requerimiento Calidad', 'Comentario']]
        data_cal.extend(rows_cal)

        col_widths = [2*cm, 8.0*cm, 5*cm, 5.0*cm]
        tabla_cal = Table(data_cal, colWidths=col_widths, hAlign='LEFT')
        tabla_cal.setStyle(TableStyle([
            ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
            ('BOX',       (0,0), (-1,-1), 0.50, colors.black),
            ('VALIGN',    (0,0), (-1,-1), 'MIDDLE'),
            ('BACKGROUND',(0,0), (-1,0),  prussian_blue),
            ('TEXTCOLOR', (0,0), (-1,0),  colors.white),
            ('FONTSIZE',  (0,0), (-1,0),  9),
            ('FONTSIZE',  (0,1), (-1,-1), 8),
        ]))

        tabla_cal.wrapOn(c, c._pagesize[0], c._pagesize[1])
        tabla_cal.drawOn(c, 20, 680 - min(520, len(data_cal)*16))
   
    c.showPage()
    # ------------ FIN NUEVA HOJA ------------------------------------------------
    
    # Agregar el encabezado en la segunda página
    c.setFont('Helvetica', 8)
    c.drawString(430, caja_iso, 'Preparado por:')
    c.drawString(420, caja_iso - 10, 'ASIST. TEC. SUBAD')
    c.drawString(525, caja_iso, 'Aprobación')
    c.drawString(520, caja_iso - 10, 'SUBD-ADTVO')
    c.drawString(50, caja_iso - 35, 'Número de documento')
    c.drawString(50, caja_iso - 45, 'SEOV-ADQ-N4-01.04')
    c.drawString(145, caja_iso - 35, 'Clasificación del documento')
    c.drawString(175, caja_iso - 45, 'No Controlado')
    c.drawString(255, caja_iso - 35, 'Nivel del documento')
    c.drawString(280, caja_iso - 45, 'N5')
    c.drawString(340, caja_iso - 35, 'Revisión No.')
    c.drawString(352, caja_iso - 45, '001')
    c.drawString(410, caja_iso - 35, 'Fecha de Emisión')
    c.drawString(425, caja_iso - 45, '03/05/23')
    c.drawString(500, caja_iso - 35, 'Fecha de Revisión')
    c.drawString(525, caja_iso - 45, '06/09/23')

    caja_proveedor = caja_iso - 65
    c.setFont('Helvetica', 12)
    c.setFillColor(prussian_blue)
    c.rect(155, 735, 250, 35, fill=True, stroke=False)  # Barra azul superior encabezado
    c.setFillColor(colors.white)
    c.setFont('Helvetica-Bold', 13)
    c.drawCentredString(280, 755, 'Requisitos en Materia de Gestión')
    c.drawCentredString(280, 740, 'de Seguridad, Salud y Medio Ambiente')
    c.drawInlineImage('static/images/logo_vordcab.jpg', 60, 735, 2 * cm, 1 * cm)  # Imagen vortec


    
    #c.setFont("Helvetica-Bold", 12)
    #c.setFillColor(prussian_blue)  # Asumiendo que prussian_blue está definido
    #c.drawCentredString(300, 680, )  # Ajusta la posición según sea necesario
    styles = getSampleStyleSheet()
    styleN = styles["Normal"]
    styleN.fontSize = 10
    #styleN.color = prussian_blue
    styleN.leading = 13  # Espaciado entre líneas
    styleN = ParagraphStyle('Justicado', parent=styles['Normal'], alignment=TA_JUSTIFY)
    styleT = styles["Normal"]
    styleT.fontSize = 13
    styleT.fontName = 'Helvetica-Bold'
    styleT.textColor = prussian_blue
    styleT.leading = 17
    styleT = ParagraphStyle('Center', parent=styles['Normal'], alignment= TA_CENTER)

    texto = """El objeto del presente escrito es poner en conocimiento, de los proveedores que realizan trabajos para nuestra empresa, 
                los requerimientos mínimos que le sean de aplicación, que debe de cumplir en lo referente a la seguridad y salud en el 
                trabajo, calidad y protección del medio ambiente. La aceptación de una orden de compra implica la responsabilidad por 
                parte del proveedor del conocimiento y aceptación de lo descrito.<br/>Nuestra empresa tiene implantado un SEOV 
                (Sistema de Excelencia Operativa Vordcab), bajo las normas ISO 9001:2015, ISO 14001:2015 e ISO 45001:2018, 
                por lo que se requiere que nuestros proveedores colaboren en el cumplimiento de los siguientes requisitos: <br/>
                """
    
    texto2 = """En todo momento el proveedor está obligado a cumplir con la legislación aplicable al servicio que está prestando.<br/>
El proveedor se compromete a garantizar el cumplimiento de lo solicitado en el pedido u orden de compra del material o del trabajo 
externo.<br/>
El proveedor debería poder garantizar la correcta gestión y control de: los residuos, emisiones atmosféricas, ruidos, 
efluentes residuales, productos peligrosos, afectación del suelo y mantenimiento de instalaciones, respecto al servicio prestado.<br/>
El proveedor deberá aplicar las medidas preventivas necesarias, para evitar situaciones de peligro o emergencia como derrame, fuga, 
incendio, etc., durante la realización del trabajo encomendado. Y si es preciso formar e informar a su personal sobre manipulación, 
almacenaje, uso y riesgos de productos o preparados peligrosos.<br/>
El proveedor deberá facilitar la documentación ambiental y de seguridad que se le solicite o requiera.<br/>
Si el proveedor tiene algún servicio que presta a nuestra empresa subcontratado, por ejemplo, transporte, deberá transmitir este 
documento a su subcontratista.<br/> """
    texto3 = """El proveedor deberá trabajar adoptando buenas prácticas ambientales y cumplir con los procedimientos internos que se 
    le hayan comunicado respecto a HSE y la Prevención de Riesgos Laborales.<br/>
Las empresas que realicen trabajos para nuestra empresa deberán cumplirse los requisitos de seguridad establecidos, uso de EPP, 
protecciones colectivas, formación, seguridad equipos,
medidas preventivas<br/>
Si durante el trabajo se deben retirar residuos peligrosos (aceites, aguas con productos peligrosos, pinturas, disolventes, sus envases)
o /y no peligrosos (escombros, embalajes), se recogerán en recipientes adecuados según la cantidad a retirar. Los recipientes llenos
serán retirados por la empresa que preste el servicio o entregados a nuestra empresa, según convenga, para su correcto almacenamiento 
y gestión posterior.<br/>
Si no sabe qué hacer con algún residuo o efluente residual o cualquier otra cosa, que se ha originado mientras prestaba el servicio,
 avise a la persona que le haya atendido o al Responsable de seguridad. No tomará decisiones.<br/>
Si detecta cualquier situación de riesgo/emergencia (ambiental, personal, de seguridad), lo comunicará de inmediato a cualquier persona
de la empresa o a la que le haya atendido o al Responsable de seguridad. Nunca actué. Deberá seguir las normas de emergencia que se le
 han facilitado a la entrada.<br/>
Si mientras trabaja se produce una situación de emergencia (derrame o fuga de producto peligrosos), y le han facilitado la formación y 
medios necesarios para actuar, proceda según la instrucción que se le ha facilitado, en caso contrario avise de inmediato a cualquier 
persona de la empresa o a la que le ha atendido o al Responsable de seguridad.<br/>"""
    titulo1 = "REQUISITOS OBLIGATORIOS PARA PROVEEDORES DE GRUPO VORDCAB"
    titulo2 = """REQUISITOS GENERALES PARA TODOS LOS PROVEEDORES"""
    titulo3 = """REQUISITOS SI USTED PRESTA EL SERVICIO O PARTE DE SERVICIO DENTRO <br/> 
                DE LAS INSTALACIONES DE GRUPO VORDCAB"""
    
    titulo1 = Paragraph(titulo1, styleT)
    parrafo = Paragraph(texto, styleN)
    titulo2 = Paragraph(titulo2, styleT)
    parrafo2 = Paragraph(texto2, styleN)
    titulo3 = Paragraph(titulo3, styleT)
    parrafo3 = Paragraph(texto3, styleN)

    ancho, alto = letter  # Asegúrate de tener estas dimensiones definidas
    #frame = Frame(120, 720, ancho - 100, alto - 100, id='frameTextoConstante')  # Ajusta las dimensiones según sea necesario
    elementos = [titulo1, Spacer(1,25), parrafo, Spacer(1,25),titulo2,Spacer(1,25), parrafo2, Spacer(1,25),titulo3,Spacer(1,25), parrafo3]
    frame = Frame(30, 0, width-50, height-100, id='frameTextoConstante')
    frame.addFromList(elementos, c)

    c.save()
    buf.seek(0)
    return buf


def generar_pdf_nueva(compra):
    #Configuration of the PDF object
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    #doc = SimpleDocTemplate(buf, pagesize=letter)
    #Here ends conf.
    #compra = Compra.objects.get(id=pk)
    productos = ArticuloComprado.objects.filter(oc=compra.id)

    # Define estilos para las tablas (antes de usarlos)
    styles = getSampleStyleSheet()
    style_desc = styles["BodyText"]
    style_desc.wordWrap = 'CJK'
    style_desc.fontSize = 6
    style_desc.leading = 8

    # Azul Vordcab
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    # Encabezado superior
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)

    # Bloque 1: Logotipo y Nombre de la Empresa (Izquierda)
    c.drawInlineImage('static/images/logo_vordcab.jpg',45,730, 3 * cm, 1.5 * cm)  # Imagen vordcab

    # Bloque 2: Título y Datos de Control (Centro Superior)
    caja_iso = 760
    # Dibujar el título principal
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(150,750,250,20, fill=True, stroke=False) #Barra azul superior Orden de Compra
    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(280,755,'Orden de compra')
    c.setFillColor(black)

    # Datos de control
    c.setFont('Helvetica',8)
    c.drawString(150,caja_iso-20,'Número de documento')
    c.drawString(150,caja_iso-30,'SEOV-ADQ-N4-01.02')
    c.drawString(245,caja_iso-20,'Clasificación del documento')
    c.drawString(275,caja_iso-30,'Controlado')
    c.drawString(355,caja_iso-20,'Nivel del documento')
    c.drawString(380,caja_iso-30, 'N5')
    c.drawString(440,caja_iso-20,'Revisión No.')
    c.drawString(452,caja_iso-30,'003')
    c.drawString(510,caja_iso-20,'Fecha de Emisión')
    c.drawString(525,caja_iso-30,'13/11/2017')

    # Bloque 3: Datos de Elaboración y Aprobación (Derecha Superior)
    c.drawString(430,caja_iso,'Preparado por:')
    c.drawString(405,caja_iso-10,'SUPT. DE ADQUISICIONES')
    c.drawString(520,caja_iso,'Aprobación')
    c.drawString(515,caja_iso-10,'SUBD ADTVO')

    # --- SECCIÓN PRINCIPAL DEL FORMULARIO ---
    # Definir las posiciones para las columnas principales con mejor distribución
    col_inicio = 30
    col_total_ancho = 530  # Aproximadamente el ancho disponible

    # Dividir en 2 columnas principales (cada panel de la fila 1 y 2)
    col_ancho = col_total_ancho // 2 - 5  # Aproximadamente 260 cada una con separación
    col_separacion = 10

    # Coordenadas Y para las secciones principales
    seccion_inicio = 700
    seccion_alto = 100  # Ajustable según contenido

    # Fila 1: ORDEN DE COMPRA Y DATOS DEL PROVEEDOR
    # Panel Izquierdo: Orden de compra (Etiqueta centrada en fondo azul oscuro)
    panel_inicio_y = seccion_inicio - 25
    c.setFillColor(prussian_blue)
    c.rect(col_inicio, panel_inicio_y, col_ancho, 20, fill=True, stroke=False)  # Fondo azul
    c.setFillColor(white)
    c.setFont('Helvetica-Bold', 9)  # Tamaño de fuente más pequeño para evitar superposición
    c.drawCentredString(col_inicio + col_ancho//2, panel_inicio_y + 12, 'Orden de compra')

    # Contenido del panel izquierdo
    c.setFillColor(black)
    c.setFont('Helvetica', 7)  # Tamaño de fuente más pequeño para más campos
    campo_y = panel_inicio_y - 8  # Espaciado reducido
    c.drawString(col_inicio + 3, campo_y, 'Folio:')
    c.drawString(col_inicio + 80,campo_y, str(compra.folio))

    campo_y -= 12
    c.drawString(col_inicio + 3, campo_y, 'Fecha de emisión:')
    c.drawString(col_inicio + 80, campo_y, compra.created_at.strftime("%d/%m/%Y"))

    campo_y -= 12
    c.drawString(col_inicio + 3, campo_y, 'Proceso:')
    # c.line(col_inicio + 45, campo_y - 1, col_inicio + col_ancho - 5, campo_y - 1)  # Línea de llenado eliminada

    campo_y -= 12
    c.drawString(col_inicio + 3, campo_y, 'Elaboró:')
    c.drawString(col_inicio + 80, campo_y,compra.creada_por.staff.staff.first_name + ' ' +compra.creada_por.staff.staff.last_name)

    campo_y -= 12
    c.drawString(col_inicio + 3, campo_y, 'Aut. Sptte.:')
    if compra.autorizado1:
        c.drawString(col_inicio + 80, campo_y,compra.oc_autorizada_por.staff.staff.first_name + ' ' +compra.oc_autorizada_por.staff.staff.last_name)
    

    campo_y -= 12
    c.drawString(col_inicio + 3, campo_y, 'Aut. Gerente:')
    if compra.autorizado2:
        c.drawString(col_inicio + 80, campo_y,compra.oc_autorizada_por2.staff.staff.first_name + ' ' + compra.oc_autorizada_por2.staff.staff.last_name)
    # Panel Derecho: Datos del proveedor (Etiqueta centrada en fondo azul oscuro)
    panel_derecho_inicio = col_inicio + col_ancho + col_separacion
    c.setFillColor(prussian_blue)
    c.rect(panel_derecho_inicio, panel_inicio_y, col_ancho, 20, fill=True, stroke=False)  # Fondo azul
    c.setFillColor(white)
    c.setFont('Helvetica-Bold', 9)  # Tamaño de fuente más pequeño
    c.drawCentredString(panel_derecho_inicio + col_ancho//2, panel_inicio_y + 12, 'Datos del proveedor')

    # Contenido del panel derecho
    c.setFillColor(black)
    c.setFont('Helvetica', 7)  # Tamaño de fuente más pequeño
    campo_y = panel_inicio_y - 8
    c.drawString(panel_derecho_inicio + 3, campo_y, 'Nombre:')
    if compra.proveedor.nombre.razon_social == 'COLABORADOR':
        c.drawString(panel_derecho_inicio + 80,campo_y, compra.deposito_comprador.staff.staff.first_name+' '+compra.deposito_comprador.staff.staff.last_name)
        c.drawString(panel_derecho_inicio + 80,campo_y- 24, compra.deposito_comprador.banco.nombre)
        c.drawString(panel_derecho_inicio + 80,campo_y- 36, compra.deposito_comprador.cuenta_bancaria)
        c.drawString(panel_derecho_inicio + 80,campo_y- 48, compra.deposito_comprador.clabe)
    else:
        c.drawString(panel_derecho_inicio + 80,campo_y, compra.proveedor.nombre.razon_social)
        c.drawString(panel_derecho_inicio + 80,campo_y- 24, compra.proveedor.banco.nombre)
        c.drawString(panel_derecho_inicio + 80,campo_y- 36, compra.proveedor.cuenta)
        c.drawString(panel_derecho_inicio + 80,campo_y- 48, compra.proveedor.clabe)

    campo_y -= 12
    c.drawString(panel_derecho_inicio + 3, campo_y, 'RFC:')
    c.drawString(panel_derecho_inicio + 80, campo_y, compra.proveedor.nombre.rfc)
        
    campo_y -= 12
    c.drawString(panel_derecho_inicio + 3, campo_y, 'Banco:')
    # c.line(panel_derecho_inicio + 30, campo_y - 1, panel_derecho_inicio + col_ancho - 5, campo_y - 1)  # Línea de llenado eliminada

    campo_y -= 12
    c.drawString(panel_derecho_inicio + 3, campo_y, 'Cuenta:')
    # c.line(panel_derecho_inicio + 40, campo_y - 1, panel_derecho_inicio + col_ancho - 5, campo_y - 1)  # Línea de llenado eliminada

    campo_y -= 12
    c.drawString(panel_derecho_inicio + 3, campo_y, 'CLABE:')
    # c.line(panel_derecho_inicio + 35, campo_y - 1, panel_derecho_inicio + col_ancho - 5, campo_y - 1)  # Línea de llenado eliminada

    campo_y -= 12
    c.drawString(panel_derecho_inicio + 3, campo_y, 'Contacto:') # <<< este campo no se renderiza
    # c.line(panel_derecho_inicio + 45, campo_y - 1, panel_derecho_inicio + col_ancho - 5, campo_y - 1)  # Línea de llenado eliminada

    # Fila 2: TRAZABILIDAD Y TÉRMINOS Y CONDICIONES
    # Reducir el espacio entre filas para acercar las secciones
    espacio_entre_filas = 80  # Reducido de 150 a 80 para acercar las secciones
    panel_inicio_y = panel_inicio_y - espacio_entre_filas  # Mayor separación

    # Panel Izquierdo: Trazabilidad/ Datos de requisición (Etiqueta centrada en fondo azul oscuro)
    c.setFillColor(prussian_blue)
    c.rect(col_inicio, panel_inicio_y, col_ancho, 20, fill=True, stroke=False)  # Fondo azul
    c.setFillColor(white)
    c.setFont('Helvetica-Bold', 9)  # Tamaño de fuente más pequeño
    c.drawCentredString(col_inicio + col_ancho//2, panel_inicio_y + 12, 'Trazabilidad/ Datos de requisición')

    # Contenido del panel izquierdo
    c.setFillColor(black)
    c.setFont('Helvetica', 7)  # Tamaño de fuente más pequeño
    campo_y = panel_inicio_y - 8
    c.drawString(col_inicio + 3, campo_y, 'Proyecto:')
    c.drawString(col_inicio + 80, campo_y,compra.req.orden.proyecto.nombre)

    campo_y -= 12
    c.drawString(col_inicio + 3, campo_y, 'Subproyecto:')
    c.drawString(col_inicio + 80, campo_y,compra.req.orden.subproyecto.nombre)

    campo_y -= 12
    c.drawString(col_inicio + 3, campo_y, 'Folio requisición:')
    c.drawString(col_inicio + 80, campo_y, str(compra.req.folio))

    campo_y -= 12
    c.drawString(col_inicio + 3, campo_y, 'Fecha autorización:')
    c.drawString(col_inicio + 80, campo_y,compra.req.orden.approved_at.strftime("%d/%m/%Y"))
    campo_y -= 12
    c.drawString(col_inicio + 3, campo_y, 'Solicitado por:')
    c.drawString(col_inicio + 80, campo_y, compra.req.orden.staff.staff.staff.first_name +' '+ compra.req.orden.staff.staff.staff.last_name)

    campo_y -= 12
    c.drawString(col_inicio + 3, campo_y, 'Activo Fijo:')
    try:
        if compra.req.orden.activo is not None:
            eco_unidad = compra.req.orden.activo.eco_unidad
            descripcion = compra.req.orden.activo.descripcion
            serie = compra.req.orden.activo.serie
            c.drawString(col_inicio + 80, campo_y, f'{eco_unidad} {descripcion}')
            c.drawString(col_inicio + 80, campo_y - 12, serie)
        else:
            c.drawString(col_inicio + 80, campo_y, 'NA')
            c.drawString(col_inicio + 80, campo_y - 12 , 'NA')
    except Activo.DoesNotExist:
        c.drawString(col_inicio + 80, campo_y, 'NA')
        c.drawString(col_inicio + 80, campo_y - 12, 'NA')

    campo_y -= 12
    c.drawString(col_inicio + 3, campo_y, 'N° de serie:')
    # c.line(col_inicio + 55, campo_y - 1, col_inicio + col_ancho - 5, campo_y - 1)  # Línea de llenado eliminada

    campo_y -= 12
    c.drawString(col_inicio + 3, campo_y, 'Comentarios:')
    #Comentario de solicitud
    if compra.comentario_solicitud:
        paragraph_content = compra.req.orden.comentario
    else:
        paragraph_content = "NA" 

    styles = getSampleStyleSheet()
    base = styles["Normal"]
    base.fontSize = 7
    #styleN.color = prussian_blue
    base.leading = 6  # Espaciado entre líneas
    styleN = ParagraphStyle('Justicado', parent=base, alignment=TA_JUSTIFY)

    if paragraph_content is None:
        paragraph_content = " "    
    conditional_paragraph = Paragraph(paragraph_content, styleN)

    # Crear un nuevo frame similar al anterior pero ajustando la posición y/o tamaño si es necesario
    # Asumiendo 'width' y 'height' ya están definidos como antes
    new_frame = Frame(col_inicio + 74, campo_y - 185 , 200, 200, id='conditional_frame')
    
    # Agregar el párrafo al nuevo marco
    new_frame.addFromList([conditional_paragraph], c)

    # Panel Derecho: Términos y condiciones (Etiqueta centrada en fondo azul oscuro)
    c.setFillColor(prussian_blue)
    c.rect(panel_derecho_inicio, panel_inicio_y, col_ancho, 20, fill=True, stroke=False)  # Fondo azul
    c.setFillColor(white)
    c.setFont('Helvetica-Bold', 9)  # Tamaño de fuente más pequeño
    c.drawCentredString(panel_derecho_inicio + col_ancho//2, panel_inicio_y + 12, 'Términos y condiciones')

    # Contenido del panel derecho
    c.setFillColor(black)
    c.setFont('Helvetica', 7)  # Tamaño de fuente más pequeño
    campo_y = panel_inicio_y - 8
    c.drawString(panel_derecho_inicio + 3, campo_y, 'Condición pago:')
    c.drawString(panel_derecho_inicio + 80, campo_y, compra.cond_de_pago.nombre)
    

    campo_y -= 12
    c.drawString(panel_derecho_inicio + 3, campo_y, 'Plazo Crédito:')
    if compra.dias_de_credito:
        c.drawString(panel_derecho_inicio + 80, campo_y, str(compra.dias_de_credito) + ' días')
    else:
        c.drawString(panel_derecho_inicio + 80, campo_y, 'No Especificado')

    campo_y -= 12
    c.drawString(panel_derecho_inicio + 3, campo_y, 'Moneda:')
    c.drawString(panel_derecho_inicio + 80, campo_y,compra.moneda.nombre)

    campo_y -= 12
    c.drawString(panel_derecho_inicio + 3, campo_y, 'Uso CFDI:')
    c.drawString(panel_derecho_inicio + 80, campo_y, compra.uso_del_cfdi.descripcion)

    campo_y -= 12
    c.drawString(panel_derecho_inicio + 3, campo_y, 'Enviar factura:')
    c.drawString(panel_derecho_inicio + 80,campo_y, compra.creada_por.staff.staff.email)

    campo_y -= 12
    c.drawString(panel_derecho_inicio + 3, campo_y, 'Tiempo entrega:')
    if compra.dias_de_entrega:
        c.drawString(panel_derecho_inicio + 80, campo_y, str(compra.dias_de_entrega)+' '+'días hábiles')

    campo_y -= 12
    c.drawString(panel_derecho_inicio + 3, campo_y, 'Lugar entrega:')
    # c.line(panel_derecho_inicio + 70, campo_y - 1, panel_derecho_inicio + col_ancho - 5, campo_y - 1)  # Línea de llenado eliminada

    campo_y -= 12
    c.drawString(panel_derecho_inicio + 3, campo_y, 'Comentarios:')
    if compra.comentarios:
        comentarios = compra.comentarios
    else:
        comentarios = "NA"

    if comentarios is None:
        comentarios = " "    
    conditional_paragraph = Paragraph(comentarios, styleN)

    # Crear un nuevo frame similar al anterior pero ajustando la posición y/o tamaño si es necesario
    # Asumiendo 'width' y 'height' ya están definidos como antes
    new_frame = Frame(col_inicio + 340, campo_y - 185 , 200, 200, id='conditional_frame')
    
    # Agregar el párrafo al nuevo marco
    new_frame.addFromList([conditional_paragraph], c)

    # --- SECCIÓN NOTA IMPORTANTE ---
    # Se calcula la posición Y para la nota, debajo de los paneles de contenido.
    nota_y_inicio = panel_inicio_y - 100 # Ajustado para bajar la nota
    
    # Define styles for the note
    style_nota_titulo = ParagraphStyle('nota_titulo', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, leading=10)
    style_nota_cuerpo = ParagraphStyle('nota_cuerpo', parent=styles['Normal'], fontSize=7, leading=9, alignment=TA_JUSTIFY)

    # Create Paragraphs for the note
    titulo_texto = "<strong>IMPORTANTE: Requisito de Acceso para Personal Externo</strong>"
    cuerpo_texto = """
    Todo personal externo (proveedores, contratistas y subcontratistas) solo podrá ingresar a realizar
    trabajos o servicios en las instalaciones de GRUPO VORDCAB presentando su constancia de vigencia
    de derechos del IMSS con una antigüedad no mayor a 30 días naturales a partir de su fecha de emisión.
    """
    titulo_p = Paragraph(titulo_texto, style_nota_titulo)
    cuerpo_p = Paragraph(cuerpo_texto, style_nota_cuerpo)

    # Create and draw a Frame to hold the note content with a border
    elementos_nota = [titulo_p, Spacer(1, 4), cuerpo_p]
    nota_frame = Frame(col_inicio, nota_y_inicio - 55, col_total_ancho, 55, showBoundary=1, leftPadding=5, rightPadding=5, topPadding=5, bottomPadding=5)
    nota_frame.addFromList(elementos_nota, c)


    # NUEVA SECCIÓN: TABLA DE PARTIDAS - Colocada debajo de la nota
    # La posición de la tabla ahora se calcula en relación a la nueva posición de la nota.
    tabla_inicio_y = nota_y_inicio - 65

    # Sección de la tabla de productos
    data = []
    data.append(['''Código''', '''Producto / Servicio''', '''Cantidad''', '''Unidad''', '''P. U.''', '''Importe'''])

    for producto in productos:
        importe = producto.precio_unitario * producto.cantidad
        importe_rounded = round(importe, 4)
        descripcion = Paragraph(producto.producto.producto.articulos.producto.producto.nombre, style_desc)
        precio_unitario = f"{producto.precio_unitario:,.4f}"
        data.append([
            producto.producto.producto.articulos.producto.producto.codigo,
            descripcion,
            producto.cantidad,
            producto.producto.producto.articulos.producto.producto.unidad,
            precio_unitario,
            f"{importe_rounded:,.4f}"
        ])

    c.setFillColor(black)
    c.setFont('Helvetica', 8)

    # Dibujar la tabla de productos en la nueva ubicación
    table = Table(data, colWidths=[1.2 * cm, 12.5 * cm, 1.5 * cm, 1.2 * cm, 1.5 * cm, 1.5 * cm])
    table_style = TableStyle([  # estilos de la tabla
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.white),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        # ENCABEZADO
        ('TEXTCOLOR', (0, 0), (-1, 0), white),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 0), (-1, 0), prussian_blue),
        # CUERPO
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
    ])
    table.setStyle(table_style)

    # Calcular la altura de la tabla para posicionarla adecuadamente
    table.wrapOn(c, c._pagesize[0], c._pagesize[1])

    # Posicionar la tabla más abajo para evitar superposición
    table_y = tabla_inicio_y - (len(data) * 18)  # Ajustar la posición basado en el número de filas

    # Asegurarse de que la tabla no se posicione demasiado abajo
    if table_y < 200:  # Si la tabla se va muy abajo, ajustar
        table_y = 400 - (len(data) * 18)

    table.drawOn(c, 20, table_y)

    # --- PIE DE PÁGINA FIJO ---
    footer_y = 120  # Posición Y fija para el inicio del pie de página
    c.setLineWidth(1)
    c.line(20, footer_y, 585, footer_y)  # Línea divisoria del pie de página

    # Sección Izquierda: Total con letra
    c.setFillColor(black)
    c.setFont('Helvetica-Bold', 9)
    c.drawString(30, footer_y - 20, 'Total con letra:')
    
    c.setFont('Helvetica', 8)
    
    def convertir_a_reales(valor):
        partes = f"{valor:.2f}".split(".")
        reales = num2words(int(partes[0]), lang='pt_BR')
        centavos = num2words(int(partes[1]), lang='pt_BR')
        return f"{reales} reais e {centavos} centavos"

    total_en_letra = ""
    if compra.moneda.nombre == "PESOS":
        total_en_letra = num2words(compra.costo_plus_adicionales, lang='es', to='currency', currency='MXN')
    elif compra.moneda.nombre == "DOLARES":
        total_en_letra = num2words(compra.costo_plus_adicionales, lang='es', to='currency', currency='USD')
    elif compra.moneda.nombre == "REAIS":
        total_en_letra = convertir_a_reales(compra.costo_plus_adicionales)
    
    # Usar Paragraph para el texto con ajuste de línea
    style_letra = ParagraphStyle('total_letra', parent=styles['Normal'], fontSize=8, leading=10)
    p_letra = Paragraph(total_en_letra.upper(), style_letra)
    
    # Dibujar el párrafo en un Frame para controlar el ancho
    frame_letra = Frame(30, footer_y - 70, 350, 50, id='frame_letra', showBoundary=0)
    frame_letra.addFromList([p_letra], c)


    # Sección Derecha: Totales numéricos
    montos_align_x = 570
    
    # Sub Total
    c.setFont('Helvetica-Bold', 9)
    c.drawRightString(montos_align_x - 80, footer_y - 20, 'Sub Total:')
    c.setFont('Helvetica', 9)
    subtotal = compra.costo_oc - compra.costo_iva
    c.drawRightString(montos_align_x, footer_y - 20, f"${subtotal:,.2f}")

    # IVA
    c.setFont('Helvetica-Bold', 9)
    c.drawRightString(montos_align_x - 80, footer_y - 40, 'IVA:')
    c.setFont('Helvetica', 9)
    c.drawRightString(montos_align_x, footer_y - 40, f"${compra.costo_iva:,.2f}")

    # Total (resaltado en azul)
    c.setFillColor(prussian_blue)
    c.setFont('Helvetica-Bold', 11)
    c.drawRightString(montos_align_x - 80, footer_y - 60, 'Total:')
    c.drawRightString(montos_align_x, footer_y - 60, f"${compra.costo_plus_adicionales:,.2f}")

    c.setFillColor(black)
    width, height = letter

    # Use the same style from earlier
    styleN = getSampleStyleSheet()["BodyText"]
    styleN.fontSize = 6



    # The duplicate table code has been removed to prevent overlap issues

    # --- NUEVA HOJA: Requerimientos de Calidad por producto ---
    rows_cal = []   # solo filas con requerimientos

    # (opcional) deduplicar por producto si la OC puede traer el mismo producto varias veces
    vistos = set()

    for ac in productos:  # Using mock data instead of Django query
        try:
            base_product = ac.producto.producto.articulos.producto.producto
        except Exception:
            continue
        if not base_product:
            continue

        # (opcional) evita duplicados por Product
        if getattr(base_product, 'id', None) in vistos:
            continue
        vistos.add(base_product.id)

        # intenta tomar el OneToOne
        try:
            producto_calidad = base_product.producto_calidad
        except:  # Using generic exception for mock data instead of Producto_Calidad.DoesNotExist
            continue

        if not producto_calidad:
            continue

        reqs = producto_calidad.requerimientos_calidad.select_related('requerimiento').all()
        if not reqs:
            continue

        codigo = base_product.codigo
        desc_par = Paragraph(base_product.nombre or '', style_desc)

        for req in reqs:
            req_nombre = req.requerimiento.nombre if req.requerimiento else ''
            comentario = req.comentarios or ''
            rows_cal.append([codigo, desc_par, req_nombre, comentario])

    # Si no hay nada que mostrar, NO crear la hoja
    if rows_cal:
        c.showPage()
        # encabezado hoja
        c.setFont('Helvetica', 12)
        c.setFillColor(prussian_blue)
        c.rect(20, 750, 565, 24, fill=True, stroke=False)
        c.setFillColor(white)
        c.setFont('Helvetica-Bold', 13)
        c.drawCentredString(300, 756, 'Requerimientos de Calidad por Producto')

        data_cal = [['Código', 'Producto', 'Requerimiento Calidad', 'Comentario']]
        data_cal.extend(rows_cal)

        col_widths = [2*cm, 8.0*cm, 5*cm, 5.0*cm]
        tabla_cal = Table(data_cal, colWidths=col_widths, hAlign='LEFT')
        tabla_cal.setStyle(TableStyle([
            ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
            ('BOX',       (0,0), (-1,-1), 0.50, colors.black),
            ('VALIGN',    (0,0), (-1,-1), 'MIDDLE'),
            ('BACKGROUND',(0,0), (-1,0),  prussian_blue),
            ('TEXTCOLOR', (0,0), (-1,0),  colors.white),
            ('FONTSIZE',  (0,0), (-1,0),  9),
            ('FONTSIZE',  (0,1), (-1,-1), 8),
        ]))

        tabla_cal.wrapOn(c, c._pagesize[0], c._pagesize[1])
        tabla_cal.drawOn(c, 20, 680 - min(520, len(data_cal)*16))

    c.showPage()
    # ------------ FIN NUEVA HOJA ------------------------------------------------

    # Agregar el encabezado en la segunda página
    c.setFont('Helvetica', 8)
    c.drawString(430, caja_iso, 'Preparado por:')
    c.drawString(420, caja_iso - 10, 'ASIST. TEC. SUBAD')
    c.drawString(525, caja_iso, 'Aprobación')
    c.drawString(520, caja_iso - 10, 'SUBD-ADTVO')
    c.drawString(50, caja_iso - 35, 'Número de documento')
    c.drawString(50, caja_iso - 45, 'SEOV-ADQ-N4-01.04')
    c.drawString(145, caja_iso - 35, 'Clasificación del documento')
    c.drawString(175, caja_iso - 45, 'No Controlado')
    c.drawString(255, caja_iso - 35, 'Nivel del documento')
    c.drawString(280, caja_iso - 45, 'N5')
    c.drawString(340, caja_iso - 35, 'Revisión No.')
    c.drawString(352, caja_iso - 45, '001')
    c.drawString(410, caja_iso - 35, 'Fecha de Emisión')
    c.drawString(425, caja_iso - 45, '03/05/23')
    c.drawString(500, caja_iso - 35, 'Fecha de Revisión')
    c.drawString(525, caja_iso - 45, '06/09/23')

    caja_proveedor = caja_iso - 65
    c.setFont('Helvetica', 12)
    c.setFillColor(prussian_blue)
    c.rect(155, 735, 250, 35, fill=True, stroke=False)  # Barra azul superior encabezado
    c.setFillColor(colors.white)
    c.setFont('Helvetica-Bold', 13)
    c.drawCentredString(280, 755, 'Requisitos en Materia de Gestión')
    c.drawCentredString(280, 740, 'de Seguridad, Salud y Medio Ambiente')
    # c.drawInlineImage('static/images/logo_vordcab.jpg', 60, 735, 2 * cm, 1 * cm)  # Imagen vortec  # This might fail if image doesn't exist



    #c.setFont("Helvetica-Bold", 12)
    #c.setFillColor(prussian_blue)  # Asumiendo que prussian_blue está definido
    #c.drawCentredString(300, 680, )  # Ajusta la posición según sea necesario
    styles = getSampleStyleSheet()
    styleN = styles["Normal"]
    styleN.fontSize = 10
    #styleN.color = prussian_blue
    styleN.leading = 13  # Espaciado entre líneas
    styleN = ParagraphStyle('Justicado', parent=styles['Normal'], alignment=TA_JUSTIFY)
    styleT = styles["Normal"]
    styleT.fontSize = 13
    styleT.fontName = 'Helvetica-Bold'
    styleT.textColor = prussian_blue
    styleT.leading = 17
    styleT = ParagraphStyle('Center', parent=styles['Normal'], alignment= TA_CENTER)

    texto = """El objeto del presente escrito es poner en conocimiento, de los proveedores que realizan trabajos para nuestra empresa,
                los requerimientos mínimos que le sean de aplicación, que debe de cumplir en lo referente a la seguridad y salud en el
                trabajo, calidad y protección del medio ambiente. La aceptación de una orden de compra implica la responsabilidad por
                parte del proveedor del conocimiento y aceptación de lo descrito.<br/>Nuestra empresa tiene implantado un SEOV
                (Sistema de Excelencia Operativa Vordcab), bajo las normas ISO 9001:2015, ISO 14001:2015 e ISO 45001:2018,
                por lo que se requiere que nuestros proveedores colaboren en el cumplimiento de los siguientes requisitos: <br/>
                """

    texto2 = """En todo momento el proveedor está obligado a cumplir con la legislación aplicable al servicio que está prestando.<br/>
El proveedor se compromete a garantizar el cumplimiento de lo solicitado en el pedido u orden de compra del material o del trabajo
externo.<br/>
El proveedor debería poder garantizar la correcta gestión y control de: los residuos, emisiones atmosféricas, ruidos,
efluentes residuales, productos peligrosos, afectación del suelo y mantenimiento de instalaciones, respecto al servicio prestado.<br/>
El proveedor deberá aplicar las medidas preventivas necesarias, para evitar situaciones de peligro o emergencia como derrame, fuga,
incendio, etc., durante la realización del trabajo encomendado. Y si es preciso formar e informar a su personal sobre manipulación,
almacenaje, uso y riesgos de productos o preparados peligrosos.<br/>
El proveedor deberá facilitar la documentación ambiental y de seguridad que se le solicite o requiera.<br/>
Si el proveedor tiene algún servicio que presta a nuestra empresa subcontratado, por ejemplo, transporte, deberá transmitir este
documento a su subcontratista.<br/> """
    texto3 = """El proveedor deberá trabajar adoptando buenas prácticas ambientales y cumplir con los procedimientos internos que se
    le hayan comunicado respecto a HSE y la Prevención de Riesgos Laborales.<br/>
Las empresas que realicen trabajos para nuestra empresa deberán cumplirse los requisitos de seguridad establecidos, uso de EPP,
protecciones colectivas, formación, seguridad equipos,
medidas preventivas<br/>
Si durante el trabajo se deben retirar residuos peligrosos (aceites, aguas con productos peligrosos, pinturas, disolventes, sus envases)
o /y no peligrosos (escombros, embalajes), se recogerán en recipientes adecuados según la cantidad a retirar. Los recipientes llenos
serán retirados por la empresa que preste el servicio o entregados a nuestra empresa, según convenga, para su correcto almacenamiento
y gestión posterior.<br/>
Si no sabe qué hacer con algún residuo o efluente residual o cualquier otra cosa, que se ha originado mientras prestaba el servicio,
 avise a la persona que le haya atendido o al Responsable de seguridad. No tomará decisiones.<br/>
Si detecta cualquier situación de riesgo/emergencia (ambiental, personal, de seguridad), lo comunicará de inmediato a cualquier persona
de la empresa o a la que le haya atendido o al Responsable de seguridad. Nunca actué. Deberá seguir las normas de emergencia que se le
 han facilitado a la entrada.<br/>
Si mientras trabaja se produce una situación de emergencia (derrame o fuga de producto peligrosos), y le han facilitado la formación y
medios necesarios para actuar, proceda según la instrucción que se le ha facilitado, en caso contrario avise de inmediato a cualquier
persona de la empresa o a la que le ha atendido o al Responsable de seguridad.<br/>"""
    titulo1 = "REQUISITOS OBLIGATORIOS PARA PROVEEDORES DE GRUPO VORDCAB"
    titulo2 = """REQUISITOS GENERALES PARA TODOS LOS PROVEEDORES"""
    titulo3 = """REQUISITOS SI USTED PRESTA EL SERVICIO O PARTE DE SERVICIO DENTRO <br/>
                DE LAS INSTALACIONES DE GRUPO VORDCAB"""

    titulo1 = Paragraph(titulo1, styleT)
    parrafo = Paragraph(texto, styleN)
    titulo2 = Paragraph(titulo2, styleT)
    parrafo2 = Paragraph(texto2, styleN)
    titulo3 = Paragraph(titulo3, styleT)
    parrafo3 = Paragraph(texto3, styleN)

    ancho, alto = letter  # Asegúrate de tener estas dimensiones definidas
    #frame = Frame(120, 720, ancho - 100, alto - 100, id='frameTextoConstante')  # Ajusta las dimensiones según sea necesario
    elementos = [titulo1, Spacer(1,25), parrafo, Spacer(1,25),titulo2,Spacer(1,25), parrafo2, Spacer(1,25),titulo3,Spacer(1,25), parrafo3]
    frame = Frame(30, 0, width-50, height-100, id='frameTextoConstante')
    frame.addFromList(elementos, c)

    c.save()
    buf.seek(0)
    return buf

def generar_pdf_proveedor(request, pk):
    #Configuration of the PDF object
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    #doc = SimpleDocTemplate(buf, pagesize=letter)
    #Here ends conf.
    #compra = Compra.objects.get(id=pk)
    proveedor = Proveedor_direcciones.objects.get(id=pk)

    #Azul Vordcab
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)
    caja_iso = 760
    #Elaborar caja
    #c.line(caja_iso,500,caja_iso,720)



    c.drawString(430,caja_iso,'Preparado por:')
    c.drawString(405,caja_iso-10,'Auditoría de Proveedores')
    c.drawString(520,caja_iso,'Aprobación')
    c.drawString(515,caja_iso-10,'SUBD QHSE')
    c.drawString(150,caja_iso-20,'Número de documento')
    c.drawString(150,caja_iso-30,'SEOV-CTE-N5-05.01')
    c.drawString(245,caja_iso-20,'Clasificación del documento')
    c.drawString(275,caja_iso-30,'Controlado')
    c.drawString(355,caja_iso-20,'Nivel del documento')
    c.drawString(380,caja_iso-30, 'N5')
    c.drawString(440,caja_iso-20,'Revisión No.')
    c.drawString(452,caja_iso-30,'000')
    c.drawString(510,caja_iso-20,'Fecha de Emisión')
    c.drawString(525,caja_iso-30,'09/01/2024')


    c.drawString(500,caja_iso - 45,'Fecha:')
    c.drawString(5400,caja_iso - 45, str(proveedor.modificado_fecha))
    c.setFont('Helvetica-Bold',12)
    c.drawString(500,caja_iso-60,'FOLIO:')

    c.setFillColor(rojo)
    c.setFont('Helvetica-Bold',12)
    c.drawString(540,caja_iso-60, str(proveedor.id))

    #Primera Tabla
    caja_proveedor = caja_iso - 85
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(140,750,260,20, fill=True, stroke=False) #Barra azul superior Título
    c.rect(20,caja_proveedor - 8,565,20, fill=True, stroke=False) #Barra azul superior | Datos Bancarios
    c.rect(20,caja_proveedor - 100,565,5, fill=True, stroke=False) #Linea posterior horizontal
    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(270,755,'Carta de Contactos y Datos Bancarios')
    c.setLineWidth(.3) #Grosor
    #c.line(20,caja_proveedor-8,20,520) #Eje Y donde empieza, Eje X donde empieza, donde termina eje y,donde termina eje x (LINEA 1 contorno)
    #c.line(585,caja_proveedor-8,585,520) #Linea 2 contorno
    c.drawInlineImage('static/images/logo_vordcab.jpg',45,730, 3 * cm, 1.5 * cm) #Imagen vortec

    c.setFillColor(white)
    c.setFont('Helvetica-Bold',11)
    c.drawString(200,caja_proveedor,'DATOS BANCARIOS MONEDA NACIONAL')
    inicio_central = 300
    #c.line(inicio_central,caja_proveedor-25,inicio_central,520) #Linea Central de caja Proveedor | Detalle
    c.setFillColor(black)
    c.setFont('Helvetica-Bold',9)
    #Primera columna
    c.drawString(30,caja_proveedor-20,'Nombre:')
    c.drawString(30,caja_proveedor-40,'Banco:')
    c.drawString(30,caja_proveedor-60,'Clabe:')
    c.drawString(30,caja_proveedor-80,'Convenio:')
    #Segunda Columna
    c.drawString(300,caja_proveedor-40,'Cuenta:')
    c.drawString(300,caja_proveedor-60,'Referencia:')
   


    c.setFillColor(black)
    c.setFont('Helvetica',9)
   
    #Primera columna
    c.drawString(100,caja_proveedor-20, proveedor.nombre.razon_social)
    c.drawString(100,caja_proveedor-40, proveedor.banco.nombre)
    c.drawString(100,caja_proveedor-60, proveedor.clabe)
    if proveedor.contratocie is None:
        c.drawString(100,caja_proveedor-80, '')
    else:
        c.drawString(100,caja_proveedor-80, proveedor.contratocie)
    #Segunda columna
    c.drawString(370,caja_proveedor-40, proveedor.cuenta)
    c.drawString(370,caja_proveedor-60, 'NR')

    
    #Segunda tabla
    segunda_tabla = caja_proveedor - 150
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(20,  segunda_tabla - 8,565,20, fill=True, stroke=False) #Barra azul superior | Datos Bancarios Dólares
    c.rect(20, segunda_tabla - 80,565,5, fill=True, stroke=False) #Linea posterior horizontal
   

    c.setFillColor(white)
    c.setFont('Helvetica-Bold',11)
    c.drawString(220, segunda_tabla,'DATOS BANCARIOS DÓLARES')
    c.setFillColor(black)
    c.setFont('Helvetica-Bold',9)
    #Primera columna
    c.drawString(30,segunda_tabla-20,'Nombre:')
    c.drawString(30,segunda_tabla-40,'Banco:')
    c.drawString(30,segunda_tabla-60,'Clabe Spid:')
    #Segunda Columna
    c.drawString(300,segunda_tabla-40,'Cuenta:')
    c.drawString(300,segunda_tabla-60,'Codigo Swift:')
   
    c.setFillColor(black)
    c.setFont('Helvetica',9)
   
    #Primera columna
    c.drawString(100,segunda_tabla-20, proveedor.nombre.razon_social)
    c.drawString(100,segunda_tabla-40, proveedor.banco.nombre)
    if proveedor.spid is None:
        c.drawString(100,segunda_tabla-60, '')
    else:
        c.drawString(100,segunda_tabla-60, proveedor.spid)
    #Segunda columna
    c.drawString(370,segunda_tabla-40, proveedor.cuenta)
    if proveedor.swift:
        c.drawString(370,segunda_tabla-60, proveedor.swift)
    else:
        c.drawString(370,segunda_tabla-60, 'NA')

    #Tercera tabla
    tercera_tabla = segunda_tabla - 140
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(20,   tercera_tabla - 8,565,20, fill=True, stroke=False) #Barra azul superior | Datos Bancarios Dólares
    c.rect(20,  tercera_tabla - 35,565,5, fill=True, stroke=False) #Linea posterior horizontal
   

    c.setFillColor(white)
    c.setFont('Helvetica-Bold',11)
    c.drawString(230,  tercera_tabla,'CONDICIONES DE COMPRA')
    c.setFillColor(black)
    c.setFont('Helvetica-Bold',9)
    #Única Fila
    c.drawString(30,tercera_tabla-20,'Crédito:')
    c.drawString(150,tercera_tabla-20,'Contado:')
    c.drawString(300,tercera_tabla-20,'Días de Crédito:')
   
    c.setFillColor(black)
    c.setFont('Helvetica',9)
   
    #Primera columna
    if proveedor.financiamiento:
        c.setFont('Helvetica-Bold', 12)
        c.drawString(70,tercera_tabla-20, "X")
    else:
        c.setFont('Helvetica-Bold', 12)
        c.drawString(200,tercera_tabla-20, "X")
    c.drawString(380,tercera_tabla-20, str(proveedor.dias_credito))
   

    #Cuarta tabla
    cuarta_tabla = tercera_tabla - 80
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(20, cuarta_tabla - 8,565,20, fill=True, stroke=False) #Barra azul superior | Datos Bancarios Dólares
    c.rect(20, cuarta_tabla - 120,565,5, fill=True, stroke=False) #Linea posterior horizontal
   

    c.setFillColor(white)
    c.setFont('Helvetica-Bold',11)
    c.drawString(220, cuarta_tabla,'DATOS DE PROVEEDOR Y CONTACTO')
    c.setFillColor(black)
    c.setFont('Helvetica-Bold',9)
    #Primera columna
    c.drawString(30,cuarta_tabla-20,'Razón Social:')
    c.drawString(30,cuarta_tabla-40,'Dirección:')
    c.drawString(30,cuarta_tabla-60,'Estado:')
    c.drawString(30,cuarta_tabla-80,'RFC:')
    c.drawString(30,cuarta_tabla-100,'Contacto:')
    #Segunda Columna
    c.drawString(300,cuarta_tabla-60,'Teléfono:')
    c.drawString(300,cuarta_tabla-80,'Email:')
    
   
    c.setFillColor(black)
    c.setFont('Helvetica',9)
   
    #Primera columna
    c.drawString(100,cuarta_tabla-20, proveedor.nombre.razon_social)
    c.drawString(100,cuarta_tabla-40, proveedor.domicilio)
    if proveedor.estado:
        c.drawString(100,cuarta_tabla-60, str(proveedor.estado.nombre))
    else:
        c.drawString(100,cuarta_tabla-60, 'No registro')
    c.drawString(100,cuarta_tabla-80, proveedor.nombre.rfc)
    c.drawString(100,cuarta_tabla-100, proveedor.contacto)
    #Segunda columna
    c.drawString(370,cuarta_tabla-60, proveedor.telefono)
    c.drawString(370,cuarta_tabla-80, proveedor.email)
   


    texto = """El objeto del presente escrito es poner en conocimiento, de los proveedores que realizan trabajos para nuestra empresa, 
                los requerimientos mínimos que le sean de aplicación, que debe de cumplir en lo referente a la seguridad y salud en el 
                trabajo, calidad y protección del medio ambiente. La aceptación de una orden de compra implica la responsabilidad por 
                parte del proveedor del conocimiento y aceptación de lo descrito.<br/>Nuestra empresa tiene implantado un SEOV 
                (Sistema de Excelencia Operativa Vordcab), bajo las normas ISO 9001:2015, ISO 14001:2015 e ISO 45001:2018, 
                por lo que se requiere que nuestros proveedores colaboren en el cumplimiento de los siguientes requisitos: <br/>
                """
    
    texto2 = """En todo momento el proveedor está obligado a cumplir con la legislación aplicable al servicio que está prestando.<br/>
El proveedor se compromete a garantizar el cumplimiento de lo solicitado en el pedido u orden de compra del material o del trabajo 
externo.<br/>
El proveedor debería poder garantizar la correcta gestión y control de: los residuos, emisiones atmosféricas, ruidos, 
efluentes residuales, productos peligrosos, afectación del suelo y mantenimiento de instalaciones, respecto al servicio prestado.<br/>
El proveedor deberá aplicar las medidas preventivas necesarias, para evitar situaciones de peligro o emergencia como derrame, fuga, 
incendio, etc., durante la realización del trabajo encomendado. Y si es preciso formar e informar a su personal sobre manipulación, 
almacenaje, uso y riesgos de productos o preparados peligrosos.<br/>
El proveedor deberá facilitar la documentación ambiental y de seguridad que se le solicite o requiera.<br/>
Si el proveedor tiene algún servicio que presta a nuestra empresa subcontratado, por ejemplo, transporte, deberá transmitir este 
documento a su subcontratista.<br/> """
    texto3 = """El proveedor deberá trabajar adoptando buenas prácticas ambientales y cumplir con los procedimientos internos que se 
    le hayan comunicado respecto a HSE y la Prevención de Riesgos Laborales.<br/>
Las empresas que realicen trabajos para nuestra empresa deberán cumplirse los requisitos de seguridad establecidos, uso de EPP, 
protecciones colectivas, formación, seguridad equipos,
medidas preventivas<br/>
Si durante el trabajo se deben retirar residuos peligrosos (aceites, aguas con productos peligrosos, pinturas, disolventes, sus envases)
o /y no peligrosos (escombros, embalajes), se recogerán en recipientes adecuados según la cantidad a retirar. Los recipientes llenos
serán retirados por la empresa que preste el servicio o entregados a nuestra empresa, según convenga, para su correcto almacenamiento 
y gestión posterior.<br/>
Si no sabe qué hacer con algún residuo o efluente residual o cualquier otra cosa, que se ha originado mientras prestaba el servicio,
 avise a la persona que le haya atendido o al Responsable de seguridad. No tomará decisiones.<br/>
Si detecta cualquier situación de riesgo/emergencia (ambiental, personal, de seguridad), lo comunicará de inmediato a cualquier persona
de la empresa o a la que le haya atendido o al Responsable de seguridad. Nunca actué. Deberá seguir las normas de emergencia que se le
 han facilitado a la entrada.<br/>
Si mientras trabaja se produce una situación de emergencia (derrame o fuga de producto peligrosos), y le han facilitado la formación y 
medios necesarios para actuar, proceda según la instrucción que se le ha facilitado, en caso contrario avise de inmediato a cualquier 
persona de la empresa o a la que le ha atendido o al Responsable de seguridad.<br/>"""
    titulo1 = "REQUISITOS OBLIGATORIOS PARA PROVEEDORES DE GRUPO VORDCAB"
    titulo2 = """REQUISITOS GENERALES PARA TODOS LOS PROVEEDORES"""
    titulo3 = """REQUISITOS SI USTED PRESTA EL SERVICIO O PARTE DE SERVICIO DENTRO <br/> 
                DE LAS INSTALACIONES DE GRUPO VORDCAB"""
    
    """titulo1 = Paragraph(titulo1, styleT)
    parrafo = Paragraph(texto, styleN)
    titulo2 = Paragraph(titulo2, styleT)
    parrafo2 = Paragraph(texto2, styleN)
    titulo3 = Paragraph(titulo3, styleT)
    parrafo3 = Paragraph(texto3, styleN)

    ancho, alto = letter  # Asegúrate de tener estas dimensiones definidas
    #frame = Frame(120, 720, ancho - 100, alto - 100, id='frameTextoConstante')  # Ajusta las dimensiones según sea necesario
    elementos = [titulo1, Spacer(1,25), parrafo, Spacer(1,25),titulo2,Spacer(1,25), parrafo2, Spacer(1,25),titulo3,Spacer(1,25), parrafo3]
    frame = Frame(30, 0, width-50, height-100, id='frameTextoConstante')
    frame.addFromList(elementos, c)"""

    c.save()
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename='Carta_Proveedor' + str(proveedor.id) +'.pdf')


def convert_excel_matriz_compras(compras, num_requis_atendidas, num_approved_requis, start_date, end_date):
    print('conteo compras:', compras.count())
    # Crea un objeto BytesIO para guardar el archivo Excel
    output = BytesIO()

    # Crea un libro de trabajo y añade una hoja
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet("Matriz_Compras")

     
    #date_format = workbook.add_format({'num_format': 'dd/mm/yyyy'})
    # Define los estilos
    head_style = workbook.add_format({'bold': True, 'font_color': 'FFFFFF', 'bg_color': '333366', 'font_name': 'Arial', 'font_size': 11})
    body_style = workbook.add_format({'font_name': 'Calibri', 'font_size': 10})
    money_style = workbook.add_format({'num_format': '$ #,##0.00', 'font_name': 'Calibri', 'font_size': 10})
    date_style = workbook.add_format({'num_format': 'dd/mm/yyyy', 'font_name': 'Calibri', 'font_size': 10})
    percent_style = workbook.add_format({'num_format': '0.00%', 'font_name': 'Calibri', 'font_size': 10})
    messages_style = workbook.add_format({'font_name':'Arial Narrow', 'font_size':11})

    columns = ['Compra', 'Requisición', 'Solicitud','Distrito', 'Proyecto', 'Subproyecto', 'Área', 'Solicitante','Comprador', 'Creado', 'Req. Autorizada', 'Proveedor',
               'Status Proveedor','Crédito/Contado', 'Costo', 'Monto Pagado', 'Status Pago','Fecha Pago', 'Status Autorización','Tipo Item', 'Días de entrega', 'Moneda',
               'Tipo de cambio', 'Entregada','Tiene Facturas', 'Activo Fijo', "Total en pesos"]

    columna_max = len(columns)+2

    worksheet.write(0, columna_max - 1, 'Reporte Creado Automáticamente por SAVIA Vordcab. UH', messages_style)
    worksheet.write(1, columna_max - 1, 'Software desarrollado por Grupo Vordcab S.A. de C.V.', messages_style)
    worksheet.set_column(columna_max - 1, columna_max, 30)  # Ajusta el ancho de las columnas nuevas
    
    # Escribir encabezados debajo de los mensajes
    worksheet.write(2, columna_max - 1, "Fecha Inicial", head_style)
    worksheet.write(3, columna_max - 1, "Fecha Final", head_style)
    worksheet.write(4, columna_max - 1, "Total de OC's", head_style)
    worksheet.write(5, columna_max - 1, "Requisiciones Aprobadas", head_style)
    worksheet.write(6, columna_max - 1, "Requisiciones Atendidas", head_style)
    worksheet.write(7, columna_max - 1, "KPI Colocadas/Aprobadas", head_style)
    worksheet.write(8, columna_max - 1, "OC Entregadas/Pagadas/Productos", head_style)
    worksheet.write(9, columna_max - 1, "OC Pagadas/Productos", head_style)
    worksheet.write(10, columna_max - 1, "KPI OC Entregadas/Total de OC", head_style)
    if num_approved_requis <= 0:
         num_approved_requis=1
    indicador = num_requis_atendidas/num_approved_requis
    letra_columna = xl_col_to_name(columna_max)
    formula = f"={letra_columna}9/{letra_columna}10"
    # Escribir datos y fórmulas
    worksheet.write(2, columna_max, start_date, date_style)  # Ejemplo de escritura de fecha
    worksheet.write(3, columna_max, end_date, date_style)
    worksheet.write_formula(4, columna_max, '=COUNTA(A:A)-1', body_style)  # Ejemplo de fórmula
    worksheet.write(5, columna_max, num_approved_requis, body_style)
    worksheet.write(6, columna_max, num_requis_atendidas, body_style)
    worksheet.write(7, columna_max, indicador, percent_style)  # Ajuste del índice de fila y columna para xlsxwriter
    worksheet.write_formula(8, columna_max, '=COUNTIFS(P:P, "Pagada", W:W, "Entregada", S:S, "PRODUCTOS")', body_style)
    # Escribir otra fórmula COUNTIF, también con el estilo corporal
    worksheet.write_formula(9, columna_max, '=COUNTIFS(P:P, "Pagada", S:S, "PRODUCTOS")', body_style)
    worksheet.write_formula(10, columna_max, formula, percent_style)

    for i, column in enumerate(columns):
        worksheet.write(0, i, column, head_style)
        worksheet.set_column(i, i, 15)  # Ajusta el ancho de las columnas

    worksheet.set_column('L:L', 12,  money_style)
    worksheet.set_column('M:M', 12, money_style) 
    # Asumiendo que ya tienes tus datos de compras
    row_num = 0
    for compra_list in compras:
        row_num += 1
        # Aquí asumimos que ya hiciste el procesamiento necesario de cada compra
        pagos = Pago.objects.filter(oc=compra_list, hecho = True).annotate(
            fecha_orden=Coalesce('pagado_real', 'pagado_date', output_field=DateField())
        ).order_by('pagado_date')

        if pagos.exists():
            primer_pago = pagos.first()
            primera_fecha_pago = primer_pago.pagado_real if primer_pago.pagado_real else primer_pago.pagado_date
            primera_fecha_pago = primera_fecha_pago.strftime('%d/%m/%Y')
        else:
            primera_fecha_pago = " "


        tipo_de_cambio_promedio_pagos = pagos.aggregate(Avg('tipo_de_cambio'))['tipo_de_cambio__avg']
        articulos = compra_list.articulocomprado_set.all()
         # Determinar el tipo de producto para la columna de tipo_producto
        todos_servicios = all(articulo.producto.producto.articulos.producto.producto.servicio for articulo in articulos)
        ningun_servicio = all(not articulo.producto.producto.articulos.producto.producto.servicio for articulo in articulos)

        if todos_servicios:
            tipo_producto = "SERVICIOS"
        elif ningun_servicio:
            tipo_producto = "PRODUCTOS"
        else:
            tipo_producto = "PRODUCTO/SERVICIOS"
    

        # Usar el tipo de cambio de los pagos, si existe. De lo contrario, usar el tipo de cambio de la compra
        tipo = tipo_de_cambio_promedio_pagos or compra_list.tipo_de_cambio
        tipo_de_cambio = '' if tipo == 0 else tipo
        created_at = compra_list.created_at.replace(tzinfo=None)
        approved_at = compra_list.req.approved_at
        try:
            activo_obj = compra_list.req.orden.activo
            activo = str(activo_obj) if activo_obj else "No definido"
        except Activo.DoesNotExist:
            activo = "No definido"

        row = [
            compra_list.folio,
            compra_list.req.folio,
            compra_list.req.orden.folio,
            compra_list.req.orden.distrito.nombre,
            compra_list.req.orden.proyecto.nombre if compra_list.req.orden.proyecto else '',
            compra_list.req.orden.subproyecto.nombre if compra_list.req.orden.subproyecto else '',
            compra_list.req.orden.operacion.nombre if compra_list.req.orden.operacion else '',
            f"{compra_list.req.orden.staff.staff.staff.first_name} {compra_list.req.orden.staff.staff.staff.last_name}",
            f"{compra_list.creada_por.staff.staff.first_name} {compra_list.creada_por.staff.staff.last_name}",
            created_at,
            approved_at,
            compra_list.proveedor.nombre.razon_social,
            compra_list.estatus_original,
            compra_list.cond_de_pago.nombre,
            compra_list.costo_oc,
            compra_list.monto_pagado,
            'Pagada' if compra_list.pagada else 'No Pagada',
            primera_fecha_pago,
            'Autorizado' if compra_list.autorizado2 else 'Cancelado' if compra_list.autorizado2 == False or compra_list.autorizado1 == False else 'Pendiente Autorización',
            tipo_producto,
            compra_list.dias_de_entrega,
            compra_list.moneda.nombre,
            tipo_de_cambio,  # Asegúrate de que tipo_de_cambio sea un valor que pueda ser escrito directamente
            'Entregada' if compra_list.entrada_completa else 'No Entregada',
            'Sí' if compra_list.facturas.exists() else 'No',
            activo,
        ]
        
        for col_num, cell_value in enumerate(row):
        # Define el formato por defecto
            cell_format = body_style

            # Aplica el formato de fecha para las columnas con fechas
            if col_num in [9, 10, 17]:  # Asume que estas son tus columnas de fechas
                cell_format = date_style
        
            # Aplica el formato de dinero para las columnas con valores monetarios
            elif col_num in [13, 14]:  # Asume que estas son tus columnas de dinero
                cell_format = money_style

            # Finalmente, escribe la celda con el valor y el formato correspondiente
            worksheet.write(row_num, col_num, cell_value, cell_format)

      
        worksheet.write_formula(row_num, 26, f'=IF(ISBLANK(W{row_num+1}), O{row_num+1}, O{row_num+1}*W{row_num+1})', money_style)
    
   
    workbook.close()

    # Construye la respuesta
    output.seek(0)

    response = HttpResponse(
        output.read(), 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    response['Content-Disposition'] = f'attachment; filename=Matriz_compras_{dt.date.today()}.xlsx'
      # Establecer una cookie para indicar que la descarga ha iniciado
    response.set_cookie('descarga_iniciada', 'true', max_age=3)  # La cookie expira en 20 segundos
    output.close()
    return response



def convert_excel_solicitud_matriz_productos(productos):
    print(productos.count())
    #response= HttpResponse(content_type = "application/ms-excel")
    #response['Content-Disposition'] = 'attachment; filename = Solicitudes_por_producto_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Compras_Producto')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)
    number_style = NamedStyle(name='number_style', number_format='#,##0.00')
    number_style.font = Font(name ='Calibri', size = 10)

    columns = ['OC','Código', 'Producto','Cantidad','Unidad','Tipo Item','Familia','Subfamilia','P.U.','Moneda','TC','Subtotal','IVA','Total','Proveedor','Status Proveedor','Fecha','Proyecto','Subproyecto','Distrito','RQ','Sol','Status','Pagada','Comentario Solicitud']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16
        if col_num == 4 or col_num == 7:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 25
        if col_num == 11:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30



    columna_max = len(columns)+2

    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por SAVIA 2.0. UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Grupo Vordcab S.A. de C.V.}')).style = messages_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 20
    
    
    rows = []
    for articulo in productos:
        #producto_id = producto.get(id)
        #articulo = ArticuloComprado.objects.get(id=producto_id)
        # Extract the needed attributes
        compra_id = articulo.oc.id
        moneda_nombre = articulo.oc.moneda.nombre
        #nombre_completo = articulo.oc.req.orden.staff.staff.staff.first_name + " " + articulo.oc.req.orden.staff.staff.staff.last_name
        proyecto_nombre = articulo.oc.req.orden.proyecto.nombre if articulo.oc.req.orden.proyecto else "Desconocido"
        subproyecto_nombre = articulo.oc.req.orden.subproyecto.nombre if articulo.oc.req.orden.subproyecto else "Desconocido"
        operacion_nombre = articulo.oc.req.orden.operacion.nombre if articulo.oc.req.orden.operacion else "Desconocido"
        fecha_creacion = articulo.created_at
        pagado_text = 'Pagada' if articulo.oc.pagada else 'No Pagada'

        # Calculate total, subtotal, and IVA using attributes from producto
        subtotal_parcial = articulo.subtotal_parcial
        iva_parcial = articulo.iva_parcial
        total = articulo.total
        if articulo.oc.autorizado2 is not None:
            status = 'Autorizado Gerente' if articulo.oc.autorizado2 else 'Cancelada'
        elif articulo.oc.autorizado1 is not None:
            status = 'Autorizado Superintendente' if articulo.oc.autorizado1 else 'Cancelada'
        else:
            status = 'Sin autorizaciones aún'
        # Handling the currency conversion logic
        pagos = Pago.objects.filter(oc_id=compra_id)
        tipo_de_cambio_promedio_pagos = pagos.aggregate(Avg('tipo_de_cambio'))['tipo_de_cambio__avg']
        tipo_de_cambio = tipo_de_cambio_promedio_pagos or articulo.oc.tipo_de_cambio

        if moneda_nombre == "DOLARES" and tipo_de_cambio:
            total = total * tipo_de_cambio

        comentarios = articulo.producto.articulos.producto.comentarios if articulo.producto.articulos.producto.comentario else "Sin comentario"

        # Constructing the row
        row = [
            articulo.oc.folio,
            articulo.producto.producto.articulos.producto.producto.codigo,
            articulo.producto.producto.articulos.producto.producto.nombre,
            articulo.cantidad,
            articulo.producto.producto.articulos.producto.producto.unidad,
            'SERVICIO' if articulo.producto.producto.articulos.producto.producto.servicio else  'PRODUCTO',
            articulo.producto.producto.articulos.producto.producto.familia.nombre,
            articulo.producto.producto.articulos.producto.producto.subfamilia.nombre if articulo.producto.producto.articulos.producto.producto.subfamilia else 'Desconocido',
            articulo.precio_unitario,
            moneda_nombre,
            tipo_de_cambio,
            subtotal_parcial,
            iva_parcial,
            total,
            articulo.oc.proveedor.nombre.razon_social,
            articulo.oc.estatus_original,
            fecha_creacion,
            #nombre_completo,
            proyecto_nombre,
            subproyecto_nombre,
            #operacion_nombre,
            articulo.oc.req.orden.distrito.nombre,
            articulo.oc.req.folio,
            articulo.oc.req.orden.folio,
            status,
            pagado_text,
            comentarios,
        ]
        rows.append(row)

    #Ahora, iteramos sobre las filas recopiladas para construir el archivo Excel:
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            (ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style
            if col_num == 5:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = body_style
            if col_num == 18:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style
            if col_num in [3]:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = number_style
            if col_num in [8, 10, 11, 12, 13]:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = money_style

    file_name='Matriz_compras_por_producto' + str(date.today()) + '.xlsx'
   
    sheet = wb['Sheet']
    wb.remove(sheet)
    output = io.BytesIO()
    wb.save(output)  # Guardar el libro de trabajo en el objeto BytesIO
    
    # Configurar la respuesta para descargar el archivo
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=%s' % file_name
    response.set_cookie('descarga_iniciada', 'true', max_age=20)  # La cookie expira en 20 segundos
    # Cerrar el objeto BytesIO
    output.close()

    return response


def convert_excel_solicitud_matriz_productos_prov(productos):
    #print(productos.count())
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet("Compras_Producto")

    # Define los estilos
    head_style = workbook.add_format({'bold': True, 'font_color': 'FFFFFF', 'bg_color': '333366', 'font_name': 'Arial', 'font_size': 11})
    body_style = workbook.add_format({'font_name': 'Calibri', 'font_size': 10})
    date_style = workbook.add_format({'num_format': 'dd/mm/yyyy', 'font_name': 'Calibri', 'font_size': 10})
    money_style = workbook.add_format({'num_format': '$ #,##0.00', 'font_name': 'Calibri', 'font_size': 10})
    number_style = workbook.add_format({'num_format': '#,##0.00', 'font_name': 'Calibri', 'font_size': 10})
    messages_style = workbook.add_format({'font_name': 'Arial Narrow', 'font_size': 11})

    columns = ['OC', 'Distrito', 'Código', 'Producto', 'Cantidad', 'Unidad', 'Tipo Item', 'Familia', 'Subfamilia', 'P.U.', 'Moneda', 'TC', 'Subtotal', 'IVA', 'Total', 'Proveedor', 'Status Proveedor','Visita', 'Dirección', 'Fecha', 'Proyecto', 'Subproyecto', 'Distrito', 'RQ', 'Sol', 'Status', 'Pagada']

    

    # Escribir encabezados
    for col_num, column in enumerate(columns):
        worksheet.write(0, col_num, column, head_style)
        worksheet.set_column(col_num, col_num, 16)
        if col_num in [4, 7]:
            worksheet.set_column(col_num, col_num, 25)
        if col_num == 11:
            worksheet.set_column(col_num, col_num, 30)

    columna_max = len(columns) + 2

    worksheet.write(0, columna_max, '{Reporte Creado Automáticamente por SAVIA 2.0. UH}', messages_style)
    worksheet.write(1, columna_max, '{Software desarrollado por Grupo Vordcab S.A. de C.V.}', messages_style)
    worksheet.set_column(columna_max, columna_max, 20)

    row_num = 1
    for articulo in productos:
        compra_id = articulo.oc.id
        moneda_nombre = articulo.oc.moneda.nombre
        proyecto_nombre = articulo.oc.req.orden.proyecto.nombre if articulo.oc.req.orden.proyecto else "Desconocido"
        subproyecto_nombre = articulo.oc.req.orden.subproyecto.nombre if articulo.oc.req.orden.subproyecto else "Desconocido"
        fecha_creacion = articulo.created_at
        pagado_text = 'Pagada' if articulo.oc.pagada else 'No Pagada'
        subtotal_parcial = articulo.subtotal_parcial
        iva_parcial = articulo.iva_parcial
        total = articulo.total
        if articulo.oc.autorizado2 is not None:
            status = 'Autorizado Gerente' if articulo.oc.autorizado2 else 'Cancelada'
        elif articulo.oc.autorizado1 is not None:
            status = 'Autorizado Superintendente' if articulo.oc.autorizado1 else 'Cancelada'
        else:
            status = 'Sin autorizaciones aún'
        pagos = Pago.objects.filter(oc_id=compra_id)
        tipo_de_cambio_promedio_pagos = pagos.aggregate(Avg('tipo_de_cambio'))['tipo_de_cambio__avg']
        tipo_de_cambio = tipo_de_cambio_promedio_pagos or articulo.oc.tipo_de_cambio
        if moneda_nombre == "DOLARES" and tipo_de_cambio:
            total = total * tipo_de_cambio

        row = [
            articulo.oc.folio,
            articulo.oc.req.orden.distrito.nombre,
            articulo.producto.producto.articulos.producto.producto.codigo,
            articulo.producto.producto.articulos.producto.producto.nombre,
            articulo.cantidad,
            articulo.producto.producto.articulos.producto.producto.unidad.nombre,
            'SERVICIO' if articulo.producto.producto.articulos.producto.producto.servicio else 'PRODUCTO',
            articulo.producto.producto.articulos.producto.producto.familia.nombre,
            articulo.producto.producto.articulos.producto.producto.subfamilia.nombre if articulo.producto.producto.articulos.producto.producto.subfamilia else 'Desconocido',
            articulo.precio_unitario,
            moneda_nombre,
            tipo_de_cambio,
            subtotal_parcial,
            iva_parcial,
            total,
            articulo.oc.proveedor.nombre.razon_social,
            articulo.oc.proveedor.estatus.nombre,
            articulo.oc.proveedor.domicilio,
            fecha_creacion,
            proyecto_nombre,
            subproyecto_nombre,
            articulo.oc.req.orden.distrito.nombre,
            articulo.oc.req.folio,
            articulo.oc.req.orden.folio,
            status,
            pagado_text,
        ]

        for col_num, cell_value in enumerate(row):
            cell_format = body_style
            if col_num in [16]:  # Fecha
                cell_format = date_style
            elif col_num in [8, 10, 11, 12, 13]:  # Dinero
                cell_format = money_style
            elif col_num in [3]:  # Números
                cell_format = number_style
            worksheet.write(row_num, col_num, cell_value, cell_format)
        row_num += 1

    workbook.close()

    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    file_name = 'Matriz_compras_por_producto_' + str(date.today()) + '.xlsx'
    response['Content-Disposition'] = f'attachment; filename={file_name}'
    response.set_cookie('descarga_iniciada', 'true', max_age=20)
    output.close()

    return response

def convert_excel_solicitud_matriz_productos_prov2(productos):
    start_time = time.time()  # Marca el tiempo de inicio
    print('Aqui comienza',productos.count())

    columns = ['OC', 'Distrito', 'Código', 'Producto', 'Cantidad', 'Unidad', 'Tipo Item', 'Familia', 'Subfamilia', 'P.U.', 'Moneda', 'TC',
                'Subtotal', 'IVA', 'Total', 'Proveedor', 'Status Proveedor', 'Dirección', 'Estado','Fecha', 'Proyecto', 'Subproyecto', 'Distrito', 
                'RQ', 'Sol', 'Status', 'Pagada', 'Comentario Solicitud','Visita']
    data = [columns]

    for articulo in productos:
        compra_id = articulo.oc.id
        moneda_nombre = articulo.oc.moneda.nombre
        proyecto_nombre = articulo.oc.req.orden.proyecto.nombre if articulo.oc.req.orden.proyecto else "Desconocido"
        subproyecto_nombre = articulo.oc.req.orden.subproyecto.nombre if articulo.oc.req.orden.subproyecto else "Desconocido"
        fecha_creacion = articulo.oc.created_at.date()
        pagado_text = 'Pagada' if articulo.oc.pagada else 'No Pagada'
        subtotal_parcial = articulo.subtotal_parcial
        iva_parcial = articulo.iva_parcial
        total = articulo.total
        if articulo.oc.autorizado2 is not None:
            status = 'Autorizado Gerente' if articulo.oc.autorizado2 else 'Cancelada'
        elif articulo.oc.autorizado1 is not None:
            status = 'Autorizado Superintendente' if articulo.oc.autorizado1 else 'Cancelada'
        else:
            status = 'Sin autorizaciones aún'
        pagos = Pago.objects.filter(oc_id=compra_id)
        tipo_de_cambio_promedio_pagos = pagos.aggregate(Avg('tipo_de_cambio'))['tipo_de_cambio__avg']
        tipo_de_cambio = tipo_de_cambio_promedio_pagos or articulo.oc.tipo_de_cambio
        if moneda_nombre == "DOLARES" and tipo_de_cambio:
            total = total * tipo_de_cambio
        if articulo.oc.proveedor.nombre.visita == True:
            visita = 'Si'
        else:
            visita = 'No'
        comentarios = articulo.producto.producto.articulos.comentario if articulo.producto.producto.articulos.comentario else "Sin comentario"

        row = [
            articulo.oc.folio,
            articulo.oc.req.orden.distrito.nombre,
            articulo.producto.producto.articulos.producto.producto.codigo,
            articulo.producto.producto.articulos.producto.producto.nombre,
            articulo.cantidad,
            articulo.producto.producto.articulos.producto.producto.unidad.nombre,
            'SERVICIO' if articulo.producto.producto.articulos.producto.producto.servicio else 'PRODUCTO',
            articulo.producto.producto.articulos.producto.producto.familia.nombre,
            articulo.producto.producto.articulos.producto.producto.subfamilia.nombre if articulo.producto.producto.articulos.producto.producto.subfamilia else 'Desconocido',
            articulo.precio_unitario,
            moneda_nombre,
            tipo_de_cambio,
            subtotal_parcial,
            iva_parcial,
            total,
            articulo.oc.proveedor.nombre.razon_social,
            articulo.oc.proveedor.estatus.nombre,
            articulo.oc.proveedor.domicilio,
            articulo.oc.proveedor.estado.nombre if articulo.oc.proveedor.estado else 'No Identificado',
            fecha_creacion,
            proyecto_nombre,
            subproyecto_nombre,
            articulo.oc.req.orden.distrito.nombre,
            articulo.oc.req.folio,
            articulo.oc.req.orden.folio,
            status,
            pagado_text,
            comentarios,
            visita,
        ]
        data.append(row)

    # Crear el archivo Excel usando pyexcelerate
    wb = Workbook()
    ws = wb.new_sheet("Compras_Producto", data=data)

     # Aplicar estilos a los encabezados
    header_style = Style(
        font=Font(bold=True, color=PXColor(255, 255, 255)),
        fill=Fill(background=PXColor(51, 51, 102)),
        alignment=Alignment(horizontal='center', vertical='center')
    )

    

   # Aplicar estilos a las celdas de datos
    date_style = Style(
        format=Format('dd/mm/yyyy'),
        alignment=Alignment(horizontal='right')
    )
    #format_obj = pyexcelerate.Format('$#,##0.00')
    money_style = Style(
        format= Format('$##,##0.00'),
        alignment=Alignment(horizontal='right')
    )
    body_style = Style(
        alignment=Alignment(horizontal='left')
    )


    #for col_num, cell_value in enumerate(row):
    #        cell_format = body_style
    #        if col_num in [16]:  # Fecha
    #            cell_format = date_style
    
    for col_num in range(1, len(columns) + 1):
        if col_num in [20]:  # Fecha
            ws.set_col_style(col_num, date_style)
        elif col_num in [10, 12, 13, 14, 15]:  # Dinero
            ws.set_col_style(col_num, money_style)
        else:
            ws.set_col_style(col_num, body_style)

    for col_num in range(1, len(columns) + 1):
        ws[1][col_num].style = header_style

    output = io.BytesIO()
    wb.save(output)  # Guardar el libro de trabajo en el objeto BytesIO

    # Configurar la respuesta para descargar el archivo
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    file_name = 'Matriz_compras_por_producto_' + str(date.today()) + '.xlsx'
    response['Content-Disposition'] = f'attachment; filename={file_name}'
    response.set_cookie('descarga_iniciada', 'true', max_age=20)

    # Cerrar el objeto BytesIO
    output.close()
    #end_time = time.time()  # Marca el tiempo de finalización
    #total_time = end_time - start_time  # Calcula el tiempo total
    #print(f"Tiempo total para generar el archivo: {total_time} segundos")

    return response


def convert_excel_solicitud_matriz_productos_quick(productos): 
    start_time = time.time()  # Start timing
    print('Starting count:', productos.count())

    # Define headers and initial data array
    columns = ['OC', 'Distrito', 'Código', 'Producto', 'Cantidad', 'Unidad', 'Tipo Item', 'Familia', 'Subfamilia', 'P.U.', 'Moneda', 'TC', 'Subtotal', 'IVA', 'Total', 'Proveedor', 'Status Proveedor', 'Dirección', 'Fecha', 'Proyecto', 'Subproyecto', 'Distrito', 'RQ', 'Sol', 'Status', 'Pagada', 'Comentario Solicitud','Visita']
    data = [columns]

    # Populate data rows
    for articulo in productos:
        compra_id = articulo.oc.id
        moneda_nombre = articulo.oc.moneda.nombre
        proyecto_nombre = articulo.oc.req.orden.proyecto.nombre if articulo.oc.req.orden.proyecto else "Desconocido"
        subproyecto_nombre = articulo.oc.req.orden.subproyecto.nombre if articulo.oc.req.orden.subproyecto else "Desconocido"
        fecha_creacion = articulo.oc.created_at.date()
        pagado_text = 'Pagada' if articulo.oc.pagada else 'No Pagada'
        subtotal_parcial = articulo.subtotal_parcial
        iva_parcial = articulo.iva_parcial
        total = articulo.total
        status = (
            'Autorizado Gerente' if articulo.oc.autorizado2 else 'Cancelada' if articulo.oc.autorizado2 is False else
            'Autorizado Superintendente' if articulo.oc.autorizado1 else 'Cancelada' if articulo.oc.autorizado1 is False else
            'Sin autorizaciones aún'
        )
        subfamilia_nombre = (
            articulo.producto.producto.articulos.producto.producto.subfamilia.nombre
            if articulo.producto.producto.articulos.producto.producto.subfamilia else "Desconocido"
        )

        pagos = Pago.objects.filter(oc_id=compra_id)
        tipo_de_cambio_promedio_pagos = pagos.aggregate(Avg('tipo_de_cambio'))['tipo_de_cambio__avg']
        tipo_de_cambio = tipo_de_cambio_promedio_pagos or articulo.oc.tipo_de_cambio
        if moneda_nombre == "DOLARES" and tipo_de_cambio:
            total *= tipo_de_cambio
        if articulo.oc.proveedor.nombre.visita == True:
            visita = 'Si'
        else:
            visita = 'No'
        comentarios = articulo.producto.producto.articulos.comentario or "Sin comentario"

        row = [
            articulo.oc.folio,
            articulo.oc.req.orden.distrito.nombre,
            articulo.producto.producto.articulos.producto.producto.codigo,
            articulo.producto.producto.articulos.producto.producto.nombre,
            articulo.cantidad,
            articulo.producto.producto.articulos.producto.producto.unidad.nombre,
            'SERVICIO' if articulo.producto.producto.articulos.producto.producto.servicio else 'PRODUCTO',
            articulo.producto.producto.articulos.producto.producto.familia.nombre,
            subfamilia_nombre,
            articulo.precio_unitario,
            moneda_nombre,
            tipo_de_cambio,
            subtotal_parcial,
            iva_parcial,
            total,
            articulo.oc.proveedor.nombre.razon_social,
            articulo.oc.proveedor.estatus.nombre,
            articulo.oc.proveedor.domicilio,
            fecha_creacion,
            proyecto_nombre,
            subproyecto_nombre,
            articulo.oc.req.orden.distrito.nombre,
            articulo.oc.req.folio,
            articulo.oc.req.orden.folio,
            status,
            pagado_text,
            comentarios,
            visita,
        ]
        data.append(row)

    # Create Excel workbook and worksheet
    wb = Workbook()
    ws = wb.new_sheet("Compras_Producto", data=data)

    # Define styles
    header_style = Style(
        font=Font(bold=True),
        alignment=Alignment(horizontal='center', vertical='center')
    )
    date_style = Style(format=Format('dd/mm/yyyy'))
    money_style = Style(format=Format('$#,##0.00'))

    # Apply header style
    for col_num in range(1, len(columns) + 1):
        ws[1][col_num].style = header_style

    # Apply styles for specific columns
    for row_num in range(2, len(data) + 1):
        ws[row_num][19].style = date_style  # Fecha
        ws[row_num][10].style = money_style  # P.U.
        ws[row_num][13].style = money_style  # IVA
        ws[row_num][14].style = money_style  # Total

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)

    # Set up response for download
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    file_name = f'Matriz_compras_por_producto_{date.today()}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={file_name}'
    response.set_cookie('descarga_iniciada', 'true', max_age=3)
    output.close()

    print(f"Total time to generate file: {time.time() - start_time} seconds")
    return response


def generar_politica_antisoborno():
    #Configuration of the PDF object
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    #doc = SimpleDocTemplate(buf, pagesize=letter)
    #Here ends conf.
    #compra = Compra.objects.get(id=pk)
    width, height = letter
    #Azul Vordcab
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)
    caja_iso = 760
    #Elaborar caja
    #c.line(caja_iso,500,caja_iso,720)

    #c.drawString(430,caja_iso,'Preparado por:')
    #c.drawString(405,caja_iso-10,'Auditoría de Proveedores')

    #Primera Tabla
    caja_proveedor = caja_iso - 85
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(140,750,260,20, fill=True, stroke=False) #Barra azul superior Título
    c.rect(20,caja_proveedor - 7,565,20, fill=True, stroke=False) #Barra azul superior |Objetivo
    c.rect(20,caja_proveedor - 130,565,5, fill=True, stroke=False) #Linea posterior horizontal
    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(270,755,'Política Antisoborno')
    c.setLineWidth(.3) #Grosor
    #c.line(20,caja_proveedor-8,20,520) #Eje Y donde empieza, Eje X donde empieza, donde termina eje y,donde termina eje x (LINEA 1 contorno)
    #c.line(585,caja_proveedor-8,585,520) #Linea 2 contorno
    c.drawInlineImage('static/images/logo_vordcab.jpg',45,730, 3 * cm, 1.5 * cm) #Imagen vortec

    c.setFillColor(white)
    c.setFont('Helvetica-Bold',11)
    #c.drawString(200,caja_proveedor,'Objetivo')
    inicio_central = 300
    #c.line(inicio_central,caja_proveedor-25,inicio_central,520) #Linea Central de caja Proveedor | Detalle
    c.setFillColor(black)
    c.setFont('Helvetica-Bold',9)
    

    #Segundo Parrafo
    segundo_parrafo = caja_proveedor - 150
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(20,   segundo_parrafo - 8,565,20, fill=True, stroke=False) #Barra azul superior | Datos Bancarios Dólares
    c.rect(20,  segundo_parrafo - 160,565,5, fill=True, stroke=False) #Linea posterior horizontal
   
   
    c.setFillColor(black)
    c.setFont('Helvetica',9)

    #Cuarta tabla
    cuarta_tabla = segundo_parrafo - 400
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(20, cuarta_tabla - 8,565,20, fill=True, stroke=False) #Barra azul superior | Datos Bancarios Dólares
    c.rect(20, cuarta_tabla - 100,565,5, fill=True, stroke=False) #Linea posterior horizontal
   

   
    
   
    c.setFillColor(black)
    c.setFont('Helvetica',9)
   
   


    texto = """La presente política tiene como objetivo principal establecer principios y lineamientos de actuación que deberán 
    adoptarse por todos y cada uno de los miembros que formamos parte de VORDCAB en el ejercicio de sus funciones y promover ante
    terceros las buenas prácticas comerciales y corporativas de vigilancia permanente fundamentadas en la ética y los valores, 
    que eviten de manera definitiva la comisión de delitos como la corrupción, el soborno y la extorsión, o cualquier práctica en
    la cual se vea involucrado un conflicto de interés, implementando herramientas y mecanismos de comunicación constante entre 
    sus colaboradores para su debida aplicación con posición de “cero tolerancia” frente a cualquier acto que vulnere la legalidad
    y los buenos principios. <br/>
                """
    
    texto2 = """Esta política es de observancia y aplicación estricta de todos y cada uno de los empleados, trabajadores, representantes,
colaboradores, proveedores, distribuidores o cualquier tercero relacionado con las actividades comerciales de VORDCAB en nuestro país, 
así como en aquellos países en los que cuenta con presencia operativa.<br/>
VORDCAB contará con un Registro en relación con terceros, en donde se localicen sus nombres, términos y condiciones de los acuerdos que 
tomen con VORDCAB, así como los pagos realizados a los terceros contratados por la Empresa, relacionados a transacciones con organismos 
públicos o empresas estatales o privadas.<br/>
En caso de sociedades conjuntas (joint ventures) o consorcios, contratistas y proveedores, deberá constar el consentimiento del tercero 
para adoptar políticas anticorrupción y asegurarse de su cumplimiento, de acuerdo con los estándares comerciales aceptados, en apego a 
la transparencia.<br/> """
    texto3 = """El presente documento, SEOV N1 11 Política Antisoborno Antisoborno, se encuentra disponible en su versión original en 
    medios electrónicos para mayor referencia, para lo cual podrá visitar la página www.grupovordcab.com, facilitando su debido cumplimiento.
    En caso de existir alguna duda o comentario en la relación con la presente, podrá contactarse al siguiente correo 
    contactointerno@grupovordcab.com, directamente con el área de Jurídico, de GRUPO VORDCAB, S.A. DE C.V..<br/>"""
    titulo1 = """OBJETIVO"""
    titulo2 = """ALCANCE"""
    titulo3 = """NOTAS"""
    
    styles = getSampleStyleSheet()
    styleN = styles["BodyText"]
    styleN.fontSize = 10
    styleN.alignment = TA_JUSTIFY
   
    styleT = styles["Normal"]
    styleT.textColor = white
    styleT.alignment = TA_JUSTIFY

    styleItalic = ParagraphStyle(
        'Title',
        parent = styles["BodyText"],
        fontName = 'Helvetica-Oblique',
        fontSize = 8,
        alignment = TA_JUSTIFY
    )

    titulo1 = Paragraph(titulo1, styleT)
    parrafo = Paragraph(texto, styleN)
    titulo2 = Paragraph(titulo2, styleT)
    parrafo2 = Paragraph(texto2, styleN)
    titulo3 = Paragraph(titulo3, styleT)
    parrafo3 = Paragraph(texto3, styleItalic)

    ancho, alto = letter  # Asegúrate de tener estas dimensiones definidas
    #frame = Frame(120, 720, ancho - 100, alto - 100, id='frameTextoConstante')  # Ajusta las dimensiones según sea necesario
    elementos = [titulo1, Spacer(1,25), parrafo, Spacer(1,25),titulo2,Spacer(1,25), parrafo2, Spacer(1,250),titulo3,Spacer(1,25), parrafo3]
    frame = Frame(30, 0, width-50, height-100, id='frameTextoConstante')
    frame.addFromList(elementos, c)

    c.save()
    buf.seek(0)
    return buf
   


def generar_codigo_etica():
    #Configuration of the PDF object
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    #doc = SimpleDocTemplate(buf, pagesize=letter)
    #Here ends conf.
    #compra = Compra.objects.get(id=pk)
    width, height = letter
    #Azul Vordcab
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)
    caja_iso = 760
    #Elaborar caja
    #c.line(caja_iso,500,caja_iso,720)

    #c.drawString(430,caja_iso,'Preparado por:')
    #c.drawString(405,caja_iso-10,'Auditoría de Proveedores')

    #Primera Tabla
    caja_proveedor = caja_iso - 85
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(140,750,260,20, fill=True, stroke=False) #Barra azul superior Título
    c.rect(20,caja_proveedor - 7,565,20, fill=True, stroke=False) #Barra azul superior |Objetivo
    c.rect(20,caja_proveedor - 300,565,5, fill=True, stroke=False) #Linea posterior horizontal
    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(270,755,'Código de Ética')
    c.setLineWidth(.3) #Grosor
    #c.line(20,caja_proveedor-8,20,520) #Eje Y donde empieza, Eje X donde empieza, donde termina eje y,donde termina eje x (LINEA 1 contorno)
    #c.line(585,caja_proveedor-8,585,520) #Linea 2 contorno
    c.drawInlineImage('static/images/logo_vordcab.jpg',45,730, 3 * cm, 1.5 * cm) #Imagen vortec

    c.setFillColor(white)
    c.setFont('Helvetica-Bold',11)
    #c.drawString(200,caja_proveedor,'Objetivo')
    inicio_central = 300
    #c.line(inicio_central,caja_proveedor-25,inicio_central,520) #Linea Central de caja Proveedor | Detalle
    c.setFillColor(black)
    c.setFont('Helvetica-Bold',9)
    

    #Segundo Parrafo
    segundo_parrafo = caja_proveedor - 150
   
   
   
    c.setFillColor(black)
    c.setFont('Helvetica',9)

    #Cuarta tabla
    cuarta_tabla = segundo_parrafo - 400
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(20, cuarta_tabla - 8,565,20, fill=True, stroke=False) #Barra azul superior | Datos Bancarios Dólares
    c.rect(20, cuarta_tabla - 100,565,5, fill=True, stroke=False) #Linea posterior horizontal
   

   
    
   
    c.setFillColor(black)
    c.setFont('Helvetica',9)
   
   


    texto = """El desarrollo de relaciones de confianza con los proveedores ha contribuido a que Grupo Vordcab se encuentre como la empresa líder en
    fabricación y servicios de Sistemas Artificiales de Producción en México. Por tal motivo buscamos siempre que la relación con nuestros proveedores se base en
    la transparencia, la confianza y la confidencialidad entre nosotros.<br/>
Tenemos el compromiso de otorgar a los proveedores las mismas oportunidades de contratarlos ya que la evaluación de las ofertas para selección de proveedores se basa
 en criterios establecidos por nuestra empresa. Las propuestas son revisadas de manera integral considerando precio, valor agregado y calidad de los productos y 
 servicios que se ofrecen.<br/>
Las negociaciones son llevadas a cabo de manera honesta y equitativa; todo proveedor es tratado con respeto, transparencia y justicia.<br/>
El respeto de los acuerdos, los términos, licencias y compromisos establecidos en nuestros contratos son un principio en Grupo Vordcab.<br/>
Estamos comprometidos con la confidencialidad de los datos de nuestros proveedores, respetamos los derechos de propiedad intelectual e industrial es por ello que 
siempre buscamos establecer relaciones con contratistas o proveedores que demuestren que están debidamente autorizados para el uso o comercialización de productos 
o servicios.<br/>
Establecer acuerdos claros en términos y condiciones de pago, así como establecer procesos estables y transparentes de pago oportuno nos permite cumplir nuestros 
compromisos con nuestros proveedores.<br/>
En Grupo Vordcab se encuentra estrictamente prohibida cualquier situación de corrupción, por tanto, no aceptamos dinero, regalos, servicios, descuentos, viajes, 
entretenimientos o cualquier otro bien que pudiera poner en entre dicho nuestra transparencia en los procesos licitatorios o de compras. Sin embargo, también 
entendemos la necesidad de nuestros proveedores de hacer esfuerzos por promover sus marcas por lo que sí está permitida la entrega de productos promocionales 
siempre y cuando el valor de los mismos sea simbólico. <br/>
"""
    
    texto3 = """El presente documento, SEOV-N1-04 Código de Ética, se encuentra disponible en su versión original en medios electrónicos para mayor referencia,
      para lo cual podrá visitar la página www.grupovordcab.com, facilitando su debido cumplimiento. En caso de existir alguna duda o comentario en la relación con la 
      presente, podrá contactarse al siguiente correo contactointerno@grupovordcab.com, directamente con el área de Jurídico, de GRUPO VORDCAB, S.A. DE C.V.<br/> """
    
    titulo1 = """Con los proveedores"""
    titulo3 = """NOTAS"""
    
    styles = getSampleStyleSheet()
    styleN = styles["BodyText"]
    styleN.fontSize = 10
    styleN.alignment = TA_JUSTIFY
   
    styleT = styles["Normal"]
    styleT.textColor = white
    styleT.alignment = TA_JUSTIFY

    styleItalic = ParagraphStyle(
        'Title',
        parent = styles["BodyText"],
        fontName = 'Helvetica-Oblique',
        fontSize = 8,
        alignment = TA_JUSTIFY
    )

    titulo1 = Paragraph(titulo1, styleT)
    parrafo = Paragraph(texto, styleN)
    titulo3 = Paragraph(titulo3, styleT)
    parrafo3 = Paragraph(texto3, styleItalic)

    ancho, alto = letter  # Asegúrate de tener estas dimensiones definidas
    #frame = Frame(120, 720, ancho - 100, alto - 100, id='frameTextoConstante')  # Ajusta las dimensiones según sea necesario
    elementos = [
        titulo1, Spacer(1,25),
        parrafo, Spacer(1,270), 
        titulo3,Spacer(1,25), 
        parrafo3
    ]
    frame = Frame(30, 0, width-50, height-100, id='frameTextoConstante')
    frame.addFromList(elementos, c)

    c.save()
    buf.seek(0)
    return buf 


def generar_aviso_privacidad():
    #Configuration of the PDF object
    buf = io.BytesIO()
  
    c = canvas.Canvas(buf, pagesize=letter)
    #Here ends conf.
    width, height = letter
    #Azul Vordcab
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)
    caja_iso = 760
    #Elaborar caja
    #c.line(caja_iso,500,caja_iso,720)
    c.drawInlineImage('static/images/logo_vordcab.jpg',45,742, 2 * cm, 1.0 * cm) #Imagen vortec
    c.drawString(425,caja_iso,'Preparado por:')
    c.drawString(435,caja_iso-10,'SUBAD')
    c.drawString(520,caja_iso,'Aprobación')
    c.drawString(515,caja_iso-10,'SUBD ADTVO')
    c.drawString(35,caja_iso-20,'Número de documento')
    c.drawString(36,caja_iso-30,'SEOV-ADQ-N4-01.08')
    c.drawString(145,caja_iso-20,'Clasificación del documento')
    c.drawString(175,caja_iso-30,'No Controlado')
    c.drawString(255,caja_iso-20,'Nivel del documento')
    c.drawString(280,caja_iso-30, 'N5')
    c.drawString(340,caja_iso-20,'Revisión No.')
    c.drawString(352,caja_iso-30,'001')
    c.drawString(410,caja_iso-20,'Fecha de Emisión')
    c.drawString(425,caja_iso-30,'14/02/2022')
    c.drawString(510,caja_iso-20,'Fecha de Revisión')
    c.drawString(525,caja_iso-30,'12/09/2023')
    #Primera Tabla
    caja_proveedor = caja_iso - 85
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(140,750,260,20, fill=True, stroke=False) #Barra azul superior Título
    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(270,755,'Aviso de Privacidad para Proveedores')
    c.setLineWidth(.3) #Grosor
    

    c.setFillColor(white)
    c.setFont('Helvetica-Bold',11)
    #c.drawString(200,caja_proveedor,'Objetivo')
    inicio_central = 300
    #c.line(inicio_central,caja_proveedor-25,inicio_central,520) #Linea Central de caja Proveedor | Detalle
    c.setFillColor(black)
    c.setFont('Helvetica-Bold',9)
    

    #Segundo Parrafo
    segundo_parrafo = caja_proveedor - 65
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(20,   segundo_parrafo - 8,565,20, fill=True, stroke=False) #Barra azul superior 
    
   
   
    c.setFillColor(black)
    c.setFont('Helvetica',9)

    #Cuarta tabla
    cuarta_tabla = segundo_parrafo - 127
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(20, cuarta_tabla - 8,565,20, fill=True, stroke=False) #Barra azul superior 
    #c.rect(20, cuarta_tabla - 100,565,5, fill=True, stroke=False) #Linea posterior horizontal
   
   
    c.setFillColor(black)
    c.setFont('Helvetica',9)
   

    texto = """El presente aviso de privacidad en lo sucesivo "EL AVISO" se emite en cumplimiento a lo dispuesto por el artículo 15 de 
    la Ley Federal de Protección de Datos Personales en Posesión de los Particulares en lo sucesivo "LA LEY", y sus correlativos del 
    Reglamento de la Ley Federal de Protección de Datos Personales en Posesión de los Particulares en lo sucesivo "EL REGLAMENTO", y se
    pone a disposición de los proveedores que entreguen datos o información de carácter personal a GRUPO VORDCAB, S.A. DE C.V. <br/>
    """
    
    texto2 = """GRUPO VORDCAB, S.A. DE C.V. con domicilio en <u>Avenida Araucarias No.55, Col. Fuentes de las Animas, C.P. 91190, Xalapa de Enríquez, 
    Veracruz,</u> es responsable del tratamiento de sus datos personales.<br/>
    GRUPO VORDCAB, S.A. DE C.V. en su carácter de responsable del tratamiento de datos personales observa los principios de licitud, calidad, consentimiento, 
    información, finalidad, lealtad, proporcionalidad y responsabilidad, previstos en la Ley Federal de Protección de Datos Personales en Posesión de los Particulares.<br/>
 """
    texto3 = """ Al tener una relación comercial con GRUPO VORDCAB S.A. DE C.V., otorga su consentimiento expreso para el tratamiento de estos
    datos conforme a lo dispuesto en el artículo 8 de “LA LEY”.<br/>
    Los datos personales que recabamos de usted para dar cumplimiento a los fines descritos en el presente "AVISO", son recabados cuando usted nos 
los proporciona a través de nuestros empleados, aplicaciones móviles, de manera telefónica, por correo electrónico, y cuando obtenemos información 
a través de otras fuentes de acceso público permitidas por "LA LEY" y “EL REGLAMENTO”.<br/>
Los datos que se solicitarán serán los siguientes:<br/>
A) Datos generales:<br/>
o Acta constitutiva si es Persona Moral o Alta de Hacienda si es Persona Física.<br/>
o Constancia de situación Fiscal.<br/>
o Comprobante de Domicilio.<br/>
o Carta con datos bancarios para recibir pagos.<br/>
o Datos de contacto como correo electrónico y teléfono.<br/>
o Constancia de cumplimiento de obligaciones (32 D).<br/>
B) Adicional y solo para Proveedores de Servicios:<br/>
o Currículo de servicios.<br/>
o Cotización de precios unitarios de sus servicios<br/>
o Carta Garantía de los servicios solo aplica para servicios de infraestructura y mantenimiento.<br/>
o Opinión de cumplimento de obligaciones de seguridad social para proveedores de servicios<br/>
o Lista de activos con la que desarrollara las actividades<br/>
o Requisitos en Materia de Gestión de Seguridad, Salud y Medio Ambiente.<br/>
C) Adicional y solo para Proveedores de Muebles:<br/>
o Factura o contrato de mutuo con la que comprueba la propiedad de mueble.<br/>
o Tarjeta de circulación.<br/>
o Verificaciones y póliza de seguro de responsabilidad amplia.<br/>
"""
    texto4 = """Sus datos personales serán utilizados para las siguientes finalidades, finalidades que dieron origen y son necesarias para la existencia, 
    mantenimiento y cumplimiento de la relación comercial entre usted y GRUPO VORDCAB:<br/>

o Identificación y contacto.<br/>
o Participar en el proceso de evaluación, selección de proveedores.<br/>
o Verificar y confirmar su identidad como proveedor, así como la autenticidad de la información que nos proporciona, incluyendo la de sus terceros autorizados,
 tales como sus referencias, obligados solidarios, avales o fiadores y empleados del proveedor, según resulte aplicable.<br/>
o Realizar procesos de investigación internos y externos, y realizar auditorías.<br/>
o Cotizar productos y/o servicios.<br/>
o Elaborar, verificar y dar seguimiento al cumplimiento del objeto del contrato celebrado, y, en su caso, la renovación correspondiente.<br/>
o Para el control, vigilancia y acceso a las instalaciones de GRUPO VORDCAB.<br/>
o Facturación o pago derivado de la relación contractual.<br/>
o Creación, actualización, personalización, mantenimiento y autenticación de su cuenta de usuario.<br/>
o Atención de dudas, quejas, comentarios, sugerencias, aclaraciones y seguimiento a las mismas.<br/>
o Notificar cambios de condiciones y mantenimiento de la relación comercial.<br/>
o Cumplimiento de obligaciones legales y normativas, así como de requerimientos de autoridades gubernamentales o judiciales Federales, Estatales o Municipales 
y/o entidades regulatorias.<br/>
o Realizar tratamientos de técnicas de análisis masivo de datos para realizar actividades de perfilamiento.<br/>
"""
   
    texto5 = """Hacemos de su conocimiento que sus datos personales serán resguardados bajo estrictas medidas de seguridad administrativas, técnicas y físicas 
    las cuales han sido implementadas con el objeto de proteger sus datos personales contra daño, pérdida, alteración, destrucción o el uso, acceso o tratamiento 
    no autorizados.<br/>
"""
    texto6 = """GRUPO VORDCAB podrá dar tratamiento a datos personales de identificación y contacto, tales como nombre completo, correo electrónico y número telefónico, 
    de terceras personas, como los empleados del proveedor, autorizados del proveedor, referencias del proveedor, obligados solidarios, avales o fiadores, cuando usted 
    los proporcione para el cumplimiento de la relación contractual, o para participar en el proceso de evaluación y selección de proveedores, por lo que usted al 
    proporcionarlos reconoce haber informado a dichos terceros sobre el uso de sus datos personales, haber obtenido de forma previa el consentimiento de estos últimos 
    para que GRUPO VORDCAB pueda tratar sus datos personales para los fines antes señalados y haberles informado sobre el presente aviso de privacidad.<br/>
"""
    texto7 = """GRUPO VORDCAB no recabará ni dará tratamiento a datos personales sensibles de ninguna clase.<br/>
"""
    texto8 = """Usted o su representante legal debidamente acreditado podrán ejercer, cuando procedan, los derechos de acceso, rectificación, cancelación u oposición en 
    lo sucesivo "DERECHOS ARCO" que la "LEY" prevé.<br/><br/>

El ejercicio de los "DERECHOS ARCO" así como la revocación de su consentimiento para el tratamiento de sus datos personales se realizará a través de la presentación de
la solicitud respectiva, que por escrito deba presentar el titular de los datos personales, su apoderado o representante legal o bien al correo electrónico de contacto
de nuestros empleados.<br/><br/>

Su solicitud deberá indicar nombre completo y apellidos, copia simple, legible y vigente de su identificación oficial, en caso de utilizar medios electrónicos, deberá 
adjuntar la versión digitalizada de la misma (escaneo), correo electrónico o domicilio que designe para notificaciones y algún número telefónico de contacto. Si su 
solicitud es presentada por su apoderado o representante legal, deberá adicionalmente acompañar, los documentos oficiales que acrediten dicha representación.<br/><br/>

En caso de que la información proporcionada en su solicitud sea errónea o insuficiente para atenderla, o bien no se acompañen los documentos de acreditación 
correspondientes, se le hará un requerimiento dentro de los cinco días hábiles siguientes a la recepción de su solicitud, para que aporte los elementos o documentos 
necesarios para dar trámite a la misma. Usted contará con un plazo de diez días hábiles para atender dicho requerimiento, en caso de no dar respuesta a dicho 
requerimiento en el plazo otorgado, su solicitud se tendrá por no presentada.<br/>
"""
    texto9 = """Sus datos personales no serán transferidos a terceros sin su consentimiento, salvo por lo dispuesto en el artículo 37 de "LA LEY":<br/><br/>
o Cuando la transferencia se realice entre sociedades controladoras, subsidiarias o afiliadas bajo el control común de GRUPO VORDCAB, o a una sociedad matriz o a 
    cualquier sociedad del mismo grupo de GRUPO VORDCAB que opere bajo los mismos procesos y políticas internas.<br/>
o Cuando la transferencia sea necesaria para la prevención o el diagnóstico médico, la prestación de asistencia sanitaria, tratamiento médico o la gestión de servicios 
sanitarios.<br/>
o Cuando la transferencia sea necesaria por virtud de un contrato celebrado o por celebrar que sea en su interés, por GRUPO VORDCAB y un tercero.<br/>
o Cuando la transferencia sea precisa para el mantenimiento o cumplimiento de una relación jurídica entre GRUPO VORDCAB y usted, tales como las instituciones bancarias
 y crediticias, cámaras de comercio, socios comerciales, entre otras.<br/>
o Cuando la transferencia sea necesaria o legalmente exigida para la salvaguarda de un interés público, o para la procuración o administración de justicia, o cuando sea 
solicitado por autoridades competentes.<br/>
"""
    texto10 = """GRUPO VORDCAB conservará su información durante el tiempo que la necesite para el propósito para el que se recabo, a menos que nos solicite la 
    eliminación de la misma, y, en tal caso, siempre que GRUPO VORDCAB ya no tenga la necesidad de conservar su información por otros motivos. GRUPO VORDCAB podrá 
    conservar sus datos personales por períodos más prolongados que los que requieren las leyes aplicables, si es de nuestro interés comercial legítimo y las leyes 
    no lo prohíben.<br/>"""
    texto11 = """Tenga en cuenta que cuando hayamos recabado su información personal en función de su consentimiento y usted retire dicho consentimiento, o ejerza sus
    derechos ARCO (a suprimir su información personal), es posible que mantengamos su información bloqueada y disponible tanto tiempo como se requiera para cumplir con 
    las leyes aplicables y para que GRUPO VORDCAB cumpla con sus responsabilidades derivadas del procesamiento de sus datos.<br/>"""
    texto12 = """GRUPO VORDCAB se reserva el derecho de efectuar en cualquier tiempo modificaciones o actualizaciones al presente "AVISO". Las modificaciones que se 
    efectúen se pondrán a su disposición a través de alguno o algunos de los siguientes medios: anuncios visibles en nuestras instalaciones, aplicaciones móviles y/o 
    vía correo electrónico a la dirección más reciente que tengamos de usted.<br/>
    """
    texto13 = """El hecho de que usted nos proporcione por cualquier medio sus datos implica que otorga su consentimiento libre, específico, informado e inequívoco para
    el tratamiento de estos, en los términos del presente aviso de privacidad, sin perjuicio de la facultad que usted tiene de ejercer sus derechos ARCO. """

    texto14 = """ """

  
    titulo1 = """ """
    titulo2 = """1. Responsable de la protección de sus Datos Personales"""
    titulo3 = """2. Datos personales que recabamos y medios de obtención"""
    titulo4 = """3. Finalidades del Tratamiento de los Datos Personales."""
    titulo5 = """4. Medios para limitar el uso o divulgación de sus datos personales."""
    titulo6 = """5. Datos Personales de terceros."""
    titulo7 = """6. Datos personales sensibles."""
    titulo8 = """7. Derechos ARCO."""
    titulo9 = """8. Transferencia de datos."""
    titulo10 = """9. Conservación de datos"""
    titulo11 = """10. Modificaciones al presente aviso de privacidad"""
    titulo12 = """11. Consentimiento"""
    
    styles = getSampleStyleSheet()
    styleN = styles["BodyText"]
    styleN.fontSize = 10
    styleN.alignment = TA_JUSTIFY
   
    styleT = styles["Normal"]
    styleT.textColor = white
    styleT.alignment = TA_JUSTIFY

    styleItalic = ParagraphStyle(
        'Title',
        parent = styles["BodyText"],
        fontName = 'Helvetica-Oblique',
        fontSize = 8,
        alignment = TA_JUSTIFY
    )

    titulo1 = Paragraph(titulo1, styleT)
    parrafo = Paragraph(texto, styleN)
    titulo2 = Paragraph(titulo2, styleT)
    parrafo2 = Paragraph(texto2, styleN)
    titulo3 = Paragraph(titulo3, styleT)
    parrafo3 = Paragraph(texto3, styleN)
    parrafo4 = Paragraph(texto4, styleN)
    titulo4 = Paragraph(titulo4, styleT)
    parrafo5 = Paragraph(texto5, styleN)
    titulo5 = Paragraph(titulo5, styleT)
    parrafo6 = Paragraph(texto6, styleN)
    titulo6 = Paragraph(titulo6, styleT)
    parrafo7 = Paragraph(texto7, styleN)
    titulo7 = Paragraph(titulo7, styleT)
    parrafo8 = Paragraph(texto8, styleN)
    titulo8 = Paragraph(titulo8, styleT)
    parrafo9 = Paragraph(texto9, styleN)
    titulo9 = Paragraph(titulo9, styleT)
    parrafo10 = Paragraph(texto10, styleN)
    titulo10 = Paragraph(titulo10, styleT)
    parrafo11 = Paragraph(texto11, styleN)
    parrafo12 = Paragraph(texto12, styleN)
    titulo11 = Paragraph(titulo11, styleT)
    parrafo13 = Paragraph(texto13, styleN)
    titulo12 = Paragraph(titulo12, styleT)
    parrafo14 = Paragraph(texto14, styleN)
    


    ancho, alto = letter  # Asegúrate de tener estas dimensiones definidas
    #frame = Frame(120, 720, ancho - 100, alto - 100, id='frameTextoConstante')  # Ajusta las dimensiones según sea necesario
    elementos = [
        titulo1, Spacer(1,25), 
        parrafo, Spacer(1,25),
        titulo2,Spacer(1,25), 
        parrafo2, Spacer(1,25),
        titulo3,Spacer(1,25), 
        parrafo3,
        ]
    
  
    # Continuar con SimpleDocTemplate
    frame = Frame(30, 0, width-50, height-50, id='frameTextoConstante')
    frame.addFromList(elementos, c)
    
    # Finalizar la página actual y comenzar una nueva
    c.showPage()

    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)
    caja_iso = 760
    c.drawInlineImage('static/images/logo_vordcab.jpg',45,742, 2 * cm, 1.0 * cm) #Imagen vortec
    c.drawString(425,caja_iso,'Preparado por:')
    c.drawString(435,caja_iso-10,'SUBAD')
    c.drawString(520,caja_iso,'Aprobación')
    c.drawString(515,caja_iso-10,'SUBD ADTVO')
    c.drawString(35,caja_iso-20,'Número de documento')
    c.drawString(36,caja_iso-30,'SEOV-ADQ-N4-01.08')
    c.drawString(145,caja_iso-20,'Clasificación del documento')
    c.drawString(175,caja_iso-30,'No Controlado')
    c.drawString(255,caja_iso-20,'Nivel del documento')
    c.drawString(280,caja_iso-30, 'N5')
    c.drawString(340,caja_iso-20,'Revisión No.')
    c.drawString(352,caja_iso-30,'001')
    c.drawString(410,caja_iso-20,'Fecha de Emisión')
    c.drawString(425,caja_iso-30,'14/02/2022')
    c.drawString(510,caja_iso-20,'Fecha de Revisión')
    c.drawString(525,caja_iso-30,'12/09/2023')
    #Primera Tabla
    caja_proveedor = caja_iso - 120
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(140,750,260,20, fill=True, stroke=False) #Barra azul superior Título
    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(270,755,'Aviso de Privacidad para Proveedores')
    c.setLineWidth(.3) #Grosor

    primer_parrafo = 605 
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(20, primer_parrafo + 80,565,20, fill=True, stroke=False) #Barra azul superior | 3. Finalidades del Tratamiento de los Datos Personales.
    c.rect(20, primer_parrafo - 218,565,20, fill=True, stroke=False) #Barra azul superior | 4. Medios para limitar el uso o divulgación de sus datos personales.
    c.rect(20, primer_parrafo - 320,565,20, fill=True, stroke=False) #Barra azul superior |5. Datos Personales de terceros.
    c.rect(20, primer_parrafo - 475,565,20, fill=True, stroke=False) #Barra azul superior | 6. Datos personales sensibles
    
    # Agregar contenido a la segunda página
    elementos_segunda_pagina = [
        titulo4, Spacer(1,25),
        parrafo4, Spacer(1,25),
        titulo5, Spacer(1,25),
        parrafo5, Spacer(1,25),
        titulo6, Spacer(1,25),
        parrafo6, Spacer(1,25),
        titulo7, Spacer(1,25),
        parrafo7, Spacer(1,25),
    ]

    # Crear frame para la segunda página y agregar elementos
    frame_segunda_pagina = Frame(30, 0, width-50, height-85, id='frameTextoConstante2')
    frame_segunda_pagina.addFromList(elementos_segunda_pagina, c)

    c.showPage()

    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)
    caja_iso = 760
    c.drawInlineImage('static/images/logo_vordcab.jpg',45,742, 2 * cm, 1.0 * cm) #Imagen vortec
    c.drawString(425,caja_iso,'Preparado por:')
    c.drawString(435,caja_iso-10,'SUBAD')
    c.drawString(520,caja_iso,'Aprobación')
    c.drawString(515,caja_iso-10,'SUBD ADTVO')
    c.drawString(35,caja_iso-20,'Número de documento')
    c.drawString(36,caja_iso-30,'SEOV-ADQ-N4-01.08')
    c.drawString(145,caja_iso-20,'Clasificación del documento')
    c.drawString(175,caja_iso-30,'Controlado')
    c.drawString(255,caja_iso-20,'Nivel del documento')
    c.drawString(280,caja_iso-30, 'N5')
    c.drawString(340,caja_iso-20,'Revisión No.')
    c.drawString(352,caja_iso-30,'000')
    c.drawString(410,caja_iso-20,'Fecha de Emisión')
    c.drawString(425,caja_iso-30,'14/02/2022')
    c.drawString(510,caja_iso-20,'Fecha de Revisión')
    c.drawString(525,caja_iso-30,'')
    #Primera Tabla
    caja_proveedor = caja_iso - 150
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(140,750,260,20, fill=True, stroke=False) #Barra azul superior Título
    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(270,755,'Aviso de Privacidad para Proveedores')
    c.setLineWidth(.3) #Grosor

    primer_parrafo = 605 
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(20, primer_parrafo + 75,565,20, fill=True, stroke=False) #Barra azul superior | 7. Derechos ARCO
    c.rect(20, primer_parrafo - 200,565,20, fill=True, stroke=False) #Barra azul superior | 8. Transferencia de datos.
    c.rect(20, primer_parrafo - 445,565,20, fill=True, stroke=False) #Barra azul superior | 9. Conservación de datos
    
    # Agregar contenido a la segunda página
    elementos_tercera_pagina = [
        titulo8, Spacer(1,25),
        parrafo8, Spacer(1,25),
        titulo9, Spacer(1,25),
        parrafo9, Spacer(1,25),
        titulo10, Spacer(1,25),
        parrafo10, Spacer(1,25),
    ]
     # Crear frame para la segunda página y agregar elementos
    frame_tercera_pagina = Frame(30, 0, width-50, height-90, id='frameTextoConstante3')
    frame_tercera_pagina.addFromList(elementos_tercera_pagina, c)


    c.showPage()

    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)
    caja_iso = 760
    c.drawInlineImage('static/images/logo_vordcab.jpg',45,742, 2 * cm, 1.0 * cm) #Imagen vortec
    c.drawString(425,caja_iso,'Preparado por:')
    c.drawString(435,caja_iso-10,'SUBAD')
    c.drawString(520,caja_iso,'Aprobación')
    c.drawString(515,caja_iso-10,'SUBD ADTVO')
    c.drawString(35,caja_iso-20,'Número de documento')
    c.drawString(36,caja_iso-30,'SEOV-ADQ-N4-01.08')
    c.drawString(145,caja_iso-20,'Clasificación del documento')
    c.drawString(175,caja_iso-30,'Controlado')
    c.drawString(255,caja_iso-20,'Nivel del documento')
    c.drawString(280,caja_iso-30, 'N5')
    c.drawString(340,caja_iso-20,'Revisión No.')
    c.drawString(352,caja_iso-30,'000')
    c.drawString(410,caja_iso-20,'Fecha de Emisión')
    c.drawString(425,caja_iso-30,'14/02/2022')
    c.drawString(510,caja_iso-20,'Fecha de Revisión')
    c.drawString(525,caja_iso-30,'')
    
    #Primera Tabla
    caja_proveedor = caja_iso - 150
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(140,750,260,20, fill=True, stroke=False) #Barra azul superior Título
    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(270,755,'Aviso de Privacidad para Proveedores')
    c.setLineWidth(.3) #Grosor

    primer_parrafo = 605 
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(20, primer_parrafo + 21,565,20, fill=True, stroke=False) #Barra azul superior |10.	Modificaciones al presente aviso de privacidad
    c.rect(20, primer_parrafo - 96,565,20, fill=True, stroke=False) #Barra azul superior | Datos Bancarios Dólares
    
    # Agregar contenido a la segunda página
    elementos_cuarta_pagina = [
        parrafo11, Spacer(1,25),
        titulo11, Spacer(1,25),
        parrafo12, Spacer(1,25),
        titulo12, Spacer(1,25),
        parrafo13, Spacer(1,25),
        
        parrafo14, Spacer(1,25),
    ]

    c.setFont('Helvetica',10)
    c.setFillColor(black)
    c.drawString(150,caja_iso-350,'RECIBÍ DE CONFORMIDAD Y DOY MI CONSENTIMIENTO')
    c.drawString(25,caja_iso-380,'Nombre y Firma')
    c.drawString(25,caja_iso-400,'Fecha')
    c.line(150,caja_iso - 380,420, caja_iso - 380)
    c.line(150,caja_iso - 400,420, caja_iso - 400)

    # Crear frame para la segunda página y agregar elementos
    frame_cuarta_pagina = Frame(30, 0, width-50, height-70, id='frameTextoConstante4')
    frame_cuarta_pagina.addFromList(elementos_cuarta_pagina, c)
    # Guardar el canvas
    c.save()
    buf.seek(0)
    return buf 

def pdf_formato_comparativo(request, pk):
    # Aquí va el código para generar el PDF
    compra = Compra.objects.get(id = pk)
    comparativo = Comparativo.objects.get(id = compra.comparativo_model.id)
    productos = Item_Comparativo.objects.filter(comparativo=comparativo)



    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)

    # Colores
    azul = Color(0.0859375,0.1953125,0.30859375) #prussian_blue 
    blanco = colors.white
    negro = colors.black

    # Cabecera azul
    c.setFillColor(azul)
    c.rect(35, 715, 550, 50, stroke=0, fill=1)
    c.setFont("Helvetica-Bold", 16)
    c.setFillColor(white)
    c.drawCentredString(306, 730, "Tabla Comparativa")

    # Logo
    c.drawInlineImage('static/images/logo_vordcab.jpg',40,600, 6 * cm, 3 * cm) #Imagen vortec
    #Estilo de parrafo
    objective_style = ParagraphStyle(
        name='ObjectiveStyle',
        fontName='Helvetica',
        fontSize=12,
        textColor=colors.black,
        spaceBefore=10,
        spaceAfter=10,
        leftIndent=10,
        rightIndent=10,
        alignment=0  # Alineación justificada
    )
    styles = getSampleStyleSheet()
    style = styles["BodyText"]

    # Justificación
    c.setFillColor(negro)
    c.setFont("Helvetica-Bold", 12)
    c.drawString(40, 705, "Comparativa:")
    c.drawString(390, 680, "Fecha de comparativa:")
    c.drawString(220, 660, "Justificación/Observación:")
    #c.rect(400, 610, 180, 40, stroke=1, fill=0)  # Caja de justificación vacía

    c.setFont("Helvetica", 10)
    parrafo = Paragraph(comparativo.nombre, style)
    ancho_parrafo, altura_parrafo = parrafo.wrap(440, 0)
    inicio_x = 125
    inicio_y = 715 - altura_parrafo  # Ajusta el inicio para que crezca hacia abajo
    parrafo.drawOn(c, inicio_x, inicio_y)  # Dibuja el párrafo en las coordenadas (x, y)
    #c.drawString(125, 700, comparativo.nombre + 'aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa')
    fecha_formateada = comparativo.created_at.strftime("%d-%m-%Y")
    c.drawString(530, 680, fecha_formateada)
    text = comparativo.comentarios
    parrafo = Paragraph(text, style)
    # Calcular el ancho y la altura del párrafo
    ancho_parrafo, altura_parrafo = parrafo.wrap(340, 0)  # Limita el ancho del párrafo a 340
    inicio_x = 220
    inicio_y = 650 - altura_parrafo  # Ajusta el inicio para que crezca hacia abajo
    parrafo.drawOn(c, inicio_x, inicio_y)  # Dibuja el párrafo en las coordenadas (x, y)

    c.setFont('Helvetica-Bold',12)
    c.drawString(45, 580, "Proveedor A:")
    c.drawString(45, 550, "Proveedor B:")
    c.drawString(45, 520, "Proveedor C:")
    c.setFont('Helvetica',12)
    c.drawString(125, 580, comparativo.proveedor.razon_social)
    if comparativo.cotizacion:
        c.drawString(125, 565, 'Tiene archivo cotización')
    else:
        c.drawString(125, 565, 'No tiene archivo cotización')
    c.drawString(125, 550, comparativo.proveedor2.razon_social)
    if comparativo.cotizacion2:
        c.drawString(125, 535, 'Tiene archivo cotización')
    else:
        c.drawString(125, 535, 'No tiene archivo cotización')
    c.drawString(125, 520, comparativo.proveedor3.razon_social)
    if comparativo.cotizacion3:
        c.drawString(125, 505, 'Tiene archivo cotización')
    else:
        c.drawString(125, 505, 'No tiene archivo cotización')

    # Definir encabezado y estilos de la tabla
    c.setFont("Helvetica-Bold", 10)
    encabezado = [["Unidad", "Producto", "Código", "Marca", "Modelo", "Precio A", "Precio B", "Precio C"]]
    datos_restantes = []
    altura_maxima = 340  # Altura máxima permitida
    altura_actual = 0    # Altura acumulada de la tabla
    
    # Definir encabezado y estilos de la tabla
    styles = getSampleStyleSheet()
    style_normal = styles["BodyText"]  # Estilo de párrafo base
    style_normal.fontSize = 7  # Ajustar el tamaño de la fuente

    # Añadir cada producto como una fila en la tabla
    for producto in productos:
        if producto.marca:
            valor1 = str(producto.marca)
        else:
            valor1 = ''
        if producto.modelo:
            valor2 = str(producto.modelo)
        else: 
            valor2 = ''
        fila = [
            Paragraph(str(producto.producto.producto.unidad), style_normal),
            Paragraph(str(producto.producto.producto.nombre), style_normal),
            Paragraph(str(producto.producto.producto.codigo), style_normal),
            Paragraph(valor1, style_normal),
            Paragraph(valor2, style_normal),
            Paragraph('$' + f"{producto.precio:,.3f}", style_normal),
            Paragraph('$' + f"{producto.precio2:,.3f}", style_normal),
            Paragraph('$' + f"{producto.precio3:,.3f}", style_normal),
        ]
        # Crear una tabla temporal para calcular la altura de la fila
        tabla_temp = Table([fila], colWidths=[1.4 * cm, 6.5 * cm, 1.5 * cm, 2 * cm, 2 * cm, 2 * cm, 2 * cm, 2 * cm])
        ancho_fila, altura_fila = tabla_temp.wrap(0, 0)

        # Verificar si se supera la altura máxima
        if altura_actual + altura_fila > altura_maxima:
            datos_restantes.append(fila)  # Guardar en datos_restantes
        else:
            encabezado.append(fila)       # Añadir a la tabla
            altura_actual += altura_fila  # Actualizar la altura acumulada


    # Crear la tabla
    tabla = Table(encabezado, colWidths=[1.4 * cm, 6.5 * cm, 1.5 * cm, 2 * cm, 2 * cm, 2 * cm, 2 * cm, 2 * cm])
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), azul),
        ("TEXTCOLOR", (0, 0), (-1, 0), blanco),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, 0), 6),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("BACKGROUND", (0, 1), (-1, -1), blanco),
        ("GRID", (0, 0), (-1, -1), 1, negro),
    ]))

    # Calcular la altura de la tabla
    ancho_tabla, altura_tabla = tabla.wrap(0, 0)  # Calcular el tamaño necesario para la tabla

    # Ajustar la posición inicial para que la tabla crezca hacia abajo
    inicio_x = 35
    inicio_y = 480 - altura_actual

    # Dibujar la tabla en el PDF en la posición ajustada
    tabla.drawOn(c, inicio_x, inicio_y)

    # Área de firmas
    c.setFont("Helvetica", 8)
    c.drawString(120, 100, "Gerencia")
    c.line(70, 110, 200, 110)
    c.drawString(285, 100, "Comprador")
    c.line(240, 110, 370, 110)
    c.drawString(470, 100, "Superintendente")
    c.line(430, 110, 560, 110)
    #Nombres
    if compra.oc_autorizada_por:
        autorizado1 = str(compra.oc_autorizada_por.staff.staff.first_name) + ' ' + str(compra.oc_autorizada_por.staff.staff.last_name)
    else:
        autorizado1 = ''
    if compra.creada_por:
        comprador = str(compra.creada_por.staff.staff.first_name) + ' ' + str(compra.creada_por.staff.staff.last_name)
    else:
        comprador = ''
    if compra.oc_autorizada_por2:
        autorizado2 = str(compra.oc_autorizada_por2.staff.staff.first_name) + ' ' + str(compra.oc_autorizada_por2.staff.staff.last_name)
    else:
        autorizado2 = ''
    c.drawCentredString(140, 115, autorizado2)
    c.drawCentredString(305, 115, comprador)
    c.drawCentredString(490, 115, autorizado1)
    # Pie de página
    c.setFillColor(azul)
    c.rect(35, 25, 550, 60, stroke=0, fill=1)
    c.setFillColor(white)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(110, 70, "Preparado por:")
    c.drawCentredString(110, 60, "SUPTE ADQ")
    c.drawCentredString(110, 45, "Nivel:")
    c.drawCentredString(110, 35, "N5")
    c.drawCentredString(240, 70, "Aprobado:")
    c.drawCentredString(240, 60, "SUB ADM")
    c.drawCentredString(240, 45, "Rev:")
    c.drawCentredString(240, 35, "000")
    c.drawCentredString(370, 70, "No. Documento:")
    c.drawCentredString(370, 60, "SEOV-ADQ-N4-01.05")
    c.drawCentredString(370, 45, "Fecha emisión:")
    c.drawCentredString(370, 35, "03-MAY-2023")
    c.drawCentredString(520, 70, "Clasificación:")
    c.drawCentredString(520, 60, "Controlado")
    c.drawCentredString(520, 45, "Fecha revisión")
    # Configuración de la altura máxima por página
    altura_maxima_pagina = 550
    # Lista para almacenar todos los elementos del documento
    elementos = []
    while datos_restantes:
        encabezado_restante = []  # Encabezado de la tabla actual
        altura_actual = 0         # Altura acumulada de la tabla

        # Añadir filas a la tabla mientras no supere la altura máxima
        i = 0
        while i < len(datos_restantes):
            fila = datos_restantes[i]
            
            # Crear tabla temporal para calcular la altura de la fila
            tabla_temp = Table([fila], colWidths=[1.4 * cm, 6.5 * cm, 1.5 * cm, 2 * cm, 2 * cm, 2 * cm, 2 * cm, 2 * cm])
            _, altura_fila = tabla_temp.wrap(0, 0)

            # Verificar si la fila cabe en la página
            if altura_actual + altura_fila > altura_maxima_pagina:
                break  # Salir del bucle y procesar los datos restantes en la siguiente página
            
            # Añadir la fila a la tabla actual
            encabezado_restante.append(fila)
            altura_actual += altura_fila
            i += 1

        # Eliminar las filas procesadas de datos_restantes
        datos_restantes = datos_restantes[i:]

        # Crear la tabla con los datos procesados
        tabla_adicional = Table(encabezado_restante, colWidths=[1.4 * cm, 6.5 * cm, 1.5 * cm, 2 * cm, 2 * cm, 2 * cm, 2 * cm, 2 * cm])
        tabla_adicional.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), blanco),
            ("TEXTCOLOR", (0, 0), (-1, 0), blanco),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, 0), 6),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
            ("BACKGROUND", (0, 1), (-1, -1), blanco),
            ("GRID", (0, 0), (-1, -1), 1, negro),
        ]))

        # Añadir un salto de página y luego la tabla al documento
        c.showPage()
        # Calcular la altura de la tabla
        ancho_tabla, altura_tabla = tabla_adicional.wrap(0, 0)  # Calcular el tamaño necesario para la tabla

        # Ajustar la posición inicial para que la tabla crezca hacia abajo
        inicio_x = 35
        inicio_y = 750 - altura_actual

        # Dibujar la tabla en el PDF en la posición ajustada
        tabla_adicional.drawOn(c, inicio_x, inicio_y)
        c.setFillColor(azul)
        c.rect(35, 25, 550, 60, stroke=0, fill=1)
        c.setFillColor(white)
        c.setFont("Helvetica-Bold", 8)
        c.drawCentredString(110, 70, "Preparado por:")
        c.drawCentredString(110, 60, "SUPTE ADQ")
        c.drawCentredString(110, 45, "Nivel:")
        c.drawCentredString(110, 35, "N5")
        c.drawCentredString(240, 70, "Aprobado:")
        c.drawCentredString(240, 60, "SUB ADM")
        c.drawCentredString(240, 45, "Rev:")
        c.drawCentredString(240, 35, "000")
        c.drawCentredString(370, 70, "No. Documento:")
        c.drawCentredString(370, 60, "SEOV-ADQ-N4-01.05")
        c.drawCentredString(370, 45, "Fecha emisión:")
        c.drawCentredString(370, 35, "03-MAY-2023")
        c.drawCentredString(520, 70, "Clasificación:")
        c.drawCentredString(520, 60, "Controlado")
        c.drawCentredString(520, 45, "Fecha revisión")
    # Guardar PDF
    c.showPage()
    c.save()
    buf.seek(0)
    return FileResponse(buf, as_attachment=True, filename=f'Comparativo_{pk}.pdf')


def ver_politica_pdf(request):
    filename = 'politica_antisoborno.pdf'
    carpeta = os.path.join(settings.MEDIA_ROOT, 'politicas')
    filepath = os.path.join(carpeta, filename)

    # Si el archivo ya existe, solo regresamos la URL
    if os.path.exists(filepath):
        url = os.path.join(settings.MEDIA_URL, 'politicas', filename)
        return JsonResponse({'url': url})

    # Si no existe, lo generamos y lo guardamos
    os.makedirs(carpeta, exist_ok=True)
    pdf_buffer = generar_politica_antisoborno()

    with open(filepath, 'wb') as f:
        f.write(pdf_buffer.getbuffer())

    url = os.path.join(settings.MEDIA_URL, 'politicas', filename)
    return JsonResponse({'url': url})

def ver_politica_proveedores(request):
    filename = 'politica_proveedores.pdf'
    carpeta = os.path.join(settings.MEDIA_ROOT, 'politicas')
    filepath = os.path.join(carpeta, filename)

    # Si el archivo ya existe, solo regresamos la URL
    if os.path.exists(filepath):
        url = os.path.join(settings.MEDIA_URL, 'politicas', filename)
        return JsonResponse({'url': url})

    # Si no existe, lo generamos y lo guardamos
    os.makedirs(carpeta, exist_ok=True)
    pdf_buffer = generar_politica_proveedores()

    with open(filepath, 'wb') as f:
        f.write(pdf_buffer.getbuffer())

    url = os.path.join(settings.MEDIA_URL, 'politicas', filename)
    return JsonResponse({'url': url})

def ver_aviso_privacidad(request):
    filename = 'aviso_privacidad.pdf'
    carpeta = os.path.join(settings.MEDIA_ROOT, 'politicas')
    filepath = os.path.join(carpeta, filename)

    # Si el archivo ya existe, solo regresamos la URL
    if os.path.exists(filepath):
        url = os.path.join(settings.MEDIA_URL, 'politicas', filename)
        return JsonResponse({'url': url})

    # Si no existe, lo generamos y lo guardamos
    os.makedirs(carpeta, exist_ok=True)
    pdf_buffer = generar_aviso_privacidad()

    with open(filepath, 'wb') as f:
        f.write(pdf_buffer.getbuffer())

    url = os.path.join(settings.MEDIA_URL, 'politicas', filename)
    return JsonResponse({'url': url})

def ver_codigo_etica(request):
    filename = 'codigo_etica.pdf'
    carpeta = os.path.join(settings.MEDIA_ROOT, 'politicas')
    filepath = os.path.join(carpeta, filename)

    # Si el archivo ya existe, solo regresamos la URL
    if os.path.exists(filepath):
        url = os.path.join(settings.MEDIA_URL, 'politicas', filename)
        return JsonResponse({'url': url})

    # Si no existe, lo generamos y lo guardamos
    os.makedirs(carpeta, exist_ok=True)
    pdf_buffer = generar_codigo_etica()

    with open(filepath, 'wb') as f:
        f.write(pdf_buffer.getbuffer())

    url = os.path.join(settings.MEDIA_URL, 'politicas', filename)
    return JsonResponse({'url': url})


@login_required
def politicas_pendientes(request):
    perfil_id = request.session.get('selected_profile_id')
    perfil = Profile.objects.get(id=perfil_id)
    proveedor = Proveedor.objects.get(id=perfil.proveedor.id)
    politicas = []
    print('proveedor:',proveedor)
    if proveedor:
        if not proveedor.acepto_politica:
            politicas.append({
                "nombre": "Política Antisoborno",
                "url": reverse('ver-politica-pdf'),
                "clave": "antisoborno"
            })

        if not proveedor.acepto_politica_proveedor:
            politicas.append({
                "nombre": "Política de Proveedores",
                "url": reverse('ver-politica-proveedores'),
                "clave": "proveedores"
            })
        
        if not proveedor.acepto_aviso_privacidad:
            politicas.append({
                "nombre": "Aviso de Privacidad",
                "url": reverse('ver-aviso-privacidad'),
                "clave": "privacidad"
            })
        
        if not proveedor.acepto_codigo_etica:
            politicas.append({
                "nombre": "Código de Ética",
                "url": reverse('ver-codigo-etica'),
                "clave": "etica"
            })

    return JsonResponse(politicas, safe=False)


def generar_politica_proveedores():
    #Configuration of the PDF object
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    width, height = letter
    #Azul Vordcab
    prussian_blue = Color(0.0859375,0.1953125,0.30859375)
    rojo = Color(0.59375, 0.05859375, 0.05859375)
    #Encabezado
    c.setFillColor(black)
    c.setLineWidth(.2)
    c.setFont('Helvetica',8)
    caja_iso = 760
    #Elaborar caja
    #c.line(caja_iso,500,caja_iso,720)

    #c.drawString(430,caja_iso,'Preparado por:')
    #c.drawString(405,caja_iso-10,'Auditoría de Proveedores')

    #Primera Tabla
    caja_proveedor = caja_iso - 85
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(170,750,260,20, fill=True, stroke=False) #Barra azul superior Título
    c.rect(20,caja_proveedor - 7,565,20, fill=True, stroke=False) #Barra azul superior |Objetivo
    c.rect(20,caja_proveedor - 122,565,20, fill=True, stroke=False) #Linea posterior horizontal
    c.rect(20,caja_proveedor - 227,565,20, fill=True, stroke=False) #Linea posterior horizontal
    c.setFillColor(white)
    c.setLineWidth(.2)
    c.setFont('Helvetica-Bold',14)
    c.drawCentredString(300,755,'Politica de Proveedores')
    c.setLineWidth(.3) #Grosor
    #c.line(20,caja_proveedor-8,20,520) #Eje Y donde empieza, Eje X donde empieza, donde termina eje y,donde termina eje x (LINEA 1 contorno)
    #c.line(585,caja_proveedor-8,585,520) #Linea 2 contorno
    c.drawInlineImage('static/images/logo_vordcab.jpg',45,730, 3 * cm, 1.5 * cm) #Imagen vortec

    c.setFillColor(white)
    c.setFont('Helvetica-Bold',11)
    #c.drawString(200,caja_proveedor,'Objetivo')
    inicio_central = 300
    #c.line(inicio_central,caja_proveedor-25,inicio_central,520) #Linea Central de caja Proveedor | Detalle
    c.setFillColor(black)
    c.setFont('Helvetica-Bold',9)
    

    #Segundo Parrafo
    segundo_parrafo = caja_proveedor - 150
   
   
   
    c.setFillColor(black)
    c.setFont('Helvetica',9)

    #Cuarta tabla
    cuarta_tabla = segundo_parrafo - 400
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    
   
    c.setFillColor(black)
    c.setFont('Helvetica',9)
   

    texto = """El objetivo de esta política es establecer un marco para la gestión responsable y sostenible de las relaciones con los proveedores, 
    buscando el desarrollo de la cadena de valor a lo largo de toda la cadena de suministro. Nuestras operaciones comerciales con proveedores deben 
    mantener altos estándares de calidad, seguridad y respeto al medio ambiente, reflejando nuestro compromiso con la transparencia y confiabilidad.<br/>
    """
    
    texto2 = """Esta Política de Proveedores aplica a todos los proveedores de GRUPO VORDCAB y a todos los empleados involucrados en el proceso de adquisición.
      Asegura que nuestras relaciones comerciales mantengan altos estándares de calidad, ética y sostenibilidad.<br/> """
    

    texto3 = """•	Código de Ética y Políticas: Comunicar nuestro código de ética, política antisoborno y aviso de privacidad a todos nuestros proveedores.<br/>
    •	Igualdad de Oportunidades: Ofrecer igualdad de oportunidades a todas las empresas mediante procesos de evaluación objetivos y confidenciales, con retroalimentación
    y análisis continuo del desempeño.<br/>
    •	Prácticas Comerciales Responsables: Evaluar integralmente los productos y servicios de nuestros proveedores, considerando costo, calidad, tiempo de respuesta,
    capacidad de adaptación, prestaciones, garantías y compromiso con la ética y el desarrollo sustentable.<br/>
    •	Confianza y Respeto: Fomentar relaciones de confianza y respeto mutuo con los proveedores, basadas en la comunicación abierta y transparente.<br/>
    •	Prevención de Delitos: Implementar medidas de control interno para evitar delitos como lavado de activos, corrupción, financiamiento del terrorismo y cohecho.<br/> 
    •	Correo para denuncias: denuncia@grupovordcab.com<br/>"""

    texto4 = """•	Promoción de Calidad, Eficiencia, Ética y Sostenibilidad: Brindar apoyo a los proveedores para mejorar continuamente la calidad de sus productos y servicios,
    y adoptar prácticas más eficientes, éticas y sostenibles.<br/>"""

    texto5 = """•	Comunicación Transparente y Abierta: Mantener una comunicación transparente y abierta con los proveedores, informándoles oportunamente sobre nuestras expectativas
    y requerimientos.<br/>
    •	Cumplimiento de Contratos y Acuerdos: Cumplir con los términos y condiciones de los contratos y acuerdos establecidos.<br/>
    •	Condiciones de Negociación Equitativas: Promover condiciones de negociación equitativas que beneficien a ambas partes.<br/>"""


    titulo1 = """1. Objetivo"""
    titulo2 = """2. Alcance"""
    titulo3 = """3. Lineamiento"""
    subtitulo31 = """3.1 Relaciones con los Proveedores"""
    subtitulo32 = """3.2 Desarrollo de Proveedores"""
    subtitulo33 = """3.3 Negociaciones"""
    
    styles = getSampleStyleSheet()
    styleN = styles["BodyText"]
    styleN.fontSize = 10
    styleN.alignment = TA_JUSTIFY
   
    styleT = styles["Normal"]
    styleT.textColor = white
    styleT.alignment = TA_JUSTIFY

    styleItalic = ParagraphStyle(
        'Title',
        parent = styles["BodyText"],
        fontName = 'Helvetica-Oblique',
        fontSize = 10,
        alignment = TA_JUSTIFY
    )

    titulo1 = Paragraph(titulo1, styleT)
    parrafo = Paragraph(texto, styleN)
    titulo2 = Paragraph(titulo2, styleT)
    parrafo2 = Paragraph(texto2, styleN)
    titulo3 = Paragraph(titulo3, styleT)
    titulo31 = Paragraph(subtitulo31, styleItalic)   
    parrafo3 = Paragraph(texto3, styleN)
    titulo32 = Paragraph(subtitulo32, styleItalic)   
    parrafo4 = Paragraph(texto4, styleN)
    titulo33 = Paragraph(subtitulo33, styleItalic)
    parrafo5 = Paragraph(texto5, styleN)

    ancho, alto = letter  # Asegúrate de tener estas dimensiones definidas
    #frame = Frame(120, 720, ancho - 100, alto - 100, id='frameTextoConstante')  # Ajusta las dimensiones según sea necesario
    elementos = [
        titulo1, Spacer(1,25),
        parrafo, Spacer(1,25), 
        titulo2,Spacer(1,25), 
        parrafo2, Spacer(1,25),
        titulo3,Spacer(1,25),
        titulo31, Spacer(1,10),
        parrafo3, Spacer(1,25),
        titulo32, Spacer(1,10),
        parrafo4, Spacer(1,25),
        titulo33, Spacer(1,10),
        parrafo5, Spacer(1,25),
    ]
    frame = Frame(30, 0, width-50, height-100, id='frameTextoConstante')
    frame.addFromList(elementos, c)

    c.showPage()

    caja_proveedor = caja_iso - 85
    c.setFont('Helvetica',12)
    c.setFillColor(prussian_blue)
    # REC (Dist del eje Y, Dist del eje X, LARGO DEL RECT, ANCHO DEL RECT)
    c.rect(170,750,260,20, fill=True, stroke=False) #Barra azul superior Título
    c.setFillColor(white)
    c.setFont('Helvetica-Bold',14)
    c.setLineWidth(.2)
    c.drawCentredString(300,755,'Politica de Proveedores')
    c.setLineWidth(.3) #Grosor
    
    c.drawInlineImage('static/images/logo_vordcab.jpg',45,730, 3 * cm, 1.5 * cm) #Imagen Vordcab

    c.setFillColor(black)
    c.setFont('Helvetica',9)

    texto6 = """•	Metas de Compra: Realizar el 80% del importe total de compras autorizadas a proveedores nacionales.<br/>
    • Sostenibilidad y Responsabilidad Social: Fomentar prácticas de sostenibilidad y responsabilidad social entre los proveedores locales.
    <br/>"""

    texto7 = """•	Evaluación de Proveedores: Evaluar periódicamente a los proveedores para asegurar el cumplimiento de los principios y 
    requisitos establecidos en esta política.<br/>
    •	Monitoreo del Desempeño: Monitorear el desempeño de los proveedores y proporcionarles retroalimentación para mejorar.<br/>
    •	Canales de Denuncia: Mantener canales de comunicación abiertos, confiables y confidenciales para que los proveedores puedan 
    reportar incumplimientos a nuestras políticas y normas.<br/>
    """

    texto8 = """Esta política será revisada y actualizada periódicamente para reflejar los cambios en las necesidades de la empresa y 
    el entorno empresarial.<br/>"""

    texto9 = """Esta política será comunicada a todos los empleados de la empresa y a nuestros proveedores.<br/>"""

    texto10 = """El cumplimiento de esta política es obligatorio para todos los proveedores de la empresa.<br/>
    El incumplimiento puede resultar en la suspensión o terminación de la relación comercial.<br/>"""

    texto11 = """GRUPO VORDCAB se compromete a implementar los mecanismos necesarios para asegurar el cumplimiento de esta política. Esto incluye:<br/>
    •	Desarrollo de Procedimientos: Crear y establecer procedimientos específicos para la gestión de proveedores según los lineamientos de esta política.<br/>
    •	Capacitación y Sensibilización: Brindar capacitación continua a los empleados sobre la política de proveedores y sus responsabilidades.<br/>
    •	Canales de Comunicación: Establecer y mantener canales de comunicación efectivos con los proveedores para asegurar la transparencia y la colaboración.<br/>
    •	Asignación de Recursos: Proveer los recursos humanos y financieros necesarios para la implementación y mantenimiento de esta política.<br/>
    •	Monitoreo y Evaluación: Monitorear y evaluar periódicamente la efectividad de los mecanismos implementados para asegurar el cumplimiento de esta política.<br/>
    """

    subtitulo34 = """3.4 Consumo Local"""
    subtitulo35 = """3.5 Evaluación y Monitoreo"""
    subtitulo36 = """3.6 Revisión y Actualización"""
    subtitulo37 = """3.7 Comunicación"""
    subtitulo38 = """3.8 Cumplimiento"""
    subtitulo39 = """3.9 Obligaciones de GRUPO VORDCAB"""


    subtitulo34 = Paragraph(subtitulo34, styleItalic)
    parrafo6 = Paragraph(texto6, styleN)
    subtitulo35 = Paragraph(subtitulo35, styleItalic)
    parrafo7 = Paragraph(texto7, styleN)
    subtitulo36 = Paragraph(subtitulo36, styleItalic)   
    parrafo8 = Paragraph(texto8, styleN)
    subtitulo37 = Paragraph(subtitulo37, styleItalic)   
    parrafo9 = Paragraph(texto9, styleN)
    subtitulo38 = Paragraph(subtitulo38, styleItalic)
    parrafo10 = Paragraph(texto10, styleN)
    subtitulo39 = Paragraph(subtitulo39, styleItalic)
    parrafo11 = Paragraph(texto11, styleN)

    ancho, alto = letter  # Asegúrate de tener estas dimensiones definidas
    #frame = Frame(120, 720, ancho - 100, alto - 100, id='frameTextoConstante')  # Ajusta las dimensiones según sea necesario
    elementos = [
        subtitulo34, Spacer(1,10),
        parrafo6, Spacer(1,25), 
        subtitulo35,Spacer(1,10), 
        parrafo7, Spacer(1,25),
        subtitulo36,Spacer(1,10),
        parrafo8, Spacer(1,25),
        subtitulo37, Spacer(1,10),
        parrafo9, Spacer(1,25),
        subtitulo38, Spacer(1,10),
        parrafo10, Spacer(1,25),
        subtitulo39, Spacer(1,10),
        parrafo11, Spacer(1,25),
    ]
    frame = Frame(30, 0, width-50, height-100, id='frameTextoConstante')
    frame.addFromList(elementos, c)


    c.save()
    buf.seek(0)
    return buf 

@login_required
def debida_diligencia_create(request):
    if request.method == 'POST':
        form = DebidaDiligenciaForm(request.POST)
        if form.is_valid():
            debida_diligencia = form.save()
            
            accionsita_formset = AccionsitaFormSet(request.POST, instance=debida_diligencia, prefix='accionsita')
            miembro_formset = MiembroAltaDireccionFormSet(request.POST, instance=debida_diligencia, prefix='miembro')
            funcionario_formset = FuncionarioPublicoRelacionadoFormSet(request.POST, instance=debida_diligencia, prefix='funcionario')
            relacion_formset = RelacionServidorPublicoFormSet(request.POST, instance=debida_diligencia, prefix='relacion')
            responsable_formset = ResponsableInteraccionFormSet(request.POST, instance=debida_diligencia, prefix='responsable')

            if (accionsita_formset.is_valid() and miembro_formset.is_valid() and 
                funcionario_formset.is_valid() and relacion_formset.is_valid() and 
                responsable_formset.is_valid()):
                
                accionsita_formset.save()
                miembro_formset.save()
                funcionario_formset.save()
                relacion_formset.save()
                responsable_formset.save()
                
                return redirect('alguna-url')
    else:
        form = DebidaDiligenciaForm()
        accionsita_formset = AccionsitaFormSet(prefix='accionsita')
        miembro_formset = MiembroAltaDireccionFormSet(prefix='miembro')
        funcionario_formset = FuncionarioPublicoRelacionadoFormSet(prefix='funcionario')
        relacion_formset = RelacionServidorPublicoFormSet(prefix='relacion')
        responsable_formset = ResponsableInteraccionFormSet(prefix='responsable')

    return render(request, 'ruta/template.html', {
        'form': form,
        'accionsita_formset': accionsita_formset,
        'miembro_formset': miembro_formset,
        'funcionario_formset': funcionario_formset,
        'relacion_formset': relacion_formset,
        'responsable_formset': responsable_formset
    })
