# tasks.py
from django.http import JsonResponse, HttpResponse 
import os
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from celery import shared_task
from .models import Compra, ArticuloComprado
from tesoreria.models import Pago
from django.db.models import F, Avg 
# Import Excel Stuff
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill
from openpyxl.utils import get_column_letter
import datetime as dt
from datetime import date

@shared_task
def convert_excel_matriz_compras_task(compras, requis_atendidas, requis_aprobadas, start_date, end_date):
    #response= HttpResponse(content_type = "application/ms-excel")
    #response['Content-Disposition'] = 'attachment; filename = Matriz_compras_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Compras')
    #Comenzar en la fila 1
    row_num = 1
    #data = json.loads(request.body)
    #compras = data['compras']


    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)
    percent_style = NamedStyle(name='percent_style', number_format='0.00%')
    percent_style.font = Font(name ='Calibri', size = 12)
    wb.add_named_style(percent_style)

    columns = ['Compra','Requisición','Solicitud','Proyecto','Subproyecto','Área','Solicitante','Creado','Req. Autorizada','Proveedor',
               'Crédito/Contado','Costo','Monto_Pagado','Status Pago','Status Autorización','Días de entrega','Moneda',
               'Tipo de cambio','Entregada',"Total en pesos",]

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16
        if col_num == 5:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 25

    columna_max = len(columns)+2

    # Agregar los mensajes
    ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por Savia Vordcab. UH}').style = messages_style
    ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}').style = messages_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 30
    ws.column_dimensions[get_column_letter(columna_max + 1)].width = 30

    # Agregar los encabezados de las nuevas columnas debajo de los mensajes
    ws.cell(row=3, column = columna_max, value="Fecha Inicial").style = head_style
    ws.cell(row=4, column = columna_max, value="Fecha Final").style = head_style
    ws.cell(row=5, column = columna_max, value="Total de OC's").style = head_style
    ws.cell(row=6, column = columna_max, value="Requisiciones Aprobadas").style = head_style
    ws.cell(row=7, column = columna_max, value="Requisiciones Colocadas").style = head_style
    ws.cell(row=8, column = columna_max, value="KPI Colocadas/Aprobadas").style = head_style
    ws.cell(row=9, column = columna_max, value="OC Entregadas").style = head_style
    ws.cell(row=10, column = columna_max, value="OC Autorizadas").style = head_style
    ws.cell(row=11, column = columna_max, value="KPI OC Entregadas/Total de OC").style = head_style
    #ws.cell(row=6, column = columna_max, value="Monto total de OC's").style = head_style

    indicador = requis_atendidas/requis_aprobadas
    letra_columna = get_column_letter(columna_max + 1 )
    formula = f"={letra_columna}9/{letra_columna}10"

    # Asumiendo que las filas de datos comienzan en la fila 2 y terminan en row_num
    ws.cell(row=3, column=columna_max + 1, value=start_date).style = date_style
    ws.cell(row=4, column=columna_max + 1, value=end_date).style = date_style
    ws.cell(row=5, column=columna_max + 1, value=f"=COUNTA(A:A)-1").style = body_style
    ws.cell(row=6, column=columna_max + 1, value=requis_aprobadas).style = body_style
    ws.cell(row=7, column=columna_max + 1, value=requis_atendidas).style = body_style
    ws.cell(row=8, column=columna_max + 1, value=indicador).style = percent_style
    ws.cell(row=9, column=columna_max + 1, value=f'=COUNTIF(S:S,"Entregada")').style = body_style
    ws.cell(row=10, column = columna_max + 1, value=f'=COUNTIF(O:O,"Autorizado")').style = body_style
    ws.cell(row=11, column=columna_max + 1, value=formula).style = percent_style
    #ws.cell(row=4, column=columna_max + 1, value=f"=COUNTIF({get_column_letter(len(columns)-1)}:{get_column_letter(len(columns)-1)}, \"<=3\")").style = body_style
    #ws.cell(row=5, column=columna_max + 1, value=f"={get_column_letter(columna_max+1)}4/{get_column_letter(columna_max+1)}3").style = percent_style
    #ws.cell(row=6, column=columna_max + 1, value=f"=SUM({get_column_letter(len(columns))}:{get_column_letter(len(columns))})").style = money_resumen_style

    rows = []
    for compra_list in compras:
        # Obtén todos los pagos relacionados con esta compra
        oc_id = compra_list.get('id')
        compra = Compra.objects.get(id=oc_id)
        pagos = Pago.objects.filter(oc=oc_id)

        # Calcula el tipo de cambio promedio de estos pagos
        tipo_de_cambio_promedio_pagos = pagos.aggregate(Avg('tipo_de_cambio'))['tipo_de_cambio__avg']

        # Usar el tipo de cambio de los pagos, si existe. De lo contrario, usar el tipo de cambio de la compra
        tipo_de_cambio = tipo_de_cambio_promedio_pagos or compra.tipo_de_cambio
        autorizado_text = 'Autorizado' if compra.autorizado2 else 'No Autorizado' if compra.autorizado2 == False or compra.autorizado1 == False else 'Pendiente Autorización'
        
        entrada_text = 'Entregada' if compra.entrada_completa else 'No Entregada' 

        row = [
        compra.folio,
        compra.req.folio,
        compra.req.orden.folio,
        compra.req.orden.proyecto.nombre if compra.req.orden.proyecto else '',
        compra.req.orden.subproyecto.nombre if compra.req.orden.subproyecto else '',
        compra.req.orden.operacion.nombre if compra.req.orden.operacion else '',
        f"{compra.req.orden.staff.staff.staff.first_name} {compra.req.orden.staff.staff.staff.last_name}",
        compra.created_at.replace(tzinfo=None),
        compra.req.approved_at,
        compra.proveedor.nombre.razon_social,
        compra.cond_de_pago.nombre,
        compra.costo_oc,
        compra.monto_pagado,
        pagado_text,
        autorizado_text,
        compra.dias_de_entrega,
        compra.moneda.nombre,
        tipo_de_cambio,
        entrada_text,   
    ]
        if row[16] == "DOLARES":
            if row[17] is None or row[17] < 15:
                row[17] = 17  # o compra.pago_oc.tipo_de_cambio si así es como obtienes el valor correcto de tipo_de_cambio
        elif row[17] is None:  # por si acaso, aún manejar el caso donde 'tipo_de_cambio' es None
            row[17] = ""

        rows.append(row)

    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            (ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style
            if col_num == 8 or col_num == 7:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style
            if col_num in [7, 8, 10, 11, 12, 19]:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = money_style
        # Agregamos la fórmula DATEDIF. Asumiendo que las columnas 'Creado' y 'Req. Autorizada'
        # están en las posiciones 8 y 9 respectivamente (empezando desde 0), las posiciones en Excel serán 9 y 10 (empezando desde 1).
        #ws.cell(row=row_num, column=len(columns)-1, value=f"=NETWORKDAYS(I{row_num}, H{row_num})").style = body_style
        # Agregar la fórmula de "Total en pesos"
        ws.cell(row=row_num, column = len(columns), value=f"=IF(ISBLANK(R{row_num}), L{row_num}, L{row_num}*R{row_num})").style = money_style
        #Agregar la columna para entregas
      
    
    file_name='Matriz_compras_' + str(date.today()) + '.xlsx'
    file_storage_location = os.path.join(settings.MEDIA_ROOT,'reportes',file_name)
    sheet = wb['Sheet']
    wb.remove(sheet)
    #wb.save(response)
    # Guardar el archivo en el sistema de archivos del servidor.
    wb.save(file_storage_location)
    # Guarda el archivo usando el sistema de almacenamiento predeterminado de Django.
    fs = FileSystemStorage()
    with open(file_storage_location, 'rb') as excel_file:
        filename = fs.save('reportes/' + file_name, excel_file)
        file_url = fs.url(filename)

    # Puedes devolver el identificador único (nombre del archivo) o la URL para descargar.
    # return JsonResponse({'report_id': filename}) # Devolver el identificador único.
    return {'file_url': file_url}  # Devolver la URL para descargar.
    #return(response)

@shared_task
def convert_excel_solicitud_matriz_productos_task(productos):
    #response= HttpResponse(content_type = "application/ms-excel")
    #response['Content-Disposition'] = 'attachment; filename = Solicitudes_por_producto_' + str(dt.date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Compras_Producto')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)


    columns = ['OC','Código', 'Producto','Cantidad','Unidad','Familia','Subfamilia','P.U.','Moneda','TC','Subtotal','IVA','Total','Proveedor','Status Proveedor','Fecha','Proyecto','Subproyecto','Distrito','RQ','Sol','Status','Pagada']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        ws.column_dimensions[get_column_letter(col_num + 1)].width = 16
        if col_num == 4 or col_num == 7:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 25
        if col_num == 11:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30



    columna_max = len(columns)+2

    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por SAVIA 2.0. UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Grupo Vordcab S.A. de C.V.}')).style = messages_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 20
    
    
    rows = []
    for producto in productos:
        producto_id = producto.get('id')
        articulo = ArticuloComprado.objects.get(id=producto_id)
        # Extract the needed attributes
        compra_id = articulo.oc.id
        moneda_nombre = articulo.oc.moneda.nombre
        #nombre_completo = articulo.oc.req.orden.staff.staff.staff.first_name + " " + articulo.oc.req.orden.staff.staff.staff.last_name
        proyecto_nombre = articulo.oc.req.orden.proyecto.nombre if articulo.oc.req.orden.proyecto else "Desconocido"
        subproyecto_nombre = articulo.oc.req.orden.subproyecto.nombre if articulo.oc.req.orden.subproyecto else "Desconocido"
        operacion_nombre = articulo.oc.req.orden.operacion.nombre if articulo.oc.req.orden.operacion else "Desconocido"
        fecha_creacion = articulo.created_at.replace(tzinfo=None)
        pagado_text = 'Pagada' if articulo.oc.pagada else 'No Pagada'

        # Calculate total, subtotal, and IVA using attributes from producto
        subtotal_parcial = articulo.subtotal_parcial
        iva_parcial = articulo.iva_parcial
        total = articulo.total
        if articulo.oc.autorizado2 is not None:
            status = 'Autorizado Gerente' if articulo.oc.autorizado2 else 'Cancelada'
        elif articulo.oc.autorizado1 is not None:
            status = 'Autorizado Superintendente' if articulo.oc.autorizado1 else 'Cancelada'
        else:
            status = 'Sin autorizaciones aún'
        # Handling the currency conversion logic
        pagos = Pago.objects.filter(oc_id=compra_id)
        tipo_de_cambio_promedio_pagos = pagos.aggregate(Avg('tipo_de_cambio'))['tipo_de_cambio__avg']
        tipo_de_cambio = tipo_de_cambio_promedio_pagos or articulo.oc.tipo_de_cambio

        if moneda_nombre == "DOLARES" and tipo_de_cambio:
            total = total * tipo_de_cambio

        # Constructing the row
        row = [
            articulo.oc.folio,
            articulo.producto.producto.articulos.producto.producto.codigo,
            articulo.producto.producto.articulos.producto.producto.nombre,
            articulo.cantidad,
            articulo.producto.producto.articulos.producto.producto.unidad,
            articulo.producto.producto.articulos.producto.producto.familia.nombre,
            articulo.producto.producto.articulos.producto.producto.subfamilia.nombre,
            articulo.precio_unitario,
            moneda_nombre,
            tipo_de_cambio,
            subtotal_parcial,
            iva_parcial,
            total,
            articulo.oc.proveedor.nombre.razon_social,
            articulo.oc.proveedor.estatus.nombre,
            fecha_creacion,
            #nombre_completo,
            proyecto_nombre,
            subproyecto_nombre,
            #operacion_nombre,
            articulo.oc.req.orden.distrito.nombre,
            articulo.oc.req.folio,
            articulo.oc.req.orden.folio,
            status,
            pagado_text,
        ]
        rows.append(row)

    #Ahora, iteramos sobre las filas recopiladas para construir el archivo Excel:
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            (ws.cell(row = row_num, column = col_num+1, value=str(row[col_num]))).style = body_style
            if col_num == 5:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = body_style
            if col_num == 15:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = date_style
            if col_num in [7, 10, 11, 12, 16, 17]:
                (ws.cell(row = row_num, column = col_num+1, value=row[col_num])).style = money_style

    file_name='Matriz_compras_por_producto' + str(date.today()) + '.xlsx'
    file_storage_location = os.path.join(settings.MEDIA_ROOT,'reportes',file_name)
    sheet = wb['Sheet']
    wb.remove(sheet)
     # Guardar el archivo en el sistema de archivos del servidor.
    wb.save(file_storage_location)
    # Guarda el archivo usando el sistema de almacenamiento predeterminado de Django.
    fs = FileSystemStorage()
    with open(file_storage_location, 'rb') as excel_file:
        filename = fs.save('reportes/' + file_name, excel_file)
        file_url_productos = fs.url(filename)

    return {'file_url_productos': file_url_productos}  # Devolver la URL para descargar.